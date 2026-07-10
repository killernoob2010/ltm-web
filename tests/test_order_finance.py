"""订单融资进度监控 — Excel parser and import tests."""
from pathlib import Path
import asyncio
import json
import os
import sys
from datetime import date, timedelta

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from app import db
from app import order_finance
from app.order_finance import (
    DuplicateOrderFinanceError,
    create_manual_order_finance_record,
    derive_business_status,
    find_order_finance_duplicates,
    import_order_finance_directory,
    import_order_finance_upload,
    list_order_finance_records,
    parse_order_finance_directory,
    summarize_order_finance,
    update_management_fields,
    build_order_finance_capital_view,
    build_order_finance_progress_view,
)


LEDGER_DIR = Path("/Users/wangjingze/建龙/贸易处/订单融资合同汇总")
NEW_LEDGER_WORKBOOK = Path("/Users/wangjingze/建龙/贸易处/YOLANDA和香港建龙出口钢材信用证台账.xlsx")


def build_three_sheet_workbook(path: Path) -> Path:
    book = Workbook()
    order = book.active
    order.title = "订单"
    order.append(["订单融资台账"])
    order.append(["单位：人民币元"])
    order.append([
        "项次", "合同数量(吨)", "品名", "供应商简称", "合同编号", "贷款行",
        "贷款人民币金额", "利率", "借款日期", "原到期日", "展期天数", "新到期日",
        "最迟装船日", "提单日期", "交单日期", "还款日", "状态",
    ])
    order.append([
        "Y-2026-1", 1000, "热轧圆钢 42CrMo", "北满", "26BM001", "UOB",
        10_000_000, 0.031, "2026-06-01", "2026-08-01", 5, "2026-08-06",
        "2026-07-20", "2026-06-20", "2026-06-23", None, "存续",
    ])
    order.append([
        "Y-2026-2", 2000, "方坯 Q235", "东钢", "26DG002", "OCBC",
        20_000_000, 0.032, "2026-05-01", "2026-07-01", 0, "2026-07-01",
        None, None, None, None, "结案",
    ])
    order.append([
        "Y-2026-1", 0, "热轧圆钢 42CrMo", "北满", "26BM001", "UOB",
        1_000_000, 0.033, "2026-06-05", "2026-08-01", 5, "2026-08-06",
        "2026-07-15", "2026-06-20", "2026-06-24", None, "存续",
    ])
    order.append([
        "Y-2026-3", 500, "方坯", "西林", "26XL003", "中信唐山",
        5_000_000, 0.031, "2026-06-01", "2026-07-01", 10, None,
        "2026-07-25", None, None, None, "存续",
    ])

    quota = book.create_sheet("额度")
    quota.append(["订单融资额度", None, None, None, None, None])
    quota.append(["单位：万元", None, None, None, None, None])
    quota.append([None, None, None, None, "UOB", "OCBC"])
    quota.append(["限定工厂", None, None, None, "集团内钢厂", "东钢"])
    quota.append(["信用证要求", None, None, None, "不能接受FCR", "可接受FCR"])
    quota.append(["提单要求", None, None, None, "to order", "客户银行"])
    quota.append(["授信额度", None, None, None, 13_400, 40_200])
    quota.append(["目前占用额度", None, None, None, 1_100, 2_000])
    quota.append(["订单融资比例", None, None, None, "90%", "80%-85%"])
    quota.append(["期限", None, None, None, "90天", "90天"])
    quota.append(["目前可用额度", None, None, None, 12_300, 38_200])

    alert = book.create_sheet("预警")
    alert.append(["一、银行交单后待回款预警"])
    alert.append(["项次", "合同编号", "情况说明"])
    alert.append(["Y-2026-1", "26BM001", "请跟进银行回款"])

    conflicting = book.create_sheet("数据合并")
    conflicting.append(["项次", "合同编号", "贷款人民币金额"])
    conflicting.append(["CONFLICT-DATA-MERGE", "WRONG", 999_999_999])
    book.create_sheet("合作")
    book.save(path)
    return path


def progress_record(item_no: str, status: str, **overrides):
    record = {
        "id": overrides.pop("id", 1),
        "business_key": f"ITEM|{item_no}|1",
        "subsidiary": "北满",
        "source_file": "test.xlsx",
        "source_sheet": "订单",
        "purchase_contract_no": f"C-{item_no}",
        "system_contract_no": "",
        "product_name": "钢材",
        "contract_quantity_mt": 1000,
        "finance_bank": "UOB",
        "finance_amount_actual": 10_000_000,
        "finance_drawdown_date": "2026-06-01",
        "finance_due_date": "2099-08-01",
        "business_status": status,
        "finance_status": status,
        "bill_of_lading_date": "",
        "document_submission_date": "",
        "tail_payment_date": "",
        "collection_date": "",
        "import_warnings_json": "[]",
        "source_json": json.dumps({"item_no": item_no}, ensure_ascii=False),
    }
    record.update(overrides)
    return record


def use_temp_db(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setattr(db, "DATA_DIR", tmp_path)
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "order_finance.db")
    db.init_db()


def test_parse_order_finance_directory_reads_new_workbook_all_years():
    result = parse_order_finance_directory(NEW_LEDGER_WORKBOOK)

    assert result["summary"]["files_read"] == 1
    assert result["summary"]["record_count"] == 70
    years = {record["source_snapshot_date"][:4] for record in result["records"]}
    assert years == {"2025", "2026"}
    subsidiaries = {record["subsidiary"] for record in result["records"]}
    assert {"东钢", "北满", "抚顺", "西林", "阿城"}.issubset(subsidiaries)


def test_parse_order_finance_keeps_management_fields_empty_for_import():
    result = parse_order_finance_directory(NEW_LEDGER_WORKBOOK)
    sample = next(record for record in result["records"] if record["purchase_contract_no"])

    assert "planned_drawdown_date" not in sample
    assert "manager_note" not in sample
    assert sample["business_key"]
    assert sample["source_file"].endswith(".xlsx")


def test_three_sheet_parser_uses_order_as_only_order_source(tmp_path):
    workbook = build_three_sheet_workbook(tmp_path / "three-sheet.xlsx")

    result = parse_order_finance_directory(workbook)

    assert result["files"][0]["sheet"] == "订单"
    assert result["files"][0]["sheets"] == {"订单": True, "额度": True, "预警": True}
    assert {record["source_sheet"] for record in result["records"]} == {"订单"}
    assert all("CONFLICT-DATA-MERGE" not in record["source_json"] for record in result["records"])
    first = next(record for record in result["records"] if json.loads(record["source_json"])["item_no"] == "Y-2026-1")
    assert first["bill_of_lading_date"] == "2026-06-20"
    assert first["latest_shipment_date"] == "2026-07-20"
    assert first["document_submission_date"] == "2026-06-23"
    assert first["tail_payment_date"] == ""
    assert first["business_status"] == "存续"
    assert first["source_sheet"] == "订单"


def test_three_sheet_parser_keeps_multiple_financings_without_duplicate_warning(tmp_path):
    workbook = build_three_sheet_workbook(tmp_path / "three-sheet.xlsx")

    result = parse_order_finance_directory(workbook)

    y1_rows = [record for record in result["records"] if json.loads(record["source_json"])["item_no"] == "Y-2026-1"]
    assert len(y1_rows) == 2
    warnings = [warning for record in y1_rows for warning in json.loads(record["import_warnings_json"])]
    assert not any("重复项次" in warning["message"] for warning in warnings)
    assert any("银行交单后待回款预警" in warning["message"] for warning in warnings)
    capital = next(json.loads(record["source_json"]).get("workbook_capital") for record in result["records"] if json.loads(record["source_json"]).get("workbook_capital"))
    assert capital["total_credit"] == 536_000_000
    assert capital["used_credit"] == 31_000_000
    assert {bank["bank"] for bank in capital["banks"]} == {"UOB", "OCBC"}


def test_parser_uses_original_due_plus_extension_when_new_due_is_empty(tmp_path):
    workbook = build_three_sheet_workbook(tmp_path / "three-sheet.xlsx")

    result = parse_order_finance_directory(workbook)
    record = next(record for record in result["records"] if json.loads(record["source_json"])["item_no"] == "Y-2026-3")

    assert record["finance_due_date"] == "2026-07-11"


def test_progress_uses_earliest_latest_shipment_and_real_repayment_timing(tmp_path):
    workbook = build_three_sheet_workbook(tmp_path / "three-sheet.xlsx")
    records = parse_order_finance_directory(workbook)["records"]
    records.append(progress_record(
        "CLOSED",
        "结案",
        id=99,
        finance_due_date="2026-03-10",
        tail_payment_date="2026-03-08",
        import_warnings_json=json.dumps([{"field": "item_no", "level": "高", "message": "历史异常"}], ensure_ascii=False),
    ))

    view = build_order_finance_progress_view(records)
    by_item = {item["item_no"]: item for item in view["contracts"]}

    assert by_item["Y-2026-1"]["latest_shipment_date"] == "2026-07-15"
    assert by_item["CLOSED"]["repayment_timing"] == "提前 2 天还款"
    assert view["summary"]["data_issues"] == sum(item["data_issue_count"] for item in view["contracts"] if item["stage"] != "已完成")


def test_multiple_financing_and_active_stage_do_not_raise_risk():
    due = (date.today() + timedelta(days=45)).isoformat()
    shipment = (date.today() + timedelta(days=40)).isoformat()
    first = progress_record("MULTI", "存续", finance_due_date=due, latest_shipment_date=shipment)
    second = progress_record("MULTI", "存续", id=2, business_key="ITEM|MULTI|2", finance_due_date=due, latest_shipment_date=shipment)

    view = build_order_finance_progress_view([first, second])
    item = view["contracts"][0]

    assert item["financing_count"] == 2
    assert item["risk"] == "低"


def test_order_finance_schema_adds_manual_shipment_confirmation_columns(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    with db.connect() as conn:
        columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_finance_progress)").fetchall()}

    assert {"shipment_confirmed_date", "shipment_confirmed_by", "shipment_confirmed_at"}.issubset(columns)


def test_manual_shipment_confirmation_updates_group_and_survives_reimport(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")

    result = order_finance.set_shipment_confirmation(
        "H-2026-3",
        confirmed=True,
        shipment_confirmed_date="2026-07-10",
        updated_by="pytest",
    )
    confirmed = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]

    assert result["updated"] == 2
    assert {row["shipment_confirmed_date"] for row in confirmed} == {"2026-07-10"}
    assert {row["shipment_confirmed_by"] for row in confirmed} == {"pytest"}

    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    reimported = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]
    assert {row["shipment_confirmed_date"] for row in reimported} == {"2026-07-10"}

    order_finance.set_shipment_confirmation("H-2026-3", confirmed=False, updated_by="pytest")
    undone = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]
    assert {row["shipment_confirmed_date"] for row in undone} == {None}


def test_contract_reminder_updates_group_survives_reimport_and_clears(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")

    result = order_finance.set_contract_reminder(
        "H-2026-3",
        manager_note="7月20日确认工厂进度",
        next_follow_up_date="2026-07-20",
        updated_by="pytest",
    )
    saved = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]

    assert result == {
        "item_no": "H-2026-3",
        "manager_note": "7月20日确认工厂进度",
        "next_follow_up_date": "2026-07-20",
        "updated": 2,
    }
    assert {row["manager_note"] for row in saved} == {"7月20日确认工厂进度"}
    assert {row["next_follow_up_date"] for row in saved} == {"2026-07-20"}

    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    reimported = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]
    assert {row["manager_note"] for row in reimported} == {"7月20日确认工厂进度"}
    assert {row["next_follow_up_date"] for row in reimported} == {"2026-07-20"}

    cleared = order_finance.set_contract_reminder("H-2026-3", manager_note="", next_follow_up_date=None, updated_by="pytest")
    rows_after_clear = [row for row in list_order_finance_records() if json.loads(row["source_json"])["item_no"] == "H-2026-3"]
    assert cleared["updated"] == 2
    assert {row["manager_note"] for row in rows_after_clear} == {""}
    assert {row["next_follow_up_date"] for row in rows_after_clear} == {None}

    try:
        order_finance.set_contract_reminder("H-2026-3", manager_note="日期错误", next_follow_up_date="2026-99-99", updated_by="pytest")
    except ValueError as exc:
        assert str(exc) == "跟进日期格式不正确"
    else:
        raise AssertionError("invalid follow-up date should be rejected")


def test_indicator_risks_color_only_the_fields_that_cause_risk():
    today = date.today()
    records = [
        progress_record("MISSING-SHIP", "存续", id=1, finance_due_date=(today + timedelta(days=45)).isoformat()),
        progress_record("SHIP-SOON", "存续", id=2, finance_due_date=(today + timedelta(days=45)).isoformat(), latest_shipment_date=(today + timedelta(days=5)).isoformat()),
        progress_record("DUE-SOON", "存续", id=3, finance_due_date=(today + timedelta(days=14)).isoformat(), latest_shipment_date=(today + timedelta(days=45)).isoformat()),
        progress_record("REPAID-ACTIVE", "存续", id=4, finance_due_date=(today + timedelta(days=45)).isoformat(), latest_shipment_date=(today + timedelta(days=45)).isoformat(), tail_payment_date=today.isoformat()),
        progress_record(
            "MANUAL-SHIP",
            "存续",
            id=5,
            finance_due_date=(today + timedelta(days=45)).isoformat(),
            shipment_confirmed_date=today.isoformat(),
            import_warnings_json=json.dumps(
                [{"field": "latest_shipment_date", "level": "高", "message": "缺少最迟装船日"}],
                ensure_ascii=False,
            ),
        ),
        progress_record("DOCUMENTED", "存续", id=6, finance_due_date=(today + timedelta(days=45)).isoformat(), document_submission_date=today.isoformat()),
        progress_record("DONE", "结案", id=7, finance_due_date=(today - timedelta(days=45)).isoformat(), import_warnings_json=json.dumps([{"field": "excel_alert", "level": "高", "message": "最迟装船预警"}], ensure_ascii=False)),
    ]

    view = build_order_finance_progress_view(records)
    items = {item["item_no"]: item for item in view["contracts"]}

    assert items["MISSING-SHIP"]["indicator_risks"] == {"shipment": "高", "finance_due": "低", "repayment": "低", "confirmation": "低", "reminder": "低"}
    assert items["MISSING-SHIP"]["risk"] == "高"
    assert items["SHIP-SOON"]["indicator_risks"]["shipment"] == "中"
    assert items["DUE-SOON"]["indicator_risks"]["finance_due"] == "中"
    assert items["REPAID-ACTIVE"]["indicator_risks"]["repayment"] == "中"
    assert items["MANUAL-SHIP"]["indicator_risks"]["shipment"] == "低"
    assert items["MANUAL-SHIP"]["stage"] == "已装船待回款"
    assert items["DOCUMENTED"]["indicator_risks"]["shipment"] == "低"
    assert items["DONE"]["indicator_risks"] == {"shipment": "低", "finance_due": "低", "repayment": "低", "confirmation": "低", "reminder": "低"}
    assert items["DONE"]["risk"] == "已完成"


def test_weekly_focus_uses_rolling_ten_day_actions():
    today = date.today()
    future_due = (today + timedelta(days=45)).isoformat()
    future_shipment = (today + timedelta(days=45)).isoformat()
    records = [
        progress_record(
            "HIGH",
            "存续",
            id=1,
            finance_due_date=future_due,
            next_follow_up_date=(today + timedelta(days=10)).isoformat(),
        ),
        progress_record("SHIP-10", "存续", id=2, finance_due_date=future_due, latest_shipment_date=(today + timedelta(days=10)).isoformat()),
        progress_record("SHIP-11", "存续", id=3, finance_due_date=future_due, latest_shipment_date=(today + timedelta(days=11)).isoformat()),
        progress_record("REMINDER-10", "存续", id=4, finance_due_date=future_due, latest_shipment_date=future_shipment, next_follow_up_date=(today + timedelta(days=10)).isoformat(), manager_note="十天后跟进"),
        progress_record("REMINDER-11", "存续", id=5, finance_due_date=future_due, latest_shipment_date=future_shipment, next_follow_up_date=(today + timedelta(days=11)).isoformat(), manager_note="十一天后跟进"),
        progress_record("REMINDER-OVERDUE", "存续", id=6, finance_due_date=future_due, latest_shipment_date=future_shipment, next_follow_up_date=(today - timedelta(days=2)).isoformat(), manager_note="口头约定"),
        progress_record("NOTE-ONLY", "存续", id=7, finance_due_date=future_due, latest_shipment_date=future_shipment, manager_note="只记录备注"),
        progress_record("DONE-REMINDER", "结案", id=8, finance_due_date=future_due, next_follow_up_date=(today - timedelta(days=2)).isoformat(), manager_note="历史备注"),
    ]

    view = build_order_finance_progress_view(records)
    items = {item["item_no"]: item for item in view["contracts"]}

    assert items["SHIP-10"]["indicator_risks"]["shipment"] == "中"
    assert items["SHIP-10"]["is_weekly_focus"] is True
    assert items["SHIP-10"]["weekly_focus_reasons"] == ["shipment_follow_up"]
    assert items["SHIP-11"]["indicator_risks"]["shipment"] == "低"
    assert items["SHIP-11"]["is_weekly_focus"] is False
    assert items["REMINDER-10"]["indicator_risks"]["reminder"] == "中"
    assert items["REMINDER-10"]["risk"] == "中"
    assert items["REMINDER-10"]["is_weekly_focus"] is True
    assert items["REMINDER-11"]["indicator_risks"]["reminder"] == "低"
    assert items["REMINDER-11"]["risk"] == "低"
    assert items["REMINDER-11"]["is_weekly_focus"] is False
    assert items["REMINDER-OVERDUE"]["indicator_risks"]["reminder"] == "中"
    assert items["REMINDER-OVERDUE"]["is_weekly_focus"] is True
    assert items["NOTE-ONLY"]["is_weekly_focus"] is False
    assert items["DONE-REMINDER"]["indicator_risks"]["reminder"] == "低"
    assert items["DONE-REMINDER"]["is_weekly_focus"] is False
    assert items["HIGH"]["weekly_focus_reasons"] == ["high_risk", "manual_follow_up"]
    assert items["REMINDER-10"]["manager_note"] == "十天后跟进"
    assert items["REMINDER-10"]["next_follow_up_date"] == (today + timedelta(days=10)).isoformat()
    assert view["summary"]["focus_risk"] == 4


def test_explicit_status_and_finance_milestones_drive_lifecycle():
    active_repaid = progress_record("Y-1", "存续", tail_payment_date="2026-07-01")
    closed_without_repayment = progress_record("Y-2", "结案", id=2)
    active_documented = progress_record("Y-3", "存续", id=3, document_submission_date="2026-06-20")
    active_billed = progress_record("Y-4", "存续", id=4, bill_of_lading_date="2026-06-18")

    view = build_order_finance_progress_view([active_repaid, closed_without_repayment, active_documented, active_billed])
    stages = {item["item_no"]: item["stage"] for item in view["contracts"]}

    assert stages == {
        "Y-1": "已还款待结案",
        "Y-2": "已完成",
        "Y-3": "已交单待回款",
        "Y-4": "已装船待回款",
    }
    assert view["summary"]["completed"] == 1
    assert view["summary"]["collected_unrepaid"] == 1


def test_capital_view_uses_quota_sheet_metadata_and_cross_checks_order_usage():
    capital = {
        "total_credit": 100_000_000,
        "used_credit": 40_000_000,
        "available_credit": 60_000_000,
        "banks": [{
            "bank": "UOB", "limit": 100_000_000, "used": 40_000_000,
            "available": 60_000_000, "note": "集团内钢厂", "lc_requirement": "",
            "bill_requirement": "to order", "finance_ratio": "90%", "term": "90天",
        }],
    }
    record = progress_record(
        "Y-1",
        "存续",
        finance_amount_actual=30_000_000,
        source_json=json.dumps({"item_no": "Y-1", "workbook_capital": capital}, ensure_ascii=False),
    )

    view = build_order_finance_capital_view([record])

    assert view["summary"]["total_credit"] == 100_000_000
    assert view["summary"]["used_credit"] == 40_000_000
    assert view["summary"]["order_used_credit"] == 30_000_000
    assert view["summary"]["usage_difference"] == 10_000_000
    assert view["bank_usage"][0]["order_used"] == 30_000_000


def test_derive_business_status_for_drawdown_without_bill_of_lading():
    status = derive_business_status(
        {
            "finance_drawdown_date": "2026-06-15",
            "finance_due_date": "2026-07-30",
            "bill_of_lading_date": "",
            "collection_date": "",
            "remark": "",
        }
    )

    assert status["business_status"] == "已放款待装船"
    assert status["next_action"] == "跟进船期和提单"


def test_import_order_finance_directory_preserves_management_fields(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    first = import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    assert first["summary"]["record_count"] == 70

    records = list_order_finance_records()
    target = records[0]
    update_management_fields(
        target["id"],
        {
            "planned_drawdown_date": "2026-07-01",
            "planned_finance_amount": 12345678,
            "repayment_requirement": "放款当日先回25%陈欠",
            "next_action": "等领导确认",
            "manager_note": "pytest manual note",
        },
    )

    second = import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    updated = next(item for item in list_order_finance_records() if item["id"] == target["id"])

    assert second["summary"]["record_count"] == 70
    assert updated["planned_drawdown_date"] == "2026-07-01"
    assert updated["planned_finance_amount"] == 12345678
    assert updated["repayment_requirement"] == "放款当日先回25%陈欠"
    assert updated["next_action"] == "等领导确认"
    assert updated["manager_note"] == "pytest manual note"


def test_imported_record_list_keeps_overseas_entity_for_existing_card_header(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    records = list_order_finance_records()

    assert records
    assert {record["overseas_entity"] for record in records} >= {"YOLANDA", "香港建龙"}


def test_import_new_workbook_archives_previous_excel_source_records(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    old_import = import_order_finance_directory(LEDGER_DIR, imported_by="pytest")
    assert old_import["summary"]["record_count"] >= 25
    assert len(list_order_finance_records()) == old_import["summary"]["record_count"]

    new_import = import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    records = list_order_finance_records()

    assert new_import["summary"]["record_count"] == 70
    assert len(records) == 70
    assert {record["source_file"] for record in records} == {NEW_LEDGER_WORKBOOK.name}


def test_upload_import_preserves_original_file_name(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    class DummyRequest:
        async def body(self):
            return NEW_LEDGER_WORKBOOK.read_bytes()

    uploaded_name = "业务手动导入台账.xlsx"
    result = asyncio.run(import_order_finance_upload(DummyRequest(), uploaded_name, imported_by="pytest"))
    records = list_order_finance_records()

    assert result["summary"]["record_count"] == 70
    assert result["files"][0]["file"] == uploaded_name
    assert {record["source_file"] for record in records} == {uploaded_name}


def test_parse_order_finance_default_path_requires_real_workbook_when_file_is_missing(tmp_path, monkeypatch):
    missing_workbook = tmp_path / "YOLANDA和香港建龙出口钢材信用证台账.xlsx"
    monkeypatch.setattr(order_finance, "LOCAL_DEFAULT_LEDGER_WORKBOOK", missing_workbook)

    with pytest.raises(ValueError, match="目录不存在"):
        parse_order_finance_directory(missing_workbook)


def test_progress_view_groups_contract_items_and_multi_financing():
    records = parse_order_finance_directory(NEW_LEDGER_WORKBOOK)["records"]

    view = build_order_finance_progress_view(records)

    assert view["summary"]["total_contracts"] < len(records)
    assert view["summary"]["completed"] > 0
    assert view["summary"]["open_contracts"] > 0
    multi = [item for item in view["contracts"] if item["financing_count"] > 1]
    assert multi
    sample = multi[0]
    assert sample["stage"] in {"已放款待装船", "已装船待回款", "已交单待回款", "已还款待结案", "已完成"}
    assert len(sample["financings"]) == sample["financing_count"]


def test_capital_view_contains_bank_limits_and_due_buckets():
    records = parse_order_finance_directory(NEW_LEDGER_WORKBOOK)["records"]

    view = build_order_finance_capital_view(records)

    assert view["summary"]["total_credit"] > 0
    assert view["summary"]["used_credit"] > 0
    assert any(row["bank"] == "UOB" and row["limit"] > 0 for row in view["bank_usage"])
    assert {bucket["label"] for bucket in view["due_buckets"]} == {"7天内", "8-30天", "31-60天", "60天以上", "已逾期"}


def test_record_summary_excludes_explicit_closed_orders_from_active_count():
    active = progress_record("Y-1", "存续")
    closed = progress_record("Y-2", "结案", id=2)

    summary = summarize_order_finance([active, closed])

    assert summary["total_count"] == 2
    assert summary["active_count"] == 1


def test_current_workbook_order_amount_unit_reconciles_to_quota_usage():
    records = parse_order_finance_directory(NEW_LEDGER_WORKBOOK)["records"]

    view = build_order_finance_capital_view(records)

    assert abs(view["summary"]["order_used_credit"] - view["summary"]["used_credit"]) < 1
    assert abs(view["summary"]["usage_difference"]) < 1


def test_manual_create_blocks_exact_contract_duplicate(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    payload = {
        "subsidiary": "北满",
        "purchase_contract_no": "MANUAL-PC-001",
        "system_contract_no": "MANUAL-SC-001",
        "terminal_customer": "MOLYCOP",
        "product_name": "Hot Rolled Alloy Steel Round Bars",
        "contract_quantity_mt": 1000,
        "finance_bank": "UOB",
        "finance_due_date": "2026-08-30",
        "planned_finance_amount": 12000000,
        "next_action": "确认放款安排",
    }

    created = create_manual_order_finance_record(payload, created_by="pytest")
    assert created["source_file"] == "手动新增"
    assert created["business_key"] == "北满|MANUAL-PC-001|MANUAL-SC-001"
    assert created["product_name"] == "Hot Rolled Alloy Steel Round Bars"
    assert created["contract_quantity_mt"] == 1000
    assert created["finance_bank"] == "UOB"

    duplicates = find_order_finance_duplicates(payload)
    assert duplicates["exact"]
    assert duplicates["exact"]["id"] == created["id"]

    try:
        create_manual_order_finance_record(payload, created_by="pytest")
    except DuplicateOrderFinanceError as exc:
        assert exc.existing["id"] == created["id"]
    else:
        raise AssertionError("expected exact duplicate to be blocked")


def test_manual_create_reports_similar_duplicates_without_blocking(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    first = create_manual_order_finance_record(
        {
            "subsidiary": "承德",
            "terminal_customer": "SAMSUNG",
            "product_name": "钢材",
            "finance_due_date": "2026-09-15",
            "planned_finance_amount": 8880000,
        },
        created_by="pytest",
    )
    second_payload = {
        "subsidiary": "承德",
        "terminal_customer": "SAMSUNG",
        "product_name": "钢材",
        "finance_due_date": "2026-09-15",
        "planned_finance_amount": 8880000,
    }

    duplicates = find_order_finance_duplicates(second_payload)
    assert duplicates["exact"] is None
    assert [item["id"] for item in duplicates["similar"]] == [first["id"]]

    second = create_manual_order_finance_record(second_payload, created_by="pytest")
    assert second["id"] != first["id"]
    assert second["business_key"].startswith("承德|手动新增|")


def test_item_key_import_does_not_silently_merge_manual_contract_match(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    excel_sample = next(record for record in parse_order_finance_directory(NEW_LEDGER_WORKBOOK)["records"] if record["purchase_contract_no"])
    manual = create_manual_order_finance_record(
        {
            "subsidiary": excel_sample["subsidiary"],
            "purchase_contract_no": excel_sample["purchase_contract_no"],
            "system_contract_no": excel_sample["system_contract_no"],
            "terminal_customer": "手工临时客户",
            "product_name": "手工临时货物",
            "planned_finance_amount": 7654321,
            "repayment_requirement": "先回陈欠",
            "manager_note": "导入前手工备注",
        },
        created_by="pytest",
    )

    import_order_finance_directory(NEW_LEDGER_WORKBOOK, imported_by="pytest")
    records = list_order_finance_records()
    preserved_manual = next(item for item in records if item["id"] == manual["id"])
    imported = [item for item in records if item["source_file"].endswith(".xlsx")]

    assert imported
    assert preserved_manual["source_file"] == "手动新增"
    assert preserved_manual["planned_finance_amount"] == 7654321
    assert preserved_manual["repayment_requirement"] == "先回陈欠"
    assert preserved_manual["manager_note"] == "导入前手工备注"
