"""Validate and transactionally import the confirmed iron-ore basis workbook."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import hashlib
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from . import db


RESULT_HEADERS = [
    "日期", "周次", "年份", "港口", "品种", "湿吨现货价", "质量升贴水",
    "品牌升贴水", "主力连续收盘价", "基差", "数据状态",
]
DETAIL_HEADERS = [
    "日期", "周次", "年份", "港口", "品种", "EBC指标编码", "EBC原始指标名",
    "EBC价格规格Fe", "湿吨现货价", "参数年份", "参数类型", "Fe", "SiO2",
    "Al2O3", "P", "S", "H2O", "S缺失默认0", "价格代理指标",
    "价格规格与参数规格不同", "Fe调整X", "品牌升贴水", "期货序列",
    "主力连续收盘价", "Fe升贴水", "SiO2升贴水", "Al2O3升贴水",
    "P升贴水", "S升贴水", "质量升贴水", "干吨现货价", "标准化现货价",
    "基差", "数据状态", "备注", "规则版本", "参数来源", "参数版本",
    "EBC原始港口名",
]


@dataclass(frozen=True)
class BasisWorkbookData:
    workbook_name: str
    workbook_sha256: str
    results: list[dict[str, Any]]
    details: list[dict[str, Any]]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sheet_rows(sheet, required_headers: list[str]) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        actual_headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"{sheet.title}为空") from exc
    missing = [header for header in required_headers if header not in actual_headers]
    if missing:
        raise ValueError(f"{sheet.title}缺少字段: {', '.join(missing)}")
    output = []
    for row_number, values in enumerate(rows, start=2):
        if not any(value is not None and value != "" for value in values):
            continue
        item = dict(zip(actual_headers, values))
        item["__row__"] = row_number
        output.append(item)
    return output


def _iso_date(value: Any, *, row: int, field: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip()).isoformat()
        except ValueError:
            pass
    raise ValueError(f"第{row}行{field}不是有效日期")


def _text(value: Any, *, row: int, field: str, allow_blank: bool = False) -> str | None:
    if value is None or str(value).strip() == "":
        if allow_blank:
            return None
        raise ValueError(f"第{row}行{field}为空")
    return str(value).strip()


def _number(value: Any, *, row: int, field: str) -> float:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"第{row}行{field}不是有效数值")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第{row}行{field}不是有效数值") from exc
    if not math.isfinite(number):
        raise ValueError(f"第{row}行{field}不是有限数值")
    return number


def _integer(value: Any, *, row: int, field: str) -> int:
    number = _number(value, row=row, field=field)
    if not number.is_integer():
        raise ValueError(f"第{row}行{field}不是整数")
    return int(number)


def _boolean(value: Any) -> int:
    if value in (True, 1, "1", "是", "true", "TRUE"):
        return 1
    return 0


def _week_number(label: str, *, row: int) -> int:
    match = re.search(r"W(\d{1,2})$", label, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"第{row}行周次格式无效")
    week = int(match.group(1))
    if not 1 <= week <= 53:
        raise ValueError(f"第{row}行周次超出范围")
    return week


def _core_key(business_date: str, port: str, product: str) -> str:
    return "|".join((business_date, port, product))


def _business_key(core_key: str, rule_version: str, parameter_version: str) -> str:
    return "|".join((core_key, rule_version, parameter_version))


def _same_number(left: Any, right: Any) -> bool:
    try:
        return math.isclose(float(left), float(right), rel_tol=0, abs_tol=1e-6)
    except (TypeError, ValueError):
        return False


def validate_basis_workbook(
    path: str | Path,
    expected_sha256: str | None = None,
) -> BasisWorkbookData:
    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.is_file():
        raise ValueError(f"工作簿不存在: {workbook_path}")
    workbook_sha256 = _sha256(workbook_path)
    if expected_sha256 and workbook_sha256.lower() != expected_sha256.lower():
        raise ValueError(
            f"工作簿SHA-256不一致: expected={expected_sha256.lower()} actual={workbook_sha256}"
        )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        missing_sheets = [name for name in ("期现数据", "计算明细") if name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError(f"工作簿缺少页签: {', '.join(missing_sheets)}")
        raw_results = _sheet_rows(workbook["期现数据"], RESULT_HEADERS)
        raw_details = _sheet_rows(workbook["计算明细"], DETAIL_HEADERS)
    finally:
        workbook.close()

    result_by_core: dict[str, dict[str, Any]] = {}
    for raw in raw_results:
        row = raw["__row__"]
        business_date = _iso_date(raw["日期"], row=row, field="日期")
        port = _text(raw["港口"], row=row, field="港口")
        product = _text(raw["品种"], row=row, field="品种")
        core = _core_key(business_date, port, product)
        if core in result_by_core:
            raise ValueError(f"期现数据存在重复业务记录: {core}")
        week_label = _text(raw["周次"], row=row, field="周次")
        result_by_core[core] = {
            "__row__": row,
            "business_date": business_date,
            "business_week": _week_number(week_label, row=row),
            "week_label": week_label,
            "business_year": _integer(raw["年份"], row=row, field="年份"),
            "port": port,
            "product": product,
            "wet_spot_price": _number(raw["湿吨现货价"], row=row, field="湿吨现货价"),
            "quality_adjustment": _number(raw["质量升贴水"], row=row, field="质量升贴水"),
            "brand_adjustment": _number(raw["品牌升贴水"], row=row, field="品牌升贴水"),
            "futures_close": _number(raw["主力连续收盘价"], row=row, field="主力连续收盘价"),
            "basis": _number(raw["基差"], row=row, field="基差"),
            "data_status": _text(raw["数据状态"], row=row, field="数据状态"),
        }

    detail_by_core: dict[str, dict[str, Any]] = {}
    for raw in raw_details:
        row = raw["__row__"]
        business_date = _iso_date(raw["日期"], row=row, field="日期")
        port = _text(raw["港口"], row=row, field="港口")
        product = _text(raw["品种"], row=row, field="品种")
        core = _core_key(business_date, port, product)
        if core in detail_by_core:
            raise ValueError(f"计算明细存在重复业务记录: {core}")
        week_label = _text(raw["周次"], row=row, field="周次")
        rule_version = _text(raw["规则版本"], row=row, field="规则版本")
        parameter_version = _text(raw["参数版本"], row=row, field="参数版本")
        detail_by_core[core] = {
            "__row__": row,
            "business_key": _business_key(core, rule_version, parameter_version),
            "business_date": business_date,
            "week_label": week_label,
            "business_year": _integer(raw["年份"], row=row, field="年份"),
            "port": port,
            "product": product,
            "ebc_indicator_code": _text(raw["EBC指标编码"], row=row, field="EBC指标编码", allow_blank=True),
            "ebc_indicator_name": _text(raw["EBC原始指标名"], row=row, field="EBC原始指标名", allow_blank=True),
            "ebc_price_fe": _number(raw["EBC价格规格Fe"], row=row, field="EBC价格规格Fe"),
            "wet_spot_price": _number(raw["湿吨现货价"], row=row, field="湿吨现货价"),
            "parameter_year": _integer(raw["参数年份"], row=row, field="参数年份"),
            "parameter_type": _text(raw["参数类型"], row=row, field="参数类型"),
            "fe": _number(raw["Fe"], row=row, field="Fe"),
            "sio2": _number(raw["SiO2"], row=row, field="SiO2"),
            "al2o3": _number(raw["Al2O3"], row=row, field="Al2O3"),
            "phosphorus": _number(raw["P"], row=row, field="P"),
            "sulfur": _number(raw["S"], row=row, field="S"),
            "h2o": _number(raw["H2O"], row=row, field="H2O"),
            "sulfur_defaulted": _boolean(raw["S缺失默认0"]),
            "price_proxy_indicator": _text(raw["价格代理指标"], row=row, field="价格代理指标", allow_blank=True),
            "price_parameter_spec_diff": _boolean(raw["价格规格与参数规格不同"]),
            "fe_adjustment_x": _number(raw["Fe调整X"], row=row, field="Fe调整X"),
            "brand_adjustment": _number(raw["品牌升贴水"], row=row, field="品牌升贴水"),
            "futures_series": _text(raw["期货序列"], row=row, field="期货序列"),
            "futures_close": _number(raw["主力连续收盘价"], row=row, field="主力连续收盘价"),
            "fe_adjustment": _number(raw["Fe升贴水"], row=row, field="Fe升贴水"),
            "sio2_adjustment": _number(raw["SiO2升贴水"], row=row, field="SiO2升贴水"),
            "al2o3_adjustment": _number(raw["Al2O3升贴水"], row=row, field="Al2O3升贴水"),
            "phosphorus_adjustment": _number(raw["P升贴水"], row=row, field="P升贴水"),
            "sulfur_adjustment": _number(raw["S升贴水"], row=row, field="S升贴水"),
            "quality_adjustment": _number(raw["质量升贴水"], row=row, field="质量升贴水"),
            "dry_spot_price": _number(raw["干吨现货价"], row=row, field="干吨现货价"),
            "standardized_spot_price": _number(raw["标准化现货价"], row=row, field="标准化现货价"),
            "basis": _number(raw["基差"], row=row, field="基差"),
            "data_status": _text(raw["数据状态"], row=row, field="数据状态"),
            "note": _text(raw["备注"], row=row, field="备注", allow_blank=True),
            "rule_version": rule_version,
            "parameter_source": _text(raw["参数来源"], row=row, field="参数来源"),
            "parameter_version": parameter_version,
            "ebc_original_port": _text(raw["EBC原始港口名"], row=row, field="EBC原始港口名", allow_blank=True),
        }

    if set(result_by_core) != set(detail_by_core):
        missing_result = sorted(set(detail_by_core) - set(result_by_core))[:3]
        missing_detail = sorted(set(result_by_core) - set(detail_by_core))[:3]
        raise ValueError(
            f"期现数据与计算明细业务记录不一致: missing_result={missing_result} missing_detail={missing_detail}"
        )

    results: list[dict[str, Any]] = []
    details: list[dict[str, Any]] = []
    compare_fields = (
        "business_year", "week_label", "wet_spot_price", "quality_adjustment",
        "brand_adjustment", "futures_close", "basis", "data_status",
    )
    numeric_fields = {
        "wet_spot_price", "quality_adjustment", "brand_adjustment", "futures_close", "basis"
    }
    seen_business_keys: set[str] = set()
    for core in result_by_core:
        result = result_by_core[core]
        detail = detail_by_core[core]
        for field in compare_fields:
            matches = _same_number(result[field], detail[field]) if field in numeric_fields else result[field] == detail[field]
            if not matches:
                raise ValueError(f"期现数据与计算明细不一致: {core} 字段={field}")
        business_key = detail["business_key"]
        if business_key in seen_business_keys:
            raise ValueError(f"计算明细存在重复业务唯一键: {business_key}")
        seen_business_keys.add(business_key)
        result.update(
            {
                "business_key": business_key,
                "standardized_spot_price": detail["standardized_spot_price"],
                "futures_series": detail["futures_series"],
                "rule_version": detail["rule_version"],
                "parameter_version": detail["parameter_version"],
                "source_workbook_name": workbook_path.name,
                "source_workbook_sha256": workbook_sha256,
            }
        )
        detail.update(
            {
                "source_workbook_name": workbook_path.name,
                "source_workbook_sha256": workbook_sha256,
            }
        )
        result.pop("__row__", None)
        detail.pop("__row__", None)
        results.append(result)
        details.append(detail)

    return BasisWorkbookData(workbook_path.name, workbook_sha256, results, details)


RESULT_COLUMNS = [
    "business_key", "business_date", "business_week", "week_label", "business_year",
    "port", "product", "wet_spot_price", "quality_adjustment", "brand_adjustment",
    "standardized_spot_price", "futures_series", "futures_close", "basis", "data_status",
    "rule_version", "parameter_version", "source_workbook_name", "source_workbook_sha256",
]
DETAIL_COLUMNS = [
    "result_id", "business_key", "business_date", "week_label", "business_year", "port",
    "product", "ebc_indicator_code", "ebc_indicator_name", "ebc_price_fe", "wet_spot_price",
    "parameter_year", "parameter_type", "fe", "sio2", "al2o3", "phosphorus", "sulfur",
    "h2o", "sulfur_defaulted", "price_proxy_indicator", "price_parameter_spec_diff",
    "fe_adjustment_x", "brand_adjustment", "futures_series", "futures_close",
    "fe_adjustment", "sio2_adjustment", "al2o3_adjustment", "phosphorus_adjustment",
    "sulfur_adjustment", "quality_adjustment", "dry_spot_price", "standardized_spot_price",
    "basis", "data_status", "note", "rule_version", "parameter_source", "parameter_version",
    "ebc_original_port", "source_workbook_name", "source_workbook_sha256",
]


def _upsert_sql(table: str, columns: list[str], conflict_column: str) -> str:
    placeholders = ", ".join("?" for _ in columns)
    updates = ", ".join(
        f"{column}=EXCLUDED.{column}"
        for column in columns
        if column not in {"result_id", conflict_column}
    )
    return (
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders}) "
        f"ON CONFLICT({conflict_column}) DO UPDATE SET {updates}, updated_at=CURRENT_TIMESTAMP"
    )


def import_basis_workbook(
    path: str | Path,
    apply: bool = False,
    expected_sha256: str | None = None,
) -> dict[str, Any]:
    data = validate_basis_workbook(path, expected_sha256=expected_sha256)
    summary: dict[str, Any] = {
        "workbook": data.workbook_name,
        "sha256": data.workbook_sha256,
        "result_rows": len(data.results),
        "detail_rows": len(data.details),
        "applied": False,
    }
    if not apply:
        return summary

    db.init_db()
    with db.connect() as conn:
        cur = conn.cursor()
        db._executemany(
            cur,
            _upsert_sql("iron_ore_basis_results", RESULT_COLUMNS, "business_key"),
            [tuple(row[column] for column in RESULT_COLUMNS) for row in data.results],
        )
        result_rows = db._exec(
            cur,
            "SELECT id, business_key FROM iron_ore_basis_results WHERE source_workbook_sha256 = ?",
            (data.workbook_sha256,),
        ).fetchall()
        result_ids = {row["business_key"]: row["id"] for row in result_rows}
        if any(row["business_key"] not in result_ids for row in data.details):
            raise RuntimeError("结果表写入后无法解析全部result_id")
        detail_values = []
        for row in data.details:
            persisted = {"result_id": result_ids[row["business_key"]], **row}
            detail_values.append(tuple(persisted[column] for column in DETAIL_COLUMNS))
        db._executemany(
            cur,
            _upsert_sql("iron_ore_basis_details", DETAIL_COLUMNS, "business_key"),
            detail_values,
        )

    summary["applied"] = True
    return summary
