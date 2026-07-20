"""交易管理模块。

P0 只读事实、业务归类和业务开平关系均通过本独立路由扩展。
"""
from datetime import date, datetime
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import base64
import binascii
import hashlib
import json
from pathlib import Path
import re
import uuid
from typing import Any, Optional

from fastapi import APIRouter, Depends, Header, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel

from . import db
from .permissions import require_permission
from .trading_settlement import parse_settlement_statement
from .trading_valuation import (
    SH_JUNNENG_RULE_VERSION,
    calculate_sh_junneng_settlement,
)


router = APIRouter()

TRADING_MODULES = {
    "overview": "trading_overview",
    "positions": "trading_positions",
    "junneng": "trading_sh_junneng",
    "options": "trading_options",
    "export": "trading_export",
}

TRADE_REQUIRED_HEADERS = {"交易所", "合约", "买卖", "开平", "手数", "成交价", "手续费"}
CLOSE_REQUIRED_HEADERS = {"交易所", "合约", "开仓日期", "买入", "卖出", "手数", "价格", "逐笔平仓盈亏"}
POSITION_REQUIRED_HEADERS = {"交易所", "开仓日期", "合约", "买卖", "手数", "价格", "占用保证金"}
OPTION_CONTRACT_RE = re.compile(
    r"^(?P<underlying>[a-z]+[0-9]+)-(?P<kind>c|p)-(?P<strike>[0-9]+(?:\.[0-9]+)?)$",
    re.IGNORECASE,
)


def _text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any, default: Optional[float] = 0.0) -> Optional[float]:
    if value in (None, ""):
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _is_date_marker(value: Any) -> bool:
    text = _text(value)
    return len(text) == 8 and text.isdigit() and text.startswith("20")


def _asset_type(contract: str) -> str:
    normalized = contract.upper()
    return "option" if "-C-" in normalized or "-P-" in normalized else "future"


def _open_close(value: Any) -> str:
    text = _text(value)
    if text.startswith("开"):
        return "开仓"
    if text.startswith("平"):
        return "平仓"
    return text


def _read_grouped_sheet(path: Path, sheet_name: str, required_headers: set[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少必要工作表：{sheet_name}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration as exc:
            raise ValueError(f"工作表为空：{sheet_name}") from exc
        headers = [_text(value) for value in header_values]
        missing = sorted(required_headers - set(headers))
        if missing:
            raise ValueError(f"{sheet_name}缺少必要字段：{', '.join(missing)}")
        current_date = ""
        records: list[dict[str, Any]] = []
        for row_no, values in enumerate(rows, start=2):
            if _is_date_marker(values[0] if values else None):
                current_date = _text(values[0])
                continue
            if not any(value not in (None, "") for value in values):
                continue
            record = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            record["_date"] = current_date
            record["_row_no"] = row_no
            records.append(record)
        return records
    finally:
        workbook.close()


def _raw_data(row: dict[str, Any]) -> dict[str, Any]:
    result = {}
    for key, value in row.items():
        if key.startswith("_"):
            continue
        result[key] = _text(value) if isinstance(value, (date, datetime)) else value
    return result


def parse_trade_workbook(path: Path) -> list[dict[str, Any]]:
    rows = _read_grouped_sheet(Path(path), "成交记录", TRADE_REQUIRED_HEADERS)
    result = []
    for row in rows:
        contract = _text(row.get("合约"))
        if not contract:
            continue
        result.append(
            {
                "date": row["_date"],
                "trade_time": _text(row.get("成交时间")),
                "exchange": _text(row.get("交易所")).upper(),
                "contract": contract.lower(),
                "asset_type": _asset_type(contract),
                "side": _text(row.get("买卖")),
                "open_close_raw": _text(row.get("开平")),
                "open_close": _open_close(row.get("开平")),
                "quantity": _number(row.get("手数")),
                "price": _number(row.get("成交价")),
                "turnover": _number(row.get("成交额")),
                "source_close_pnl": _number(row.get("平仓盈亏")),
                "fee": _number(row.get("手续费")),
                "hedge_flag": _text(row.get("投保")),
                "premium_cashflow": _number(row.get("权利金收支")),
                "source_row_no": row["_row_no"],
                "raw_data": _raw_data(row),
            }
        )
    return result


def parse_close_workbook(path: Path) -> list[dict[str, Any]]:
    rows = _read_grouped_sheet(Path(path), "平仓记录", CLOSE_REQUIRED_HEADERS)
    result = []
    pending: Optional[dict[str, Any]] = None
    for row in rows:
        contract = _text(row.get("合约"))
        if contract:
            pending = row
            continue
        if pending is None or not (row.get("买入") or row.get("卖出")):
            continue
        contract = _text(pending.get("合约"))
        open_side = "买" if pending.get("买入") not in (None, "", 0) else "卖"
        close_side = "买" if row.get("买入") not in (None, "", 0) else "卖"
        result.append(
            {
                "open_date": _text(pending.get("开仓日期")),
                "close_date": row["_date"],
                "exchange": _text(pending.get("交易所")).upper(),
                "contract": contract.lower(),
                "asset_type": _asset_type(contract),
                "open_side": open_side,
                "close_side": close_side,
                "quantity": _number(row.get("手数")),
                "open_price": _number(pending.get("价格")),
                "close_price": _number(row.get("价格")),
                "fact_close_pnl": _number(row.get("逐笔平仓盈亏")),
                "mark_close_pnl": _number(row.get("盯市平仓盈亏")),
                "fee": _number(row.get("手续费"), None),
                "premium_cashflow": _number(row.get("权利金收支")),
                "source_row_no": row["_row_no"],
                "raw_data": {"open": _raw_data(pending), "close": _raw_data(row)},
            }
        )
        pending = None
    return result


def parse_position_workbook(path: Path) -> list[dict[str, Any]]:
    rows = _read_grouped_sheet(Path(path), "期末持仓", POSITION_REQUIRED_HEADERS)
    result = []
    for row in rows:
        contract = _text(row.get("合约"))
        if not contract:
            continue
        result.append(
            {
                "snapshot_date": row["_date"],
                "exchange": _text(row.get("交易所")).upper(),
                "open_date": _text(row.get("开仓日期")),
                "contract": contract.lower(),
                "asset_type": _asset_type(contract),
                "direction": _text(row.get("买卖")),
                "quantity": _number(row.get("手数")),
                "average_price": _number(row.get("价格")),
                "source_floating_pnl": _number(row.get("浮动盈亏"), None),
                "mark_pnl": _number(row.get("盯市盈亏"), None),
                "margin": _number(row.get("占用保证金")),
                "hedge_flag": _text(row.get("投保")),
                "valuation_price": None,
                "floating_pnl": None,
                "valuation_status": "pending_calculation",
                "source_row_no": row["_row_no"],
                "raw_data": _raw_data(row),
            }
        )
    return result


def _canonical(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float, Decimal)) or str(value).replace(".", "", 1).isdigit():
        try:
            decimal = Decimal(str(value).replace(",", "").strip()).normalize()
            return format(decimal, "f")
        except InvalidOperation:
            pass
    return _text(value).strip().lower()


SIGNATURE_FIELDS = {
    "trade": ("date", "trade_time", "exchange", "contract", "side", "open_close_raw", "quantity", "price", "turnover", "fee"),
    "close": ("open_date", "close_date", "exchange", "contract", "open_side", "close_side", "quantity", "open_price", "close_price", "fact_close_pnl"),
    "position": ("snapshot_date", "exchange", "contract", "direction", "open_date", "quantity", "average_price", "margin"),
}


def build_fact_signature(fact_type: str, account_code: str, row: dict[str, Any]) -> str:
    if fact_type not in SIGNATURE_FIELDS:
        raise ValueError(f"未知事实类型：{fact_type}")
    values = [_canonical(account_code)] + [_canonical(row.get(field)) for field in SIGNATURE_FIELDS[fact_type]]
    return hashlib.sha256("|".join(values).encode("utf-8")).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def preview_trading_import(
    account_id: int,
    trade_path: Optional[Path],
    close_path: Optional[Path],
    position_path: Optional[Path],
    actor: str,
) -> dict[str, Any]:
    if not trade_path or not close_path or not position_path:
        raise ValueError("成交、平仓、持仓三表必须齐全")
    trade_path, close_path, position_path = map(Path, (trade_path, close_path, position_path))
    trades = parse_trade_workbook(trade_path)
    closes = parse_close_workbook(close_path)
    positions = parse_position_workbook(position_path)
    dates = [row["date"] for row in trades] + [row["close_date"] for row in closes] + [row["snapshot_date"] for row in positions]
    dates = [value for value in dates if value]
    range_start = min(dates) if dates else None
    range_end = max(dates) if dates else None
    snapshot_dates = [row["snapshot_date"] for row in positions if row["snapshot_date"]]
    summary = {
        "paths": {"trade": str(trade_path), "close": str(close_path), "position": str(position_path)},
        "counts": {"trade": len(trades), "close": len(closes), "position": len(positions)},
        "range_start": range_start,
        "range_end": range_end,
    }
    with db.connect() as conn:
        cur = conn.cursor()
        account = db._exec(
            cur,
            "SELECT id FROM trading_accounts WHERE id = ? AND is_active = 1",
            (account_id,),
        ).fetchone()
        if not account:
            raise ValueError("交易账户不存在或已停用")
        overlaps = (
            db._exec(
                cur,
                """
                SELECT id, range_start, range_end FROM trading_import_batches
                WHERE account_id = ? AND status = 'active'
                  AND NOT (range_end < ? OR range_start > ?)
                """,
                (account_id, range_start, range_end),
            ).fetchall()
            if range_start and range_end
            else []
        )
        if any(row["range_start"] != range_start or row["range_end"] != range_end for row in overlaps):
            raise ValueError("新批次与有效批次日期范围部分重叠")
        batch_id = db._last_insert_id(
            cur,
            """
            INSERT INTO trading_import_batches
                (account_id, range_start, range_end, position_snapshot_date, status,
                 trade_file_name, close_file_name, position_file_name,
                 trade_file_sha256, close_file_sha256, position_file_sha256,
                 trade_count, close_count, position_count, parse_summary, created_by)
            VALUES (?, ?, ?, ?, 'preview', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                account_id, range_start, range_end, max(snapshot_dates) if snapshot_dates else None,
                trade_path.name, close_path.name, position_path.name,
                _file_sha256(trade_path), _file_sha256(close_path), _file_sha256(position_path),
                len(trades), len(closes), len(positions), json.dumps(summary, ensure_ascii=False), actor,
            ),
        )
        conn.commit()
    return {"preview_batch_id": batch_id, **summary}


def _statement_fact_signature(
    fact_type: str, account_code: str, row: dict[str, Any]
) -> str:
    if fact_type == "trade" and row.get("transaction_no"):
        values = (
            account_code,
            row.get("date"),
            row.get("exchange"),
            row.get("trading_code"),
            row.get("transaction_no"),
        )
    elif fact_type == "close":
        values = (
            account_code, row.get("open_date"), row.get("close_date"),
            row.get("exchange"), row.get("contract"), row.get("open_side"),
            row.get("close_side"), row.get("quantity"),
        )
    elif fact_type == "option_event":
        values = (
            account_code, row.get("event_date"), row.get("exchange"),
            row.get("contract"), row.get("side"), row.get("event_type"),
            row.get("quantity"), row.get("exercise_price"),
        )
    else:
        values = (
            account_code, row.get("snapshot_date"), row.get("exchange"),
            row.get("contract"), row.get("direction"), row.get("open_date"),
            row.get("average_price"), row.get("hedge_flag"),
        )
    payload = "|".join(_canonical(value) for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _rows_with_statement_keys(
    fact_type: str, account_code: str, rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    signatures = [
        _statement_fact_signature(fact_type, account_code, row) for row in rows
    ]
    totals: dict[str, int] = {}
    occurrences: dict[str, int] = {}
    for signature in signatures:
        totals[signature] = totals.get(signature, 0) + 1
    result = []
    for row, signature in zip(rows, signatures):
        occurrences[signature] = occurrences.get(signature, 0) + 1
        stable_key = (
            signature
            if totals[signature] == 1
            else f"{signature}#{occurrences[signature]}"
        )
        result.append({**row, "base_signature": signature, "stable_key": stable_key})
    return result


def _statement_continuity(
    cur, account_id: int, statement: dict[str, Any]
) -> dict[str, Any]:
    range_start = statement["metadata"]["range_start"]
    previous_date_row = db._exec(
        cur,
        """
        SELECT MAX(ps.snapshot_date) AS snapshot_date
        FROM trading_position_snapshots ps
        JOIN trading_fact_identities fi ON fi.id = ps.identity_id
        WHERE fi.account_id = ? AND ps.is_current = 1 AND ps.snapshot_date < ?
        """,
        (account_id, range_start),
    ).fetchone()
    previous_date = previous_date_row["snapshot_date"] if previous_date_row else None
    if not previous_date:
        return {
            "status": "unverified",
            "message": "期初未建立，连续性未校验",
            "previous_snapshot_date": None,
            "difference_lots": None,
        }
    previous = db._exec(
        cur,
        """
        SELECT ps.contract, ps.direction, SUM(ps.quantity) AS quantity
        FROM trading_position_snapshots ps
        JOIN trading_fact_identities fi ON fi.id = ps.identity_id
        WHERE fi.account_id = ? AND ps.is_current = 1 AND ps.snapshot_date = ?
        GROUP BY ps.contract, ps.direction
        """,
        (account_id, previous_date),
    ).fetchall()
    expected: dict[tuple[str, str], float] = {
        (row["contract"], row["direction"]): float(row["quantity"] or 0)
        for row in previous
    }
    for trade in statement["trades"]:
        if trade["open_close"] == "开仓":
            key = (trade["contract"], trade["side"])
            expected[key] = expected.get(key, 0.0) + trade["quantity"]
        elif trade["open_close"] == "平仓":
            direction = "买" if trade["side"] == "卖" else "卖"
            key = (trade["contract"], direction)
            expected[key] = expected.get(key, 0.0) - trade["quantity"]
    for event in statement["exercises"]:
        key = (event["contract"], event["side"])
        expected[key] = expected.get(key, 0.0) - event["quantity"]
    actual: dict[tuple[str, str], float] = {}
    for position in statement["positions"]:
        key = (position["contract"], position["direction"])
        actual[key] = actual.get(key, 0.0) + position["quantity"]
    keys = set(expected) | set(actual)
    differences = {
        f"{contract}|{direction}": round(actual.get((contract, direction), 0.0) - expected.get((contract, direction), 0.0), 8)
        for contract, direction in keys
        if abs(actual.get((contract, direction), 0.0) - expected.get((contract, direction), 0.0)) > 1e-8
    }
    return {
        "status": "passed" if not differences else "failed",
        "message": "期初至期末持仓连续性通过" if not differences else "期初至期末持仓存在差异",
        "previous_snapshot_date": previous_date,
        "difference_lots": round(sum(abs(value) for value in differences.values()), 8),
        "differences": differences,
    }


def preview_settlement_import(
    account_id: int,
    filename: str,
    content: bytes,
    actor: str,
) -> dict[str, Any]:
    if Path(filename).suffix.lower() != ".txt":
        raise ValueError("仅支持 txt 结算单")
    statement = parse_settlement_statement(content, filename)
    metadata = statement["metadata"]
    digest = hashlib.sha256(content).hexdigest()
    with db.connect() as conn:
        cur = conn.cursor()
        account = db._exec(
            cur,
            """
            SELECT id, account_code, statement_account_code
            FROM trading_accounts WHERE id = ? AND is_active = 1
            """,
            (account_id,),
        ).fetchone()
        if not account:
            raise ValueError("交易账户不存在或已停用")
        if (
            account["statement_account_code"]
            and account["statement_account_code"] != metadata["account_code"]
        ):
            raise ValueError("结算单账户与所选交易账户不一致")
        duplicate = db._exec(
            cur,
            """
            SELECT id FROM trading_import_batches
            WHERE account_id = ? AND statement_file_sha256 = ?
              AND status IN ('active', 'confirmed')
            ORDER BY id DESC LIMIT 1
            """,
            (account_id, digest),
        ).fetchone()
        if duplicate:
            return {
                "duplicate_batch_id": duplicate["id"],
                "statement_type": metadata["statement_type"],
                "range_start": metadata["range_start"],
                "range_end": metadata["range_end"],
                "counts": statement["counts"],
            }

        upload_dir = db.DATA_DIR / "trading_import_uploads" / uuid.uuid4().hex
        upload_dir.mkdir(parents=True, exist_ok=False)
        path = upload_dir / "statement.txt"
        path.write_bytes(content)
        priority = 200 if metadata["statement_type"] == "monthly" else 100
        safe_metadata = {
            key: value
            for key, value in metadata.items()
            if key != "account_code"
        }
        summary = {
            "statement_path": str(path),
            "metadata": safe_metadata,
            "counts": statement["counts"],
            "account_summary": statement["account_summary"],
            "binding_required": not bool(account["statement_account_code"]),
            "warnings": statement["warnings"],
            "continuity": _statement_continuity(cur, account_id, statement),
        }
        batch_id = db._last_insert_id(
            cur,
            """
            INSERT INTO trading_import_batches
                (account_id, range_start, range_end, position_snapshot_date, status,
                 statement_type, statement_file_name, statement_file_sha256,
                 statement_account_code_masked, source_priority,
                 trade_count, close_count, position_count, parse_summary, created_by)
            VALUES (?, ?, ?, ?, 'preview', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                account_id, metadata["range_start"], metadata["range_end"],
                metadata["range_end"], metadata["statement_type"], filename, digest,
                metadata["account_code_masked"], priority,
                statement["counts"]["trade"], statement["counts"]["close"],
                statement["counts"]["position"],
                json.dumps(summary, ensure_ascii=False), actor,
            ),
        )
        conn.commit()
    return {
        "preview_batch_id": batch_id,
        "statement_type": metadata["statement_type"],
        "range_start": metadata["range_start"],
        "range_end": metadata["range_end"],
        "account_code_masked": metadata["account_code_masked"],
        "binding_required": summary["binding_required"],
        "counts": statement["counts"],
        "account_summary": statement["account_summary"],
        "warnings": statement["warnings"],
        "continuity": summary["continuity"],
    }


def _rows_with_stable_keys(fact_type: str, account_code: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    signatures = [build_fact_signature(fact_type, account_code, row) for row in rows]
    counts: dict[str, int] = {}
    totals: dict[str, int] = {}
    for signature in signatures:
        totals[signature] = totals.get(signature, 0) + 1
    result = []
    for row, signature in zip(rows, signatures):
        counts[signature] = counts.get(signature, 0) + 1
        stable_key = signature if totals[signature] == 1 else f"{signature}#{counts[signature]}"
        result.append({**row, "base_signature": signature, "stable_key": stable_key})
    return result


def _identity_id(cur, account_id: int, fact_type: str, stable_key: str) -> int:
    existing = db._exec(
        cur,
        "SELECT id FROM trading_fact_identities WHERE account_id = ? AND fact_type = ? AND stable_key = ?",
        (account_id, fact_type, stable_key),
    ).fetchone()
    if existing:
        return existing["id"]
    return db._last_insert_id(
        cur,
        "INSERT INTO trading_fact_identities (account_id, fact_type, stable_key) VALUES (?, ?, ?)",
        (account_id, fact_type, stable_key),
    )


def _insert_source_row(cur, batch_id: int, source_type: str, source_file: str, source_sheet: str, row: dict[str, Any]) -> int:
    raw_json = json.dumps(row["raw_data"], ensure_ascii=False, sort_keys=True)
    return db._last_insert_id(
        cur,
        """
        INSERT INTO trading_source_rows
            (batch_id, source_type, source_file, source_sheet, source_row_no, raw_hash, raw_json)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            batch_id,
            source_type,
            source_file,
            source_sheet,
            row["source_row_no"],
            hashlib.sha256(raw_json.encode("utf-8")).hexdigest(),
            raw_json,
        ),
    )


def _prepare_import_rows(
    cur,
    batch_id: int,
    account_id: int,
    fact_type: str,
    source_file: str,
    source_sheet: str,
    rows: list[dict[str, Any]],
) -> list[tuple[dict[str, Any], int, int]]:
    source_values = []
    for row in rows:
        raw_json = json.dumps(row["raw_data"], ensure_ascii=False, sort_keys=True)
        source_values.append((
            batch_id, fact_type, source_file, source_sheet, row["source_row_no"],
            hashlib.sha256(raw_json.encode("utf-8")).hexdigest(), raw_json,
        ))
    if source_values:
        db._executemany(
            cur,
            """
            INSERT INTO trading_source_rows
                (batch_id, source_type, source_file, source_sheet, source_row_no, raw_hash, raw_json)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            source_values,
        )
    source_ids = {
        row["source_row_no"]: row["id"]
        for row in db._exec(
            cur,
            "SELECT id, source_row_no FROM trading_source_rows WHERE batch_id = ? AND source_type = ?",
            (batch_id, fact_type),
        ).fetchall()
    }
    identities = {
        row["stable_key"]: row["id"]
        for row in db._exec(
            cur,
            "SELECT id, stable_key FROM trading_fact_identities WHERE account_id = ? AND fact_type = ?",
            (account_id, fact_type),
        ).fetchall()
    }
    missing = [
        (account_id, fact_type, row["stable_key"])
        for row in rows if row["stable_key"] not in identities
    ]
    if missing:
        db._executemany(
            cur,
            """
            INSERT INTO trading_fact_identities (account_id, fact_type, stable_key)
            VALUES (?, ?, ?) ON CONFLICT(account_id, fact_type, stable_key) DO NOTHING
            """,
            missing,
        )
        identities = {
            row["stable_key"]: row["id"]
            for row in db._exec(
                cur,
                "SELECT id, stable_key FROM trading_fact_identities WHERE account_id = ? AND fact_type = ?",
                (account_id, fact_type),
            ).fetchall()
        }
    return [(row, source_ids[row["source_row_no"]], identities[row["stable_key"]]) for row in rows]


def confirm_trading_import(preview_batch_id: int, actor: str) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        batch = db._exec(
            cur,
            "SELECT * FROM trading_import_batches WHERE id = ?",
            (preview_batch_id,),
        ).fetchone()
        if not batch or batch["status"] != "preview":
            raise ValueError("预览批次已确认或不可用")
        summary = json.loads(batch["parse_summary"] or "{}")
        paths = {key: Path(value) for key, value in summary.get("paths", {}).items()}
        if set(paths) != {"trade", "close", "position"}:
            raise ValueError("预览批次缺少三表文件信息")
        expected_hashes = {
            "trade": batch["trade_file_sha256"],
            "close": batch["close_file_sha256"],
            "position": batch["position_file_sha256"],
        }
        if any(not paths[key].exists() or _file_sha256(paths[key]) != expected_hashes[key] for key in paths):
            raise ValueError("预览文件已变化，请重新预览")

        account = db._exec(
            cur,
            "SELECT account_code FROM trading_accounts WHERE id = ? AND is_active = 1",
            (batch["account_id"],),
        ).fetchone()
        if not account:
            raise ValueError("交易账户不存在或已停用")
        account_code = account["account_code"]
        trades = _rows_with_stable_keys("trade", account_code, parse_trade_workbook(paths["trade"]))
        closes = _rows_with_stable_keys("close", account_code, parse_close_workbook(paths["close"]))
        positions = _rows_with_stable_keys("position", account_code, parse_position_workbook(paths["position"]))

        previous = db._exec(
            cur,
            """
            SELECT id FROM trading_import_batches
            WHERE account_id = ? AND status = 'active' AND range_start = ? AND range_end = ? AND id <> ?
            """,
            (batch["account_id"], batch["range_start"], batch["range_end"], preview_batch_id),
        ).fetchone()

        previous_assignments: dict[tuple[Any, ...], list[int]] = {}
        if previous:
            inheritance_rows = db._exec(
                cur,
                """
                SELECT old_tf.identity_id, old_sr.source_row_no, old_tf.trade_date,
                       old_tf.exchange, old_tf.contract, old_tf.side, old_tf.open_close
                FROM trading_trade_facts old_tf
                JOIN trading_source_rows old_sr ON old_sr.id = old_tf.source_row_id
                JOIN trading_business_assignments old_ba ON old_ba.trade_identity_id = old_tf.identity_id
                WHERE old_tf.batch_id = ?
                """,
                (previous["id"],),
            ).fetchall()
            for old_row in inheritance_rows:
                key = (
                    old_row["source_row_no"], old_row["trade_date"], old_row["exchange"],
                    old_row["contract"], old_row["side"], old_row["open_close"],
                )
                previous_assignments.setdefault(key, []).append(old_row["identity_id"])

        prepared_trades = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "trade", batch["trade_file_name"], "成交记录", trades
        )
        trade_values = []
        for row, source_row_id, identity_id in prepared_trades:
            inheritance_review = None
            if previous:
                key = (
                    row["source_row_no"], row["date"], row["exchange"],
                    row["contract"], row["side"], row["open_close"],
                )
                inheritance_review = next(
                    (old_identity_id for old_identity_id in previous_assignments.get(key, []) if old_identity_id != identity_id),
                    None,
                )
            verification_status = "inheritance_review_required" if inheritance_review else "file_imported"
            trade_values.append((
                identity_id, preview_batch_id, source_row_id, row["date"], row["trade_time"],
                row["exchange"], row["contract"], row["asset_type"], row["side"],
                row["open_close_raw"], row["open_close"], row["quantity"], row["price"],
                row["turnover"], row["fee"], row["hedge_flag"], row["premium_cashflow"],
                verification_status,
            ))
        if trade_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_trade_facts
                    (identity_id, batch_id, source_row_id, trade_date, trade_time, exchange, contract,
                     asset_type, side, open_close_raw, open_close, quantity, price, turnover, fee,
                     hedge_flag, premium_cashflow, verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                trade_values,
            )

        prepared_closes = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "close", batch["close_file_name"], "平仓记录", closes
        )
        close_values = [
            (
                identity_id, preview_batch_id, source_row_id, row["open_date"], row["close_date"],
                row["exchange"], row["contract"], row["asset_type"], row["open_side"],
                row["close_side"], row["quantity"], row["open_price"], row["close_price"],
                row["fact_close_pnl"], row["fee"], "matched" if row["fee"] is not None else "pending_match",
            )
            for row, source_row_id, identity_id in prepared_closes
        ]
        if close_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_close_facts
                    (identity_id, batch_id, source_row_id, open_date, close_date, exchange, contract,
                     asset_type, open_side, close_side, quantity, open_price, close_price,
                     fact_close_pnl, matched_fee, fee_status, verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'file_imported')
                """,
                close_values,
            )

        prepared_positions = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "position", batch["position_file_name"], "期末持仓", positions
        )
        position_values = [
            (
                identity_id, preview_batch_id, source_row_id, row["snapshot_date"], row["exchange"],
                row["contract"], row["asset_type"], row["direction"], row["open_date"],
                row["quantity"], row["average_price"], row["margin"], row["valuation_price"],
                row["floating_pnl"], row["valuation_status"],
            )
            for row, source_row_id, identity_id in prepared_positions
        ]
        if position_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_position_snapshots
                    (identity_id, batch_id, source_row_id, snapshot_date, exchange, contract, asset_type,
                     direction, open_date, quantity, average_price, margin, valuation_price, floating_pnl,
                     valuation_status, verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'file_imported')
                """,
                position_values,
            )

        if previous:
            for fact_table in (
                "trading_trade_facts",
                "trading_close_facts",
                "trading_position_snapshots",
            ):
                db._exec(
                    cur,
                    f"UPDATE {fact_table} SET is_current = 0 WHERE batch_id = ?",
                    (previous["id"],),
                )
            db._exec(cur, "UPDATE trading_import_batches SET status = 'superseded' WHERE id = ?", (previous["id"],))
        db._exec(
            cur,
            """
            UPDATE trading_import_batches
            SET status = 'active', supersedes_batch_id = ?, confirmed_by = ?, confirmed_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (previous["id"] if previous else None, actor, preview_batch_id),
        )
        conn.commit()
    return {
        "batch_id": preview_batch_id,
        "status": "active",
        "counts": {"trade": len(trades), "close": len(closes), "position": len(positions)},
        "supersedes_batch_id": previous["id"] if previous else None,
    }


def _contract_product_code(contract: str) -> str:
    match = re.match(r"([a-zA-Z]+)", contract)
    return match.group(1).lower() if match else ""


def _statement_close_pnl(cur, row: dict[str, Any]) -> float:
    product_code = _contract_product_code(row["contract"])
    spec = db._exec(
        cur,
        """
        SELECT contract_multiplier FROM trading_contract_specs
        WHERE exchange = ? AND product_code = ? AND asset_type = ? AND is_active = 1
        """,
        (row["exchange"], product_code, row["asset_type"]),
    ).fetchone()
    if not spec:
        raise ValueError(f"合约 {row['contract']} 缺少已验证乘数")
    difference = (
        row["close_price"] - row["open_price"]
        if row["open_side"] == "买"
        else row["open_price"] - row["close_price"]
    )
    return difference * row["quantity"] * float(spec["contract_multiplier"])


def _source_row_changes(old_raw_json: str, new_raw_json: str) -> list[dict[str, Any]]:
    old = json.loads(old_raw_json)
    new = json.loads(new_raw_json)
    changes: list[dict[str, Any]] = []

    def compare(path: str, old_value: Any, new_value: Any) -> None:
        if isinstance(old_value, dict) and isinstance(new_value, dict):
            for key in sorted(set(old_value) | set(new_value)):
                compare(
                    f"{path}.{key}" if path else key,
                    old_value.get(key),
                    new_value.get(key),
                )
            return
        if isinstance(old_value, list) and isinstance(new_value, list):
            for index in range(max(len(old_value), len(new_value))):
                compare(
                    f"{path}[{index}]",
                    old_value[index] if index < len(old_value) else None,
                    new_value[index] if index < len(new_value) else None,
                )
            return
        if old_value != new_value:
            changes.append({"field": path, "old": old_value, "new": new_value})

    compare("", old, new)
    return changes


def _statement_current_decision(
    cur,
    table: str,
    identity_id: int,
    source_row_id: int,
    new_priority: int,
    new_batch_id: int,
    fact_type: str,
) -> tuple[int, int]:
    current = db._exec(
        cur,
        f"""
        SELECT f.id, f.batch_id, sr.raw_hash, sr.raw_json, b.source_priority
        FROM {table} f
        JOIN trading_source_rows sr ON sr.id = f.source_row_id
        JOIN trading_import_batches b ON b.id = f.batch_id
        WHERE f.identity_id = ? AND f.is_current = 1
        ORDER BY f.id DESC LIMIT 1
        """,
        (identity_id,),
    ).fetchone()
    if not current:
        return 1, 0
    new_source = db._exec(
        cur, "SELECT raw_hash, raw_json FROM trading_source_rows WHERE id = ?", (source_row_id,)
    ).fetchone()
    if new_priority < int(current["source_priority"] or 0):
        return 0, 0
    db._exec(cur, f"UPDATE {table} SET is_current = 0 WHERE id = ?", (current["id"],))
    changed = int(current["raw_hash"] != new_source["raw_hash"])
    if changed:
        db._exec(
            cur,
            """
            INSERT INTO trading_fact_source_differences
                (identity_id, fact_type, old_batch_id, new_batch_id, diff_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                identity_id, fact_type, current["batch_id"], new_batch_id,
                json.dumps(
                    {
                        "old_raw_hash": current["raw_hash"],
                        "new_raw_hash": new_source["raw_hash"],
                        "changes": _source_row_changes(
                            current["raw_json"], new_source["raw_json"]
                        ),
                    },
                    ensure_ascii=False,
                ),
            ),
        )
    return 1, changed


def confirm_settlement_import(preview_batch_id: int, actor: str) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        batch = db._exec(
            cur,
            "SELECT * FROM trading_import_batches WHERE id = ?",
            (preview_batch_id,),
        ).fetchone()
        if not batch or batch["status"] != "preview" or not batch["statement_type"]:
            raise ValueError("结算单预览批次已确认或不可用")
        summary = json.loads(batch["parse_summary"] or "{}")
        path = Path(summary.get("statement_path", ""))
        if not path.exists() or _file_sha256(path) != batch["statement_file_sha256"]:
            raise ValueError("预览文件已变化，请重新预览")
        statement = parse_settlement_statement(path.read_bytes(), batch["statement_file_name"])
        account = db._exec(
            cur,
            """
            SELECT account_code, statement_account_code
            FROM trading_accounts WHERE id = ? AND is_active = 1
            """,
            (batch["account_id"],),
        ).fetchone()
        if not account:
            raise ValueError("交易账户不存在或已停用")
        statement_account = statement["metadata"]["account_code"]
        if account["statement_account_code"] not in (None, "", statement_account):
            raise ValueError("结算单账户与所选交易账户不一致")
        if not account["statement_account_code"]:
            db._exec(
                cur,
                "UPDATE trading_accounts SET statement_account_code = ? WHERE id = ?",
                (statement_account, batch["account_id"]),
            )

        account_code = account["account_code"]
        priority = int(batch["source_priority"] or 0)
        replacements = 0
        differences = 0

        trades = _rows_with_statement_keys("trade", account_code, statement["trades"])
        prepared_trades = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "trade",
            batch["statement_file_name"], "成交记录", trades,
        )
        trade_values = []
        for row, source_id, identity_id in prepared_trades:
            is_current, changed = _statement_current_decision(
                cur, "trading_trade_facts", identity_id, source_id,
                priority, preview_batch_id, "trade",
            )
            replacements += int(is_current and changed)
            differences += changed
            trade_values.append(
                (
                    identity_id, preview_batch_id, source_id, row["date"], row["trade_time"],
                    row["exchange"], row["contract"], row["asset_type"], row["side"],
                    row["open_close_raw"], row["open_close"], row["quantity"], row["price"],
                    row["turnover"], row["fee"], row["hedge_flag"],
                    row["premium_cashflow"], is_current,
                )
            )
        if trade_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_trade_facts
                    (identity_id, batch_id, source_row_id, trade_date, trade_time, exchange,
                     contract, asset_type, side, open_close_raw, open_close, quantity, price,
                     turnover, fee, hedge_flag, premium_cashflow, is_current,
                     verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'file_imported')
                """,
                trade_values,
            )

        closes = _rows_with_statement_keys("close", account_code, statement["closes"])
        for row in closes:
            row["fact_close_pnl"] = _statement_close_pnl(cur, row)
        prepared_closes = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "close",
            batch["statement_file_name"], "平仓明细", closes,
        )
        close_values = []
        for row, source_id, identity_id in prepared_closes:
            is_current, changed = _statement_current_decision(
                cur, "trading_close_facts", identity_id, source_id,
                priority, preview_batch_id, "close",
            )
            replacements += int(is_current and changed)
            differences += changed
            close_values.append(
                (
                    identity_id, preview_batch_id, source_id, row["open_date"],
                    row["close_date"], row["exchange"], row["contract"], row["asset_type"],
                    row["open_side"], row["close_side"], row["quantity"], row["open_price"],
                    row["close_price"], row["fact_close_pnl"], is_current,
                )
            )
        if close_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_close_facts
                    (identity_id, batch_id, source_row_id, open_date, close_date, exchange,
                     contract, asset_type, open_side, close_side, quantity, open_price,
                     close_price, fact_close_pnl, is_current, verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'file_imported')
                """,
                close_values,
            )

        positions = _rows_with_statement_keys(
            "position", account_code, statement["positions"]
        )
        prepared_positions = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "position",
            batch["statement_file_name"], "持仓明细", positions,
        )
        position_values = []
        for row, source_id, identity_id in prepared_positions:
            is_current, changed = _statement_current_decision(
                cur, "trading_position_snapshots", identity_id, source_id,
                priority, preview_batch_id, "position",
            )
            replacements += int(is_current and changed)
            differences += changed
            position_values.append(
                (
                    identity_id, preview_batch_id, source_id, row["snapshot_date"],
                    row["exchange"], row["contract"], row["asset_type"], row["direction"],
                    row["open_date"], row["quantity"], row["average_price"], row["margin"],
                    row["valuation_price"], row["floating_pnl"], is_current,
                    row["valuation_status"],
                )
            )
        if position_values:
            db._executemany(
                cur,
                """
                INSERT INTO trading_position_snapshots
                    (identity_id, batch_id, source_row_id, snapshot_date, exchange, contract,
                     asset_type, direction, open_date, quantity, average_price, margin,
                     valuation_price, floating_pnl, is_current, valuation_status,
                     verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'file_imported')
                """,
                position_values,
            )

        db._exec(
            cur,
            "INSERT INTO trading_statement_account_summaries (batch_id, summary_json) VALUES (?, ?)",
            (preview_batch_id, json.dumps(statement["account_summary"], ensure_ascii=False)),
        )
        for movement in statement["cash_movements"]:
            db._exec(
                cur,
                """
                INSERT INTO trading_statement_cash_movements
                    (batch_id, source_row_no, movement_date, movement_type,
                     deposit, withdrawal, raw_json)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    preview_batch_id, movement["source_row_no"], movement["date"],
                    movement["movement_type"], movement["deposit"], movement["withdrawal"],
                    json.dumps(movement["raw_data"], ensure_ascii=False),
                ),
            )
        events = _rows_with_statement_keys(
            "option_event", account_code, statement["exercises"]
        )
        prepared_events = _prepare_import_rows(
            cur, preview_batch_id, batch["account_id"], "option_event",
            batch["statement_file_name"], "行权明细", events,
        )
        for event, source_id, identity_id in prepared_events:
            is_current, changed = _statement_current_decision(
                cur, "trading_close_facts", identity_id, source_id,
                priority, preview_batch_id, "option_event",
            )
            replacements += int(is_current and changed)
            differences += changed
            if is_current:
                db._exec(
                    cur,
                    """UPDATE trading_statement_exercises SET is_current = 0
                       WHERE identity_id = ? AND is_current = 1""",
                    (identity_id,),
                )
            db._exec(
                cur,
                """
                INSERT INTO trading_statement_exercises
                    (batch_id, source_row_no, identity_id, source_row_id, exchange,
                     product, event_date, contract, event_type, event_type_raw, side,
                     quantity, exercise_price, exercise_amount, exercise_pnl, fee,
                     is_current, raw_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    preview_batch_id, event["source_row_no"], identity_id, source_id,
                    event["exchange"], event["product"], event["event_date"],
                    event["contract"], event["event_type"], event["event_type_raw"],
                    event["side"], event["quantity"], event["exercise_price"],
                    event["exercise_amount"], event["exercise_pnl"], event["fee"],
                    is_current,
                    json.dumps(event["raw_data"], ensure_ascii=False),
                ),
            )
            close_side = "卖" if event["side"] == "买" else "买"
            db._exec(
                cur,
                """
                INSERT INTO trading_close_facts
                    (identity_id, batch_id, source_row_id, open_date, close_date,
                     exchange, contract, asset_type, open_side, close_side, quantity,
                     open_price, close_price, fact_close_pnl, matched_fee, is_current,
                     fee_status, data_status, verification_status, settlement_type,
                     event_type_raw, exercise_price, exercise_amount,
                     statement_event_pnl, underlying_link_status)
                VALUES (?, ?, ?, NULL, ?, ?, ?, 'option', ?, ?, ?, 0, 0, 0, ?, ?,
                        'statement_event', 'file_imported', 'pending_event_match', ?, ?,
                        ?, ?, ?, ?)
                """,
                (
                    identity_id, preview_batch_id, source_id, event["event_date"],
                    event["exchange"], event["contract"], event["side"], close_side,
                    event["quantity"], event["fee"], is_current, event["event_type"],
                    event["event_type_raw"], event["exercise_price"],
                    event["exercise_amount"], event["exercise_pnl"],
                    "not_required" if event["event_type"] == "expiry_abandon" else "pending",
                ),
            )
        for item in statement["position_summary"]:
            db._exec(
                cur,
                """
                INSERT INTO trading_statement_position_summaries
                    (batch_id, source_row_no, contract, raw_json)
                VALUES (?, ?, ?, ?)
                """,
                (
                    preview_batch_id, item["source_row_no"], item["contract"],
                    json.dumps(item["raw_data"], ensure_ascii=False),
                ),
            )
        db._exec(
            cur,
            """
            UPDATE trading_import_batches
            SET status = 'active', confirmed_by = ?, confirmed_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (actor, preview_batch_id),
        )
        conn.commit()
    return {
        "batch_id": preview_batch_id,
        "status": "active",
        "counts": statement["counts"],
        "monthly_replacements": replacements if batch["statement_type"] == "monthly" else 0,
        "differences": differences,
    }


def _option_event_snapshot_opening(
    cur, account_id: int, event: dict[str, Any]
) -> Optional[tuple[str, float]]:
    previous_row = db._exec(
        cur,
        """
        SELECT MAX(ps.snapshot_date) AS snapshot_date
        FROM trading_position_snapshots ps
        JOIN trading_fact_identities fi ON fi.id = ps.identity_id
        JOIN trading_import_batches b ON b.id = ps.batch_id AND b.status = 'active'
        WHERE fi.account_id = ? AND ps.is_current = 1
          AND ps.snapshot_date < ?
        """,
        (account_id, event["close_date"]),
    ).fetchone()
    snapshot_date = previous_row["snapshot_date"] if previous_row else None
    if not snapshot_date:
        return None
    snapshot_rows = db._exec(
        cur,
        """
        SELECT ps.open_date, ps.average_price, SUM(ps.quantity) AS quantity
        FROM trading_position_snapshots ps
        JOIN trading_fact_identities fi ON fi.id = ps.identity_id
        WHERE fi.account_id = ? AND ps.is_current = 1
          AND ps.snapshot_date = ? AND ps.contract = ? AND ps.direction = ?
        GROUP BY ps.open_date, ps.average_price
        ORDER BY ps.open_date, ps.average_price
        """,
        (
            account_id, snapshot_date, event["contract"], event["open_side"],
        ),
    ).fetchall()
    remaining = {
        (row["open_date"], round(float(row["average_price"]), 8)):
            float(row["quantity"] or 0)
        for row in snapshot_rows
    }
    closed_rows = db._exec(
        cur,
        """
        SELECT cf.open_date, cf.open_price, SUM(cf.quantity) AS quantity
        FROM trading_close_facts cf
        JOIN trading_fact_identities fi ON fi.id = cf.identity_id
        WHERE fi.account_id = ? AND cf.is_current = 1
          AND cf.contract = ? AND cf.open_side = ?
          AND cf.close_date > ? AND cf.close_date <= ?
          AND cf.identity_id != ?
          AND (cf.settlement_type = 'trade_close'
               OR cf.verification_status = 'matched')
        GROUP BY cf.open_date, cf.open_price
        """,
        (
            account_id, event["contract"], event["open_side"], snapshot_date,
            event["close_date"], event["identity_id"],
        ),
    ).fetchall()
    for row in closed_rows:
        key = (row["open_date"], round(float(row["open_price"]), 8))
        if key in remaining:
            remaining[key] = max(0.0, remaining[key] - float(row["quantity"] or 0))
    needed = float(event["quantity"] or 0)
    if sum(remaining.values()) + 1e-9 < needed:
        return None
    weighted_open = 0.0
    earliest_open_date = None
    for (open_date, open_price), available in remaining.items():
        quantity = min(needed, available)
        if quantity <= 1e-9:
            continue
        weighted_open += open_price * quantity
        earliest_open_date = min(
            value for value in (earliest_open_date, open_date) if value
        )
        needed -= quantity
        if needed <= 1e-9:
            break
    return earliest_open_date, weighted_open / float(event["quantity"])


def _option_contract_parts(contract: str) -> Optional[tuple[str, str]]:
    match = OPTION_CONTRACT_RE.match((contract or "").strip())
    if not match:
        return None
    return (
        match.group("underlying").lower(),
        "call" if match.group("kind").lower() == "c" else "put",
    )


def _option_event_underlying_side(option_kind: str, open_side: str) -> str:
    if option_kind == "call":
        return open_side
    return "卖" if open_side == "买" else "买"


def _link_option_event_underlying_trades(cur, batch_id: int) -> int:
    batch = db._exec(
        cur, "SELECT account_id FROM trading_import_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if not batch:
        return 0
    events = db._exec(
        cur,
        """
        SELECT * FROM trading_close_facts
        WHERE batch_id = ? AND is_current = 1
          AND settlement_type IN ('exercise', 'assignment')
        ORDER BY close_date, source_row_id, id
        """,
        (batch_id,),
    ).fetchall()
    if not events:
        return 0
    event_ids = [int(row["identity_id"]) for row in events]
    placeholders = ",".join("?" for _ in event_ids)
    db._exec(
        cur,
        f"""DELETE FROM trading_option_event_underlying_links
             WHERE event_identity_id IN ({placeholders})""",
        tuple(event_ids),
    )
    used_rows = db._exec(
        cur,
        """
        SELECT l.underlying_trade_identity_id, SUM(l.matched_quantity) AS quantity
        FROM trading_option_event_underlying_links l
        JOIN trading_close_facts cf ON cf.identity_id = l.event_identity_id
        WHERE cf.is_current = 1
        GROUP BY l.underlying_trade_identity_id
        """,
    ).fetchall()
    used = {
        int(row["underlying_trade_identity_id"]): float(row["quantity"] or 0)
        for row in used_rows
    }
    link_count = 0
    for event in events:
        parts = _option_contract_parts(event["contract"])
        if not parts:
            db._exec(
                cur,
                "UPDATE trading_close_facts SET underlying_link_status = 'pending' WHERE id = ?",
                (event["id"],),
            )
            continue
        underlying_contract, option_kind = parts
        underlying_side = _option_event_underlying_side(
            option_kind, event["open_side"]
        )
        candidates = db._exec(
            cur,
            """
            SELECT tf.*
            FROM trading_trade_facts tf
            JOIN trading_fact_identities fi ON fi.id = tf.identity_id
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            WHERE fi.account_id = ? AND tf.is_current = 1
              AND tf.asset_type = 'future' AND tf.open_close = '开仓'
              AND tf.trade_date = ? AND tf.contract = ? AND tf.side = ?
            ORDER BY tf.trade_time, tf.source_row_id, tf.id
            """,
            (
                batch["account_id"], event["close_date"], underlying_contract,
                underlying_side,
            ),
        ).fetchall()
        available = {
            int(row["identity_id"]): max(
                0.0,
                float(row["quantity"] or 0)
                - used.get(int(row["identity_id"]), 0.0),
            )
            for row in candidates
            if abs(float(row["price"]) - float(event["exercise_price"] or 0)) < 1e-8
        }
        needed = float(event["quantity"] or 0)
        if sum(available.values()) + 1e-9 < needed:
            db._exec(
                cur,
                "UPDATE trading_close_facts SET underlying_link_status = 'pending' WHERE id = ?",
                (event["id"],),
            )
            continue
        link_rows = []
        for trade in candidates:
            identity_id = int(trade["identity_id"])
            quantity = min(needed, available.get(identity_id, 0.0))
            if quantity <= 1e-9:
                continue
            link_rows.append((event["identity_id"], identity_id, quantity))
            available[identity_id] -= quantity
            used[identity_id] = used.get(identity_id, 0.0) + quantity
            needed -= quantity
            if needed <= 1e-9:
                break
        db._executemany(
            cur,
            """
            INSERT INTO trading_option_event_underlying_links
                (event_identity_id, underlying_trade_identity_id, matched_quantity,
                 rule_version)
            VALUES (?, ?, ?, 'option-exercise-underlying-v1')
            """,
            link_rows,
        )
        link_count += len(link_rows)
        db._exec(
            cur,
            "UPDATE trading_close_facts SET underlying_link_status = 'matched' WHERE id = ?",
            (event["id"],),
        )
    return link_count


def _match_option_event_allocations(cur, batch_id: int) -> tuple[int, int]:
    batch = db._exec(
        cur, "SELECT account_id FROM trading_import_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if not batch:
        return 0, 0
    events = db._exec(
        cur,
        """
        SELECT * FROM trading_close_facts
        WHERE batch_id = ? AND is_current = 1 AND settlement_type != 'trade_close'
        ORDER BY close_date, source_row_id, id
        """,
        (batch_id,),
    ).fetchall()
    if not events:
        return 0, 0
    event_ids = [int(row["identity_id"]) for row in events]
    placeholders = ",".join("?" for _ in event_ids)
    db._exec(
        cur,
        f"DELETE FROM trading_fact_close_allocations WHERE close_identity_id IN ({placeholders})",
        tuple(event_ids),
    )
    opening_rows = db._exec(
        cur,
        """
        SELECT tf.*
        FROM trading_trade_facts tf
        JOIN trading_fact_identities fi ON fi.id = tf.identity_id
        JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
        WHERE fi.account_id = ? AND tf.is_current = 1
          AND tf.asset_type = 'option' AND tf.open_close = '开仓'
        ORDER BY tf.trade_date, tf.trade_time, tf.source_row_id, tf.id
        """,
        (batch["account_id"],),
    ).fetchall()
    used_rows = db._exec(
        cur,
        """
        SELECT fa.open_trade_identity_id, SUM(fa.matched_quantity) AS quantity
        FROM trading_fact_close_allocations fa
        JOIN trading_close_facts cf ON cf.identity_id = fa.close_identity_id
        WHERE cf.is_current = 1
        GROUP BY fa.open_trade_identity_id
        """,
    ).fetchall()
    used = {
        int(row["open_trade_identity_id"]): float(row["quantity"] or 0)
        for row in used_rows
    }
    remaining = {
        int(row["identity_id"]): max(
            0.0, float(row["quantity"] or 0) - used.get(int(row["identity_id"]), 0.0)
        )
        for row in opening_rows
    }
    allocation_count = 0
    pending_count = 0
    for event in events:
        candidates = [
            row for row in opening_rows
            if row["contract"] == event["contract"]
            and row["side"] == event["open_side"]
            and row["trade_date"] <= event["close_date"]
            and remaining.get(int(row["identity_id"]), 0.0) > 1e-9
        ]
        needed = float(event["quantity"] or 0)
        spec = db._exec(
            cur,
            """
            SELECT contract_multiplier FROM trading_contract_specs
            WHERE exchange = ? AND product_code = ? AND asset_type = 'option'
              AND is_active = 1
            """,
            (event["exchange"], _product_code(event["contract"])),
        ).fetchone()
        available_trade_quantity = sum(
            remaining[int(row["identity_id"])] for row in candidates
        )
        snapshot_opening = None
        if available_trade_quantity + 1e-9 < needed:
            snapshot_opening = _option_event_snapshot_opening(
                cur, int(batch["account_id"]), dict(event)
            )
        if not spec or (available_trade_quantity + 1e-9 < needed and not snapshot_opening):
            pending_count += 1
            continue
        multiplier = float(spec["contract_multiplier"])
        if snapshot_opening:
            open_date, open_price = snapshot_opening
            pnl = calculate_business_pnl(
                open_price, 0.0, event["open_side"], needed, multiplier
            )
            db._exec(
                cur,
                """
                UPDATE trading_close_facts
                SET open_date = ?, open_price = ?, fact_close_pnl = ?,
                    verification_status = 'matched'
                WHERE id = ?
                """,
                (open_date, open_price, round(pnl, 8), event["id"]),
            )
            continue
        allocations = []
        pnl = 0.0
        weighted_open = 0.0
        earliest_open_date = None
        for opening in candidates:
            identity_id = int(opening["identity_id"])
            quantity = min(needed, remaining[identity_id])
            if quantity <= 1e-9:
                continue
            open_price = float(opening["price"])
            allocations.append((event["identity_id"], identity_id, quantity))
            pnl += calculate_business_pnl(
                open_price, 0.0, event["open_side"], quantity, multiplier
            )
            weighted_open += open_price * quantity
            earliest_open_date = min(
                value for value in (earliest_open_date, opening["trade_date"]) if value
            )
            remaining[identity_id] -= quantity
            needed -= quantity
            if needed <= 1e-9:
                break
        db._executemany(
            cur,
            """
            INSERT INTO trading_fact_close_allocations
                (close_identity_id, open_trade_identity_id, matched_quantity,
                 match_rule_version)
            VALUES (?, ?, ?, 'option-event-fifo-v1')
            """,
            allocations,
        )
        allocation_count += len(allocations)
        db._exec(
            cur,
            """
            UPDATE trading_close_facts
            SET open_date = ?, open_price = ?, fact_close_pnl = ?,
                verification_status = 'matched'
            WHERE id = ?
            """,
            (
                earliest_open_date,
                weighted_open / float(event["quantity"]),
                round(pnl, 8),
                event["id"],
            ),
        )
    return allocation_count, pending_count


def match_imported_facts(batch_id: int) -> dict[str, int]:
    with db.connect() as conn:
        cur = conn.cursor()
        closes = db._exec(
            cur,
            """SELECT * FROM trading_close_facts
               WHERE batch_id = ? AND is_current = 1
                 AND settlement_type = 'trade_close' ORDER BY id""",
            (batch_id,),
        ).fetchall()
        trades = db._exec(
            cur,
            "SELECT * FROM trading_trade_facts WHERE batch_id = ? AND is_current = 1 ORDER BY trade_date, trade_time, source_row_id, id",
            (batch_id,),
        ).fetchall()
        db._exec(
            cur,
            """
            DELETE FROM trading_close_trade_links
            WHERE close_identity_id IN (
                SELECT identity_id FROM trading_close_facts WHERE batch_id = ?
            )
            """,
            (batch_id,),
        )
        db._exec(
            cur,
            """
            DELETE FROM trading_fact_close_allocations
            WHERE close_identity_id IN (
                SELECT identity_id FROM trading_close_facts WHERE batch_id = ?
            )
            """,
            (batch_id,),
        )

        close_trade_links = 0
        fact_allocations = 0
        pending = 0
        close_trade_rows = []
        close_fee_updates = []
        fact_allocation_rows = []
        remaining_close_trades = {row["id"]: float(row["quantity"]) for row in trades if row["open_close"] == "平仓"}
        remaining_open_trades = {row["id"]: float(row["quantity"]) for row in trades if row["open_close"] == "开仓"}
        for close in closes:
            matching_closes = [
                row for row in trades
                if row["open_close"] == "平仓"
                and row["trade_date"] == close["close_date"]
                and row["contract"] == close["contract"]
                and row["side"] == close["close_side"]
                and abs(float(row["price"]) - float(close["close_price"])) < 1e-8
                and remaining_close_trades.get(row["id"], 0) > 0
            ]
            needed = float(close["quantity"])
            if sum(remaining_close_trades[row["id"]] for row in matching_closes) + 1e-9 < needed:
                pending += 1
                continue
            allocated_fee = 0.0
            for trade in matching_closes:
                quantity = min(needed, remaining_close_trades[trade["id"]])
                if quantity <= 0:
                    continue
                fee = float(trade["fee"] or 0) * quantity / float(trade["quantity"])
                close_trade_rows.append((close["identity_id"], trade["identity_id"], quantity, fee))
                close_trade_links += 1
                allocated_fee += fee
                remaining_close_trades[trade["id"]] -= quantity
                needed -= quantity
                if needed <= 1e-9:
                    break
            close_fee_updates.append((allocated_fee, close["id"]))

            matching_opens = [
                row for row in trades
                if row["open_close"] == "开仓"
                and row["trade_date"] == close["open_date"]
                and row["contract"] == close["contract"]
                and row["side"] == close["open_side"]
                and abs(float(row["price"]) - float(close["open_price"])) < 1e-8
                and remaining_open_trades.get(row["id"], 0) > 0
            ]
            needed = float(close["quantity"])
            if sum(remaining_open_trades[row["id"]] for row in matching_opens) + 1e-9 < needed:
                pending += 1
                continue
            for trade in matching_opens:
                quantity = min(needed, remaining_open_trades[trade["id"]])
                if quantity <= 0:
                    continue
                fact_allocation_rows.append((close["identity_id"], trade["identity_id"], quantity))
                fact_allocations += 1
                remaining_open_trades[trade["id"]] -= quantity
                needed -= quantity
                if needed <= 1e-9:
                    break
        if close_trade_rows:
            db._executemany(
                cur,
                """
                INSERT INTO trading_close_trade_links
                    (close_identity_id, close_trade_identity_id, matched_quantity, allocated_fee, rule_version)
                VALUES (?, ?, ?, ?, 'wenhua-group-v1')
                """,
                close_trade_rows,
            )
        if close_fee_updates:
            db._executemany(
                cur,
                "UPDATE trading_close_facts SET matched_fee = ?, fee_status = 'matched' WHERE id = ?",
                close_fee_updates,
            )
        if fact_allocation_rows:
            db._executemany(
                cur,
                """
                INSERT INTO trading_fact_close_allocations
                    (close_identity_id, open_trade_identity_id, matched_quantity, match_rule_version)
                VALUES (?, ?, ?, 'wenhua-fifo-v1')
                """,
                fact_allocation_rows,
            )
        event_allocations, pending_events = _match_option_event_allocations(cur, batch_id)
        _link_option_event_underlying_trades(cur, batch_id)
        fact_allocations += event_allocations
        pending += pending_events
        conn.commit()
    return {
        "close_trade_links": close_trade_links,
        "fact_close_allocations": fact_allocations,
        "pending_closes": pending,
    }


@dataclass
class FactFilters:
    contract: str = ""
    direction: str = ""
    asset_type: str = ""
    open_close: str = ""
    start_date: str = ""
    end_date: str = ""
    classification: str = ""
    business_type: str = ""
    page: int = 1
    page_size: int = 20

    def __post_init__(self):
        if self.page_size not in {20, 50, 100}:
            raise ValueError("每页条数只允许 20、50、100")
        if self.business_type not in {"", *BUSINESS_TYPES}:
            raise ValueError("未知业务类型")
        self.page = max(1, self.page)


def _page_result(items: list[dict[str, Any]], summary: dict[str, Any], filters: FactFilters, data_status: str = "ok") -> dict[str, Any]:
    total = len(items)
    total_pages = max(1, (total + filters.page_size - 1) // filters.page_size)
    page = min(filters.page, total_pages)
    start = (page - 1) * filters.page_size
    return {
        "items": items[start:start + filters.page_size],
        "summary": summary,
        "page": page,
        "page_size": filters.page_size,
        "total_items": total,
        "total_pages": total_pages,
        "data_status": data_status,
    }


def _fact_page_meta(total: int, filters: FactFilters) -> tuple[int, int, int]:
    total_pages = max(1, (total + filters.page_size - 1) // filters.page_size)
    page = min(filters.page, total_pages)
    return page, total_pages, (page - 1) * filters.page_size


def _query_trade_rows_paged(cur, filters: FactFilters) -> dict[str, Any]:
    where = ["b.status = 'active'", "tf.is_current = 1"]
    params: list[Any] = []
    if filters.contract:
        where.append("LOWER(tf.contract) LIKE ?")
        params.append(f"%{filters.contract.lower()}%")
    if filters.direction:
        where.append("tf.side = ?")
        params.append(filters.direction)
    if filters.asset_type:
        where.append("tf.asset_type = ?")
        params.append(filters.asset_type)
    if filters.open_close:
        where.append("tf.open_close = ?")
        params.append(filters.open_close)
    if filters.start_date:
        where.append("tf.trade_date >= ?")
        params.append(filters.start_date)
    if filters.end_date:
        where.append("tf.trade_date <= ?")
        params.append(filters.end_date)
    if filters.classification == "classified":
        where.append("ba.id IS NOT NULL")
    elif filters.classification == "unclassified":
        where.append("ba.id IS NULL")
    where_sql = " AND ".join(where)
    close_pnl_cte = """
        WITH close_pnl AS (
            SELECT l.close_trade_identity_id,
                   SUM(cf.fact_close_pnl * l.matched_quantity / cf.quantity) AS fact_close_pnl
            FROM trading_close_trade_links l
            JOIN trading_close_facts cf ON cf.identity_id = l.close_identity_id
            JOIN trading_import_batches cb ON cb.id = cf.batch_id AND cb.status = 'active'
            WHERE cf.is_current = 1
            GROUP BY l.close_trade_identity_id
        )
    """
    summary_row = db._exec(
        cur,
        close_pnl_cte + f"""
        SELECT COUNT(*) AS record_count,
               COALESCE(SUM(tf.quantity), 0) AS quantity,
               COALESCE(SUM(tf.fee), 0) AS fee,
               COALESCE(SUM(cp.fact_close_pnl), 0) AS fact_close_pnl
        FROM trading_trade_facts tf
        JOIN trading_import_batches b ON b.id = tf.batch_id
        LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
        LEFT JOIN close_pnl cp ON cp.close_trade_identity_id = tf.identity_id
        WHERE {where_sql}
        """,
        tuple(params),
    ).fetchone()
    total = int(summary_row["record_count"] or 0)
    page, total_pages, offset = _fact_page_meta(total, filters)
    rows = db._exec(
        cur,
        close_pnl_cte + f"""
        SELECT tf.*, s.name AS business_subject, ba.business_type,
               st.name AS strategy,
               CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status,
               cp.fact_close_pnl
        FROM trading_trade_facts tf
        JOIN trading_import_batches b ON b.id = tf.batch_id
        LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
        LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
        LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
        LEFT JOIN close_pnl cp ON cp.close_trade_identity_id = tf.identity_id
        WHERE {where_sql}
        ORDER BY tf.trade_date DESC, tf.id DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [filters.page_size, offset]),
    ).fetchall()
    return {
        "items": [dict(row) for row in rows],
        "summary": {
            "record_count": total,
            "quantity": float(summary_row["quantity"] or 0),
            "fee": float(summary_row["fee"] or 0),
            "fact_close_pnl": float(summary_row["fact_close_pnl"] or 0),
        },
        "page": page, "page_size": filters.page_size,
        "total_items": total, "total_pages": total_pages, "data_status": "imported",
    }


def _query_close_rows_paged(cur, filters: FactFilters) -> dict[str, Any]:
    where = ["b.status = 'active'", "cf.is_current = 1"]
    params: list[Any] = []
    if filters.contract:
        where.append("LOWER(cf.contract) LIKE ?")
        params.append(f"%{filters.contract.lower()}%")
    if filters.direction:
        where.append("cf.open_side = ?")
        params.append(filters.direction)
    if filters.asset_type:
        where.append("cf.asset_type = ?")
        params.append(filters.asset_type)
    if filters.start_date:
        where.append("cf.close_date >= ?")
        params.append(filters.start_date)
    if filters.end_date:
        where.append("cf.close_date <= ?")
        params.append(filters.end_date)
    has_allocations = "EXISTS (SELECT 1 FROM trading_business_close_allocations a WHERE a.close_identity_id = cf.identity_id)"
    has_unclassified = "EXISTS (SELECT 1 FROM trading_business_close_allocations a LEFT JOIN trading_business_assignments xba ON xba.trade_identity_id = a.open_trade_identity_id WHERE a.close_identity_id = cf.identity_id AND xba.id IS NULL)"
    if filters.classification == "classified":
        where.extend([has_allocations, f"NOT {has_unclassified}"])
    elif filters.classification == "unclassified":
        where.append(f"(NOT {has_allocations} OR {has_unclassified})")
    where_sql = " AND ".join(where)
    summary_row = db._exec(
        cur,
        f"""
        SELECT COUNT(*) AS record_count,
               COUNT(CASE WHEN cf.settlement_type = 'trade_close' THEN 1 END)
                   AS trade_close_record_count,
               COALESCE(SUM(cf.quantity), 0) AS settlement_quantity,
               COALESCE(SUM(CASE WHEN cf.settlement_type = 'trade_close'
                                 THEN cf.quantity ELSE 0 END), 0)
                   AS transaction_close_quantity,
               COALESCE(SUM(cf.fact_close_pnl), 0) AS fact_close_pnl,
               COALESCE(SUM(cf.matched_fee), 0) AS fee
        FROM trading_close_facts cf
        JOIN trading_import_batches b ON b.id = cf.batch_id
        WHERE {where_sql}
        """,
        tuple(params),
    ).fetchone()
    total = int(summary_row["record_count"] or 0)
    page, total_pages, offset = _fact_page_meta(total, filters)
    rows = db._exec(
        cur,
        f"""
        SELECT cf.* FROM trading_close_facts cf
        JOIN trading_import_batches b ON b.id = cf.batch_id
        WHERE {where_sql}
        ORDER BY cf.close_date DESC, cf.id DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [filters.page_size, offset]),
    ).fetchall()
    items = [dict(row) for row in rows]
    close_ids = [int(item["identity_id"]) for item in items]
    if close_ids:
        placeholders = ",".join("?" for _ in close_ids)
        assignment_rows = db._exec(
            cur,
            f"""
            SELECT a.close_identity_id, ba.business_type, st.name AS strategy,
                   CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
            FROM trading_business_close_allocations a
            LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = a.open_trade_identity_id
            LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
            WHERE a.close_identity_id IN ({placeholders})
            """,
            tuple(close_ids),
        ).fetchall()
        assignments: dict[int, list[dict[str, Any]]] = {}
        for row in assignment_rows:
            item = dict(row)
            assignments.setdefault(int(item["close_identity_id"]), []).append(item)
        for item in items:
            related = assignments.get(int(item["identity_id"]), [])
            classified = [row for row in related if row["assignment_status"] == "classified"]
            item["assignment_status"] = "classified" if related and len(classified) == len(related) else "unclassified"
            item["business_type"] = classified[0]["business_type"] if classified and len({row["business_type"] for row in classified}) == 1 else None
            item["strategy"] = classified[0]["strategy"] if classified and len({row["strategy"] for row in classified}) == 1 else None
    return {
        "items": items,
        "summary": {
            "record_count": total,
            "trade_close_record_count": int(
                summary_row["trade_close_record_count"] or 0
            ),
            "quantity": float(summary_row["settlement_quantity"] or 0),
            "settlement_quantity": float(summary_row["settlement_quantity"] or 0),
            "transaction_close_quantity": float(
                summary_row["transaction_close_quantity"] or 0
            ),
            "fact_close_pnl": float(summary_row["fact_close_pnl"] or 0),
            "fee": float(summary_row["fee"] or 0),
        },
        "page": page, "page_size": filters.page_size,
        "total_items": total, "total_pages": total_pages, "data_status": "imported",
    }


def query_fact_rows(view: str, filters: FactFilters) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        if view == "trades":
            return _query_trade_rows_paged(cur, filters)
        if view == "closes":
            return _query_close_rows_paged(cur, filters)
        if view == "positions":
            snapshot_date = filters.end_date
            if not snapshot_date:
                row = db._exec(
                    cur,
                    """
                    SELECT MAX(ps.snapshot_date) AS d
                    FROM trading_position_snapshots ps
                    JOIN trading_import_batches b ON b.id = ps.batch_id
                    WHERE b.status = 'active' AND ps.is_current = 1
                    """,
                ).fetchone()
                snapshot_date = row["d"] if row else None
            rows = db._exec(
                cur,
                """
                SELECT ps.* FROM trading_position_snapshots ps
                JOIN trading_import_batches b ON b.id = ps.batch_id
                WHERE b.status = 'active' AND ps.is_current = 1 AND ps.snapshot_date = ?
                ORDER BY ps.contract, ps.direction, ps.id
                """,
                (snapshot_date,),
            ).fetchall() if snapshot_date else []
            grouped_positions: dict[tuple[str, str, str], dict[str, Any]] = {}
            for raw_row in rows:
                row = dict(raw_row)
                key = (row["contract"], row["direction"], row["asset_type"])
                group = grouped_positions.get(key)
                if not group:
                    group = dict(row)
                    group["quantity"] = 0.0
                    group["margin"] = 0.0
                    group["weighted_price"] = 0.0
                    group["source_record_count"] = 0
                    grouped_positions[key] = group
                quantity = float(row["quantity"] or 0)
                group["quantity"] += quantity
                group["margin"] += float(row["margin"] or 0)
                group["weighted_price"] += float(row["average_price"] or 0) * quantity
                group["source_record_count"] += 1
            items = list(grouped_positions.values())
            for item in items:
                item["average_price"] = item.pop("weighted_price") / item["quantity"] if item["quantity"] else 0
            assignment_rows = db._exec(
                cur,
                """
                SELECT tf.contract, tf.side AS direction, tf.asset_type,
                       ba.business_type, st.name AS strategy,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
                FROM trading_trade_facts tf
                JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                WHERE tf.is_current = 1 AND tf.open_close = '开仓'
                """,
            ).fetchall()
            assignments: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
            for assignment in assignment_rows:
                item = dict(assignment)
                assignments.setdefault((item["contract"], item["direction"], item["asset_type"]), []).append(item)
            for item in items:
                related = assignments.get((item["contract"], item["direction"], item["asset_type"]), [])
                classified = [row for row in related if row["assignment_status"] == "classified"]
                item["assignment_status"] = "classified" if related and len(classified) == len(related) else "unclassified"
                item["business_type"] = classified[0]["business_type"] if classified and len({row["business_type"] for row in classified}) == 1 else None
                item["strategy"] = classified[0]["strategy"] if classified and len({row["strategy"] for row in classified}) == 1 else None
            if filters.contract:
                items = [row for row in items if filters.contract.lower() in row["contract"].lower()]
            if filters.direction:
                items = [row for row in items if row["direction"] == filters.direction]
            if filters.asset_type:
                items = [row for row in items if row["asset_type"] == filters.asset_type]
            if filters.classification == "classified":
                items = [row for row in items if row["assignment_status"] == "classified"]
            elif filters.classification == "unclassified":
                items = [row for row in items if row["assignment_status"] == "unclassified"]
            summary = {
                "record_count": len(items),
                "source_record_count": sum(int(row.get("source_record_count") or 0) for row in items),
                "quantity": sum(float(row["quantity"]) for row in items),
                "margin": sum(float(row["margin"] or 0) for row in items),
            }
            status = "ok" if items else "no_position_snapshot"
            return _page_result(items, summary, filters, status)
        if view == "closes":
            rows = db._exec(
                cur,
                """
                SELECT cf.* FROM trading_close_facts cf
                JOIN trading_import_batches b ON b.id = cf.batch_id
                WHERE b.status = 'active' AND cf.is_current = 1
                ORDER BY cf.close_date DESC, cf.id DESC
                """,
            ).fetchall()
            items = [dict(row) for row in rows]
            allocation_rows = db._exec(
                cur,
                """
                SELECT a.close_identity_id, ba.business_type, st.name AS strategy,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
                FROM trading_business_close_allocations a
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = a.open_trade_identity_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                """,
            ).fetchall()
            close_assignments: dict[int, list[dict[str, Any]]] = {}
            for allocation in allocation_rows:
                allocation = dict(allocation)
                close_assignments.setdefault(allocation["close_identity_id"], []).append(allocation)
            for item in items:
                related = close_assignments.get(item["identity_id"], [])
                classified = [row for row in related if row["assignment_status"] == "classified"]
                item["assignment_status"] = "classified" if related and len(classified) == len(related) else "unclassified"
                item["business_type"] = classified[0]["business_type"] if classified and len({row["business_type"] for row in classified}) == 1 else None
                item["strategy"] = classified[0]["strategy"] if classified and len({row["strategy"] for row in classified}) == 1 else None
            if filters.contract:
                items = [row for row in items if filters.contract.lower() in row["contract"].lower()]
            if filters.direction:
                items = [row for row in items if row["open_side"] == filters.direction]
            if filters.asset_type:
                items = [row for row in items if row["asset_type"] == filters.asset_type]
            if filters.start_date:
                items = [row for row in items if row["close_date"] >= filters.start_date]
            if filters.end_date:
                items = [row for row in items if row["close_date"] <= filters.end_date]
            if filters.classification == "classified":
                items = [row for row in items if row["assignment_status"] == "classified"]
            elif filters.classification == "unclassified":
                items = [row for row in items if row["assignment_status"] == "unclassified"]
            summary = {
                "record_count": len(items),
                "quantity": sum(float(row["quantity"]) for row in items),
                "fact_close_pnl": sum(float(row["fact_close_pnl"] or 0) for row in items),
                "fee": sum(float(row["matched_fee"] or 0) for row in items),
            }
            return _page_result(items, summary, filters)
        if view != "trades":
            raise ValueError(f"未知事实视图：{view}")
        rows = db._exec(
            cur,
            """
            SELECT tf.*, s.name AS business_subject, ba.business_type,
                   st.name AS strategy,
                   CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status,
                   (SELECT SUM(cf.fact_close_pnl * l.matched_quantity / cf.quantity)
                    FROM trading_close_trade_links l
                    JOIN trading_close_facts cf ON cf.identity_id = l.close_identity_id
                    JOIN trading_import_batches cb ON cb.id = cf.batch_id AND cb.status = 'active'
                    WHERE cf.is_current = 1 AND l.close_trade_identity_id = tf.identity_id) AS fact_close_pnl
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id
            LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
            LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
            LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
            WHERE b.status = 'active' AND tf.is_current = 1
            ORDER BY tf.trade_date DESC, tf.id DESC
            """,
        ).fetchall()
        items = [dict(row) for row in rows]
        if filters.contract:
            items = [row for row in items if filters.contract.lower() in row["contract"].lower()]
        if filters.direction:
            items = [row for row in items if row["side"] == filters.direction]
        if filters.asset_type:
            items = [row for row in items if row["asset_type"] == filters.asset_type]
        if filters.open_close:
            items = [row for row in items if row["open_close"] == filters.open_close]
        if filters.start_date:
            items = [row for row in items if row["trade_date"] >= filters.start_date]
        if filters.end_date:
            items = [row for row in items if row["trade_date"] <= filters.end_date]
        if filters.classification == "classified":
            items = [row for row in items if row["assignment_status"] == "classified"]
        elif filters.classification == "unclassified":
            items = [row for row in items if row["assignment_status"] == "unclassified"]
        summary = {
            "record_count": len(items),
            "quantity": sum(float(row["quantity"]) for row in items),
            "fee": sum(float(row["fee"] or 0) for row in items),
            "fact_close_pnl": sum(float(row["fact_close_pnl"] or 0) for row in items),
        }
        return _page_result(items, summary, filters)


def query_trade_selection_identities(filters: FactFilters) -> dict[str, Any]:
    """返回当前筛选口径下的全部成交标识，避免前端逐页拉取完整行数据。"""
    with db.connect() as conn:
        rows = db._exec(
            conn.cursor(),
            """
            SELECT tf.identity_id, tf.contract, tf.side, tf.asset_type,
                   tf.open_close, tf.trade_date,
                   CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
            WHERE tf.is_current = 1
            ORDER BY tf.trade_date DESC, tf.id DESC
            """,
        ).fetchall()
    items = [dict(row) for row in rows]
    if filters.contract:
        items = [row for row in items if filters.contract.lower() in row["contract"].lower()]
    if filters.direction:
        items = [row for row in items if row["side"] == filters.direction]
    if filters.asset_type:
        items = [row for row in items if row["asset_type"] == filters.asset_type]
    if filters.open_close:
        items = [row for row in items if row["open_close"] == filters.open_close]
    if filters.start_date:
        items = [row for row in items if row["trade_date"] >= filters.start_date]
    if filters.end_date:
        items = [row for row in items if row["trade_date"] <= filters.end_date]
    if filters.classification in {"classified", "unclassified"}:
        items = [row for row in items if row["assignment_status"] == filters.classification]
    identity_ids = [int(row["identity_id"]) for row in items]
    return {"identity_ids": identity_ids, "total_items": len(identity_ids)}


def _build_business_type_overview(filters: FactFilters) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        common_params = (
            filters.business_type,
            filters.contract, f"%{filters.contract}%",
            filters.direction, filters.direction,
            filters.asset_type, filters.asset_type,
        )
        trade_row = db._exec(
            cur,
            """
            SELECT COUNT(*) AS record_count,
                   COALESCE(SUM(tf.quantity), 0) AS quantity,
                   COALESCE(SUM(tf.fee), 0) AS fee
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
            WHERE tf.is_current = 1 AND ba.business_type = ?
              AND (? = '' OR LOWER(tf.contract) LIKE LOWER(?))
              AND (? = '' OR tf.side = ?)
              AND (? = '' OR tf.asset_type = ?)
              AND (? = '' OR tf.trade_date >= ?)
              AND (? = '' OR tf.trade_date <= ?)
            """,
            common_params + (
                filters.start_date, filters.start_date,
                filters.end_date, filters.end_date,
            ),
        ).fetchone()
        close_rows = db._exec(
            cur,
            """
            SELECT cf.close_date,
                   a.matched_quantity,
                   cf.fact_close_pnl * a.matched_quantity / NULLIF(cf.quantity, 0)
                       AS allocated_fact_close_pnl,
                   COALESCE(cf.matched_fee, 0) * a.matched_quantity / NULLIF(cf.quantity, 0)
                       AS allocated_fee
            FROM trading_close_facts cf
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            JOIN trading_business_close_allocations a ON a.close_identity_id = cf.identity_id
            JOIN trading_business_assignments ba ON ba.trade_identity_id = a.open_trade_identity_id
            WHERE cf.is_current = 1 AND ba.business_type = ?
              AND (? = '' OR LOWER(cf.contract) LIKE LOWER(?))
              AND (? = '' OR cf.open_side = ?)
              AND (? = '' OR cf.asset_type = ?)
              AND (? = '' OR cf.close_date >= ?)
              AND (? = '' OR cf.close_date <= ?)
            ORDER BY cf.close_date
            """,
            common_params + (
                filters.start_date, filters.start_date,
                filters.end_date, filters.end_date,
            ),
        ).fetchall()
        snapshot_row = db._exec(
            cur,
            """
            SELECT MAX(ps.snapshot_date) AS snapshot_date
            FROM trading_position_snapshots ps
            JOIN trading_import_batches b ON b.id = ps.batch_id AND b.status = 'active'
            WHERE ps.is_current = 1 AND (? = '' OR ps.snapshot_date <= ?)
            """,
            (filters.end_date, filters.end_date),
        ).fetchone()
        snapshot_date = snapshot_row["snapshot_date"] if snapshot_row else None
        snapshot_rows = db._exec(
            cur,
            """
            SELECT ps.contract, ps.direction, ps.asset_type,
                   SUM(ps.quantity) AS quantity, SUM(COALESCE(ps.margin, 0)) AS margin,
                   COUNT(*) AS source_record_count
            FROM trading_position_snapshots ps
            JOIN trading_import_batches b ON b.id = ps.batch_id AND b.status = 'active'
            WHERE ps.is_current = 1 AND ps.snapshot_date = ?
            GROUP BY ps.contract, ps.direction, ps.asset_type
            """,
            (snapshot_date,),
        ).fetchall() if snapshot_date else []
        open_rows = db._exec(
            cur,
            """
            SELECT tf.contract, tf.side AS direction, tf.asset_type,
                   tf.quantity - COALESCE((
                       SELECT SUM(a.matched_quantity)
                       FROM trading_business_close_allocations a
                       JOIN trading_close_facts cf ON cf.identity_id = a.close_identity_id
                       WHERE a.open_trade_identity_id = tf.identity_id
                         AND (? = '' OR cf.close_date <= ?)
                   ), 0) AS remaining_quantity
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
            WHERE tf.is_current = 1 AND tf.open_close = '开仓'
              AND ba.business_type = ?
              AND (? = '' OR LOWER(tf.contract) LIKE LOWER(?))
              AND (? = '' OR tf.side = ?)
              AND (? = '' OR tf.asset_type = ?)
              AND (? = '' OR tf.trade_date <= ?)
            """,
            (
                filters.end_date, filters.end_date,
                filters.business_type,
                filters.contract, f"%{filters.contract}%",
                filters.direction, filters.direction,
                filters.asset_type, filters.asset_type,
                filters.end_date, filters.end_date,
            ),
        ).fetchall()
    position_groups: dict[tuple[str, str, str], dict[str, float]] = {}
    for raw in open_rows:
        remaining = float(raw["remaining_quantity"] or 0)
        if remaining <= 1e-9:
            continue
        key = (raw["contract"], raw["direction"], raw["asset_type"])
        group = position_groups.setdefault(key, {"quantity": 0.0, "source_record_count": 0.0})
        group["quantity"] += remaining
        group["source_record_count"] += 1
    snapshot_by_key = {
        (row["contract"], row["direction"], row["asset_type"]): row
        for row in snapshot_rows
    }
    margin = 0.0
    for key, group in position_groups.items():
        snapshot = snapshot_by_key.get(key)
        if snapshot and float(snapshot["quantity"] or 0) > 0:
            margin += (
                float(snapshot["margin"] or 0)
                * group["quantity"]
                / float(snapshot["quantity"])
            )
    daily: dict[str, float] = {}
    for row in close_rows:
        daily[row["close_date"]] = daily.get(row["close_date"], 0.0) + float(
            row["allocated_fact_close_pnl"] or 0
        )
    close_pnl = sum(daily.values())
    return {
        "trades": {
            "record_count": int(trade_row["record_count"] or 0),
            "quantity": float(trade_row["quantity"] or 0),
            "fee": float(trade_row["fee"] or 0),
            "fact_close_pnl": close_pnl,
        },
        "closes": {
            "record_count": len(close_rows),
            "quantity": sum(float(row["matched_quantity"] or 0) for row in close_rows),
            "settlement_quantity": sum(float(row["matched_quantity"] or 0) for row in close_rows),
            "transaction_close_quantity": sum(float(row["matched_quantity"] or 0) for row in close_rows),
            "fact_close_pnl": close_pnl,
            "fee": sum(float(row["allocated_fee"] or 0) for row in close_rows),
        },
        "positions": {
            "record_count": int(sum(row["source_record_count"] for row in position_groups.values())),
            "group_count": len(position_groups),
            "quantity": sum(row["quantity"] for row in position_groups.values()),
            "margin": margin,
            "snapshot_date": snapshot_date,
            "floating_pnl": None,
            "floating_pnl_status": "pending_calculation",
        },
        "data_status": {
            "fact": "file_imported",
            "positions": "ok" if snapshot_date else "no_position_snapshot",
        },
        "daily_close_pnl": [
            {"date": close_date, "fact_close_pnl": value}
            for close_date, value in sorted(daily.items())
        ],
    }


def build_overview(filters: FactFilters) -> dict[str, Any]:
    if filters.business_type:
        return _build_business_type_overview(filters)
    trades = query_fact_rows("trades", FactFilters(
        contract=filters.contract,
        direction=filters.direction,
        asset_type=filters.asset_type,
        open_close=filters.open_close,
        start_date=filters.start_date,
        end_date=filters.end_date,
        page=1,
        page_size=100,
    ))
    closes = query_fact_rows("closes", FactFilters(
        contract=filters.contract,
        direction=filters.direction,
        asset_type=filters.asset_type,
        start_date=filters.start_date,
        end_date=filters.end_date,
        page=1,
        page_size=100,
    ))
    positions = query_fact_rows("positions", FactFilters(
        contract=filters.contract,
        direction=filters.direction,
        asset_type=filters.asset_type,
        end_date=filters.end_date,
        page=1,
        page_size=100,
    ))
    with db.connect() as conn:
        daily_rows = db._exec(
            conn.cursor(),
            """
            SELECT cf.close_date AS date, SUM(cf.fact_close_pnl) AS fact_close_pnl
            FROM trading_close_facts cf
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            WHERE cf.is_current = 1
              AND (? = '' OR LOWER(cf.contract) LIKE LOWER(?))
              AND (? = '' OR cf.open_side = ?)
              AND (? = '' OR cf.asset_type = ?)
              AND (? = '' OR cf.close_date >= ?)
              AND (? = '' OR cf.close_date <= ?)
            GROUP BY cf.close_date
            ORDER BY cf.close_date
            """,
            (
                filters.contract, f"%{filters.contract}%", filters.direction, filters.direction,
                filters.asset_type, filters.asset_type, filters.start_date, filters.start_date,
                filters.end_date, filters.end_date,
            ),
        ).fetchall()
    snapshot_date = positions["items"][0]["snapshot_date"] if positions["items"] else None
    return {
        "trades": trades["summary"],
        "closes": closes["summary"],
        "positions": {
            **positions["summary"],
            "record_count": positions["summary"].get("source_record_count", positions["summary"]["record_count"]),
            "group_count": positions["summary"]["record_count"],
            "snapshot_date": snapshot_date,
            "floating_pnl": None,
            "floating_pnl_status": "pending_calculation",
        },
        "data_status": {
            "fact": "file_imported",
            "positions": positions["data_status"],
        },
        "daily_close_pnl": [
            {"date": row["date"], "fact_close_pnl": float(row["fact_close_pnl"] or 0)}
            for row in daily_rows
        ],
    }


BUSINESS_TYPES = ("basic_hedging", "strategic_hedging")
_REMATCH_PREVIEWS: dict[str, dict[str, Any]] = {}


class TradingUploadFile(BaseModel):
    name: str
    content_base64: str


class TradingImportPreviewIn(BaseModel):
    account_id: int
    statement_file: TradingUploadFile


class BusinessSubjectIn(BaseModel):
    name: str


class StrategyIn(BaseModel):
    name: str


class BusinessAssignmentIn(BaseModel):
    identity_ids: list[int]
    business_subject_id: int
    business_type: str
    strategy_name: str = ""
    instruction_text: str = ""


class RematchSelectionIn(BaseModel):
    open_trade_identity_id: int
    quantity: float


class RematchPreviewIn(BaseModel):
    allocation_version: int
    selections: list[RematchSelectionIn]


class RematchConfirmIn(BaseModel):
    preview_token: str
    allocation_version: int
    reason: str = ""


class RestoreDefaultIn(BaseModel):
    allocation_version: int
    reason: str = "恢复事实层默认开平关系"


def _normalized_name(value: str) -> str:
    return " ".join(str(value or "").strip().split()).lower()


def create_business_subject(name: str, actor: str) -> dict[str, Any]:
    display_name = " ".join(str(name or "").strip().split())
    normalized = _normalized_name(display_name)
    if not normalized:
        raise ValueError("业务归属名称不能为空")
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM trading_business_subjects WHERE normalized_name = ?", (normalized,)).fetchone()
        if row:
            return dict(row)
        subject_id = db._last_insert_id(
            cur,
            "INSERT INTO trading_business_subjects (name, normalized_name, created_by, updated_by) VALUES (?, ?, ?, ?)",
            (display_name, normalized, actor, actor),
        )
        conn.commit()
        row = db._exec(cur, "SELECT * FROM trading_business_subjects WHERE id = ?", (subject_id,)).fetchone()
    return dict(row)


def get_or_create_strategy(name: str, actor: str) -> Optional[dict[str, Any]]:
    display_name = " ".join(str(name or "").strip().split())
    if not display_name:
        return None
    normalized = _normalized_name(display_name)
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM trading_strategies WHERE normalized_name = ?", (normalized,)).fetchone()
        if row:
            return dict(row)
        strategy_id = db._last_insert_id(
            cur,
            "INSERT INTO trading_strategies (name, normalized_name, source, created_by, updated_by) VALUES (?, ?, 'manual', ?, ?)",
            (display_name, normalized, actor, actor),
        )
        conn.commit()
        row = db._exec(cur, "SELECT * FROM trading_strategies WHERE id = ?", (strategy_id,)).fetchone()
    return dict(row)


def list_trading_config() -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        subjects = db._exec(cur, "SELECT * FROM trading_business_subjects ORDER BY name").fetchall()
        strategies = db._exec(cur, "SELECT * FROM trading_strategies ORDER BY name").fetchall()
        accounts = db._exec(cur, "SELECT * FROM trading_accounts WHERE is_active = 1 ORDER BY display_name").fetchall()
    return {
        "business_types": list(BUSINESS_TYPES),
        "subjects": [dict(row) for row in subjects],
        "strategies": [dict(row) for row in strategies],
        "accounts": [dict(row) for row in accounts],
        "junneng_candidate_products": ["rb", "hc"],
    }


def _write_assignment_audit(cur, identity_id: int, operation_type: str, actor: str, before: Any, after: Any) -> None:
    db._exec(
        cur,
        """
        INSERT INTO operation_logs
            (module_code, entity_type, entity_id, operation_type, description, before_data, after_data)
        VALUES ('trading_positions', 'trading_business_assignment', ?, ?, ?, ?, ?)
        """,
        (
            identity_id,
            operation_type,
            f"{actor} {operation_type}",
            json.dumps(before, ensure_ascii=False) if before is not None else None,
            json.dumps(after, ensure_ascii=False) if after is not None else None,
        ),
    )


def classify_trade_identities(
    identity_ids: list[int],
    business_subject_id: int,
    business_type: str,
    strategy_name: str,
    instruction_text: str,
    actor: str,
    requested_quantity: Optional[float] = None,
) -> dict[str, Any]:
    if requested_quantity is not None:
        raise ValueError("一笔成交不允许按手数拆分归类")
    if business_type not in BUSINESS_TYPES:
        raise ValueError("业务类型只允许基础套保或战略套保")
    if not identity_ids:
        raise ValueError("请选择需要归类的完整成交")
    strategy = get_or_create_strategy(strategy_name, actor)
    identity_ids = list(dict.fromkeys(int(identity_id) for identity_id in identity_ids))
    with db.connect() as conn:
        cur = conn.cursor()
        subject = db._exec(
            cur,
            "SELECT id FROM trading_business_subjects WHERE id = ? AND is_active = 1",
            (business_subject_id,),
        ).fetchone()
        if not subject:
            raise ValueError("业务归属不存在或已停用")
        active_ids: set[int] = set()
        for start in range(0, len(identity_ids), 500):
            chunk = identity_ids[start:start + 500]
            placeholders = ",".join("?" for _ in chunk)
            active_rows = db._exec(
                cur,
                f"""
                SELECT tf.identity_id FROM trading_trade_facts tf
                JOIN trading_import_batches b ON b.id = tf.batch_id
                WHERE b.status = 'active' AND tf.is_current = 1 AND tf.identity_id IN ({placeholders})
                """,
                tuple(chunk),
            ).fetchall()
            active_ids.update(int(row["identity_id"]) for row in active_rows)
        missing_ids = [identity_id for identity_id in identity_ids if identity_id not in active_ids]
        if missing_ids:
            raise ValueError(f"成交事实不存在或不是有效版本：{missing_ids[0]}")
        strategy_id = strategy["id"] if strategy else None

        def fetch_assignments(ids: list[int]) -> dict[int, dict[str, Any]]:
            result: dict[int, dict[str, Any]] = {}
            for start in range(0, len(ids), 500):
                chunk = ids[start:start + 500]
                if not chunk:
                    continue
                placeholders = ",".join("?" for _ in chunk)
                rows = db._exec(
                    cur,
                    f"SELECT * FROM trading_business_assignments WHERE trade_identity_id IN ({placeholders})",
                    tuple(chunk),
                ).fetchall()
                result.update({int(row["trade_identity_id"]): dict(row) for row in rows})
            return result

        before = fetch_assignments(identity_ids)
        linked_close_ids: set[int] = set()
        for start in range(0, len(identity_ids), 500):
            chunk = identity_ids[start:start + 500]
            placeholders = ",".join("?" for _ in chunk)
            rows = db._exec(
                cur,
                f"SELECT DISTINCT close_identity_id FROM trading_fact_close_allocations WHERE open_trade_identity_id IN ({placeholders})",
                tuple(chunk),
            ).fetchall()
            linked_close_ids.update(int(row["close_identity_id"]) for row in rows)
        inherited_ids: set[int] = set()
        linked_close_list = sorted(linked_close_ids)
        for start in range(0, len(linked_close_list), 500):
            chunk = linked_close_list[start:start + 500]
            placeholders = ",".join("?" for _ in chunk)
            rows = db._exec(
                cur,
                f"SELECT DISTINCT close_trade_identity_id FROM trading_close_trade_links WHERE close_identity_id IN ({placeholders})",
                tuple(chunk),
            ).fetchall()
            inherited_ids.update(int(row["close_trade_identity_id"]) for row in rows)
        inherited_ids.difference_update(identity_ids)
        inherited_list = sorted(inherited_ids)
        before.update(fetch_assignments(inherited_list))

        db._executemany(
            cur,
            """
            INSERT INTO trading_business_assignments
                (trade_identity_id, business_subject_id, business_type, strategy_id,
                 instruction_text, assigned_by, updated_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(trade_identity_id) DO UPDATE SET
                business_subject_id = excluded.business_subject_id,
                business_type = excluded.business_type,
                strategy_id = excluded.strategy_id,
                instruction_text = excluded.instruction_text,
                updated_by = excluded.updated_by,
                updated_at = CURRENT_TIMESTAMP
            """,
            [(identity_id, business_subject_id, business_type, strategy_id, instruction_text or None, actor, actor)
             for identity_id in identity_ids],
        )
        if inherited_list:
            db._executemany(
                cur,
                """
                INSERT INTO trading_business_assignments
                    (trade_identity_id, business_subject_id, business_type, strategy_id, assigned_by, updated_by)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_identity_id) DO UPDATE SET
                    business_subject_id = excluded.business_subject_id,
                    business_type = excluded.business_type,
                    strategy_id = excluded.strategy_id,
                    updated_by = excluded.updated_by,
                    updated_at = CURRENT_TIMESTAMP
                """,
                [(identity_id, business_subject_id, business_type, strategy_id, actor, actor)
                 for identity_id in inherited_list],
            )
        all_changed_ids = identity_ids + inherited_list
        after = fetch_assignments(all_changed_ids)
        audit_rows = []
        explicitly_assigned_ids = set(identity_ids)
        for identity_id in all_changed_ids:
            operation_type = "业务归类" if identity_id in explicitly_assigned_ids else "业务平仓自动继承"
            audit_rows.append((
                identity_id, operation_type, f"{actor} {operation_type}",
                json.dumps(before.get(identity_id), ensure_ascii=False) if before.get(identity_id) is not None else None,
                json.dumps(after[identity_id], ensure_ascii=False),
            ))
        db._executemany(
            cur,
            """
            INSERT INTO operation_logs
                (module_code, entity_type, entity_id, operation_type, description, before_data, after_data)
            VALUES ('trading_positions', 'trading_business_assignment', ?, ?, ?, ?, ?)
            """,
            audit_rows,
        )
        conn.commit()
    return {"assigned_count": len(identity_ids)}


def remove_trade_assignment(identity_id: int, actor: str) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        before_row = db._exec(
            cur,
            "SELECT * FROM trading_business_assignments WHERE trade_identity_id = ?",
            (identity_id,),
        ).fetchone()
        if not before_row:
            return {"removed": False}
        db._exec(cur, "DELETE FROM trading_business_assignments WHERE trade_identity_id = ?", (identity_id,))
        _write_assignment_audit(cur, identity_id, "取消业务归类", actor, dict(before_row), None)
        conn.commit()
    return {"removed": True}


def calculate_business_pnl(
    open_price: float,
    close_price: float,
    side: str,
    quantity: float,
    multiplier: float,
) -> float:
    difference = close_price - open_price if side == "买" else open_price - close_price
    return round(difference * quantity * multiplier, 8)


def _product_code(contract: str) -> str:
    match = re.match(r"[a-zA-Z]+", contract or "")
    return match.group(0).lower() if match else ""


def rebuild_default_business_allocations(batch_id: int) -> dict[str, int]:
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(
            cur,
            """
            DELETE FROM trading_business_close_allocations
            WHERE source = 'fact_default' AND close_identity_id IN (
                SELECT identity_id FROM trading_close_facts WHERE batch_id = ?
            )
            """,
            (batch_id,),
        )
        fact_rows = db._exec(
            cur,
            """
            SELECT fa.close_identity_id, fa.open_trade_identity_id, fa.matched_quantity,
                   cf.open_price, cf.close_price, cf.open_side, cf.exchange, cf.contract,
                   cf.asset_type, cf.settlement_type, ot.price AS allocated_open_price
            FROM trading_fact_close_allocations fa
            JOIN trading_close_facts cf ON cf.identity_id = fa.close_identity_id
                AND cf.batch_id = ? AND cf.is_current = 1
            JOIN trading_trade_facts ot ON ot.identity_id = fa.open_trade_identity_id
                AND ot.is_current = 1
            ORDER BY fa.id
            """,
            (batch_id,),
        ).fetchall()
        specs = db._exec(
            cur,
            """
            SELECT exchange, product_code, asset_type, contract_multiplier
            FROM trading_contract_specs WHERE is_active = 1
            """,
        ).fetchall()
        spec_by_contract = {
            (row["exchange"].lower(), row["product_code"].lower(), row["asset_type"]): float(row["contract_multiplier"])
            for row in specs
        }
        count = 0
        allocation_rows = []
        for row in fact_rows:
            multiplier = spec_by_contract.get(
                (row["exchange"].lower(), _product_code(row["contract"]), row["asset_type"])
            )
            business_pnl = None
            if multiplier is not None:
                open_price = (
                    float(row["allocated_open_price"])
                    if row["settlement_type"] != "trade_close"
                    else float(row["open_price"])
                )
                business_pnl = calculate_business_pnl(
                    open_price, float(row["close_price"]), row["open_side"],
                    float(row["matched_quantity"]), multiplier,
                )
            allocation_rows.append(
                (row["close_identity_id"], row["open_trade_identity_id"], row["matched_quantity"], business_pnl)
            )
            count += 1
        if allocation_rows:
            db._executemany(
                cur,
                """
                INSERT INTO trading_business_close_allocations
                    (close_identity_id, open_trade_identity_id, matched_quantity, source,
                     business_pnl, rule_version)
                VALUES (?, ?, ?, 'fact_default', ?, 'business-default-v1')
                """,
                allocation_rows,
            )
        conn.commit()
    return {"allocation_count": count}


def _filter_business_items(items: list[dict[str, Any]], tab: str, filters: FactFilters) -> list[dict[str, Any]]:
    if filters.contract:
        items = [row for row in items if filters.contract.lower() in row["contract"].lower()]
    direction_key = "direction" if tab == "positions" else "open_side" if tab == "closes" else "side"
    date_key = "close_date" if tab == "closes" else "trade_date" if tab == "trades" else None
    if filters.direction:
        items = [row for row in items if row.get(direction_key) == filters.direction]
    if filters.classification == "classified":
        items = [row for row in items if row.get("assignment_status") == "classified"]
    elif filters.classification == "unclassified":
        items = [row for row in items if row.get("assignment_status") == "unclassified"]
    if date_key and filters.start_date:
        items = [row for row in items if row.get(date_key, "") >= filters.start_date]
    if date_key and filters.end_date:
        items = [row for row in items if row.get(date_key, "") <= filters.end_date]
    return items


def _aggregate_business_positions(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in items:
        quantity = float(row.get("remaining_quantity") or 0)
        key = (
            row["contract"], row["direction"], row["asset_type"], row.get("business_subject"),
            row.get("business_type"), row.get("strategy"), row.get("assignment_status"),
            row.get("ledger_membership"),
        )
        group = grouped.get(key)
        if not group:
            group = dict(row)
            group["quantity"] = 0.0
            group["weighted_price"] = 0.0
            group["source_record_count"] = 0
            grouped[key] = group
        group["quantity"] += quantity
        group["weighted_price"] += float(row.get("average_price") or 0) * quantity
        group["source_record_count"] += 1
    result = list(grouped.values())
    for row in result:
        row["average_price"] = row.pop("weighted_price") / row["quantity"] if row["quantity"] else 0
        row.pop("remaining_quantity", None)
    return result


def query_business_rows(view: str, tab: str, filters: FactFilters) -> dict[str, Any]:
    if view not in {"junneng", "options"} or tab not in {"positions", "closes", "trades"}:
        raise ValueError("未知业务视图")
    if tab == "positions":
        with db.connect() as conn:
            rows = db._exec(
                conn.cursor(),
                """
                SELECT tf.identity_id, tf.contract, tf.asset_type, tf.side AS direction,
                       tf.price AS average_price, tf.quantity,
                       tf.quantity - COALESCE((
                           SELECT SUM(a.matched_quantity)
                           FROM trading_business_close_allocations a
                           WHERE a.open_trade_identity_id = tf.identity_id
                       ), 0) AS remaining_quantity,
                       s.name AS business_subject, ba.business_type, st.name AS strategy,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
                FROM trading_trade_facts tf
                JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
                LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                WHERE tf.is_current = 1 AND tf.open_close = '开仓'
                ORDER BY tf.contract, tf.id
                """,
            ).fetchall()
        all_items = [
            dict(row) for row in rows
            if float(row["remaining_quantity"] or 0) > 1e-9
            and row["assignment_status"] == "classified"
        ]
        if view == "junneng":
            items = [
                {**row, "ledger_membership": "confirmed"}
                for row in all_items if row["business_subject"] == "上海钧能"
            ]
        else:
            items = [row for row in all_items if row["asset_type"] == "option"]
        items = _aggregate_business_positions(items)
        items = _filter_business_items(items, tab, filters)
        for item in items:
            item["floating_pnl"] = None
            item["floating_pnl_status"] = "pending_calculation"
        summary = {
            "record_count": len(items),
            "quantity": sum(float(row["quantity"]) for row in items),
            "business_pnl": 0,
            "floating_pnl": None,
            "floating_pnl_status": "pending_calculation",
        }
        return _page_result(items, summary, filters)

    with db.connect() as conn:
        cur = conn.cursor()
        if tab == "trades":
            rows = db._exec(
                cur,
                """
                SELECT tf.*, s.name AS business_subject, ba.business_type,
                       st.name AS strategy, ba.instruction_text,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status,
                       (SELECT SUM(CASE WHEN a.source = 'fact_default'
                                           THEN cf.fact_close_pnl * a.matched_quantity / NULLIF(cf.quantity, 0)
                                           ELSE a.business_pnl END)
                        FROM trading_close_trade_links l
                        JOIN trading_close_facts cf ON cf.identity_id = l.close_identity_id
                        JOIN trading_import_batches cb ON cb.id = cf.batch_id AND cb.status = 'active'
                        LEFT JOIN trading_business_close_allocations a ON a.close_identity_id = cf.identity_id
                        WHERE cf.is_current = 1 AND l.close_trade_identity_id = tf.identity_id) AS business_pnl
                FROM trading_trade_facts tf
                JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
                LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                WHERE tf.is_current = 1
                ORDER BY tf.trade_date DESC, tf.id DESC
                """,
            ).fetchall()
            all_items = [
                dict(row) for row in rows
                if row["assignment_status"] == "classified"
            ]
            if view == "junneng":
                items = [
                    {**row, "ledger_membership": "confirmed"}
                    for row in all_items if row["business_subject"] == "上海钧能"
                ]
            else:
                items = [row for row in all_items if row["asset_type"] == "option"]
            items = _filter_business_items(items, tab, filters)
            summary = {
                "record_count": len(items),
                "quantity": sum(float(row["quantity"]) for row in items),
                "fee": sum(float(row["fee"] or 0) for row in items),
                "business_pnl": sum(float(row["business_pnl"] or 0) for row in items),
            }
            return _page_result(items, summary, filters)
        if view == "options":
            rows = db._exec(
                cur,
                """
                SELECT cf.*,
                       CASE WHEN a.source IS NULL OR a.source = 'fact_default'
                            THEN cf.fact_close_pnl * COALESCE(a.matched_quantity, cf.quantity) / NULLIF(cf.quantity, 0)
                            ELSE a.business_pnl END AS business_pnl,
                       COALESCE(a.matched_quantity, cf.quantity) AS matched_quantity,
                       COALESCE(a.allocation_version, 1) AS allocation_version,
                       a.source AS allocation_source, s.name AS business_subject,
                       ba.business_type, st.name AS strategy,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
                FROM trading_close_facts cf
                JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
                LEFT JOIN trading_business_close_allocations a ON a.close_identity_id = cf.identity_id
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = a.open_trade_identity_id
                LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                WHERE cf.is_current = 1 AND cf.asset_type = 'option'
                ORDER BY cf.close_date DESC, cf.id DESC
                """,
            ).fetchall()
        else:
            rows = db._exec(
                cur,
                """
                SELECT cf.*,
                       CASE WHEN a.source IS NULL OR a.source = 'fact_default'
                            THEN cf.fact_close_pnl * COALESCE(a.matched_quantity, cf.quantity) / NULLIF(cf.quantity, 0)
                            ELSE a.business_pnl END AS business_pnl,
                       COALESCE(a.matched_quantity, cf.quantity) AS matched_quantity,
                       COALESCE(a.allocation_version, 1) AS allocation_version,
                       a.source AS allocation_source,
                       a.business_pnl AS allocation_business_pnl,
                       ot.trade_date AS allocated_open_date,
                       ot.price AS allocated_open_price,
                       ot.quantity AS open_trade_quantity,
                       ot.fee AS open_trade_fee,
                       s.name AS business_subject, ba.business_type, st.name AS strategy,
                       CASE WHEN ba.id IS NULL THEN 'unclassified' ELSE 'classified' END AS assignment_status
                FROM trading_close_facts cf
                JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
                LEFT JOIN trading_business_close_allocations a ON a.close_identity_id = cf.identity_id
                LEFT JOIN trading_trade_facts ot ON ot.identity_id = a.open_trade_identity_id
                    AND ot.is_current = 1
                LEFT JOIN trading_business_assignments ba ON ba.trade_identity_id = a.open_trade_identity_id
                LEFT JOIN trading_business_subjects s ON s.id = ba.business_subject_id
                LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
                WHERE cf.is_current = 1
                ORDER BY cf.close_date DESC, cf.id DESC
                """,
            ).fetchall()
            spec_rows = db._exec(
                cur,
                """
                SELECT exchange, product_code, asset_type, contract_multiplier
                FROM trading_contract_specs WHERE is_active = 1
                """,
            ).fetchall()
            spec_by_key = {
                (
                    str(row["exchange"] or "").lower(),
                    str(row["product_code"] or "").lower(),
                    row["asset_type"],
                ): float(row["contract_multiplier"])
                for row in spec_rows
            }
        all_items = [
            dict(row) for row in rows
            if row["assignment_status"] == "classified"
            and row["allocation_source"] is not None
        ]
        if view == "junneng":
            items = [
                {**row, "ledger_membership": "confirmed"}
                for row in all_items if row["business_subject"] == "上海钧能"
            ]
        else:
            items = [row for row in all_items if row["asset_type"] == "option"]
        items = _filter_business_items(items, tab, filters)
        if view == "junneng":
            for item in items:
                multiplier = spec_by_key.get(
                    (
                        str(item["exchange"] or "").lower(),
                        _product_code(item["contract"]),
                        item["asset_type"],
                    )
                )
                if multiplier is None:
                    item.update({
                        "net_close_pnl": None,
                        "fund_interest": None,
                        "settlement_80": None,
                        "settlement_20": None,
                        "settlement_rule_version": SH_JUNNENG_RULE_VERSION,
                        "settlement_status": "missing_contract_spec",
                    })
                    continue
                matched_quantity = float(item["matched_quantity"] or 0)
                allocated_open_fee = (
                    float(item["open_trade_fee"] or 0)
                    * matched_quantity
                    / float(item["open_trade_quantity"] or matched_quantity or 1)
                )
                allocated_close_fee = (
                    float(item["matched_fee"] or 0)
                    * matched_quantity
                    / float(item["quantity"] or matched_quantity or 1)
                )
                settlement = calculate_sh_junneng_settlement(
                    gross_pnl=float(
                        item["allocation_business_pnl"]
                        if item["allocation_business_pnl"] is not None
                        else item["business_pnl"] or 0
                    ),
                    allocated_open_fee=allocated_open_fee,
                    allocated_close_fee=allocated_close_fee,
                    open_date=item["allocated_open_date"] or item["open_date"],
                    close_date=item["close_date"],
                    matched_quantity=matched_quantity,
                    open_price=float(
                        item["allocated_open_price"]
                        if item["allocated_open_price"] is not None
                        else item["open_price"]
                    ),
                    multiplier=multiplier,
                )
                item.update(settlement)
                item["allocated_open_fee"] = round(allocated_open_fee, 2)
                item["allocated_close_fee"] = round(allocated_close_fee, 2)
                item["settlement_status"] = "calculated"
        summary = {
            "record_count": len(items),
            "trade_close_record_count": sum(
                1 for row in items if row["settlement_type"] == "trade_close"
            ),
            "quantity": sum(float(row["matched_quantity"]) for row in items),
            "settlement_quantity": sum(
                float(row["matched_quantity"]) for row in items
            ),
            "transaction_close_quantity": sum(
                float(row["matched_quantity"])
                for row in items if row["settlement_type"] == "trade_close"
            ),
            "business_pnl": sum(float(row["business_pnl"] or 0) for row in items),
            "fact_close_pnl": sum(float(row["fact_close_pnl"] or 0) for row in items),
            "fee": sum(float(row["matched_fee"] or 0) for row in items),
        }
        if view == "junneng":
            calculated = [
                row for row in items if row.get("settlement_status") == "calculated"
            ]
            summary.update({
                "net_close_pnl": sum(float(row["net_close_pnl"]) for row in calculated),
                "fund_interest": sum(float(row["fund_interest"]) for row in calculated),
                "settlement_80": sum(float(row["settlement_80"]) for row in calculated),
                "settlement_20": sum(float(row["settlement_20"]) for row in calculated),
                "fee": sum(
                    float(row["allocated_open_fee"]) + float(row["allocated_close_fee"])
                    for row in calculated
                ),
                "settlement_rule_version": SH_JUNNENG_RULE_VERSION,
                "settlement_status": (
                    "calculated" if len(calculated) == len(items)
                    else "partial_missing_contract_spec"
                ),
            })
        return _page_result(items, summary, filters)


def _current_allocation_version(cur, close_identity_id: int) -> int:
    row = db._exec(
        cur,
        "SELECT MAX(allocation_version) AS v FROM trading_business_close_allocations WHERE close_identity_id = ?",
        (close_identity_id,),
    ).fetchone()
    return int(row["v"] or 0)


def list_business_close_candidates(close_identity_id: int) -> list[dict[str, Any]]:
    with db.connect() as conn:
        cur = conn.cursor()
        close = db._exec(
            cur,
            """
            SELECT cf.*, fi.account_id FROM trading_close_facts cf
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = cf.identity_id
            WHERE cf.is_current = 1 AND cf.identity_id = ?
            """,
            (close_identity_id,),
        ).fetchone()
        if not close:
            raise ValueError("平仓事实不存在或不是有效版本")
        rows = db._exec(
            cur,
            """
            SELECT tf.identity_id, tf.trade_date, tf.contract, tf.side, tf.quantity, tf.price,
                   tf.quantity - COALESCE((
                       SELECT SUM(a.matched_quantity) FROM trading_business_close_allocations a
                       WHERE a.open_trade_identity_id = tf.identity_id AND a.close_identity_id <> ?
                   ), 0) AS available_quantity,
                   ba.business_subject_id, ba.business_type, ba.strategy_id,
                   s.name AS business_subject, st.name AS strategy
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = tf.identity_id
            JOIN trading_business_assignments ba ON ba.trade_identity_id = tf.identity_id
            JOIN trading_business_subjects s ON s.id = ba.business_subject_id
            LEFT JOIN trading_strategies st ON st.id = ba.strategy_id
            WHERE tf.is_current = 1
              AND fi.account_id = ? AND tf.open_close = '开仓'
              AND tf.contract = ? AND tf.side = ? AND tf.trade_date <= ?
            ORDER BY tf.trade_date, tf.id
            """,
            (close_identity_id, close["account_id"], close["contract"], close["open_side"], close["close_date"]),
        ).fetchall()
    return [dict(row) for row in rows if float(row["available_quantity"] or 0) > 1e-9]


def preview_business_rematch(
    close_identity_id: int,
    selections: list[dict[str, Any]],
    allocation_version: int,
) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        current_version = _current_allocation_version(cur, close_identity_id)
        if current_version != allocation_version:
            raise ValueError("业务开平数据已变化，请刷新后重试")
        close = db._exec(
            cur,
            "SELECT * FROM trading_close_facts WHERE is_current = 1 AND identity_id = ?",
            (close_identity_id,),
        ).fetchone()
        if not close:
            raise ValueError("平仓事实不存在")
        candidates = {row["identity_id"]: row for row in list_business_close_candidates(close_identity_id)}
        if not selections:
            raise ValueError("请选择业务开仓记录")
        selected_rows = []
        total = 0.0
        configs = set()
        for selection in selections:
            identity_id = int(selection["open_trade_identity_id"])
            quantity = float(selection["quantity"])
            candidate = candidates.get(identity_id)
            if not candidate:
                raise ValueError("只能选择同账户、同合约、同方向且时间合规的业务开仓")
            if quantity <= 0 or quantity > float(candidate["available_quantity"]) + 1e-9:
                raise ValueError("业务可平手数不足")
            configs.add((candidate["business_subject_id"], candidate["business_type"], candidate["strategy_id"]))
            selected_rows.append({**selection, "candidate": candidate})
            total += quantity
        if len(configs) != 1:
            raise ValueError("多笔开仓的业务归属、业务类型和策略必须一致")
        if abs(total - float(close["quantity"])) > 1e-9:
            raise ValueError("业务匹配手数合计必须等于平仓手数")
        spec = db._exec(
            cur,
            """
            SELECT contract_multiplier FROM trading_contract_specs
            WHERE LOWER(exchange) = LOWER(?) AND LOWER(product_code) = ? AND asset_type = ? AND is_active = 1
            """,
            (close["exchange"], _product_code(close["contract"]), close["asset_type"]),
        ).fetchone()
        if not spec:
            raise ValueError("合约参数待核验")
        before_row = db._exec(
            cur,
            "SELECT SUM(business_pnl) AS pnl FROM trading_business_close_allocations WHERE close_identity_id = ?",
            (close_identity_id,),
        ).fetchone()
        after_pnl = sum(
            calculate_business_pnl(
                float(item["candidate"]["price"]), float(close["close_price"]), close["open_side"],
                float(item["quantity"]), float(spec["contract_multiplier"]),
            )
            for item in selected_rows
        )
        token = uuid.uuid4().hex
        payload = {
            "close_identity_id": close_identity_id,
            "allocation_version": allocation_version,
            "selections": [
                {"open_trade_identity_id": int(item["open_trade_identity_id"]), "quantity": float(item["quantity"])}
                for item in selected_rows
            ],
            "business_config": list(next(iter(configs))),
            "before_business_pnl": float(before_row["pnl"] or 0),
            "after_business_pnl": after_pnl,
        }
        _REMATCH_PREVIEWS[token] = payload
    return {"preview_token": token, **payload}


def reconcile_business_pnl(close_identity_id: int) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        close = db._exec(
            cur,
            """
            SELECT cf.contract, cf.open_side, fi.account_id
            FROM trading_close_facts cf
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = cf.identity_id
            WHERE cf.is_current = 1 AND cf.identity_id = ?
            """,
            (close_identity_id,),
        ).fetchone()
        if not close:
            raise ValueError("平仓事实不存在或不是有效版本")
        open_row = db._exec(
            cur,
            """
            SELECT SUM(tf.quantity) AS quantity
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = tf.identity_id
            WHERE tf.is_current = 1
              AND fi.account_id = ? AND tf.contract = ? AND tf.side = ? AND tf.open_close = '开仓'
            """,
            (close["account_id"], close["contract"], close["open_side"]),
        ).fetchone()
        allocation_row = db._exec(
            cur,
            """
            SELECT SUM(a.matched_quantity) AS quantity, SUM(a.business_pnl) AS business_pnl
            FROM trading_business_close_allocations a
            JOIN trading_close_facts cf ON cf.identity_id = a.close_identity_id
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = cf.identity_id
            WHERE cf.is_current = 1
              AND fi.account_id = ? AND cf.contract = ? AND cf.open_side = ?
            """,
            (close["account_id"], close["contract"], close["open_side"]),
        ).fetchone()
        fact_row = db._exec(
            cur,
            """
            SELECT SUM(cf.fact_close_pnl) AS fact_pnl
            FROM trading_close_facts cf
            JOIN trading_import_batches b ON b.id = cf.batch_id AND b.status = 'active'
            JOIN trading_fact_identities fi ON fi.id = cf.identity_id
            WHERE cf.is_current = 1
              AND fi.account_id = ? AND cf.contract = ? AND cf.open_side = ?
            """,
            (close["account_id"], close["contract"], close["open_side"]),
        ).fetchone()
    open_quantity = float(open_row["quantity"] or 0)
    allocated_quantity = float(allocation_row["quantity"] or 0)
    business_pnl = float(allocation_row["business_pnl"] or 0)
    fact_pnl = float(fact_row["fact_pnl"] or 0)
    remaining_quantity = max(0.0, open_quantity - allocated_quantity)
    difference = round(business_pnl - fact_pnl, 8)
    if open_quantity <= 0 or remaining_quantity > 1e-9:
        status = "not_fully_closed"
    elif abs(difference) <= 1e-8:
        status = "reconciled"
    else:
        status = "business_pnl_reconciliation_failed"
    return {
        "status": status,
        "account_id": close["account_id"],
        "contract": close["contract"],
        "direction": close["open_side"],
        "open_quantity": open_quantity,
        "allocated_quantity": allocated_quantity,
        "remaining_quantity": remaining_quantity,
        "fact_pnl": fact_pnl,
        "business_pnl": business_pnl,
        "difference": difference,
    }


def _inherit_close_trade_assignment(cur, close_identity_id: int, business_config: tuple, actor: str) -> None:
    business_subject_id, business_type, strategy_id = business_config
    close_trade_rows = db._exec(
        cur,
        "SELECT DISTINCT close_trade_identity_id FROM trading_close_trade_links WHERE close_identity_id = ?",
        (close_identity_id,),
    ).fetchall()
    for row in close_trade_rows:
        trade_identity_id = row["close_trade_identity_id"]
        before_row = db._exec(
            cur,
            "SELECT * FROM trading_business_assignments WHERE trade_identity_id = ?",
            (trade_identity_id,),
        ).fetchone()
        before = dict(before_row) if before_row else None
        if before_row:
            db._exec(
                cur,
                """
                UPDATE trading_business_assignments
                SET business_subject_id = ?, business_type = ?, strategy_id = ?,
                    updated_by = ?, updated_at = CURRENT_TIMESTAMP
                WHERE trade_identity_id = ?
                """,
                (business_subject_id, business_type, strategy_id, actor, trade_identity_id),
            )
        else:
            db._exec(
                cur,
                """
                INSERT INTO trading_business_assignments
                    (trade_identity_id, business_subject_id, business_type, strategy_id,
                     assigned_by, updated_by)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (trade_identity_id, business_subject_id, business_type, strategy_id, actor, actor),
            )
        after = db._exec(
            cur,
            "SELECT * FROM trading_business_assignments WHERE trade_identity_id = ?",
            (trade_identity_id,),
        ).fetchone()
        _write_assignment_audit(cur, trade_identity_id, "业务平仓自动继承", actor, before, dict(after))


def confirm_business_rematch(
    close_identity_id: int,
    preview_token: str,
    allocation_version: int,
    actor: str,
    reason: str = "",
) -> dict[str, Any]:
    payload = _REMATCH_PREVIEWS.pop(preview_token, None)
    if not payload or payload["close_identity_id"] != close_identity_id:
        raise ValueError("重配预览已失效，请重新预览")
    if payload["allocation_version"] != allocation_version:
        raise ValueError("业务开平数据已变化，请刷新后重试")
    with db.connect() as conn:
        cur = conn.cursor()
        if _current_allocation_version(cur, close_identity_id) != allocation_version:
            raise ValueError("业务开平数据已变化，请刷新后重试")
        before_rows = [
            dict(row) for row in db._exec(
                cur,
                "SELECT * FROM trading_business_close_allocations WHERE close_identity_id = ? ORDER BY id",
                (close_identity_id,),
            ).fetchall()
        ]
        new_version = allocation_version + 1
        group_id = uuid.uuid4().hex
        db._exec(cur, "DELETE FROM trading_business_close_allocations WHERE close_identity_id = ?", (close_identity_id,))
        close = db._exec(cur, "SELECT * FROM trading_close_facts WHERE is_current = 1 AND identity_id = ?", (close_identity_id,)).fetchone()
        spec = db._exec(
            cur,
            "SELECT contract_multiplier FROM trading_contract_specs WHERE LOWER(exchange)=LOWER(?) AND LOWER(product_code)=? AND asset_type=? AND is_active=1",
            (close["exchange"], _product_code(close["contract"]), close["asset_type"]),
        ).fetchone()
        after_rows = []
        for selection in payload["selections"]:
            open_fact = db._exec(
                cur,
                "SELECT price FROM trading_trade_facts tf JOIN trading_import_batches b ON b.id=tf.batch_id AND b.status='active' WHERE tf.is_current=1 AND tf.identity_id=?",
                (selection["open_trade_identity_id"],),
            ).fetchone()
            pnl = calculate_business_pnl(
                float(open_fact["price"]), float(close["close_price"]), close["open_side"],
                float(selection["quantity"]), float(spec["contract_multiplier"]),
            )
            allocation_id = db._last_insert_id(
                cur,
                """
                INSERT INTO trading_business_close_allocations
                    (close_identity_id, open_trade_identity_id, matched_quantity, source,
                     override_group_id, business_pnl, rule_version, allocation_version, created_by, updated_by)
                VALUES (?, ?, ?, 'manual_override', ?, ?, 'business-manual-v1', ?, ?, ?)
                """,
                (close_identity_id, selection["open_trade_identity_id"], selection["quantity"], group_id, pnl, new_version, actor, actor),
            )
            after_rows.append({"id": allocation_id, **selection, "business_pnl": pnl})
        _inherit_close_trade_assignment(cur, close_identity_id, tuple(payload["business_config"]), actor)
        db._exec(
            cur,
            """
            INSERT INTO trading_business_allocation_audit
                (override_group_id, close_identity_id, before_allocations, after_allocations,
                 before_business_pnl, after_business_pnl, reason, operated_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                group_id, close_identity_id, json.dumps(before_rows, ensure_ascii=False),
                json.dumps(after_rows, ensure_ascii=False), payload["before_business_pnl"],
                payload["after_business_pnl"], reason or None, actor,
            ),
        )
        conn.commit()
    return {
        "allocation_version": new_version,
        "business_pnl": payload["after_business_pnl"],
        "reconciliation": reconcile_business_pnl(close_identity_id),
    }


def restore_default_business_match(
    close_identity_id: int,
    allocation_version: int,
    actor: str,
    reason: str = "恢复事实层默认开平关系",
) -> dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        if _current_allocation_version(cur, close_identity_id) != allocation_version:
            raise ValueError("业务开平数据已变化，请刷新后重试")
        close = db._exec(
            cur,
            "SELECT * FROM trading_close_facts WHERE is_current = 1 AND identity_id = ?",
            (close_identity_id,),
        ).fetchone()
        if not close:
            raise ValueError("平仓事实不存在")
        fact_rows = db._exec(
            cur,
            """
            SELECT fa.open_trade_identity_id, fa.matched_quantity, tf.price,
                   ba.business_subject_id, ba.business_type, ba.strategy_id
            FROM trading_fact_close_allocations fa
            JOIN trading_trade_facts tf ON tf.identity_id = fa.open_trade_identity_id
            JOIN trading_import_batches b ON b.id = tf.batch_id AND b.status = 'active'
            JOIN trading_business_assignments ba ON ba.trade_identity_id = fa.open_trade_identity_id
            WHERE tf.is_current = 1 AND fa.close_identity_id = ?
            ORDER BY fa.id
            """,
            (close_identity_id,),
        ).fetchall()
        if not fact_rows:
            raise ValueError("没有可恢复的事实层默认开平关系")
        configs = {
            (row["business_subject_id"], row["business_type"], row["strategy_id"])
            for row in fact_rows
        }
        if len(configs) != 1:
            raise ValueError("事实层开仓对应多个业务归属，无法自动继承到整笔平仓成交")
        spec = db._exec(
            cur,
            """
            SELECT contract_multiplier FROM trading_contract_specs
            WHERE LOWER(exchange)=LOWER(?) AND LOWER(product_code)=? AND asset_type=? AND is_active=1
            """,
            (close["exchange"], _product_code(close["contract"]), close["asset_type"]),
        ).fetchone()
        if not spec:
            raise ValueError("合约参数待核验")
        before_rows = [
            dict(row) for row in db._exec(
                cur,
                "SELECT * FROM trading_business_close_allocations WHERE close_identity_id = ? ORDER BY id",
                (close_identity_id,),
            ).fetchall()
        ]
        new_version = allocation_version + 1
        group_id = uuid.uuid4().hex
        db._exec(cur, "DELETE FROM trading_business_close_allocations WHERE close_identity_id = ?", (close_identity_id,))
        after_rows = []
        after_pnl = 0.0
        for row in fact_rows:
            pnl = calculate_business_pnl(
                float(row["price"]), float(close["close_price"]), close["open_side"],
                float(row["matched_quantity"]), float(spec["contract_multiplier"]),
            )
            allocation_id = db._last_insert_id(
                cur,
                """
                INSERT INTO trading_business_close_allocations
                    (close_identity_id, open_trade_identity_id, matched_quantity, source,
                     override_group_id, business_pnl, rule_version, allocation_version, created_by, updated_by)
                VALUES (?, ?, ?, 'fact_default', ?, ?, 'business-default-v1', ?, ?, ?)
                """,
                (close_identity_id, row["open_trade_identity_id"], row["matched_quantity"],
                 group_id, pnl, new_version, actor, actor),
            )
            after_rows.append({
                "id": allocation_id,
                "open_trade_identity_id": row["open_trade_identity_id"],
                "quantity": row["matched_quantity"],
                "business_pnl": pnl,
            })
            after_pnl += pnl
        _inherit_close_trade_assignment(cur, close_identity_id, next(iter(configs)), actor)
        before_pnl = sum(float(row.get("business_pnl") or 0) for row in before_rows)
        db._exec(
            cur,
            """
            INSERT INTO trading_business_allocation_audit
                (override_group_id, close_identity_id, before_allocations, after_allocations,
                 before_business_pnl, after_business_pnl, reason, operated_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (group_id, close_identity_id, json.dumps(before_rows, ensure_ascii=False),
             json.dumps(after_rows, ensure_ascii=False), before_pnl, after_pnl, reason, actor),
        )
        conn.commit()
    return {
        "allocation_version": new_version,
        "business_pnl": after_pnl,
        "reconciliation": reconcile_business_pnl(close_identity_id),
    }


def _actor(user: dict[str, Any]) -> str:
    return str(user.get("username") or user.get("name") or user.get("id") or "unknown")


def _as_http_error(exc: ValueError) -> HTTPException:
    return HTTPException(status_code=400, detail=str(exc))


def _decode_statement_upload(item: TradingUploadFile) -> bytes:
    if Path(item.name).suffix.lower() != ".txt":
        raise ValueError("仅支持 txt 结算单")
    try:
        content = base64.b64decode(item.content_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError(f"{item.name} 文件内容无效") from exc
    if not content:
        raise ValueError(f"{item.name} 文件为空")
    return content


def _cleanup_preview_files(batch_id: int) -> None:
    with db.connect() as conn:
        row = db._exec(
            conn.cursor(),
            "SELECT parse_summary FROM trading_import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
    if not row:
        return
    summary = json.loads(row["parse_summary"] or "{}")
    paths = summary.get("paths", {})
    if summary.get("statement_path"):
        paths = {**paths, "statement": summary["statement_path"]}
    parents = set()
    for value in paths.values():
        path = Path(value)
        parents.add(path.parent)
        path.unlink(missing_ok=True)
    for parent in parents:
        if parent.name and parent.parent.name == "trading_import_uploads":
            try:
                parent.rmdir()
            except OSError:
                pass


def _api_filters(
    contract: str = "",
    direction: str = "",
    asset_type: str = "",
    open_close: str = "",
    start_date: str = "",
    end_date: str = "",
    classification: str = "",
    business_type: str = "",
    page: int = 1,
    page_size: int = 20,
) -> FactFilters:
    return FactFilters(
        contract=contract, direction=direction, asset_type=asset_type, open_close=open_close,
        start_date=start_date, end_date=end_date, classification=classification,
        business_type=business_type, page=page, page_size=page_size,
    )


async def trading_management_current_user(
    authorization: Optional[str] = Header(default=None),
) -> dict:
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.removeprefix("Bearer ").strip()
    user = db.get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="请先登录")
    return user


@router.get("/overview")
def get_trading_overview(
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.overview", "view")
    return build_overview(filters)


def _get_trading_facts(
    view: str,
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.facts", "view")
    if view not in {"positions", "closes", "trades"}:
        raise HTTPException(status_code=404, detail="未知事实视图")
    return query_fact_rows(view, filters)


@router.get("/facts/positions")
def get_trading_positions(
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    return _get_trading_facts("positions", filters, user)


@router.get("/facts/closes")
def get_trading_closes(
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    return _get_trading_facts("closes", filters, user)


@router.get("/facts/trades")
def get_trading_trades(
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    return _get_trading_facts("trades", filters, user)


@router.get("/facts/trades/selection-identities")
def get_trading_trade_selection_identities(
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.facts", "view")
    return query_trade_selection_identities(filters)


@router.get("/imports")
def get_trading_imports(user=Depends(trading_management_current_user)):
    require_permission(user, "trading.imports", "view")
    with db.connect() as conn:
        rows = db._exec(
            conn.cursor(),
            "SELECT * FROM trading_import_batches ORDER BY created_at DESC, id DESC",
        ).fetchall()
    return {"items": [dict(row) for row in rows]}


@router.get("/imports/{batch_id}/validation")
def get_trading_import_validation(batch_id: int, user=Depends(trading_management_current_user)):
    require_permission(user, "trading.imports", "view")
    with db.connect() as conn:
        batch = db._exec(
            conn.cursor(),
            "SELECT id, status, parse_summary FROM trading_import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
    if not batch:
        raise HTTPException(status_code=404, detail="导入批次不存在")
    return {
        "batch_id": batch["id"],
        "status": batch["status"],
        "summary": json.loads(batch["parse_summary"] or "{}"),
    }


@router.post("/imports/preview")
def post_trading_import_preview(
    payload: TradingImportPreviewIn,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.imports", "import")
    try:
        content = _decode_statement_upload(payload.statement_file)
        return preview_settlement_import(
            payload.account_id, payload.statement_file.name, content, _actor(user)
        )
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/imports/{preview_batch_id}/confirm")
def post_trading_import_confirm(
    preview_batch_id: int,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.imports", "import")
    try:
        result = confirm_settlement_import(preview_batch_id, _actor(user))
        result["matching"] = match_imported_facts(result["batch_id"])
        result["business_allocations"] = rebuild_default_business_allocations(result["batch_id"])
        _cleanup_preview_files(preview_batch_id)
        return result
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.get("/config")
def get_trading_config(user=Depends(trading_management_current_user)):
    require_permission(user, "trading.config", "view")
    return list_trading_config()


@router.post("/business-subjects")
def post_business_subject(payload: BusinessSubjectIn, user=Depends(trading_management_current_user)):
    require_permission(user, "trading.config", "edit")
    try:
        return create_business_subject(payload.name, _actor(user))
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/strategies")
def post_strategy(payload: StrategyIn, user=Depends(trading_management_current_user)):
    require_permission(user, "trading.config", "edit")
    try:
        return get_or_create_strategy(payload.name, _actor(user))
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/business-assignments/batch-confirm")
def post_business_assignments(
    payload: BusinessAssignmentIn,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.config", "edit")
    try:
        return classify_trade_identities(
            payload.identity_ids, payload.business_subject_id, payload.business_type,
            payload.strategy_name, payload.instruction_text, _actor(user),
        )
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.delete("/business-assignments/{trade_identity_id}")
def delete_business_assignment(
    trade_identity_id: int,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.config", "delete")
    return remove_trade_assignment(trade_identity_id, _actor(user))


def _get_business_rows(
    view: str,
    tab: str,
    filters: FactFilters,
    user: dict[str, Any],
):
    resource = "trading.junneng" if view == "junneng" else "trading.options"
    require_permission(user, resource, "view")
    try:
        return query_business_rows(view, tab, filters)
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.get("/business/junneng/{tab}")
def get_junneng_business_rows(
    tab: str,
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    return _get_business_rows("junneng", tab, filters, user)


@router.get("/business/options/{tab}")
def get_option_business_rows(
    tab: str,
    filters: FactFilters = Depends(_api_filters),
    user=Depends(trading_management_current_user),
):
    return _get_business_rows("options", tab, filters, user)


@router.get("/business-closes/{close_identity_id}/candidates")
def get_rematch_candidates(close_identity_id: int, user=Depends(trading_management_current_user)):
    require_permission(user, "trading.config", "edit")
    try:
        return {"items": list_business_close_candidates(close_identity_id)}
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/business-closes/{close_identity_id}/preview")
def post_rematch_preview(
    close_identity_id: int,
    payload: RematchPreviewIn,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.config", "edit")
    try:
        return preview_business_rematch(
            close_identity_id,
            [item.dict() for item in payload.selections],
            payload.allocation_version,
        )
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/business-closes/{close_identity_id}/confirm")
def post_rematch_confirm(
    close_identity_id: int,
    payload: RematchConfirmIn,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.config", "edit")
    try:
        return confirm_business_rematch(
            close_identity_id, payload.preview_token, payload.allocation_version,
            _actor(user), payload.reason,
        )
    except ValueError as exc:
        raise _as_http_error(exc) from exc


@router.post("/business-closes/{close_identity_id}/restore-default")
def post_restore_default(
    close_identity_id: int,
    payload: RestoreDefaultIn,
    user=Depends(trading_management_current_user),
):
    require_permission(user, "trading.config", "edit")
    try:
        return restore_default_business_match(
            close_identity_id, payload.allocation_version, _actor(user), payload.reason,
        )
    except ValueError as exc:
        raise _as_http_error(exc) from exc
