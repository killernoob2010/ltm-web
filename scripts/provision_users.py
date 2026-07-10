#!/usr/bin/env python3
"""Dry-run-first provisioning for the initial internal user roster."""

import argparse
import getpass
import json
import os
from pathlib import Path
from typing import Callable, Optional
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


REQUIRED_HEADERS = {
    "姓名": "name",
    "部门": "department",
    "用户类型": "role",
}
OPTIONAL_HEADERS = {
    "登录账号（可空）": "username",
}


def load_roster(path: Path) -> list[dict[str, str]]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values, [])]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"名单缺少列: {', '.join(missing)}")
    indexes = {
        field: headers.index(header)
        for header, field in REQUIRED_HEADERS.items()
    }
    indexes.update({
        field: headers.index(header)
        for header, field in OPTIONAL_HEADERS.items()
        if header in headers
    })
    rows = []
    for row_no, values_row in enumerate(values, start=2):
        if not any(value not in (None, "") for value in values_row):
            continue
        item = {
            key: str(values_row[index] or "").strip()
            for key, index in indexes.items()
        }
        item.setdefault("username", "")
        if not item["name"] or not item["department"] or not item["role"]:
            raise ValueError(f"第 {row_no} 行姓名、部门和用户类型不能为空")
        rows.append(item)
    if not rows:
        raise ValueError("名单中没有用户")
    return rows


def permission_overrides(row: dict[str, str]) -> list[dict[str, str]]:
    if row["name"] == "龙云飞":
        return [
            {"module_code": "order_finance_progress", "level": "operate"},
            {"module_code": "order_finance_capital", "level": "operate"},
        ]
    return []


def preflight_roster(
    rows: list[dict[str, str]],
    preview: Callable[[dict], dict],
) -> list[dict[str, object]]:
    results = []
    conflicts = []
    usernames = set()
    for row in rows:
        payload = {**row, "permissions": permission_overrides(row)}
        item = preview(payload)
        username = item["username"]
        if not item.get("username_available", False) or username in usernames:
            conflicts.append(username)
        usernames.add(username)
        permissions = [
            {"module_code": code, "level": level}
            for code, level in item.get("final_permissions", {}).items()
            if level != "none"
        ]
        results.append({
            "name": row["name"],
            "username": username,
            "temporary_password": item["temporary_password"],
            "password_rule": item["password_rule"],
            "department": row["department"],
            "role": row["role"],
            "permission_count": len(permissions),
            "permissions": permissions,
        })
    if conflicts:
        raise ValueError(f"登录账号冲突: {', '.join(sorted(set(conflicts)))}")
    return results


class ApiClient:
    def __init__(self, base_url: str, token: str = ""):
        self.base_url = base_url.rstrip("/")
        self.token = token

    def request(
        self,
        path: str,
        payload: Optional[dict] = None,
        method: str = "POST",
    ) -> dict:
        headers = {"Content-Type": "application/json"}
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        request = Request(
            f"{self.base_url}{path}",
            data=(
                json.dumps(payload, ensure_ascii=False).encode("utf-8")
                if payload is not None
                else None
            ),
            headers=headers,
            method=method,
        )
        try:
            with urlopen(request, timeout=30) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            try:
                detail = json.loads(detail).get("detail", detail)
            except json.JSONDecodeError:
                pass
            raise ValueError(f"API {exc.code}: {detail}") from exc


def apply_roster(
    client: ApiClient,
    rows: list[dict[str, str]],
    previews: list[dict[str, object]],
) -> None:
    for row, preview in zip(rows, previews):
        client.request("/api/users", {
            **row,
            "username": preview["username"],
            "permissions": preview["permissions"],
        })


def configure_wangjingze(client: ApiClient, password: str) -> None:
    users = client.request("/api/users", method="GET")["users"]
    target = next((item for item in users if item["name"] == "王景泽"), None)
    if not target:
        raise ValueError("未找到现有用户王景泽")
    if target["department"] != "管理部门" or target["role"] != "管理员":
        raise ValueError("王景泽必须先是管理部门管理员")
    client.request(
        f"/api/users/{target['id']}/set-password",
        {"new_password": password, "password_change_recommended": False},
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="预检并创建首批内部用户（默认仅 dry-run）")
    parser.add_argument("--base-url", required=True, help="Staging 或经确认的 Production URL")
    parser.add_argument("--file", required=True, type=Path, help="Excel 用户名单")
    parser.add_argument("--apply", action="store_true", help="通过预检后真正创建用户")
    parser.add_argument(
        "--configure-wangjingze",
        action="store_true",
        help="应用时同步配置现有管理员王景泽",
    )
    args = parser.parse_args()

    username = os.getenv("LTM_ADMIN_USERNAME") or input("管理员登录账号: ").strip()
    password = os.getenv("LTM_ADMIN_PASSWORD") or getpass.getpass("管理员密码: ")
    client = ApiClient(args.base_url)
    login = client.request("/api/auth/login", {"username": username, "password": password})
    client.token = login["token"]

    rows = load_roster(args.file)
    previews = preflight_roster(rows, lambda payload: client.request("/api/users/preview", payload))
    print(json.dumps({"mode": "apply" if args.apply else "dry-run", "users": previews}, ensure_ascii=False, indent=2))
    if not args.apply:
        return

    apply_roster(client, rows, previews)
    print(f"已创建 {len(rows)} 个用户")
    if args.configure_wangjingze:
        wang_password = os.getenv("LTM_WANGJINGZE_PASSWORD") or getpass.getpass("王景泽新密码: ")
        configure_wangjingze(client, wang_password)
        print("已配置现有管理员王景泽")


if __name__ == "__main__":
    main()
