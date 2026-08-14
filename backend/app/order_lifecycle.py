"""订单全流程管理（测试版）。

本模块与旧的 ``order_finance_progress`` 平行存在：旧表和旧页面继续服务原有
订单融资功能；本模块保存一张业务主卡及其融资、合同、船舶、单据、回款等子记录。
来源解析默认只读，只有显式调用导入函数时才写入测试环境数据库。
"""
from __future__ import annotations

import base64
import binascii
from copy import deepcopy
import hashlib
import json
import re
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Optional
from uuid import uuid4

from fastapi import APIRouter, Depends, Header, HTTPException, Query
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from . import db
from .order_finance import (
    _get,
    _normalize_date,
    _normalize_text,
    _parse_date,
    _subsidiary_from_filename,
    _to_float,
    parse_order_finance_workbook,
)
from .permissions import can, require_permission


router = APIRouter()
ORDER_LIFECYCLE_MODULE = "order_lifecycle_progress"
PERMISSION_RESOURCE = "order_finance.records"
WPS_RAW_SHEETS = ("YOLANDA", "JLHK", "天津建龙")
MAIL_MILLS = ("阿城", "北满", "承德", "东钢", "抚顺", "西林")
TABLES = (
    "order_lifecycle_businesses",
    "order_lifecycle_contracts",
    "order_lifecycle_financings",
    "order_lifecycle_vessels",
    "order_lifecycle_documents",
    "order_lifecycle_customer_receipts",
    "order_lifecycle_bank_repayments",
    "order_lifecycle_source_batches",
    "order_lifecycle_source_records",
    "order_lifecycle_business_sources",
    "order_lifecycle_manual_overrides",
    "order_lifecycle_child_overrides",
    "order_lifecycle_manual_child_records",
    "order_lifecycle_data_anomalies",
    "order_lifecycle_audit",
    "order_lifecycle_match_candidates",
    "order_lifecycle_sync_state",
)

MANUAL_OVERRIDE_FIELDS = {
    "business_no",
    "business_type",
    "trade_entity",
    "supplier_steel_mill",
    "terminal_customer",
    "product_name",
    "contract_quantity_mt",
    "settlement_status",
    "settlement_date",
    "guo_danlei_special",
    "port_status",
    "port_confirmed_date",
    "shipment_status",
    "shipment_confirmed_date",
    "next_follow_up_date",
}

CHILD_TABLES = {
    "contracts": "order_lifecycle_contracts",
    "financings": "order_lifecycle_financings",
    "vessels": "order_lifecycle_vessels",
    "documents": "order_lifecycle_documents",
    "customer_receipts": "order_lifecycle_customer_receipts",
    "bank_repayments": "order_lifecycle_bank_repayments",
}

CHILD_FIELDS = {
    "contracts": {"contract_no", "purchase_contract_no", "system_contract_no", "buyer", "seller", "quantity_mt"},
    "financings": {"bank", "amount", "currency", "financing_date", "original_due_date", "extended_due_date", "repayment_date", "repayment_status"},
    "vessels": {"vessel_name", "imo", "loading_port", "discharge_port", "eta", "etb", "estimated_discharge_date", "latest_shipment_date"},
    "documents": {"document_type", "document_date"},
    "customer_receipts": {"receipt_date", "amount", "currency", "fully_received", "applicable_scope"},
    "bank_repayments": {"financing_id", "repayment_date", "amount", "currency", "completion_explicit"},
}

CHILD_COLUMN_ORDER = {
    "contracts": ("contract_no", "purchase_contract_no", "system_contract_no", "buyer", "seller", "quantity_mt"),
    "financings": ("bank", "amount", "currency", "financing_date", "original_due_date", "extended_due_date", "repayment_date", "repayment_status"),
    "vessels": ("vessel_name", "imo", "loading_port", "discharge_port", "eta", "etb", "estimated_discharge_date", "latest_shipment_date", "source"),
    "documents": ("document_type", "document_date"),
    "customer_receipts": ("receipt_date", "amount", "currency", "fully_received", "applicable_scope"),
    "bank_repayments": ("financing_id", "repayment_date", "amount", "currency", "completion_explicit"),
}


def _id_sql() -> str:
    return "SERIAL PRIMARY KEY" if db._is_pg() else "INTEGER PRIMARY KEY AUTOINCREMENT"


def initialize_schema(conn) -> None:
    """Create only the new lifecycle tables; never alters old order-finance facts."""
    ident = _id_sql()
    cur = conn.cursor()
    statements = [
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_businesses (
            id {ident},
            business_uid TEXT NOT NULL UNIQUE,
            business_key TEXT NOT NULL UNIQUE,
            business_no TEXT NOT NULL,
            business_type TEXT NOT NULL,
            trade_entity TEXT,
            supplier_steel_mill TEXT,
            terminal_customer TEXT,
            product_name TEXT,
            contract_quantity_mt DOUBLE PRECISION,
            status TEXT NOT NULL DEFAULT '待确认',
            port_status TEXT NOT NULL DEFAULT '待确认',
            port_confirmed_date TEXT,
            port_confirmed_by TEXT,
            shipment_status TEXT NOT NULL DEFAULT '待确认',
            shipment_confirmed_date TEXT,
            shipment_confirmed_by TEXT,
            next_follow_up_date TEXT,
            risk_level TEXT NOT NULL DEFAULT '低风险',
            risk_reasons_json TEXT NOT NULL DEFAULT '[]',
            anomaly_count INTEGER NOT NULL DEFAULT 0,
            fcr INTEGER NOT NULL DEFAULT 0,
            settlement_status TEXT NOT NULL DEFAULT '待结算',
            settlement_date TEXT,
            guo_danlei_special INTEGER NOT NULL DEFAULT 0,
            completed_date TEXT,
            is_cancelled INTEGER NOT NULL DEFAULT 0,
            cancelled_at TEXT,
            source_type TEXT NOT NULL,
            source_snapshot_date TEXT,
            source_version TEXT,
            source_record_key TEXT,
            source_presence_hash TEXT,
            source_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_contracts (
            id {ident},
            business_id INTEGER NOT NULL,
            contract_no TEXT,
            purchase_contract_no TEXT,
            system_contract_no TEXT,
            buyer TEXT,
            seller TEXT,
            quantity_mt DOUBLE PRECISION,
            source_key TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_financings (
            id {ident},
            business_id INTEGER NOT NULL,
            bank TEXT,
            amount DOUBLE PRECISION,
            currency TEXT,
            financing_date TEXT,
            original_due_date TEXT,
            extended_due_date TEXT,
            repayment_date TEXT,
            repayment_status TEXT,
            source_key TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_vessels (
            id {ident},
            business_id INTEGER NOT NULL,
            vessel_name TEXT,
            imo TEXT,
            loading_port TEXT,
            discharge_port TEXT,
            eta TEXT,
            etb TEXT,
            estimated_discharge_date TEXT,
            latest_shipment_date TEXT,
            source_key TEXT NOT NULL,
            source TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_documents (
            id {ident},
            business_id INTEGER NOT NULL,
            document_type TEXT NOT NULL,
            document_date TEXT,
            source_key TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_customer_receipts (
            id {ident},
            business_id INTEGER NOT NULL,
            receipt_date TEXT,
            amount DOUBLE PRECISION,
            currency TEXT,
            fully_received INTEGER NOT NULL DEFAULT 0,
            applicable_scope TEXT,
            source_key TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_bank_repayments (
            id {ident},
            business_id INTEGER NOT NULL,
            financing_id INTEGER,
            repayment_date TEXT,
            amount DOUBLE PRECISION,
            currency TEXT,
            completion_explicit INTEGER NOT NULL DEFAULT 0,
            source_key TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(business_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_source_batches (
            id {ident},
            source_type TEXT NOT NULL,
            source_locator TEXT,
            source_version TEXT,
            snapshot_date TEXT,
            source_hash TEXT,
            source_key_set_hash TEXT,
            deletion_candidate_hash TEXT,
            deletion_candidate_count INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL,
            record_count INTEGER NOT NULL DEFAULT 0,
            error_message TEXT,
            started_at TEXT DEFAULT CURRENT_TIMESTAMP,
            completed_at TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_source_records (
            id {ident},
            batch_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            source_key TEXT NOT NULL,
            business_key TEXT NOT NULL,
            source_file TEXT,
            source_sheet TEXT,
            source_row INTEGER,
            raw_json TEXT NOT NULL DEFAULT '{{}}',
            normalized_json TEXT NOT NULL DEFAULT '{{}}',
            raw_hash TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(batch_id, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_business_sources (
            id {ident},
            business_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            source_business_key TEXT NOT NULL,
            source_record_key TEXT,
            source_version TEXT,
            source_presence_hash TEXT,
            source_active INTEGER NOT NULL DEFAULT 1,
            missing_observation_hash TEXT,
            missing_observation_count INTEGER NOT NULL DEFAULT 0,
            last_seen_batch_id INTEGER,
            last_seen_at TEXT DEFAULT CURRENT_TIMESTAMP,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(source_type, source_business_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_manual_overrides (
            id {ident},
            business_id INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            value_json TEXT NOT NULL,
            note TEXT,
            modified_by TEXT NOT NULL,
            modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
            is_active INTEGER NOT NULL DEFAULT 1,
            UNIQUE(business_id, field_name)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_child_overrides (
            id {ident},
            business_id INTEGER NOT NULL,
            collection TEXT NOT NULL,
            source_key TEXT NOT NULL,
            field_name TEXT NOT NULL,
            value_json TEXT NOT NULL,
            note TEXT,
            modified_by TEXT NOT NULL,
            modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
            is_active INTEGER NOT NULL DEFAULT 1,
            UNIQUE(business_id, collection, source_key, field_name)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_manual_child_records (
            id {ident},
            business_id INTEGER NOT NULL,
            collection TEXT NOT NULL,
            source_key TEXT NOT NULL,
            record_json TEXT NOT NULL DEFAULT '{{}}',
            note TEXT,
            modified_by TEXT NOT NULL,
            modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
            is_active INTEGER NOT NULL DEFAULT 1,
            UNIQUE(business_id, collection, source_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_data_anomalies (
            id {ident},
            business_id INTEGER NOT NULL,
            anomaly_key TEXT NOT NULL,
            anomaly_type TEXT NOT NULL,
            description TEXT NOT NULL,
            details_json TEXT NOT NULL DEFAULT '{{}}',
            status TEXT NOT NULL DEFAULT 'open',
            first_seen_at TEXT DEFAULT CURRENT_TIMESTAMP,
            last_seen_at TEXT DEFAULT CURRENT_TIMESTAMP,
            resolved_at TEXT,
            resolved_by TEXT,
            UNIQUE(business_id, anomaly_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_audit (
            id {ident},
            business_id INTEGER NOT NULL,
            operation TEXT NOT NULL,
            path TEXT NOT NULL,
            old_value_json TEXT NOT NULL DEFAULT 'null',
            new_value_json TEXT NOT NULL DEFAULT 'null',
            operator TEXT NOT NULL,
            changed_at TEXT DEFAULT CURRENT_TIMESTAMP,
            note TEXT
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_match_candidates (
            id {ident},
            source_type TEXT NOT NULL,
            source_record_key TEXT NOT NULL,
            source_version TEXT,
            reason TEXT NOT NULL,
            candidate_keys_json TEXT NOT NULL DEFAULT '[]',
            raw_json TEXT NOT NULL DEFAULT '{{}}',
            status TEXT NOT NULL DEFAULT 'open',
            first_seen_at TEXT DEFAULT CURRENT_TIMESTAMP,
            last_seen_at TEXT DEFAULT CURRENT_TIMESTAMP,
            resolved_at TEXT,
            UNIQUE(source_type, source_record_key)
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS order_lifecycle_sync_state (
            id {ident},
            wps_last_success_at TEXT,
            wps_last_error TEXT,
            email_last_success_at TEXT,
            email_last_error TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            CHECK (id = 1)
        )
        """,
    ]
    if db._is_pg():
        for statement in statements:
            cur.execute(statement)
        for name, col_type in (("port_status", "TEXT NOT NULL DEFAULT '待确认'"), ("port_confirmed_date", "TEXT"), ("port_confirmed_by", "TEXT"), ("shipment_status", "TEXT NOT NULL DEFAULT '待确认'"), ("shipment_confirmed_date", "TEXT"), ("shipment_confirmed_by", "TEXT"), ("next_follow_up_date", "TEXT"), ("source_active", "INTEGER NOT NULL DEFAULT 1"), ("settlement_status", "TEXT NOT NULL DEFAULT '待结算'"), ("settlement_date", "TEXT"), ("guo_danlei_special", "INTEGER NOT NULL DEFAULT 0"), ("completed_date", "TEXT")):
            cur.execute(f"ALTER TABLE order_lifecycle_businesses ADD COLUMN IF NOT EXISTS {name} {col_type}")
        cur.execute("ALTER TABLE order_lifecycle_customer_receipts ADD COLUMN IF NOT EXISTS fully_received INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE order_lifecycle_customer_receipts ADD COLUMN IF NOT EXISTS applicable_scope TEXT")
        cur.execute("ALTER TABLE order_lifecycle_vessels ADD COLUMN IF NOT EXISTS latest_shipment_date TEXT")
        cur.execute("ALTER TABLE order_lifecycle_bank_repayments ADD COLUMN IF NOT EXISTS currency TEXT")
        cur.execute("ALTER TABLE order_lifecycle_bank_repayments ADD COLUMN IF NOT EXISTS completion_explicit INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE order_lifecycle_financings ADD COLUMN IF NOT EXISTS currency TEXT")
    else:
        for statement in statements:
            conn.execute(statement)
        existing_columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_lifecycle_businesses)").fetchall()}
        for name, col_type in (("port_status", "TEXT NOT NULL DEFAULT '待确认'"), ("port_confirmed_date", "TEXT"), ("port_confirmed_by", "TEXT"), ("shipment_status", "TEXT NOT NULL DEFAULT '待确认'"), ("shipment_confirmed_date", "TEXT"), ("shipment_confirmed_by", "TEXT"), ("next_follow_up_date", "TEXT"), ("source_active", "INTEGER NOT NULL DEFAULT 1"), ("settlement_status", "TEXT NOT NULL DEFAULT '待结算'"), ("settlement_date", "TEXT"), ("guo_danlei_special", "INTEGER NOT NULL DEFAULT 0"), ("completed_date", "TEXT")):
            if name not in existing_columns:
                conn.execute(f"ALTER TABLE order_lifecycle_businesses ADD COLUMN {name} {col_type}")
        receipt_columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_lifecycle_customer_receipts)").fetchall()}
        for name, col_type in (("fully_received", "INTEGER NOT NULL DEFAULT 0"), ("applicable_scope", "TEXT")):
            if name not in receipt_columns:
                conn.execute(f"ALTER TABLE order_lifecycle_customer_receipts ADD COLUMN {name} {col_type}")
        vessel_columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_lifecycle_vessels)").fetchall()}
        if "latest_shipment_date" not in vessel_columns:
            conn.execute("ALTER TABLE order_lifecycle_vessels ADD COLUMN latest_shipment_date TEXT")
        repayment_columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_lifecycle_bank_repayments)").fetchall()}
        for name, col_type in (("currency", "TEXT"), ("completion_explicit", "INTEGER NOT NULL DEFAULT 0")):
            if name not in repayment_columns:
                conn.execute(f"ALTER TABLE order_lifecycle_bank_repayments ADD COLUMN {name} {col_type}")
        financing_columns = {row["name"] for row in conn.execute("PRAGMA table_info(order_lifecycle_financings)").fetchall()}
        if "currency" not in financing_columns:
            conn.execute("ALTER TABLE order_lifecycle_financings ADD COLUMN currency TEXT")
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_ol_business_type_status ON order_lifecycle_businesses(business_type, status)",
        "CREATE INDEX IF NOT EXISTS idx_ol_business_risk ON order_lifecycle_businesses(risk_level)",
        "CREATE INDEX IF NOT EXISTS idx_ol_business_source ON order_lifecycle_businesses(source_type, source_record_key)",
        "CREATE INDEX IF NOT EXISTS idx_ol_contract_business ON order_lifecycle_contracts(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_financing_business ON order_lifecycle_financings(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_vessel_business ON order_lifecycle_vessels(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_document_business ON order_lifecycle_documents(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_receipt_business ON order_lifecycle_customer_receipts(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_repayment_business ON order_lifecycle_bank_repayments(business_id)",
        "CREATE INDEX IF NOT EXISTS idx_ol_source_record_business ON order_lifecycle_source_records(business_key, source_type)",
        "CREATE INDEX IF NOT EXISTS idx_ol_business_source_membership ON order_lifecycle_business_sources(business_id, source_active)",
        "CREATE INDEX IF NOT EXISTS idx_ol_source_membership_active ON order_lifecycle_business_sources(source_type, source_active)",
        "CREATE INDEX IF NOT EXISTS idx_ol_anomaly_business_status ON order_lifecycle_data_anomalies(business_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_ol_child_override_business ON order_lifecycle_child_overrides(business_id, collection, source_key)",
        "CREATE INDEX IF NOT EXISTS idx_ol_manual_child_business ON order_lifecycle_manual_child_records(business_id, collection, is_active)",
        "CREATE INDEX IF NOT EXISTS idx_ol_audit_business_time ON order_lifecycle_audit(business_id, changed_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_ol_match_candidate_status ON order_lifecycle_match_candidates(status, source_type)",
    ]
    for statement in indexes:
        cur.execute(statement) if db._is_pg() else conn.execute(statement)
    if db._is_pg():
        cur.execute(
            "INSERT INTO order_lifecycle_sync_state (id) VALUES (1) ON CONFLICT (id) DO NOTHING"
        )
        db._secure_postgres_tables(cur, TABLES)
    else:
        conn.execute("INSERT OR IGNORE INTO order_lifecycle_sync_state (id) VALUES (1)")
        # ``mixed`` is an aggregate display value, not an authoritative source
        # membership.  Deactivate stale rows left by the earlier test schema.
        conn.execute("UPDATE order_lifecycle_business_sources SET source_active = 0 WHERE LOWER(source_type) = 'mixed'")
        conn.execute(
            """
            INSERT OR IGNORE INTO order_lifecycle_business_sources
                (business_id, source_type, source_business_key, source_record_key, source_version, source_presence_hash, source_active)
            SELECT id, LOWER(source_type), business_key, source_record_key, source_version, source_presence_hash, source_active
            FROM order_lifecycle_businesses
            WHERE business_key IS NOT NULL AND source_type IS NOT NULL
              AND LOWER(source_type) IN ('wps', 'email')
              AND NOT EXISTS (
                  SELECT 1 FROM order_lifecycle_business_sources s
                  WHERE s.business_id = order_lifecycle_businesses.id
                    AND s.source_type = LOWER(order_lifecycle_businesses.source_type)
              )
            """
        )
    if db._is_pg():
        cur.execute("UPDATE order_lifecycle_business_sources SET source_active = 0 WHERE LOWER(source_type) = 'mixed'")
        cur.execute(
            """
            INSERT INTO order_lifecycle_business_sources
                (business_id, source_type, source_business_key, source_record_key, source_version, source_presence_hash, source_active)
            SELECT id, LOWER(source_type), business_key, source_record_key, source_version, source_presence_hash, source_active
            FROM order_lifecycle_businesses
            WHERE business_key IS NOT NULL AND source_type IS NOT NULL
              AND LOWER(source_type) IN ('wps', 'email')
              AND NOT EXISTS (
                  SELECT 1 FROM order_lifecycle_business_sources s
                  WHERE s.business_id = order_lifecycle_businesses.id
                    AND s.source_type = LOWER(order_lifecycle_businesses.source_type)
              )
            ON CONFLICT (source_type, source_business_key) DO NOTHING
            """
        )
        cur.execute(
            "UPDATE order_lifecycle_businesses SET source_active = CASE WHEN EXISTS (SELECT 1 FROM order_lifecycle_business_sources s WHERE s.business_id = order_lifecycle_businesses.id AND s.source_active = 1 AND LOWER(s.source_type) IN ('wps', 'email')) THEN 1 ELSE 0 END WHERE is_cancelled = 0"
        )
    else:
        conn.execute(
            "UPDATE order_lifecycle_businesses SET source_active = CASE WHEN EXISTS (SELECT 1 FROM order_lifecycle_business_sources s WHERE s.business_id = order_lifecycle_businesses.id AND s.source_active = 1 AND LOWER(s.source_type) IN ('wps', 'email')) THEN 1 ELSE 0 END WHERE is_cancelled = 0"
        )
    conn.commit()


def sync_order_lifecycle_permissions(cur) -> None:
    """Copy existing order-finance permissions for the new page only when missing."""
    users = db._exec(cur, "SELECT id FROM users").fetchall()
    for user in users:
        old = db._exec(
            cur,
            "SELECT can_view, can_edit, can_sensitive FROM module_permissions WHERE user_id = ? AND module_code = ?",
            (user["id"], "order_finance_progress"),
        ).fetchone()
        if not old:
            continue
        db._exec(
            cur,
            """
            INSERT OR IGNORE INTO module_permissions
                (user_id, module_code, can_view, can_edit, can_sensitive)
            VALUES (?, ?, ?, ?, ?)
            """,
            (user["id"], ORDER_LIFECYCLE_MODULE, old["can_view"], old["can_edit"], old["can_sensitive"]),
        )


def _compact(value: Any) -> str:
    return re.sub(r"[\s\-_/—–]+", "", _normalize_text(value)).upper()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _first_nonempty(values: Iterable[Any]) -> Any:
    for value in values:
        if _normalize_text(value):
            return value
    return None


def _short_supplier(value: Any) -> str:
    text = _normalize_text(value)
    aliases = (
        ("北满", "北满"), ("BEIMAN", "北满"), ("东钢", "东钢"), ("DONG", "东钢"),
        ("承德", "承德"), ("CHENGDE", "承德"), ("抚顺", "抚顺"), ("FUSHUN", "抚顺"),
        ("西林", "西林"), ("XILIN", "西林"), ("阿城", "阿城"), ("ACHENG", "阿城"),
    )
    upper = text.upper()
    for needle, label in aliases:
        if needle in text or needle in upper:
            return label
    return text


def _header_text(value: Any) -> str:
    return re.sub(r"\s+", "", _normalize_text(value)).replace("\n", "")


def _header_index(headers: list[Any], aliases: tuple[str, ...], start: int = 0, end: Optional[int] = None) -> Optional[int]:
    end = len(headers) if end is None else min(end, len(headers))
    normalized = [_header_text(value) for value in headers]
    for alias in aliases:
        alias_text = _header_text(alias)
        for index in range(start, end):
            if alias_text and alias_text in normalized[index]:
                return index
    return None


def _row_value(row: list[Any], headers: list[Any], aliases: tuple[str, ...], start: int = 0, end: Optional[int] = None) -> Any:
    return _get(row, _header_index(headers, aliases, start, end))


def _normalize_business_no(value: Any) -> str:
    """Normalize a real business identifier without inventing one."""
    text = _normalize_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).replace("－", "-").replace("—", "-").replace("–", "-").upper()


def _is_legacy_mill_row_business_no(value: Any) -> bool:
    """Return True for the old steel-mill-plus-row labels, never for real IDs."""
    text = _normalize_business_no(value)
    if not text or re.search(r"-\d{4}-\d+", text):
        return False
    number = r"(?:\d+|[零〇一二两三四五六七八九十百千万]+)"
    return bool(
        re.fullmatch(rf"[\u4e00-\u9fffA-Z]+杠{number}", text)
        or re.fullmatch(rf"杠{number}", text)
        or re.fullmatch(rf"[\u4e00-\u9fffA-Z]+-\d+", text)
    )


def _contract_identity(item: dict[str, Any]) -> str:
    values = (
        _compact(item.get("purchase_contract_no")),
        _compact(item.get("system_contract_no")),
        _compact(item.get("contract_no")),
    )
    return "|".join(value for value in values if value)


def _contract_child_identity(item: dict[str, Any]) -> str:
    """Use the stable purchase/system pair before source-specific display numbers."""
    purchase = _compact(item.get("purchase_contract_no"))
    system = _compact(item.get("system_contract_no"))
    if purchase or system:
        return "|".join(value for value in (purchase, system) if value)
    return _compact(item.get("contract_no"))


def _record_contract_identities(record: dict[str, Any]) -> list[str]:
    identities: list[str] = []
    for item in record.get("contracts", []):
        identity = _contract_identity(item)
        if identity and identity not in identities:
            identities.append(identity)
    return identities


def _record_purchase_contract(record: dict[str, Any]) -> str:
    for item in record.get("contracts", []):
        value = _normalize_business_no(item.get("purchase_contract_no"))
        if value:
            return value
    return ""


def _source_business_identity(record: dict[str, Any]) -> str:
    source_type = _normalize_text(record.get("source_type")).lower()
    if source_type == "wps":
        business_no = _normalize_business_no(record.get("business_no"))
        return f"business:wps:{_compact(business_no)}" if business_no else ""
    explicit = _normalize_business_no(record.get("business_no"))
    purchase = _record_purchase_contract(record)
    if record.get("business_type") == "过单":
        value = purchase or explicit
        return f"business:pass:{_compact(value)}" if value else ""
    contract_identity = _record_contract_identities(record)[0] if _record_contract_identities(record) else ""
    value = purchase or explicit or contract_identity
    return f"source:email:{_compact(value)}" if value else ""


def _business_key(source_type: str, sheet: str, item: str, purchase: Any, system: Any, row_no: int) -> str:
    """Return a source-local grouping key; the parent key is resolved later."""
    identity = _compact(item) or _compact(purchase) or _compact(system) or str(row_no)
    return f"{source_type}:{_compact(sheet)}:{identity}"


def _source_record_key(source_type: str, sheet: str, row_no: int, item: str) -> str:
    # Row numbers are retained in raw evidence but never form the business identity.
    return f"{source_type}:{_compact(sheet)}:{_compact(item) or str(row_no)}"


def _source_child_key(kind: str, item: dict[str, Any]) -> str:
    if kind == "contract":
        identity = _contract_child_identity(item)
        return f"contract:{identity}" if identity else _normalize_text(item.get("source_key"))
    if kind == "financing":
        values = (
            _compact(item.get("bank")),
            str(item.get("amount") if item.get("amount") is not None else ""),
            _normalize_date(item.get("financing_date")),
            _normalize_date(item.get("original_due_date")),
            _normalize_date(item.get("extended_due_date")),
        )
        return "finance:" + ":".join(values) if any(values) else _normalize_text(item.get("source_key"))
    if kind == "vessel":
        values = (
            _compact(item.get("imo")),
            _compact(item.get("vessel_name")),
            _compact(item.get("loading_port")),
            _compact(item.get("discharge_port")),
        )
        return "vessel:" + ":".join(values) if any(values) else _normalize_text(item.get("source_key"))
    if kind == "document":
        document_date = _normalize_date(item.get("document_date"))
        return f"document:{_compact(item.get('document_type') or '交单')}:{document_date}" if document_date else _normalize_text(item.get("source_key"))
    if kind == "receipt":
        values = (_normalize_date(item.get("receipt_date")), str(item.get("amount") if item.get("amount") is not None else ""), _compact(item.get("currency")))
        return "receipt:" + ":".join(values) if any(values) else _normalize_text(item.get("source_key"))
    if kind == "bank_repayment":
        values = (_normalize_date(item.get("repayment_date")), str(item.get("amount") if item.get("amount") is not None else ""), _compact(item.get("financing_source_key")))
        return "bank_repayment:" + ":".join(values) if any(values) else _normalize_text(item.get("source_key"))
    return _normalize_text(item.get("source_key"))


def _merge_unique_children(target: dict[str, Any], incoming: dict[str, Any]) -> None:
    for collection, kind in (
        ("contracts", "contract"),
        ("financings", "financing"),
        ("vessels", "vessel"),
        ("documents", "document"),
        ("customer_receipts", "receipt"),
        ("bank_repayments", "bank_repayment"),
    ):
        current = target.setdefault(collection, [])
        by_key = {_source_child_key(kind, item): item for item in current}
        for item in incoming.get(collection, []):
            key = _source_child_key(kind, item)
            if not key:
                key = _normalize_text(item.get("source_key")) or uuid4().hex
            if key in by_key:
                existing = by_key[key]
                for field, value in item.items():
                    if field in {"fully_received", "completion_explicit"} and value:
                        existing[field] = value
                    elif value not in (None, "", []) and existing.get(field) in (None, "", []):
                        existing[field] = value
            else:
                copied = deepcopy(item)
                copied["source_key"] = key
                current.append(copied)
                by_key[key] = copied


def _merge_source_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Deduplicate repeated continuation rows by exact contract identity."""
    grouped: dict[str, dict[str, Any]] = {}
    for record in records:
        identity = _source_business_identity(record) or _normalize_text(record.get("source_record_key"))
        if identity not in grouped:
            grouped[identity] = deepcopy(record)
            grouped[identity]["business_key"] = identity
            grouped[identity]["source_business_key"] = identity
            continue
        current = grouped[identity]
        for field in ("business_no", "trade_entity", "supplier_steel_mill", "terminal_customer", "product_name", "contract_quantity_mt"):
            if current.get(field) in (None, "") and record.get(field) not in (None, ""):
                current[field] = record[field]
        _merge_unique_children(current, record)
        raw_rows = current.setdefault("raw", {}).setdefault("rows", [])
        if record.get("raw") and record.get("raw") not in raw_rows:
            raw_rows.append(record["raw"])
    return list(grouped.values())


def _empty_record() -> dict[str, Any]:
    return {
        "business_type": "融资",
        "business_no": "",
        "business_key": "",
        "trade_entity": "",
        "supplier_steel_mill": "",
        "terminal_customer": "",
        "product_name": "",
        "contract_quantity_mt": None,
        "source_type": "",
        "source_snapshot_date": date.today().isoformat(),
        "source_version": "",
        "source_record_key": "",
        "contracts": [],
        "financings": [],
        "vessels": [],
        "documents": [],
        "customer_receipts": [],
        "bank_repayments": [],
        "settlement_status": "待结算",
        "settlement_date": "",
        "guo_danlei_special": False,
        "completed_date": "",
        "next_follow_up_date": "",
        "raw": {},
    }


def parse_wps_workbook(path: Path | str) -> dict[str, Any]:
    """Parse only the three authoritative WPS raw sheets into standard records."""
    path = Path(path)
    if not path.exists():
        raise ValueError(f"WPS文件不存在：{path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    snapshot_date = datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()
    version = _hash_file(path)
    records: list[dict[str, Any]] = []
    for sheet_name in WPS_RAW_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header_row_no = None
        headers: list[Any] = []
        for row_no, values in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), 1):
            row_headers = list(values)
            if _header_index(row_headers, ("项次",)) is not None:
                header_row_no, headers = row_no, row_headers
                break
        if not header_row_no:
            continue
        groups: dict[str, dict[str, Any]] = {}
        current_item = ""
        for row_no, values in enumerate(sheet.iter_rows(min_row=header_row_no + 1, values_only=True), header_row_no + 1):
            row = list(values)
            item = _normalize_text(_row_value(row, headers, ("项次",))) or current_item
            if not item or _compact(item) in {"TOTAL", "合计"}:
                continue
            current_item = item
            purchase = _row_value(row, headers, ("合同号",), 0, 24)
            system = _row_value(row, headers, ("系统合同号",), 0, 24)
            key = _business_key("wps", sheet_name, item, purchase, system, row_no)
            record = groups.setdefault(key, _empty_record())
            record.update({
                "business_type": "融资",
                "business_no": _normalize_business_no(item),
                "business_key": key,
                "trade_entity": {"YOLANDA": "YOLANDA", "JLHK": "香港建龙", "天津建龙": "天津建龙"}.get(sheet_name, sheet_name),
                "source_type": "wps",
                "source_snapshot_date": snapshot_date,
                "source_version": version,
                "source_record_key": _source_record_key("wps", sheet_name, row_no, item),
                "product_name": _first_nonempty((record.get("product_name"), _row_value(row, headers, ("品名", "合同品名")))),
                "supplier_steel_mill": _short_supplier(_first_nonempty((record.get("supplier_steel_mill"), _row_value(row, headers, ("供应商",))))),
                "terminal_customer": _first_nonempty((record.get("terminal_customer"), _row_value(row, headers, ("合同买方",), 24))),
                "contract_quantity_mt": _first_nonempty((record.get("contract_quantity_mt"), _to_float(_row_value(row, headers, ("合同数量",))))),
                "raw": {"sheet": sheet_name, "row_no": row_no, "row": row, "headers": headers},
            })
            purchase = _normalize_text(purchase)
            system = _normalize_text(system)
            sales_system = _normalize_text(_row_value(row, headers, ("系统合同号",), 24))
            sales_contract = _normalize_text(_row_value(row, headers, ("双方合同号",), 24))
            buyer = _normalize_text(_row_value(row, headers, ("合同买方",), 24))
            seller = _normalize_text(_row_value(row, headers, ("供应商",), 0, 24))
            if any((purchase, system, sales_system, sales_contract, buyer, seller)):
                contract_key = ":".join((str(row_no), purchase, system, sales_system, sales_contract))
                if not any(item.get("source_key") == contract_key for item in record["contracts"]):
                    record["contracts"].append({
                        "contract_no": sales_contract or system,
                        "purchase_contract_no": purchase,
                        "system_contract_no": system or sales_system,
                        "buyer": buyer,
                        "seller": seller,
                        "quantity_mt": _to_float(_row_value(row, headers, ("合同数量",))),
                        "source_key": contract_key,
                    })
            bank = _normalize_text(_row_value(row, headers, ("贷款行",), 0, 24))
            amount = _to_float(_row_value(row, headers, ("贷款人民币金额",), 0, 24))
            financing_date = _normalize_date(_row_value(row, headers, ("借款日期",), 0, 24))
            original_due = _normalize_date(_row_value(row, headers, ("原到期日",), 0, 24))
            extended_due = _normalize_date(_row_value(row, headers, ("新到期日",), 0, 24))
            repayment_date = _normalize_date(_row_value(row, headers, ("还款日",), 0, 24))
            loan_status = _normalize_text(_row_value(row, headers, ("贷款状态",), 0, 24))
            finance_key = f"{row_no}:{bank}:{amount}:{financing_date}"
            if bank or amount is not None or financing_date or original_due or extended_due or repayment_date or loan_status:
                if not any(item.get("source_key") == finance_key for item in record["financings"]):
                    record["financings"].append({
                        "bank": bank,
                        "amount": amount,
                        "currency": "CNY" if amount is not None else "",
                        "financing_date": financing_date,
                        "original_due_date": original_due,
                        "extended_due_date": extended_due,
                        "repayment_date": repayment_date,
                        "repayment_status": loan_status,
                        "source_key": finance_key,
                    })
            if repayment_date or "已还" in loan_status:
                record["bank_repayments"].append({
                    "repayment_date": repayment_date,
                    "amount": amount,
                    "currency": "CNY",
                    "financing_source_key": finance_key,
                    "source_key": f"bank-repayment:{finance_key}",
                    "completion_explicit": True,
                })
            vessel = _normalize_text(_row_value(row, headers, ("船名",), 24))
            if vessel:
                record["vessels"].append({
                    "vessel_name": vessel,
                    "imo": "",
                    "loading_port": _normalize_text(_row_value(row, headers, ("起运港",))),
                    "discharge_port": _normalize_text(_row_value(row, headers, ("目的港",))),
                    "eta": "",
                    "etb": "",
                    "estimated_discharge_date": "",
                    "latest_shipment_date": _normalize_date(_row_value(row, headers, ("最迟装船日", "最迟装船期"), 24)),
                    "source_key": f"vessel:{row_no}:{vessel}",
                    "source": "wps",
                })
            document_date = _normalize_date(_row_value(row, headers, ("交单日期",), 24))
            if document_date:
                record["documents"].append({"document_type": "交单", "document_date": document_date, "source_key": f"doc:{row_no}:{document_date}"})
            receipt_date = _normalize_date(_row_value(row, headers, ("收汇日期",), 24))
            if receipt_date:
                receipt_note = _normalize_text(_row_value(row, headers, ("情况说明", "备注"), 44))
                record["customer_receipts"].append({
                    "receipt_date": receipt_date,
                    "amount": _to_float(_row_value(row, headers, ("交单金额",), 24)),
                    "currency": _normalize_text(_row_value(row, headers, ("合同币别",), 24)),
                    "source_key": f"receipt:{row_no}:{receipt_date}",
                    "fully_received": "已收款" in receipt_note or "已回款" in receipt_note,
                    "applicable_scope": _contract_child_identity({"purchase_contract_no": purchase, "system_contract_no": system or sales_system}),
                })
        records.extend(groups.values())
    records = _merge_source_records(records)
    return {
        "source_type": "wps",
        "source_locator": str(path),
        "source_version": version,
        "snapshot_date": snapshot_date,
        "source_hash": version,
        "records": records,
        "summary": {"record_count": len(records), "sheet_count": len({r["raw"].get("sheet") for r in records})},
    }


def _mail_xlsx_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row_no = None
    headers: list[Any] = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), 1):
        row_headers = list(values)
        if _header_index(row_headers, ("序号", "货物品名", "货物")) is not None:
            header_row_no, headers = row_no, row_headers
            break
    if not header_row_no:
        return []
    records = []
    current = ""
    for row_no, values in enumerate(sheet.iter_rows(min_row=header_row_no + 1, values_only=True), header_row_no + 1):
        row = list(values)
        sequence = _normalize_text(_row_value(row, headers, ("序号",))) or current
        if not sequence:
            continue
        current = sequence
        records.append(_mail_row_record(path, sheet.title, headers, row, row_no, sequence))
    return records


def _mail_row_record(path: Path, sheet_name: str, headers: list[Any], row: list[Any], row_no: int, sequence: str) -> dict[str, Any]:
    subsidiary = _subsidiary_from_filename(path.name)
    product = _normalize_text(_row_value(row, headers, ("货物品名", "货物")))
    purchase = _normalize_text(_row_value(row, headers, ("采/销合同号", "合同号")))
    system = _normalize_text(_row_value(row, headers, ("YOLANDA合同号", "Yolanda采/销合同号", "Yolanda/Jianlong采/销合同号")))
    amount = _to_float(_row_value(row, headers, ("实际放款金额", "融资金额", "应放款金额", "放款金额")))
    financing_date = _normalize_date(_row_value(row, headers, ("放款日期",)))
    bank = _normalize_text(_row_value(row, headers, ("贷款行", "融资银行", "银行")))
    business_type = "融资" if any((amount is not None, financing_date, bank)) else "过单"
    business_key = _business_key("email", subsidiary, sequence, purchase, system, row_no)
    source_identity = _normalize_business_no(purchase or system) or sequence
    contract_scope = _contract_child_identity({"purchase_contract_no": purchase, "system_contract_no": system})
    receipt_date = _normalize_date(_row_value(row, headers, ("回款日期",)))
    doc_date = _normalize_date(_row_value(row, headers, ("交单日期",)))
    vessel = _normalize_text(_row_value(row, headers, ("船名航次", "船名")))
    record = _empty_record()
    record.update({
        "business_type": business_type,
        "business_no": _normalize_business_no(purchase) if business_type == "过单" else "",
        "business_key": business_key,
        "trade_entity": "YOLANDA",
        "supplier_steel_mill": subsidiary,
        "terminal_customer": _normalize_text(_row_value(row, headers, ("买方",))),
        "product_name": product,
        "contract_quantity_mt": _to_float(_row_value(row, headers, ("合同数量",))),
        "source_type": "email",
        "source_snapshot_date": datetime.fromtimestamp(path.stat().st_mtime).date().isoformat(),
        "source_record_key": _source_record_key("email", subsidiary, row_no, source_identity),
        "settlement_status": "待结算",
        "settlement_date": "",
        "raw": {"sheet": sheet_name, "row": row, "headers": headers, "file": path.name},
    })
    if purchase or system:
        record["contracts"].append({
            "contract_no": purchase or system,
            "purchase_contract_no": purchase,
            "system_contract_no": system,
            "buyer": _normalize_text(_row_value(row, headers, ("买方",))),
            "seller": _normalize_text(_row_value(row, headers, ("卖方",))),
            "quantity_mt": record["contract_quantity_mt"],
            "source_key": f"contract:{row_no}:{purchase}:{system}",
        })
    if record["business_type"] == "融资":
        record["financings"].append({
            "bank": bank,
            "amount": amount,
            "currency": _normalize_text(_row_value(row, headers, ("融资币种", "贷款币种", "币种"))) or ("CNY" if amount is not None else ""),
            "financing_date": financing_date,
            "original_due_date": _normalize_date(_row_value(row, headers, ("放款到期日期", "融资到期日"))),
            "extended_due_date": _normalize_date(_row_value(row, headers, ("新到期日",))),
            "repayment_date": _normalize_date(_row_value(row, headers, ("还款日期", "还款日"))),
            "repayment_status": _normalize_text(_row_value(row, headers, ("贷款状态", "状态"))),
            "source_key": f"finance:{row_no}:{bank}:{amount}:{financing_date}",
        })
        repayment_date = _normalize_date(_row_value(row, headers, ("还款日期", "还款日")))
        repayment_status = _normalize_text(_row_value(row, headers, ("贷款状态", "状态")))
        if repayment_date or "已还" in repayment_status:
            record["bank_repayments"].append({
                "repayment_date": repayment_date,
                "amount": amount,
                "currency": "CNY",
                "financing_source_key": f"finance:{row_no}:{bank}:{amount}:{financing_date}",
                "source_key": f"bank-repayment:{row_no}:{bank}:{amount}:{financing_date}",
                "completion_explicit": True,
            })
    if vessel:
        record["vessels"].append({
            "vessel_name": vessel,
            "imo": "",
            "loading_port": _normalize_text(_row_value(row, headers, ("起运港",))),
            "discharge_port": _normalize_text(_row_value(row, headers, ("目的港", "卸港"))),
            "eta": "",
            "etb": "",
            "estimated_discharge_date": "",
            "latest_shipment_date": "",
            "source_key": f"vessel:{row_no}:{vessel}",
            "source": "email",
        })
    if doc_date:
        record["documents"].append({"document_type": "交单", "document_date": doc_date, "source_key": f"doc:{row_no}:{doc_date}"})
    if receipt_date:
        record["customer_receipts"].append({"receipt_date": receipt_date, "amount": None, "currency": "", "source_key": f"receipt:{row_no}:{receipt_date}", "fully_received": False, "applicable_scope": contract_scope})
    return record


def parse_email_batch(directory: Path | str) -> dict[str, Any]:
    """Parse a complete six-mill email batch; missing attachment blocks the batch."""
    base = Path(directory)
    if not base.exists() or not base.is_dir():
        raise ValueError(f"邮件台账目录不存在：{base}")
    files = [p for p in sorted(base.iterdir()) if p.suffix.lower() in {".xls", ".xlsx"} and not p.name.startswith("~$")]
    found = {mill for mill in MAIL_MILLS if any(mill in path.name for path in files)}
    missing = [mill for mill in MAIL_MILLS if mill not in found]
    if missing:
        raise ValueError(f"邮件台账批次不完整，缺少：{'、'.join(missing)}")
    records: list[dict[str, Any]] = []
    file_results = []
    for path in files:
        if path.suffix.lower() == ".xls":
            result = parse_order_finance_workbook(path)
            parsed = [_standardize_legacy_mail_record(item, path) for item in result.get("records", [])]
        else:
            parsed = _mail_xlsx_records(path)
        records.extend(parsed)
        file_results.append({"file": path.name, "record_count": len(parsed)})
    records = _merge_source_records(records)
    version = hashlib.sha256("|".join(f"{p.name}:{_hash_file(p)}" for p in files).encode()).hexdigest()
    snapshot_date = max(datetime.fromtimestamp(p.stat().st_mtime).date().isoformat() for p in files)
    return {
        "source_type": "email",
        "source_locator": str(base),
        "source_version": version,
        "snapshot_date": snapshot_date,
        "source_hash": version,
        "records": records,
        "files": file_results,
        "summary": {"record_count": len(records), "files_read": len(files)},
    }


def _standardize_legacy_mail_record(item: dict[str, Any], path: Path) -> dict[str, Any]:
    sequence = _normalize_text(item.get("business_key")) or str(item.get("source_row_start") or "")
    purchase = _normalize_text(item.get("purchase_contract_no"))
    system = _normalize_text(item.get("system_contract_no"))
    business_type = "融资" if any((item.get("finance_amount_actual") is not None, item.get("finance_drawdown_date"), item.get("finance_bank"))) else "过单"
    item_no = _normalize_business_no(purchase) if business_type == "过单" else ""
    source_identity = _normalize_business_no(purchase or system) or sequence
    main_contract_scope = _contract_child_identity({"purchase_contract_no": purchase, "system_contract_no": system})
    raw_status = _normalize_text(item.get("business_status"))
    raw_remark = _normalize_text(item.get("remark"))
    settlement = _safe_json(item.get("settlement_json"), {})
    settlement_date = ""
    if isinstance(settlement, dict):
        settlement_date = _normalize_date(settlement.get("settlement_date") or settlement.get("date") or settlement.get("结算日期"))
    key = _business_key("email", _subsidiary_from_filename(path.name), item_no, purchase, system, int(item.get("source_row_start") or 0))
    record = _empty_record()
    record.update({
        "business_type": business_type,
        "business_no": item_no,
        "business_key": key,
        "trade_entity": "YOLANDA",
        "supplier_steel_mill": _subsidiary_from_filename(path.name),
        "terminal_customer": _normalize_text(item.get("terminal_customer") or item.get("buyer")),
        "product_name": _normalize_text(item.get("product_name")),
        "contract_quantity_mt": item.get("contract_quantity_mt"),
        "source_type": "email",
        "source_snapshot_date": item.get("source_snapshot_date") or date.today().isoformat(),
        "source_record_key": _source_record_key("email", path.name, int(item.get("source_row_start") or 0), source_identity),
        "settlement_status": "已结算" if "已结算" in raw_status or "已结算" in raw_remark else "待结算",
        "settlement_date": settlement_date,
        "next_follow_up_date": _normalize_date(item.get("next_follow_up_date")),
        "raw": item,
    })
    if purchase or system:
        record["contracts"].append({
            "contract_no": purchase or system,
            "purchase_contract_no": purchase,
            "system_contract_no": system,
            "buyer": _normalize_text(item.get("buyer")),
            "seller": _normalize_text(item.get("seller")),
            "quantity_mt": item.get("contract_quantity_mt"),
            "source_key": f"contract:{item.get('source_row_start') or 0}:{purchase}:{system}",
        })
    for child_index, child in enumerate(_safe_json(item.get("sales_contracts_json"), []), 1):
        child_purchase = _normalize_text(child.get("contract"))
        child_system = _normalize_text(child.get("system_contract"))
        if not child_purchase and not child_system:
            continue
        record["contracts"].append({
            "contract_no": child_purchase or child_system,
            "purchase_contract_no": child_purchase,
            "system_contract_no": child_system,
            "buyer": _normalize_text(child.get("buyer")),
            "seller": _normalize_text(child.get("seller")),
            "quantity_mt": None,
            "source_key": f"sales-contract:{child_index}:{child_purchase}:{child_system}",
        })
    if record["business_type"] == "融资":
        finance_key = f"finance:{item.get('source_row_start') or 0}:{item.get('finance_bank')}:{item.get('finance_amount_actual')}"
        record["financings"].append({
            "bank": _normalize_text(item.get("finance_bank")),
            "amount": item.get("finance_amount_actual") if item.get("finance_amount_actual") is not None else item.get("finance_amount_expected"),
            "currency": _normalize_text(item.get("finance_currency") or item.get("currency")) or ("CNY" if item.get("finance_amount_actual") is not None or item.get("finance_amount_expected") is not None else ""),
            "financing_date": _normalize_date(item.get("finance_drawdown_date")),
            "original_due_date": _normalize_date(item.get("finance_due_date")),
            "extended_due_date": _normalize_date(item.get("finance_due_date")),
            "repayment_date": _normalize_date(item.get("tail_payment_date")) if "还款" in _normalize_text(item.get("remark")) else "",
            "repayment_status": "已还款" if "还款" in _normalize_text(item.get("remark")) else "",
            "source_key": finance_key,
        })
        repayment_date = _normalize_date(item.get("tail_payment_date")) if "还款" in raw_remark else ""
        if repayment_date or "已还" in raw_remark:
            record["bank_repayments"].append({
                "repayment_date": repayment_date,
                "amount": item.get("finance_amount_actual"),
                "currency": "CNY",
                "financing_source_key": finance_key,
                "source_key": f"bank-repayment:{finance_key}",
                "completion_explicit": True,
            })
    for child_index, child in enumerate(_safe_json(item.get("sales_contracts_json"), []), 1):
        child_collection = _normalize_date(child.get("collection_date"))
        if child_collection:
            record["customer_receipts"].append({"receipt_date": child_collection, "amount": child.get("amount"), "currency": child.get("currency", ""), "source_key": f"receipt:{child_index}:{child_collection}", "fully_received": True, "applicable_scope": _contract_child_identity({"purchase_contract_no": child.get("contract"), "system_contract_no": child.get("system_contract")})})
    collection = _normalize_date(item.get("collection_date"))
    if collection:
        record["customer_receipts"].append({"receipt_date": collection, "amount": None, "currency": "", "source_key": f"receipt:main:{collection}", "fully_received": raw_status in {"已回款待结算", "已结算", "已回款"} or "已结算" in raw_remark, "applicable_scope": main_contract_scope})
    document = _normalize_date(item.get("document_submission_date"))
    if document:
        record["documents"].append({"document_type": "交单", "document_date": document, "source_key": f"doc:main:{document}"})
    if record["settlement_status"] == "已结算" and not record["settlement_date"]:
        record["settlement_date"] = collection
    vessel = _normalize_text(item.get("vessel_voyage"))
    if vessel:
        record["vessels"].append({
            "vessel_name": vessel,
            "imo": "",
            "loading_port": _normalize_text(item.get("origin_port")),
            "discharge_port": _normalize_text(item.get("destination_port")),
            "eta": "",
            "etb": "",
            "estimated_discharge_date": "",
            "latest_shipment_date": _normalize_date(item.get("latest_shipment_date")),
            "source_key": f"vessel:main:{vessel}",
            "source": "email",
        })
    return record


def _safe_json(value: Any, default: Any) -> Any:
    if isinstance(value, (list, dict)):
        return value
    try:
        parsed = json.loads(value or "")
        return parsed if parsed is not None else default
    except (TypeError, ValueError):
        return default


def _natural_sort_key(value: Any) -> str:
    text = _normalize_text(value).upper()
    return re.sub(r"\d+", lambda match: f"{int(match.group(0)):020d}", text)


def _completion_date(record: dict[str, Any], status: str) -> Optional[str]:
    if status == "已结算":
        return _normalize_date(record.get("settlement_date")) or None
    if status != "已完结":
        return None
    if record.get("business_type") == "融资":
        dates = [
            _normalize_date(item.get("repayment_date"))
            for item in record.get("bank_repayments", []) + record.get("financings", [])
            if item.get("repayment_date")
        ]
    else:
        dates = [
            _normalize_date(item.get("receipt_date"))
            for item in record.get("customer_receipts", [])
            if item.get("fully_received") and item.get("receipt_date")
        ]
    dates = [item for item in dates if item]
    return max(dates) if dates else None


def _weekly_focus_reasons(item: dict[str, Any]) -> list[str]:
    if item.get("status") in {"已完结", "已结算"}:
        return []
    reasons: list[str] = []
    if item.get("risk_level") == "高风险":
        reasons.append("high_risk")
    shipment_dates = [_parse_date(row.get("latest_shipment_date")) for row in item.get("vessels", []) if row.get("latest_shipment_date")]
    shipment_dates = [value for value in shipment_dates if value]
    if shipment_dates and item.get("shipment_status") != "已装船" and min((value - date.today()).days for value in shipment_dates) <= 10:
        reasons.append("shipment_follow_up")
    due_dates = [_parse_date(row.get("extended_due_date") or row.get("original_due_date")) for row in item.get("financings", []) if row.get("extended_due_date") or row.get("original_due_date")]
    due_dates = [value for value in due_dates if value]
    if due_dates and not item.get("documents") and min((value - date.today()).days for value in due_dates) <= 15:
        reasons.append("document_follow_up")
    follow_up = _parse_date(item.get("next_follow_up_date"))
    if follow_up and (follow_up - date.today()).days <= 10:
        reasons.append("manual_follow_up")
    if item.get("status") == "已回款" and item.get("outstanding_financing_amount", 0) > 0 and due_dates and min((value - date.today()).days for value in due_dates) <= 30:
        reasons.append("repayment_follow_up")
    return reasons


def calculate_business(record: dict[str, Any], manual_fcr: bool = False) -> tuple[str, str, list[dict[str, Any]]]:
    """Return status, current risk and separate data anomalies."""
    anomalies: list[dict[str, Any]] = []
    risk_reasons: list[str] = []
    required = {
        "business_no": "业务编号",
        "business_type": "业务类型",
        "trade_entity": "贸易主体",
        "supplier_steel_mill": "供应钢厂",
        "terminal_customer": "终端客户",
        "product_name": "货物品名",
        "contract_quantity_mt": "业务整体合同数量",
    }
    for field, label in required.items():
        if record.get(field) in (None, ""):
            anomalies.append({"key": f"missing:{field}", "type": "数据缺失", "description": f"缺少{label}"})
    if record.get("business_no") and record.get("business_no") == _normalize_text(record.get("source_record_key")):
        anomalies.append({"key": "invalid:business_no_source_key", "type": "数据异常", "description": "页面业务编号疑似使用来源记录键，需回到真实业务编号或真实采购合同号"})
    financings = [item for item in record.get("financings", []) if item.get("amount") is not None or item.get("financing_date") or item.get("bank")]
    repayments = [item for item in record.get("bank_repayments", []) if item.get("repayment_date") or item.get("completion_explicit")]
    receipts = [item for item in record.get("customer_receipts", []) if item.get("receipt_date")]
    fully_received = [item for item in receipts if item.get("fully_received")]
    documents = [item for item in record.get("documents", []) if item.get("document_date")]
    vessels = [item for item in record.get("vessels", []) if item.get("vessel_name")]
    if record.get("business_type") == "融资":
        for index, item in enumerate(financings):
            if not item.get("bank"):
                anomalies.append({"key": f"missing:finance_bank:{index}", "type": "数据缺失", "description": "融资明细缺少融资银行"})
            if item.get("amount") is None:
                anomalies.append({"key": f"missing:finance_amount:{index}", "type": "数据缺失", "description": "融资明细缺少融资金额"})
            if not item.get("financing_date"):
                anomalies.append({"key": f"missing:finance_date:{index}", "type": "数据缺失", "description": "融资明细缺少融资日期"})
            if "已还" in _normalize_text(item.get("repayment_status")) and not item.get("repayment_date"):
                anomalies.append({"key": f"missing:repayment_date:{index}", "type": "数据缺失", "description": "融资明细已明确还款但缺少银行实际还款日期"})
        if not financings:
            anomalies.append({"key": "missing:financing", "type": "数据缺失", "description": "融资业务暂无可确认的实际融资明细"})
    for field, label in (("financing_date", "融资日期"), ("document_date", "交单日期"), ("receipt_date", "客户回款日期"), ("repayment_date", "银行还款日期")):
        values = []
        collections = record.get("financings", []) if field == "financing_date" else record.get("documents", []) if field == "document_date" else record.get("customer_receipts", []) if field == "receipt_date" else record.get("bank_repayments", []) + record.get("financings", [])
        for collection in collections:
            value = collection.get(field)
            if value:
                values.append(value)
        for value in values:
            parsed = _parse_date(value)
            if parsed and parsed > date.today():
                anomalies.append({"key": f"future:{field}:{value}", "type": "日期异常", "description": f"{label}为未来日期：{value}"})
    has_document = bool(documents)
    has_vessel = bool(vessels)
    has_port = bool(record.get("_port_confirmed"))
    has_shipment = bool(record.get("_shipment_confirmed"))
    if receipts and not has_document:
        anomalies.append({"key": "sequence:receipt_without_document", "type": "节点矛盾", "description": "已有客户回款事实但交单事实缺失"})
    if has_document and not has_vessel:
        anomalies.append({"key": "sequence:document_without_vessel", "type": "节点矛盾", "description": "已有交单事实但装船/船舶事实缺失"})
    if record.get("business_type") == "融资":
        repayment_amounts_by_key: dict[str, float] = defaultdict(float)
        repayment_amounts_by_id: dict[Any, float] = defaultdict(float)
        repaid_finance_keys: set[str] = set()
        repaid_finance_ids: set[Any] = set()
        for repayment in repayments:
            repayment_key = _normalize_text(repayment.get("financing_source_key"))
            repayment_id = repayment.get("financing_id")
            amount = float(repayment.get("amount") or 0)
            if repayment_key:
                repayment_amounts_by_key[repayment_key] += amount
            if repayment_id is not None:
                repayment_amounts_by_id[repayment_id] += amount
            if repayment.get("completion_explicit"):
                if repayment_key:
                    repaid_finance_keys.add(repayment_key)
                if repayment_id is not None:
                    repaid_finance_ids.add(repayment_id)
        for financing in financings:
            financing_key = _normalize_text(financing.get("source_key"))
            financing_id = financing.get("id")
            financing_amount = float(financing.get("amount") or 0)
            if financing_amount > 0 and (
                (financing_key and repayment_amounts_by_key[financing_key] >= financing_amount)
                or (financing_id is not None and repayment_amounts_by_id[financing_id] >= financing_amount)
            ):
                if financing_key:
                    repaid_finance_keys.add(financing_key)
                if financing_id is not None:
                    repaid_finance_ids.add(financing_id)
        active_financings = [
            item for item in financings
            if not (
                item.get("repayment_date")
                or "已还" in _normalize_text(item.get("repayment_status"))
                or _normalize_text(item.get("source_key")) in repaid_finance_keys
                or item.get("id") in repaid_finance_ids
            )
        ]
        all_repaid = bool(financings) and all(
            item.get("repayment_date")
            or "已还" in _normalize_text(item.get("repayment_status"))
            or _normalize_text(item.get("source_key")) in repaid_finance_keys
            or item.get("id") in repaid_finance_ids
            for item in financings
        )
        if all_repaid:
            status = "已完结"
        elif fully_received and len(fully_received) >= len(receipts):
            status = "已回款"
        elif documents and not receipts:
            status = "已交单"
        elif receipts and not fully_received:
            status = "待收汇"
        elif documents:
            status = "待收汇"
        elif has_shipment:
            status = "已装船"
        elif has_port:
            status = "已集港"
        elif financings:
            status = "已放款"
        else:
            status = "待确认"
        active_due_dates = [_parse_date(item.get("extended_due_date") or item.get("original_due_date")) for item in active_financings]
        active_due_dates = [item for item in active_due_dates if item]
        if status == "已完结":
            risk = "低风险"
        elif not active_due_dates:
            risk = "中风险"
            risk_reasons.append("融资到期日缺失或无法计算")
        else:
            days = min((item - date.today()).days for item in active_due_dates)
            risk = "高风险" if days <= 7 else "中风险" if days <= 30 else "低风险"
            if days <= 7:
                risk_reasons.append(f"最近融资到期日距今天 {days} 天")
            elif days <= 30:
                risk_reasons.append(f"最近融资到期日距今天 {days} 天")
        if status != "已完结":
            shipment_dates = [_parse_date(item.get("latest_shipment_date")) for item in vessels if item.get("latest_shipment_date")]
            shipment_dates = [item for item in shipment_dates if item]
            if shipment_dates and not has_shipment:
                shipment_days = min((item - date.today()).days for item in shipment_dates)
                if shipment_days < 0:
                    risk = "高风险"
                    risk_reasons.append("最迟装船日已逾期且尚未确认装船")
                elif has_port and shipment_days <= 2:
                    risk = "高风险" if not (record.get("fcr") and has_port) else "中风险"
                    risk_reasons.append("已集港但临近最迟装船日，尚未确认装船")
                elif shipment_days <= 10 and risk == "低风险":
                    risk = "中风险"
                    risk_reasons.append("距最迟装船日不超过 10 天，尚未确认装船")
            if not shipment_dates and (has_port or status in {"已放款", "已集港"}):
                anomalies.append({"key": "missing:latest_shipment_date", "type": "数据缺失", "description": "当前执行节点缺少最迟装船日，无法判断装船风险"})
    else:
        settlement_status = _normalize_text(record.get("settlement_status"))
        all_received = bool(receipts) and len(fully_received) == len(receipts)
        if settlement_status == "已结算":
            status = "已结算"
            if not _normalize_date(record.get("settlement_date")):
                anomalies.append({"key": "missing:settlement_date", "type": "数据缺失", "description": "已标记结算但缺少结算日期"})
        elif record.get("guo_danlei_special") and all_received:
            status = "已完结"
        elif all_received:
            status = "已回款"
        else:
            status = "订单执行中"
        risk = "中风险" if (has_document or has_port) and status not in {"已完结", "已结算"} else "低风险"
        if risk == "中风险":
            risk_reasons.append("过单业务存在待确认执行节点")
    record["_risk_reasons"] = risk_reasons
    return status, risk, anomalies


def _upsert_child(cur, table: str, columns: list[str], values: list[Any], conflict_columns: list[str]) -> int:
    placeholders = ", ".join("?" for _ in values)
    fields = ", ".join(columns)
    if db._is_pg():
        updates = ", ".join(f"{column}=EXCLUDED.{column}" for column in columns if column not in conflict_columns)
        sql = f"INSERT INTO {table} ({fields}) VALUES ({placeholders}) ON CONFLICT ({', '.join(conflict_columns)}) DO UPDATE SET {updates} RETURNING id"
        cur.execute(sql.replace("?", "%s"), values)
        row = cur.fetchone()
        return int(row["id"])
    cur.execute(
        f"INSERT INTO {table} ({fields}) VALUES ({placeholders}) ON CONFLICT ({', '.join(conflict_columns)}) DO UPDATE SET "
        + ", ".join(f"{column}=excluded.{column}" for column in columns if column not in conflict_columns),
        values,
    )
    row = cur.execute(f"SELECT id FROM {table} WHERE " + " AND ".join(f"{column} = ?" for column in conflict_columns), [values[columns.index(column)] for column in conflict_columns]).fetchone()
    return int(row["id"])


def _upsert_anomalies(cur, business_id: int, anomalies: list[dict[str, Any]]) -> None:
    active_keys = {item["key"] for item in anomalies}
    current = db._exec(cur, "SELECT anomaly_key FROM order_lifecycle_data_anomalies WHERE business_id = ? AND status = 'open'", (business_id,)).fetchall()
    for row in current:
        if row["anomaly_key"] not in active_keys:
            db._exec(cur, "UPDATE order_lifecycle_data_anomalies SET status = 'resolved', resolved_at = CURRENT_TIMESTAMP WHERE business_id = ? AND anomaly_key = ?", (business_id, row["anomaly_key"]))
    for item in anomalies:
        db._exec(
            cur,
            """
            INSERT INTO order_lifecycle_data_anomalies
                (business_id, anomaly_key, anomaly_type, description, details_json, status)
            VALUES (?, ?, ?, ?, ?, 'open')
            ON CONFLICT (business_id, anomaly_key) DO UPDATE SET
                anomaly_type = excluded.anomaly_type,
                description = excluded.description,
                details_json = excluded.details_json,
                status = 'open',
                last_seen_at = CURRENT_TIMESTAMP,
                resolved_at = NULL,
                resolved_by = NULL
            """,
            (business_id, item["key"], item["type"], item["description"], _json(item)),
        )


def _audit_change(cur, business_id: int, operation: str, path: str, old_value: Any, new_value: Any, operator: str, note: str = "") -> None:
    db._exec(
        cur,
        "INSERT INTO order_lifecycle_audit (business_id, operation, path, old_value_json, new_value_json, operator, changed_at, note) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)",
        (business_id, operation, path, _json(old_value), _json(new_value), operator, note or None),
    )


def _replace_children(cur, business_id: int, record: dict[str, Any]) -> None:
    for table in ("order_lifecycle_contracts", "order_lifecycle_financings", "order_lifecycle_vessels", "order_lifecycle_documents", "order_lifecycle_customer_receipts", "order_lifecycle_bank_repayments"):
        db._exec(cur, f"DELETE FROM {table} WHERE business_id = ?", (business_id,))
    finance_ids: dict[str, int] = {}
    for item in record.get("contracts", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("contract", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        _upsert_child(cur, "order_lifecycle_contracts", ["business_id", "contract_no", "purchase_contract_no", "system_contract_no", "buyer", "seller", "quantity_mt", "source_key"], [business_id, item.get("contract_no"), item.get("purchase_contract_no"), item.get("system_contract_no"), item.get("buyer"), item.get("seller"), item.get("quantity_mt"), source_key], ["business_id", "source_key"])
    for item in record.get("financings", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("financing", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        finance_id = _upsert_child(cur, "order_lifecycle_financings", ["business_id", "bank", "amount", "currency", "financing_date", "original_due_date", "extended_due_date", "repayment_date", "repayment_status", "source_key"], [business_id, item.get("bank"), item.get("amount"), item.get("currency"), item.get("financing_date"), item.get("original_due_date"), item.get("extended_due_date"), item.get("repayment_date"), item.get("repayment_status"), source_key], ["business_id", "source_key"])
        finance_ids[_normalize_text(item.get("source_key"))] = finance_id
        finance_ids[source_key] = finance_id
    for item in record.get("vessels", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("vessel", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        _upsert_child(cur, "order_lifecycle_vessels", ["business_id", "vessel_name", "imo", "loading_port", "discharge_port", "eta", "etb", "estimated_discharge_date", "latest_shipment_date", "source_key", "source"], [business_id, item.get("vessel_name"), item.get("imo"), item.get("loading_port"), item.get("discharge_port"), item.get("eta"), item.get("etb"), item.get("estimated_discharge_date"), item.get("latest_shipment_date"), source_key, item.get("source")], ["business_id", "source_key"])
    for item in record.get("documents", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("document", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        _upsert_child(cur, "order_lifecycle_documents", ["business_id", "document_type", "document_date", "source_key"], [business_id, item.get("document_type", "交单"), item.get("document_date"), source_key], ["business_id", "source_key"])
    for item in record.get("customer_receipts", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("receipt", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        _upsert_child(cur, "order_lifecycle_customer_receipts", ["business_id", "receipt_date", "amount", "currency", "fully_received", "applicable_scope", "source_key"], [business_id, item.get("receipt_date"), item.get("amount"), item.get("currency"), 1 if item.get("fully_received") else 0, item.get("applicable_scope"), source_key], ["business_id", "source_key"])
    for item in record.get("bank_repayments", []):
        if item.get("manual_record"):
            continue
        source_key = _source_child_key("bank_repayment", item) or _normalize_text(item.get("source_key")) or uuid4().hex
        financing_source_key = _normalize_text(item.get("financing_source_key"))
        _upsert_child(cur, "order_lifecycle_bank_repayments", ["business_id", "financing_id", "repayment_date", "amount", "currency", "completion_explicit", "source_key"], [business_id, finance_ids.get(financing_source_key), item.get("repayment_date"), item.get("amount"), item.get("currency"), 1 if item.get("completion_explicit") else 0, source_key], ["business_id", "source_key"])


def _record_source_business_key(record: dict[str, Any]) -> str:
    return _normalize_text(
        record.get("source_business_key")
        or record.get("business_key")
        or record.get("source_record_key")
    )


def _record_match_tokens(record: dict[str, Any]) -> set[str]:
    tokens: set[str] = set()
    source_type = _normalize_text(record.get("source_type")).lower()
    if source_type == "wps" or (record.get("business_type") == "融资" and record.get("business_no")):
        business_no = _normalize_business_no(record.get("business_no"))
        if business_no:
            tokens.add(f"business:{_compact(business_no)}")
    for contract in record.get("contracts", []):
        values = [_compact(contract.get(field)) for field in ("purchase_contract_no", "system_contract_no", "contract_no")]
        values = [value for value in values if value]
        tokens.update(f"contract:{value}" for value in values)
        if len(values) >= 2:
            tokens.add(f"chain:{'|'.join(values)}")
    return tokens


def _parent_match_tokens(row: dict[str, Any], children: dict[str, list[dict[str, Any]]]) -> set[str]:
    tokens: set[str] = set()
    if row.get("source_type") == "wps" or row.get("business_type") == "融资":
        business_no = _normalize_business_no(row.get("business_no"))
        if business_no:
            tokens.add(f"business:{_compact(business_no)}")
    for contract in children.get("contracts", []):
        values = [_compact(contract.get(field)) for field in ("purchase_contract_no", "system_contract_no", "contract_no")]
        values = [value for value in values if value]
        tokens.update(f"contract:{value}" for value in values)
        if len(values) >= 2:
            tokens.add(f"chain:{'|'.join(values)}")
    return tokens


def _load_source_child_keys(cur, parent_key: str, source_type: str, source_business_key: str = "") -> dict[str, set[str]]:
    keys = {collection: set() for collection in CHILD_TABLES}
    rows = db._exec(
        cur,
        "SELECT source_key, normalized_json FROM order_lifecycle_source_records WHERE source_type = ? AND business_key = ? ORDER BY id DESC",
        (source_type, parent_key),
    ).fetchall()
    matching_rows = []
    for row in rows:
        record = _safe_json(row["normalized_json"], {})
        if not isinstance(record, dict):
            continue
        if source_business_key and source_business_key not in {_record_source_business_key(record), _normalize_text(row["source_key"])}:
            continue
        matching_rows.append((row, record))
    if source_business_key and not matching_rows and len(rows) == 1:
        record = _safe_json(rows[0]["normalized_json"], {})
        if isinstance(record, dict):
            matching_rows.append((rows[0], record))
    for row, record in matching_rows:
        for collection, kind in (("contracts", "contract"), ("financings", "financing"), ("vessels", "vessel"), ("documents", "document"), ("customer_receipts", "receipt"), ("bank_repayments", "bank_repayment")):
            for item in record.get(collection, []):
                child_key = _source_child_key(kind, item)
                if child_key:
                    keys[collection].add(child_key)
    return keys


def _remove_source_children(record: dict[str, Any], source_keys: dict[str, set[str]]) -> None:
    for collection, kind in (("contracts", "contract"), ("financings", "financing"), ("vessels", "vessel"), ("documents", "document"), ("customer_receipts", "receipt"), ("bank_repayments", "bank_repayment")):
        record[collection] = [
            item for item in record.get(collection, [])
            if item.get("manual_record") or _source_child_key(kind, item) not in source_keys.get(collection, set())
        ]


def _other_active_source_child_keys(cur, business_id: int, source_type: str, source_business_key: str) -> dict[str, set[str]]:
    other_keys = {collection: set() for collection in CHILD_TABLES}
    active_rows = db._exec(
        cur,
        "SELECT source_type, source_business_key FROM order_lifecycle_business_sources WHERE business_id = ? AND source_active = 1 AND NOT (source_type = ? AND source_business_key = ?)",
        (business_id, source_type, source_business_key),
    ).fetchall()
    parent = db._exec(cur, "SELECT business_key FROM order_lifecycle_businesses WHERE id = ?", (business_id,)).fetchone()
    if not parent:
        return other_keys
    for row in active_rows:
        contribution = _load_source_child_keys(cur, parent["business_key"], row["source_type"], row["source_business_key"])
        for collection in CHILD_TABLES:
            other_keys[collection].update(contribution[collection])
    return other_keys


def _remove_missing_source_children(cur, business_id: int, source_type: str, source_business_key: str) -> None:
    parent = db._exec(cur, "SELECT business_key FROM order_lifecycle_businesses WHERE id = ?", (business_id,)).fetchone()
    if not parent:
        return
    target_keys = _load_source_child_keys(cur, parent["business_key"], source_type, source_business_key)
    if not any(target_keys.values()):
        return
    active_source = db._exec(
        cur,
        "SELECT 1 FROM order_lifecycle_business_sources WHERE business_id = ? AND source_active = 1 AND NOT (source_type = ? AND source_business_key = ?) LIMIT 1",
        (business_id, source_type, source_business_key),
    ).fetchone()
    if not active_source:
        return
    other_keys = _other_active_source_child_keys(cur, business_id, source_type, source_business_key)
    for collection, table in CHILD_TABLES.items():
        removable = sorted(target_keys[collection] - other_keys[collection])
        if removable:
            placeholders = ", ".join("?" for _ in removable)
            db._exec(cur, f"DELETE FROM {table} WHERE business_id = ? AND source_key IN ({placeholders})", [business_id, *removable])
    _recalculate_business(cur, business_id)


def _load_parent_match_cache(cur) -> tuple[dict[int, dict[str, Any]], dict[str, set[int]], dict[tuple[str, str], int]]:
    parents: dict[int, dict[str, Any]] = {}
    token_index: dict[str, set[int]] = defaultdict(set)
    rows = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE is_cancelled = 0").fetchall()
    for row in rows:
        row_dict = _row_json(row)
        children = _load_business_children(cur, row_dict["id"])
        parents[row_dict["id"]] = {"row": row_dict, "record": {**row_dict, **children}, "tokens": _parent_match_tokens(row_dict, children)}
        for token in parents[row_dict["id"]]["tokens"]:
            token_index[token].add(row_dict["id"])
    membership_rows = db._exec(
        cur,
        "SELECT source_type, source_business_key, business_id FROM order_lifecycle_business_sources",
    ).fetchall()
    memberships = {(row["source_type"], row["source_business_key"]): row["business_id"] for row in membership_rows}
    return parents, token_index, memberships


def _match_source_record(
    record: dict[str, Any],
    parents: dict[int, dict[str, Any]],
    token_index: dict[str, set[int]],
    memberships: dict[tuple[str, str], int],
) -> dict[str, Any]:
    source_type = _normalize_text(record.get("source_type")).lower()
    source_key = _record_source_business_key(record)
    direct_id = memberships.get((source_type, source_key)) if source_key else None
    token_candidates: set[int] = set()
    ambiguous_tokens: list[str] = []
    for token in _record_match_tokens(record):
        candidates = token_index.get(token, set())
        if len(candidates) > 1:
            ambiguous_tokens.append(token)
        token_candidates.update(candidates)
    if direct_id:
        if token_candidates and (token_candidates != {direct_id} or ambiguous_tokens):
            token_candidates.add(direct_id)
            return {"status": "pending", "reason": "来源标识指向不同主卡，禁止自动合并", "candidate_ids": sorted(token_candidates)}
        return {"status": "matched", "business_id": direct_id, "candidate_ids": [direct_id], "reason": "同来源稳定标识"}
    if ambiguous_tokens or len(token_candidates) > 1:
        return {"status": "pending", "reason": "精确合同标识指向多个主卡，禁止自动合并", "candidate_ids": sorted(token_candidates)}
    if len(token_candidates) == 1:
        business_id = next(iter(token_candidates))
        return {"status": "matched", "business_id": business_id, "candidate_ids": [business_id], "reason": "精确业务编号或合同标识"}
    return {"status": "unmatched", "candidate_ids": [], "reason": "未找到精确对应主卡"}


def _canonical_business_key(record: dict[str, Any], parent: Optional[dict[str, Any]] = None) -> str:
    if parent:
        source_type = _normalize_text(record.get("source_type")).lower()
        business_no = _normalize_business_no(record.get("business_no"))
        if source_type == "wps" and business_no:
            return f"business:wps:{_compact(business_no)}"
        return parent["row"].get("business_key") or parent["record"].get("business_key") or ""
    identity = _source_business_identity(record)
    if identity:
        return identity
    return _normalize_text(record.get("business_key"))


def _merge_parent_record(parent_record: dict[str, Any], incoming: dict[str, Any], source_type: str) -> dict[str, Any]:
    merged = deepcopy(parent_record)
    for field in ("business_no", "trade_entity", "supplier_steel_mill", "terminal_customer", "product_name", "contract_quantity_mt"):
        incoming_value = incoming.get(field)
        if incoming_value in (None, ""):
            continue
        if merged.get(field) in (None, "") or (source_type == "wps" and field in {"business_no", "trade_entity", "supplier_steel_mill", "product_name", "contract_quantity_mt"}):
            merged[field] = incoming_value
    if incoming.get("business_type") == "融资" or merged.get("business_type") not in {"融资", "过单"}:
        merged["business_type"] = incoming.get("business_type") or merged.get("business_type")
    for field in ("settlement_status", "settlement_date", "guo_danlei_special", "completed_date"):
        if incoming.get(field) not in (None, "", False) and (field != "settlement_status" or incoming.get(field) == "已结算"):
            merged[field] = incoming[field]
    _merge_unique_children(merged, incoming)
    return merged


def _upsert_business_source(cur, business_id: int, record: dict[str, Any], batch_id: int, presence_hash: str) -> None:
    source_type = _normalize_text(record.get("source_type")).lower()
    source_key = _record_source_business_key(record)
    if not source_type or not source_key:
        return
    db._exec(
        cur,
        """
        INSERT INTO order_lifecycle_business_sources
            (business_id, source_type, source_business_key, source_record_key, source_version, source_presence_hash, source_active, missing_observation_hash, missing_observation_count, last_seen_batch_id, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?, 1, NULL, 0, ?, CURRENT_TIMESTAMP)
        ON CONFLICT (source_type, source_business_key) DO UPDATE SET
            business_id = excluded.business_id,
            source_record_key = excluded.source_record_key,
            source_version = excluded.source_version,
            source_presence_hash = excluded.source_presence_hash,
            source_active = 1,
            missing_observation_hash = NULL,
            missing_observation_count = 0,
            last_seen_batch_id = excluded.last_seen_batch_id,
            last_seen_at = CURRENT_TIMESTAMP
        """,
        (business_id, source_type, source_key, record.get("source_record_key"), record.get("source_version"), presence_hash, batch_id),
    )


def _save_match_candidate(cur, record: dict[str, Any], reason: str, candidate_ids: list[int], parents: dict[int, dict[str, Any]]) -> None:
    source_type = _normalize_text(record.get("source_type")).lower()
    source_key = _record_source_business_key(record)
    if not source_type or not source_key:
        return
    candidate_keys = [parents[item]["row"].get("business_key") for item in candidate_ids if item in parents]
    db._exec(
        cur,
        """
        INSERT INTO order_lifecycle_match_candidates
            (source_type, source_record_key, source_version, reason, candidate_keys_json, raw_json, status, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?, 'open', CURRENT_TIMESTAMP)
        ON CONFLICT (source_type, source_record_key) DO UPDATE SET
            source_version = excluded.source_version,
            reason = excluded.reason,
            candidate_keys_json = excluded.candidate_keys_json,
            raw_json = excluded.raw_json,
            status = 'open',
            last_seen_at = CURRENT_TIMESTAMP,
            resolved_at = NULL
        """,
        (source_type, source_key, record.get("source_version"), reason, _json(candidate_keys), _json(record)),
    )


def _resolve_match_candidate(cur, record: dict[str, Any]) -> None:
    source_type = _normalize_text(record.get("source_type")).lower()
    source_key = _record_source_business_key(record)
    if source_type and source_key:
        db._exec(
            cur,
            "UPDATE order_lifecycle_match_candidates SET status = 'resolved', resolved_at = CURRENT_TIMESTAMP, last_seen_at = CURRENT_TIMESTAMP WHERE source_type = ? AND source_record_key = ? AND status = 'open'",
            (source_type, source_key),
        )


def apply_source_batch(batch: dict[str, Any], imported_by: str = "system", complete_snapshot: bool = True) -> dict[str, Any]:
    """Apply an isolated source snapshot with exact cross-source reconciliation.

    A source key is only a membership key.  The parent card is resolved through
    exact business-number/contract identifiers; ambiguous or unmatched financing
    mail remains a candidate and never becomes a temporary business card.
    """
    with db.connect() as conn:
        initialize_schema(conn)
        cur = conn.cursor()
        records = batch.get("records", [])
        source_type = _normalize_text(batch.get("source_type")).lower()
        source_keys = sorted({_record_source_business_key(item) for item in records if _record_source_business_key(item)})
        key_hash = hashlib.sha256("|".join(source_keys).encode()).hexdigest()
        batch_id = db._last_insert_id(
            cur,
            "INSERT INTO order_lifecycle_source_batches (source_type, source_locator, source_version, snapshot_date, source_hash, source_key_set_hash, status, record_count, completed_at) VALUES (?, ?, ?, ?, ?, ?, 'success', ?, CURRENT_TIMESTAMP)",
            (source_type, batch.get("source_locator"), batch.get("source_version"), batch.get("snapshot_date"), batch.get("source_hash"), key_hash, len(records)),
        )
        membership_rows = db._exec(
            cur,
            "SELECT source_business_key, business_id FROM order_lifecycle_business_sources WHERE source_type = ? AND source_active = 1",
            (source_type,),
        ).fetchall()
        missing_keys = sorted({row["source_business_key"] for row in membership_rows} - set(source_keys)) if complete_snapshot else []
        deletion_hash = hashlib.sha256("|".join(missing_keys).encode()).hexdigest() if missing_keys else ""
        previous = db._exec(
            cur,
            "SELECT deletion_candidate_hash FROM order_lifecycle_source_batches WHERE source_type = ? AND id <> ? ORDER BY id DESC LIMIT 1",
            (source_type, batch_id),
        ).fetchone()
        confirmed_missing = set(missing_keys) if missing_keys and previous and previous["deletion_candidate_hash"] == deletion_hash else set()
        if missing_keys:
            db._exec(
                cur,
                "UPDATE order_lifecycle_source_batches SET deletion_candidate_hash = ?, deletion_candidate_count = ? WHERE id = ?",
                (deletion_hash, len(missing_keys), batch_id),
            )

        pending_candidates = 0
        matched_records = 0
        created_businesses = 0
        deactivated_business_ids: set[int] = set()
        for record in records:
            parents, token_index, memberships = _load_parent_match_cache(cur)
            source_key = _record_source_business_key(record)
            resolution = _match_source_record(record, parents, token_index, memberships)
            source_kind = _normalize_text(record.get("source_type")).lower()
            if source_kind == "wps" and _is_legacy_mill_row_business_no(record.get("business_no")):
                _save_match_candidate(
                    cur,
                    record,
                    "WPS业务编号仅为钢厂+行号，禁止生成业务主卡；请回读 XYZ-年份-序号格式的真实编号",
                    [],
                    parents,
                )
                pending_candidates += 1
                continue
            business_id = resolution.get("business_id") if resolution.get("status") == "matched" else None
            if resolution["status"] == "pending" or (
                resolution["status"] == "unmatched"
                and source_kind == "email"
                and record.get("business_type") == "融资"
                and not _normalize_business_no(record.get("business_no"))
            ):
                reason = resolution["reason"] if resolution["status"] == "pending" else "融资邮件未找到可精确匹配的WPS主卡"
                _save_match_candidate(cur, record, reason, resolution.get("candidate_ids", []), parents)
                pending_candidates += 1
                pending_key = f"pending:{source_kind}:{source_key or record.get('source_record_key') or uuid4().hex}"
                db._exec(
                    cur,
                    "INSERT OR IGNORE INTO order_lifecycle_source_records (batch_id, source_type, source_key, business_key, source_file, source_sheet, source_row, raw_json, normalized_json, raw_hash) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (batch_id, source_kind, record.get("source_record_key") or source_key, pending_key, record.get("raw", {}).get("file"), record.get("raw", {}).get("sheet"), record.get("raw", {}).get("row_no"), _json(record.get("raw", {})), _json(record), hashlib.sha256(_json(record.get("raw", {})).encode()).hexdigest()),
                )
                continue
            if business_id is None:
                canonical_key = _canonical_business_key(record)
                if source_kind == "wps" and not _normalize_business_no(record.get("business_no")):
                    _save_match_candidate(cur, record, "WPS记录缺少真实业务编号，禁止建临时主卡", [], parents)
                    pending_candidates += 1
                    continue
                if not canonical_key:
                    _save_match_candidate(cur, record, "缺少可用于归并的真实业务编号或合同标识", [], parents)
                    pending_candidates += 1
                    continue
                business_id = None
                existing_key_owner = next((item_id for item_id, item in parents.items() if item["row"].get("business_key") == canonical_key), None)
                if existing_key_owner is not None:
                    _save_match_candidate(cur, record, "规范化主卡编号已被其他主卡占用，禁止覆盖", [existing_key_owner], parents)
                    pending_candidates += 1
                    continue
            existing_parent = parents.get(business_id) if business_id else None
            if existing_parent:
                base_record = deepcopy(existing_parent["record"])
                previous_source_keys = _load_source_child_keys(cur, existing_parent["row"].get("business_key"), source_kind, source_key)
                other_source_keys = _other_active_source_child_keys(cur, business_id, source_kind, source_key)
                removable_source_keys = {
                    collection: previous_source_keys[collection] - other_source_keys[collection]
                    for collection in CHILD_TABLES
                }
                _remove_source_children(base_record, removable_source_keys)
                effective_record = _merge_parent_record(base_record, record, source_kind)
                effective_record["business_key"] = _canonical_business_key(record, existing_parent)
                effective_record["source_type"] = "mixed" if existing_parent["row"].get("source_type") not in {source_kind, "mixed"} else existing_parent["row"].get("source_type")
                effective_record["_port_confirmed"] = existing_parent["row"].get("port_status") == "已集港"
                effective_record["_shipment_confirmed"] = existing_parent["row"].get("shipment_status") == "已装船"
            else:
                effective_record = deepcopy(record)
                effective_record["business_key"] = _canonical_business_key(record)
                effective_record["_port_confirmed"] = False
                effective_record["_shipment_confirmed"] = False
            anomalies: list[dict[str, Any]] = []
            if business_id:
                override_rows = db._exec(cur, "SELECT field_name, value_json FROM order_lifecycle_manual_overrides WHERE business_id = ? AND is_active = 1", (business_id,)).fetchall()
                for override in override_rows:
                    field_name = override["field_name"]
                    if field_name not in MANUAL_OVERRIDE_FIELDS and field_name != "fcr":
                        continue
                    manual_value = _safe_json(override["value_json"], None)
                    source_value = record.get(field_name)
                    if field_name != "fcr" and source_value not in (None, "") and source_value != manual_value:
                        source_hash = hashlib.sha256(_json(source_value).encode()).hexdigest()[:16]
                        anomalies.append({"key": f"manual_conflict:{field_name}:{source_hash}", "type": "来源冲突", "description": f"人工值与来源值不同：{field_name}", "details": {"field": field_name, "manual_value": manual_value, "source_value": source_value}})
                    if field_name != "fcr":
                        effective_record[field_name] = manual_value
                        if field_name == "port_status":
                            effective_record["_port_confirmed"] = manual_value == "已集港"
                        elif field_name == "shipment_status":
                            effective_record["_shipment_confirmed"] = manual_value == "已装船"
            effective_record["_port_status"] = "已集港" if effective_record.get("_port_confirmed") else "待确认"
            effective_record["_shipment_status"] = "已装船" if effective_record.get("_shipment_confirmed") else "待确认"
            status, risk, calculated_anomalies = calculate_business(effective_record)
            completion_date = _completion_date(effective_record, status)
            anomalies.extend(calculated_anomalies)
            if business_id:
                owner = db._exec(cur, "SELECT id FROM order_lifecycle_businesses WHERE business_key = ? AND id <> ?", (effective_record["business_key"], business_id)).fetchone()
                if owner:
                    _save_match_candidate(cur, record, "规范化主卡编号与现有主卡冲突，禁止自动合并", [business_id, owner["id"]], parents)
                    pending_candidates += 1
                    continue
                db._exec(
                    cur,
                    """UPDATE order_lifecycle_businesses SET business_key = ?, business_no = ?, business_type = ?, trade_entity = ?, supplier_steel_mill = ?, terminal_customer = ?, product_name = ?, contract_quantity_mt = ?, status = ?, port_status = ?, shipment_status = ?, risk_level = ?, risk_reasons_json = ?, anomaly_count = ?, settlement_status = ?, settlement_date = ?, guo_danlei_special = ?, completed_date = ?, next_follow_up_date = ?, source_type = ?, source_snapshot_date = ?, source_version = ?, source_record_key = ?, source_presence_hash = ?, source_active = 1, updated_at = CURRENT_TIMESTAMP WHERE id = ?""",
                    (effective_record["business_key"], effective_record.get("business_no"), effective_record.get("business_type"), effective_record.get("trade_entity"), effective_record.get("supplier_steel_mill"), effective_record.get("terminal_customer"), effective_record.get("product_name"), effective_record.get("contract_quantity_mt"), status, effective_record["_port_status"], effective_record["_shipment_status"], risk, _json(effective_record.get("_risk_reasons", [])), len(anomalies), effective_record.get("settlement_status") or "待结算", effective_record.get("settlement_date") or None, 1 if effective_record.get("guo_danlei_special") else 0, completion_date, effective_record.get("next_follow_up_date") or None, effective_record.get("source_type") or source_kind, record.get("source_snapshot_date"), record.get("source_version"), record.get("source_record_key"), key_hash, business_id),
                )
                matched_records += 1
            else:
                business_uid = uuid4().hex
                business_id = db._last_insert_id(
                    cur,
                    """INSERT INTO order_lifecycle_businesses (business_uid, business_key, business_no, business_type, trade_entity, supplier_steel_mill, terminal_customer, product_name, contract_quantity_mt, status, port_status, shipment_status, risk_level, risk_reasons_json, anomaly_count, settlement_status, settlement_date, guo_danlei_special, completed_date, next_follow_up_date, source_type, source_snapshot_date, source_version, source_record_key, source_presence_hash, source_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (business_uid, effective_record["business_key"], effective_record.get("business_no"), effective_record.get("business_type"), effective_record.get("trade_entity"), effective_record.get("supplier_steel_mill"), effective_record.get("terminal_customer"), effective_record.get("product_name"), effective_record.get("contract_quantity_mt"), status, effective_record["_port_status"], effective_record["_shipment_status"], risk, _json(effective_record.get("_risk_reasons", [])), len(anomalies), effective_record.get("settlement_status") or "待结算", effective_record.get("settlement_date") or None, 1 if effective_record.get("guo_danlei_special") else 0, completion_date, effective_record.get("next_follow_up_date") or None, source_kind, record.get("source_snapshot_date"), record.get("source_version"), record.get("source_record_key"), key_hash, 1),
                )
                created_businesses += 1
            _replace_children(cur, business_id, effective_record)
            _upsert_anomalies(cur, business_id, anomalies)
            _upsert_business_source(cur, business_id, record, batch_id, hashlib.sha256(_json(record).encode()).hexdigest())
            _resolve_match_candidate(cur, record)
            db._exec(
                cur,
                "INSERT OR IGNORE INTO order_lifecycle_source_records (batch_id, source_type, source_key, business_key, source_file, source_sheet, source_row, raw_json, normalized_json, raw_hash) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (batch_id, source_kind, record.get("source_record_key") or source_key, effective_record["business_key"], record.get("raw", {}).get("file"), record.get("raw", {}).get("sheet"), record.get("raw", {}).get("row_no"), _json(record.get("raw", {})), _json(record), hashlib.sha256(_json(record.get("raw", {})).encode()).hexdigest()),
            )

        for source_key in confirmed_missing:
            source_row = db._exec(
                cur,
                "SELECT business_id FROM order_lifecycle_business_sources WHERE source_type = ? AND source_business_key = ?",
                (source_type, source_key),
            ).fetchone()
            db._exec(
                cur,
                "UPDATE order_lifecycle_business_sources SET source_active = 0, missing_observation_hash = ?, missing_observation_count = missing_observation_count + 1 WHERE source_type = ? AND source_business_key = ?",
                (deletion_hash, source_type, source_key),
            )
            if source_row:
                _remove_missing_source_children(cur, source_row["business_id"], source_type, source_key)
        if missing_keys:
            for source_key in missing_keys:
                if source_key not in confirmed_missing:
                    db._exec(
                        cur,
                        "UPDATE order_lifecycle_business_sources SET missing_observation_hash = ?, missing_observation_count = CASE WHEN missing_observation_hash = ? THEN missing_observation_count + 1 ELSE 1 END WHERE source_type = ? AND source_business_key = ?",
                        (deletion_hash, deletion_hash, source_type, source_key),
                    )
        db._exec(
            cur,
            "UPDATE order_lifecycle_businesses SET source_active = CASE WHEN EXISTS (SELECT 1 FROM order_lifecycle_business_sources s WHERE s.business_id = order_lifecycle_businesses.id AND s.source_active = 1) THEN 1 ELSE 0 END WHERE is_cancelled = 0",
        )
        for row in db._exec(cur, "SELECT id FROM order_lifecycle_businesses WHERE is_cancelled = 0 AND source_active = 0").fetchall():
            deactivated_business_ids.add(row["id"])
        now_field = "wps_last_success_at" if source_type == "wps" else "email_last_success_at"
        db._exec(cur, f"UPDATE order_lifecycle_sync_state SET {now_field} = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = 1")
        conn.commit()
    return {"batch_id": batch_id, "record_count": len(records), "matched_records": matched_records, "created_businesses": created_businesses, "pending_match_candidates": pending_candidates, "deletion_candidates": len(missing_keys), "deleted_businesses": len(deactivated_business_ids), "source_type": source_type}


def _row_json(row: Any) -> dict[str, Any]:
    if not row:
        return {}
    item = dict(row)
    for key, value in item.items():
        if key.endswith("_at"):
            item[key] = _seconds_text(value)
    return item


def _seconds_text(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).replace("T", " ")
    match = re.match(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", text)
    return match.group(1) if match else value


def _load_business_children(cur, business_id: int) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    parent = db._exec(cur, "SELECT business_key FROM order_lifecycle_businesses WHERE id = ?", (business_id,)).fetchone()
    source_meta: dict[tuple[str, str], dict[str, Any]] = {}
    if parent:
        source_rows = db._exec(
            cur,
            "SELECT r.source_type, r.source_key, r.normalized_json, b.source_version AS batch_source_version, b.snapshot_date AS batch_snapshot_date, b.completed_at, r.created_at FROM order_lifecycle_source_records r LEFT JOIN order_lifecycle_source_batches b ON b.id = r.batch_id WHERE r.business_key = ? ORDER BY r.id DESC",
            (parent["business_key"],),
        ).fetchall()
        for source_row in source_rows:
            normalized = _safe_json(source_row["normalized_json"], {})
            if not isinstance(normalized, dict):
                continue
            metadata = {
                "source": normalized.get("source_type") or source_row["source_type"],
                "source_type": normalized.get("source_type") or source_row["source_type"],
                "source_record_key": source_row["source_key"],
                "source_version": normalized.get("source_version") or source_row["batch_source_version"],
                "source_snapshot_date": normalized.get("source_snapshot_date") or source_row["batch_snapshot_date"],
                "source_updated_at": _seconds_text(source_row["completed_at"] or source_row["created_at"]),
            }
            for collection, kind in (("contracts", "contract"), ("financings", "financing"), ("vessels", "vessel"), ("documents", "document"), ("customer_receipts", "receipt"), ("bank_repayments", "bank_repayment")):
                for child in normalized.get(collection, []):
                    child_key = _source_child_key(kind, child)
                    if child_key:
                        source_meta.setdefault((collection, child_key), metadata)
    override_rows = db._exec(
        cur,
        "SELECT collection, source_key, field_name, value_json FROM order_lifecycle_child_overrides WHERE business_id = ? AND is_active = 1",
        (business_id,),
    ).fetchall()
    overrides = {(row["collection"], row["source_key"], row["field_name"]): _safe_json(row["value_json"], None) for row in override_rows}
    override_keys = {(row["collection"], row["source_key"], row["field_name"]) for row in override_rows}
    for key, table in CHILD_TABLES.items():
        rows = [_row_json(item) for item in db._exec(cur, f"SELECT * FROM {table} WHERE business_id = ? ORDER BY id", (business_id,)).fetchall()]
        for item in rows:
            for field in CHILD_FIELDS[key]:
                override = overrides.get((key, item.get("source_key"), field))
                if (key, item.get("source_key"), field) in override_keys:
                    item[field] = override
            if key == "financings":
                original_due = _parse_date(item.get("original_due_date"))
                extended_due = _parse_date(item.get("extended_due_date"))
                item["extension_days"] = max((extended_due - original_due).days, 0) if original_due and extended_due else 0
                item["current_due_date"] = _normalize_date(item.get("extended_due_date") or item.get("original_due_date")) or None
            metadata = source_meta.get((key, item.get("source_key")))
            if metadata:
                for field, value in metadata.items():
                    if value not in (None, ""):
                        item[field] = value
        manual_rows = db._exec(
            cur,
            "SELECT source_key, record_json, modified_by, modified_at, note FROM order_lifecycle_manual_child_records WHERE business_id = ? AND collection = ? AND is_active = 1 ORDER BY id",
            (business_id, key),
        ).fetchall()
        for manual in manual_rows:
            item = _safe_json(manual["record_json"], {})
            if not isinstance(item, dict):
                item = {}
            item["source_key"] = manual["source_key"]
            item["source"] = item.get("source") or "人工"
            item["manual_record"] = True
            item["modified_by"] = manual["modified_by"]
            item["modified_at"] = manual["modified_at"]
            item["note"] = manual["note"]
            rows.append(item)
        result[key] = rows
    return result


def _recalculate_business(cur, business_id: int) -> tuple[str, str, list[dict[str, Any]]]:
    row = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
    if not row:
        raise KeyError(business_id)
    record = {field: row[field] for field in ("business_no", "business_type", "trade_entity", "supplier_steel_mill", "terminal_customer", "product_name", "contract_quantity_mt", "settlement_status", "settlement_date", "guo_danlei_special", "next_follow_up_date", "fcr")}
    parent_overrides = db._exec(
        cur,
        "SELECT field_name, value_json FROM order_lifecycle_manual_overrides WHERE business_id = ? AND is_active = 1",
        (business_id,),
    ).fetchall()
    for override in parent_overrides:
        if override["field_name"] in MANUAL_OVERRIDE_FIELDS:
            record[override["field_name"]] = _safe_json(override["value_json"], None)
    record.update(_load_business_children(cur, business_id))
    record["_port_confirmed"] = row["port_status"] == "已集港"
    record["_shipment_confirmed"] = row["shipment_status"] == "已装船"
    status, risk, anomalies = calculate_business(record)
    completion_date = _completion_date(record, status)
    db._exec(
        cur,
        "UPDATE order_lifecycle_businesses SET status = ?, risk_level = ?, risk_reasons_json = ?, anomaly_count = ?, completed_date = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (status, risk, _json(record.get("_risk_reasons", [])), len(anomalies), completion_date, business_id),
    )
    _upsert_anomalies(cur, business_id, anomalies)
    return status, risk, anomalies


def _serialize_business(row: Any, children: dict[str, list[dict[str, Any]]], anomaly_rows: list[dict[str, Any]], can_sensitive: bool = False) -> dict[str, Any]:
    item = _row_json(row)
    item.update(children)
    item["risk_reasons"] = _safe_json(item.pop("risk_reasons_json", "[]"), [])
    item["anomalies"] = anomaly_rows
    item["anomaly_count"] = len(anomaly_rows)
    item["can_sensitive"] = can_sensitive
    item["fcr"] = bool(item.get("fcr"))
    item["financing_count"] = len(item.get("financings", []))
    item["bank_repayment_count"] = len(item.get("bank_repayments", []))
    item["customer_receipt_count"] = len(item.get("customer_receipts", []))
    item["financing_amount"] = sum(float(child.get("amount") or 0) for child in item.get("financings", []))
    repayment_amounts_by_id: dict[Any, float] = defaultdict(float)
    repaid_financing_ids = set()
    for repayment in item.get("bank_repayments", []):
        repayment_id = repayment.get("financing_id")
        if repayment_id is not None:
            repayment_amounts_by_id[repayment_id] += float(repayment.get("amount") or 0)
        if repayment.get("completion_explicit") and repayment_id is not None:
            repaid_financing_ids.add(repayment_id)
    for financing in item.get("financings", []):
        financing_id = financing.get("id")
        financing_amount = float(financing.get("amount") or 0)
        if financing_id is not None and financing_amount > 0 and repayment_amounts_by_id[financing_id] >= financing_amount:
            repaid_financing_ids.add(financing_id)
    repaid_financings = [
        child for child in item.get("financings", [])
        if child.get("repayment_date")
        or "已还" in _normalize_text(child.get("repayment_status"))
        or child.get("id") in repaid_financing_ids
    ]
    active_financings = [child for child in item.get("financings", []) if child not in repaid_financings]
    item["outstanding_financing_amount"] = sum(float(child.get("amount") or 0) for child in active_financings)
    due_rows = []
    extension_days = []
    for child in active_financings:
        current_due = _normalize_date(child.get("extended_due_date") or child.get("original_due_date"))
        parsed_current_due = _parse_date(current_due)
        if current_due and parsed_current_due:
            due_rows.append((parsed_current_due, current_due))
        original_due = _parse_date(child.get("original_due_date"))
        extended_due = _parse_date(child.get("extended_due_date"))
        if original_due and extended_due:
            extension_days.append(max((extended_due - original_due).days, 0))
    item["current_due_date"] = min(due_rows)[1] if due_rows else None
    item["due_days"] = (min(due_rows)[0] - date.today()).days if due_rows else None
    item["extension_days"] = max(extension_days, default=0)
    item["repayment_progress"] = {
        "repaid_financing_count": len(repaid_financings),
        "total_financing_count": len(item.get("financings", [])),
        "repaid_amount": sum(float(child.get("amount") or 0) for child in repaid_financings),
        "bank_repayment_count": len(item.get("bank_repayments", [])),
    }
    item["customer_receipt_progress"] = {
        "received_count": sum(1 for child in item.get("customer_receipts", []) if child.get("fully_received")),
        "total_count": len(item.get("customer_receipts", [])),
    }
    if anomaly_rows:
        item["next_action"] = "人工判断并修正数据异常"
    else:
        item["next_action"] = {
            "已放款": "确认集港进度",
            "已集港": "确认装船进度",
            "已装船": "确认交单事实",
            "待收汇": "跟进客户回款",
            "客户已回款": "确认银行还款",
            "已回款": "确认银行还款",
            "订单执行中": "跟进下一执行节点",
            "待确认": "补齐必要业务字段",
            "已完结": "无",
            "已结算": "无",
        }.get(item.get("status"), "人工确认下一步")
    return item


def _load_business_children_batch(cur, business_ids: list[int]) -> dict[int, dict[str, list[dict[str, Any]]]]:
    """Load card facts in one query per collection instead of one query per business."""
    result = {business_id: {collection: [] for collection in CHILD_TABLES} for business_id in business_ids}
    if not business_ids:
        return result
    placeholders = ", ".join("?" for _ in business_ids)
    for collection, table in CHILD_TABLES.items():
        rows = db._exec(
            cur,
            f"SELECT * FROM {table} WHERE business_id IN ({placeholders}) ORDER BY business_id, id",
            business_ids,
        ).fetchall()
        for row in rows:
            business_id = row["business_id"]
            if business_id in result:
                result[business_id][collection].append(_row_json(row))
    return result


def _load_open_anomalies_batch(cur, business_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    result = {business_id: [] for business_id in business_ids}
    if not business_ids:
        return result
    placeholders = ", ".join("?" for _ in business_ids)
    rows = db._exec(
        cur,
        f"SELECT * FROM order_lifecycle_data_anomalies WHERE status = 'open' AND business_id IN ({placeholders}) ORDER BY business_id, id",
        business_ids,
    ).fetchall()
    for row in rows:
        if row["business_id"] in result:
            result[row["business_id"]].append(_row_json(row))
    return result


def _serialize_business_card(row: Any, children: dict[str, list[dict[str, Any]]], anomaly_rows: list[dict[str, Any]], can_sensitive: bool = False) -> dict[str, Any]:
    """Serialize only the wide-card facts; full child/source/audit rows stay in detail."""
    item = _serialize_business(row, children, anomaly_rows, can_sensitive)
    item["contract_count"] = len(children.get("contracts", []))
    item["contract_numbers"] = [
        value
        for value in dict.fromkeys(
            _normalize_text(child.get("purchase_contract_no") or child.get("contract_no") or child.get("system_contract_no"))
            for child in children.get("contracts", [])
        )
        if value
    ]
    item["financing_banks"] = [
        value for value in dict.fromkeys(_normalize_text(child.get("bank")) for child in children.get("financings", [])) if value
    ]
    item["vessel_count"] = len(children.get("vessels", []))
    item["document_count"] = len(children.get("documents", []))
    shipment_dates = [
        _normalize_date(child.get("latest_shipment_date"))
        for child in children.get("vessels", [])
        if child.get("latest_shipment_date")
    ]
    item["latest_shipment_date"] = min((value for value in shipment_dates if value), default=None)
    for collection in CHILD_TABLES:
        item.pop(collection, None)
    item["children_loaded"] = False
    return item


def list_businesses(filters: dict[str, Any]) -> dict[str, Any]:
    clauses = ["b.is_cancelled = 0", "b.source_active = 1"]
    params: list[Any] = []
    def _values(name: str) -> list[str]:
        value = filters.get(name)
        if isinstance(value, (list, tuple, set)):
            return [_normalize_text(item) for item in value if _normalize_text(item)]
        text = _normalize_text(value)
        if text == "__none__":
            return ["__none__"]
        return [_normalize_text(item) for item in text.split(",") if _normalize_text(item)]

    def _append_in_filter(column: str, values: list[str]) -> bool:
        if values == ["__none__"]:
            clauses.append("1 = 0")
            return True
        if values:
            clauses.append(f"{column} IN ({', '.join('?' for _ in values)})")
            params.extend(values)
            return True
        return False

    for field, column in (("business_types", "b.business_type"), ("risk_levels", "b.risk_level"), ("statuses", "b.status")):
        values = _values(field)
        if not values:
            values = _values(field.removesuffix("s"))
        if _append_in_filter(column, values):
            continue
    if filters.get("business_type") and not _values("business_types"):
        clauses.append("b.business_type = ?")
        params.append(filters["business_type"])
    if filters.get("risk_level") and not _values("risk_levels"):
        clauses.append("b.risk_level = ?")
        params.append(filters["risk_level"])
    if filters.get("status") and not _values("statuses"):
        clauses.append("b.status = ?")
        params.append(filters["status"])
    fcr_values = _values("fcr")
    if fcr_values == ["__none__"]:
        clauses.append("1 = 0")
    elif fcr_values and set(fcr_values) != {"FCR", "非FCR"}:
        clauses.append("b.fcr = ?")
        params.append(1 if "FCR" in fcr_values else 0)
    anomaly_types = _values("anomaly_types")
    if anomaly_types:
        if anomaly_types == ["__none__"]:
            clauses.append("1 = 0")
        elif "无异常" in anomaly_types:
            clauses.append("((b.id NOT IN (SELECT business_id FROM order_lifecycle_data_anomalies WHERE status = 'open')) OR EXISTS (SELECT 1 FROM order_lifecycle_data_anomalies a0 WHERE a0.business_id = b.id AND a0.status = 'open' AND a0.anomaly_type IN (" + ", ".join("?" for _ in anomaly_types if _ != "无异常") + ")))" if any(item != "无异常" for item in anomaly_types) else "b.id NOT IN (SELECT business_id FROM order_lifecycle_data_anomalies WHERE status = 'open')")
            params.extend([item for item in anomaly_types if item != "无异常"])
        else:
            clauses.append("EXISTS (SELECT 1 FROM order_lifecycle_data_anomalies a1 WHERE a1.business_id = b.id AND a1.status = 'open' AND a1.anomaly_type IN (" + ", ".join("?" for _ in anomaly_types) + "))")
            params.extend(anomaly_types)
    keyword = _normalize_text(filters.get("keyword"))
    if keyword:
        like = f"%{keyword.lower()}%"
        search_field = _normalize_text(filters.get("search_field")) or "all"
        expressions = {
            "business_no": "b.business_no",
            "trade_entity": "b.trade_entity",
            "supplier_steel_mill": "b.supplier_steel_mill",
            "terminal_customer": "b.terminal_customer",
            "product_name": "b.product_name",
            "bank": "(SELECT GROUP_CONCAT(COALESCE(f.bank, '')) FROM order_lifecycle_financings f WHERE f.business_id = b.id)",
            "contract_no": "(SELECT GROUP_CONCAT(COALESCE(c.contract_no, '') || COALESCE(c.purchase_contract_no, '') || COALESCE(c.system_contract_no, '')) FROM order_lifecycle_contracts c WHERE c.business_id = b.id)",
            "date": "(SELECT GROUP_CONCAT(COALESCE(f.financing_date, '') || COALESCE(f.extended_due_date, '') || COALESCE(f.repayment_date, '') || COALESCE(r.receipt_date, '')) FROM order_lifecycle_financings f LEFT JOIN order_lifecycle_customer_receipts r ON r.business_id = f.business_id WHERE f.business_id = b.id)",
            "amount": "(SELECT GROUP_CONCAT(CAST(COALESCE(f.amount, 0) AS TEXT)) FROM order_lifecycle_financings f WHERE f.business_id = b.id)",
        }
        if search_field in expressions:
            clauses.append(f"LOWER(REPLACE(REPLACE(COALESCE({expressions[search_field]}, ''), ' ', ''), '-', '')) LIKE ?")
            params.append(like.replace(" ", "").replace("-", ""))
        else:
            clauses.append("LOWER(REPLACE(REPLACE(COALESCE(b.business_no, '') || COALESCE(b.product_name, '') || COALESCE(b.terminal_customer, '') || COALESCE(b.supplier_steel_mill, ''), ' ', ''), '-', '')) LIKE ?")
            params.append(like.replace(" ", "").replace("-", ""))
    where = " AND ".join(clauses)
    page = max(int(filters.get("page") or 1), 1)
    page_size = min(max(int(filters.get("page_size") or 20), 1), 100)
    with db.connect() as conn:
        cur = conn.cursor()
        all_rows = db._exec(cur, f"SELECT b.* FROM order_lifecycle_businesses b WHERE {where}", params).fetchall()
        business_ids = [row["id"] for row in all_rows]
        children_by_business = _load_business_children_batch(cur, business_ids)
        anomalies_by_business = _load_open_anomalies_batch(cur, business_ids)
        all_cards = [
            _serialize_business_card(
                row,
                children_by_business.get(row["id"], {collection: [] for collection in CHILD_TABLES}),
                anomalies_by_business.get(row["id"], []),
                bool(filters.get("can_sensitive")),
            )
            for row in all_rows
        ]
        for item in all_cards:
            item["weekly_focus_reasons"] = _weekly_focus_reasons(
                {
                    **item,
                    **children_by_business.get(item["id"], {}),
                }
            )
        risk_order = {"高风险": 0, "中风险": 1, "低风险": 2}
        all_cards.sort(key=lambda item: (
            1 if item.get("status") in {"已完结", "已结算"} else 0,
            risk_order.get(item.get("risk_level"), 3),
            _natural_sort_key(item.get("business_no")),
        ))
        view = _normalize_text(filters.get("view")) or "overview"
        if view == "focus":
            all_cards = [item for item in all_cards if item.get("weekly_focus_reasons")]
        total = len(all_cards)
        cards = all_cards[(page - 1) * page_size: page * page_size]
        summary = {
            "存续业务": 0,
            "其中进行中": 0,
            "已完结业务": 0,
            "存续融资金额": 0,
            "高风险": 0,
            "中风险": 0,
            "低风险": 0,
            "已完结": 0,
            "数据异常": 0,
            "更新异常": 0,
        }
        for item in all_cards:
            completed = item.get("status") in {"已完结", "已结算"}
            if completed:
                summary["已完结业务"] += 1
                summary["已完结"] += 1
            else:
                summary["存续业务"] += 1
                summary["存续融资金额"] += float(item.get("outstanding_financing_amount") or 0)
                if item.get("status") != "待确认":
                    summary["其中进行中"] += 1
                if item.get("risk_level") in {"高风险", "中风险", "低风险"}:
                    summary[item["risk_level"]] += 1
        sync = _row_json(db._exec(cur, "SELECT * FROM order_lifecycle_sync_state WHERE id = 1").fetchone())
        summary["数据异常"] = sum(1 for item in all_cards if item.get("anomalies"))
        summary["更新异常数"] = int(bool(sync.get("wps_last_error"))) + int(bool(sync.get("email_last_error")))
        summary["更新异常"] = summary.pop("更新异常数")
    return {"records": cards, "total": int(total), "page": page, "page_size": page_size, "summary": summary, "sync_status": sync}


def set_manual_fcr(business_id: int, enabled: bool, user: dict, note: str = "") -> dict[str, Any]:
    require_permission(user, PERMISSION_RESOURCE, "manage")
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT fcr FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not row:
            raise KeyError(business_id)
        old_value = bool(row["fcr"])
        db._exec(cur, "UPDATE order_lifecycle_businesses SET fcr = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (1 if enabled else 0, business_id))
        db._exec(cur, "INSERT INTO order_lifecycle_manual_overrides (business_id, field_name, value_json, note, modified_by, modified_at, is_active) VALUES (?, 'fcr', ?, ?, ?, CURRENT_TIMESTAMP, 1) ON CONFLICT (business_id, field_name) DO UPDATE SET value_json = excluded.value_json, note = excluded.note, modified_by = excluded.modified_by, modified_at = CURRENT_TIMESTAMP, is_active = 1", (business_id, _json(bool(enabled)), note or None, user.get("name") or user.get("username") or "unknown"))
        _audit_change(cur, business_id, "人工修改", "fcr", old_value, bool(enabled), user.get("name") or user.get("username") or "unknown", note)
        status, risk, _ = _recalculate_business(cur, business_id)
        conn.commit()
    db.log_operation(user["id"], ORDER_LIFECYCLE_MODULE, "人工修改", f"人工设置订单全流程FCR：{business_id}={bool(enabled)}")
    return {"business_id": business_id, "fcr": bool(enabled), "status": status, "risk_level": risk}


def _lifecycle_user(authorization: Optional[str] = Header(default=None)):
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.removeprefix("Bearer ").strip()
    user = db.get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="请先登录")
    require_permission(user, PERMISSION_RESOURCE, "view")
    return user


class ManualFcrRequest(BaseModel):
    enabled: bool
    note: str = Field(default="", max_length=500)


class ManualOverrideRequest(BaseModel):
    field_name: str = Field(min_length=1)
    value: Any
    note: str = Field(default="", max_length=500)


class LifecycleSettlementRequest(BaseModel):
    settled: bool
    date: str = Field(default="")
    note: str = Field(default="", max_length=500)


class LifecycleSpecialRequest(BaseModel):
    enabled: bool
    note: str = Field(default="", max_length=500)


class LifecycleChildOverrideRequest(BaseModel):
    collection: str = Field(min_length=1)
    source_key: str = Field(min_length=1)
    field_name: str = Field(min_length=1)
    value: Any
    note: str = Field(default="", max_length=500)


class LifecycleChildRecordRequest(BaseModel):
    collection: str = Field(min_length=1)
    source_key: str = Field(min_length=1)
    value: dict[str, Any] = Field(default_factory=dict)
    note: str = Field(default="", max_length=500)


class NodeConfirmationRequest(BaseModel):
    node: str = Field(pattern="^(集港|装船)$")
    confirmed: bool
    date: str = Field(default="")
    note: str = Field(default="", max_length=500)


class LocalImportRequest(BaseModel):
    path: str = Field(min_length=1)
    source_type: str = Field(pattern="^(wps|email)$")


class LifecycleUploadFile(BaseModel):
    file_name: str
    file_data: str


class LifecycleUploadRequest(BaseModel):
    source_type: str = Field(pattern="^(wps|email)$")
    files: list[LifecycleUploadFile] = Field(min_length=1, max_length=6)


@router.get("/order-lifecycle/progress")
def order_lifecycle_progress(
    keyword: str = "",
    search_field: str = "",
    business_type: str = "",
    risk_level: str = "",
    status: str = "",
    fcr: str = "",
    view: str = "overview",
    business_types: str = "",
    risk_levels: str = "",
    statuses: str = "",
    anomaly_types: str = "",
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    user: dict = Depends(_lifecycle_user),
):
    return list_businesses({"keyword": keyword, "search_field": search_field, "business_type": business_type, "risk_level": risk_level, "status": status, "fcr": fcr, "view": view, "business_types": business_types, "risk_levels": risk_levels, "statuses": statuses, "anomaly_types": anomaly_types, "page": page, "page_size": page_size, "can_sensitive": can(user, PERMISSION_RESOURCE, "manage")})


@router.get("/order-lifecycle/sync-status")
def order_lifecycle_sync_status(user: dict = Depends(_lifecycle_user)):
    with db.connect() as conn:
        row = db._exec(conn.cursor(), "SELECT * FROM order_lifecycle_sync_state WHERE id = 1").fetchone()
    return _row_json(row)


@router.get("/order-lifecycle/businesses/{business_id}")
def order_lifecycle_business_detail(business_id: int, user: dict = Depends(_lifecycle_user)):
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0 AND source_active = 1", (business_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="业务不存在")
        children = _load_business_children(cur, business_id)
        anomalies = [_row_json(item) for item in db._exec(cur, "SELECT * FROM order_lifecycle_data_anomalies WHERE business_id = ? ORDER BY id", (business_id,)).fetchall()]
        audit = [_row_json(item) for item in db._exec(cur, "SELECT * FROM order_lifecycle_audit WHERE business_id = ? ORDER BY changed_at DESC, id DESC", (business_id,)).fetchall()]
        overrides = [_row_json(item) for item in db._exec(cur, "SELECT * FROM order_lifecycle_manual_overrides WHERE business_id = ? ORDER BY modified_at DESC, id DESC", (business_id,)).fetchall()]
        item = _serialize_business(row, children, anomalies, can(user, PERMISSION_RESOURCE, "manage"))
        item["audit"] = audit
        item["manual_overrides"] = overrides
        item["sections"] = ["业务基本信息", "合同明细", "融资明细与银行还款", "执行进度 / 船舶明细", "单据 / 交单明细", "客户回款明细", "当前风险原因", "数据异常与人工修改记录"]
        return item


@router.patch("/order-lifecycle/businesses/{business_id}/settlement")
def order_lifecycle_settlement(business_id: int, request: LifecycleSettlementRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.date and not _parse_date(request.date):
        raise HTTPException(status_code=400, detail="结算日期格式无法识别")
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT settlement_status, settlement_date, business_type FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="业务不存在")
        old = {"settlement_status": row["settlement_status"], "settlement_date": row["settlement_date"]}
        new_status = "已结算" if request.settled else "待结算"
        new_date = request.date if request.settled else None
        db._exec(cur, "UPDATE order_lifecycle_businesses SET settlement_status = ?, settlement_date = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (new_status, new_date, business_id))
        _audit_change(cur, business_id, "人工修改", "settlement", old, {"settlement_status": new_status, "settlement_date": new_date}, user.get("name") or user.get("username") or "unknown", request.note)
        status, risk, _ = _recalculate_business(cur, business_id)
        conn.commit()
    return {"business_id": business_id, "settlement_status": new_status, "settlement_date": new_date, "status": status, "risk_level": risk}


@router.patch("/order-lifecycle/businesses/{business_id}/guo-danlei")
def order_lifecycle_special_flag(business_id: int, request: LifecycleSpecialRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT guo_danlei_special FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="业务不存在")
        old = bool(row["guo_danlei_special"])
        db._exec(cur, "UPDATE order_lifecycle_businesses SET guo_danlei_special = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (1 if request.enabled else 0, business_id))
        _audit_change(cur, business_id, "敏感人工标记", "guo_danlei_special", old, request.enabled, user.get("name") or user.get("username") or "unknown", request.note)
        status, risk, _ = _recalculate_business(cur, business_id)
        conn.commit()
    return {"business_id": business_id, "guo_danlei_special": request.enabled, "status": status, "risk_level": risk}


@router.patch("/order-lifecycle/businesses/{business_id}/child-override")
def order_lifecycle_child_override(business_id: int, request: LifecycleChildOverrideRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.collection not in CHILD_TABLES or request.field_name not in CHILD_FIELDS[request.collection]:
        raise HTTPException(status_code=400, detail="该子记录字段不支持人工覆盖")
    operator = user.get("name") or user.get("username") or "unknown"
    with db.connect() as conn:
        cur = conn.cursor()
        business = db._exec(cur, "SELECT id FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not business:
            raise HTTPException(status_code=404, detail="业务不存在")
        existing = db._exec(cur, f"SELECT {request.field_name} FROM {CHILD_TABLES[request.collection]} WHERE business_id = ? AND source_key = ?", (business_id, request.source_key)).fetchone()
        old_value = existing[request.field_name] if existing else None
        db._exec(
            cur,
            "INSERT INTO order_lifecycle_child_overrides (business_id, collection, source_key, field_name, value_json, note, modified_by, modified_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 1) ON CONFLICT (business_id, collection, source_key, field_name) DO UPDATE SET value_json = excluded.value_json, note = excluded.note, modified_by = excluded.modified_by, modified_at = CURRENT_TIMESTAMP, is_active = 1",
            (business_id, request.collection, request.source_key, request.field_name, _json(request.value), request.note or None, operator),
        )
        _audit_change(cur, business_id, "人工修改", f"{request.collection}.{request.source_key}.{request.field_name}", old_value, request.value, operator, request.note)
        _recalculate_business(cur, business_id)
        conn.commit()
    return {"business_id": business_id, "collection": request.collection, "source_key": request.source_key, "field_name": request.field_name, "value": request.value}


@router.post("/order-lifecycle/businesses/{business_id}/child-record")
def order_lifecycle_child_record(business_id: int, request: LifecycleChildRecordRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.collection not in CHILD_TABLES:
        raise HTTPException(status_code=400, detail="该子记录类型不支持人工新增")
    operator = user.get("name") or user.get("username") or "unknown"
    record = deepcopy(request.value)
    record["source_key"] = request.source_key
    with db.connect() as conn:
        cur = conn.cursor()
        business = db._exec(cur, "SELECT id FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not business:
            raise HTTPException(status_code=404, detail="业务不存在")
        db._exec(
            cur,
            "INSERT INTO order_lifecycle_manual_child_records (business_id, collection, source_key, record_json, note, modified_by, modified_at, is_active) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 1) ON CONFLICT (business_id, collection, source_key) DO UPDATE SET record_json = excluded.record_json, note = excluded.note, modified_by = excluded.modified_by, modified_at = CURRENT_TIMESTAMP, is_active = 1",
            (business_id, request.collection, request.source_key, _json(record), request.note or None, operator),
        )
        _audit_change(cur, business_id, "人工新增子记录", f"{request.collection}.{request.source_key}", None, record, operator, request.note)
        _recalculate_business(cur, business_id)
        conn.commit()
    return {"business_id": business_id, "collection": request.collection, "source_key": request.source_key, "record": record}


@router.post("/order-lifecycle/businesses/{business_id}/fcr")
def order_lifecycle_fcr(business_id: int, request: ManualFcrRequest, user: dict = Depends(_lifecycle_user)):
    try:
        return set_manual_fcr(business_id, request.enabled, user, request.note)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="业务不存在") from exc


@router.patch("/order-lifecycle/businesses/{business_id}/node-confirmation")
def order_lifecycle_node_confirmation(business_id: int, request: NodeConfirmationRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.date and not _parse_date(request.date):
        raise HTTPException(status_code=400, detail="节点日期格式无法识别")
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="业务不存在")
        operator = user.get("name") or user.get("username") or "unknown"
        if request.node == "集港":
            db._exec(cur, "UPDATE order_lifecycle_businesses SET port_status = ?, port_confirmed_date = ?, port_confirmed_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", ("已集港" if request.confirmed else "待确认", request.date if request.confirmed else None, operator if request.confirmed else None, business_id))
        else:
            db._exec(cur, "UPDATE order_lifecycle_businesses SET shipment_status = ?, shipment_confirmed_date = ?, shipment_confirmed_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", ("已装船" if request.confirmed else "待确认", request.date if request.confirmed else None, operator if request.confirmed else None, business_id))
        refreshed = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE id = ?", (business_id,)).fetchone()
        status, risk, _ = _recalculate_business(cur, business_id)
        _audit_change(cur, business_id, "人工修改", f"{request.node}_status", None, {"status": "已集港" if request.node == "集港" and request.confirmed else "已装船" if request.node == "装船" and request.confirmed else "待确认", "date": request.date if request.confirmed else None}, operator, request.note)
        conn.commit()
    db.log_operation(user["id"], ORDER_LIFECYCLE_MODULE, "人工确认", f"人工确认订单全流程节点：{business_id}/{request.node}/{request.confirmed}")
    return {"business_id": business_id, "node": request.node, "confirmed": request.confirmed, "date": request.date or None, "status": status, "risk_level": risk}


@router.patch("/order-lifecycle/businesses/{business_id}/override")
def order_lifecycle_override(business_id: int, request: ManualOverrideRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.field_name not in MANUAL_OVERRIDE_FIELDS:
        raise HTTPException(status_code=400, detail="该字段不支持人工覆盖")
    value = request.value
    if request.field_name == "business_no":
        value = _normalize_business_no(value)
        if not value:
            raise HTTPException(status_code=400, detail="真实业务编号不能为空")
    elif request.field_name == "business_type":
        value = _normalize_text(value)
        if value not in {"融资", "过单"}:
            raise HTTPException(status_code=400, detail="业务类型只能是融资或过单")
    elif request.field_name in {"port_status", "shipment_status"}:
        value = _normalize_text(value)
        if value not in {"待确认", "已集港", "已装船"}:
            raise HTTPException(status_code=400, detail="节点状态值无法识别")
    elif request.field_name in {"settlement_status"}:
        value = _normalize_text(value)
        if value not in {"待结算", "已结算"}:
            raise HTTPException(status_code=400, detail="结算状态值无法识别")
    elif request.field_name in {"port_confirmed_date", "shipment_confirmed_date", "settlement_date", "next_follow_up_date"}:
        value = _normalize_date(value)
        if request.value not in (None, "") and not value:
            raise HTTPException(status_code=400, detail="日期格式无法识别")
    elif request.field_name == "contract_quantity_mt" and value not in (None, ""):
        try:
            value = float(value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="合同数量必须是数字") from exc
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM order_lifecycle_businesses WHERE id = ? AND is_cancelled = 0", (business_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="业务不存在")
        if request.field_name == "business_no":
            owner = db._exec(cur, "SELECT id FROM order_lifecycle_businesses WHERE business_no = ? AND id <> ? AND is_cancelled = 0", (value, business_id)).fetchone()
            if owner:
                raise HTTPException(status_code=409, detail="该真实业务编号已被其他主卡使用")
            if value == _normalize_business_no(row["source_record_key"]):
                raise HTTPException(status_code=400, detail="不能使用来源记录键作为页面业务编号")
        old_value = row[request.field_name] if request.field_name in row.keys() else None
        operator = user.get("name") or user.get("username") or "unknown"
        db._exec(cur, "INSERT INTO order_lifecycle_manual_overrides (business_id, field_name, value_json, note, modified_by, modified_at, is_active) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 1) ON CONFLICT (business_id, field_name) DO UPDATE SET value_json = excluded.value_json, note = excluded.note, modified_by = excluded.modified_by, modified_at = CURRENT_TIMESTAMP, is_active = 1", (business_id, request.field_name, _json(value), request.note or None, operator))
        _audit_change(cur, business_id, "人工修改", request.field_name, old_value, value, operator, request.note)
        db._exec(cur, f"UPDATE order_lifecycle_businesses SET {request.field_name} = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (value, business_id))
        status, risk, _ = _recalculate_business(cur, business_id)
        conn.commit()
    db.log_operation(user["id"], ORDER_LIFECYCLE_MODULE, "人工修改", f"人工覆盖订单全流程字段：{business_id}/{request.field_name}")
    return {"business_id": business_id, "field_name": request.field_name, "value": value, "status": status, "risk_level": risk}


@router.post("/order-lifecycle/import-local")
def order_lifecycle_import_local(request: LocalImportRequest, user: dict = Depends(_lifecycle_user)):
    require_permission(user, PERMISSION_RESOURCE, "manage")
    path = Path(request.path)
    try:
        batch = parse_wps_workbook(path) if request.source_type == "wps" else parse_email_batch(path)
        return apply_source_batch(batch, imported_by=user.get("name") or "user")
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/order-lifecycle/import-upload")
def order_lifecycle_import_upload(request: LifecycleUploadRequest, user: dict = Depends(_lifecycle_user)):
    """Import a controlled WPS workbook or complete six-file mail snapshot in Staging."""
    require_permission(user, PERMISSION_RESOURCE, "manage")
    if request.source_type == "wps" and len(request.files) != 1:
        raise HTTPException(status_code=400, detail="WPS 快照只能上传一个 .xlsx 文件")
    if request.source_type == "email" and len(request.files) < len(MAIL_MILLS):
        raise HTTPException(status_code=400, detail="邮件台账必须一次上传六个钢厂附件")

    temp_dir = Path(tempfile.mkdtemp(prefix="order-lifecycle-upload-"))
    paths: list[Path] = []
    try:
        seen_names: set[str] = set()
        for item in request.files:
            name = Path(item.file_name or "").name
            suffix = Path(name).suffix.lower()
            if not name or name in seen_names:
                raise HTTPException(status_code=400, detail="上传文件名为空或重复")
            if suffix not in {".xls", ".xlsx"}:
                raise HTTPException(status_code=400, detail=f"不支持的文件格式：{name}")
            if request.source_type == "wps" and suffix != ".xlsx":
                raise HTTPException(status_code=400, detail="WPS 快照仅支持 .xlsx 文件")
            try:
                content = base64.b64decode(item.file_data, validate=True)
            except (binascii.Error, ValueError) as exc:
                raise HTTPException(status_code=400, detail=f"文件内容无效：{name}") from exc
            if not content:
                raise HTTPException(status_code=400, detail=f"文件为空：{name}")
            if len(content) > 30 * 1024 * 1024:
                raise HTTPException(status_code=400, detail=f"文件超过 30MB：{name}")
            target = temp_dir / name
            target.write_bytes(content)
            paths.append(target)
            seen_names.add(name)

        batch = parse_wps_workbook(paths[0]) if request.source_type == "wps" else parse_email_batch(temp_dir)
        batch["source_locator"] = f"staging-upload://{request.source_type}"
        result = apply_source_batch(batch, imported_by=user.get("name") or "user")
        return {"status": "success", "source_type": request.source_type, **result}
    except HTTPException:
        raise
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        for path in paths:
            path.unlink(missing_ok=True)
        temp_dir.rmdir()
