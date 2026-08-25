"""现货业务台账的来源适配、完整扫描、调度和历史迁移。"""

from __future__ import annotations

import ast
import base64
from dataclasses import dataclass, field
from datetime import date, datetime, time as day_time
from html.parser import HTMLParser
import json
import os
from pathlib import Path
import re
import threading
import time
from typing import Any, Callable, Optional
from urllib.parse import parse_qs, quote, urlparse
from uuid import uuid4
from zoneinfo import ZoneInfo

from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import padding
import requests
from openpyxl import load_workbook

from . import db
from .spot_ledger import (
    FIELD_CODES,
    FIELD_NAME_TO_CODE,
    FIELD_DEFINITIONS,
    MANUAL_FIELDS,
    SHANGHAI_GROUPS,
    SYSTEM_PRIORITY_FIELDS,
    calculate_derived_fields,
    initialize_schema,
    missing_required_fields,
    normalize_sales_contract_record,
    record_to_public,
)


SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
SPOT_LEDGER_SYNC_TIMES = tuple(day_time(hour, 0) for hour in range(9, 19))
CANDIDATE_SOURCE_URL = "https://tds-report.ejianlong.com/jmreport/show"
CANDIDATE_REPORT_ID = "1055351755192311808"
JIANLONG_AUTH_BASE_URL = "https://server-auth.ejianlong.com"
JIANLONG_TDS_API_BASE_URL = "https://tds-api.ejianlong.com"
OFFICIAL_SALES_CONTRACT_LIST_URL = f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/saleContractList"
OFFICIAL_SETTLEMENT_QUERY_URL = f"{JIANLONG_TDS_API_BASE_URL}/tdsSettle/queryJiesuan?sheetCode=G01112"
OFFICIAL_RESOURCE_CATALOG_URL = f"{JIANLONG_TDS_API_BASE_URL}/tradeing/sale/list?sheetCode=G01003"
OFFICIAL_DICTIONARY_URL = f"{JIANLONG_TDS_API_BASE_URL}/system/dict/data/type"
OFFICIAL_DEMAND_LIST_URL = f"{JIANLONG_TDS_API_BASE_URL}/tradeing/demand/list?sheetCode=G01002"
OFFICIAL_DEMAND_DETAIL_URL = f"{JIANLONG_TDS_API_BASE_URL}/tradeing/demand?sheetCode=G01002"
OFFICIAL_RESOURCE_DETAIL_URL = f"{JIANLONG_TDS_API_BASE_URL}/tradeing/sale?sheetCode=G01003"
OFFICIAL_CHAIN_SALE_CONTRACT_LIST_URL = (
    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/saleContractList?sheetCode=G01004"
)
JIANLONG_TDS_APP_ID = "2d948bd76f7b432193b6bb2823eee6a5"
JIANLONG_TDS_REDIRECT_URI = "https://tds.ejianlong.com/"
JIANLONG_SOURCE_USER_AGENT = "ltm-spot-ledger/1.0"
CONFIRMED_SOURCE_FIELD_MAP = {
    "detail_id": "销售合同商品明细id",
    "spot_type": "期现货",
    "contract_status": "合同状态",
    "quantity_group": "量归属组",
    "profit_group": "业务毛利归属组",
    "business_category_code": "业务类别",
    "operation_title": "公司",
    "resource_date": "资源日期",
    "product_name": "物资名称",
    "port": "合同卸货港",
    "mode": "定价模式",
    "vessel_name": "中文船名",
    "contract_quantity": "合同数量",
    "settlement_quantity": "结算数量",
    "is_closed": "结案状态",
    "purchase_price": "资源单单价",
    "supplier": "资源方",
    "purchase_business": "资源业务员",
    "purchase_execution": "初始资源单创建人",
    "signed_date": "签订日期",
    "sales_price": "合同单价",
    "demander": "需求方",
    "contract_number": "销售合同号",
    "sales_business": "需求业务员",
    "sales_execution": "合同创建人",
}
_scheduler_lock = threading.Lock()
_scheduler_started = False


class SalesContractSourceError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        stage: Optional[str] = None,
        http_status: Optional[int] = None,
    ):
        self.code = code
        self.stage = stage or code
        self.http_status = http_status
        super().__init__(f"{code}: {message}")


@dataclass
class FullScanResult:
    records: list[dict[str, Any]]
    page_count: int
    expected_page_count: Optional[int]
    total_count: int
    complete: bool
    errors: list[Any] = field(default_factory=list)
    source_mode: str = "fixture"
    diagnostics: dict[str, Any] = field(default_factory=dict)


class SalesContractSource:
    def fetch_full_scan(self) -> FullScanResult:
        raise NotImplementedError


class _LoginPageParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.inputs: dict[str, str] = {}
        self.scripts: list[str] = []
        self._in_script = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, Optional[str]]]) -> None:
        if tag.lower() == "script":
            self._in_script = True
            return
        if tag.lower() != "input":
            return
        values = {key: value for key, value in attrs}
        name = values.get("name")
        if name:
            self.inputs[name] = values.get("value") or ""

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "script":
            self._in_script = False

    def handle_data(self, data: str) -> None:
        if self._in_script:
            self.scripts.append(data)


def _parse_login_contract(html: str) -> tuple[dict[str, str], str]:
    parser = _LoginPageParser()
    parser.feed(html)
    required = ("redirectUri", "appId", "companyId")
    if any(name not in parser.inputs for name in required):
        raise SalesContractSourceError("auth_unavailable", "统一认证登录参数不可用", stage="login_contract")
    script = "\n".join(parser.scripts)
    match = re.search(r"\bvar\s+publicKey\s*=\s*(.*?);", script, flags=re.DOTALL)
    if not match:
        raise SalesContractSourceError("auth_unavailable", "统一认证公钥不可用", stage="login_contract")
    literals = re.findall(r"'(?:\\.|[^'\\])*'|\"(?:\\.|[^\"\\])*\"", match.group(1))
    try:
        public_key = "".join(ast.literal_eval(item) for item in literals)
    except (SyntaxError, ValueError) as exc:
        raise SalesContractSourceError("auth_unavailable", "统一认证公钥不可用", stage="login_contract") from exc
    if not public_key.startswith("-----BEGIN PUBLIC KEY-----") or not public_key.rstrip().endswith(
        "-----END PUBLIC KEY-----"
    ):
        raise SalesContractSourceError("auth_unavailable", "统一认证公钥不可用", stage="login_contract")
    return parser.inputs, public_key


def _encrypt_login_password(public_key: str, password: str) -> str:
    try:
        key = serialization.load_pem_public_key(public_key.encode("ascii"))
        encrypted = key.encrypt(password.encode("utf-8"), padding.PKCS1v15())
    except Exception as exc:
        raise SalesContractSourceError(
            "auth_unavailable",
            "统一认证密码加密失败",
            stage="password_encryption",
        ) from exc
    return base64.b64encode(encrypted).decode("ascii")


class JianlongPasswordAuthProvider:
    """Use the confirmed Jianlong SSO password flow without persisting credentials."""

    def __init__(self, username: str, password: str, *, http: Any = None, timeout_seconds: float = 30):
        self._username = username.strip()
        self._password = password
        self.http = http or requests.Session()
        headers = getattr(self.http, "headers", None)
        if hasattr(headers, "update"):
            headers.update({"User-Agent": JIANLONG_SOURCE_USER_AGENT})
        self.timeout_seconds = timeout_seconds
        self._token: Optional[str] = None
        self._lock = threading.Lock()

    @staticmethod
    def _json(response: Any, *, stage: str) -> dict[str, Any]:
        try:
            payload = response.json()
        except Exception as exc:
            raise SalesContractSourceError("auth_unavailable", "统一认证响应无效", stage=stage) from exc
        if not isinstance(payload, dict):
            raise SalesContractSourceError("auth_unavailable", "统一认证响应无效", stage=stage)
        return payload

    @staticmethod
    def _check_http(response: Any, *, stage: str) -> None:
        status = int(getattr(response, "status_code", 200))
        if status >= 400:
            raise SalesContractSourceError(
                "auth_unavailable",
                "统一认证请求失败",
                stage=stage,
                http_status=status,
            )

    def _request(self, method: str, url: str, *, stage: str, **kwargs: Any) -> Any:
        try:
            return getattr(self.http, method)(url, **kwargs)
        except SalesContractSourceError:
            raise
        except Exception as exc:
            raise SalesContractSourceError("auth_unavailable", "统一认证请求失败", stage=stage) from exc

    def _login(self) -> str:
        if not self._username or not self._password:
            raise SalesContractSourceError("auth_unavailable", "个人账号认证未配置", stage="credentials_missing")
        try:
            clear_cookies = getattr(getattr(self.http, "cookies", None), "clear", None)
            if callable(clear_cookies):
                clear_cookies()
            login_page = self._request(
                "get",
                f"{JIANLONG_AUTH_BASE_URL}/login",
                stage="login_page_request",
                params={"appId": JIANLONG_TDS_APP_ID, "redirectUri": JIANLONG_TDS_REDIRECT_URI},
                timeout=self.timeout_seconds,
            )
            self._check_http(login_page, stage="login_page_http")
            inputs, public_key = _parse_login_contract(str(getattr(login_page, "text", "")))
            password = _encrypt_login_password(public_key, self._password)
            login_response = self._request(
                "post",
                f"{JIANLONG_AUTH_BASE_URL}/login/pwd",
                stage="password_login_request",
                json={
                    "redirectUri": inputs["redirectUri"],
                    "appId": inputs["appId"],
                    "companyId": inputs["companyId"],
                    "username": self._username,
                    "password": password,
                },
                timeout=self.timeout_seconds,
            )
            self._check_http(login_response, stage="password_login_http")
            login_payload = self._json(login_response, stage="password_login_response")
            if str(login_payload.get("code")) != "200" or not isinstance(login_payload.get("data"), str):
                raise SalesContractSourceError("auth_unavailable", "个人账号认证失败", stage="credentials_rejected")
            code = parse_qs(urlparse(login_payload["data"]).query).get("code", [""])[0]
            if not code:
                raise SalesContractSourceError("auth_unavailable", "统一认证未返回登录票据", stage="login_ticket_missing")
            token_response = self._request(
                "get",
                f"{JIANLONG_TDS_API_BASE_URL}/login",
                stage="token_exchange_request",
                params={"code": code},
                headers={
                    "Origin": JIANLONG_TDS_REDIRECT_URI.rstrip("/"),
                    "Referer": JIANLONG_TDS_REDIRECT_URI,
                },
                timeout=self.timeout_seconds,
            )
            self._check_http(token_response, stage="token_exchange_http")
            token_payload = self._json(token_response, stage="token_exchange_response")
            token = token_payload.get("data")
            if str(token_payload.get("code")) != "200" or not isinstance(token, str) or not token:
                raise SalesContractSourceError("auth_unavailable", "统一认证未返回访问令牌", stage="token_missing")
            return token
        except SalesContractSourceError:
            raise
        except Exception as exc:
            raise SalesContractSourceError("auth_unavailable", "统一认证请求失败", stage="auth_request") from exc

    def __call__(self) -> dict[str, str]:
        with self._lock:
            if not self._token:
                self._token = self._login()
            return {"Authorization": f"Bearer {self._token}"}

    def refresh(self) -> dict[str, str]:
        with self._lock:
            self._token = self._login()
            return {"Authorization": f"Bearer {self._token}"}


def build_candidate_request_body(report_date: Any, *, page_no: int = 1, page_size: int = 20) -> dict[str, Any]:
    """Build the request body confirmed from the logged-in sales report page.

    This contains only the non-secret report identifier and business filters. It does
    not create or copy authentication material; the caller still needs an explicit,
    supported auth provider before the HTTP adapter can run.
    """
    if page_no < 1 or page_size < 1:
        raise ValueError("page_no 和 page_size 必须为正整数")
    period_date = report_date.isoformat() if isinstance(report_date, (date, datetime)) else str(report_date)
    params = {
        "pageNo": page_no,
        "periodDate": period_date,
        "releaseDate": period_date,
        "TJJLYSHZ__期现货": "现货",
        "TJJLYSHZ__合同状态": "生效",
        "TJJLYSHZ__业务毛利归属组": ",".join(SHANGHAI_GROUPS),
        "pageSize": str(page_size),
    }
    return {
        "id": CANDIDATE_REPORT_ID,
        "apiUrl": "",
        "params": json.dumps(params, ensure_ascii=False, separators=(",", ":")),
    }


def build_candidate_source_profile(report_date: Any, *, page_size: int = 20) -> dict[str, Any]:
    """Return the request/response profile confirmed from the JSON report response."""
    return {
        "url": CANDIDATE_SOURCE_URL,
        "request_body": build_candidate_request_body(report_date, page_no=1, page_size=page_size),
        "records_path": "result.dataList.TJJLYSHZ.list",
        "total_path": "result.dataList.TJJLYSHZ.count",
        "page_count_path": "result.dataList.TJJLYSHZ.total",
        "field_map": dict(CONFIRMED_SOURCE_FIELD_MAP),
        "pagination": {
            "params_key": "params",
            "page_number_key": "pageNo",
            "page_size_key": "pageSize",
            "page_size": page_size,
        },
    }


def _now(value: Optional[datetime] = None) -> datetime:
    current = value or datetime.now(SHANGHAI_TZ)
    return current.astimezone(SHANGHAI_TZ) if current.tzinfo else current.replace(tzinfo=SHANGHAI_TZ)


def _timestamp(value: Optional[datetime] = None) -> str:
    return _now(value).isoformat(timespec="seconds")


def _raw_execute(cur, sql: str, params: tuple[Any, ...] = ()):
    if db._is_pg():
        sql = sql.replace("?", "%s")
    cur.execute(sql, params)
    return cur


def _empty(value: Any) -> bool:
    return value is None or value == ""


def _fixture_records(payload: dict[str, Any]) -> FullScanResult:
    raw_records = payload.get("records")
    if not isinstance(raw_records, list):
        raise SalesContractSourceError("parse_error", "fixture.records 必须是数组")
    normalized = [normalize_sales_contract_record(item) for item in raw_records if isinstance(item, dict)]
    records = [item for item in normalized if item.get("eligible")]
    return FullScanResult(
        records=records,
        page_count=int(payload.get("page_count") or 1),
        expected_page_count=(int(payload["expected_page_count"]) if payload.get("expected_page_count") is not None else None),
        total_count=len(records),
        complete=bool(payload.get("complete", True)),
        errors=list(payload.get("errors") or []),
        source_mode="fixture",
    )


class FixtureSalesContractSource(SalesContractSource):
    """明确标记为本地 fixture 的只读 source。"""

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def fetch_full_scan(self) -> FullScanResult:
        if not self.path.exists():
            raise SalesContractSourceError("fixture_missing", f"fixture 不存在: {self.path}")
        try:
            payload = json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise SalesContractSourceError("fixture_parse_error", "fixture 读取失败") from exc
        if payload.get("source_mode") != "fixture":
            raise SalesContractSourceError("fixture_unmarked", "fixture 必须明确 source_mode=fixture")
        return validate_full_scan(_fixture_records(payload))


def _extract_path(payload: Any, path: str) -> Any:
    current = payload
    for part in path.split(".") if path else []:
        if isinstance(current, dict):
            if part not in current:
                raise KeyError(path)
            current = current[part]
        elif isinstance(current, list) and part.isdigit():
            current = current[int(part)]
        else:
            raise KeyError(path)
    return current


class ProfiledSalesContractSource(SalesContractSource):
    """仅执行外部提供的、已验证的 request/response profile。

    profile 不能省略认证 provider、请求体、分页路径和字段映射；本类不会根据网页
    或字段名称猜测协议。认证 provider 由调用方在运行时传入，不会写入仓库或日志。
    """

    def __init__(
        self,
        profile: Optional[dict[str, Any]],
        http: Any = requests,
        auth_provider: Optional[Callable[[], dict[str, str]]] = None,
    ):
        self.profile = profile or {}
        self.http = http
        self.auth_provider = auth_provider

    @classmethod
    def from_env(cls) -> "ProfiledSalesContractSource":
        profile_path = (os.getenv("SPOT_LEDGER_SOURCE_PROFILE") or "").strip()
        profile: dict[str, Any] = build_candidate_source_profile(datetime.now(SHANGHAI_TZ).date())
        if profile_path:
            try:
                profile = json.loads(Path(profile_path).read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise SalesContractSourceError("auth_unavailable", "source profile 读取失败") from exc
        username = (os.getenv("SPOT_LEDGER_SOURCE_USERNAME") or "").strip()
        password = os.getenv("SPOT_LEDGER_SOURCE_PASSWORD") or ""
        if username and password:
            session = requests.Session()
            provider = JianlongPasswordAuthProvider(username, password, http=session)
            return cls(profile, http=session, auth_provider=provider)
        return cls(profile)

    def _validate_profile(self) -> None:
        required = ("request_body", "records_path", "total_path", "page_count_path", "field_map", "pagination")
        if self.profile.get("url") != CANDIDATE_SOURCE_URL or any(key not in self.profile for key in required):
            raise SalesContractSourceError("auth_unavailable", "真实源 profile 或认证方式尚未确认")
        if not callable(self.auth_provider):
            raise SalesContractSourceError(
                "auth_unavailable",
                "无人值守认证 provider 尚未提供",
                stage="auth_provider_missing",
            )
        pagination = self.profile.get("pagination")
        if not isinstance(pagination, dict) or any(
            not isinstance(pagination.get(key), str) or not pagination.get(key)
            for key in ("params_key", "page_number_key", "page_size_key")
        ):
            raise SalesContractSourceError("auth_unavailable", "真实源分页 profile 尚未确认")

    def _request_body_for_page(self, page_no: int) -> dict[str, Any]:
        body = json.loads(json.dumps(self.profile["request_body"], ensure_ascii=False))
        pagination = self.profile["pagination"]
        params_key = pagination["params_key"]
        params = body.get(params_key)
        params_was_string = isinstance(params, str)
        if params_was_string:
            try:
                params = json.loads(params)
            except json.JSONDecodeError as exc:
                raise SalesContractSourceError("parse_error", "真实源分页参数不是有效 JSON") from exc
        if not isinstance(params, dict):
            raise SalesContractSourceError("parse_error", "真实源分页参数不是对象")
        params[pagination["page_number_key"]] = page_no
        if pagination.get("page_size") is not None:
            params[pagination["page_size_key"]] = str(pagination["page_size"])
        body[params_key] = json.dumps(params, ensure_ascii=False, separators=(",", ":")) if params_was_string else params
        return body

    @staticmethod
    def _needs_reauthentication(response: Any) -> bool:
        if getattr(response, "status_code", 200) in {401, 403}:
            return True
        response_url = str(getattr(response, "url", ""))
        parsed = urlparse(response_url)
        return parsed.hostname == "server-auth.ejianlong.com" and parsed.path.rstrip("/") == "/login"

    def _fetch_page(self, request_body: dict[str, Any]) -> tuple[list[dict[str, Any]], int, int]:
        try:
            headers = self.auth_provider() or {}
            response = self.http.post(
                self.profile["url"],
                json=request_body,
                headers=headers,
                timeout=float(self.profile.get("timeout_seconds", 30)),
            )
            if self._needs_reauthentication(response):
                refresh = getattr(self.auth_provider, "refresh", None)
                if callable(refresh):
                    response = self.http.post(
                        self.profile["url"],
                        json=request_body,
                        headers=refresh() or {},
                        timeout=float(self.profile.get("timeout_seconds", 30)),
                    )
        except SalesContractSourceError:
            raise
        except Exception as exc:
            raise SalesContractSourceError("source_request", "真实源请求失败") from exc
        if self._needs_reauthentication(response):
            raise SalesContractSourceError("auth_unavailable", "真实源认证失败", stage="report_auth")
        if getattr(response, "status_code", 200) >= 400:
            raise SalesContractSourceError("source_request", "真实源返回错误")
        try:
            payload = response.json()
            external_records = _extract_path(payload, self.profile["records_path"])
            total_count = int(_extract_path(payload, self.profile["total_path"]))
            page_count = int(_extract_path(payload, self.profile["page_count_path"]))
        except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
            raise SalesContractSourceError("parse_error", "真实源响应不符合已确认 profile") from exc
        if not isinstance(external_records, list):
            raise SalesContractSourceError("parse_error", "真实源记录路径不是数组")
        return external_records, total_count, page_count

    def _normalize_external_records(self, external_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        field_map = self.profile["field_map"]
        if not isinstance(field_map, dict) or not field_map:
            raise SalesContractSourceError("auth_unavailable", "真实源字段映射尚未确认")
        standard_records: list[dict[str, Any]] = []
        for external in external_records:
            if not isinstance(external, dict):
                raise SalesContractSourceError("parse_error", "真实源记录不是对象")
            try:
                standard_records.append({target: _extract_path(external, path) for target, path in field_map.items()})
            except (KeyError, TypeError) as exc:
                raise SalesContractSourceError("parse_error", "真实源字段映射无法读取记录") from exc
        records = [normalize_sales_contract_record(item) for item in standard_records]
        return [item for item in records if item.get("eligible")]

    def fetch_full_scan(self) -> FullScanResult:
        self._validate_profile()
        first_records, total_count, page_count = self._fetch_page(self._request_body_for_page(1))
        if page_count < 1 or total_count < 0:
            raise SalesContractSourceError("parse_error", "真实源分页统计值无效")
        max_pages = int(self.profile.get("max_pages", 1000))
        if page_count > max_pages:
            raise SalesContractSourceError("parse_error", "真实源分页数超过安全上限")
        external_records = list(first_records)
        for page_no in range(2, page_count + 1):
            page_records, page_total, page_count_again = self._fetch_page(self._request_body_for_page(page_no))
            if page_total != total_count or page_count_again != page_count:
                raise SalesContractSourceError("parse_error", "真实源分页统计在扫描过程中发生变化")
            external_records.extend(page_records)
        if len(external_records) != total_count:
            raise SalesContractSourceError("parse_error", "真实源分页记录数与总数不一致")
        records = self._normalize_external_records(external_records)
        return validate_full_scan(
            FullScanResult(
                records=records,
                page_count=page_count,
                expected_page_count=page_count,
                total_count=total_count,
                complete=True,
                errors=[],
                source_mode="profiled_http",
            )
        )


class OfficialJsonSalesContractSource(SalesContractSource):
    """Read the confirmed Jianlong JSON APIs and build normalized ledger rows.

    The adapter is deliberately read-only. It uses only the documented list/detail
    relationships confirmed against the official frontend and never persists the
    bearer token or source credentials.
    """

    DICTIONARY_TYPES = (
        "quantity_attribution",
        "profit_attribution",
        "source_type",
        "price_mode",
    )

    def __init__(
        self,
        *,
        http: Any = requests,
        auth_provider: Optional[Callable[[], dict[str, str]]] = None,
        page_size: int = 50,
        max_pages: int = 1000,
        timeout_seconds: float = 30,
    ):
        if page_size < 1 or max_pages < 1:
            raise ValueError("page_size 和 max_pages 必须为正整数")
        self.http = http
        self.auth_provider = auth_provider
        self.page_size = page_size
        self.max_pages = max_pages
        self.timeout_seconds = timeout_seconds

    @classmethod
    def from_env(cls) -> "OfficialJsonSalesContractSource":
        username = (os.getenv("SPOT_LEDGER_SOURCE_USERNAME") or "").strip()
        password = os.getenv("SPOT_LEDGER_SOURCE_PASSWORD") or ""
        if not username or not password:
            return cls()
        session = requests.Session()
        provider = JianlongPasswordAuthProvider(username, password, http=session)
        return cls(http=session, auth_provider=provider)

    def _headers(self, *, refresh: bool = False) -> dict[str, str]:
        if not callable(self.auth_provider):
            raise SalesContractSourceError(
                "auth_unavailable",
                "个人账号认证未配置",
                stage="auth_provider_missing",
            )
        provider = self.auth_provider
        if refresh:
            refresh_provider = getattr(provider, "refresh", None)
            if not callable(refresh_provider):
                raise SalesContractSourceError(
                    "auth_unavailable",
                    "真实源认证已失效且不能刷新",
                    stage="source_auth_refresh",
                )
            supplied = refresh_provider() or {}
        else:
            supplied = provider() or {}
        if not isinstance(supplied, dict):
            raise SalesContractSourceError(
                "auth_unavailable",
                "真实源认证 provider 返回值无效",
                stage="source_auth_headers",
            )
        return {
            **supplied,
            "Origin": JIANLONG_TDS_REDIRECT_URI.rstrip("/"),
            "Referer": JIANLONG_TDS_REDIRECT_URI,
        }

    def _request_json(self, method: str, url: str, *, stage: str, **kwargs: Any) -> dict[str, Any]:
        request = getattr(self.http, method)

        def send(headers: dict[str, str]) -> Any:
            return request(
                url,
                headers=headers,
                timeout=self.timeout_seconds,
                **kwargs,
            )

        try:
            response = send(self._headers())
            if ProfiledSalesContractSource._needs_reauthentication(response):
                response = send(self._headers(refresh=True))
        except SalesContractSourceError:
            raise
        except Exception as exc:
            raise SalesContractSourceError(
                "source_request",
                "正式 JSON 数据源请求失败",
                stage=f"{stage}_request",
            ) from exc
        if ProfiledSalesContractSource._needs_reauthentication(response):
            raise SalesContractSourceError(
                "auth_unavailable",
                "真实源认证失败",
                stage=f"{stage}_auth",
                http_status=int(getattr(response, "status_code", 401)),
            )
        status = int(getattr(response, "status_code", 200))
        if status >= 400:
            raise SalesContractSourceError(
                "source_request",
                "正式 JSON 数据源返回错误",
                stage=f"{stage}_http",
                http_status=status,
            )
        try:
            payload = response.json()
        except Exception as exc:
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 数据源响应无效",
                stage=f"{stage}_response",
            ) from exc
        if not isinstance(payload, dict) or str(payload.get("code") or "") not in {"", "200"}:
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 数据源响应不符合已确认协议",
                stage=f"{stage}_response",
            )
        return payload

    @staticmethod
    def _paged_rows(payload: dict[str, Any], *, stage: str) -> tuple[list[dict[str, Any]], int]:
        data = payload.get("data")
        rows = data.get("rows") if isinstance(data, dict) else None
        total = data.get("total") if isinstance(data, dict) else None
        if not isinstance(rows, list):
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 分页响应缺少 rows",
                stage=f"{stage}_rows",
            )
        try:
            total_count = int(total)
        except (TypeError, ValueError) as exc:
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 分页响应缺少有效 total",
                stage=f"{stage}_total",
            ) from exc
        if total_count < 0 or any(not isinstance(row, dict) for row in rows):
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 分页响应内容无效",
                stage=f"{stage}_rows",
            )
        return rows, total_count

    def _fetch_active_contracts(self) -> tuple[list[dict[str, Any]], int]:
        rows: list[dict[str, Any]] = []
        expected_total: Optional[int] = None
        page_count = 1
        for page_no in range(1, self.max_pages + 1):
            payload = self._request_json(
                "post",
                OFFICIAL_SALES_CONTRACT_LIST_URL,
                stage="official_contract_list",
                params={"sheetCode": "G01009", "pageNum": page_no, "pageSize": self.page_size},
                json={"status": "70", "isQryAll": "N"},
            )
            page_rows, total = self._paged_rows(payload, stage="official_contract_list")
            if expected_total is None:
                expected_total = total
                page_count = max(1, (total + self.page_size - 1) // self.page_size)
                if page_count > self.max_pages:
                    raise SalesContractSourceError(
                        "parse_error",
                        "正式销售合同分页数超过安全上限",
                        stage="official_contract_list_pages",
                    )
            elif total != expected_total:
                raise SalesContractSourceError(
                    "parse_error",
                    "正式销售合同分页统计在扫描过程中发生变化",
                    stage="official_contract_list_changed",
                )
            rows.extend(page_rows)
            if page_no >= page_count:
                break
        if expected_total is None or len(rows) != expected_total:
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同分页记录数与总数不一致",
                stage="official_contract_list_count",
            )
        return rows, page_count

    def _fetch_settlements(self) -> dict[str, float | int]:
        rows: list[dict[str, Any]] = []
        expected_total: Optional[int] = None
        page_count = 1
        for page_no in range(1, self.max_pages + 1):
            payload = self._request_json(
                "post",
                OFFICIAL_SETTLEMENT_QUERY_URL,
                stage="official_settlement",
                params={"pageNum": page_no, "pageSize": self.page_size},
                json={"status": "70"},
            )
            page_rows, total = self._paged_rows(payload, stage="official_settlement")
            if expected_total is None:
                expected_total = total
                page_count = max(1, (total + self.page_size - 1) // self.page_size)
                if page_count > self.max_pages:
                    raise SalesContractSourceError(
                        "parse_error",
                        "正式结算分页数超过安全上限",
                        stage="official_settlement_pages",
                    )
            elif total != expected_total:
                raise SalesContractSourceError(
                    "parse_error",
                    "正式结算分页统计在扫描过程中发生变化",
                    stage="official_settlement_changed",
                )
            rows.extend(page_rows)
            if page_no >= page_count:
                break
        if expected_total is None or len(rows) != expected_total:
            raise SalesContractSourceError(
                "parse_error",
                "正式结算分页记录数与总数不一致",
                stage="official_settlement_count",
            )
        totals: dict[str, float | int] = {}
        for row in rows:
            detail_id = str(row.get("saleContractMxId") or "").strip()
            quantity = row.get("countQuantity")
            if not detail_id or quantity in (None, ""):
                continue
            try:
                numeric = float(str(quantity).replace(",", ""))
            except ValueError:
                continue
            value: float | int = int(numeric) if numeric.is_integer() else numeric
            totals[detail_id] = totals.get(detail_id, 0) + value
        return totals

    def _fetch_dictionaries(self) -> dict[str, dict[str, str]]:
        dictionaries: dict[str, dict[str, str]] = {}
        for dictionary_type in self.DICTIONARY_TYPES:
            payload = self._request_json(
                "get",
                OFFICIAL_DICTIONARY_URL,
                stage=f"official_dictionary_{dictionary_type}",
                params={"dictType": dictionary_type},
            )
            rows = payload.get("data")
            if not isinstance(rows, list):
                raise SalesContractSourceError(
                    "parse_error",
                    "正式源字典响应不是数组",
                    stage=f"official_dictionary_{dictionary_type}_rows",
                )
            values: dict[str, str] = {}
            for row in rows:
                if not isinstance(row, dict):
                    continue
                code = str(row.get("dictValue") or "").strip()
                label = str(row.get("dictLabel") or "").strip()
                if code and label:
                    values[code] = label
            dictionaries[dictionary_type] = values
        return dictionaries

    @staticmethod
    def _dictionary_label(values: dict[str, str], raw: Any) -> tuple[str, bool]:
        text = str(raw or "").strip()
        if not text:
            return "", False
        if text in values:
            return values[text], True
        if text in values.values():
            return text, True
        return text, False

    def _get_data_dict(self, url: str, *, stage: str, **kwargs: Any) -> dict[str, Any]:
        payload = self._request_json("get", url, stage=stage, **kwargs)
        data = payload.get("data")
        if not isinstance(data, dict):
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 详情响应不是对象",
                stage=f"{stage}_data",
            )
        return data

    def _get_data_list(self, url: str, *, stage: str, **kwargs: Any) -> list[dict[str, Any]]:
        payload = self._request_json("get", url, stage=stage, **kwargs)
        data = payload.get("data")
        if not isinstance(data, list) or any(not isinstance(row, dict) for row in data):
            raise SalesContractSourceError(
                "parse_error",
                "正式 JSON 列表响应不是数组",
                stage=f"{stage}_data",
            )
        return data

    def _purchase_lines(self, contract_id: str) -> dict[str, dict[str, Any]]:
        relevance = self._get_data_list(
            (
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/getRelevanceContract/"
                f"{quote(contract_id, safe='')}?sheetCode=G01009"
            ),
            stage="official_relevance",
        )
        purchase_lines: dict[str, dict[str, Any]] = {}
        for row in relevance:
            purchase_id = str(row.get("purchaseContractId") or "").strip()
            if not purchase_id:
                continue
            detail = self._get_data_dict(
                (
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/purchaseContract/"
                    f"{quote(purchase_id, safe='')}?sheetCode=G01008"
                ),
                stage="official_purchase_detail",
            )
            lines = detail.get("tdsPurchaseContractMxVos")
            if not isinstance(lines, list):
                lines = detail.get("purchaseContractMxList")
            if not isinstance(lines, list):
                lines = []
            for line in lines:
                if not isinstance(line, dict):
                    continue
                line_id = str(line.get("purchaseContractMxId") or "").strip()
                if line_id:
                    purchase_lines[line_id] = line
        return purchase_lines

    def fetch_full_scan(self) -> FullScanResult:
        contract_rows, contract_page_count = self._fetch_active_contracts()
        settlements = self._fetch_settlements()
        dictionaries = self._fetch_dictionaries()
        normalized_records: list[dict[str, Any]] = []
        scan_errors: list[dict[str, str]] = []
        source_detail_count = 0
        out_of_scope_count = 0
        ambiguous_match_count = 0
        missing_match_count = 0
        demand_cache: dict[str, dict[str, Any]] = {}
        resource_cache: dict[str, dict[str, Any]] = {}

        for contract_row in contract_rows:
            contract_id = str(contract_row.get("saleContractId") or "").strip()
            if not contract_id:
                scan_errors.append({"type": "missing_contract_id"})
                continue
            detail = self._get_data_dict(
                (
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/"
                    f"{quote(contract_id, safe='')}?sheetCode=G01009"
                ),
                stage="official_contract_detail",
            )
            lines = detail.get("tdsSaleContractMxVos")
            if not isinstance(lines, list):
                lines = detail.get("saleContractMxList")
            if not isinstance(lines, list):
                scan_errors.append({"type": "missing_sale_lines", "contract_id": contract_id})
                continue
            lines = [line for line in lines if isinstance(line, dict)]
            source_detail_count += len(lines)
            purchase_lines = self._purchase_lines(contract_id)
            traders_id = str(detail.get("syncTradersId") or "").strip()
            match_rows = (
                self._get_data_list(
                    (
                        f"{JIANLONG_TDS_API_BASE_URL}/chain/goods/matchResult/"
                        f"{quote(traders_id, safe='')}?sheetCode=G01004"
                    ),
                    stage="official_match_result",
                )
                if traders_id
                else []
            )
            matches_by_goods: dict[str, list[dict[str, Any]]] = {}
            for match in match_rows:
                goods_code = str(match.get("goodsCode") or "").strip()
                if goods_code:
                    matches_by_goods.setdefault(goods_code, []).append(match)
            sale_goods_counts: dict[str, int] = {}
            for line in lines:
                goods_code = str(line.get("goodsCode") or "").strip()
                if goods_code:
                    sale_goods_counts[goods_code] = sale_goods_counts.get(goods_code, 0) + 1

            for line in lines:
                detail_id = str(line.get("saleContractMxId") or line.get("id") or "").strip()
                goods_code = str(line.get("goodsCode") or "").strip()
                candidates = matches_by_goods.get(goods_code, []) if goods_code else []
                if not detail_id:
                    scan_errors.append({"type": "missing_detail_id", "contract_id": contract_id})
                    continue
                if not candidates:
                    missing_match_count += 1
                    scan_errors.append({"type": "missing_resource_match", "detail_id": detail_id})
                    continue
                if len(candidates) != 1 or sale_goods_counts.get(goods_code, 0) != 1:
                    ambiguous_match_count += 1
                    scan_errors.append({"type": "ambiguous_resource_match", "detail_id": detail_id})
                    continue
                match = candidates[0]
                demand_id = str(match.get("demandId") or "").strip()
                if not demand_id:
                    missing_match_count += 1
                    scan_errors.append({"type": "missing_demand_link", "detail_id": detail_id})
                    continue
                if demand_id not in demand_cache:
                    demand_cache[demand_id] = self._get_data_dict(
                        OFFICIAL_DEMAND_DETAIL_URL,
                        stage="official_demand_detail",
                        params={"demandId": demand_id},
                    )
                demand = demand_cache[demand_id]
                quantity_group, quantity_group_known = self._dictionary_label(
                    dictionaries["quantity_attribution"],
                    demand.get("quantityAttribution"),
                )
                profit_group, profit_group_known = self._dictionary_label(
                    dictionaries["profit_attribution"],
                    demand.get("profitAttribution"),
                )
                source_type, source_type_known = self._dictionary_label(
                    dictionaries["source_type"],
                    demand.get("sourceType"),
                )
                if not quantity_group_known or not profit_group_known or not source_type_known:
                    scan_errors.append({"type": "missing_dictionary_mapping", "detail_id": detail_id})
                    continue
                status = str(detail.get("status") or contract_row.get("status") or "").strip()
                contract_active = status in {"70", "生效"}
                if source_type != "现货" or not contract_active or quantity_group not in SHANGHAI_GROUPS:
                    out_of_scope_count += 1
                    continue

                resource_id = str(match.get("saleId") or "").strip()
                resource: dict[str, Any] = {}
                record_errors: list[dict[str, str]] = []
                if resource_id:
                    if resource_id not in resource_cache:
                        resource_cache[resource_id] = self._get_data_dict(
                            OFFICIAL_RESOURCE_DETAIL_URL,
                            stage="official_resource_detail",
                            params={"saleId": resource_id},
                        )
                    resource = resource_cache[resource_id]
                else:
                    record_errors.append(
                        {
                            "type": "missing_resource_detail",
                            "field": "G",
                            "message": "资源匹配结果缺少资源单 ID",
                        }
                    )

                purchase_line_id = str(line.get("upContractMxId") or "").strip()
                purchase_line = purchase_lines.get(purchase_line_id, {})
                price_mode, _ = self._dictionary_label(dictionaries["price_mode"], line.get("priceMode"))
                standard = {
                    "detail_id": detail_id,
                    "spot_type": source_type,
                    "contract_status": "生效" if contract_active else status,
                    "quantity_group": quantity_group,
                    "profit_group": profit_group,
                    "business_category_code": demand.get("businessType"),
                    "operation_title": detail.get("workCompName") or demand.get("workCompName"),
                    "resource_date": resource.get("sourceDate"),
                    "product_name": line.get("goodsName"),
                    "product_category": line.get("goodsName"),
                    "port": detail.get("dischargePortName"),
                    "mode": price_mode,
                    "vessel_name": resource.get("chineseShipName") or detail.get("chineseShipName"),
                    "contract_quantity": line.get("countQuantity")
                    if line.get("countQuantity") not in (None, "")
                    else line.get("signQuantity"),
                    "settlement_quantity": settlements.get(detail_id),
                    "is_closed": detail_id in settlements,
                    "purchase_price": match.get("matchPrice"),
                    "supplier": resource.get("supplierName")
                    or match.get("startSupplierName")
                    or match.get("lastSupplierName"),
                    "purchase_business": resource.get("workManName") or purchase_line.get("workManName"),
                    "purchase_execution": resource.get("createBy") or purchase_line.get("createBy"),
                    "signed_date": detail.get("signingDate"),
                    "sales_price": line.get("unitPrice")
                    if line.get("unitPrice") not in (None, "")
                    else line.get("taxPrice") or line.get("price"),
                    "demander": detail.get("coustomName"),
                    "contract_number": detail.get("contractCode"),
                    "sales_business": detail.get("workManName"),
                    "sales_execution": detail.get("createBy"),
                }
                record = normalize_sales_contract_record(standard)
                record["sync_errors"].extend(record_errors)
                if record.get("eligible"):
                    normalized_records.append(record)
                else:
                    out_of_scope_count += 1

        diagnostics = {
            "active_contract_count": len(contract_rows),
            "source_detail_count": source_detail_count,
            "eligible_record_count": len(normalized_records),
            "out_of_scope_record_count": out_of_scope_count,
            "ambiguous_resource_match_count": ambiguous_match_count,
            "missing_resource_match_count": missing_match_count,
        }
        return validate_full_scan(
            FullScanResult(
                records=normalized_records,
                page_count=contract_page_count,
                expected_page_count=contract_page_count,
                total_count=len(normalized_records),
                complete=not scan_errors,
                errors=scan_errors,
                source_mode="official_json",
                diagnostics=diagnostics,
            )
        )


def _safe_error_type(error: Any) -> str:
    value = error.get("type") if isinstance(error, dict) else error
    text = str(value or "").strip().lower()
    return text if re.fullmatch(r"[a-z][a-z0-9_]{0,63}", text) else "unknown"


def _error_type_counts(errors: list[Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for error in errors:
        error_type = _safe_error_type(error)
        counts[error_type] = counts.get(error_type, 0) + 1
    return dict(sorted(counts.items()))


def summarize_official_source_scan(scan: FullScanResult) -> dict[str, Any]:
    """Return dry-run evidence without exposing source records or identifiers."""
    diagnostic_defaults = {
        "active_contract_count": 0,
        "source_detail_count": 0,
        "eligible_record_count": len(scan.records),
        "out_of_scope_record_count": 0,
        "ambiguous_resource_match_count": 0,
        "missing_resource_match_count": 0,
    }
    counts: dict[str, int] = {}
    for name, default in diagnostic_defaults.items():
        value = (scan.diagnostics or {}).get(name, default)
        try:
            counts[name] = max(0, int(value))
        except (TypeError, ValueError):
            counts[name] = default
    field_coverage = {
        code: {
            "filled_count": sum(not _empty(record.get(code)) for record in scan.records),
            "total_count": len(scan.records),
        }
        for code in FIELD_CODES
    }
    record_errors = [
        error
        for record in scan.records
        for error in (record.get("sync_errors") or [])
    ]
    return {
        "ok": bool(scan.complete),
        "source_mode": scan.source_mode,
        "page_count": scan.page_count,
        "expected_page_count": scan.expected_page_count,
        "counts": counts,
        "field_coverage": field_coverage,
        "scan_error_types": _error_type_counts(list(scan.errors or [])),
        "record_error_types": _error_type_counts(record_errors),
    }


def run_official_source_dry_run(
    source: Optional[OfficialJsonSalesContractSource] = None,
) -> dict[str, Any]:
    """Execute a full read-only source scan and return aggregate evidence only."""
    scan = (source or OfficialJsonSalesContractSource.from_env()).fetch_full_scan()
    return summarize_official_source_scan(scan)


def run_profiled_source_dry_run(
    source: Optional[ProfiledSalesContractSource] = None,
) -> dict[str, Any]:
    """Read the confirmed JSON report without persisting records or source values."""
    scan = (source or ProfiledSalesContractSource.from_env()).fetch_full_scan()
    return summarize_official_source_scan(scan)


def probe_official_scope_filters(
    *,
    source: Optional[OfficialJsonSalesContractSource] = None,
) -> dict[str, Any]:
    """Confirm server-side scope paths using aggregate, read-only evidence only."""
    active_source = source or OfficialJsonSalesContractSource.from_env()
    dictionaries = active_source._fetch_dictionaries()
    quantity_codes = {
        label: code
        for code, label in dictionaries.get("quantity_attribution", {}).items()
        if label in SHANGHAI_GROUPS
    }
    spot_code = next(
        (code for code, label in dictionaries.get("source_type", {}).items() if label == "现货"),
        "",
    )
    if not spot_code or len(quantity_codes) != len(SHANGHAI_GROUPS):
        raise SalesContractSourceError(
            "parse_error",
            "正式源范围字典不完整",
            stage="official_scope_dictionary",
        )

    group_counts: dict[str, int] = {}
    sampled_group_count = 0
    sample_match_count = 0
    demand_schema_paths: set[str] = set()
    for group in SHANGHAI_GROUPS:
        group_code = quantity_codes[group]
        payload = active_source._request_json(
            "get",
            OFFICIAL_DEMAND_LIST_URL,
            stage="official_scope_demand_list",
            params={
                "pageNum": 1,
                "pageSize": 10,
                "sourceType": spot_code,
                "quantityAttribution": group_code,
            },
        )
        demand_schema_paths.update(_schema_paths(payload))
        rows, total = active_source._paged_rows(payload, stage="official_scope_demand_list")
        group_counts[group] = total
        if not rows:
            continue
        sampled_group_count += 1
        sample = rows[0]
        demand_id = str(sample.get("demandId") or "").strip()
        if sample.get("sourceType") in (None, "") or sample.get("quantityAttribution") in (None, ""):
            if not demand_id:
                continue
            sample = active_source._get_data_dict(
                OFFICIAL_DEMAND_DETAIL_URL,
                stage="official_scope_demand_detail",
                params={"demandId": demand_id},
            )
        sample_source, source_known = active_source._dictionary_label(
            dictionaries["source_type"],
            sample.get("sourceType"),
        )
        sample_group, group_known = active_source._dictionary_label(
            dictionaries["quantity_attribution"],
            sample.get("quantityAttribution"),
        )
        if source_known and group_known and sample_source == "现货" and sample_group == group:
            sample_match_count += 1

    local_page_size = 500
    local_rows: list[dict[str, Any]] = []
    local_total: Optional[int] = None
    local_page_count = 1
    for page_no in range(1, active_source.max_pages + 1):
        payload = active_source._request_json(
            "get",
            OFFICIAL_DEMAND_LIST_URL,
            stage="official_scope_local_demand_list",
            params={"pageNum": page_no, "pageSize": local_page_size},
        )
        demand_schema_paths.update(_schema_paths(payload))
        rows, total = active_source._paged_rows(
            payload,
            stage="official_scope_local_demand_list",
        )
        if local_total is None:
            local_total = total
            local_page_count = max(1, (total + local_page_size - 1) // local_page_size)
            if local_page_count > active_source.max_pages:
                raise SalesContractSourceError(
                    "parse_error",
                    "正式需求分页数超过安全上限",
                    stage="official_scope_local_demand_pages",
                )
        elif total != local_total:
            raise SalesContractSourceError(
                "parse_error",
                "正式需求分页统计在扫描过程中发生变化",
                stage="official_scope_local_demand_changed",
            )
        local_rows.extend(rows)
        if page_no >= local_page_count:
            break
    if local_total is None or len(local_rows) != local_total:
        raise SalesContractSourceError(
            "parse_error",
            "正式需求分页记录数与总数不一致",
            stage="official_scope_local_demand_count",
        )

    local_group_counts = {group: 0 for group in SHANGHAI_GROUPS}
    spot_demand_count = 0
    seven_group_spot_demand_count = 0
    duplicate_demand_id_count = 0
    seen_demand_ids: set[str] = set()
    in_scope_demand_ids: list[str] = []
    for row in local_rows:
        source_type, source_known = active_source._dictionary_label(
            dictionaries["source_type"],
            row.get("sourceType"),
        )
        quantity_group, quantity_group_known = active_source._dictionary_label(
            dictionaries["quantity_attribution"],
            row.get("quantityAttribution"),
        )
        if source_known and source_type == "现货":
            spot_demand_count += 1
        if not (
            source_known
            and quantity_group_known
            and source_type == "现货"
            and quantity_group in SHANGHAI_GROUPS
        ):
            continue
        seven_group_spot_demand_count += 1
        local_group_counts[quantity_group] += 1
        demand_id = str(row.get("demandId") or "").strip()
        if not demand_id:
            continue
        if demand_id in seen_demand_ids:
            duplicate_demand_id_count += 1
            continue
        seen_demand_ids.add(demand_id)
        in_scope_demand_ids.append(demand_id)

    local_scope_scan = {
        "source_total_count": local_total,
        "page_count": local_page_count,
        "scanned_row_count": len(local_rows),
        "duplicate_demand_id_count": duplicate_demand_id_count,
        "spot_demand_count": spot_demand_count,
        "seven_group_spot_demand_count": seven_group_spot_demand_count,
        "group_counts": local_group_counts,
    }

    related_chains: list[dict[str, Any]] = []
    related_chain_schema_paths: set[str] = set()
    related_sales: list[dict[str, Any]] = []
    related_sale_schema_paths: set[str] = set()
    related_chain_sample_attempt_count = 0
    for demand_id in in_scope_demand_ids[:20]:
        related_chain_sample_attempt_count += 1
        related_payload = active_source._request_json(
            "get",
            (
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/relatedToDemand/"
                f"{quote(demand_id, safe='')}?sheetCode=G01004"
            ),
            stage="official_scope_related_chain",
        )
        related_chain_schema_paths.update(_schema_paths(related_payload))
        related_data = related_payload.get("data")
        if isinstance(related_data, list):
            related_chains = [row for row in related_data if isinstance(row, dict)]
        if related_chains:
            break

    if related_chains:
        chain_id = next(
            (str(row.get("chainId")) for row in related_chains if row.get("chainId") not in (None, "")),
            "",
        )
        if chain_id:
            sale_payload = active_source._request_json(
                "post",
                OFFICIAL_CHAIN_SALE_CONTRACT_LIST_URL,
                stage="official_scope_chain_sales",
                json={"chainId": chain_id, "tradersId": ""},
            )
            related_sale_schema_paths.update(_schema_paths(sale_payload))
            sale_data = sale_payload.get("data")
            if isinstance(sale_data, list):
                related_sales = [row for row in sale_data if isinstance(row, dict)]

    settlement_baseline_payload = active_source._request_json(
        "post",
        OFFICIAL_SETTLEMENT_QUERY_URL,
        stage="official_scope_settlement_baseline",
        params={"pageNum": 1, "pageSize": 10},
        json={"status": "70"},
    )
    settlement_rows, settlement_baseline_total = active_source._paged_rows(
        settlement_baseline_payload,
        stage="official_scope_settlement_baseline",
    )
    settlement_filters: dict[str, dict[str, Any]] = {}
    settlement_sample = settlement_rows[0] if settlement_rows else {}
    for field_name in ("salesContractNo", "saleContractId", "saleContractMxId"):
        value = settlement_sample.get(field_name)
        if value in (None, ""):
            settlement_filters[field_name] = {
                "total": 0,
                "sample_count": 0,
                "sample_match_count": 0,
                "effective": False,
            }
            continue
        filtered_payload = active_source._request_json(
            "post",
            OFFICIAL_SETTLEMENT_QUERY_URL,
            stage=f"official_scope_settlement_{field_name}",
            params={"pageNum": 1, "pageSize": 10},
            json={"status": "70", field_name: value},
        )
        filtered_rows, filtered_total = active_source._paged_rows(
            filtered_payload,
            stage=f"official_scope_settlement_{field_name}",
        )
        match_count = sum(str(row.get(field_name) or "") == str(value) for row in filtered_rows)
        settlement_filters[field_name] = {
            "total": filtered_total,
            "sample_count": len(filtered_rows),
            "sample_match_count": match_count,
            "effective": bool(
                filtered_rows
                and filtered_total < settlement_baseline_total
                and match_count == len(filtered_rows)
            ),
        }

    active_related_sales = sum(str(row.get("status") or "") in {"70", "生效"} for row in related_sales)
    filter_confirmed = any(item["effective"] for item in settlement_filters.values())
    return {
        "ok": bool(
            sampled_group_count
            and sample_match_count == sampled_group_count
            and related_chains
            and related_sales
            and filter_confirmed
        ),
        "source_mode": "official_json",
        "demand_filter": {
            "group_counts": group_counts,
            "sampled_group_count": sampled_group_count,
            "sample_match_count": sample_match_count,
        },
        "local_scope_scan": local_scope_scan,
        "demand_schema_paths": sorted(demand_schema_paths)[:300],
        "related_chain_sample_attempt_count": related_chain_sample_attempt_count,
        "related_chain_count": len(related_chains),
        "related_chain_schema_paths": sorted(related_chain_schema_paths)[:300],
        "related_sale_contract_count": len(related_sales),
        "related_active_sale_contract_count": active_related_sales,
        "related_sale_contract_schema_paths": sorted(related_sale_schema_paths)[:300],
        "settlement_filter_baseline_total": settlement_baseline_total,
        "settlement_filters": settlement_filters,
    }


def _schema_paths(value: Any, prefix: str = "", depth: int = 0) -> set[str]:
    if depth > 6:
        return set()
    paths: set[str] = set()
    if isinstance(value, dict):
        for key in sorted(value):
            path = f"{prefix}.{key}" if prefix else str(key)
            paths.add(path)
            paths.update(_schema_paths(value[key], path, depth + 1))
    elif isinstance(value, list):
        path = f"{prefix}[]"
        paths.add(path)
        if value:
            paths.update(_schema_paths(value[0], path, depth + 1))
    return paths


def _probe_json_request(
    source: ProfiledSalesContractSource,
    method: str,
    url: str,
    *,
    headers: dict[str, str],
    stage: str,
    description: str,
    **kwargs: Any,
) -> tuple[dict[str, Any], str]:
    try:
        response = getattr(source.http, method)(url, headers=headers, timeout=30, **kwargs)
    except Exception as exc:
        raise SalesContractSourceError(
            "source_request",
            f"{description}请求失败",
            stage=f"{stage}_request",
        ) from exc
    status = int(getattr(response, "status_code", 200))
    if status >= 400:
        raise SalesContractSourceError(
            "source_request",
            f"{description}返回错误",
            stage=f"{stage}_http",
            http_status=status,
        )
    try:
        payload = response.json()
    except Exception as exc:
        raise SalesContractSourceError(
            "parse_error",
            f"{description}响应无效",
            stage=f"{stage}_response",
        ) from exc
    if not isinstance(payload, dict):
        raise SalesContractSourceError(
            "parse_error",
            f"{description}响应无效",
            stage=f"{stage}_response",
        )
    return payload, str(payload.get("code") or "")


def _payload_rows(payload: dict[str, Any], *paths: tuple[str, ...]) -> list[dict[str, Any]]:
    for path in paths:
        current: Any = payload
        for key in path:
            current = current.get(key) if isinstance(current, dict) else None
        if isinstance(current, list):
            return [row for row in current if isinstance(row, dict)]
    return []


def _payload_total(payload: dict[str, Any]) -> int:
    data = payload.get("data")
    value = data.get("total") if isinstance(data, dict) else None
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return 0


def _nonempty_ids(rows: list[dict[str, Any]], key: str) -> set[str]:
    return {str(row[key]) for row in rows if row.get(key) not in (None, "")}


def _linkage_counts(
    detail_payload: dict[str, Any],
    purchase_payload: dict[str, Any],
    match_payload: dict[str, Any],
    demand_goods_payload: dict[str, Any],
    settlement_payload: dict[str, Any],
) -> dict[str, int]:
    sale_lines = _payload_rows(
        detail_payload,
        ("data", "tdsSaleContractMxVos"),
        ("data", "saleContractMxList"),
    )
    purchase_lines = _payload_rows(
        purchase_payload,
        ("data", "tdsPurchaseContractMxVos"),
        ("data", "purchaseContractMxList"),
    )
    match_rows = _payload_rows(match_payload, ("data",))
    demand_goods = _payload_rows(demand_goods_payload, ("data",))
    settlement_rows = _payload_rows(settlement_payload, ("data", "rows"))
    sale_ids = {
        "detail": _nonempty_ids(sale_lines, "saleContractMxId"),
        "business": _nonempty_ids(sale_lines, "businessDetailId"),
        "chain_good": _nonempty_ids(sale_lines, "chainGoodId"),
        "relevance": _nonempty_ids(sale_lines, "relevanceId"),
        "up_contract": _nonempty_ids(sale_lines, "upContractMxId"),
        "goods_code": _nonempty_ids(sale_lines, "goodsCode"),
    }
    sale_goods_specs = {
        (str(row.get("goodsCode")), str(row.get("specification")))
        for row in sale_lines
        if row.get("goodsCode") not in (None, "") and row.get("specification") not in (None, "")
    }
    demand_goods_specs = {
        (str(row.get("goodsCode")), str(row.get("spec")))
        for row in demand_goods
        if row.get("goodsCode") not in (None, "") and row.get("spec") not in (None, "")
    }
    return {
        "sale_lines": len(sale_lines),
        "purchase_lines": len(purchase_lines),
        "settlement_rows": len(settlement_rows),
        "match_rows": len(match_rows),
        "settlement_to_sale_detail": sum(
            str(row.get("saleContractMxId")) in sale_ids["detail"]
            for row in settlement_rows
            if row.get("saleContractMxId") not in (None, "")
        ),
        "match_demand_to_sale_detail": sum(
            str(row.get("demandId")) in sale_ids["detail"]
            for row in match_rows
            if row.get("demandId") not in (None, "")
        ),
        "match_demand_to_sale_business_detail": sum(
            str(row.get("demandId")) in sale_ids["business"]
            for row in match_rows
            if row.get("demandId") not in (None, "")
        ),
        "match_demand_to_sale_chain_good": sum(
            str(row.get("demandId")) in sale_ids["chain_good"]
            for row in match_rows
            if row.get("demandId") not in (None, "")
        ),
        "match_demand_to_sale_relevance": sum(
            str(row.get("demandId")) in sale_ids["relevance"]
            for row in match_rows
            if row.get("demandId") not in (None, "")
        ),
        "match_goods_to_sale_goods_code": sum(
            str(row.get("goodsCode")) in sale_ids["goods_code"]
            for row in match_rows
            if row.get("goodsCode") not in (None, "")
        ),
        "match_goods_and_specs_to_sale_line": sum(
            (str(row.get("goodsCode")), str(row.get("specs"))) in sale_goods_specs
            for row in match_rows
            if row.get("goodsCode") not in (None, "") and row.get("specs") not in (None, "")
        ),
        "match_goods_and_specs_to_demand_goods": sum(
            (str(row.get("goodsCode")), str(row.get("specs"))) in demand_goods_specs
            for row in match_rows
            if row.get("goodsCode") not in (None, "") and row.get("specs") not in (None, "")
        ),
        "demand_goods_to_sale_goods_code": sum(
            str(row.get("goodsCode")) in sale_ids["goods_code"]
            for row in demand_goods
            if row.get("goodsCode") not in (None, "")
        ),
        "demand_goods_and_spec_to_sale_line": sum(
            (str(row.get("goodsCode")), str(row.get("spec"))) in sale_goods_specs
            for row in demand_goods
            if row.get("goodsCode") not in (None, "") and row.get("spec") not in (None, "")
        ),
        "demand_goods_detail_to_sale_business_detail": sum(
            str(row.get("goodsDetailId")) in sale_ids["business"]
            for row in demand_goods
            if row.get("goodsDetailId") not in (None, "")
        ),
        "demand_goods_chain_good_to_sale_chain_good": sum(
            str(row.get("chainGoodId")) in sale_ids["chain_good"]
            for row in demand_goods
            if row.get("chainGoodId") not in (None, "")
        ),
        "purchase_to_sale_up_contract": sum(
            str(row.get("purchaseContractMxId")) in sale_ids["up_contract"]
            for row in purchase_lines
            if row.get("purchaseContractMxId") not in (None, "")
        ),
        "purchase_business_detail_to_sale_business_detail": sum(
            str(row.get("businessDetailId")) in sale_ids["business"]
            for row in purchase_lines
            if row.get("businessDetailId") not in (None, "")
        ),
        "purchase_chain_good_to_sale_chain_good": sum(
            str(row.get("chainGoodId")) in sale_ids["chain_good"]
            for row in purchase_lines
            if row.get("chainGoodId") not in (None, "")
        ),
        "purchase_relevance_to_sale_relevance": sum(
            str(row.get("relevanceId")) in sale_ids["relevance"]
            for row in purchase_lines
            if row.get("relevanceId") not in (None, "")
        ),
    }


def _group_dictionary_coverage(payload: dict[str, Any]) -> int:
    labels = {
        str(row.get("dictLabel") or "").strip()
        for row in _payload_rows(payload, ("data",))
    }
    return len(labels.intersection(SHANGHAI_GROUPS))


def _demand_group_coverage(payload: dict[str, Any]) -> dict[str, int]:
    data = payload.get("data")
    return {
        "quantity": int(isinstance(data, dict) and data.get("quantityAttribution") not in (None, "")),
        "profit": int(isinstance(data, dict) and data.get("profitAttribution") not in (None, "")),
    }


def probe_official_sales_contract_api(
    *,
    source: Optional[ProfiledSalesContractSource] = None,
) -> dict[str, Any]:
    active_source = source or ProfiledSalesContractSource.from_env()
    if not callable(active_source.auth_provider):
        raise SalesContractSourceError(
            "auth_unavailable",
            "无人值守认证 provider 尚未提供",
            stage="auth_provider_missing",
        )
    try:
        headers = {
            **(active_source.auth_provider() or {}),
            "Origin": JIANLONG_TDS_REDIRECT_URI.rstrip("/"),
            "Referer": JIANLONG_TDS_REDIRECT_URI,
        }
        response = active_source.http.post(
            OFFICIAL_SALES_CONTRACT_LIST_URL,
            params={"sheetCode": "G01009", "pageNum": 1, "pageSize": 10},
            json={"status": "70", "isQryAll": "N"},
            headers=headers,
            timeout=30,
        )
    except SalesContractSourceError:
        raise
    except Exception as exc:
        raise SalesContractSourceError(
            "source_request",
            "正式销售合同 JSON 接口请求失败",
            stage="official_list_request",
        ) from exc
    status = int(getattr(response, "status_code", 200))
    if status >= 400:
        raise SalesContractSourceError(
            "source_request",
            "正式销售合同 JSON 接口返回错误",
            stage="official_list_http",
            http_status=status,
        )
    try:
        payload = response.json()
    except Exception as exc:
        raise SalesContractSourceError(
            "parse_error",
            "正式销售合同 JSON 接口响应无效",
            stage="official_list_response",
        ) from exc
    if not isinstance(payload, dict):
        raise SalesContractSourceError(
            "parse_error",
            "正式销售合同 JSON 接口响应无效",
            stage="official_list_response",
        )
    response_code = str(payload.get("code") or "")
    rows = (payload.get("data") or {}).get("rows") if isinstance(payload.get("data"), dict) else None
    detail_payload: dict[str, Any] = {}
    detail_response_code = ""
    relevance_payload: dict[str, Any] = {}
    relevance_response_code = ""
    purchase_payload: dict[str, Any] = {}
    purchase_response_code = ""
    chain_payload: dict[str, Any] = {}
    chain_response_code = ""
    resource_list_payload: dict[str, Any] = {}
    resource_list_response_code = ""
    match_payload: dict[str, Any] = {}
    match_response_code = ""
    resource_detail_payload: dict[str, Any] = {}
    resource_detail_response_code = ""
    demand_detail_payload: dict[str, Any] = {}
    demand_detail_response_code = ""
    demand_goods_payload: dict[str, Any] = {}
    demand_goods_response_code = ""
    settlement_payload: dict[str, Any] = {}
    settlement_response_code = ""
    resource_catalog_payload: dict[str, Any] = {}
    resource_catalog_response_code = ""
    quantity_group_dictionary_payload: dict[str, Any] = {}
    quantity_group_dictionary_response_code = ""
    profit_group_dictionary_payload: dict[str, Any] = {}
    profit_group_dictionary_response_code = ""
    sampled_contract_count = 0
    resource_id: Any = None
    detail_schema_path_set: set[str] = set()
    chain_schema_path_set: set[str] = set()
    resource_list_schema_path_set: set[str] = set()
    match_schema_path_set: set[str] = set()
    seen_chain_ids: set[str] = set()
    seen_traders_ids: set[str] = set()
    if isinstance(rows, list) and rows:
        sampled_contract_count = 1
        contract_id = rows[0].get("saleContractId") if isinstance(rows[0], dict) else None
        if not contract_id:
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同 JSON 列表缺少合同 ID",
                stage="official_list_contract_id",
            )
        try:
            detail_response = active_source.http.get(
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/{quote(str(contract_id), safe='')}?sheetCode=G01009",
                headers=headers,
                timeout=30,
            )
        except Exception as exc:
            raise SalesContractSourceError(
                "source_request",
                "正式销售合同 JSON 详情请求失败",
                stage="official_detail_request",
            ) from exc
        detail_status = int(getattr(detail_response, "status_code", 200))
        if detail_status >= 400:
            raise SalesContractSourceError(
                "source_request",
                "正式销售合同 JSON 详情返回错误",
                stage="official_detail_http",
                http_status=detail_status,
            )
        try:
            detail_payload = detail_response.json()
        except Exception as exc:
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同 JSON 详情响应无效",
                stage="official_detail_response",
            ) from exc
        if not isinstance(detail_payload, dict):
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同 JSON 详情响应无效",
                stage="official_detail_response",
            )
        detail_response_code = str(detail_payload.get("code") or "")
        detail_schema_path_set.update(_schema_paths(detail_payload))
        try:
            relevance_response = active_source.http.get(
                (
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/getRelevanceContract/"
                    f"{quote(str(contract_id), safe='')}?sheetCode=G01009"
                ),
                headers=headers,
                timeout=30,
            )
        except Exception as exc:
            raise SalesContractSourceError(
                "source_request",
                "正式销售合同 JSON 关联合同请求失败",
                stage="official_relevance_request",
            ) from exc
        relevance_status = int(getattr(relevance_response, "status_code", 200))
        if relevance_status >= 400:
            raise SalesContractSourceError(
                "source_request",
                "正式销售合同 JSON 关联合同返回错误",
                stage="official_relevance_http",
                http_status=relevance_status,
            )
        try:
            relevance_payload = relevance_response.json()
        except Exception as exc:
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同 JSON 关联合同响应无效",
                stage="official_relevance_response",
            ) from exc
        if not isinstance(relevance_payload, dict):
            raise SalesContractSourceError(
                "parse_error",
                "正式销售合同 JSON 关联合同响应无效",
                stage="official_relevance_response",
            )
        relevance_response_code = str(relevance_payload.get("code") or "")
        relevance_rows = relevance_payload.get("data")
        purchase_contract_id = next(
            (
                row.get("purchaseContractId")
                for row in relevance_rows
                if isinstance(row, dict) and row.get("purchaseContractId")
            ),
            None,
        ) if isinstance(relevance_rows, list) else None
        if purchase_contract_id:
            try:
                purchase_response = active_source.http.get(
                    (
                        f"{JIANLONG_TDS_API_BASE_URL}/tradeing/purchaseContract/"
                        f"{quote(str(purchase_contract_id), safe='')}?sheetCode=G01008"
                    ),
                    headers=headers,
                    timeout=30,
                )
            except Exception as exc:
                raise SalesContractSourceError(
                    "source_request",
                    "正式采购合同 JSON 详情请求失败",
                    stage="official_purchase_detail_request",
                ) from exc
            purchase_status = int(getattr(purchase_response, "status_code", 200))
            if purchase_status >= 400:
                raise SalesContractSourceError(
                    "source_request",
                    "正式采购合同 JSON 详情返回错误",
                    stage="official_purchase_detail_http",
                    http_status=purchase_status,
                )
            try:
                purchase_payload = purchase_response.json()
            except Exception as exc:
                raise SalesContractSourceError(
                    "parse_error",
                    "正式采购合同 JSON 详情响应无效",
                    stage="official_purchase_detail_response",
                ) from exc
            if not isinstance(purchase_payload, dict):
                raise SalesContractSourceError(
                    "parse_error",
                    "正式采购合同 JSON 详情响应无效",
                    stage="official_purchase_detail_response",
                )
            purchase_response_code = str(purchase_payload.get("code") or "")
        detail_data = detail_payload.get("data")
        chain_id = detail_data.get("chainId") if isinstance(detail_data, dict) else None
        if chain_id:
            seen_chain_ids.add(str(chain_id))
            quoted_chain_id = quote(str(chain_id), safe="")
            chain_payload, chain_response_code = _probe_json_request(
                active_source,
                "get",
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/getById/{quoted_chain_id}?sheetCode=G01004",
                headers=headers,
                stage="official_chain_detail",
                description="正式交易链 JSON 详情",
            )
            chain_schema_path_set.update(_schema_paths(chain_payload))
            resource_list_payload, resource_list_response_code = _probe_json_request(
                active_source,
                "post",
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/saleList?sheetCode=G01004",
                headers=headers,
                stage="official_resource_list",
                description="正式销售资源 JSON 列表",
                json={"chainId": str(chain_id), "tradersId": ""},
            )
            resource_list_schema_path_set.update(_schema_paths(resource_list_payload))
            resource_rows = resource_list_payload.get("data")
            resource_id = next(
                (row.get("saleId") for row in resource_rows if isinstance(row, dict) and row.get("saleId")),
                None,
            ) if isinstance(resource_rows, list) else None
        sync_traders_id = detail_data.get("syncTradersId") if isinstance(detail_data, dict) else None
        if not resource_id and sync_traders_id:
            seen_traders_ids.add(str(sync_traders_id))
            match_payload, match_response_code = _probe_json_request(
                active_source,
                "get",
                (
                    f"{JIANLONG_TDS_API_BASE_URL}/chain/goods/matchResult/"
                    f"{quote(str(sync_traders_id), safe='')}?sheetCode=G01004"
                ),
                headers=headers,
                stage="official_match_result",
                description="正式交易链资源匹配 JSON",
            )
            match_schema_path_set.update(_schema_paths(match_payload))
            match_rows = match_payload.get("data")
            resource_id = next(
                (row.get("saleId") for row in match_rows if isinstance(row, dict) and row.get("saleId")),
                None,
            ) if isinstance(match_rows, list) else None
        if not resource_id:
            for extra_row in rows[1:5]:
                extra_contract_id = extra_row.get("saleContractId") if isinstance(extra_row, dict) else None
                if not extra_contract_id:
                    continue
                sampled_contract_count += 1
                extra_detail_payload, _ = _probe_json_request(
                    active_source,
                    "get",
                    (
                        f"{JIANLONG_TDS_API_BASE_URL}/tradeing/saleContract/"
                        f"{quote(str(extra_contract_id), safe='')}?sheetCode=G01009"
                    ),
                    headers=headers,
                    stage="official_sample_detail",
                    description="正式销售合同 JSON 抽样详情",
                )
                detail_schema_path_set.update(_schema_paths(extra_detail_payload))
                extra_detail_data = extra_detail_payload.get("data")
                extra_chain_id = extra_detail_data.get("chainId") if isinstance(extra_detail_data, dict) else None
                if not extra_chain_id or str(extra_chain_id) in seen_chain_ids:
                    continue
                seen_chain_ids.add(str(extra_chain_id))
                quoted_chain_id = quote(str(extra_chain_id), safe="")
                chain_payload, chain_response_code = _probe_json_request(
                    active_source,
                    "get",
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/getById/{quoted_chain_id}?sheetCode=G01004",
                    headers=headers,
                    stage="official_sample_chain_detail",
                    description="正式交易链 JSON 抽样详情",
                )
                chain_schema_path_set.update(_schema_paths(chain_payload))
                resource_list_payload, resource_list_response_code = _probe_json_request(
                    active_source,
                    "post",
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/chain/saleList?sheetCode=G01004",
                    headers=headers,
                    stage="official_sample_resource_list",
                    description="正式销售资源 JSON 抽样列表",
                    json={"chainId": str(extra_chain_id), "tradersId": ""},
                )
                resource_list_schema_path_set.update(_schema_paths(resource_list_payload))
                resource_rows = resource_list_payload.get("data")
                resource_id = next(
                    (row.get("saleId") for row in resource_rows if isinstance(row, dict) and row.get("saleId")),
                    None,
                ) if isinstance(resource_rows, list) else None
                extra_sync_traders_id = (
                    extra_detail_data.get("syncTradersId") if isinstance(extra_detail_data, dict) else None
                )
                if (
                    not resource_id
                    and extra_sync_traders_id
                    and str(extra_sync_traders_id) not in seen_traders_ids
                ):
                    seen_traders_ids.add(str(extra_sync_traders_id))
                    match_payload, match_response_code = _probe_json_request(
                        active_source,
                        "get",
                        (
                            f"{JIANLONG_TDS_API_BASE_URL}/chain/goods/matchResult/"
                            f"{quote(str(extra_sync_traders_id), safe='')}?sheetCode=G01004"
                        ),
                        headers=headers,
                        stage="official_sample_match_result",
                        description="正式交易链资源匹配 JSON 抽样",
                    )
                    match_schema_path_set.update(_schema_paths(match_payload))
                    match_rows = match_payload.get("data")
                    resource_id = next(
                        (row.get("saleId") for row in match_rows if isinstance(row, dict) and row.get("saleId")),
                        None,
                    ) if isinstance(match_rows, list) else None
                if resource_id:
                    break
        if resource_id:
            resource_detail_payload, resource_detail_response_code = _probe_json_request(
                active_source,
                "get",
                f"{JIANLONG_TDS_API_BASE_URL}/tradeing/sale?sheetCode=G01003",
                headers=headers,
                stage="official_resource_detail",
                description="正式销售资源 JSON 详情",
                params={"saleId": str(resource_id)},
            )
        match_rows = match_payload.get("data")
        demand_id = next(
            (row.get("demandId") for row in match_rows if isinstance(row, dict) and row.get("demandId")),
            None,
        ) if isinstance(match_rows, list) else None
        if demand_id:
            demand_detail_payload, demand_detail_response_code = _probe_json_request(
                active_source,
                "get",
                OFFICIAL_DEMAND_DETAIL_URL,
                headers=headers,
                stage="official_demand_detail",
                description="正式需求资源 JSON 详情",
                params={"demandId": str(demand_id)},
            )
            demand_goods_payload, demand_goods_response_code = _probe_json_request(
                active_source,
                "get",
                (
                    f"{JIANLONG_TDS_API_BASE_URL}/tradeing/goods/list/"
                    f"{quote(str(demand_id), safe='')}?sheetCode=G01002"
                ),
                headers=headers,
                stage="official_demand_goods",
                description="正式需求资源物资 JSON 列表",
            )
    settlement_payload, settlement_response_code = _probe_json_request(
        active_source,
        "post",
        OFFICIAL_SETTLEMENT_QUERY_URL,
        headers=headers,
        stage="official_settlement_query",
        description="正式结算 JSON 查询",
        params={"pageNum": 1, "pageSize": 10},
        json={"status": "70"},
    )
    resource_catalog_payload, resource_catalog_response_code = _probe_json_request(
        active_source,
        "get",
        OFFICIAL_RESOURCE_CATALOG_URL,
        headers=headers,
        stage="official_resource_catalog",
        description="正式销售资源 JSON 台账",
        params={
            "saleNo": "",
            "workCompId": "",
            "coustomName": "",
            "workDeptList": "",
            "workMan": "",
            "status": "70",
            "pageNum": 1,
            "pageSize": 10,
            "sheetCode": "G01003",
        },
    )
    quantity_group_dictionary_payload, quantity_group_dictionary_response_code = _probe_json_request(
        active_source,
        "get",
        OFFICIAL_DICTIONARY_URL,
        headers=headers,
        stage="official_quantity_group_dictionary",
        description="正式量归属组字典",
        params={"dictType": "quantity_attribution"},
    )
    profit_group_dictionary_payload, profit_group_dictionary_response_code = _probe_json_request(
        active_source,
        "get",
        OFFICIAL_DICTIONARY_URL,
        headers=headers,
        stage="official_profit_group_dictionary",
        description="正式毛利归属组字典",
        params={"dictType": "profit_attribution"},
    )
    return {
        "ok": (
            response_code in {"", "200"}
            and detail_response_code in {"", "200"}
            and relevance_response_code in {"", "200"}
            and purchase_response_code in {"", "200"}
            and chain_response_code in {"", "200"}
            and resource_list_response_code in {"", "200"}
            and match_response_code in {"", "200"}
            and resource_detail_response_code in {"", "200"}
            and demand_detail_response_code in {"", "200"}
            and demand_goods_response_code in {"", "200"}
            and settlement_response_code in {"", "200"}
            and resource_catalog_response_code in {"", "200"}
            and quantity_group_dictionary_response_code in {"", "200"}
            and profit_group_dictionary_response_code in {"", "200"}
        ),
        "source_mode": "official_json",
        "http_status": status,
        "response_code": response_code,
        "detail_response_code": detail_response_code,
        "relevance_response_code": relevance_response_code,
        "purchase_response_code": purchase_response_code,
        "chain_response_code": chain_response_code,
        "resource_list_response_code": resource_list_response_code,
        "match_response_code": match_response_code,
        "resource_detail_response_code": resource_detail_response_code,
        "demand_detail_response_code": demand_detail_response_code,
        "demand_goods_response_code": demand_goods_response_code,
        "settlement_response_code": settlement_response_code,
        "resource_catalog_response_code": resource_catalog_response_code,
        "quantity_group_dictionary_response_code": quantity_group_dictionary_response_code,
        "profit_group_dictionary_response_code": profit_group_dictionary_response_code,
        "sampled_contract_count": sampled_contract_count,
        "active_contract_total": _payload_total(payload),
        "settlement_row_total": _payload_total(settlement_payload),
        "schema_paths": sorted(_schema_paths(payload))[:300],
        "detail_schema_paths": sorted(detail_schema_path_set)[:300],
        "relevance_schema_paths": sorted(_schema_paths(relevance_payload))[:300],
        "purchase_schema_paths": sorted(_schema_paths(purchase_payload))[:300],
        "chain_schema_paths": sorted(chain_schema_path_set)[:300],
        "resource_list_schema_paths": sorted(resource_list_schema_path_set)[:300],
        "match_schema_paths": sorted(match_schema_path_set)[:300],
        "resource_detail_schema_paths": sorted(_schema_paths(resource_detail_payload))[:300],
        "demand_detail_schema_paths": sorted(_schema_paths(demand_detail_payload))[:300],
        "demand_goods_schema_paths": sorted(_schema_paths(demand_goods_payload))[:300],
        "settlement_schema_paths": sorted(_schema_paths(settlement_payload))[:300],
        "resource_catalog_schema_paths": sorted(_schema_paths(resource_catalog_payload))[:300],
        "quantity_group_dictionary_schema_paths": sorted(_schema_paths(quantity_group_dictionary_payload))[:300],
        "profit_group_dictionary_schema_paths": sorted(_schema_paths(profit_group_dictionary_payload))[:300],
        "group_dictionary_coverage": {
            "quantity": _group_dictionary_coverage(quantity_group_dictionary_payload),
            "profit": _group_dictionary_coverage(profit_group_dictionary_payload),
        },
        "sampled_demand_group_coverage": _demand_group_coverage(demand_detail_payload),
        "linkage_counts": _linkage_counts(
            detail_payload,
            purchase_payload,
            match_payload,
            demand_goods_payload,
            settlement_payload,
        ),
    }


def validate_full_scan(scan: FullScanResult) -> FullScanResult:
    errors = list(scan.errors or [])
    ids = [str(record.get("source_detail_id") or "") for record in scan.records]
    if scan.expected_page_count is not None and scan.page_count != scan.expected_page_count:
        errors.append("page_count_mismatch")
    if scan.total_count != len(scan.records):
        errors.append("total_count_mismatch")
    if any(not detail_id for detail_id in ids):
        errors.append("missing_detail_id")
    if len(ids) != len(set(ids)):
        errors.append("duplicate_detail_id")
    complete = bool(scan.complete and not errors)
    return FullScanResult(
        records=scan.records,
        page_count=scan.page_count,
        expected_page_count=scan.expected_page_count,
        total_count=scan.total_count,
        complete=complete,
        errors=errors,
        source_mode=scan.source_mode,
        diagnostics=dict(scan.diagnostics or {}),
    )


def _record_id(detail_id: str) -> str:
    return f"spot:{detail_id}"


def _payload_for_storage(record: dict[str, Any]) -> str:
    return json.dumps({key: value for key, value in record.items() if key != "sync_errors"}, ensure_ascii=False, default=str)


def _merge_record(incoming: dict[str, Any], existing: Optional[dict[str, Any]]) -> dict[str, Any]:
    merged = dict(incoming)
    if existing:
        for field in MANUAL_FIELDS:
            if field in existing and not _empty(existing.get(field)):
                merged[field] = existing[field]
        # K 是系统优先补录字段：源有有效船名时覆盖，否则保留既有人工值。
        if _empty(incoming.get("K")) and not _empty(existing.get("K")):
            merged["K"] = existing["K"]
        merged["record_id"] = existing.get("record_id") or merged.get("record_id")
    return calculate_derived_fields(merged)


def _record_columns() -> list[str]:
    return [
        "record_id", "source_detail_id", "source_closed_state", "record_source_type", *FIELD_CODES, "long_contract_object", "eligible",
        "is_active", "supplement_status", "missing_fields", "sync_status", "last_synced_at", "sync_error_summary",
        "source_payload_json", "source_mode",
    ]


def _record_values(record: dict[str, Any], *, existing: Optional[dict[str, Any]], source_mode: str, timestamp: str) -> list[Any]:
    serialized = dict(record)
    serialized["record_id"] = record.get("record_id") or _record_id(str(record["source_detail_id"]))
    serialized["source_detail_id"] = record.get("source_detail_id")
    serialized["source_closed_state"] = record.get("source_closed_state") or "未结案"
    serialized["record_source_type"] = "现货同步"
    serialized["long_contract_object"] = record.get("long_contract_object") or ""
    missing = missing_required_fields(record)
    sync_errors = record.get("sync_errors") or []
    serialized["eligible"] = 1 if record.get("eligible") else 0
    serialized["is_active"] = 1
    serialized["supplement_status"] = "待补录" if missing else "已完成"
    serialized["missing_fields"] = json.dumps(missing, ensure_ascii=False)
    serialized["sync_status"] = "异常" if sync_errors else "正常"
    serialized["last_synced_at"] = timestamp
    serialized["sync_error_summary"] = json.dumps(sync_errors, ensure_ascii=False) if sync_errors else ""
    serialized["source_payload_json"] = _payload_for_storage(record)
    serialized["source_mode"] = source_mode
    return [serialized.get(column, "") for column in _record_columns()]


def _upsert_record(cur, record: dict[str, Any], source_mode: str, timestamp: str) -> tuple[bool, bool]:
    detail_id = str(record.get("source_detail_id") or "")
    existing_row = _raw_execute(cur, "SELECT * FROM spot_ledger_records WHERE source_detail_id = ?", (detail_id,)).fetchone()
    existing = dict(existing_row) if existing_row else None
    record = _merge_record({**record, "record_id": (existing or {}).get("record_id") or _record_id(detail_id)}, existing)
    columns = _record_columns()
    values = _record_values(record, existing=existing, source_mode=source_mode, timestamp=timestamp)
    quoted_columns = [f'"{column}"' if column in FIELD_CODES else column for column in columns]
    if existing:
        assignments = ", ".join(f"{column} = ?" for column in quoted_columns if column != "record_id")
        update_values = [value for column, value in zip(quoted_columns, values) if column != "record_id"]
        _raw_execute(cur, f"UPDATE spot_ledger_records SET {assignments}, updated_at = ? WHERE record_id = ?", (*update_values, timestamp, existing["record_id"]))
        return False, bool(record.get("sync_errors"))
    _raw_execute(
        cur,
        f"INSERT INTO spot_ledger_records ({', '.join(quoted_columns)}, created_at, updated_at) VALUES ({', '.join('?' for _ in columns)}, ?, ?)",
        (*values, timestamp, timestamp),
    )
    return True, bool(record.get("sync_errors"))


def _insert_run(cur, slot_key: str, started_at: str, finished_at: str, scan: FullScanResult, result: dict[str, Any]) -> None:
    status = "成功" if scan.complete else "异常"
    errors = list(scan.errors or []) + list(result.get("record_errors") or [])
    _raw_execute(
        cur,
        """
        INSERT INTO spot_ledger_sync_runs
            (id, slot_key, started_at, finished_at, status, source_mode, page_count,
             expected_page_count, total_count, inserted_count, updated_count, hidden_count,
             error_count, error_summary)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            uuid4().hex, slot_key, started_at, finished_at, status, scan.source_mode,
            scan.page_count, scan.expected_page_count, scan.total_count, result.get("inserted", 0),
            result.get("updated", 0), result.get("hidden", 0), len(errors), json.dumps(errors, ensure_ascii=False),
        ),
    )


def apply_full_scan(scan: FullScanResult, slot_key: str, now: Optional[datetime] = None) -> dict[str, Any]:
    """Apply one normalized full scan atomically; only a validated scan may soft-hide."""
    scan = validate_full_scan(scan)
    started_at = _timestamp(now)
    finished_at = _timestamp(now)
    result: dict[str, Any] = {"status": "success" if scan.complete else "error", "inserted": 0, "updated": 0, "hidden": 0, "record_errors": []}
    initialize_needed = True
    with db.connect() as conn:
        if initialize_needed:
            initialize_schema(conn)
        cur = conn.cursor()
        for incoming in scan.records:
            if not incoming.get("eligible"):
                continue
            inserted, has_error = _upsert_record(cur, incoming, scan.source_mode, finished_at)
            if inserted:
                result["inserted"] += 1
            else:
                result["updated"] += 1
            if has_error:
                result["record_errors"].append(incoming.get("sync_errors") or [])
        if scan.complete:
            ids = [str(record["source_detail_id"]) for record in scan.records if record.get("eligible")]
            if ids:
                marks = ", ".join("?" for _ in ids)
                hide_sql = f"UPDATE spot_ledger_records SET is_active = 0, updated_at = ? WHERE record_source_type = '现货同步' AND is_active = 1 AND source_detail_id NOT IN ({marks})"
                cursor = _raw_execute(cur, hide_sql, (finished_at, *ids))
            else:
                cursor = _raw_execute(cur, "UPDATE spot_ledger_records SET is_active = 0, updated_at = ? WHERE record_source_type = '现货同步' AND is_active = 1", (finished_at,))
            result["hidden"] = max(0, int(cursor.rowcount or 0))
        _insert_run(cur, slot_key, started_at, finished_at, scan, result)
    return result


def get_active_records() -> list[dict[str, Any]]:
    with db.connect() as conn:
        rows = db._exec(conn.cursor(), "SELECT * FROM spot_ledger_records WHERE is_active = 1 ORDER BY \"U\" DESC, record_id").fetchall()
    return [record_to_public(dict(row)) for row in rows]


def get_sync_runs(limit: int = 20) -> list[dict[str, Any]]:
    with db.connect() as conn:
        rows = db._exec(
            conn.cursor(), "SELECT * FROM spot_ledger_sync_runs ORDER BY started_at DESC LIMIT ?", (max(1, min(limit, 100)),)
        ).fetchall()
    return [record_to_public(dict(row)) for row in rows]


def _slot_key(current: datetime, slot: day_time) -> str:
    return datetime.combine(current.date(), slot, SHANGHAI_TZ).isoformat(timespec="minutes")


def due_spot_ledger_slots(now: datetime, attempted_slots: Optional[set[str]] = None) -> list[str]:
    current = _now(now)
    attempted = attempted_slots or set()
    return [
        key for slot in SPOT_LEDGER_SYNC_TIMES
        if (key := _slot_key(current, slot)) not in attempted and datetime.fromisoformat(key) <= current
    ]


def scheduler_due_slots(now: datetime, attempted_slots: set[str], *, startup: bool) -> list[str]:
    current = _now(now)
    due = due_spot_ledger_slots(current, attempted_slots)
    if current.hour > SPOT_LEDGER_SYNC_TIMES[-1].hour:
        attempted_slots.update(due)
        return []
    if startup and due:
        attempted_slots.update(due[:-1])
        return due[-1:]
    return due


def _source_from_env() -> SalesContractSource:
    mode = (os.getenv("SPOT_LEDGER_SOURCE_MODE") or "profiled_http").strip().lower()
    if mode == "fixture":
        path = (os.getenv("SPOT_LEDGER_FIXTURE_PATH") or "").strip()
        if not path:
            raise SalesContractSourceError("fixture_missing", "SPOT_LEDGER_FIXTURE_PATH 未配置")
        return FixtureSalesContractSource(path)
    return ProfiledSalesContractSource.from_env()


def run_spot_ledger_sync_once(slot_key: str, source: Optional[SalesContractSource] = None) -> dict[str, Any]:
    try:
        scan = (source or _source_from_env()).fetch_full_scan()
        return apply_full_scan(scan, slot_key)
    except SalesContractSourceError as exc:
        now = _timestamp()
        initialize_schema_for_error = True
        with db.connect() as conn:
            if initialize_schema_for_error:
                initialize_schema(conn)
            _insert_run(
                conn.cursor(), slot_key, now, now,
                FullScanResult([], 0, None, 0, False, [exc.code], "profiled_http"),
                {"inserted": 0, "updated": 0, "hidden": 0, "record_errors": []},
            )
        raise


def _scheduler_loop(interval_seconds: int) -> None:
    attempted: set[str] = set()
    startup = True
    while True:
        current = _now()
        for slot_key in scheduler_due_slots(current, attempted, startup=startup):
            attempted.add(slot_key)
            try:
                run_spot_ledger_sync_once(slot_key)
            except Exception:
                # 错误已写入 sync_runs；调度线程继续等待下一 slot。
                pass
        startup = False
        cutoff = (current.date().toordinal() - 3)
        attempted = {slot for slot in attempted if datetime.fromisoformat(slot).date().toordinal() >= cutoff}
        time.sleep(max(1, interval_seconds))


def start_spot_ledger_sync_scheduler(interval_seconds: int = 30) -> bool:
    global _scheduler_started
    if (os.getenv("SPOT_LEDGER_AUTO_SYNC_ENABLED") or "").strip().lower() != "true":
        return False
    with _scheduler_lock:
        if _scheduler_started:
            return False
        _scheduler_started = True
        thread = threading.Thread(target=_scheduler_loop, args=(interval_seconds,), daemon=True, name="spot-ledger-sync")
        thread.start()
    return True


def _history_value(row: dict[str, Any], code: str) -> Any:
    if code in row:
        return row[code]
    return row.get(FIELD_BY_CODE_NAME.get(code, ""), "")


FIELD_BY_CODE_NAME = {item["code"]: item["name"] for item in FIELD_DEFINITIONS}


def _usable_history_value(value: Any) -> bool:
    if value is None or value == "":
        return False
    return str(value).strip() not in {"—", "——", "-", "--", "***"}


def _history_row_to_values(headers: list[Any], values: tuple[Any, ...]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, value in zip(headers, values):
        header_text = str(header or "").strip()
        code = header_text if header_text in FIELD_CODES else FIELD_NAME_TO_CODE.get(header_text)
        if code:
            row[code] = value
        elif header_text in {"长协对象", "long_contract_object"}:
            row["long_contract_object"] = value
        elif header_text == "销售合同商品明细 ID":
            row["source_detail_id"] = value
    return row


def _matches_history(row: dict[str, Any], candidate: dict[str, Any]) -> bool:
    contract = row.get("AD")
    product = row.get("H")
    price = row.get("Z")
    quantity = row.get("X") if _usable_history_value(row.get("X")) else row.get("L")
    if not all(_usable_history_value(value) for value in (contract, product, price, quantity)):
        return False
    if str(candidate.get("AD") or "").strip() != str(contract).strip():
        return False
    if str(candidate.get("H") or "").strip() != str(product).strip():
        return False
    try:
        if abs(float(candidate.get("Z")) - float(price)) > 0.000001:
            return False
        candidate_quantity = candidate.get("X") if candidate.get("X") is not None else candidate.get("L")
        return abs(float(candidate_quantity) - float(quantity)) <= 0.000001
    except (TypeError, ValueError):
        return False


def _split_long_contract(value: Any, explicit_object: Any) -> tuple[str, str]:
    explicit = str(explicit_object).strip() if _usable_history_value(explicit_object) else ""
    text = str(value).strip() if _usable_history_value(value) else ""
    if text in {"是", "否"}:
        return text, explicit
    if text.startswith("是"):
        object_value = explicit or re.sub(r"^[是：:、,，\s]+", "", text)
        return "是", object_value
    if text.startswith("否"):
        return "否", explicit
    return "", explicit


def migrate_history_workbook(path: str | Path, apply: bool = False) -> dict[str, Any]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    sheet = workbook["现货业务台账"] if "现货业务台账" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows, ()))
    history_rows = [_history_row_to_values(headers, values) for values in rows if any(_usable_history_value(value) for value in values)]
    summary: dict[str, Any] = {"matched": 0, "updated": 0, "ambiguous": 0, "unmatched": 0, "dry_run": not apply, "errors": []}
    with db.connect() as conn:
        initialize_schema(conn)
        cur = conn.cursor()
        candidates = [record_to_public(dict(row)) for row in db._exec(cur, "SELECT * FROM spot_ledger_records WHERE source_detail_id IS NOT NULL").fetchall()]
        updates: list[tuple[dict[str, Any], dict[str, Any]]] = []
        for history in history_rows:
            matches = [candidate for candidate in candidates if _matches_history(history, candidate)]
            if len(matches) != 1:
                summary["ambiguous" if len(matches) > 1 else "unmatched"] += 1
                continue
            summary["matched"] += 1
            candidate = matches[0]
            update: dict[str, Any] = {}
            for field in MANUAL_FIELDS:
                if field == "long_contract_object":
                    continue
                value = history.get(field)
                if _usable_history_value(value):
                    update[field] = value
            p_value, object_value = _split_long_contract(history.get("P"), history.get("long_contract_object"))
            if p_value:
                update["P"] = p_value
            if object_value:
                update["long_contract_object"] = object_value
            if not update:
                continue
            updates.append((candidate, update))
            if apply:
                projected = {**candidate, **update}
                missing = missing_required_fields(projected)
                assignments = []
                values: list[Any] = []
                for field, value in update.items():
                    column = f'"{field}"' if field in FIELD_CODES else field
                    assignments.append(f"{column} = ?")
                    values.append(value)
                assignments.extend(["missing_fields = ?", "supplement_status = ?", "updated_at = ?"])
                values.extend([json.dumps(missing, ensure_ascii=False), "待补录" if missing else "已完成", _timestamp()])
                values.append(candidate["record_id"])
                _raw_execute(cur, f"UPDATE spot_ledger_records SET {', '.join(assignments)} WHERE record_id = ?", tuple(values))
                summary["updated"] += 1
        if not apply:
            summary["candidate_updates"] = len(updates)
    workbook.close()
    return summary
