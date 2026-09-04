from __future__ import annotations

import base64
from contextlib import contextmanager
import json
import os
from datetime import datetime
from pathlib import Path
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "spot_ledger_sales_contract_fixture.json"


@pytest.fixture
def ledger_db(tmp_path, monkeypatch):
    from app import db

    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setattr(db, "DATA_DIR", tmp_path)
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "spot-ledger.db")
    db.init_db()
    return db


def load_fixture_scan():
    from app.spot_ledger_sync import FixtureSalesContractSource

    return FixtureSalesContractSource(FIXTURE_PATH).fetch_full_scan()


def test_fixture_scan_accepts_only_seven_groups_spot_and_effective_records():
    from app.spot_ledger import SHANGHAI_GROUPS

    scan = load_fixture_scan()
    assert set(record["E"] for record in scan.records) == set(SHANGHAI_GROUPS)
    assert all(record["eligible"] for record in scan.records)
    assert len({record["source_detail_id"] for record in scan.records}) == len(scan.records)
    assert sum(record["AD"] == "C-100" for record in scan.records) == 2
    assert {record["D"] for record in scan.records} >= {"B05", "B06", "B07", "B09"}


def test_full_scan_upsert_is_idempotent_and_preserves_manual_fields(ledger_db):
    from app.spot_ledger_sync import apply_full_scan

    scan = load_fixture_scan()
    first = apply_full_scan(scan, "2026-08-24T09:00+08:00")
    second = apply_full_scan(scan, "2026-08-24T10:00+08:00")
    assert first["inserted"] == len(scan.records)
    assert second["inserted"] == 0
    assert second["updated"] == len(scan.records)
    with ledger_db.connect() as conn:
        count = conn.execute("SELECT COUNT(*) AS c FROM spot_ledger_records WHERE record_source_type = '现货同步'").fetchone()["c"]
        detail = conn.execute("SELECT source_detail_id FROM spot_ledger_records WHERE \"AD\" = 'C-100' ORDER BY source_detail_id").fetchall()
    assert count == len(scan.records)
    assert len(detail) == 2


def test_sync_runs_expose_compact_metadata_without_raw_error_details(ledger_db):
    from app.spot_ledger_sync import apply_full_scan, get_sync_runs

    apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")
    with ledger_db.connect() as conn:
        stored = conn.execute(
            "SELECT error_summary FROM spot_ledger_sync_runs ORDER BY started_at DESC LIMIT 1"
        ).fetchone()

    runs = get_sync_runs()
    assert stored["error_summary"]
    assert runs[0]["error_count"] > 0
    assert "error_summary" not in runs[0]


def test_full_scan_commits_schema_setup_before_opening_data_transaction(ledger_db, monkeypatch):
    from app import spot_ledger_sync as sync

    original_connect = sync.db.connect
    original_initialize = sync.initialize_schema
    events = []

    @contextmanager
    def tracked_connect():
        events.append("connect_enter")
        with original_connect() as conn:
            yield conn
        events.append("connect_exit")

    def tracked_initialize(conn):
        events.append("initialize_schema")
        original_initialize(conn)

    monkeypatch.setattr(sync.db, "connect", tracked_connect)
    monkeypatch.setattr(sync, "initialize_schema", tracked_initialize)

    sync.apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")

    assert events[:4] == ["connect_enter", "initialize_schema", "connect_exit", "connect_enter"]


def test_full_scan_prefetches_existing_records_without_one_select_per_record(ledger_db, monkeypatch):
    from app import spot_ledger_sync as sync

    original_execute = sync._raw_execute
    single_record_selects = []

    def tracked_execute(cur, sql, params=()):
        if "SELECT * FROM spot_ledger_records WHERE source_detail_id = ?" in " ".join(sql.split()):
            single_record_selects.append(params)
        return original_execute(cur, sql, params)

    monkeypatch.setattr(sync, "_raw_execute", tracked_execute)
    sync.apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")

    assert single_record_selects == []


def test_sync_slot_is_not_repeated_after_process_restart(ledger_db):
    from app import spot_ledger_sync as sync

    slot_key = "2026-08-24T09:00+08:00"
    sync.apply_full_scan(load_fixture_scan(), slot_key)

    class Source:
        def __init__(self):
            self.calls = 0

        def fetch_full_scan(self):
            self.calls += 1
            return load_fixture_scan()

    source = Source()
    result = sync.run_spot_ledger_sync_once(slot_key, source)

    assert result == {"status": "skipped", "reason": "slot_already_recorded", "slot_key": slot_key}
    assert source.calls == 0


def test_incomplete_scan_does_not_hide_existing_record(ledger_db):
    from app.spot_ledger_sync import FullScanResult, apply_full_scan, get_active_records

    first = load_fixture_scan()
    apply_full_scan(first, "2026-08-24T09:00+08:00")
    broken = FullScanResult(
        records=first.records[:1], page_count=1, expected_page_count=2,
        total_count=1, complete=False, errors=["page_missing"], source_mode="fixture",
    )
    result = apply_full_scan(broken, "2026-08-24T10:00+08:00")
    assert result["hidden"] == 0
    assert len(get_active_records()) == len(first.records)


def test_complete_scan_soft_hides_missing_record_only_after_success(ledger_db):
    from app.spot_ledger_sync import FullScanResult, apply_full_scan, get_active_records

    first = load_fixture_scan()
    apply_full_scan(first, "2026-08-24T09:00+08:00")
    smaller = FullScanResult(
        records=first.records[:1], page_count=1, expected_page_count=1,
        total_count=1, complete=True, errors=[], source_mode="fixture",
    )
    result = apply_full_scan(smaller, "2026-08-24T10:00+08:00")
    assert result["hidden"] == len(first.records) - 1
    assert len(get_active_records()) == 1


def test_duplicate_detail_ids_make_scan_incomplete():
    from app.spot_ledger_sync import FullScanResult, validate_full_scan

    scan = load_fixture_scan()
    duplicate = FullScanResult(
        records=[scan.records[0], scan.records[0]], page_count=1,
        expected_page_count=1, total_count=2, complete=True, errors=[], source_mode="fixture",
    )
    result = validate_full_scan(duplicate)
    assert not result.complete
    assert any("duplicate_detail_id" in str(error) for error in result.errors)


def test_unattended_source_without_profile_is_explicitly_blocked(monkeypatch):
    from app.spot_ledger_sync import ProfiledSalesContractSource, SalesContractSourceError

    monkeypatch.delenv("SPOT_LEDGER_SOURCE_PROFILE", raising=False)
    monkeypatch.delenv("SPOT_LEDGER_SOURCE_USERNAME", raising=False)
    monkeypatch.delenv("SPOT_LEDGER_SOURCE_PASSWORD", raising=False)
    source = ProfiledSalesContractSource.from_env()
    with pytest.raises(SalesContractSourceError, match="auth_unavailable") as error:
        source.fetch_full_scan()
    assert error.value.stage == "auth_provider_missing"


def test_password_auth_uses_official_rsa_code_exchange_and_caches_bearer_token():
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric import padding, rsa

    from app.spot_ledger_sync import JianlongPasswordAuthProvider

    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    public_pem = private_key.public_key().public_bytes(
        serialization.Encoding.PEM,
        serialization.PublicFormat.SubjectPublicKeyInfo,
    ).decode("ascii")
    public_key_expression = " +\n".join(repr(line) for line in public_pem.splitlines(keepends=True))
    login_html = f"""
        <form>
          <input name="redirectUri" value="https://tds.ejianlong.com/">
          <input name="appId" value="confirmed-app-id">
          <input name="companyId" value="">
        </form>
        <script>var publicKey = {public_key_expression};</script>
    """

    class Response:
        status_code = 200

        def __init__(self, *, text="", payload=None):
            self.text = text
            self.payload = payload

        def json(self):
            return self.payload

    class Session:
        def __init__(self):
            self.calls = []
            self.headers = {}
            self.cookies = type("CookieJar", (), {"clear_calls": 0})()
            self.cookies.clear = lambda: setattr(
                self.cookies, "clear_calls", self.cookies.clear_calls + 1
            )

        def get(self, url, **kwargs):
            self.calls.append(("GET", url, kwargs))
            if kwargs.get("params", {}).get("code"):
                return Response(payload={"code": 200, "data": "bearer-token"})
            return Response(text=login_html)

        def post(self, url, **kwargs):
            self.calls.append(("POST", url, kwargs))
            return Response(
                payload={
                    "code": 200,
                    "data": "https://tds.ejianlong.com/?code=one-time-code#/sign/saleContract",
                }
            )

    session = Session()
    provider = JianlongPasswordAuthProvider("employee-id", "personal-password", http=session)

    assert provider() == {"Authorization": "Bearer bearer-token"}
    assert provider() == {"Authorization": "Bearer bearer-token"}
    assert session.headers["User-Agent"] == "ltm-spot-ledger/1.0"
    assert len(session.calls) == 3
    password_payload = session.calls[1][2]["json"]["password"]
    decrypted = private_key.decrypt(base64.b64decode(password_payload), padding.PKCS1v15())
    assert decrypted == b"personal-password"
    assert session.calls[1][2]["json"]["username"] == "employee-id"
    assert session.calls[2][1] == "https://tds-api.ejianlong.com/login"
    assert session.calls[2][2]["params"] == {"code": "one-time-code"}
    assert session.calls[2][2]["headers"] == {
        "Origin": "https://tds.ejianlong.com",
        "Referer": "https://tds.ejianlong.com/",
    }
    assert "personal-password" not in repr(session.calls)
    assert session.cookies.clear_calls == 1
    assert provider.refresh() == {"Authorization": "Bearer bearer-token"}
    assert session.cookies.clear_calls == 2


def test_password_auth_errors_never_expose_credentials_or_source_response():
    from app.spot_ledger_sync import JianlongPasswordAuthProvider, SalesContractSourceError

    class Response:
        status_code = 503
        text = "personal-password bearer-token"

        def json(self):
            return {"code": 503, "msg": self.text}

    class Session:
        def get(self, _url, **_kwargs):
            return Response()

    provider = JianlongPasswordAuthProvider("employee-id", "personal-password", http=Session())

    with pytest.raises(SalesContractSourceError) as error:
        provider()
    assert "personal-password" not in str(error.value)
    assert "bearer-token" not in str(error.value)
    assert error.value.stage == "login_page_http"
    assert error.value.http_status == 503


def test_profiled_source_from_env_uses_one_session_for_login_and_report(monkeypatch):
    from app import spot_ledger_sync as sync

    session = object()
    monkeypatch.setenv("SPOT_LEDGER_SOURCE_USERNAME", "employee-id")
    monkeypatch.setenv("SPOT_LEDGER_SOURCE_PASSWORD", "personal-password")
    monkeypatch.delenv("SPOT_LEDGER_SOURCE_PROFILE", raising=False)
    monkeypatch.setattr(sync.requests, "Session", lambda: session)

    source = sync.ProfiledSalesContractSource.from_env()

    assert source.http is session
    assert isinstance(source.auth_provider, sync.JianlongPasswordAuthProvider)


def test_source_factory_selects_official_json_mode(monkeypatch):
    from app import spot_ledger_sync as sync

    session = object()
    monkeypatch.setenv("SPOT_LEDGER_SOURCE_MODE", "official_json")
    monkeypatch.setenv("SPOT_LEDGER_SOURCE_USERNAME", "employee-id")
    monkeypatch.setenv("SPOT_LEDGER_SOURCE_PASSWORD", "personal-password")
    monkeypatch.setattr(sync.requests, "Session", lambda: session)

    source = sync._source_from_env()

    assert isinstance(source, sync.OfficialJsonSalesContractSource)
    assert source.http is session
    assert isinstance(source.auth_provider, sync.JianlongPasswordAuthProvider)
    assert source.enrich_sales_type_labels is True


def test_official_contract_scope_starts_from_locally_filtered_demands():
    from app import spot_ledger_sync as sync

    class Source(sync.OfficialJsonSalesContractSource):
        def __init__(self):
            super().__init__(auth_provider=lambda: {}, page_size=2)
            self.calls = []

        def _request_json(self, method, url, *, stage, **kwargs):
            self.calls.append((method, url, stage, kwargs))
            if "/tradeing/demand/list" in url:
                page = kwargs["params"]["pageNum"]
                rows = {
                    1: [
                        {
                            "demandId": "demand-in-scope",
                            "sourceType": "10",
                            "quantityAttribution": "Q1",
                        },
                        {
                            "demandId": "demand-other-group",
                            "sourceType": "10",
                            "quantityAttribution": "Q2",
                        },
                    ],
                    2: [
                        {
                            "demandId": "demand-futures",
                            "sourceType": "20",
                            "quantityAttribution": "Q1",
                        },
                        {
                            "demandId": "demand-without-group",
                            "sourceType": "10",
                            "quantityAttribution": "",
                        },
                    ],
                }[page]
                return {"code": 200, "data": {"rows": rows, "total": 4}}
            if "/relatedToDemand/" in url:
                assert "demand-in-scope" in url
                return {
                    "code": 200,
                    "data": [{"chainId": "chain-1"}, {"chainId": "chain-1"}],
                }
            if "/tradeing/chain/saleContractList" in url:
                assert kwargs["json"] == {"chainId": "chain-1", "tradersId": ""}
                return {
                    "code": 200,
                    "data": [
                        {"saleContractId": "contract-active", "status": "70"},
                        {"saleContractId": "contract-inactive", "status": "60"},
                        {"saleContractId": "contract-active", "status": "70"},
                    ],
                }
            raise AssertionError((method, url, stage, kwargs))

    dictionaries = {
        "source_type": {"10": "现货", "20": "期货"},
        "quantity_attribution": {"Q1": "大客户组", "Q2": "其他组"},
    }
    source = Source()

    scope = source._fetch_contract_scope(dictionaries)

    assert scope.page_count == 2
    assert set(scope.demands) == {
        "demand-in-scope",
        "demand-other-group",
        "demand-futures",
        "demand-without-group",
    }
    assert scope.in_scope_demand_ids == {"demand-in-scope"}
    assert scope.active_contracts == [{"saleContractId": "contract-active", "status": "70"}]
    assert scope.errors == []
    assert scope.diagnostics == {
        "source_demand_count": 4,
        "spot_demand_count": 3,
        "in_scope_demand_count": 1,
        "duplicate_demand_id_count": 0,
        "unclassified_demand_scope_count": 0,
        "related_chain_count": 1,
        "related_sale_contract_count": 3,
        "active_contract_count": 1,
    }
    assert not any("/tradeing/saleContract/saleContractList" in call[1] for call in source.calls)


@pytest.mark.parametrize(
    ("report_sales_business", "expected_missing_sales_business"),
    [("需求业务员A", False), ("", True)],
)
def test_official_json_source_fetches_all_pages_and_maps_confirmed_relations(
    report_sales_business, expected_missing_sales_business
):
    from app import spot_ledger_sync as sync

    class Response:
        status_code = 200
        url = "https://tds-api.ejianlong.com/"

        def __init__(self, payload):
            self.payload = payload

        def json(self):
            return self.payload

    class Http:
        def __init__(self):
            self.calls = []

        def post(self, url, **kwargs):
            self.calls.append(("POST", url, kwargs))
            if url == sync.CANDIDATE_SOURCE_URL:
                return Response(
                    {
                        "code": 200,
                        "result": {
                            "dataList": {
                                "TJJLYSHZ": {
                                    "list": [
                                        {
                                            "销售合同商品明细id": "sale-line-1",
                                            "业务类别": "贸易-港口现货-市场加价-B07",
                                            "需求业务员": report_sales_business,
                                        }
                                    ],
                                    "count": 1,
                                    "total": 1,
                                }
                            }
                        },
                    }
                )
            if "/tradeing/chain/saleContractList" in url:
                assert kwargs["json"] == {"chainId": "chain-1", "tradersId": ""}
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {"saleContractId": "contract-1", "status": "70"},
                            {"saleContractId": "contract-2", "status": "60"},
                        ],
                    }
                )
            if "/tdsSettle/queryJiesuan" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "rows": [
                                {
                                    "saleContractMxId": "sale-line-1",
                                    "countQuantity": 90,
                                }
                            ],
                            "total": 1,
                        },
                    }
                )
            raise AssertionError(url)

        def get(self, url, **kwargs):
            self.calls.append(("GET", url, kwargs))
            if "/tradeing/demand/list" in url:
                page = kwargs["params"]["pageNum"]
                rows = [
                    {
                        "demandId": f"demand-{page}",
                        "sourceType": "10",
                        "businessType": "B07" if page == 1 else "B05",
                        "businessTypeName": "贸易-港口现货-市场加价-B07" if page == 1 else "贸易-落地-B05",
                        "quantityAttribution": "Q1" if page == 1 else "Q2",
                        "profitAttribution": "P1" if page == 1 else "P2",
                    }
                ]
                return Response({"code": 200, "data": {"rows": rows, "total": 2}})
            if "/relatedToDemand/" in url:
                assert "demand-1" in url
                return Response({"code": 200, "data": [{"chainId": "chain-1"}]})
            if "/system/dict/data/type" in url:
                dictionaries = {
                    "quantity_attribution": [
                        {"dictValue": "Q1", "dictLabel": "大客户组"},
                        {"dictValue": "Q2", "dictLabel": "其他组"},
                    ],
                    "profit_attribution": [
                        {"dictValue": "P1", "dictLabel": "东北组"},
                        {"dictValue": "P2", "dictLabel": "其他组"},
                    ],
                    "source_type": [{"dictValue": "10", "dictLabel": "现货"}],
                    "price_mode": [{"dictValue": "20", "dictLabel": "固定价"}],
                }
                return Response({"code": 200, "data": dictionaries[kwargs["params"]["dictType"]]})
            if "/getRelevanceContract/" in url:
                contract_id = url.split("/getRelevanceContract/", 1)[1].split("?", 1)[0]
                return Response({"code": 200, "data": [{"purchaseContractId": f"purchase-{contract_id[-1]}"}]})
            if "/purchaseContract/" in url:
                suffix = url.split("/purchaseContract/", 1)[1].split("?", 1)[0]
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "purchaseContractId": suffix,
                            "tdsPurchaseContractMxVos": [
                                {
                                    "purchaseContractMxId": f"purchase-line-{suffix[-1]}",
                                    "countQuantity": 100,
                                }
                            ],
                        },
                    }
                )
            if "/saleContract/" in url:
                contract_id = url.split("/saleContract/", 1)[1].split("?", 1)[0]
                index = contract_id[-1]
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "saleContractId": contract_id,
                            "contractCode": f"XS-{index}",
                            "status": "70",
                            "signingDate": "2026-08-20",
                            "workCompName": "操作抬头A",
                            "workManName": "合同经办人A",
                            "createBy": "001_销售执行员",
                            "coustomName": "客户A",
                            "dischargePortName": "曹妃甸港",
                            "syncTradersId": f"traders-{index}",
                            "tdsSaleContractMxVos": [
                                {
                                    "saleContractMxId": f"sale-line-{index}",
                                    "goodsCode": f"GOODS-{index}",
                                    "goodsName": "铁矿石",
                                    "countQuantity": 100,
                                    "unitPrice": 760,
                                    "taxPrice": 858.8,
                                    "priceMode": "20",
                                    "upContractMxId": f"purchase-line-{index}",
                                }
                            ],
                        },
                    }
                )
            if "/chain/goods/matchResult/" in url:
                index = url.split("/matchResult/traders-", 1)[1].split("?", 1)[0]
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {
                                "demandId": f"demand-{index}",
                                "goodsCode": f"GOODS-{index}",
                                "matchPrice": 700,
                                "saleId": f"resource-{index}",
                            }
                        ],
                    }
                )
            if "/tradeing/demand?" in url:
                index = kwargs["params"]["demandId"][-1]
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "demandId": f"demand-{index}",
                            "sourceType": "10",
                            "businessType": "B07" if index == "1" else "B05",
                            "businessTypeName": "贸易-港口现货-市场加价-B07" if index == "1" else "贸易-落地-B05",
                            "quantityAttribution": "Q1" if index == "1" else "Q2",
                            "profitAttribution": "P1" if index == "1" else "P2",
                        },
                    }
                )
            if "/tradeing/sale?" in url:
                index = kwargs["params"]["saleId"][-1]
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "saleId": f"resource-{index}",
                            "sourceDate": "2026-08-18",
                            "chineseShipName": "海运一号",
                            "supplierName": "供应商A",
                            "workManName": "采购业务员",
                            "createBy": "采购执行员",
                        },
                    }
                )
            raise AssertionError(url)

    class Provider:
        def __call__(self):
            return {"Authorization": "Bearer bearer-token"}

    source = sync.OfficialJsonSalesContractSource(
        http=Http(),
        auth_provider=Provider(),
        page_size=1,
        enrich_sales_type_labels=True,
    )

    scan = source.fetch_full_scan()

    assert scan.complete is True
    assert scan.page_count == 2
    assert scan.total_count == 1
    assert scan.source_mode == "official_json"
    assert scan.diagnostics == {
        "source_demand_count": 2,
        "spot_demand_count": 2,
        "in_scope_demand_count": 1,
        "duplicate_demand_id_count": 0,
        "unclassified_demand_scope_count": 0,
        "related_chain_count": 1,
        "related_sale_contract_count": 2,
        "active_contract_count": 1,
        "source_detail_count": 1,
        "eligible_record_count": 1,
        "out_of_scope_record_count": 0,
        "ambiguous_resource_match_count": 0,
        "missing_resource_match_count": 0,
        "source_sales_type_label_count": 1,
        "source_sales_business_label_count": 0 if expected_missing_sales_business else 1,
    }
    assert len(scan.records) == 1
    record = scan.records[0]
    assert record["source_detail_id"] == "sale-line-1"
    assert record["eligible"] is True
    assert record["D"] == "贸易-港口现货-市场加价-B07"
    assert record["E"] == "大客户组"
    assert record["AP"] == "东北组"
    assert record["AQ"] == "是"
    assert record["L"] == record["X"] == 90
    assert record["M"] == 700
    assert record["Z"] == 858.8
    assert record["U"] == "2026-08-20"
    assert record["AD"] == "XS-1"
    assert record["K"] == "海运一号"
    assert record["AF"] == ("" if expected_missing_sales_business else "需求业务员A")
    assert any(
        error.get("type") == "missing_source_demand_salesperson"
        for error in record["sync_errors"]
    ) is expected_missing_sales_business
    assert record["source_closed_state"] == "已结案"
    assert not any(error.get("type") == "group_mismatch" for error in record["sync_errors"])
    demand_calls = [call for call in source.http.calls if "/tradeing/demand/list" in call[1]]
    assert [call[2]["params"]["pageNum"] for call in demand_calls] == [1, 2]
    assert not any("/tradeing/saleContract/saleContractList" in call[1] for call in source.http.calls)


def test_official_report_enrichment_rejects_conflicting_demand_salespeople(monkeypatch):
    from app.spot_ledger_sync import OfficialJsonSalesContractSource, SalesContractSourceError

    source = OfficialJsonSalesContractSource(auth_provider=lambda: {}, page_size=20)
    payload = {
        "code": 200,
        "result": {
            "dataList": {
                "TJJLYSHZ": {
                    "list": [
                        {
                            "销售合同商品明细id": "sale-line-1",
                            "业务类别": "贸易-港口现货-市场加价-B07",
                            "需求业务员": "需求业务员A",
                        },
                        {
                            "销售合同商品明细id": "sale-line-1",
                            "业务类别": "贸易-港口现货-市场加价-B07",
                            "需求业务员": "需求业务员B",
                        },
                    ],
                    "count": 2,
                    "total": 1,
                }
            }
        },
    }
    monkeypatch.setattr(source, "_request_json", lambda *_args, **_kwargs: payload)

    with pytest.raises(SalesContractSourceError) as error:
        source._fetch_report_enrichment()

    assert error.value.stage == "official_sales_type_report_duplicate"


def test_official_json_source_marks_duplicate_goods_match_as_ambiguous():
    from app import spot_ledger_sync as sync

    class Source(sync.OfficialJsonSalesContractSource):
        def __init__(self):
            super().__init__(auth_provider=lambda: {})

        def _fetch_contract_scope(self, _dictionaries):
            return sync.OfficialContractScope(
                active_contracts=[{"saleContractId": "contract-1", "status": "70"}],
                demands={
                    "demand-1": {
                        "demandId": "demand-1",
                        "sourceType": "10",
                        "businessType": "B07",
                        "quantityAttribution": "Q1",
                        "profitAttribution": "P1",
                    }
                },
                in_scope_demand_ids={"demand-1"},
                page_count=1,
                diagnostics={"active_contract_count": 1},
            )

        def _fetch_settlements(self):
            return {}

        def _fetch_dictionaries(self):
            return {
                "quantity_attribution": {"Q1": "大客户组"},
                "profit_attribution": {"P1": "东北组"},
                "source_type": {"10": "现货"},
                "price_mode": {"20": "固定价"},
            }

        def _purchase_lines(self, _contract_id):
            return {}

        def _get_data_dict(self, _url, *, stage, **_kwargs):
            assert stage == "official_contract_detail"
            return {
                "status": "70",
                "syncTradersId": "traders-1",
                "tdsSaleContractMxVos": [
                    {"saleContractMxId": "sale-line-1", "goodsCode": "GOODS-1"},
                    {"saleContractMxId": "sale-line-2", "goodsCode": "GOODS-1"},
                ],
            }

        def _get_data_list(self, _url, *, stage, **_kwargs):
            assert stage == "official_match_result"
            return [{"demandId": "demand-1", "goodsCode": "GOODS-1", "saleId": "resource-1"}]

    scan = Source().fetch_full_scan()

    assert scan.complete is False
    assert scan.records == []
    assert scan.diagnostics["ambiguous_resource_match_count"] == 2
    assert scan.diagnostics["missing_resource_match_count"] == 0
    assert [error["type"] for error in scan.errors] == [
        "ambiguous_resource_match",
        "ambiguous_resource_match",
    ]


def test_official_contract_bundle_isolates_purchase_detail_parse_error():
    from app import spot_ledger_sync as sync

    class Source(sync.OfficialJsonSalesContractSource):
        def __init__(self):
            super().__init__(auth_provider=lambda: {})

        def _get_data_dict(self, _url, *, stage, **_kwargs):
            assert stage == "official_contract_detail"
            return {
                "syncTradersId": "traders-1",
                "tdsSaleContractMxVos": [
                    {"saleContractMxId": "sale-line-1", "goodsCode": "GOODS-1"}
                ],
            }

        def _purchase_lines(self, _contract_id):
            raise sync.SalesContractSourceError(
                "parse_error",
                "sensitive purchase response",
                stage="official_purchase_detail_response",
            )

        def _get_data_list(self, _url, *, stage, **_kwargs):
            assert stage == "official_match_result"
            return [{"demandId": "demand-1", "goodsCode": "GOODS-1"}]

    bundle = Source()._fetch_contract_bundle(
        {"saleContractId": "contract-sensitive", "status": "70"}
    )

    assert len(bundle.lines) == 1
    assert len(bundle.match_rows) == 1
    assert bundle.purchase_lines == {}
    assert bundle.errors == [{"type": "official_purchase_detail_response"}]
    assert "sensitive purchase response" not in repr(bundle)


def test_official_source_dry_run_summary_contains_only_aggregate_metadata():
    from app import spot_ledger_sync as sync

    scan = sync.FullScanResult(
        records=[
            {
                "source_detail_id": "sensitive-detail-id",
                "D": "现货-市场加价",
                "E": "大客户组",
                "AP": "东北组",
                "AB": "敏感客户名称",
                "sync_errors": [
                    {
                        "type": "conversion_mapping",
                        "field": "F",
                        "message": "包含敏感操作抬头",
                    }
                ],
            }
        ],
        page_count=1,
        expected_page_count=1,
        total_count=1,
        complete=False,
        errors=[{"type": "ambiguous_resource_match", "detail_id": "sensitive-detail-id"}],
        source_mode="official_json",
        diagnostics={
            "source_demand_count": 3,
            "in_scope_demand_count": 1,
            "related_chain_count": 1,
            "active_contract_count": 1,
            "eligible_record_count": 1,
        },
    )

    result = sync.summarize_official_source_scan(scan)
    serialized = json.dumps(result, ensure_ascii=False)

    assert result["ok"] is False
    assert result["counts"]["source_demand_count"] == 3
    assert result["counts"]["in_scope_demand_count"] == 1
    assert result["counts"]["related_chain_count"] == 1
    assert result["counts"]["eligible_record_count"] == 1
    assert result["field_coverage"]["AB"] == {"filled_count": 1, "total_count": 1}
    assert result["scan_error_types"] == {"ambiguous_resource_match": 1}
    assert result["record_error_types"] == {"conversion_mapping": 1}
    assert "sensitive-detail-id" not in serialized
    assert "敏感客户名称" not in serialized
    assert "包含敏感操作抬头" not in serialized


def test_profiled_report_dry_run_returns_aggregate_only():
    from app import spot_ledger_sync as sync

    class Source:
        def fetch_full_scan(self):
            return sync.FullScanResult(
                records=[
                    {
                        "source_detail_id": "sensitive-detail-id",
                        "E": "大客户组",
                        "AB": "敏感客户名称",
                        "sync_errors": [],
                    }
                ],
                page_count=1,
                expected_page_count=1,
                total_count=1,
                complete=True,
                source_mode="profiled_http",
            )

    result = sync.run_profiled_source_dry_run(source=Source())
    serialized = json.dumps(result, ensure_ascii=False)

    assert result["ok"] is True
    assert result["source_mode"] == "profiled_http"
    assert result["field_coverage"]["E"] == {"filled_count": 1, "total_count": 1}
    assert "sensitive-detail-id" not in serialized
    assert "敏感客户名称" not in serialized


def test_official_contract_scope_dry_run_returns_counts_without_identifiers():
    from app import spot_ledger_sync as sync

    class Source:
        def _fetch_dictionaries(self):
            return {}

        def _fetch_contract_scope(self, _dictionaries):
            return sync.OfficialContractScope(
                active_contracts=[{"saleContractId": "sensitive-contract-id"}],
                demands={"sensitive-demand-id": {}},
                in_scope_demand_ids={"sensitive-demand-id"},
                page_count=2,
                errors=[
                    {
                        "type": "unclassified_demand_scope",
                        "demand_id": "sensitive-demand-id",
                    }
                ],
                diagnostics={
                    "source_demand_count": 10,
                    "in_scope_demand_count": 1,
                    "active_contract_count": 1,
                },
            )

    result = sync.run_official_contract_scope_dry_run(source=Source())
    serialized = json.dumps(result, ensure_ascii=False)

    assert result == {
        "ok": False,
        "source_mode": "official_json",
        "page_count": 2,
        "counts": {
            "source_demand_count": 10,
            "in_scope_demand_count": 1,
            "active_contract_count": 1,
        },
        "error_types": {"unclassified_demand_scope": 1},
    }
    assert "sensitive" not in serialized


def test_official_scope_probe_confirms_demand_and_settlement_filters_without_values():
    from app import spot_ledger_sync as sync

    class Source(sync.OfficialJsonSalesContractSource):
        def __init__(self):
            super().__init__(auth_provider=lambda: {})

        def _fetch_dictionaries(self):
            return {
                "quantity_attribution": {
                    f"Q{index}": group for index, group in enumerate(sync.SHANGHAI_GROUPS, start=1)
                },
                "profit_attribution": {},
                "source_type": {"10": "现货"},
                "price_mode": {},
            }

        def _request_json(self, method, url, *, stage, **kwargs):
            if "/tradeing/demand/list" in url:
                if "quantityAttribution" not in kwargs["params"]:
                    return {
                        "code": 200,
                        "data": {
                            "rows": [
                                {
                                    "demandId": f"demand-sensitive-{index}",
                                    "sourceType": "10",
                                    "quantityAttribution": f"Q{index}",
                                }
                                for index in range(1, 8)
                            ],
                            "total": 7,
                        },
                    }
                group_code = kwargs["params"]["quantityAttribution"]
                index = int(group_code[1:])
                return {
                    "code": 200,
                    "data": {
                        "rows": [
                            {
                                "demandId": f"demand-sensitive-{index}",
                                "sourceType": "10",
                                "quantityAttribution": group_code,
                            }
                        ],
                        "total": index,
                    },
                }
            if "/relatedToDemand/" in url:
                return {"code": 200, "data": [{"chainId": "chain-sensitive"}]}
            if "/tradeing/chain/saleContractList" in url:
                assert kwargs["json"] == {"chainId": "chain-sensitive", "tradersId": ""}
                return {
                    "code": 200,
                    "data": [
                        {"saleContractId": "sale-sensitive-1", "status": "70"},
                        {"saleContractId": "sale-sensitive-2", "status": "60"},
                    ],
                }
            if "/tdsSettle/queryJiesuan" in url:
                body = kwargs["json"]
                row = {
                    "salesContractNo": "contract-sensitive",
                    "saleContractId": "sale-sensitive-1",
                    "saleContractMxId": "line-sensitive-1",
                }
                filter_names = [name for name in row if name in body]
                if filter_names:
                    assert body[filter_names[0]] == row[filter_names[0]]
                    return {"code": 200, "data": {"rows": [row], "total": 1}}
                return {"code": 200, "data": {"rows": [row], "total": 100}}
            raise AssertionError((method, url, stage, kwargs))

    result = sync.probe_official_scope_filters(source=Source())
    serialized = json.dumps(result, ensure_ascii=False)

    assert result["ok"] is True
    assert result["demand_filter"]["sampled_group_count"] == 7
    assert result["demand_filter"]["sample_match_count"] == 7
    assert result["demand_filter"]["group_counts"]["大客户组"] == 1
    assert result["local_scope_scan"] == {
        "source_total_count": 7,
        "page_count": 1,
        "scanned_row_count": 7,
        "duplicate_demand_id_count": 0,
        "spot_demand_count": 7,
        "seven_group_spot_demand_count": 7,
        "group_counts": {group: 1 for group in sync.SHANGHAI_GROUPS},
    }
    assert result["related_chain_sample_attempt_count"] == 1
    assert result["related_chain_count"] == 1
    assert result["related_sale_contract_count"] == 2
    assert result["related_active_sale_contract_count"] == 1
    assert result["settlement_filter_baseline_total"] == 100
    assert all(item["effective"] for item in result["settlement_filters"].values())
    assert "sensitive" not in serialized


def test_official_scope_probe_paginates_local_demand_rows_and_deduplicates_ids():
    from app import spot_ledger_sync as sync

    class Source(sync.OfficialJsonSalesContractSource):
        def __init__(self):
            super().__init__(auth_provider=lambda: {})

        def _fetch_dictionaries(self):
            return {
                "quantity_attribution": {
                    f"Q{index}": group for index, group in enumerate(sync.SHANGHAI_GROUPS, start=1)
                },
                "profit_attribution": {},
                "source_type": {"10": "现货"},
                "price_mode": {},
            }

        def _request_json(self, method, url, *, stage, **kwargs):
            if "/tradeing/demand/list" in url:
                params = kwargs["params"]
                if "quantityAttribution" in params:
                    return {
                        "code": 200,
                        "data": {
                            "rows": [
                                {
                                    "demandId": f"filtered-{params['quantityAttribution']}",
                                    "sourceType": "10",
                                    "quantityAttribution": params["quantityAttribution"],
                                }
                            ],
                            "total": 1,
                        },
                    }
                if params["pageNum"] == 1:
                    rows = [
                        {
                            "demandId": f"demand-{index}",
                            "sourceType": "10",
                            "quantityAttribution": "Q1",
                        }
                        for index in range(500)
                    ]
                else:
                    rows = [
                        {
                            "demandId": "demand-0",
                            "sourceType": "10",
                            "quantityAttribution": "Q1",
                        }
                    ]
                return {"code": 200, "data": {"rows": rows, "total": 501}}
            if "/relatedToDemand/" in url:
                if url.endswith("/demand-0?sheetCode=G01004"):
                    return {"code": 200, "data": []}
                return {"code": 200, "data": [{"chainId": "chain-sensitive"}]}
            if "/tradeing/chain/saleContractList" in url:
                return {
                    "code": 200,
                    "data": [{"saleContractId": "sale-sensitive", "status": "70"}],
                }
            if "/tdsSettle/queryJiesuan" in url:
                row = {
                    "salesContractNo": "contract-sensitive",
                    "saleContractId": "sale-sensitive",
                    "saleContractMxId": "line-sensitive",
                }
                total = 1 if len(kwargs["json"]) > 1 else 100
                return {"code": 200, "data": {"rows": [row], "total": total}}
            raise AssertionError((method, url, stage, kwargs))

    result = sync.probe_official_scope_filters(source=Source())

    assert result["local_scope_scan"]["page_count"] == 2
    assert result["local_scope_scan"]["scanned_row_count"] == 501
    assert result["local_scope_scan"]["duplicate_demand_id_count"] == 1
    assert result["local_scope_scan"]["seven_group_spot_demand_count"] == 501
    assert result["related_chain_sample_attempt_count"] == 2


def test_official_json_probe_uses_confirmed_post_and_returns_schema_only():
    from app import spot_ledger_sync as sync

    class Response:
        status_code = 200

        def __init__(self, payload):
            self.payload = payload

        def json(self):
            return self.payload

    class Http:
        def __init__(self):
            self.calls = []

        def post(self, url, **kwargs):
            self.calls.append((url, kwargs))
            if "/tradeing/chain/saleList" in url:
                return Response({"code": 200, "data": []})
            if "/tdsSettle/queryJiesuan" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "rows": [
                                {
                                    "settleObjectDetailId": "must-not-be-returned-settlement-id",
                                    "saleContractId": "must-not-be-returned-id",
                                    "saleContractMxId": "must-not-be-returned",
                                    "countQuantity": 99,
                                }
                            ],
                            "total": 1,
                        },
                    }
                )
            return Response(
                {
                    "code": 200,
                    "data": {
                        "rows": [
                            {
                                "saleContractId": "must-not-be-returned-id",
                                "contractNo": "must-not-be-returned",
                                "goods": [{"saleContractMxId": "must-not-be-returned"}],
                            },
                            {"saleContractId": "must-not-be-returned-id-2"},
                        ],
                        "total": 1,
                    },
                }
            )

        def get(self, url, **kwargs):
            self.calls.append((url, kwargs))
            if "/system/dict/data/type" in url:
                dictionary = kwargs["params"]["dictType"]
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {
                                "dictLabel": group,
                                "dictValue": f"must-not-be-returned-{dictionary}-{index}",
                            }
                            for index, group in enumerate(sync.SHANGHAI_GROUPS)
                        ],
                    }
                )
            if "/tradeing/sale/list?" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "rows": [
                                {
                                    "saleId": "must-not-be-returned-resource-id",
                                    "quantityAttribution": "must-not-be-returned-quantity-group",
                                    "profitAttribution": "must-not-be-returned-profit-group",
                                }
                            ],
                            "total": 1,
                        },
                    }
                )
            if "/tradeing/demand?" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "demandId": "must-not-be-returned-demand-id",
                            "quantityAttribution": "must-not-be-returned-quantity-group",
                            "profitAttribution": "must-not-be-returned-profit-group",
                        },
                    }
                )
            if "/tradeing/goods/list/" in url:
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {
                                "goodsDetailId": "must-not-be-returned-business-detail-id",
                                "chainGoodId": "must-not-be-returned-chain-good-id",
                                "goodsCode": "must-not-be-returned-goods-code",
                                "spec": "must-not-be-returned-specification",
                            }
                        ],
                    }
                )
            if "/tradeing/chain/getById/" in url:
                chain_id = "must-not-be-returned-chain-id-2" if "chain-id-2" in url else "must-not-be-returned-chain-id"
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "chainId": chain_id,
                            "profitAttribution": "must-not-be-returned-profit-group",
                            "quantityAttribution": "must-not-be-returned-quantity-group",
                        },
                    }
                )
            if "/tradeing/saleContract/must-not-be-returned-id-2" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "chainId": "must-not-be-returned-chain-id-2",
                            "saleContractId": "must-not-be-returned-id-2",
                            "syncTradersId": "must-not-be-returned-traders-id",
                            "tdsSaleContractMxVos": [],
                        },
                    }
                )
            if "/chain/goods/matchResult/" in url:
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {
                                "saleId": "must-not-be-returned-resource-id",
                                "saleNo": "must-not-be-returned-resource-no",
                                "sourceDate": "must-not-be-returned-resource-date",
                                "demandId": "must-not-be-returned-demand-id",
                                "goodsCode": "must-not-be-returned-goods-code",
                                "specs": "must-not-be-returned-specification",
                            }
                        ],
                    }
                )
            if "/tradeing/sale?" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "saleId": "must-not-be-returned-resource-id",
                            "sourceDate": "must-not-be-returned-resource-date",
                            "supplierName": "must-not-be-returned-resource-supplier",
                            "tdsGoodsList": [{"price": 66}],
                        },
                    }
                )
            if "/tradeing/purchaseContract/" in url:
                return Response(
                    {
                        "code": 200,
                        "data": {
                            "purchaseContractId": "must-not-be-returned-purchase-id",
                            "supplierName": "must-not-be-returned-supplier",
                            "tdsPurchaseContractMxVos": [
                                {
                                    "purchaseContractMxId": "must-not-be-returned-purchase-line-id",
                                    "businessDetailId": "must-not-be-returned-purchase-business-detail-id",
                                    "chainGoodId": "must-not-be-returned-chain-good-id",
                                    "relevanceId": "must-not-be-returned-relevance-id",
                                    "price": 88,
                                }
                            ],
                        },
                    }
                )
            if "/getRelevanceContract/" in url:
                return Response(
                    {
                        "code": 200,
                        "data": [
                            {
                                "purchaseContractId": "must-not-be-returned-purchase-id",
                                "purchaseContractMxId": "must-not-be-returned-purchase-line-id",
                                "supplierName": "must-not-be-returned-supplier",
                            }
                        ],
                    }
                )
            return Response(
                {
                    "code": 200,
                    "data": {
                        "chainId": "must-not-be-returned-chain-id",
                        "saleContractId": "must-not-be-returned-id",
                        "syncTradersId": "must-not-be-returned-traders-id",
                        "tdsSaleContractMxVos": [
                            {
                                "saleContractMxId": "must-not-be-returned",
                                "businessDetailId": "must-not-be-returned-business-detail-id",
                                "chainGoodId": "must-not-be-returned-chain-good-id",
                                "goodsCode": "must-not-be-returned-goods-code",
                                "relevanceId": "must-not-be-returned-relevance-id",
                                "specification": "must-not-be-returned-specification",
                                "upContractMxId": "must-not-be-returned-purchase-line-id",
                                "contractQuantity": 100,
                            }
                        ],
                    },
                }
            )

    class Provider:
        def __call__(self):
            return {"Authorization": "Bearer bearer-token"}

    source = sync.ProfiledSalesContractSource(
        sync.build_candidate_source_profile("2026-08-25"),
        http=Http(),
        auth_provider=Provider(),
    )

    result = sync.probe_official_sales_contract_api(source=source)

    assert result == {
        "ok": True,
        "source_mode": "official_json",
        "http_status": 200,
        "response_code": "200",
        "detail_response_code": "200",
        "relevance_response_code": "200",
        "purchase_response_code": "200",
        "chain_response_code": "200",
        "resource_list_response_code": "200",
        "match_response_code": "200",
        "resource_detail_response_code": "200",
        "demand_detail_response_code": "200",
        "demand_goods_response_code": "200",
        "settlement_response_code": "200",
        "resource_catalog_response_code": "200",
        "quantity_group_dictionary_response_code": "200",
        "profit_group_dictionary_response_code": "200",
        "sampled_contract_count": 1,
        "active_contract_total": 1,
        "settlement_row_total": 1,
        "schema_paths": [
            "code",
            "data",
            "data.rows",
            "data.rows[]",
            "data.rows[].contractNo",
            "data.rows[].goods",
            "data.rows[].goods[]",
            "data.rows[].goods[].saleContractMxId",
            "data.rows[].saleContractId",
            "data.total",
        ],
        "detail_schema_paths": [
            "code",
            "data",
            "data.chainId",
            "data.saleContractId",
            "data.syncTradersId",
            "data.tdsSaleContractMxVos",
            "data.tdsSaleContractMxVos[]",
            "data.tdsSaleContractMxVos[].businessDetailId",
            "data.tdsSaleContractMxVos[].chainGoodId",
            "data.tdsSaleContractMxVos[].contractQuantity",
            "data.tdsSaleContractMxVos[].goodsCode",
            "data.tdsSaleContractMxVos[].relevanceId",
            "data.tdsSaleContractMxVos[].saleContractMxId",
            "data.tdsSaleContractMxVos[].specification",
            "data.tdsSaleContractMxVos[].upContractMxId",
        ],
        "relevance_schema_paths": [
            "code",
            "data",
            "data[]",
            "data[].purchaseContractId",
            "data[].purchaseContractMxId",
            "data[].supplierName",
        ],
        "purchase_schema_paths": [
            "code",
            "data",
            "data.purchaseContractId",
            "data.supplierName",
            "data.tdsPurchaseContractMxVos",
            "data.tdsPurchaseContractMxVos[]",
            "data.tdsPurchaseContractMxVos[].businessDetailId",
            "data.tdsPurchaseContractMxVos[].chainGoodId",
            "data.tdsPurchaseContractMxVos[].price",
            "data.tdsPurchaseContractMxVos[].purchaseContractMxId",
            "data.tdsPurchaseContractMxVos[].relevanceId",
        ],
        "chain_schema_paths": [
            "code",
            "data",
            "data.chainId",
            "data.profitAttribution",
            "data.quantityAttribution",
        ],
        "resource_list_schema_paths": [
            "code",
            "data",
            "data[]",
        ],
        "match_schema_paths": [
            "code",
            "data",
            "data[]",
            "data[].demandId",
            "data[].goodsCode",
            "data[].saleId",
            "data[].saleNo",
            "data[].sourceDate",
            "data[].specs",
        ],
        "resource_detail_schema_paths": [
            "code",
            "data",
            "data.saleId",
            "data.sourceDate",
            "data.supplierName",
            "data.tdsGoodsList",
            "data.tdsGoodsList[]",
            "data.tdsGoodsList[].price",
        ],
        "demand_detail_schema_paths": [
            "code",
            "data",
            "data.demandId",
            "data.profitAttribution",
            "data.quantityAttribution",
        ],
        "demand_goods_schema_paths": [
            "code",
            "data",
            "data[]",
            "data[].chainGoodId",
            "data[].goodsCode",
            "data[].goodsDetailId",
            "data[].spec",
        ],
        "settlement_schema_paths": [
            "code",
            "data",
            "data.rows",
            "data.rows[]",
            "data.rows[].countQuantity",
            "data.rows[].saleContractId",
            "data.rows[].saleContractMxId",
            "data.rows[].settleObjectDetailId",
            "data.total",
        ],
        "resource_catalog_schema_paths": [
            "code",
            "data",
            "data.rows",
            "data.rows[]",
            "data.rows[].profitAttribution",
            "data.rows[].quantityAttribution",
            "data.rows[].saleId",
            "data.total",
        ],
        "quantity_group_dictionary_schema_paths": [
            "code",
            "data",
            "data[]",
            "data[].dictLabel",
            "data[].dictValue",
        ],
        "profit_group_dictionary_schema_paths": [
            "code",
            "data",
            "data[]",
            "data[].dictLabel",
            "data[].dictValue",
        ],
        "group_dictionary_coverage": {"quantity": 7, "profit": 7},
        "sampled_demand_group_coverage": {"quantity": 1, "profit": 1},
        "linkage_counts": {
            "sale_lines": 1,
            "purchase_lines": 1,
            "settlement_rows": 1,
            "match_rows": 1,
            "settlement_to_sale_detail": 1,
            "match_demand_to_sale_detail": 0,
            "match_demand_to_sale_business_detail": 0,
            "match_demand_to_sale_chain_good": 0,
            "match_demand_to_sale_relevance": 0,
            "match_goods_to_sale_goods_code": 1,
            "match_goods_and_specs_to_sale_line": 1,
            "match_goods_and_specs_to_demand_goods": 1,
            "demand_goods_to_sale_goods_code": 1,
            "demand_goods_and_spec_to_sale_line": 1,
            "demand_goods_detail_to_sale_business_detail": 1,
            "demand_goods_chain_good_to_sale_chain_good": 1,
            "purchase_to_sale_up_contract": 1,
            "purchase_business_detail_to_sale_business_detail": 0,
            "purchase_chain_good_to_sale_chain_good": 1,
            "purchase_relevance_to_sale_relevance": 1,
        },
    }
    assert source.http.calls == [
        (
            "https://tds-api.ejianlong.com/tradeing/saleContract/saleContractList",
            {
                "params": {"sheetCode": "G01009", "pageNum": 1, "pageSize": 10},
                "json": {"status": "70", "isQryAll": "N"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/saleContract/must-not-be-returned-id?sheetCode=G01009",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/saleContract/getRelevanceContract/must-not-be-returned-id?sheetCode=G01009",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/purchaseContract/must-not-be-returned-purchase-id?sheetCode=G01008",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/chain/getById/must-not-be-returned-chain-id?sheetCode=G01004",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/chain/saleList?sheetCode=G01004",
            {
                "json": {"chainId": "must-not-be-returned-chain-id", "tradersId": ""},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/chain/goods/matchResult/must-not-be-returned-traders-id?sheetCode=G01004",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/sale?sheetCode=G01003",
            {
                "params": {"saleId": "must-not-be-returned-resource-id"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/demand?sheetCode=G01002",
            {
                "params": {"demandId": "must-not-be-returned-demand-id"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/goods/list/must-not-be-returned-demand-id?sheetCode=G01002",
            {
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tdsSettle/queryJiesuan?sheetCode=G01112",
            {
                "params": {"pageNum": 1, "pageSize": 10},
                "json": {"status": "70"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/tradeing/sale/list?sheetCode=G01003",
            {
                "params": {
                    "saleNo": "",
                    "workCompId": "",
                    "coustomName": "",
                    "workDeptList": "",
                    "workMan": "",
                    "status": "70",
                    "pageNum": 1,
                    "pageSize": 10,
                    "sheetCode": "G01003",
                },
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/system/dict/data/type",
            {
                "params": {"dictType": "quantity_attribution"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
        (
            "https://tds-api.ejianlong.com/system/dict/data/type",
            {
                "params": {"dictType": "profit_attribution"},
                "headers": {
                    "Authorization": "Bearer bearer-token",
                    "Origin": "https://tds.ejianlong.com",
                    "Referer": "https://tds.ejianlong.com/",
                },
                "timeout": 30,
            },
        ),
    ]
    assert "must-not-be-returned" not in repr(result)


@pytest.mark.parametrize(
    ("expired_status", "expired_url"),
    [
        (401, "https://tds-report.ejianlong.com/jmreport/show"),
        (200, "https://server-auth.ejianlong.com/login?appId=confirmed-app-id"),
    ],
)
def test_profiled_source_reauthenticates_once_after_bearer_expiry(expired_status, expired_url):
    from app.spot_ledger_sync import ProfiledSalesContractSource, build_candidate_source_profile

    class Response:
        def __init__(self, status_code, payload=None, *, url="https://tds-report.ejianlong.com/jmreport/show"):
            self.status_code = status_code
            self.payload = payload or {}
            self.url = url

        def json(self):
            return self.payload

    class Http:
        def __init__(self):
            self.calls = []

        def post(self, _url, **kwargs):
            self.calls.append(kwargs["headers"])
            if len(self.calls) == 1:
                return Response(expired_status, url=expired_url)
            row = {
                "销售合同商品明细id": "DETAIL-1",
                "期现货": "现货",
                "合同状态": "生效",
                "量归属组": "大客户组",
                "业务毛利归属组": "大客户组",
                "业务类别": "B07",
                "公司": "操作抬头",
                "资源日期": "2026-08-20",
                "物资名称": "铁矿石",
                "合同卸货港": "曹妃甸港",
                "定价模式": "固定价",
                "中文船名": "",
                "合同数量": 100,
                "结算数量": None,
                "结案状态": "未结案",
                "资源单单价": 700,
                "资源方": "供应商",
                "资源业务员": "采购员",
                "初始资源单创建人": "执行员",
                "签订日期": "2026-08-21",
                "合同单价": 750,
                "需求方": "客户",
                "销售合同号": "XS-1",
                "需求业务员": "销售员",
                "合同创建人": "销售执行员",
            }
            return Response(
                200,
                {
                    "code": 200,
                    "result": {
                        "dataList": {
                            "TJJLYSHZ": {"list": [row], "count": 1, "total": 1}
                        }
                    },
                },
            )

    class Provider:
        def __init__(self):
            self.refresh_count = 0

        def __call__(self):
            return {"Authorization": "Bearer old-token"}

        def refresh(self):
            self.refresh_count += 1
            return {"Authorization": "Bearer new-token"}

    http = Http()
    provider = Provider()
    source = ProfiledSalesContractSource(
        build_candidate_source_profile("2026-08-25", page_size=1),
        http=http,
        auth_provider=provider,
    )

    scan = source.fetch_full_scan()

    assert scan.complete is True
    assert provider.refresh_count == 1
    assert http.calls == [
        {"Authorization": "Bearer old-token"},
        {"Authorization": "Bearer new-token"},
    ]


def test_profiled_source_reports_http_status_without_response_body():
    from app.spot_ledger_sync import (
        ProfiledSalesContractSource,
        SalesContractSourceError,
        build_candidate_source_profile,
    )

    class Response:
        status_code = 502
        url = "https://tds-report.ejianlong.com/jmreport/show"
        text = "sensitive upstream response"

    class Http:
        def post(self, _url, **_kwargs):
            return Response()

    source = ProfiledSalesContractSource(
        build_candidate_source_profile("2026-08-25", page_size=1),
        http=Http(),
        auth_provider=lambda: {"Authorization": "Bearer sensitive-token"},
    )

    with pytest.raises(SalesContractSourceError) as error:
        source.fetch_full_scan()

    assert error.value.stage == "report_http"
    assert error.value.http_status == 502
    assert "sensitive" not in str(error.value)


def test_profiled_source_classifies_connection_error_without_message():
    import requests

    from app.spot_ledger_sync import (
        ProfiledSalesContractSource,
        SalesContractSourceError,
        build_candidate_source_profile,
    )

    class Http:
        def post(self, _url, **_kwargs):
            raise requests.ConnectionError("sensitive upstream address")

    source = ProfiledSalesContractSource(
        build_candidate_source_profile("2026-08-25", page_size=1),
        http=Http(),
        auth_provider=lambda: {"Authorization": "Bearer sensitive-token"},
    )

    with pytest.raises(SalesContractSourceError) as error:
        source.fetch_full_scan()

    assert error.value.stage == "report_request_connection"
    assert "sensitive" not in str(error.value)


def test_scheduler_startup_uses_only_latest_due_hour():
    from app.spot_ledger_sync import scheduler_due_slots

    now = datetime.fromisoformat("2026-08-24T18:15:42+08:00")
    attempted = set()
    latest = scheduler_due_slots(now, attempted, startup=True)
    assert latest == ["2026-08-24T18:00+08:00"]
    assert len(attempted) == 9
    attempted.update(latest)
    assert scheduler_due_slots(now, attempted, startup=False) == []
    assert len(scheduler_due_slots(now, set(), startup=False)) == 10


def test_scheduler_startup_after_sync_window_does_not_run_catch_up():
    from app.spot_ledger_sync import scheduler_due_slots

    attempted = set()
    assert scheduler_due_slots(
        datetime.fromisoformat("2026-08-24T19:00:00+08:00"),
        attempted,
        startup=True,
    ) == []
    assert len(attempted) == 10


def test_confirmed_candidate_request_body_matches_observed_contract():
    from app.spot_ledger_sync import build_candidate_request_body

    body = build_candidate_request_body("2026-08-24", page_no=1, page_size=20)
    assert body == {
        "id": "1055351755192311808",
        "apiUrl": "",
        "params": json.dumps(
            {
                "pageNo": 1,
                "periodDate": "2026-08-24",
                "releaseDate": "2026-08-24",
                "TJJLYSHZ__期现货": "现货",
                "TJJLYSHZ__合同状态": "生效",
                "TJJLYSHZ__业务毛利归属组": "大客户组,东北组,山东组,黄骅组,天津组,唐山组,南方组",
                "pageSize": "20",
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ),
    }


def test_profiled_source_paginates_confirmed_json_request_without_guessing_response_shape():
    from app.spot_ledger_sync import CANDIDATE_SOURCE_URL, ProfiledSalesContractSource, build_candidate_request_body

    class Response:
        status_code = 200

        def __init__(self, rows):
            self.rows = rows

        def json(self):
            return {"result": {"rows": self.rows, "total": 2, "pageCount": 2}}

    class Http:
        def __init__(self):
            self.calls = []

        def post(self, url, **kwargs):
            request_body = kwargs["json"]
            self.calls.append({"url": url, "json": request_body, "headers": kwargs["headers"], "timeout": kwargs["timeout"]})
            params = json.loads(request_body["params"])
            row = {
                "detail_id": f"D{params['pageNo']}",
                "spot_type": "现货",
                "contract_status": "生效",
                "quantity_group": "大客户组",
                "profit_group": "大客户组",
                "contract_number": f"C-{params['pageNo']}",
                "product_name": "铁矿石",
                "signed_date": "2026-08-20",
                "resource_date": "2026-08-15",
                "contract_quantity": 100,
                "business_category_code": "B07",
            }
            return Response([row])

    http = Http()
    profile = {
        "url": CANDIDATE_SOURCE_URL,
        "request_body": build_candidate_request_body("2026-08-24", page_no=1, page_size=1),
        "records_path": "result.rows",
        "total_path": "result.total",
        "page_count_path": "result.pageCount",
        "field_map": {
            "detail_id": "detail_id",
            "spot_type": "spot_type",
            "contract_status": "contract_status",
            "quantity_group": "quantity_group",
            "profit_group": "profit_group",
            "contract_number": "contract_number",
            "product_name": "product_name",
            "signed_date": "signed_date",
            "resource_date": "resource_date",
            "contract_quantity": "contract_quantity",
            "business_category_code": "business_category_code",
        },
        "pagination": {
            "params_key": "params",
            "page_number_key": "pageNo",
            "page_size_key": "pageSize",
        },
    }
    source = ProfiledSalesContractSource(profile, http=http, auth_provider=lambda: {"X-Test": "fixture"})

    scan = source.fetch_full_scan()

    assert scan.complete is True
    assert scan.source_mode == "profiled_http"
    assert scan.total_count == 2
    assert [record["source_detail_id"] for record in scan.records] == ["D1", "D2"]
    assert [json.loads(call["json"]["params"])["pageNo"] for call in http.calls] == [1, 2]
    assert all(call["url"] == CANDIDATE_SOURCE_URL for call in http.calls)


def test_confirmed_profile_reads_observed_jmreport_shape_and_maps_system_fields():
    from app.spot_ledger_sync import ProfiledSalesContractSource, build_candidate_source_profile

    class Response:
        status_code = 200

        def __init__(self, page_no):
            self.page_no = page_no

        def json(self):
            row = {
                "销售合同商品明细id": f"REAL-{self.page_no}",
                "期现货": "现货",
                "合同状态": "生效",
                "量归属组": "山东组",
                "业务毛利归属组": "唐山组",
                "业务类别": "贸易-代理落地-B09",
                "公司": "操作抬头A",
                "资源日期": "2026-08-20 00:00:00",
                "物资名称": "铁矿石",
                "合同卸货港": "日照港",
                "定价模式": "固定价",
                "中文船名": "测试船",
                "合同数量": "120.0000",
                "结算数量": "100.0000",
                "结案状态": "已结案",
                "资源单单价": "770.50",
                "资源方": "供应商A",
                "资源业务员": "采购业务员",
                "初始资源单创建人": "采购执行员",
                "签订日期": "2026-08-21 00:00:00",
                "合同单价": "820.25",
                "需求方": "签约客户",
                "销售合同号": f"XS-{self.page_no}",
                "需求业务员": "销售业务员",
                "合同创建人": "920109_销售执行员",
            }
            return {
                "success": True,
                "message": "",
                "code": 200,
                "result": {
                    "id": "1055351755192311808",
                    "dataList": {
                        "expData": {},
                        "replaceParams": {"pageNo": self.page_no, "pageSize": 1},
                        "TJJLYSHZ": {
                            "total": 2,
                            "count": 2,
                            "isPage": "1",
                            "isList": "1",
                            "dbType": "mysql",
                            "list": [row],
                            "linkList": None,
                        },
                    },
                },
                "timestamp": 1787587200000,
            }

    class Http:
        def post(self, _url, **kwargs):
            page_no = json.loads(kwargs["json"]["params"])["pageNo"]
            return Response(page_no)

    profile = build_candidate_source_profile("2026-08-25", page_size=1)
    source = ProfiledSalesContractSource(profile, http=Http(), auth_provider=lambda: {"X-Test-Auth": "fixture"})

    scan = source.fetch_full_scan()

    assert scan.complete is True
    assert scan.total_count == 2
    assert scan.page_count == 2
    assert [record["source_detail_id"] for record in scan.records] == ["REAL-1", "REAL-2"]
    first = scan.records[0]
    assert first["D"] == "贸易-代理落地-B09"
    assert first["E"] == "山东组" and first["AP"] == "唐山组" and first["AQ"] == "是"
    assert first["F"] == "操作抬头A"
    assert first["G"] == "2026-08-20" and first["U"] == "2026-08-21"
    assert first["H"] == "铁矿石" and first["I"] == "日照港" and first["J"] == "固定价"
    assert first["K"] == "测试船" and first["L"] == first["X"] == 100
    assert first["M"] == 770.5 and first["Z"] == 820.25
    assert first["Q"] == "供应商A" and first["S"] == "采购业务员" and first["T"] == "采购执行员"
    assert first["AB"] == "签约客户" and first["AD"] == "XS-1"
    assert first["AF"] == "销售业务员" and first["AG"] == "销售执行员"
    assert first["source_closed_state"] == "已结案"


def test_due_slots_are_daily_nine_to_eighteen_and_second_free():
    from app.spot_ledger_sync import due_spot_ledger_slots

    slots = due_spot_ledger_slots(datetime.fromisoformat("2026-08-24T18:15:42+08:00"))
    assert len(slots) == 10
    assert slots[0].endswith("09:00+08:00")
    assert slots[-1].endswith("18:00+08:00")
    assert all("." not in slot and ":42" not in slot for slot in slots)


def test_history_migration_only_updates_unique_match_and_defaults_to_dry_run(ledger_db, tmp_path):
    from openpyxl import Workbook, load_workbook
    from app.spot_ledger_sync import apply_full_scan, migrate_history_workbook

    scan = load_fixture_scan()
    apply_full_scan(scan, "2026-08-24T09:00+08:00")
    path = tmp_path / "history.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "现货业务台账"
    sheet.append(["销售合同号", "商品名称", "销售价格（元/吨）", "销售数量（吨）", "销售日期", "备注"])
    sheet.append(["C-100", "铁矿石", 800, 100, "2026-08-20", "历史人工备注"])
    sheet.append(["不存在", "铁矿石", 800, 100, "2026-08-20", "不应写入"])
    workbook.save(path)
    preview = migrate_history_workbook(path)
    assert preview["matched"] == 1
    assert preview["updated"] == 0
    applied = migrate_history_workbook(path, apply=True)
    assert applied["updated"] == 1
    with ledger_db.connect() as conn:
        row = conn.execute("SELECT \"AM\" FROM spot_ledger_records WHERE \"AD\" = 'C-100' ORDER BY source_detail_id LIMIT 1").fetchone()
    assert row["AM"] == "历史人工备注"


def test_history_migration_skips_non_numeric_text_in_numeric_manual_fields(ledger_db, tmp_path):
    from openpyxl import Workbook
    from app.spot_ledger_sync import apply_full_scan, migrate_history_workbook

    apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")
    path = tmp_path / "history-invalid-numeric.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "现货业务台账"
    sheet.append(["销售合同号", "商品名称", "销售价格（元/吨）", "销售数量（吨）", "销售日期", "期货量", "备注"])
    sheet.append(["C-100", "铁矿石", 800, 100, "2026-08-20", "无期货业务", "合法备注仍应迁移"])
    workbook.save(path)

    result = migrate_history_workbook(path, apply=True)

    assert result["updated"] == 1
    with ledger_db.connect() as conn:
        row = conn.execute(
            "SELECT \"AJ\", \"AM\" FROM spot_ledger_records WHERE \"AD\" = 'C-100' ORDER BY source_detail_id LIMIT 1"
        ).fetchone()
    assert row["AJ"] in (None, "")
    assert row["AM"] == "合法备注仍应迁移"


def test_history_migration_detects_third_row_headers_and_excel_parenthesis_variants(ledger_db, tmp_path):
    from openpyxl import Workbook
    from app.spot_ledger_sync import apply_full_scan, migrate_history_workbook

    scan = load_fixture_scan()
    apply_full_scan(scan, "2026-08-24T09:00+08:00")
    path = tmp_path / "history-with-title-rows.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "现货业务台账"
    sheet.append(["现货台账"])
    sheet.append(["现货业务台账", "", "采购", "", "销售"])
    sheet.append([
        "销售合同号",
        "商品名称",
        "销售价格\n（元/吨）",
        "销售数量(吨）",
        "销售日期",
        "实物含税盈亏 (万元）",
        "备注",
        "利润合计校验",
    ])
    sheet.append(["C-100", "铁矿石", 800, 100, "2026-08-20", 12.5, "历史人工备注", "一致"])
    sheet.append([None, None, None, None, None, None, None, 0])
    workbook.save(path)

    preview = migrate_history_workbook(path)

    assert preview["matched"] == 1
    assert preview["candidate_updates"] == 1
    assert preview["unmatched"] == 0


def test_stored_mapping_reconciliation_canonicalizes_supplier_and_product_without_touching_other_errors(ledger_db):
    from app.spot_ledger import SUPPLIER_MAPPINGS
    from app.spot_ledger_sync import apply_full_scan, reconcile_stored_mapping_state

    apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")
    source_supplier = next(iter(SUPPLIER_MAPPINGS))
    with ledger_db.connect() as conn:
        conn.execute(
            'UPDATE spot_ledger_records SET "Q" = ?, "AU" = ?, sync_status = ?, sync_error_summary = ? WHERE "AD" = ?',
            (
                SUPPLIER_MAPPINGS[source_supplier], "SFGB粉", "异常",
                '[{"type":"conversion_mapping","field":"Q"},{"type":"category_mapping","field":"AU"},{"type":"source","field":"AD"}]',
                "C-101",
            ),
        )

    preview = reconcile_stored_mapping_state()
    assert preview["dry_run"] is True
    assert preview["candidate_updates"] == 1
    assert preview["supplier_canonicalized"] == 1
    assert preview["category_reclassified"] == 1
    assert preview["errors_cleared"] == 2

    applied = reconcile_stored_mapping_state(apply=True)
    assert applied["updated"] == 1
    with ledger_db.connect() as conn:
        row = conn.execute(
            'SELECT "Q", "AU", sync_status, sync_error_summary FROM spot_ledger_records WHERE "AD" = ?',
            ("C-101",),
        ).fetchone()
    assert row["Q"] == source_supplier
    assert row["AU"] == "铁矿石"
    assert row["sync_status"] == "异常"
    assert json.loads(row["sync_error_summary"])[0]["field"] == "AD"


def test_history_migration_is_2026_only_and_does_not_overwrite_nonempty_conflicts(ledger_db, tmp_path):
    from openpyxl import Workbook
    from app.spot_ledger_sync import apply_full_scan, migrate_history_workbook

    apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")
    path = tmp_path / "history-focus-and-conflict.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "现货业务台账"
    sheet.append(["销售合同号", "商品名称", "销售价格（元/吨）", "销售数量（吨）", "销售日期", "船名", "备注"])
    sheet.append(["C-102", "铁矿石", 820, 100, "2026-08-18", "补录船", "应写入"])
    sheet.append(["C-101", "铁矿石", 795, 80, "2026-08-19", "Excel船", None])
    sheet.append(["C-103", "铁矿石", 790, 60, "2025-08-17", "历史船", "历史不写入"])
    workbook.save(path)

    preview = migrate_history_workbook(path)
    assert preview["skipped_historical"] == 1
    assert preview["conflicts"] == 1
    assert preview["candidate_updates"] == 1

    applied = migrate_history_workbook(path, apply=True)
    assert applied["updated"] == 1
    with ledger_db.connect() as conn:
        rows = conn.execute(
            'SELECT "AD", "K", "AM" FROM spot_ledger_records WHERE "AD" IN (?, ?, ?) ORDER BY "AD"',
            ("C-101", "C-102", "C-103"),
        ).fetchall()
    by_contract = {row["AD"]: row for row in rows}
    assert by_contract["C-102"]["K"] == "补录船"
    assert by_contract["C-102"]["AM"] == "应写入"
    assert by_contract["C-101"]["K"] == "船B"
    assert by_contract["C-103"]["AM"] in (None, "")


def test_history_sales_type_prefers_system_value_and_uses_excel_only_when_system_is_blank(ledger_db, tmp_path):
    from openpyxl import Workbook
    from app.spot_ledger_sync import apply_full_scan, migrate_history_workbook

    apply_full_scan(load_fixture_scan(), "2026-08-24T09:00+08:00")
    with ledger_db.connect() as conn:
        conn.execute('UPDATE spot_ledger_records SET "D" = ? WHERE "AD" = ?', ("", "C-103"))

    path = tmp_path / "history-sales-type.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "现货业务台账"
    sheet.append(["销售合同号", "商品名称", "销售价格（元/吨）", "销售数量（吨）", "销售日期", "销售类型"])
    sheet.append(["C-102", "铁矿石", 820, 100, "2026-08-18", "B05"])
    sheet.append(["C-103", "铁矿石", 790, 60, "2026-08-17", "贸易-落地-固定价-B05"])
    workbook.save(path)

    preview = migrate_history_workbook(path)
    assert preview["conflicts"] == 1
    assert preview["candidate_updates"] == 1
    applied = migrate_history_workbook(path, apply=True)
    assert applied["updated"] == 1

    with ledger_db.connect() as conn:
        rows = conn.execute(
            'SELECT "AD", "D" FROM spot_ledger_records WHERE "AD" IN (?, ?) ORDER BY "AD"',
            ("C-102", "C-103"),
        ).fetchall()
    by_contract = {row["AD"]: row for row in rows}
    assert by_contract["C-102"]["D"] == "B09"
    assert by_contract["C-103"]["D"] == "贸易-落地-固定价-B05"
