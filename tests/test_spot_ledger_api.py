from __future__ import annotations

import asyncio
import io
import os
from pathlib import Path
import sys

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from app import db


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "spot_ledger_sales_contract_fixture.json"


@pytest.fixture
def ledger_context(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setattr(db, "DATA_DIR", tmp_path)
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "spot-ledger-api.db")
    db.init_db()
    from app.spot_ledger_sync import FixtureSalesContractSource, apply_full_scan

    apply_full_scan(FixtureSalesContractSource(FIXTURE_PATH).fetch_full_scan(), "2026-08-24T09:00+08:00")
    with db.connect() as conn:
        admin = dict(conn.execute("SELECT * FROM users WHERE role = '管理员' ORDER BY id LIMIT 1").fetchone())
        trade = db._exec(
            conn.cursor(),
            "INSERT INTO users (name, username, department, password_hash, role) VALUES (?, ?, ?, ?, ?)",
            ("贸易测试", "trade-test", "贸易处", db.password_hash("pass"), "用户"),
        )
        trade_id = trade.lastrowid
        db._exec(
            conn.cursor(),
            "INSERT INTO module_permissions (user_id, module_code, can_view, can_edit, can_sensitive) VALUES (?, ?, 1, 1, 0)",
            (trade_id, "spot_ledger"),
        )
        trade_user = dict(conn.execute("SELECT * FROM users WHERE id = ?", (trade_id,)).fetchone())
    return admin, trade_user


def test_records_support_combined_filters_and_expose_all_field_definitions(ledger_context):
    from app.spot_ledger import FIELD_CODES, get_records, field_definitions

    admin, _ = ledger_context
    result = get_records(
        sales_group="山东组", profit_group="唐山组", sales_type="船货-落地", product_name="铁矿石",
        port="日照港", operation_title="操作抬头A", supplier="供应商A", customer="客户C",
        contract_number="C-102", user=admin,
    )
    assert [row["source_detail_id"] for row in result["records"]] == ["D1004"]
    assert [field["code"] for field in field_definitions(user=admin)["fields"]] == list(FIELD_CODES)


def test_closed_state_filter_uses_source_settlement_state_instead_of_ledger_eligibility(ledger_context):
    from app.spot_ledger import get_records

    admin, _ = ledger_context
    closed = get_records(closed_state="已结案", user=admin)["records"]
    open_records = get_records(closed_state="未结案", user=admin)["records"]

    assert [row["source_detail_id"] for row in closed] == ["D1004"]
    assert "D1004" not in {row["source_detail_id"] for row in open_records}
    assert "D1001" in {row["source_detail_id"] for row in open_records}


def test_pending_and_sync_error_views_are_explicit(ledger_context):
    from app.spot_ledger import get_pending, get_sync_errors

    admin, _ = ledger_context
    pending = get_pending(user=admin)
    errors = get_sync_errors(user=admin)
    assert any(row["source_detail_id"] == "D1004" for row in pending["records"])
    assert any(row["source_detail_id"] == "D1009" for row in errors["records"])
    assert any(error["type"] == "conversion_mapping" for error in errors["records"][0]["sync_error_summary"])


def test_source_readiness_requires_administrator_role(ledger_context):
    from app.spot_ledger import source_readiness_view

    _, trade_user = ledger_context
    with pytest.raises(HTTPException) as denied:
        source_readiness_view(user=trade_user)
    assert denied.value.status_code == 403


def test_source_readiness_returns_only_official_json_schema_metadata(ledger_context, monkeypatch):
    from app import spot_ledger_sync as sync
    from app.spot_ledger import source_readiness_view

    admin, _ = ledger_context
    probe = {
        "ok": True,
        "source_mode": "official_json",
        "http_status": 200,
        "response_code": "200",
        "schema_paths": ["code", "data.rows[]", "data.rows[].saleContractId"],
        "detail_response_code": "200",
        "detail_schema_paths": ["code", "data.saleContractMxList[]"],
    }
    monkeypatch.setattr(sync, "probe_official_sales_contract_api", lambda: probe)

    assert source_readiness_view(user=admin) == probe


@pytest.mark.parametrize(
    ("source_error", "expected_detail"),
    [
        pytest.param(
            lambda: __import__("app.spot_ledger_sync", fromlist=["SalesContractSourceError"]).SalesContractSourceError(
                "auth_unavailable",
                "personal-password bearer-token source-response",
                stage="login_page_http",
                http_status=403,
            ),
            {"code": "auth_unavailable", "stage": "login_page_http", "http_status": 403},
            id="known-source-error",
        ),
        pytest.param(
            lambda: RuntimeError("personal-password bearer-token source-response"),
            {"code": "source_probe_failed"},
            id="unexpected-error",
        ),
    ],
)
def test_source_readiness_redacts_source_errors(ledger_context, monkeypatch, source_error, expected_detail):
    from app import spot_ledger_sync as sync
    from app.spot_ledger import source_readiness_view

    admin, _ = ledger_context

    def fail_probe():
        raise source_error()

    monkeypatch.setattr(sync, "probe_official_sales_contract_api", fail_probe)

    with pytest.raises(HTTPException) as failed:
        source_readiness_view(user=admin)
    assert failed.value.status_code == 503
    assert failed.value.detail == expected_detail
    assert "personal-password" not in str(failed.value)
    assert "bearer-token" not in str(failed.value)


def test_source_dry_run_requires_administrator_role(ledger_context):
    from app.spot_ledger import source_dry_run_view

    _, trade_user = ledger_context
    with pytest.raises(HTTPException) as denied:
        source_dry_run_view(user=trade_user)
    assert denied.value.status_code == 403


def test_source_dry_run_returns_only_aggregate_result(ledger_context, monkeypatch):
    from app import spot_ledger_sync as sync
    from app.spot_ledger import source_dry_run_view

    admin, _ = ledger_context
    aggregate = {
        "ok": True,
        "source_mode": "official_json",
        "page_count": 2,
        "counts": {"active_contract_count": 12, "eligible_record_count": 7},
        "field_coverage": {"AD": {"filled_count": 7, "total_count": 7}},
        "scan_error_types": {},
        "record_error_types": {},
    }
    monkeypatch.setattr(sync, "run_official_source_dry_run", lambda: aggregate)

    assert source_dry_run_view(user=admin) == aggregate


def test_manual_edit_requires_sensitive_permission_and_cannot_change_system_field(ledger_context):
    from app.spot_ledger import SpotLedgerPatch, patch_record

    admin, trade_user = ledger_context
    with pytest.raises(HTTPException) as denied:
        patch_record("spot:D1001", SpotLedgerPatch(values={"C": "自主建仓"}), user=trade_user)
    assert denied.value.status_code == 403
    updated = patch_record(
        "spot:D1001",
        SpotLedgerPatch(values={"C": "自主建仓", "N": 0, "O": -1, "Y": 0, "AM": "人工补录"}),
        user=admin,
    )
    assert updated["record"]["C"] == "自主建仓"
    assert updated["record"]["O"] == -1
    with pytest.raises(HTTPException) as readonly:
        patch_record("spot:D1001", SpotLedgerPatch(values={"AD": "不能改合同号"}), user=admin)
    assert readonly.value.status_code == 400


def test_strategy_hedging_requires_complete_open_close_and_rejects_partial_close(ledger_context):
    from app.spot_ledger import StrategicHedgingIn, create_strategic_hedging

    admin, _ = ledger_context
    payload = StrategicHedgingIn(
        group_name="大客户组", account="模拟账户", contract="I2609", open_direction="多",
        opened_at="2026-08-24 09:00:00", open_quantity=10, quantity_unit="吨",
        open_price=800, price_currency="元/吨", closed_at="2026-08-25 09:00:00",
        close_quantity=10, close_price=820, remark="本地 fixture 手工记录",
    )
    created = create_strategic_hedging(payload, user=admin)
    assert created["record"]["strategic_status"] == "已平仓"
    partial = payload.model_copy(update={"close_quantity": 5})
    with pytest.raises(HTTPException) as invalid:
        create_strategic_hedging(partial, user=admin)
    assert invalid.value.status_code == 400


def test_export_defaults_to_a_to_ay_and_adds_technical_key_only_when_requested(ledger_context):
    from app.spot_ledger import export_records

    admin, _ = ledger_context

    async def read_response(response):
        return b"".join([chunk async for chunk in response.body_iterator])

    default_bytes = asyncio.run(read_response(export_records(user=admin)))
    technical_bytes = asyncio.run(read_response(export_records(include_technical_key=True, user=admin)))
    default_headers = list(load_workbook(io.BytesIO(default_bytes), read_only=True).active.iter_rows(values_only=True))[0]
    technical_headers = list(load_workbook(io.BytesIO(technical_bytes), read_only=True).active.iter_rows(values_only=True))[0]
    assert len(default_headers) == 51
    assert "销售合同商品明细 ID" not in default_headers
    assert len(technical_headers) == 52
    assert technical_headers[-1] == "销售合同商品明细 ID"


def test_spot_ledger_routes_are_registered_in_main_app():
    from app.main import app

    paths = {route.path for route in app.routes}
    assert "/api/spot-ledger/records" in paths
    assert "/api/spot-ledger/export" in paths
    assert "/api/spot-ledger/strategic-hedging" in paths
    assert "/api/spot-ledger/source-readiness" in paths
    assert "/api/spot-ledger/source-dry-run" in paths
    assert not any(path.endswith("/sync-now") for path in paths)
