"""交易管理模块的数据结构与业务规则测试。"""
import importlib.util
import os
import sys

import pytest
from openpyxl import Workbook, load_workbook


sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from app import db, trading_management


TRADING_TABLES = {
    "trading_accounts",
    "trading_import_batches",
    "trading_source_rows",
    "trading_fact_identities",
    "trading_trade_facts",
    "trading_close_facts",
    "trading_position_snapshots",
    "trading_contract_specs",
    "trading_fact_close_allocations",
    "trading_close_trade_links",
    "trading_business_subjects",
    "trading_strategies",
    "trading_business_assignments",
    "trading_business_close_allocations",
    "trading_business_allocation_audit",
}


def use_temp_db(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setattr(db, "DATA_DIR", tmp_path)
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "trading-management.db")
    db.init_db()


def test_trading_management_schema_contains_isolated_tables(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    with db.connect() as conn:
        actual = {
            row["name"]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name LIKE 'trading_%'"
            ).fetchall()
        }

    assert TRADING_TABLES <= actual
    assert not any(name.startswith("sh_junneng") for name in TRADING_TABLES)


def test_trading_management_seeds_verified_sample_reference_data(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    with db.connect() as conn:
        account = conn.execute(
            "SELECT * FROM trading_accounts WHERE account_code = 'hongyuan_futures'"
        ).fetchone()
        specs = conn.execute(
            "SELECT exchange, product_code, asset_type, contract_multiplier FROM trading_contract_specs"
        ).fetchall()

    assert account["display_name"] == "宏源期货账户"
    assert {(row["exchange"], row["product_code"], row["asset_type"], row["contract_multiplier"]) for row in specs} >= {
        ("上期所", "rb", "future", 10),
        ("上期所", "hc", "future", 10),
        ("大商所", "i", "future", 100),
        ("大商所", "i", "option", 100),
        ("大商所", "j", "future", 100),
    }


def test_new_trading_menu_permissions_are_added_without_overwriting_existing_choices(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    with db.connect() as conn:
        futures_user = conn.execute(
            """
            INSERT INTO users (name, username, department, password_hash, role)
            VALUES ('期货测试', 'futures-test', '期货组', 'x', '用户')
            """
        ).lastrowid
        leader_user = conn.execute(
            """
            INSERT INTO users (name, username, department, password_hash, role)
            VALUES ('领导测试', 'leader-test', '公司领导', 'x', '领导')
            """
        ).lastrowid
        db.sync_trading_module_permissions(conn.cursor())
        futures_permissions = conn.execute(
            "SELECT can_view, can_edit, can_sensitive FROM module_permissions WHERE user_id = ? AND module_code LIKE 'trading_%'",
            (futures_user,),
        ).fetchall()
        leader_permissions = conn.execute(
            "SELECT can_view, can_edit, can_sensitive FROM module_permissions WHERE user_id = ? AND module_code LIKE 'trading_%'",
            (leader_user,),
        ).fetchall()
        conn.execute(
            "UPDATE module_permissions SET can_edit = 0 WHERE user_id = ? AND module_code = 'trading_overview'",
            (futures_user,),
        )
        db.sync_trading_module_permissions(conn.cursor())
        preserved = conn.execute(
            "SELECT can_edit FROM module_permissions WHERE user_id = ? AND module_code = 'trading_overview'",
            (futures_user,),
        ).fetchone()["can_edit"]

    assert len(futures_permissions) == 5
    assert {(row["can_view"], row["can_edit"], row["can_sensitive"]) for row in futures_permissions} == {(1, 1, 0)}
    assert len(leader_permissions) == 5
    assert {(row["can_view"], row["can_edit"], row["can_sensitive"]) for row in leader_permissions} == {(1, 0, 0)}
    assert preserved == 0


def test_trading_management_router_module_exists():
    assert importlib.util.find_spec("app.trading_management") is not None


def test_postgres_executemany_uses_batched_protocol(monkeypatch):
    calls = []

    class Cursor:
        def executemany(self, sql, rows):
            raise AssertionError("PostgreSQL should not send one execute per row")

    monkeypatch.setattr(db, "_is_pg", lambda: True)
    monkeypatch.setattr(
        db.psycopg2.extras,
        "execute_batch",
        lambda cur, sql, rows, page_size: calls.append((sql, list(rows), page_size)),
    )

    db._executemany(Cursor(), "INSERT INTO sample (value) VALUES (?)", [(1,), (2,)])

    assert calls == [("INSERT INTO sample (value) VALUES (%s)", [(1,), (2,)], 1000)]


def test_trading_management_modules_are_new_first_level_menu():
    modules = [item for item in db.MODULES if item[0] == "交易管理"]

    assert modules == [
        ("交易管理", "trading_overview", "总览"),
        ("交易管理", "trading_positions", "持仓与交易"),
        ("交易管理", "trading_sh_junneng", "上海钧能台账"),
        ("交易管理", "trading_options", "期权台账"),
        ("交易管理", "trading_export", "汇总与导出"),
    ]


def build_trade_workbook(path):
    book = Workbook()
    sheet = book.active
    sheet.title = "成交记录"
    sheet.append([
        "交易所", "合约", "买卖", "开平", "手数", "成交价", "成交额",
        "平仓盈亏", "手续费", "投保", "权利金收支",
    ])
    sheet.append(["20260620"])
    sheet.append(["SHFE", "rb2610", "买", "开", 2, 3100, 62000, 0, 6, "保", 0])
    sheet.append(["20260630"])
    sheet.append(["SHFE", "rb2610", "卖", "平今", 2, 3120, 62400, 999, 6, "保", 0])
    sheet.append(["DCE", "i2609-C-700", "买", "开", 1, 4.2, 420, 0, 1.2, "投", -420])
    book.save(path)
    return path


def build_close_workbook(path):
    book = Workbook()
    sheet = book.active
    sheet.title = "平仓记录"
    sheet.append([
        "交易所", "合约", "开仓日期", "买入", "卖出", "手数", "价格",
        "逐笔平仓盈亏", "盯市平仓盈亏", "手续费", "权利金收支",
    ])
    sheet.append(["20260630"])
    sheet.append(["SHFE", "rb2610", "20260620", 2, None, None, 3100, None, None, None, None])
    sheet.append([None, None, None, None, 2, 2, 3120, 1200, 1100, None, 0])
    book.save(path)
    return path


def build_position_workbook(path):
    book = Workbook()
    sheet = book.active
    sheet.title = "期末持仓"
    sheet.append([
        "交易所", "开仓日期", "合约", "买卖", "手数", "价格", "浮动盈亏",
        "盯市盈亏", "占用保证金", "投保",
    ])
    sheet.append(["20260630"])
    sheet.append(["SHFE", "20260630", "rb2610", "买", 10, 3100, 2000, 1800, 50000, "保"])
    book.save(path)
    return path


def test_parser_normalizes_grouped_wenhua_workbooks(tmp_path):
    trades = trading_management.parse_trade_workbook(build_trade_workbook(tmp_path / "trades.xlsx"))
    closes = trading_management.parse_close_workbook(build_close_workbook(tmp_path / "closes.xlsx"))
    positions = trading_management.parse_position_workbook(build_position_workbook(tmp_path / "positions.xlsx"))

    assert trades[0]["open_close"] == "开仓"
    assert trades[1]["open_close"] == "平仓"
    assert trades[2]["asset_type"] == "option"
    assert closes[0]["fact_close_pnl"] == 1200
    assert closes[0]["open_side"] == "买"
    assert positions[0]["margin"] == 50000
    assert positions[0]["valuation_status"] == "pending_calculation"


def test_parser_rejects_unknown_trade_headers(tmp_path):
    book = Workbook()
    book.active.title = "成交记录"
    book.active.append(["合约", "未知数量"])
    book.save(tmp_path / "bad.xlsx")

    with pytest.raises(ValueError, match="缺少必要字段"):
        trading_management.parse_trade_workbook(tmp_path / "bad.xlsx")


def test_fact_signature_is_stable_after_normalization():
    first = {
        "date": "20260630",
        "exchange": " shfe ",
        "contract": "RB2610",
        "side": "买",
        "open_close_raw": "开",
        "quantity": 2,
        "price": 3100.0,
        "turnover": 62000,
        "fee": 6,
    }
    second = dict(first, exchange="SHFE", contract="rb2610", price="3100.000000")

    assert trading_management.build_fact_signature("trade", "A001", first) == trading_management.build_fact_signature(
        "trade", "A001", second
    )


def test_import_preview_requires_all_three_files(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    with db.connect() as conn:
        conn.execute(
            "INSERT INTO trading_accounts (account_code, display_name) VALUES (?, ?)",
            ("A001", "测试账户"),
        )

    with pytest.raises(ValueError, match="成交、平仓、持仓三表必须齐全"):
        trading_management.preview_trading_import(
            account_id=1,
            trade_path=build_trade_workbook(tmp_path / "trades.xlsx"),
            close_path=None,
            position_path=build_position_workbook(tmp_path / "positions.xlsx"),
            actor="tester",
        )


def test_import_preview_rejects_missing_account(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    with pytest.raises(ValueError, match="交易账户不存在或已停用"):
        trading_management.preview_trading_import(
            account_id=999,
            trade_path=build_trade_workbook(tmp_path / "trades.xlsx"),
            close_path=build_close_workbook(tmp_path / "closes.xlsx"),
            position_path=build_position_workbook(tmp_path / "positions.xlsx"),
            actor="tester",
        )


def test_import_preview_rejects_partially_overlapping_active_batch(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    with db.connect() as conn:
        conn.execute(
            "INSERT INTO trading_accounts (account_code, display_name) VALUES (?, ?)",
            ("A001", "测试账户"),
        )
        conn.execute(
            """
            INSERT INTO trading_import_batches
                (account_id, range_start, range_end, status)
            VALUES (1, '20260601', '20260630', 'active')
            """
        )

    with pytest.raises(ValueError, match="日期范围部分重叠"):
        trading_management.preview_trading_import(
            account_id=1,
            trade_path=build_trade_workbook(tmp_path / "trades.xlsx"),
            close_path=build_close_workbook(tmp_path / "closes.xlsx"),
            position_path=build_position_workbook(tmp_path / "positions.xlsx"),
            actor="tester",
        )


def create_preview_batch(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)
    with db.connect() as conn:
        conn.execute(
            "INSERT INTO trading_accounts (account_code, display_name) VALUES (?, ?)",
            ("A001", "测试账户"),
        )
    return trading_management.preview_trading_import(
        account_id=1,
        trade_path=build_trade_workbook(tmp_path / "trades.xlsx"),
        close_path=build_close_workbook(tmp_path / "closes.xlsx"),
        position_path=build_position_workbook(tmp_path / "positions.xlsx"),
        actor="tester",
    )


def test_confirm_import_writes_immutable_fact_versions(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)

    result = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    assert result["status"] == "active"
    assert result["counts"] == {"trade": 3, "close": 1, "position": 1}
    with db.connect() as conn:
        batch = conn.execute(
            "SELECT status, confirmed_by FROM trading_import_batches WHERE id = ?",
            (preview["preview_batch_id"],),
        ).fetchone()
        assert batch["status"] == "active"
        assert batch["confirmed_by"] == "tester"
        assert conn.execute("SELECT COUNT(*) AS c FROM trading_source_rows").fetchone()["c"] == 5
        assert conn.execute("SELECT COUNT(*) AS c FROM trading_trade_facts").fetchone()["c"] == 3
        raw_rows = [row["raw_json"] for row in conn.execute("SELECT raw_json FROM trading_source_rows")]
        assert any("平今" in raw for raw in raw_rows)

    with pytest.raises(ValueError, match="预览批次已确认或不可用"):
        trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")


def test_confirm_same_range_supersedes_old_batch_and_keeps_source_rows(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    first = trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    with db.connect() as conn:
        first_identity_ids = {
            row["id"] for row in conn.execute("SELECT id FROM trading_fact_identities").fetchall()
        }

    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    second = trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")

    with db.connect() as conn:
        statuses = {
            row["id"]: row["status"]
            for row in conn.execute("SELECT id, status FROM trading_import_batches").fetchall()
        }
        second_identity_ids = {
            row["id"] for row in conn.execute("SELECT id FROM trading_fact_identities").fetchall()
        }
        source_count = conn.execute("SELECT COUNT(*) AS c FROM trading_source_rows").fetchone()["c"]

    assert statuses[first["batch_id"]] == "superseded"
    assert statuses[second["batch_id"]] == "active"
    assert first_identity_ids == second_identity_ids
    assert source_count == 10


def test_overwrite_trade_pnl_uses_only_active_close_fact_version(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    first = trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(first["batch_id"])
    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    second = trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")
    trading_management.match_imported_facts(second["batch_id"])

    trades = trading_management.query_fact_rows(
        "trades", trading_management.FactFilters(page=1, page_size=20)
    )

    assert trades["summary"]["fact_close_pnl"] == 1200


def test_same_stable_identity_keeps_business_assignment_after_reimport(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    with db.connect() as conn:
        trade_identity_id = conn.execute(
            "SELECT identity_id FROM trading_trade_facts ORDER BY id LIMIT 1"
        ).fetchone()["identity_id"]
        subject_id = conn.execute(
            "SELECT id FROM trading_business_subjects WHERE normalized_name = '上海钧能'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO trading_business_assignments
                (trade_identity_id, business_subject_id, business_type, assigned_by)
            VALUES (?, ?, 'basic_hedging', 'tester')
            """,
            (trade_identity_id, subject_id),
        )

    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")

    with db.connect() as conn:
        assignment = conn.execute(
            "SELECT trade_identity_id, business_type FROM trading_business_assignments"
        ).fetchone()

    assert assignment["trade_identity_id"] == trade_identity_id
    assert assignment["business_type"] == "basic_hedging"


def test_changed_assigned_trade_is_not_force_inherited(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    with db.connect() as conn:
        old_identity_id = conn.execute(
            "SELECT identity_id FROM trading_trade_facts ORDER BY id LIMIT 1"
        ).fetchone()["identity_id"]
        subject_id = conn.execute(
            "SELECT id FROM trading_business_subjects WHERE normalized_name = '上海钧能'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO trading_business_assignments
                (trade_identity_id, business_subject_id, business_type, assigned_by)
            VALUES (?, ?, 'basic_hedging', 'tester')
            """,
            (old_identity_id, subject_id),
        )

    workbook = load_workbook(tmp_path / "trades.xlsx")
    workbook["成交记录"]["F3"] = 3110
    workbook.save(tmp_path / "trades.xlsx")
    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")

    with db.connect() as conn:
        active_row = conn.execute(
            """
            SELECT tf.identity_id, tf.verification_status
            FROM trading_trade_facts tf
            JOIN trading_import_batches b ON b.id = tf.batch_id
            WHERE b.status = 'active' AND tf.source_row_id = (
                SELECT MIN(source_row_id) FROM trading_trade_facts WHERE batch_id = b.id
            )
            """
        ).fetchone()
        inherited = conn.execute(
            "SELECT COUNT(*) AS c FROM trading_business_assignments WHERE trade_identity_id = ?",
            (active_row["identity_id"],),
        ).fetchone()["c"]

    assert active_row["identity_id"] != old_identity_id
    assert inherited == 0
    assert active_row["verification_status"] == "inheritance_review_required"


def test_fact_matching_links_close_trade_and_open_trade(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    result = trading_management.match_imported_facts(confirmed["batch_id"])

    assert result == {"close_trade_links": 1, "fact_close_allocations": 1, "pending_closes": 0}
    with db.connect() as conn:
        close = conn.execute(
            "SELECT fact_close_pnl, matched_fee, fee_status FROM trading_close_facts"
        ).fetchone()
        link = conn.execute(
            "SELECT matched_quantity, allocated_fee FROM trading_close_trade_links"
        ).fetchone()
        allocation = conn.execute(
            "SELECT matched_quantity FROM trading_fact_close_allocations"
        ).fetchone()

    assert close["fact_close_pnl"] == 1200
    assert close["matched_fee"] == 6
    assert close["fee_status"] == "matched"
    assert link["matched_quantity"] == 2
    assert allocation["matched_quantity"] == 2


def test_fact_query_uses_same_filters_for_items_and_summary(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(confirmed["batch_id"])

    result = trading_management.query_fact_rows(
        "trades",
        trading_management.FactFilters(contract="rb", page=1, page_size=20),
    )

    assert result["total_items"] == 2
    assert result["summary"]["record_count"] == 2
    assert result["summary"]["quantity"] == 4
    assert result["page"] == 1
    assert result["total_pages"] == 1
    close_trade = next(item for item in result["items"] if item["open_close"] == "平仓")
    assert close_trade["fact_close_pnl"] == 1200


def test_fact_positions_report_missing_historical_snapshot(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    result = trading_management.query_fact_rows(
        "positions",
        trading_management.FactFilters(end_date="20260629", page=1, page_size=20),
    )

    assert result["items"] == []
    assert result["data_status"] == "no_position_snapshot"


def test_close_query_and_overview_use_fact_pnl(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(confirmed["batch_id"])

    closes = trading_management.query_fact_rows(
        "closes",
        trading_management.FactFilters(contract="rb", page=1, page_size=20),
    )
    overview = trading_management.build_overview(
        trading_management.FactFilters(start_date="20260601", end_date="20260630")
    )

    assert closes["summary"]["record_count"] == 1
    assert closes["summary"]["fact_close_pnl"] == 1200
    assert closes["summary"]["fee"] == 6
    assert overview["trades"]["record_count"] == 3
    assert overview["closes"]["fact_close_pnl"] == 1200
    assert overview["positions"]["margin"] == 50000
    assert overview["positions"]["snapshot_date"] == "20260630"
    assert overview["positions"]["floating_pnl_status"] == "pending_calculation"


def test_overview_returns_daily_close_pnl_from_active_facts(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(confirmed["batch_id"])

    overview = trading_management.build_overview(
        trading_management.FactFilters(page=1, page_size=20)
    )

    assert overview["daily_close_pnl"] == [{"date": "20260630", "fact_close_pnl": 1200.0}]


def test_overview_daily_close_pnl_ignores_superseded_batch(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    first = trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(first["batch_id"])
    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    second = trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")
    trading_management.match_imported_facts(second["batch_id"])

    overview = trading_management.build_overview(
        trading_management.FactFilters(page=1, page_size=20)
    )

    assert overview["daily_close_pnl"] == [{"date": "20260630", "fact_close_pnl": 1200.0}]


def test_fact_matching_batches_database_writes_instead_of_round_trip_per_close(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    batch_id = confirmed["batch_id"]
    with db.connect() as conn:
        conn.execute("UPDATE trading_trade_facts SET quantity = 100 WHERE batch_id = ?", (batch_id,))
        template = conn.execute(
            "SELECT * FROM trading_close_facts WHERE batch_id = ?", (batch_id,)
        ).fetchone()
        conn.execute("UPDATE trading_close_facts SET quantity = 0.05 WHERE id = ?", (template["id"],))
        for index in range(19):
            identity_id = conn.execute(
                "INSERT INTO trading_fact_identities (account_id, fact_type, stable_key) VALUES (1, 'close', ?)",
                (f"bulk-close-{index}",),
            ).lastrowid
            conn.execute(
                """
                INSERT INTO trading_close_facts
                    (identity_id, batch_id, source_row_id, open_date, close_date, exchange, contract,
                     asset_type, open_side, close_side, quantity, open_price, close_price,
                     fact_close_pnl, matched_fee, fee_status, verification_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0.05, ?, ?, ?, ?, ?, ?)
                """,
                (
                    identity_id, batch_id, template["source_row_id"], template["open_date"],
                    template["close_date"], template["exchange"], template["contract"],
                    template["asset_type"], template["open_side"], template["close_side"],
                    template["open_price"], template["close_price"], template["fact_close_pnl"],
                    template["matched_fee"], template["fee_status"], template["verification_status"],
                ),
            )
        conn.commit()

    calls = {"execute": 0, "executemany": 0}
    original_exec = db._exec
    original_executemany = db._executemany

    def counted_exec(*args, **kwargs):
        calls["execute"] += 1
        return original_exec(*args, **kwargs)

    def counted_executemany(*args, **kwargs):
        calls["executemany"] += 1
        return original_executemany(*args, **kwargs)

    monkeypatch.setattr(db, "_exec", counted_exec)
    monkeypatch.setattr(db, "_executemany", counted_executemany)

    result = trading_management.match_imported_facts(batch_id)

    assert result["fact_close_allocations"] == 20
    assert calls["execute"] + calls["executemany"] <= 12


def test_default_business_allocations_batch_specs_and_writes(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    batch_id = confirmed["batch_id"]
    trading_management.match_imported_facts(batch_id)
    with db.connect() as conn:
        template = conn.execute(
            "SELECT * FROM trading_fact_close_allocations WHERE close_identity_id IN "
            "(SELECT identity_id FROM trading_close_facts WHERE batch_id = ?)",
            (batch_id,),
        ).fetchone()
        for _ in range(19):
            conn.execute(
                """
                INSERT INTO trading_fact_close_allocations
                    (close_identity_id, open_trade_identity_id, matched_quantity, match_rule_version)
                VALUES (?, ?, ?, 'wenhua-fifo-v1')
                """,
                (template["close_identity_id"], template["open_trade_identity_id"], 0.01),
            )
        conn.commit()

    calls = {"execute": 0, "executemany": 0}
    original_exec = db._exec
    original_executemany = db._executemany

    def counted_exec(*args, **kwargs):
        calls["execute"] += 1
        return original_exec(*args, **kwargs)

    def counted_executemany(*args, **kwargs):
        calls["executemany"] += 1
        return original_executemany(*args, **kwargs)

    monkeypatch.setattr(db, "_exec", counted_exec)
    monkeypatch.setattr(db, "_executemany", counted_executemany)

    result = trading_management.rebuild_default_business_allocations(batch_id)

    assert result["allocation_count"] == 20
    assert calls["execute"] + calls["executemany"] <= 8


def test_overwrite_import_prefetches_business_inheritance_once(tmp_path, monkeypatch):
    first_preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(first_preview["preview_batch_id"], actor="tester")
    second_preview = trading_management.preview_trading_import(
        account_id=1,
        trade_path=tmp_path / "trades.xlsx",
        close_path=tmp_path / "closes.xlsx",
        position_path=tmp_path / "positions.xlsx",
        actor="tester2",
    )
    inheritance_queries = []
    original_exec = db._exec

    def counted_exec(cur, sql, params=None):
        if "JOIN trading_business_assignments old_ba" in sql:
            inheritance_queries.append((sql, params))
        return original_exec(cur, sql, params)

    monkeypatch.setattr(db, "_exec", counted_exec)

    trading_management.confirm_trading_import(second_preview["preview_batch_id"], actor="tester2")

    assert len(inheritance_queries) == 1


def test_confirm_import_batches_source_identity_and_fact_inserts(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    calls = {"single_insert": 0, "batch_insert": 0}
    original_last_insert = db._last_insert_id
    original_executemany = db._executemany

    def counted_last_insert(*args, **kwargs):
        calls["single_insert"] += 1
        return original_last_insert(*args, **kwargs)

    def counted_executemany(*args, **kwargs):
        calls["batch_insert"] += 1
        return original_executemany(*args, **kwargs)

    monkeypatch.setattr(db, "_last_insert_id", counted_last_insert)
    monkeypatch.setattr(db, "_executemany", counted_executemany)

    result = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    assert result["counts"] == {"trade": 3, "close": 1, "position": 1}
    assert calls["single_insert"] == 0
    assert calls["batch_insert"] >= 6


def test_fact_api_routes_are_registered():
    paths = {route.path for route in trading_management.router.routes}

    assert {
        "/overview",
        "/facts/positions",
        "/facts/closes",
        "/facts/trades",
        "/imports",
        "/imports/{batch_id}/validation",
    } <= paths


def test_business_config_supports_controlled_subjects_and_reusable_strategies(tmp_path, monkeypatch):
    use_temp_db(tmp_path, monkeypatch)

    subject = trading_management.create_business_subject(" 上海钧能 ", actor="tester")
    same_subject = trading_management.create_business_subject("上海钧能", actor="tester")
    strategy = trading_management.get_or_create_strategy("代内部公司套保", actor="tester")
    same_strategy = trading_management.get_or_create_strategy(" 代内部公司套保 ", actor="tester")
    config = trading_management.list_trading_config()

    assert subject["id"] == same_subject["id"]
    assert strategy["id"] == same_strategy["id"]
    assert config["business_types"] == ["basic_hedging", "strategic_hedging"]
    assert "上海钧能" in {item["name"] for item in config["subjects"]}
    assert "期货组" in {item["name"] for item in config["subjects"]}
    assert "代内部公司套保" in {item["name"] for item in config["strategies"]}
    assert "战略套保-期权结构化套利" in {item["name"] for item in config["strategies"]}


def test_classification_assigns_complete_trade_identities(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    subject = trading_management.create_business_subject("上海钧能", actor="tester")
    with db.connect() as conn:
        trade_ids = [
            row["identity_id"]
            for row in conn.execute(
                "SELECT identity_id FROM trading_trade_facts ORDER BY id LIMIT 2"
            ).fetchall()
        ]

    result = trading_management.classify_trade_identities(
        trade_ids,
        business_subject_id=subject["id"],
        business_type="basic_hedging",
        strategy_name="代内部公司套保",
        instruction_text="上海钧能钢材套保",
        actor="tester",
    )

    assert result["assigned_count"] == 2
    with db.connect() as conn:
        assignments = conn.execute(
            "SELECT * FROM trading_business_assignments ORDER BY trade_identity_id"
        ).fetchall()
        audit_count = conn.execute(
            "SELECT COUNT(*) AS c FROM operation_logs WHERE module_code = 'trading_positions'"
        ).fetchone()["c"]
    assert len(assignments) == 2
    assert {row["business_type"] for row in assignments} == {"basic_hedging"}
    assert audit_count == 2


def test_classification_rejects_partial_quantity_and_disabled_subject(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    subject = trading_management.create_business_subject("上海钧能", actor="tester")
    with db.connect() as conn:
        trade_id = conn.execute(
            "SELECT identity_id FROM trading_trade_facts ORDER BY id LIMIT 1"
        ).fetchone()["identity_id"]

    with pytest.raises(ValueError, match="不允许按手数拆分"):
        trading_management.classify_trade_identities(
            [trade_id], subject["id"], "basic_hedging", "", "", "tester", requested_quantity=1
        )

    with db.connect() as conn:
        conn.execute("UPDATE trading_business_subjects SET is_active = 0 WHERE id = ?", (subject["id"],))
    with pytest.raises(ValueError, match="业务归属不存在或已停用"):
        trading_management.classify_trade_identities(
            [trade_id], subject["id"], "basic_hedging", "", "", "tester"
        )


def test_remove_business_assignment_keeps_audit(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    subject = trading_management.create_business_subject("上海钧能", actor="tester")
    with db.connect() as conn:
        trade_id = conn.execute(
            "SELECT identity_id FROM trading_trade_facts ORDER BY id LIMIT 1"
        ).fetchone()["identity_id"]
    trading_management.classify_trade_identities(
        [trade_id], subject["id"], "basic_hedging", "", "", "tester"
    )

    result = trading_management.remove_trade_assignment(trade_id, actor="tester")

    assert result["removed"] is True
    with db.connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS c FROM trading_business_assignments").fetchone()["c"] == 0
        assert conn.execute(
            "SELECT COUNT(*) AS c FROM operation_logs WHERE operation_type = '取消业务归类'"
        ).fetchone()["c"] == 1


def setup_classified_business_sample(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    confirmed = trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")
    trading_management.match_imported_facts(confirmed["batch_id"])
    subject = trading_management.create_business_subject("上海钧能", actor="tester")
    with db.connect() as conn:
        rb_ids = [
            row["identity_id"] for row in conn.execute(
                "SELECT identity_id FROM trading_trade_facts WHERE contract = 'rb2610' ORDER BY id"
            ).fetchall()
        ]
        conn.execute(
            """
            INSERT INTO trading_contract_specs
                (exchange, product_code, asset_type, contract_multiplier, price_tick)
            VALUES ('SHFE', 'rb', 'future', 10, 1)
            """
        )
    trading_management.classify_trade_identities(
        rb_ids, subject["id"], "basic_hedging", "代内部公司套保", "上海钧能钢材套保", "tester"
    )
    return confirmed


def test_default_business_allocations_preserve_fact_pnl(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)

    result = trading_management.rebuild_default_business_allocations(confirmed["batch_id"])

    assert result["allocation_count"] == 1
    with db.connect() as conn:
        allocation = conn.execute("SELECT * FROM trading_business_close_allocations").fetchone()
        fact_pnl = conn.execute("SELECT fact_close_pnl FROM trading_close_facts").fetchone()["fact_close_pnl"]
    assert allocation["source"] == "fact_default"
    assert allocation["business_pnl"] == 400
    assert fact_pnl == 1200


def test_business_views_separate_junneng_candidates_and_all_options(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)
    trading_management.rebuild_default_business_allocations(confirmed["batch_id"])

    junneng_closes = trading_management.query_business_rows(
        "junneng", "closes", trading_management.FactFilters(page=1, page_size=20)
    )
    junneng_trades = trading_management.query_business_rows(
        "junneng", "trades", trading_management.FactFilters(page=1, page_size=20)
    )
    option_trades = trading_management.query_business_rows(
        "options", "trades", trading_management.FactFilters(page=1, page_size=20)
    )

    assert junneng_closes["summary"]["business_pnl"] == 400
    assert junneng_closes["summary"]["fact_close_pnl"] == 1200
    assert len(junneng_trades["items"]) == 2
    assert all(item["business_subject"] == "上海钧能" for item in junneng_trades["items"])
    assert len(option_trades["items"]) == 1
    assert option_trades["items"][0]["contract"] == "i2609-c-700"
    assert option_trades["items"][0]["assignment_status"] == "unclassified"


def test_unclassified_rb_hc_appear_in_junneng_until_assigned_elsewhere(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    result = trading_management.query_business_rows(
        "junneng", "trades", trading_management.FactFilters(page=1, page_size=20)
    )

    assert len(result["items"]) == 2
    assert result["summary"]["record_count"] == 2
    assert all(row["assignment_status"] == "unclassified" for row in result["items"])
    assert all(row["ledger_membership"] == "candidate" for row in result["items"])
    assert result["candidates"]["record_count"] == 2


def test_fact_trade_classification_filter_returns_assignment_metadata(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)

    assigned = trading_management.query_fact_rows(
        "trades", trading_management.FactFilters(classification="classified", page=1, page_size=20)
    )
    pending = trading_management.query_fact_rows(
        "trades", trading_management.FactFilters(classification="unclassified", page=1, page_size=20)
    )

    assert assigned["items"]
    assert all(row["assignment_status"] == "classified" for row in assigned["items"])
    assert all(row["business_type"] == "basic_hedging" for row in assigned["items"])
    assert pending["items"]
    assert all(row["assignment_status"] == "unclassified" for row in pending["items"])


def test_option_business_positions_include_unclassified_open_options(tmp_path, monkeypatch):
    preview = create_preview_batch(tmp_path, monkeypatch)
    trading_management.confirm_trading_import(preview["preview_batch_id"], actor="tester")

    result = trading_management.query_business_rows(
        "options", "positions", trading_management.FactFilters(page=1, page_size=20)
    )

    assert result["summary"]["record_count"] == 1
    assert result["summary"]["quantity"] == 1
    assert result["items"][0]["contract"] == "i2609-c-700"
    assert result["items"][0]["average_price"] == 4.2
    assert result["items"][0]["floating_pnl_status"] == "pending_calculation"


def add_later_classified_open(batch_id, subject_id):
    with db.connect() as conn:
        source_row_id = conn.execute(
            """
            INSERT INTO trading_source_rows
                (batch_id, source_type, source_file, source_sheet, source_row_no, raw_hash, raw_json)
            VALUES (?, 'trade', 'manual-test.xlsx', '成交记录', 99, 'later-open', '{}')
            """,
            (batch_id,),
        ).lastrowid
        identity_id = conn.execute(
            """
            INSERT INTO trading_fact_identities (account_id, fact_type, stable_key)
            VALUES (1, 'trade', 'later-open-rb2610')
            """
        ).lastrowid
        conn.execute(
            """
            INSERT INTO trading_trade_facts
                (identity_id, batch_id, source_row_id, trade_date, exchange, contract, asset_type,
                 side, open_close_raw, open_close, quantity, price, fee, verification_status)
            VALUES (?, ?, ?, '20260625', 'SHFE', 'rb2610', 'future',
                    '买', '开', '开仓', 2, 3110, 6, 'file_imported')
            """,
            (identity_id, batch_id, source_row_id),
        )
        strategy_id = conn.execute(
            """
            INSERT INTO trading_strategies (name, normalized_name, source, created_by)
            VALUES ('晚开策略', '晚开策略', 'manual', 'tester')
            """
        ).lastrowid
        conn.execute(
            """
            INSERT INTO trading_business_assignments
                (trade_identity_id, business_subject_id, business_type, strategy_id, assigned_by)
            VALUES (?, ?, 'basic_hedging', ?, 'tester')
            """,
            (identity_id, subject_id, strategy_id),
        )
    return identity_id


def test_manual_rematch_moves_closed_and_open_business_quantities_atomically(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)
    trading_management.rebuild_default_business_allocations(confirmed["batch_id"])
    with db.connect() as conn:
        subject_id = conn.execute(
            "SELECT id FROM trading_business_subjects WHERE name = '上海钧能'"
        ).fetchone()["id"]
        close_id = conn.execute("SELECT identity_id FROM trading_close_facts").fetchone()["identity_id"]
        old_open_id = conn.execute(
            "SELECT open_trade_identity_id FROM trading_business_close_allocations"
        ).fetchone()["open_trade_identity_id"]
    new_open_id = add_later_classified_open(confirmed["batch_id"], subject_id)

    candidates = trading_management.list_business_close_candidates(close_id)
    preview = trading_management.preview_business_rematch(
        close_id,
        [{"open_trade_identity_id": new_open_id, "quantity": 2}],
        allocation_version=1,
    )
    result = trading_management.confirm_business_rematch(
        close_id, preview["preview_token"], allocation_version=1, actor="tester", reason="实际平晚开仓"
    )

    assert {row["identity_id"] for row in candidates} >= {old_open_id, new_open_id}
    assert preview["before_business_pnl"] == 400
    assert preview["after_business_pnl"] == 200
    assert result["allocation_version"] == 2
    with db.connect() as conn:
        allocation = conn.execute("SELECT * FROM trading_business_close_allocations").fetchone()
        fact_allocation = conn.execute("SELECT * FROM trading_fact_close_allocations").fetchone()
        audit = conn.execute("SELECT * FROM trading_business_allocation_audit").fetchone()
        close_assignment = conn.execute(
            """
            SELECT ba.*, s.name AS strategy
            FROM trading_close_trade_links l
            JOIN trading_business_assignments ba ON ba.trade_identity_id = l.close_trade_identity_id
            LEFT JOIN trading_strategies s ON s.id = ba.strategy_id
            WHERE l.close_identity_id = ?
            """,
            (close_id,),
        ).fetchone()
    assert allocation["open_trade_identity_id"] == new_open_id
    assert allocation["source"] == "manual_override"
    assert fact_allocation["open_trade_identity_id"] == old_open_id
    assert audit["before_business_pnl"] == 400
    assert audit["after_business_pnl"] == 200
    assert close_assignment["strategy"] == "晚开策略"

    positions = trading_management.query_business_rows(
        "junneng", "positions", trading_management.FactFilters(page=1, page_size=20)
    )
    assert positions["summary"]["quantity"] == 2
    assert positions["items"][0]["identity_id"] == old_open_id


def test_manual_rematch_rejects_stale_version_and_cross_contract(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)
    trading_management.rebuild_default_business_allocations(confirmed["batch_id"])
    with db.connect() as conn:
        close_id = conn.execute("SELECT identity_id FROM trading_close_facts").fetchone()["identity_id"]
        option_id = conn.execute(
            "SELECT identity_id FROM trading_trade_facts WHERE asset_type = 'option'"
        ).fetchone()["identity_id"]

    with pytest.raises(ValueError, match="同账户、同合约"):
        trading_management.preview_business_rematch(
            close_id, [{"open_trade_identity_id": option_id, "quantity": 2}], allocation_version=1
        )
    with pytest.raises(ValueError, match="数据已变化"):
        trading_management.preview_business_rematch(
            close_id, [], allocation_version=99
        )


def test_restore_default_business_match_reverts_allocation_and_close_inheritance(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)
    trading_management.rebuild_default_business_allocations(confirmed["batch_id"])
    with db.connect() as conn:
        subject_id = conn.execute(
            "SELECT id FROM trading_business_subjects WHERE name = '上海钧能'"
        ).fetchone()["id"]
        close_id = conn.execute("SELECT identity_id FROM trading_close_facts").fetchone()["identity_id"]
        fact_open_id = conn.execute(
            "SELECT open_trade_identity_id FROM trading_fact_close_allocations"
        ).fetchone()["open_trade_identity_id"]
    new_open_id = add_later_classified_open(confirmed["batch_id"], subject_id)
    preview = trading_management.preview_business_rematch(
        close_id, [{"open_trade_identity_id": new_open_id, "quantity": 2}], allocation_version=1
    )
    trading_management.confirm_business_rematch(
        close_id, preview["preview_token"], allocation_version=1, actor="tester"
    )

    restored = trading_management.restore_default_business_match(close_id, allocation_version=2, actor="tester")

    assert restored["allocation_version"] == 3
    assert restored["business_pnl"] == 400
    with db.connect() as conn:
        allocation = conn.execute("SELECT * FROM trading_business_close_allocations").fetchone()
        close_assignment = conn.execute(
            """
            SELECT ba.strategy_id
            FROM trading_close_trade_links l
            JOIN trading_business_assignments ba ON ba.trade_identity_id = l.close_trade_identity_id
            WHERE l.close_identity_id = ?
            """,
            (close_id,),
        ).fetchone()
        open_assignment = conn.execute(
            "SELECT strategy_id FROM trading_business_assignments WHERE trade_identity_id = ?",
            (fact_open_id,),
        ).fetchone()
    assert allocation["open_trade_identity_id"] == fact_open_id
    assert allocation["source"] == "fact_default"
    assert close_assignment["strategy_id"] == open_assignment["strategy_id"]


def test_fully_closed_contract_reports_business_pnl_reconciliation_difference(tmp_path, monkeypatch):
    confirmed = setup_classified_business_sample(tmp_path, monkeypatch)
    trading_management.rebuild_default_business_allocations(confirmed["batch_id"])
    with db.connect() as conn:
        close_id = conn.execute("SELECT identity_id FROM trading_close_facts").fetchone()["identity_id"]

    failed = trading_management.reconcile_business_pnl(close_id)
    with db.connect() as conn:
        conn.execute("UPDATE trading_close_facts SET fact_close_pnl = 400")
    reconciled = trading_management.reconcile_business_pnl(close_id)

    assert failed["status"] == "business_pnl_reconciliation_failed"
    assert failed["difference"] == -800
    assert reconciled["status"] == "reconciled"
    assert reconciled["difference"] == 0


def test_trading_management_exposes_complete_p0_api_surface():
    routes = {(route.path, method) for route in trading_management.router.routes for method in route.methods}

    expected = {
        ("/imports/preview", "POST"),
        ("/imports/{preview_batch_id}/confirm", "POST"),
        ("/config", "GET"),
        ("/business-assignments/batch-confirm", "POST"),
        ("/business/junneng/{tab}", "GET"),
        ("/business/options/{tab}", "GET"),
        ("/business-closes/{close_identity_id}/candidates", "GET"),
        ("/business-closes/{close_identity_id}/preview", "POST"),
        ("/business-closes/{close_identity_id}/confirm", "POST"),
        ("/business-closes/{close_identity_id}/restore-default", "POST"),
    }
    assert expected <= routes
