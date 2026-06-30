"""数据可视化管理 — 单元测试"""
import sys, os
# Make backend/app a package by adding backend/ to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
from app.data_visualization import (
    MAINSTREAM_PRODUCT_ORDER,
    MAINSTREAM_PRODUCTS,
    _mainstream_status,
    compute_business_week,
    _parse_integrated_excel,
    _import_integrated_points,
    _save_integrated_points,
    _load_integrated_preview_cache,
    _save_integrated_preview_cache,
    _split_filter_values,
    _to_date,
    _to_float,
    build_integrated_workbook_bytes,
    get_filters,
    get_chart,
    get_table,
    integrate_mysteel_files,
    _local_mysteel_files,
)

from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from io import BytesIO
import asyncio


def _insert_integrated_point_direct(db_module, point):
    with db_module.connect() as conn:
        cur = conn.cursor()
        batch_id = db_module._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'completed', 1, 0, '{}', ?)""",
            (point.get("source_file", "old.xlsx"), "pytest"),
        )
        db_module._exec(
            cur,
            """INSERT INTO dv_integrated_points
               (batch_id, week_start, week_end, business_year, business_week, week_label,
                display_date, metric_type, source_country, product,
                category, mainstream_status, value, unit, source_file,
                source_sheet, source_section, is_calculable, validation_status, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                batch_id,
                point["week_start"],
                point["week_end"],
                point["business_year"],
                point["business_week"],
                point["week_label"],
                point["display_date"],
                point["metric_type"],
                point["source_country"],
                point["product"],
                point["category"],
                point["mainstream_status"],
                point["value"],
                point.get("unit", "万吨"),
                point.get("source_file", "old.xlsx"),
                point.get("source_sheet", "历史"),
                point.get("source_section", "历史"),
                point.get("is_calculable", 0),
                point.get("validation_status", "ok"),
                point.get("note", ""),
            ),
        )


def test_compute_business_week_w01():
    bw = compute_business_week(date(2025, 1, 1))
    assert bw["year"] == 2025, f"year expected 2025 got {bw['year']}"
    assert bw["week_no"] == 1, f"week_no expected 1 got {bw['week_no']}"

def test_compute_business_week_w02():
    bw = compute_business_week(date(2025, 1, 6))
    assert bw["year"] == 2025
    assert bw["week_no"] == 2

def test_compute_business_week_dec31():
    bw = compute_business_week(date(2024, 12, 31))
    assert bw["year"] == 2025
    assert bw["week_no"] == 1

def test_compute_business_week_dec30():
    bw = compute_business_week(date(2024, 12, 30))
    assert bw["year"] == 2025
    assert bw["week_no"] == 1

def test_to_date():
    import datetime as dtmod
    assert _to_date(dtmod.datetime(2025, 6, 15)) == date(2025, 6, 15)
    assert _to_date(date(2025, 6, 15)) == date(2025, 6, 15)
    assert _to_date("2025-06-15") is None

def test_to_float():
    assert _to_float(3.14) == 3.14
    assert _to_float(None) is None
    assert _to_float("abc") is None


def test_split_filter_values_supports_multi_select():
    assert _split_filter_values("主流,非主流") == ["主流", "非主流"]
    assert _split_filter_values(" 主流 , , 非主流 ") == ["主流", "非主流"]
    assert _split_filter_values("") == []


EXPECTED_MAINSTREAM_PRODUCT_ORDER = [
    "PB粉", "麦克粉", "纽曼粉", "金布巴粉", "超特粉", "混合粉", "卡粉",
    "巴混", "SP10粉", "几内亚粉", "杨迪粉", "罗伊山粉",
    "罗伊山MB粉", "印度（粉矿）", "IOC6", "罗伊山MB块", "PMI块",
    "印度（球团）", "乌克兰（精粉）", "卡拉拉精粉",
]


def test_mainstream_products_follow_business_definition():
    assert MAINSTREAM_PRODUCT_ORDER == EXPECTED_MAINSTREAM_PRODUCT_ORDER
    assert MAINSTREAM_PRODUCTS == {
        "PB粉", "麦克粉", "纽曼粉", "金布巴粉", "超特粉", "混合粉", "卡粉",
        "巴混", "SP10粉", "几内亚粉", "杨迪粉", "罗伊山粉",
        "罗伊山MB粉", "印度", "IOC6", "罗伊山MB块", "PMI块", "乌克兰", "卡拉拉精粉",
    }
    for skipped in ["RTX", "RTX块", "高锰粉矿", "未知", "库宾粉", "PB块", "纽曼块"]:
        assert skipped not in MAINSTREAM_PRODUCTS
    assert _mainstream_status("印度", "粉矿") == "主流"
    assert _mainstream_status("印度", "球团") == "主流"
    assert _mainstream_status("印度", "精粉") == "非主流"
    assert _mainstream_status("乌克兰", "精粉") == "主流"
    assert _mainstream_status("乌克兰", "球团") == "非主流"


def test_parse_integrated_excel_reports_invalid_rows(tmp_path):
    path = tmp_path / "bad_integrated.xlsx"
    headers = [
        "统计周一", "统计周日", "业务年份", "业务周次", "周次标签", "展示日期",
        "数据类型", "来源/国家", "品种", "种类", "主流/非主流", "数值",
        "单位", "来源文件", "来源Sheet", "来源区域", "是否参与表需", "校验状态", "备注",
    ]
    row = [
        "2026-06-01", "2026-06-07", 2026, 23, "2026 W23", "2026-06-02",
        "库存", "澳洲", "PB粉", "粉矿", "主流", 100,
        "万吨", "file.xlsx", "sheet", "区域", "是", "ok", "",
    ]
    invalid_metric = list(row)
    invalid_metric[6] = "非法类型"
    invalid_value = list(row)
    invalid_value[8] = "纽曼粉"
    invalid_value[11] = "abc"

    wb = Workbook()
    ws = wb.active
    ws.title = "整合明细"
    ws.append(headers)
    ws.append(row)
    ws.append(list(row))
    ws.append(invalid_metric)
    ws.append(invalid_value)
    wb.save(path)

    result = _parse_integrated_excel(path)
    messages = [item["message"] for item in result["errors"]]
    assert result["summary"]["duplicate_key_count"] == 1
    assert any("数据类型无效" in message for message in messages)
    assert any("数值不是数字" in message for message in messages)
    assert len(result["rows"]) == 2


def test_parse_integrated_excel_accepts_arrival_metric(tmp_path):
    path = tmp_path / "arrival_integrated.xlsx"
    headers = [
        "统计周一", "统计周日", "业务年份", "业务周次", "周次标签", "展示日期",
        "数据类型", "来源/国家", "品种", "种类", "主流/非主流", "数值",
        "单位", "来源文件", "来源Sheet", "来源区域", "是否参与表需", "校验状态", "备注",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "整合明细"
    ws.append(headers)
    ws.append([
        "2026-06-01", "2026-06-07", 2026, 23, "2026 W23", "2026-06-02",
        "到港", "澳洲", "PB粉", "粉矿", "主流", 100,
        "万吨", "file.xlsx", "sheet", "区域", "是", "ok", "",
    ])
    wb.save(path)

    result = _parse_integrated_excel(path)

    assert result["errors"] == []
    assert result["rows"][0]["metric_type"] == "arrival"
    assert result["summary"]["arrival_count"] == 1


def test_parse_integrated_excel_normalizes_excel_datetimes(tmp_path):
    path = tmp_path / "datetime_integrated.xlsx"
    headers = [
        "统计周一", "统计周日", "业务年份", "业务周次", "周次标签", "展示日期",
        "数据类型", "来源/国家", "品种", "种类", "主流/非主流", "数值",
        "单位", "来源文件", "来源Sheet", "来源区域", "是否参与表需", "校验状态", "备注",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "整合明细"
    ws.append(headers)
    ws.append([
        datetime(2026, 6, 1), datetime(2026, 6, 7), 2026, 23, "2026 W23", datetime(2026, 6, 2),
        "库存", "澳洲", "PB粉", "粉矿", "主流", 100,
        "万吨", "file.xlsx", "sheet", "区域", "是", "ok", "",
    ])
    wb.save(path)

    result = _parse_integrated_excel(path)

    assert result["errors"] == []
    assert result["rows"][0]["week_start"] == "2026-06-01"
    assert result["rows"][0]["week_end"] == "2026-06-07"
    assert result["rows"][0]["display_date"] == "2026-06-02"


def test_integrated_import_replaces_previous_batch(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()
    rows = [
        {
            "week_start": "2026-06-01",
            "week_end": "2026-06-07",
            "business_year": 2026,
            "business_week": 23,
            "week_label": "2026 W23",
            "display_date": "2026-06-02",
            "metric_type": "inventory",
            "source_country": "澳洲",
            "product": "PB粉",
            "category": "粉矿",
            "mainstream_status": "主流",
            "value": 100.0,
            "unit": "万吨",
            "source_file": "first.xlsx",
            "source_sheet": "整合明细",
            "source_section": "测试",
            "is_calculable": 1,
            "validation_status": "ok",
            "note": "",
        }
    ]
    _import_integrated_points(rows, "first.xlsx", "pytest")
    second_rows = [dict(rows[0], value=120.0, source_file="second.xlsx")]

    _import_integrated_points(second_rows, "second.xlsx", "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        point_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integrated_points").fetchone()["c"]
        batch_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integration_batches").fetchone()["c"]
        point = db._exec(cur, "SELECT value, source_file FROM dv_integrated_points").fetchone()

    assert point_count == 1
    assert batch_count == 1
    assert point["value"] == 120.0
    assert point["source_file"] == "second.xlsx"


def test_save_integrated_points_merges_weekly_uploads(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "澳洲",
        "product": "PB粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "history.xlsx",
        "source_sheet": "库存",
        "source_section": "历史",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([base_point], ["history.xlsx"], "pytest")

    changed_duplicate = dict(base_point, value=120.0, source_file="week1.xlsx", source_section="周度修正")
    blank_duplicate = dict(base_point, value=None, source_file="blank.xlsx", source_section="空值")
    new_week = dict(base_point, business_week=24, week_start="2026-06-08", week_end="2026-06-14",
                    week_label="2026 W24", display_date="2026-06-09", value=130.0,
                    source_file="week1.xlsx")
    batch_id = _save_integrated_points([changed_duplicate, blank_duplicate, new_week], ["week1.xlsx"], "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        points = db._exec(
            cur,
            """SELECT business_week, value, source_file, source_section
               FROM dv_integrated_points ORDER BY business_week""",
        ).fetchall()
        batch = db._exec(
            cur,
            "SELECT point_count, validation_summary FROM dv_integration_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

    summary = __import__("json").loads(batch["validation_summary"])
    assert len(points) == 2
    assert points[0]["business_week"] == 23
    assert points[0]["value"] == 120.0
    assert points[0]["source_file"] == "week1.xlsx"
    assert points[1]["business_week"] == 24
    assert points[1]["value"] == 130.0
    assert batch["point_count"] == 2
    assert summary["inserted"] == 1
    assert summary["updated"] == 1
    assert summary["skipped"] == 1
    assert summary["skipped_blank_overwrite"] == 1


def test_save_integrated_points_migrates_legacy_pellet_inventory_names(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    legacy_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "乌克兰",
        "product": "乌克兰",
        "category": "球团",
        "mainstream_status": "非主流",
        "value": 0.000015,
        "unit": "万吨",
        "source_file": "old_mysteel.xlsx",
        "source_sheet": "球团",
        "source_section": "总计行",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    with db.connect() as conn:
        cur = conn.cursor()
        old_batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'completed', 1, 0, '{}', ?)""",
            ("old_mysteel.xlsx", "pytest"),
        )
        db._exec(
            cur,
            """INSERT INTO dv_integrated_points
               (batch_id, week_start, week_end, business_year, business_week, week_label,
                display_date, metric_type, source_country, product,
                category, mainstream_status, value, unit, source_file,
                source_sheet, source_section, is_calculable, validation_status, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                old_batch_id,
                legacy_point["week_start"],
                legacy_point["week_end"],
                legacy_point["business_year"],
                legacy_point["business_week"],
                legacy_point["week_label"],
                legacy_point["display_date"],
                legacy_point["metric_type"],
                legacy_point["source_country"],
                "乌克兰球",
                legacy_point["category"],
                "主流",
                legacy_point["value"],
                legacy_point["unit"],
                legacy_point["source_file"],
                legacy_point["source_sheet"],
                legacy_point["source_section"],
                legacy_point["is_calculable"],
                legacy_point["validation_status"],
                legacy_point["note"],
            ),
        )
        conn.commit()

    corrected_point = dict(
        legacy_point,
        product="乌克兰",
        mainstream_status="非主流",
        value=0.0,
        source_file="new_mysteel.xlsx",
    )
    batch_id = _save_integrated_points([corrected_point], ["new_mysteel.xlsx"], "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        points = db._exec(
            cur,
            """SELECT source_country, product, mainstream_status, value, source_file
               FROM dv_integrated_points
               WHERE metric_type = 'inventory'
                 AND source_country = '乌克兰'
                 AND category = '球团'
                 AND business_year = 2026
                 AND business_week = 23""",
        ).fetchall()
        batch = db._exec(
            cur,
            "SELECT validation_summary FROM dv_integration_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

    summary = __import__("json").loads(batch["validation_summary"])
    assert len(points) == 1
    assert points[0]["product"] == "乌克兰"
    assert points[0]["mainstream_status"] == "非主流"
    assert points[0]["value"] == 0.0
    assert points[0]["source_file"] == "new_mysteel.xlsx"
    assert summary["updated"] == 1
    assert summary["inserted"] == 0


def test_save_integrated_points_migrates_changed_mainstream_status(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "arrival",
        "source_country": "澳洲",
        "product": "SP10粉",
        "category": "粉矿",
        "mainstream_status": "非主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "old_mysteel.xlsx",
        "source_sheet": "澳洲预计到达中国锚地量",
        "source_section": "历史",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    with db.connect() as conn:
        cur = conn.cursor()
        old_batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'completed', 1, 0, '{}', ?)""",
            ("old_mysteel.xlsx", "pytest"),
        )
        db._exec(
            cur,
            """INSERT INTO dv_integrated_points
               (batch_id, week_start, week_end, business_year, business_week, week_label,
                display_date, metric_type, source_country, product,
                category, mainstream_status, value, unit, source_file,
                source_sheet, source_section, is_calculable, validation_status, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                old_batch_id,
                point["week_start"],
                point["week_end"],
                point["business_year"],
                point["business_week"],
                point["week_label"],
                point["display_date"],
                point["metric_type"],
                point["source_country"],
                point["product"],
                point["category"],
                point["mainstream_status"],
                point["value"],
                point["unit"],
                point["source_file"],
                point["source_sheet"],
                point["source_section"],
                point["is_calculable"],
                point["validation_status"],
                point["note"],
            ),
        )
        conn.commit()

    batch_id = _save_integrated_points([
        dict(point, mainstream_status="主流", value=120.0, source_file="new_mysteel.xlsx"),
    ], ["new_mysteel.xlsx"], "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        points = db._exec(
            cur,
            """SELECT mainstream_status, value, source_file
               FROM dv_integrated_points
               WHERE metric_type = 'arrival'
                 AND source_country = '澳洲'
                 AND product = 'SP10粉'
                 AND category = '粉矿'
                 AND business_year = 2026
                 AND business_week = 23""",
        ).fetchall()
        batch = db._exec(
            cur,
            "SELECT validation_summary FROM dv_integration_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

    summary = __import__("json").loads(batch["validation_summary"])
    assert len(points) == 1
    assert points[0]["mainstream_status"] == "主流"
    assert points[0]["value"] == 120.0
    assert points[0]["source_file"] == "new_mysteel.xlsx"
    assert summary["updated"] == 1
    assert summary["inserted"] == 0


def test_save_integrated_points_recalculates_apparent_demand_from_full_history(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "澳洲",
        "product": "PB粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "history.xlsx",
        "source_sheet": "库存",
        "source_section": "历史",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, week_start="2026-06-08", week_end="2026-06-14",
             business_week=24, week_label="2026 W24", display_date="2026-06-09", value=90.0),
    ], ["history.xlsx"], "pytest")
    _save_integrated_points([
        dict(base_point, week_start="2026-06-08", week_end="2026-06-14",
             business_week=24, week_label="2026 W24", display_date="2026-06-08",
             metric_type="arrival", value=20.0, source_file="mysteel.xlsx",
             source_sheet="澳洲预计到达中国锚地量", source_section="到港"),
    ], ["mysteel.xlsx"], "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        apparent = db._exec(
            cur,
            """SELECT value, source_file, source_sheet
               FROM dv_integrated_points
               WHERE metric_type = 'apparent_demand'
                 AND product = 'PB粉'
                 AND business_year = 2026
                 AND business_week = 24""",
        ).fetchone()

    assert apparent is not None
    assert apparent["value"] == 30.0
    assert apparent["source_file"] == "系统计算"
    assert apparent["source_sheet"] == "表需"


def test_get_table_groups_same_business_week_with_mixed_date_formats(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    point = {
        "week_start": "2026-06-01T00:00:00",
        "week_end": "2026-06-07T00:00:00",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-01T00:00:00",
        "metric_type": "arrival",
        "source_country": "巴西",
        "product": "卡粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "history.xlsx",
        "source_sheet": "整合明细",
        "source_section": "历史",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([point], ["history.xlsx"], "pytest")
    _save_integrated_points([
        dict(point, week_start="2026-06-01", week_end="2026-06-07",
             display_date="2026-06-01", source_country="澳洲", product="PB粉",
             value=200.0, source_file="mysteel.xlsx"),
    ], ["mysteel.xlsx"], "pytest")

    result = asyncio.run(get_table(metric="arrival", years="2026", mainstream_status="主流", user={"role": "管理员"}))

    assert len(result["data"]) == 1
    assert result["data"][0]["week"] == "2026 W23"
    assert result["data"][0]["卡粉"]["value"] == 100.0
    assert result["data"][0]["PB粉"]["value"] == 200.0


def test_filters_reflect_mainstream_status_written_by_integration(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "arrival",
        "source_country": "澳洲",
        "product": "SP10粉",
        "category": "粉矿",
        "mainstream_status": "非主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "mysteel.xlsx",
        "source_sheet": "澳洲预计到达中国锚地量",
        "source_section": "测试",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, source_country="巴西", product="巴混", category="粉矿",
             metric_type="inventory", is_calculable=0, value=40.0),
    ], ["mysteel.xlsx"], "pytest")

    filters = asyncio.run(get_filters(user={"role": "管理员"}))

    assert "SP10粉" in filters["product_pools"]["mainstream"]
    assert "SP10粉" not in filters["product_pools"]["non_mainstream"]
    assert "巴混" in filters["product_pools"]["mainstream"]
    assert "巴混" not in filters["product_pools"]["non_mainstream"]


def test_mainstream_pool_filters_and_orders_products_by_business_definition(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "shipment",
        "source_country": "澳洲",
        "product": "罗伊山粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 10.0,
        "unit": "万吨",
        "source_file": "mysteel.xlsx",
        "source_sheet": "澳巴发运",
        "source_section": "测试",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, product="PB粉", value=20.0),
        dict(base_point, product="巴混", source_country="巴西", value=30.0),
        dict(base_point, product="金布巴粉", value=40.0),
        dict(base_point, product="罗伊山MB粉", value=41.0),
        dict(base_point, source_country="印度", product="印度", category="粉矿", value=42.0),
        dict(base_point, source_country="巴西", product="IOC6", category="粉矿", value=43.0),
        dict(base_point, product="罗伊山MB块", category="块矿", value=44.0),
        dict(base_point, product="PMI块", category="块矿", value=45.0),
        dict(base_point, source_country="印度", product="印度", category="球团", value=46.0),
        dict(base_point, source_country="乌克兰", product="乌克兰", category="精粉", value=47.0),
        dict(base_point, product="卡拉拉精粉", category="精粉", value=48.0),
        dict(base_point, product="南非", source_country="南非", category="全品种",
             mainstream_status="非主流", value=50.0),
    ], ["mysteel.xlsx"], "pytest")
    with db.connect() as conn:
        cur = conn.cursor()
        stale_batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'completed', 1, 0, '{}', ?)""",
            ("old_mysteel.xlsx", "pytest"),
        )
        db._exec(
            cur,
            """INSERT INTO dv_integrated_points
               (batch_id, week_start, week_end, business_year, business_week, week_label,
                display_date, metric_type, source_country, product,
                category, mainstream_status, value, unit, source_file,
                source_sheet, source_section, is_calculable, validation_status, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                stale_batch_id,
                base_point["week_start"],
                base_point["week_end"],
                base_point["business_year"],
                base_point["business_week"],
                base_point["week_label"],
                base_point["display_date"],
                base_point["metric_type"],
                "澳洲",
                "PB块",
                "块矿",
                "主流",
                60.0,
                base_point["unit"],
                "old_mysteel.xlsx",
                base_point["source_sheet"],
                base_point["source_section"],
                base_point["is_calculable"],
                base_point["validation_status"],
                base_point["note"],
            ),
        )

    filters = asyncio.run(get_filters(user={"role": "管理员"}))
    table = asyncio.run(get_table(
        metric="shipment",
        years="2026",
        product_pool="mainstream",
        user={"role": "管理员"},
    ))
    chart = asyncio.run(get_chart(
        metric="shipment",
        years="2026",
        product_pool="mainstream",
        user={"role": "管理员"},
    ))

    expected_products = [
        "PB粉", "金布巴粉", "巴混", "罗伊山粉", "罗伊山MB粉", "印度（粉矿）",
        "IOC6", "罗伊山MB块", "PMI块", "印度（球团）", "乌克兰（精粉）", "卡拉拉精粉",
    ]
    assert filters["product_pools"]["mainstream"] == expected_products
    assert table["products"] == expected_products
    assert "PB块" not in table["products"]
    assert "南非（全品种）" not in table["products"]
    assert list(chart["series"]) == expected_products


def test_syncs_historical_mainstream_labels_for_db_aggregates_and_export(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "shipment",
        "source_country": "巴西",
        "product": "巴混",
        "category": "粉矿",
        "mainstream_status": "非主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "old.xlsx",
        "source_sheet": "历史",
        "source_section": "历史",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _insert_integrated_point_direct(db, base_point)
    _insert_integrated_point_direct(db, dict(base_point, mainstream_status="主流", value=120.0, source_file="new.xlsx"))
    _insert_integrated_point_direct(db, dict(
        base_point,
        source_country="澳洲",
        product="PB块",
        category="块矿",
        mainstream_status="主流",
        value=70.0,
        source_file="old_blocks.xlsx",
    ))

    filters = asyncio.run(get_filters(user={"role": "管理员"}))
    mainstream_table = asyncio.run(get_table(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="主流矿合计",
        user={"role": "管理员"},
    ))
    non_mainstream_table = asyncio.run(get_table(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="非主流矿合计",
        user={"role": "管理员"},
    ))
    workbook_bytes = build_integrated_workbook_bytes()
    wb = load_workbook(BytesIO(workbook_bytes), read_only=True)
    ws = wb["整合明细"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    with db.connect() as conn:
        cur = conn.cursor()
        points = [
            dict(row) for row in db._exec(
                cur,
                """SELECT product, category, mainstream_status, value
                   FROM dv_integrated_points
                   WHERE metric_type = 'shipment'
                   ORDER BY product, category, mainstream_status""",
            ).fetchall()
        ]

    assert "巴混" in filters["product_pools"]["mainstream"]
    assert "PB块" not in filters["product_pools"]["mainstream"]
    assert points == [
        {"product": "PB块", "category": "块矿", "mainstream_status": "非主流", "value": 70.0},
        {"product": "巴混", "category": "粉矿", "mainstream_status": "主流", "value": 120.0},
    ]
    assert mainstream_table["data"][0]["主流矿合计"]["value"] == 120.0
    assert non_mainstream_table["data"][0]["非主流矿合计"]["value"] == 70.0

    header = rows[0]
    product_idx = header.index("品种")
    status_idx = header.index("主流/非主流")
    exported = {
        row[product_idx]: row[status_idx]
        for row in rows[1:]
        if row[product_idx] in {"巴混", "PB块"}
    }
    assert exported == {"巴混": "主流", "PB块": "非主流"}


def test_mainstream_status_does_not_make_product_apparent_demand_calculable(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "澳洲",
        "product": "PMI块",
        "category": "块矿",
        "mainstream_status": "非主流",
        "value": 12.0,
        "unit": "万吨",
        "source_file": "inventory.xlsx",
        "source_sheet": "块矿",
        "source_section": "总计行",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([point], ["inventory.xlsx"], "pytest")

    with db.connect() as conn:
        cur = conn.cursor()
        rows = [
            dict(row) for row in db._exec(
                cur,
                """SELECT product, category, mainstream_status, is_calculable
                   FROM dv_integrated_points
                   ORDER BY id""",
            ).fetchall()
        ]

    assert rows == [{
        "product": "PMI块",
        "category": "块矿",
        "mainstream_status": "主流",
        "is_calculable": 0,
    }]


def test_aggregate_product_filter_applies_to_table_and_chart(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "shipment",
        "source_country": "澳洲",
        "product": "PB粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "mysteel.xlsx",
        "source_sheet": "澳洲发货量",
        "source_section": "测试",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, source_country="南非", product="南非", category="全品种",
             mainstream_status="非主流", value=40.0),
    ], ["mysteel.xlsx"], "pytest")

    table = asyncio.run(get_table(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="主流矿合计",
        user={"role": "管理员"},
    ))
    chart = asyncio.run(get_chart(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="主流矿合计",
        user={"role": "管理员"},
    ))

    assert table["products"] == ["主流矿合计"]
    assert table["data"][0]["主流矿合计"]["value"] == 100.0
    assert "非主流矿合计" not in table["data"][0]
    assert sorted(chart["series"]) == ["主流矿合计"]


def test_aggregate_product_filter_ignores_conflicting_mainstream_status(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "shipment",
        "source_country": "澳洲",
        "product": "PB粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "mysteel.xlsx",
        "source_sheet": "澳洲发货量",
        "source_section": "测试",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, source_country="南非", product="南非", category="全品种",
             mainstream_status="非主流", value=40.0),
    ], ["mysteel.xlsx"], "pytest")

    table = asyncio.run(get_table(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="主流矿合计",
        mainstream_status="非主流",
        user={"role": "管理员"},
    ))
    chart = asyncio.run(get_chart(
        metric="shipment",
        years="2026",
        product_pool="aggregate",
        products="主流矿合计",
        mainstream_status="非主流",
        user={"role": "管理员"},
    ))

    assert table["products"] == ["主流矿合计"]
    assert table["data"][0]["主流矿合计"]["value"] == 100.0
    assert sorted(chart["series"]) == ["主流矿合计"]


def test_get_table_and_chart_disambiguate_same_product_category_across_sources(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "澳洲",
        "product": "其他",
        "category": "粉矿",
        "mainstream_status": "非主流",
        "value": 12.0,
        "unit": "万吨",
        "source_file": "mysteel.xlsx",
        "source_sheet": "粗粉",
        "source_section": "总计行",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, source_country="巴西", value=34.0),
        dict(base_point, source_country="其他", value=56.0),
    ], ["mysteel.xlsx"], "pytest")

    table = asyncio.run(get_table(metric="inventory", years="2026", products="其他", categories="粉矿", user={"role": "管理员"}))
    row = table["data"][0]

    assert table["products"] == ["澳洲（其他粉矿）", "巴西（其他粉矿）", "其他（粉矿）"]
    assert row["澳洲（其他粉矿）"]["value"] == 12.0
    assert row["巴西（其他粉矿）"]["value"] == 34.0
    assert row["其他（粉矿）"]["value"] == 56.0

    chart = asyncio.run(get_chart(metric="inventory", years="2026", products="其他", categories="粉矿", user={"role": "管理员"}))

    assert sorted(chart["series"]) == ["其他（粉矿）", "巴西（其他粉矿）", "澳洲（其他粉矿）"]


def test_country_level_shipments_display_as_all_products(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "shipment",
        "source_country": "俄罗斯",
        "product": "俄罗斯",
        "category": "全品种",
        "mainstream_status": "非主流",
        "value": 88.0,
        "unit": "万吨",
        "source_file": "global.xlsx",
        "source_sheet": "全球铁矿石发运量",
        "source_section": "铁矿石全球发运量",
        "is_calculable": 0,
        "validation_status": "record_only",
        "note": "",
    }
    _save_integrated_points([point], ["global.xlsx"], "pytest")

    table = asyncio.run(get_table(metric="shipment", years="2026", user={"role": "管理员"}))
    chart = asyncio.run(get_chart(metric="shipment", years="2026", user={"role": "管理员"}))

    assert table["products"] == ["俄罗斯（全品种）"]
    assert table["data"][0]["俄罗斯（全品种）"]["value"] == 88.0
    assert sorted(chart["series"]) == ["俄罗斯（全品种）"]


def test_filters_use_display_labels_for_country_level_non_mainstream_products(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "俄罗斯",
        "product": "俄罗斯",
        "category": "球团",
        "mainstream_status": "非主流",
        "value": 11.0,
        "unit": "万吨",
        "source_file": "inventory.xlsx",
        "source_sheet": "球团",
        "source_section": "总计行",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, category="精粉", source_sheet="精粉", value=12.0),
        dict(base_point, metric_type="shipment", category="全品种",
             source_file="global.xlsx", source_sheet="全球铁矿石发运量", value=13.0),
    ], ["inventory.xlsx", "global.xlsx"], "pytest")

    filters = asyncio.run(get_filters(user={"role": "管理员"}))

    assert filters["source_countries"] == ["俄罗斯"]
    assert set(filters["categories"]) == {"全品种", "球团", "精粉"}
    assert "俄罗斯" not in filters["product_pools"]["non_mainstream"]
    assert set(filters["product_pools"]["non_mainstream"]) == {
        "俄罗斯（全品种）",
        "俄罗斯（球团）",
        "俄罗斯（精粉）",
    }


def test_display_label_product_filter_maps_back_to_integrated_identity(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    base_point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "俄罗斯",
        "product": "俄罗斯",
        "category": "球团",
        "mainstream_status": "非主流",
        "value": 11.0,
        "unit": "万吨",
        "source_file": "inventory.xlsx",
        "source_sheet": "球团",
        "source_section": "总计行",
        "is_calculable": 0,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([
        base_point,
        dict(base_point, category="精粉", source_sheet="精粉", value=12.0),
    ], ["inventory.xlsx"], "pytest")

    table = asyncio.run(get_table(
        metric="inventory",
        years="2026",
        products="俄罗斯（精粉）",
        user={"role": "管理员"},
    ))
    chart = asyncio.run(get_chart(
        metric="inventory",
        years="2026",
        products="俄罗斯（精粉）",
        user={"role": "管理员"},
    ))

    assert table["products"] == ["俄罗斯（精粉）"]
    assert table["data"][0]["俄罗斯（精粉）"]["value"] == 12.0
    assert "俄罗斯（球团）" not in table["data"][0]
    assert sorted(chart["series"]) == ["俄罗斯（精粉）"]


def test_integrated_export_downloads_current_full_history(tmp_path, monkeypatch):
    from app import db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "app.db")
    db.init_db()

    point = {
        "week_start": "2026-06-01",
        "week_end": "2026-06-07",
        "business_year": 2026,
        "business_week": 23,
        "week_label": "2026 W23",
        "display_date": "2026-06-02",
        "metric_type": "inventory",
        "source_country": "澳洲",
        "product": "PB粉",
        "category": "粉矿",
        "mainstream_status": "主流",
        "value": 100.0,
        "unit": "万吨",
        "source_file": "history.xlsx",
        "source_sheet": "库存",
        "source_section": "历史",
        "is_calculable": 1,
        "validation_status": "ok",
        "note": "",
    }
    _save_integrated_points([point], ["history.xlsx"], "pytest")
    _save_integrated_points([
        dict(point, business_week=24, week_start="2026-06-08", week_end="2026-06-14",
             week_label="2026 W24", display_date="2026-06-09", value=130.0,
             source_file="week1.xlsx")
    ], ["week1.xlsx"], "pytest")

    workbook_bytes = build_integrated_workbook_bytes()
    wb = load_workbook(BytesIO(workbook_bytes), read_only=True)
    ws = wb["整合明细"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    assert len(rows) == 3
    assert [row[3] for row in rows[1:]] == [23, 24]


def test_integrated_preview_cache_roundtrip():
    result = {
        "rows": [{"week_start": "2026-06-01", "metric_type": "inventory"}],
        "errors": [],
        "summary": {"total_points": 1},
    }

    preview_id = _save_integrated_preview_cache(result, "historical.xlsx")
    cached = _load_integrated_preview_cache(preview_id)

    assert cached["file_name"] == "historical.xlsx"
    assert cached["rows"] == result["rows"]
    assert cached["errors"] == []
    assert cached["summary"]["total_points"] == 1


def test_integrate_mysteel_files_local_templates():
    files = _local_mysteel_files()
    assert len(files) == 3
    result = integrate_mysteel_files(files)
    metrics = result["summary"]["metrics"]
    sample = result["points"][0]
    assert metrics["inventory"] > 0
    assert metrics["shipment"] > 0
    assert metrics["arrival"] > 0
    assert metrics["apparent_demand"] > 0
    assert all(point["metric_type"] != "shipment" for point in result["points"] if "到中国" in point["source_section"])
    assert sample["week_end"]
    assert sample["business_year"]
    assert sample["business_week"]
    assert sample["week_label"]
    assert not result["summary"]["warnings"]


def test_integrate_mysteel_files_reads_brazil_card_powder_shipments(tmp_path):
    path = tmp_path / "brazil.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "巴西发货量"
    ws.append(["占位"])
    for _ in range(25):
        ws.append([None])
    ws.append(["日期", "卡粉"])
    ws.append(["2026/6/1 - 2026/6/7", 307.233126])
    ws.append(["2026/6/8 - 2026/6/14", 357.908361])
    wb.save(path)

    result = integrate_mysteel_files([path])
    shipments = [
        point for point in result["points"]
        if point["metric_type"] == "shipment" and point["source_country"] == "巴西" and point["product"] == "卡粉"
    ]

    assert [point["week_start"] for point in shipments] == ["2026-06-01", "2026-06-08"]
    assert shipments[0]["value"] == 307.233126
    assert shipments[0]["source_section"] == "卡粉发运量"


def test_integrate_mysteel_files_treats_listed_blank_values_as_zero(tmp_path):
    path = tmp_path / "blank_mysteel.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("澳洲发货量")
    ws.append(["日期", "PB粉", "PB块", "混合粉"])
    ws.append(["2026/6/1 - 2026/6/7", None, 1, 2])

    ws = wb.create_sheet("澳洲预计到达中国锚地量")
    ws.append(["日期", "PB粉", "PB块", "混合粉"])
    ws.append(["2026/6/1 - 2026/6/7", None, 1, 2])

    ws = wb.create_sheet("巴西发货量")
    ws.append(["日期", "中国大陆", "卡粉"])
    ws.append(["2026/6/1 - 2026/6/7", None, None])

    ws = wb.create_sheet("全球铁矿石发运量")
    ws.append(["日期", "澳大利亚", "巴西", "南非"])
    ws.append(["2026/6/1 - 2026/6/7", 1, 2, None])

    ws = wb.create_sheet("粗粉")
    ws.append([None, None, None, "PB粉"])
    ws.append([date(2026, 6, 2), "总计", None, None])
    wb.save(path)

    result = integrate_mysteel_files([path])

    def values(metric_type, source_country, product):
        return [
            point["value"] for point in result["points"]
            if point["metric_type"] == metric_type
            and point["source_country"] == source_country
            and point["product"] == product
        ]

    assert values("shipment", "澳洲", "PB粉") == [0.0]
    assert values("arrival", "澳洲", "PB粉") == [0.0]
    assert values("shipment", "巴西", "卡粉") == [0.0]
    assert values("arrival", "巴西", "卡粉") == [0.0]
    assert values("shipment", "南非", "南非") == [0.0]
    assert values("inventory", "澳洲", "PB粉") == [0.0]


def test_integrate_mysteel_files_calculates_demand_with_zero_values_only_when_pair_exists(tmp_path):
    path = tmp_path / "zero_demand_mysteel.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("澳洲预计到达中国锚地量")
    ws.append(["日期", "PB粉", "PB块", "混合粉"])
    ws.append(["2026/6/1 - 2026/6/7", 5, 1, 2])
    ws.append(["2026/6/8 - 2026/6/14", None, 1, 2])

    ws = wb.create_sheet("粗粉")
    ws.append([None, None, None, "PB粉"])
    ws.append([date(2026, 6, 2), "总计", None, 10])
    ws.append([date(2026, 6, 9), "总计", None, None])
    ws.append([date(2026, 6, 16), "总计", None, 7])
    wb.save(path)

    result = integrate_mysteel_files([path])

    arrival = [
        point for point in result["points"]
        if point["metric_type"] == "arrival"
        and point["source_country"] == "澳洲"
        and point["product"] == "PB粉"
    ]
    inventory = [
        point for point in result["points"]
        if point["metric_type"] == "inventory"
        and point["source_country"] == "澳洲"
        and point["product"] == "PB粉"
    ]
    apparent_demand = [
        point for point in result["points"]
        if point["metric_type"] == "apparent_demand"
        and point["source_country"] == "澳洲"
        and point["product"] == "PB粉"
    ]

    assert [(point["week_start"], point["value"]) for point in arrival] == [
        ("2026-06-01", 5.0),
        ("2026-06-08", 0.0),
    ]
    assert [(point["week_start"], point["value"]) for point in inventory] == [
        ("2026-06-01", 10.0),
        ("2026-06-08", 0.0),
        ("2026-06-15", 7.0),
    ]
    assert [(point["week_start"], point["value"]) for point in apparent_demand] == [
        ("2026-06-08", 10.0),
    ]


def test_integrate_mysteel_files_normalizes_pellet_inventory_country_headers(tmp_path):
    path = tmp_path / "pellet_inventory_mysteel.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("球团")
    ws.append([None, None, None, "乌克兰", "印度"])
    ws.append([date(2026, 6, 2), "总计", None, 0.00, 0.00])
    wb.save(path)

    result = integrate_mysteel_files([path])
    inventory = {
        (point["source_country"], point["product"], point["category"]): point
        for point in result["points"]
        if point["metric_type"] == "inventory"
    }

    assert inventory[("乌克兰", "乌克兰", "球团")]["value"] == 0.0
    assert inventory[("印度", "印度", "球团")]["value"] == 0.0


def test_integrate_mysteel_files_normalizes_inventory_country_and_generic_headers(tmp_path):
    path = tmp_path / "inventory_headers_mysteel.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("粗粉")
    ws.append([None, None, None, "印度", "其他澳粉", "其他巴粗", "其他"])
    ws.append([date(2026, 6, 2), "总计", None, 1, 2, 3, 4])

    ws = wb.create_sheet("块矿")
    ws.append([None, None, None, "南非", "澳块", "巴块", "PMI块", "其他"])
    ws.append([date(2026, 6, 2), "总计", None, 5, 6, 7, 8, 9])

    ws = wb.create_sheet("精粉")
    ws.append([None, None, None, "乌克兰", "其他巴西精粉", "其他"])
    ws.append([date(2026, 6, 2), "总计", None, 10, 11, 12])
    wb.save(path)

    result = integrate_mysteel_files([path])
    inventory = {
        (point["source_country"], point["product"], point["category"]): point["value"]
        for point in result["points"]
        if point["metric_type"] == "inventory"
    }

    assert inventory[("印度", "印度", "粉矿")] == 1.0
    assert inventory[("澳洲", "其他", "粉矿")] == 2.0
    assert inventory[("巴西", "其他", "粉矿")] == 3.0
    assert inventory[("其他", "其他", "粉矿")] == 4.0
    assert inventory[("南非", "南非", "块矿")] == 5.0
    assert inventory[("澳洲", "其他", "块矿")] == 6.0
    assert inventory[("巴西", "巴西", "块矿")] == 7.0
    assert inventory[("澳洲", "PMI块", "块矿")] == 8.0
    assert inventory[("其他", "其他", "块矿")] == 9.0
    assert inventory[("乌克兰", "乌克兰", "精粉")] == 10.0
    assert inventory[("巴西", "其他", "精粉")] == 11.0
    assert inventory[("其他", "其他", "精粉")] == 12.0


def test_integrate_mysteel_files_calculates_demand_only_for_exact_arrival_inventory_pairs(tmp_path):
    path = tmp_path / "demand_scope_mysteel.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("巴西发货量")
    ws.append(["占位"])
    for _ in range(25):
        ws.append([None])
    ws.append(["日期", "中国大陆", "卡粉"])
    ws.append(["2026/5/11 - 2026/5/17", 120, 20])
    ws.append(["2026/5/18 - 2026/5/24", 160, 30])

    ws = wb.create_sheet("粗粉")
    ws.append([None, None, None, "卡粉", "巴混"])
    ws.append([date(2026, 6, 23), "总计", None, 100, 40])
    ws.append([date(2026, 6, 30), "总计", None, 90, 35])
    wb.save(path)

    result = integrate_mysteel_files([path])
    apparent = [
        point for point in result["points"]
        if point["metric_type"] == "apparent_demand"
        and point["source_country"] == "巴西"
    ]

    assert [(point["product"], point["week_start"], point["value"]) for point in apparent] == [
        ("卡粉", "2026-06-29", 160 * 0.75 + 100 - 90),
    ]


def test_integrate_mysteel_files_covers_current_template_sections():
    result = integrate_mysteel_files(_local_mysteel_files())
    australia_shipments = [
        point for point in result["points"]
        if point["metric_type"] == "shipment" and point["source_country"] == "澳洲"
    ]
    australia_arrivals = [
        point for point in result["points"]
        if point["metric_type"] == "arrival" and point["source_country"] == "澳洲"
    ]
    brazil_arrivals = [
        point for point in result["points"]
        if point["metric_type"] == "arrival" and point["source_country"] == "巴西"
    ]
    brazil_shipments = [
        point for point in result["points"]
        if point["metric_type"] == "shipment" and point["source_country"] == "巴西" and point["product"] == "卡粉"
    ]

    assert len(australia_shipments) == 174
    assert len(australia_arrivals) == 203
    assert len(brazil_arrivals) == 6
    assert len(brazil_shipments) == 6
    assert sorted({point["week_start"] for point in australia_shipments}) == [
        "2026-05-11", "2026-05-18", "2026-05-25",
        "2026-06-01", "2026-06-08", "2026-06-15",
    ]
    assert any(
        point["product"] == "PB粉"
        and point["week_start"] == "2026-06-01"
        and point["source_section"] == "澳洲发货量（分品种）"
        for point in australia_shipments
    )
    assert any(
        point["product"] == "澳大利亚球团"
        and point["week_start"] == "2026-06-08"
        and point["value"] == 0
        for point in australia_shipments
    )
    assert sorted({point["week_start"] for point in brazil_arrivals}) == [
        "2026-06-22", "2026-06-29", "2026-07-06",
        "2026-07-13", "2026-07-20", "2026-07-27",
    ]
    assert sorted({point["week_start"] for point in brazil_shipments}) == [
        "2026-05-11", "2026-05-18", "2026-05-25",
        "2026-06-01", "2026-06-08", "2026-06-15",
    ]

print("All tests passed!")
