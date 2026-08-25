#!/usr/bin/env python3
"""Dry-run or safely apply the approved 2026 spot-ledger workbook fields to Staging.

The script deliberately has no Production mode. It only accepts the exact Render
Staging host, never writes Q/AU, and patches blank fields after an exact match.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date
import json
import os
from pathlib import Path
import re
import sys
from typing import Any, Iterable
from urllib.parse import quote, urlparse

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.spot_ledger import (  # noqa: E402
    FIELD_DEFINITIONS,
    MANUAL_FIELDS,
    NUMERIC_FIELDS,
    SPOT_LEDGER_FOCUS_START_DATE,
    SYSTEM_PRIORITY_FIELDS,
    _normalize_date,
    _number,
)


STAGING_BASE_URL = "https://ltm-web-staging.onrender.com"
BACKFILL_FIELDS = MANUAL_FIELDS | SYSTEM_PRIORITY_FIELDS
SOURCE_KEY_FIELDS = {"AD", "H", "Z", "X", "L", "U"}
PLACEHOLDERS = {"", "-", "--", "—", "——", "***"}
HEADER_TO_CODE = {
    re.sub(r"\s+", "", str(item["name"])).translate(str.maketrans({"(": "（", ")": "）"})): item["code"]
    for item in FIELD_DEFINITIONS
}
HEADER_TO_CODE.update({"长协对象": "long_contract_object", "long_contract_object": "long_contract_object"})


def _usable(value: Any) -> bool:
    return value is not None and str(value).strip() not in PLACEHOLDERS


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalized_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).translate(str.maketrans({"(": "（", ")": "）"}))


def _date_status(value: Any) -> str:
    normalized = _normalize_date(value)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        return "invalid"
    try:
        date.fromisoformat(normalized)
    except ValueError:
        return "invalid"
    return "historical" if normalized < SPOT_LEDGER_FOCUS_START_DATE else "focus"


def validate_staging_base_url(value: str) -> str:
    parsed = urlparse(value.strip())
    if parsed.scheme != "https" or parsed.netloc != urlparse(STAGING_BASE_URL).netloc or parsed.path not in {"", "/"} or parsed.query or parsed.fragment:
        raise ValueError("只允许使用 https://ltm-web-staging.onrender.com，禁止其他环境或带查询参数地址")
    return STAGING_BASE_URL


def _header_code(value: Any) -> str:
    text = str(value or "").strip()
    return text if text in SOURCE_KEY_FIELDS or text in BACKFILL_FIELDS else HEADER_TO_CODE.get(_normalized_header(text), "")


def read_workbook_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["现货业务台账"] if "现货业务台账" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers: list[Any] | None = None
    for row_number, values in enumerate(iterator, start=1):
        codes = {_header_code(value) for value in values}
        if {"AD", "H", "Z"}.issubset(codes) and ({"X", "L"} & codes):
            headers = list(values)
            break
        if row_number >= 20:
            break
    if headers is None:
        workbook.close()
        raise ValueError("工作簿前 20 行未找到销售合同号、商品、价格和数量表头")
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row: dict[str, Any] = {}
        for header, value in zip(headers, values):
            code = _header_code(header)
            if code:
                row[code] = value
        if any(_usable(row.get(code)) for code in ("AD", "H", "Z", "X", "L")):
            rows.append(row)
    title = sheet.title
    workbook.close()
    return title, rows


def _quantity(value: dict[str, Any]) -> Any:
    return value.get("X") if _usable(value.get("X")) else value.get("L")


def _value_equal(field: str, current: Any, incoming: Any) -> bool:
    if not _usable(current) and not _usable(incoming):
        return True
    if field in NUMERIC_FIELDS:
        left = _number(current)
        right = _number(incoming)
        return left is not None and right is not None and abs(left - right) <= 0.000001
    return _text(current) == _text(incoming)


def _candidate_matches(source: dict[str, Any], candidate: dict[str, Any], detail: dict[str, Any]) -> bool:
    if _text(source.get("AD")) != _text(candidate.get("AD")):
        return False
    if _text(source.get("H")) != _text(candidate.get("H")):
        return False
    if _normalize_date(source.get("U")) != _normalize_date(candidate.get("U")):
        return False
    if not _value_equal("X", _quantity(source), _quantity(candidate)):
        return False
    return _value_equal("Z", source.get("Z"), detail.get("Z"))


def build_backfill_plan(
    source_rows: Iterable[dict[str, Any]],
    records: list[dict[str, Any]],
    details: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    rows = list(source_rows)
    by_contract_product: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_contract_product[(_text(record.get("AD")), _text(record.get("H")))].append(record)
    result: dict[str, Any] = {
        "source_rows": len(rows),
        "focus_rows": 0,
        "records": len(records),
        "skipped_historical": 0,
        "skipped_invalid_date": 0,
        "skipped_invalid_key": 0,
        "unique_matches": 0,
        "ambiguous": 0,
        "unmatched": 0,
        "candidate_updates": 0,
        "conflicts": 0,
        "conflict_field_counts": {},
        "field_updates": {},
        "plans": [],
    }
    for source in rows:
        status = _date_status(source.get("U"))
        if status == "historical":
            result["skipped_historical"] += 1
            continue
        if status != "focus":
            result["skipped_invalid_date"] += 1
            continue
        result["focus_rows"] += 1
        if not all(_usable(source.get(field)) for field in ("AD", "H", "Z")) or not _usable(_quantity(source)):
            result["skipped_invalid_key"] += 1
            continue
        candidates = [
            candidate
            for candidate in by_contract_product[(_text(source.get("AD")), _text(source.get("H")))]
            if _candidate_matches(source, candidate, details.get(str(candidate.get("record_id")), {}))
        ]
        if len(candidates) != 1:
            result["ambiguous" if len(candidates) > 1 else "unmatched"] += 1
            continue
        result["unique_matches"] += 1
        candidate = candidates[0]
        detail = details.get(str(candidate.get("record_id")), {})
        values: dict[str, Any] = {}
        current_values: dict[str, Any] = {}
        for field in sorted(BACKFILL_FIELDS):
            incoming = source.get(field)
            if not _usable(incoming):
                continue
            if field in NUMERIC_FIELDS:
                incoming = _number(incoming)
                if incoming is None:
                    continue
            current = detail.get(field)
            if not _usable(current):
                values[field] = incoming
                current_values[field] = current
                result["field_updates"][field] = result["field_updates"].get(field, 0) + 1
            elif _value_equal(field, current, incoming):
                continue
            else:
                result["conflicts"] += 1
                field_counts = result["conflict_field_counts"]
                field_counts[field] = field_counts.get(field, 0) + 1
        if values:
            result["candidate_updates"] += 1
            result["plans"].append({"record_id": str(candidate["record_id"]), "values": values, "current_values": current_values})
    return result


def candidate_detail_ids(source_rows: Iterable[dict[str, Any]], records: list[dict[str, Any]]) -> list[str]:
    source_keys = {
        (_text(row.get("AD")), _text(row.get("H")), _normalize_date(row.get("U")), _text(_quantity(row)))
        for row in source_rows
        if _date_status(row.get("U")) == "focus" and any(_usable(row.get(field)) for field in BACKFILL_FIELDS)
    }
    return [
        str(record["record_id"])
        for record in records
        if record.get("record_id") and (
            _text(record.get("AD")), _text(record.get("H")), _normalize_date(record.get("U")), _text(_quantity(record))
        ) in source_keys
    ]


class StagingLedgerClient:
    def __init__(self, base_url: str, username: str, password: str, timeout: float = 30, session: requests.Session | None = None):
        self.base_url = validate_staging_base_url(base_url)
        self.username = username
        self.password = password
        self.timeout = timeout
        self.session = session or requests.Session()
        self.token = ""

    def _request(self, method: str, path: str, **kwargs: Any) -> dict[str, Any]:
        headers = dict(kwargs.pop("headers", {}))
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        response = self.session.request(method, f"{self.base_url}{path}", headers=headers, timeout=self.timeout, **kwargs)
        if not 200 <= response.status_code < 300:
            raise RuntimeError(f"Staging API HTTP {response.status_code}: {method} {path}")
        try:
            return response.json()
        except ValueError as exc:
            raise RuntimeError(f"Staging API returned invalid JSON: {method} {path}") from exc

    def login(self) -> None:
        result = self._request("POST", "/api/auth/login", json={"username": self.username, "password": self.password})
        self.token = str(result.get("token") or "")
        if not self.token:
            raise RuntimeError("Staging login returned no token")

    def list_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        offset = 0
        while True:
            result = self._request("GET", "/api/spot-ledger/records", params={"limit": 100, "offset": offset})
            page = result.get("records") or []
            records.extend(item for item in page if isinstance(item, dict))
            total = int(result.get("count") or len(records))
            if not page or len(records) >= total:
                return records
            offset += len(page)

    def get_detail(self, record_id: str) -> dict[str, Any]:
        return self._request("GET", f"/api/spot-ledger/records/{quote(record_id, safe='')}").get("record") or {}

    def patch(self, record_id: str, values: dict[str, Any]) -> dict[str, Any]:
        return self._request("PATCH", f"/api/spot-ledger/records/{quote(record_id, safe='')}", json={"values": values})

    def reconcile_mappings(self, apply: bool = False) -> dict[str, Any]:
        return self._request("POST", "/api/spot-ledger/reconcile-mappings", params={"apply": str(apply).lower()})

    def get_details_concurrent(self, record_ids: Iterable[str], max_workers: int = 8) -> dict[str, dict[str, Any]]:
        ids = list(record_ids)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            details = executor.map(self.get_detail, ids)
        return dict(zip(ids, details))


def apply_backfill_plan(client: StagingLedgerClient, result: dict[str, Any], change_log: Path) -> dict[str, Any]:
    changes: list[dict[str, Any]] = []
    applied = 0
    skipped_race = 0
    failed = 0
    for plan in result.get("plans", []):
        record_id = str(plan["record_id"])
        current = client.get_detail(record_id)
        if any(not _value_equal(field, current.get(field), expected) for field, expected in plan.get("current_values", {}).items()):
            skipped_race += 1
            continue
        changes.append({"record_id": record_id, "values": plan["values"], "previous": plan.get("current_values", {})})
    change_log.parent.mkdir(parents=True, exist_ok=True)
    change_log.write_text(json.dumps(changes, ensure_ascii=False, indent=2), encoding="utf-8")
    for change in changes:
        try:
            client.patch(change["record_id"], change["values"])
            applied += 1
        except (OSError, RuntimeError, requests.RequestException):
            failed += 1
            break
    return {"change_log": str(change_log), "applied": applied, "skipped_race": skipped_race, "failed": failed}


def safe_summary(result: dict[str, Any]) -> str:
    hidden = {"plans", "password", "token", "username", "current_values", "previous"}
    public = {key: value for key, value in result.items() if key not in hidden}
    return json.dumps(public, ensure_ascii=False, separators=(",", ":"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry-run or apply the 2026 spot-ledger workbook to Staging.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--base-url", default=STAGING_BASE_URL)
    parser.add_argument("--apply", action="store_true", help="Explicitly write the dry-run-approved blank fields to Staging.")
    parser.add_argument("--change-log", type=Path, default=Path("/tmp/spot-ledger-staging-backfill-change-log.json"))
    parser.add_argument("--reconcile-mappings", action="store_true", help="Run the admin-only stored mapping reconciliation first.")
    parser.add_argument("--reconcile-only", action="store_true", help="Only run mapping reconciliation; do not read or backfill the workbook.")
    parser.add_argument("--timeout", type=float, default=30)
    args = parser.parse_args()
    username = os.getenv("STAGING_LEDGER_USERNAME", "")
    password = os.getenv("STAGING_LEDGER_PASSWORD", "")
    if not username or not password:
        print(json.dumps({"ok": False, "error": "需要设置 STAGING_LEDGER_USERNAME 和 STAGING_LEDGER_PASSWORD（不通过命令行传递）"}, ensure_ascii=False))
        return 2
    try:
        client = StagingLedgerClient(args.base_url, username, password, timeout=args.timeout)
        client.login()
        if args.reconcile_mappings:
            print(json.dumps({"ok": True, "mapping_reconciliation": json.loads(safe_summary(client.reconcile_mappings(apply=args.apply)))}, ensure_ascii=False, separators=(",", ":")))
            if args.reconcile_only:
                return 0
        sheet, source_rows = read_workbook_rows(args.workbook)
        records = client.list_records()
        details = client.get_details_concurrent(candidate_detail_ids(source_rows, records))
        result = build_backfill_plan(source_rows, records, details)
        result["sheet"] = sheet
        result["dry_run"] = not args.apply
        print(json.dumps({"ok": True, "backfill": json.loads(safe_summary(result))}, ensure_ascii=False, separators=(",", ":")))
        if not args.apply or not result["plans"]:
            return 0
        applied = apply_backfill_plan(client, result, args.change_log)
        print(json.dumps({"ok": applied["failed"] == 0, "backfill_apply": applied}, ensure_ascii=False, separators=(",", ":")))
        return 0 if applied["failed"] == 0 else 1
    except (OSError, ValueError, RuntimeError, requests.RequestException) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, separators=(",", ":")))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
