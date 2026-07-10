import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


def test_roster_preflight_reads_required_columns_and_stops_before_apply_on_conflict(tmp_path):
    from scripts import provision_users

    path = tmp_path / "users.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.append(["姓名", "部门", "用户类型", "登录账号（可空）"])
    sheet.append(["张三", "贸易处", "用户", "zhangsan"])
    sheet.append(["李雷", "期货组", "领导", ""])
    book.save(path)

    rows = provision_users.load_roster(path)
    calls = []

    def preview(payload):
        calls.append(payload)
        username = payload["username"] or "lilei"
        return {
            "name": payload["name"],
            "username": username,
            "temporary_password": f"{username}123",
            "password_rule": "test_rule",
            "username_available": username != "zhangsan",
            "final_permissions": {"info_summary": "operate"},
        }

    try:
        provision_users.preflight_roster(rows, preview)
    except ValueError as exc:
        assert "zhangsan" in str(exc)
    else:
        raise AssertionError("conflicting roster must fail before any create call")

    assert len(calls) == 2
    assert rows[1]["role"] == "领导"


def test_roster_preflight_returns_preview_without_writing(tmp_path):
    from scripts import provision_users

    path = tmp_path / "users.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.append(["姓名", "部门", "用户类型", "登录账号（可空）"])
    sheet.append(["王芳", "财企处", "用户", "wangfang"])
    book.save(path)

    rows = provision_users.load_roster(path)
    result = provision_users.preflight_roster(
        rows,
        lambda payload: {
            **payload,
            "username": "wangfang",
            "temporary_password": "wangfang123",
            "password_rule": "username_suffix",
            "username_available": True,
            "final_permissions": {"order_finance_progress": "operate"},
        },
    )

    assert result == [{
        "name": "王芳",
        "username": "wangfang",
        "temporary_password": "wangfang123",
        "password_rule": "username_suffix",
        "department": "财企处",
        "role": "用户",
        "permission_count": 1,
        "permissions": [
            {"module_code": "order_finance_progress", "level": "operate"},
        ],
    }]


def test_three_column_roster_normalizes_whitespace_and_adds_long_yunfei_override(tmp_path):
    from scripts import provision_users

    path = tmp_path / "users.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.append(["姓名", "部门", "用户类型"])
    sheet.append(["龙云飞", "贸易处", "用户 "])
    sheet.append(["张胜根", "公司领导", "领导"])
    book.save(path)

    rows = provision_users.load_roster(path)
    assert rows == [
        {"name": "龙云飞", "department": "贸易处", "role": "用户", "username": ""},
        {"name": "张胜根", "department": "公司领导", "role": "领导", "username": ""},
    ]
    assert provision_users.permission_overrides(rows[0]) == [
        {"module_code": "order_finance_progress", "level": "operate"},
        {"module_code": "order_finance_capital", "level": "operate"},
    ]
    assert provision_users.permission_overrides(rows[1]) == []


def test_preflight_preserves_password_rule_and_permissions_for_apply():
    from scripts import provision_users

    row = {"name": "龙云飞", "department": "贸易处", "role": "用户", "username": ""}
    result = provision_users.preflight_roster(
        [row],
        lambda payload: {
            "username": "longyunfei",
            "temporary_password": "longyunfei",
            "password_rule": "trade_or_futures_plain",
            "username_available": True,
            "final_permissions": {
                "order_finance_progress": "operate",
                "order_finance_capital": "operate",
            },
        },
    )

    assert result[0]["password_rule"] == "trade_or_futures_plain"
    assert result[0]["permissions"] == [
        {"module_code": "order_finance_progress", "level": "operate"},
        {"module_code": "order_finance_capital", "level": "operate"},
    ]


def test_apply_roster_uses_preflight_permissions():
    from scripts import provision_users

    calls = []

    class Client:
        def request(self, path, payload=None, method="POST"):
            calls.append((path, payload, method))
            return {"id": 8}

    rows = [{"name": "龙云飞", "department": "贸易处", "role": "用户", "username": ""}]
    previews = [{
        "username": "longyunfei",
        "permissions": [
            {"module_code": "order_finance_progress", "level": "operate"},
            {"module_code": "order_finance_capital", "level": "operate"},
        ],
    }]

    provision_users.apply_roster(Client(), rows, previews)

    assert calls == [(
        "/api/users",
        {
            "name": "龙云飞",
            "department": "贸易处",
            "role": "用户",
            "username": "longyunfei",
            "permissions": previews[0]["permissions"],
        },
        "POST",
    )]


def test_configure_wangjingze_validates_existing_admin_and_hides_secret(capsys):
    from scripts import provision_users

    calls = []

    class Client:
        def request(self, path, payload=None, method="POST"):
            calls.append((path, payload, method))
            if path == "/api/users":
                return {
                    "users": [{
                        "id": 7,
                        "name": "王景泽",
                        "department": "管理部门",
                        "role": "管理员",
                    }],
                }
            return {"ok": True}

    secret = "".join(["test", "-only", "-secret"])
    provision_users.configure_wangjingze(Client(), secret)

    assert calls == [
        ("/api/users", None, "GET"),
        (
            "/api/users/7/set-password",
            {"new_password": secret, "password_change_recommended": False},
            "POST",
        ),
    ]
    captured = capsys.readouterr()
    assert secret not in captured.out + captured.err


def test_configure_wangjingze_rejects_wrong_department_or_role():
    from scripts import provision_users

    class Client:
        def request(self, path, payload=None, method="POST"):
            assert path == "/api/users"
            return {
                "users": [{
                    "id": 7,
                    "name": "王景泽",
                    "department": "贸易处",
                    "role": "用户",
                }],
            }

    try:
        provision_users.configure_wangjingze(Client(), "".join(["test", "-only"]))
    except ValueError as exc:
        assert str(exc) == "王景泽必须先是管理部门管理员"
    else:
        raise AssertionError("非管理部门管理员不应设置密码")


def test_api_client_get_has_no_request_body(monkeypatch):
    from scripts import provision_users

    seen = {}

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def read(self):
            return b'{"users": []}'

    def fake_urlopen(request, timeout):
        seen["method"] = request.get_method()
        seen["data"] = request.data
        seen["timeout"] = timeout
        return Response()

    monkeypatch.setattr(provision_users, "urlopen", fake_urlopen)
    result = provision_users.ApiClient("https://example.invalid").request(
        "/api/users", method="GET"
    )

    assert result == {"users": []}
    assert seen == {"method": "GET", "data": None, "timeout": 30}
