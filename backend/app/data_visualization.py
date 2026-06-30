"""
数据可视化管理 — 卡粉 MVP。
Excel 解析、业务周计算、表需计算、导入预检/确认、数据查询与编辑、图表数据。
"""
from __future__ import annotations

import base64
import io
import json
import os
import tempfile
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, Header, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel

from . import db

router = APIRouter()

# ── 常量 ──────────────────────────────────────────────────────────────
PRODUCT = "卡粉"
DV_PRODUCTS = ["卡粉", "纽曼粉", "麦克粉", "PB粉", "金布巴粉", "超特粉", "混合粉", "杨迪粉", "罗伊山粉", "巴西粗粉", "乌克兰精粉", "俄罗斯精粉", "IOC6"]
SHIPMENT_SHEET = "发运"
INVENTORY_SHEET = "卡粉库存"
SHIPMENT_DATE_COL = 2   # B 列
SHIPMENT_VAL_COL = 51   # AY 列
INVENTORY_DATE_COL = 1  # A 列
INVENTORY_VAL_COL = 2   # B 列
LOCAL_MYSTEEL_DIR = Path("/Users/wangjingze/建龙/期货组/本地数据库")


_INTEGRATED_FIELDS = {
    "统计周一": "week_start", "统计周日": "week_end",
    "业务年份": "business_year", "业务周次": "business_week",
    "周次标签": "week_label", "展示日期": "display_date",
    "数据类型": "metric_type", "来源/国家": "source_country",
    "品种": "product", "种类": "category",
    "主流/非主流": "mainstream_status", "数值": "value",
    "单位": "unit", "来源文件": "source_file",
    "来源Sheet": "source_sheet", "来源区域": "source_section",
    "是否参与表需": "is_calculable", "校验状态": "validation_status",
    "备注": "note",
}

_METRIC_TYPE_CN = {
    "库存": "inventory", "发运": "shipment", "发运/到港": "shipment",
    "到港": "arrival", "表需": "apparent_demand",
}
_VALID_INTEGRATED_METRIC_TYPES = {"inventory", "shipment", "arrival", "apparent_demand"}

_REQUIRED_INTEGRATED_FIELDS = [
    "week_start", "display_date", "metric_type",
    "source_country", "product", "category", "mainstream_status",
]
_INTEGRATED_PREVIEW_CACHE_DIR = Path(tempfile.gettempdir()) / "ltm_dv_import_previews"
_INTEGRATED_PREVIEW_FILE_DIR = _INTEGRATED_PREVIEW_CACHE_DIR / "files"
_INTEGRATED_PREVIEW_JOBS: Dict[str, Dict[str, Any]] = {}
_INTEGRATED_IMPORT_JOBS: Dict[str, Dict[str, Any]] = {}

class ImportRequest(BaseModel):
    file_data: Optional[str] = None
    file_name: str
    preview_id: Optional[str] = None
    overwrite_manual_ids: List[int] = []


class ManualEditRequest(BaseModel):
    data_point_id: int
    new_value: float


class IntegrationUploadFile(BaseModel):
    file_name: str
    file_data: str


class IntegrationFilesRequest(BaseModel):
    files: List[IntegrationUploadFile]

# ── 权限（自包含，避免循环导入 main.py）───────────────────────────────


async def dv_current_user(authorization: Optional[str] = Header(default=None)):
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.removeprefix("Bearer ").strip()
    user = db.get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="请先登录")
    return user


def dv_require_edit(module_code: str, user: dict):
    if user["role"] == "管理员":
        return
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(
            cur,
            "SELECT can_edit FROM module_permissions WHERE user_id = ? AND module_code = ?",
            (user["id"], module_code),
        ).fetchone()
    if not row or not row["can_edit"]:
        raise HTTPException(status_code=403, detail="没有编辑权限")


# ── 工具函数 ──────────────────────────────────────────────────────────


def _to_date(val: Any) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _row_to_dict(row) -> dict:
    return dict(row) if row else {}


def _week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _week_end(d: date) -> date:
    return _week_start(d) + timedelta(days=6)


def _normalize_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError:
        return text


def _normalize_integrated_point(point: Dict[str, Any]) -> Dict[str, Any]:
    item = dict(point)
    for field in ("week_start", "week_end", "display_date"):
        item[field] = _normalize_date_value(item.get(field))
    if item.get("metric_type") == "inventory":
        source_country, product, category = _canonical_inventory_identity(
            item.get("source_sheet", ""),
            item.get("product", ""),
            item.get("category", ""),
            item.get("source_country", ""),
        )
        item["source_country"] = source_country
        item["product"] = product
        item["category"] = category
    if item.get("product"):
        item["mainstream_status"] = _mainstream_status(item["product"], item.get("category", ""))
    if item.get("metric_type") in {"arrival", "inventory"}:
        item["is_calculable"] = 1 if _is_apparent_demand_candidate(
            item.get("source_country", ""),
            item.get("product", ""),
            item.get("category", ""),
        ) else 0
    return item


def _week_map_key(row: Dict[str, Any]) -> str:
    business_year = row.get("business_year")
    business_week = row.get("business_week")
    if business_year and business_week:
        return f"{business_year}-W{int(business_week):02d}"
    return _normalize_date_value(row.get("week_start"))


def _parse_period_start(value: Any) -> Optional[date]:
    if not value:
        return None
    try:
        left = str(value).split("-")[0].strip()
        return datetime.strptime(left, "%Y/%m/%d").date()
    except (ValueError, TypeError):
        return None


def _clean_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


AUSTRALIA_ARRIVAL_HEADER_ROW = 20
AUSTRALIA_ARRIVAL_START_ROW = 21
AUSTRALIA_ARRIVAL_END_ROW = 27
BRAZIL_DESTINATION_HEADER_ROW = 39
BRAZIL_DESTINATION_START_ROW = 40
BRAZIL_DESTINATION_END_ROW = 45


AUSTRALIA_PRODUCT_MAP: Dict[str, Optional[tuple]] = {
    "PB粉": ("PB粉", "粉矿", "粗粉:PB粉"),
    "PB块": ("PB块", "块矿", "块矿:PB块"),
    "杨迪粉": ("杨迪粉", "粉矿", "粗粉:杨迪粉"),
    "大杨迪": ("杨迪粉", "粉矿", "粗粉:杨迪粉"),
    "混合粉": ("混合粉", "粉矿", "粗粉:混合粉"),
    "纽曼块": ("纽曼块", "块矿", "块矿:纽曼块"),
    "超特粉": ("超特粉", "粉矿", "粗粉:超特粉"),
    "纽曼粉": ("纽曼粉", "粉矿", "粗粉:纽曼粉"),
    "麦克粉": ("麦克粉", "粉矿", "粗粉:麦克粉"),
    "金布巴粉": ("金布巴粉", "粉矿", "粗粉:金布巴粉"),
    "罗布河粉": ("罗布河粉", "粉矿", "粗粉:罗布河粉"),
    "国王粉": ("国王粉", "粉矿", "粗粉:国王粉"),
    "Roy Hill粉": ("罗伊山粉", "粉矿", "粗粉:罗伊山粉"),
    "ATLAS粉": ("Atlas粉", "粉矿", "粗粉:Atlas粉"),
    "罗布河块": ("罗布河块", "块矿", "块矿:罗布河块"),
    "卡拉拉粉": ("卡拉拉精粉", "精粉", "精粉:卡拉拉精粉"),
    "Roy Hill块": ("罗伊山块", "块矿", "块矿:罗伊山块"),
    "球团": ("澳大利亚球团", "球团", "球团:澳大利亚"),
    "FMG 块矿": ("FMG块", "块矿", "块矿:FMG块"),
    "西皮尔巴拉粉": ("西皮尔巴拉粉", "粉矿", "粗粉:西皮尔巴拉粉"),
    "SP10粉": ("SP10粉", "粉矿", "粗粉:RTX粉(SP10粉)"),
    "SP10块": ("SP10块", "块矿", "块矿:SP10块"),
    "库宾块": ("库宾块", "块矿", "块矿:库宾块"),
    "库兰精粉": ("库兰粉", "粉矿", "粗粉:库兰粉"),
    "一刚粉": ("一钢粉", "粉矿", "粗粉:一钢粉"),
    "一刚块": ("一钢块", "块矿", "块矿:一钢块"),
    "中信泰富精粉": ("泰富精粉", "精粉", "精粉:泰富精粉"),
    "铁桥精粉": ("铁桥精粉", "精粉", "精粉:铁桥精粉"),
    "丝路粉": ("丝路粉", "粉矿", "粗粉:丝路粉"),
    "RTX": None,
    "RTX块": None,
    "高锰粉矿": None,
    "未知": None,
    "库宾粉": None,
}


MAINSTREAM_PRODUCT_IDENTITIES = [
    ("PB粉", None, "PB粉"),
    ("麦克粉", None, "麦克粉"),
    ("纽曼粉", None, "纽曼粉"),
    ("金布巴粉", None, "金布巴粉"),
    ("超特粉", None, "超特粉"),
    ("混合粉", None, "混合粉"),
    ("卡粉", None, "卡粉"),
    ("巴混", None, "巴混"),
    ("SP10粉", None, "SP10粉"),
    ("几内亚", "粉矿", "几内亚（粉矿）"),
    ("杨迪粉", None, "杨迪粉"),
    ("罗伊山粉", None, "罗伊山粉"),
    ("罗伊山MB粉", None, "罗伊山MB粉"),
    ("印度", "粉矿", "印度（粉矿）"),
    ("IOC6", None, "IOC6"),
    ("PB块", None, "PB块"),
    ("纽曼块", None, "纽曼块"),
    ("SP10块", None, "SP10块"),
    ("南非", "块矿", "南非（块矿）"),
    ("罗伊山MB块", None, "罗伊山MB块"),
    ("PMI块", None, "PMI块"),
    ("印度", "球团", "印度（球团）"),
    ("乌克兰", "精粉", "乌克兰（精粉）"),
    ("卡拉拉精粉", None, "卡拉拉精粉"),
]
MAINSTREAM_PRODUCT_ORDER = [label for _product, _category, label in MAINSTREAM_PRODUCT_IDENTITIES]
MAINSTREAM_PRODUCT_ORDER_INDEX = {
    product: index for index, product in enumerate(MAINSTREAM_PRODUCT_ORDER)
}
MAINSTREAM_PRODUCTS = {product for product, _category, _label in MAINSTREAM_PRODUCT_IDENTITIES}
MAINSTREAM_PRODUCT_WILDCARDS = {
    product for product, category, _label in MAINSTREAM_PRODUCT_IDENTITIES
    if category is None
}
MAINSTREAM_PRODUCT_CATEGORY_LABELS = {
    (product, category): label
    for product, category, label in MAINSTREAM_PRODUCT_IDENTITIES
    if category is not None
}


INVENTORY_SHEETS = {"粗粉", "块矿", "球团", "精粉"}
INVENTORY_CATEGORIES = {"粉矿", "块矿", "球团", "精粉"}

INVENTORY_RENAME = {
    ("粗粉", "RTX粉(SP10粉)"): ("SP10粉", "粉矿", "澳洲"),
    ("粗粉", "其他澳粉"): ("其他", "粉矿", "澳洲"),
    ("粗粉", "其他巴粗"): ("其他", "粉矿", "巴西"),
    ("粗粉", "其他粉矿"): ("其他", "粉矿", "其他"),
    ("粗粉", "PMI粉"): ("PMI粉", "粉矿", "澳洲"),
    ("粗粉", "RTBF/RTGF"): ("RTBF/RTGF", "粉矿", "澳洲"),
    ("粗粉", "金宝粉"): ("金宝粉", "粉矿", "澳洲"),
    ("粗粉", "哈扬粉"): ("哈扬粉", "粉矿", "澳洲"),
    ("粗粉", "IOH4"): ("IOH4", "粉矿", "巴西"),
    ("粗粉", "IOC6"): ("IOC6", "粉矿", "巴西"),
    ("粗粉", "特卡粉"): ("特卡粉", "粉矿", "巴西"),
    ("粗粉", "其他CSN粗粉"): ("其他CSN粗粉", "粉矿", "巴西"),
    ("块矿", "澳块"): ("其他", "块矿", "澳洲"),
    ("块矿", "巴块"): ("巴西", "块矿", "巴西"),
    ("块矿", "PMI块"): ("PMI块", "块矿", "澳洲"),
    ("块矿", "其他"): ("其他", "块矿", "其他"),
    ("球团", "澳大利亚"): ("澳大利亚球团", "球团", "澳洲"),
    ("球团", "其他"): ("其他", "球团", "其他"),
    ("精粉", "澳大利亚其他精粉"): ("其他", "精粉", "澳洲"),
    ("精粉", "其他巴西精粉"): ("其他", "精粉", "巴西"),
    ("精粉", "泰富精粉"): ("泰富精粉", "精粉", "澳洲"),
    ("精粉", "铁桥精粉"): ("铁桥精粉", "精粉", "澳洲"),
    ("精粉", "米纳斯精粉"): ("米纳斯精粉", "精粉", "巴西"),
    ("精粉", "其他"): ("其他", "精粉", "其他"),
}

INVENTORY_COUNTRY_HEADERS = {
    "印度", "南非", "毛里塔尼亚", "秘鲁", "伊朗", "新西兰", "马来西亚",
    "印尼", "智利", "西班牙", "墨西哥", "塞拉利昂", "委内瑞拉",
    "加拿大", "俄罗斯", "瑞典", "阿曼", "土耳其", "巴西", "乌克兰",
}

INVENTORY_COUNTRY_ALIAS = {
    "几内亚粉": "几内亚",
}

LEGACY_INVENTORY_IDENTITY_RENAME = {
    ("inventory", "球团", "乌克兰球"): ("乌克兰", "乌克兰", "球团"),
    ("inventory", "球团", "印球"): ("印度", "印度", "球团"),
}


GLOBAL_SKIP_COUNTRIES = {"澳大利亚", "巴西", "总计"}


def _category_for_inventory(sheet_name: str, header: str) -> str:
    if sheet_name == "粗粉":
        return "粉矿"
    if sheet_name == "块矿":
        return "块矿"
    if sheet_name == "球团":
        return "球团"
    if sheet_name == "精粉":
        return "精粉"
    if "块" in header:
        return "块矿"
    if "球" in header:
        return "球团"
    if "精粉" in header:
        return "精粉"
    return "粉矿"


def _source_country_for_product(product: str, category: str, sheet_name: str = "") -> str:
    australia_tokens = (
        "PB", "纽曼", "麦克", "金布巴", "超特", "混合", "杨迪", "罗布河",
        "国王", "罗伊山", "Atlas", "ATLAS", "FMG", "SP10", "库宾",
        "库兰", "一钢", "泰富", "铁桥", "丝路", "卡拉拉", "西皮尔巴拉", "澳大利亚",
    )
    brazil_tokens = ("卡粉", "巴混", "巴粗", "CSN", "托克", "米纳斯")
    if any(token in product for token in australia_tokens):
        return "澳洲"
    if any(token in product for token in brazil_tokens):
        return "巴西"
    return product if sheet_name in {"球团", "精粉"} and len(product) <= 8 else "其他"


AUSTRALIA_ARRIVAL_PRODUCTS = {
    (mapped[0], mapped[1])
    for mapped in AUSTRALIA_PRODUCT_MAP.values()
    if mapped is not None
}


def _is_apparent_demand_candidate(source_country: str, product: str, category: str) -> bool:
    if source_country == "澳洲" and (product, category) in AUSTRALIA_ARRIVAL_PRODUCTS:
        return True
    return source_country == "巴西" and product == "卡粉" and category == "粉矿"


def _canonical_inventory_identity(
    sheet_name: str,
    product: str,
    category: str = "",
    source_country: str = "",
) -> tuple:
    sheet = sheet_name if sheet_name in INVENTORY_SHEETS else ""
    inventory_category = category or _category_for_inventory(sheet, product)

    legacy = LEGACY_INVENTORY_IDENTITY_RENAME.get(("inventory", inventory_category, product))
    if legacy:
        return legacy

    if sheet:
        renamed = INVENTORY_RENAME.get((sheet, product))
        if renamed:
            return renamed[2], renamed[0], renamed[1]
        inventory_category = _category_for_inventory(sheet, product)
    else:
        for candidate_sheet in INVENTORY_SHEETS:
            if _category_for_inventory(candidate_sheet, product) != inventory_category:
                continue
            renamed = INVENTORY_RENAME.get((candidate_sheet, product))
            if renamed:
                return renamed[2], renamed[0], renamed[1]

    country = INVENTORY_COUNTRY_ALIAS.get(product)
    if country:
        return country, country, inventory_category
    if product in INVENTORY_COUNTRY_HEADERS:
        return product, product, inventory_category

    source = source_country or _source_country_for_product(product, inventory_category, sheet)
    return source, product, inventory_category


def _mainstream_status(product: str, category: str = "") -> str:
    if (product, category) in MAINSTREAM_PRODUCT_CATEGORY_LABELS:
        return "主流"
    return "主流" if product in MAINSTREAM_PRODUCT_WILDCARDS else "非主流"


def _make_point(
    *,
    week_start: date,
    display_date: date,
    metric_type: str,
    source_country: str,
    product: str,
    category: str,
    value: float,
    source_file: str,
    source_sheet: str,
    source_section: str,
    is_calculable: bool,
    validation_status: str = "ok",
    note: str = "",
) -> Dict[str, Any]:
    business_week = compute_business_week(week_start)
    week_end = _week_end(week_start)
    return {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "business_year": business_week["year"],
        "business_week": business_week["week_no"],
        "week_label": f"{business_week['year']} W{business_week['week_no']:02d}",
        "display_date": display_date.isoformat(),
        "metric_type": metric_type,
        "source_country": source_country,
        "product": product,
        "category": category,
        "mainstream_status": _mainstream_status(product, category),
        "value": value,
        "unit": "万吨",
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_section": source_section,
        "is_calculable": 1 if is_calculable else 0,
        "validation_status": validation_status,
        "note": note,
    }


def _migrate_legacy_integrated_product_key(cur, point: Dict[str, Any]) -> None:
    canonical_identity = (point["source_country"], point["product"], point["category"])
    legacy_keys = [
        (metric_type, legacy_category, legacy_product)
        for (metric_type, legacy_category, legacy_product), renamed_identity
        in LEGACY_INVENTORY_IDENTITY_RENAME.items()
        if metric_type == point["metric_type"] and renamed_identity == canonical_identity
    ]
    if not legacy_keys:
        return

    canonical_params = (
        point["metric_type"],
        point["source_country"],
        point["product"],
        point["category"],
        point["mainstream_status"],
        point["business_year"],
        point["business_week"],
    )
    canonical = db._exec(
        cur,
        """SELECT id
           FROM dv_integrated_points
           WHERE metric_type = ?
             AND source_country = ?
             AND product = ?
             AND category = ?
             AND mainstream_status = ?
             AND business_year = ?
             AND business_week = ?
           ORDER BY id DESC
           LIMIT 1""",
        canonical_params,
    ).fetchone()

    legacy_rows = []
    for metric_type, legacy_category, legacy_product in legacy_keys:
        legacy_rows.extend(db._exec(
            cur,
            """SELECT id
               FROM dv_integrated_points
               WHERE metric_type = ?
                 AND product = ?
                 AND category = ?
                 AND business_year = ?
                 AND business_week = ?
               ORDER BY id DESC""",
            (
                metric_type,
                legacy_product,
                legacy_category,
                point["business_year"],
                point["business_week"],
            ),
        ).fetchall())
    if not legacy_rows:
        return

    legacy_ids = [row["id"] for row in legacy_rows]
    if canonical:
        placeholders = ",".join("?" for _ in legacy_ids)
        db._exec(cur, f"DELETE FROM dv_integrated_points WHERE id IN ({placeholders})", tuple(legacy_ids))
        return

    keep_id, *delete_ids = legacy_ids
    db._exec(
        cur,
        """UPDATE dv_integrated_points
           SET source_country = ?, product = ?, category = ?, mainstream_status = ?
           WHERE id = ?""",
        (point["source_country"], point["product"], point["category"], point["mainstream_status"], keep_id),
    )
    if delete_ids:
        placeholders = ",".join("?" for _ in delete_ids)
        db._exec(cur, f"DELETE FROM dv_integrated_points WHERE id IN ({placeholders})", tuple(delete_ids))


def _migrate_integrated_mainstream_status(cur, point: Dict[str, Any]) -> None:
    key_params = (
        point["metric_type"],
        point["source_country"],
        point["product"],
        point["category"],
        point["business_year"],
        point["business_week"],
    )
    canonical = db._exec(
        cur,
        """SELECT id
           FROM dv_integrated_points
           WHERE metric_type = ?
             AND source_country = ?
             AND product = ?
             AND category = ?
             AND business_year = ?
             AND business_week = ?
             AND mainstream_status = ?
           ORDER BY id DESC
           LIMIT 1""",
        (*key_params, point["mainstream_status"]),
    ).fetchone()
    old_status_rows = db._exec(
        cur,
        """SELECT id
           FROM dv_integrated_points
           WHERE metric_type = ?
             AND source_country = ?
             AND product = ?
             AND category = ?
             AND business_year = ?
             AND business_week = ?
             AND mainstream_status != ?
           ORDER BY id DESC""",
        (*key_params, point["mainstream_status"]),
    ).fetchall()
    if not old_status_rows:
        return

    old_status_ids = [row["id"] for row in old_status_rows]
    if canonical:
        placeholders = ",".join("?" for _ in old_status_ids)
        db._exec(cur, f"DELETE FROM dv_integrated_points WHERE id IN ({placeholders})", tuple(old_status_ids))
        return

    keep_id, *delete_ids = old_status_ids
    db._exec(
        cur,
        """UPDATE dv_integrated_points
           SET mainstream_status = ?
           WHERE id = ?""",
        (point["mainstream_status"], keep_id),
    )
    if delete_ids:
        placeholders = ",".join("?" for _ in delete_ids)
        db._exec(cur, f"DELETE FROM dv_integrated_points WHERE id IN ({placeholders})", tuple(delete_ids))


def _refresh_latest_integration_batch_counts(cur) -> None:
    batch = db._exec(
        cur,
        "SELECT id FROM dv_integration_batches ORDER BY created_at DESC, id DESC LIMIT 1",
    ).fetchone()
    if not batch:
        return
    counts = db._exec(
        cur,
        """SELECT COUNT(*) AS total,
                  SUM(CASE WHEN metric_type = 'apparent_demand' THEN 1 ELSE 0 END) AS apparent_demand_count
           FROM dv_integrated_points""",
    ).fetchone()
    db._exec(
        cur,
        "UPDATE dv_integration_batches SET point_count = ?, apparent_demand_count = ? WHERE id = ?",
        (counts["total"] or 0, counts["apparent_demand_count"] or 0, batch["id"]),
    )


def _sync_integrated_mainstream_statuses(cur) -> int:
    rows = db._exec(
        cur,
        """SELECT id, metric_type, source_country, product, category,
                  mainstream_status, business_year, business_week
           FROM dv_integrated_points
           ORDER BY id DESC""",
    ).fetchall()
    changed = 0
    changed_metrics = set()
    for raw_row in rows:
        row = _row_to_dict(raw_row)
        expected_status = _mainstream_status(row["product"], row["category"])
        if row["mainstream_status"] == expected_status:
            continue

        key_params = (
            row["metric_type"],
            row["source_country"],
            row["product"],
            row["category"],
            row["business_year"],
            row["business_week"],
            expected_status,
            row["id"],
        )
        canonical = db._exec(
            cur,
            """SELECT id
               FROM dv_integrated_points
               WHERE metric_type = ?
                 AND source_country = ?
                 AND product = ?
                 AND category = ?
                 AND business_year = ?
                 AND business_week = ?
                 AND mainstream_status = ?
                 AND id != ?
               ORDER BY id DESC
               LIMIT 1""",
            key_params,
        ).fetchone()
        if canonical:
            db._exec(cur, "DELETE FROM dv_integrated_points WHERE id = ?", (row["id"],))
        else:
            db._exec(
                cur,
                "UPDATE dv_integrated_points SET mainstream_status = ? WHERE id = ?",
                (expected_status, row["id"]),
            )
        changed += 1
        changed_metrics.add(row["metric_type"])

    if changed:
        if changed_metrics & {"inventory", "arrival", "apparent_demand"}:
            batch = db._exec(
                cur,
                "SELECT id FROM dv_integration_batches ORDER BY created_at DESC, id DESC LIMIT 1",
            ).fetchone()
            if batch:
                _recalculate_integrated_apparent_demand(cur, batch["id"])
        _refresh_latest_integration_batch_counts(cur)
    return changed


def _find_col(ws, header_row: int, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, col).value or "").strip() == header_name:
            return col
    return None


def _iter_section_rows(ws, header_row: int):
    row = header_row + 1
    while row <= ws.max_row:
        first_cell = ws.cell(row, 1).value
        if str(first_cell or "").strip() == "日期":
            break
        period_start = _parse_period_start(first_cell)
        if period_start is None:
            break
        yield row, period_start
        row += 1


def _iter_total_rows(ws) -> List[int]:
    return [row for row in range(1, ws.max_row + 1) if ws.cell(row, 2).value == "总计"]


def _extract_australia_arrivals(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "澳洲预计到达中国锚地量" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["澳洲预计到达中国锚地量"]
    points: List[Dict[str, Any]] = []
    header_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() != "日期":
            continue
        mapped_count = sum(
            1
            for col in range(2, ws.max_column + 1)
            if str(ws.cell(row, col).value or "").strip() in AUSTRALIA_PRODUCT_MAP
        )
        if mapped_count >= 3:
            header_row = row
            break
    if header_row is None:
        wb.close()
        return points
    headers = {col: str(ws.cell(header_row, col).value or "").strip() for col in range(2, ws.max_column + 1)}
    for row, period_start in _iter_section_rows(ws, header_row):
        for col, raw_product in headers.items():
            if not raw_product or raw_product == "总计":
                continue
            mapping = AUSTRALIA_PRODUCT_MAP.get(raw_product)
            if mapping is None:
                continue
            value = _clean_number(ws.cell(row, col).value)
            if value is None:
                continue
            product, category, _inventory_key = mapping
            points.append(_make_point(
                week_start=period_start,
                display_date=period_start,
                metric_type="arrival",
                source_country="澳洲",
                product=product,
                category=category,
                value=value,
                source_file=path.name,
                source_sheet="澳洲预计到达中国锚地量",
                source_section="预计到中国锚地量（品种）",
                is_calculable=True,
                note="澳洲直接采用预计到中国锚地量，归类为到港",
            ))
    wb.close()
    return points


def _extract_australia_shipments(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "澳洲发货量" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["澳洲发货量"]
    points: List[Dict[str, Any]] = []
    header_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() != "日期":
            continue
        mapped_count = sum(
            1
            for col in range(2, ws.max_column + 1)
            if str(ws.cell(row, col).value or "").strip() in AUSTRALIA_PRODUCT_MAP
        )
        if mapped_count >= 3:
            header_row = row
            break
    if header_row is None:
        wb.close()
        return points
    headers = {col: str(ws.cell(header_row, col).value or "").strip() for col in range(2, ws.max_column + 1)}
    for row, period_start in _iter_section_rows(ws, header_row):
        for col, raw_product in headers.items():
            if not raw_product or raw_product == "总计":
                continue
            mapping = AUSTRALIA_PRODUCT_MAP.get(raw_product)
            if mapping is None:
                continue
            value = _clean_number(ws.cell(row, col).value)
            if value is None:
                continue
            product, category, _inventory_key = mapping
            points.append(_make_point(
                week_start=period_start,
                display_date=period_start,
                metric_type="shipment",
                source_country="澳洲",
                product=product,
                category=category,
                value=value,
                source_file=path.name,
                source_sheet="澳洲发货量",
                source_section="澳洲发货量（分品种）",
                is_calculable=False,
                validation_status="record_only",
                note="澳洲发运原始值；到港和表需继续使用预计到中国锚地量",
            ))
    wb.close()
    return points


def _extract_brazil_estimated_arrivals(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "巴西发货量" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["巴西发货量"]
    points: List[Dict[str, Any]] = []
    destination_header_row = None
    china_col = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() != "日期":
            continue
        col = _find_col(ws, row, "中国大陆")
        if col is not None:
            destination_header_row = row
            china_col = col
            break
    if destination_header_row is None or china_col is None:
        wb.close()
        return points
    for row, period_start in _iter_section_rows(ws, destination_header_row):
        value = _clean_number(ws.cell(row, china_col).value)
        if value is None:
            continue
        arrival_week = period_start + timedelta(weeks=6)
        points.append(_make_point(
            week_start=arrival_week,
            display_date=arrival_week,
            metric_type="arrival",
            source_country="巴西",
            product="卡粉",
            category="粉矿",
            value=value * 0.75,
            source_file=path.name,
            source_sheet="巴西发货量",
            source_section="中国大陆发运量×75%（6周后到港估算）",
            is_calculable=True,
            note="巴西卡粉按中国大陆发运量的75%折算到港，滞后6周",
        ))
    wb.close()
    return points


def _extract_brazil_card_powder_shipments(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "巴西发货量" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["巴西发货量"]
    points: List[Dict[str, Any]] = []
    header_row = None
    card_col = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() != "日期":
            continue
        col = _find_col(ws, row, "卡粉")
        if col is not None:
            header_row = row
            card_col = col
            break
    if header_row is None or card_col is None:
        wb.close()
        return points
    for row, period_start in _iter_section_rows(ws, header_row):
        value = _clean_number(ws.cell(row, card_col).value)
        if value is None:
            continue
        points.append(_make_point(
            week_start=period_start,
            display_date=period_start,
            metric_type="shipment",
            source_country="巴西",
            product="卡粉",
            category="粉矿",
            value=value,
            source_file=path.name,
            source_sheet="巴西发货量",
            source_section="卡粉发运量",
            is_calculable=True,
            note="巴西卡粉采用Mysteel卡粉发运量",
        ))
    wb.close()
    return points


def _extract_global_shipments(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "全球铁矿石发运量" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["全球铁矿石发运量"]
    points: List[Dict[str, Any]] = []
    header_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() != "日期":
            continue
        headers = {str(ws.cell(row, col).value or "").strip() for col in range(2, ws.max_column + 1)}
        if {"澳大利亚", "巴西", "南非"}.issubset(headers):
            header_row = row
            break
    if header_row is None:
        wb.close()
        return points
    headers = {col: str(ws.cell(header_row, col).value or "").strip() for col in range(2, ws.max_column + 1)}
    for row, period_start in _iter_section_rows(ws, header_row):
        for col, country in headers.items():
            if not country or country in GLOBAL_SKIP_COUNTRIES:
                continue
            value = _clean_number(ws.cell(row, col).value)
            if value is None:
                continue
            points.append(_make_point(
                week_start=period_start,
                display_date=period_start,
                metric_type="shipment",
                source_country=country,
                product=country,
                category="全品种",
                value=value,
                source_file=path.name,
                source_sheet="全球铁矿石发运量",
                source_section="铁矿石全球发运量",
                is_calculable=False,
                validation_status="record_only",
                note="非澳巴全球发运只记录，不参与表需计算",
            ))
    wb.close()
    return points


def _inventory_product(sheet_name: str, header: str) -> Optional[tuple]:
    if not header or "总计" in header:
        return None
    source_country, product, category = _canonical_inventory_identity(sheet_name, header)
    return product, category, source_country


def _extract_inventory(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    points: List[Dict[str, Any]] = []
    for sheet_name in ["粗粉", "块矿", "球团", "精粉"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {col: str(ws.cell(1, col).value or "").strip() for col in range(4, ws.max_column + 1)}
        for row in _iter_total_rows(ws):
            raw_date = _to_date(ws.cell(row, 1).value)
            if raw_date is None:
                continue
            week_start = _week_start(raw_date)
            for col, header in headers.items():
                product_info = _inventory_product(sheet_name, header)
                if product_info is None:
                    continue
                value = _clean_number(ws.cell(row, col).value)
                if value is None:
                    continue
                product, category, source_country = product_info
                points.append(_make_point(
                    week_start=week_start,
                    display_date=raw_date,
                    metric_type="inventory",
                    source_country=source_country,
                    product=product,
                    category=category,
                    value=value,
                    source_file=path.name,
                    source_sheet=sheet_name,
                    source_section="总计行",
                    is_calculable=_is_apparent_demand_candidate(source_country, product, category),
                    validation_status="ok",
                ))
    wb.close()
    return points


def _summarize_points(points: List[Dict[str, Any]], warnings: List[str]) -> Dict[str, Any]:
    metrics: Dict[str, int] = {}
    products: set = set()
    weeks: set = set()
    for point in points:
        metrics[point["metric_type"]] = metrics.get(point["metric_type"], 0) + 1
        products.add(point["product"])
        weeks.add(point["week_start"])
    return {
        "total_points": len(points),
        "metrics": metrics,
        "product_count": len(products),
        "week_count": len(weeks),
        "warnings": warnings,
        "samples": points[:20],
    }


def _add_apparent_demand(points: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = list(points)
    inv: Dict[tuple, Dict[str, Any]] = {}
    ship: Dict[tuple, Dict[str, Any]] = {}
    for point in points:
        key = (point["week_start"], point["source_country"], point["product"], point["category"])
        if point["metric_type"] == "inventory":
            inv[key] = point
        elif point["metric_type"] == "arrival" and point["is_calculable"]:
            ship[key] = point

    product_keys = sorted({(p["source_country"], p["product"], p["category"]) for p in points if p["is_calculable"]})
    for source_country, product, category in product_keys:
        week_starts = sorted({
            p["week_start"] for p in points
            if p["source_country"] == source_country and p["product"] == product and p["category"] == category
        })
        for index, week_start in enumerate(week_starts):
            cur_key = (week_start, source_country, product, category)
            if cur_key not in ship or cur_key not in inv or index == 0:
                continue
            prev_week = week_starts[index - 1]
            prev_key = (prev_week, source_country, product, category)
            if prev_key not in inv:
                continue
            shipment_value = float(ship[cur_key]["value"] or 0)
            inv_prev = float(inv[prev_key]["value"] or 0)
            inv_cur = float(inv[cur_key]["value"] or 0)
            value = shipment_value + inv_prev - inv_cur
            base = ship[cur_key]
            result.append(_make_point(
                week_start=date.fromisoformat(week_start),
                display_date=date.fromisoformat(week_start),
                metric_type="apparent_demand",
                source_country=source_country,
                product=product,
                category=category,
                value=value,
                source_file=base["source_file"],
                source_sheet="系统计算",
                source_section="表需",
                is_calculable=True,
                validation_status="ok",
                note="表需=到港/估算到港+上周库存-本周库存",
            ))
    return result


def integrate_mysteel_files(file_paths: List[Path]) -> Dict[str, Any]:
    points: List[Dict[str, Any]] = []
    warnings: List[str] = []
    used = {"australia": False, "brazil": False, "global": False, "inventory": False}
    for path in file_paths:
        if not path.exists():
            warnings.append(f"文件不存在: {path.name}")
            continue
        australia_shipments = _extract_australia_shipments(path)
        australia = _extract_australia_arrivals(path)
        brazil = _extract_brazil_estimated_arrivals(path)
        global_points = _extract_global_shipments(path)
        inventory = _extract_inventory(path)
        if australia or australia_shipments:
            used["australia"] = True
        if brazil:
            used["brazil"] = True
        if global_points:
            used["global"] = True
        if inventory:
            used["inventory"] = True
        points.extend(australia_shipments)
        points.extend(australia)
        points.extend(brazil)
        points.extend(_extract_brazil_card_powder_shipments(path))
        points.extend(global_points)
        points.extend(inventory)
    for key, label in [
        ("australia", "澳洲到港"),
        ("brazil", "巴西中国大陆发运"),
        ("global", "全球其他国家发运"),
        ("inventory", "库存明细"),
    ]:
        if not used[key]:
            warnings.append(f"未识别到{label}数据")
    points = _add_apparent_demand(points)
    return {"points": points, "summary": _summarize_points(points, warnings)}


def _local_mysteel_files() -> List[Path]:
    return sorted(LOCAL_MYSTEEL_DIR.glob("Mysteel*.xlsx"))


def _write_uploads_to_tmp(files: List[IntegrationUploadFile]) -> List[Path]:
    paths: List[Path] = []
    for item in files:
        suffix = ".xlsx" if item.file_name.lower().endswith(".xlsx") else ".xls"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(base64.b64decode(item.file_data))
            paths.append(Path(tmp.name))
    return paths


def _cleanup_tmp(paths: List[Path]) -> None:
    for path in paths:
        try:
            path.unlink()
        except OSError:
            pass


def _split_filter_values(value: str) -> List[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


AGGREGATE_PRODUCT_ORDER = ["主流矿合计", "非主流矿合计"]
AGGREGATE_PRODUCT_STATUS = {
    "主流矿合计": "主流",
    "非主流矿合计": "非主流",
}


def _aggregate_labels_and_statuses(product_list: List[str], mainstream_list: List[str]) -> tuple:
    labels = [
        label for label in AGGREGATE_PRODUCT_ORDER
        if not product_list or label in product_list
    ]
    if product_list and not labels:
        return [], []
    pairs = [(label, AGGREGATE_PRODUCT_STATUS[label]) for label in labels]
    if mainstream_list:
        pairs = [(label, status) for label, status in pairs if status in mainstream_list]
    return [label for label, _status in pairs], [status for _label, status in pairs]


def _effective_mainstream_statuses(product_pool: str, mainstream_list: List[str]) -> List[str]:
    if product_pool == "mainstream":
        return ["主流"]
    if product_pool == "non_mainstream":
        return ["非主流"]
    return mainstream_list


def _ordered_mainstream_labels(labels: List[str], label_map: Dict[tuple, str]) -> List[str]:
    original_index = {label: index for index, label in enumerate(labels)}
    identity_label_by_label: Dict[str, str] = {}
    for (_source_country, product, category), label in label_map.items():
        identity_label_by_label.setdefault(
            label,
            MAINSTREAM_PRODUCT_CATEGORY_LABELS.get((product, category), product),
        )
    return sorted(
        labels,
        key=lambda label: (
            MAINSTREAM_PRODUCT_ORDER_INDEX.get(
                label,
                MAINSTREAM_PRODUCT_ORDER_INDEX.get(
                    identity_label_by_label.get(label, label),
                    len(MAINSTREAM_PRODUCT_ORDER),
                ),
            ),
            original_index[label],
        ),
    )


def _uses_mainstream_product_order(product_pool: str, mainstream_list: List[str]) -> bool:
    return product_pool == "mainstream" or set(mainstream_list) == {"主流"}


def _integrated_product_labels(rows: List[Dict[str, Any]]) -> tuple:
    product_categories: Dict[str, set] = {}
    product_category_sources: Dict[tuple, set] = {}
    for row in rows:
        product_categories.setdefault(row["product"], set()).add(row["category"])
        product_category_sources.setdefault((row["product"], row["category"]), set()).add(row["source_country"])

    label_map: Dict[tuple, str] = {}
    products_ordered: List[str] = []
    for row in rows:
        key = (row["source_country"], row["product"], row["category"])
        if key in label_map:
            continue
        source_country, product, category = key
        if product == "其他":
            if source_country == "其他":
                label = f"其他（{category}）"
            else:
                label = f"{source_country}（其他{category}）"
        elif source_country == product and category in INVENTORY_CATEGORIES | {"全品种"}:
            label = f"{product}（{category}）"
        elif len(product_category_sources[(product, category)]) > 1:
            label = f"{source_country} / {product}（{category}）"
        elif len(product_categories[product]) > 1:
            label = f"{product}（{category}）"
        else:
            label = product
        label_map[key] = label
        products_ordered.append(label)
    return label_map, products_ordered


def _integrated_product_pool_labels(rows: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    label_map, products_ordered = _integrated_product_labels(rows)
    status_by_label: Dict[str, set] = {}
    for row in rows:
        label = label_map[(row["source_country"], row["product"], row["category"])]
        status_by_label.setdefault(label, set()).add(row["mainstream_status"] or "非主流")
    mainstream_labels = [label for label in products_ordered if "主流" in status_by_label.get(label, set())]
    return {
        "mainstream": _ordered_mainstream_labels(mainstream_labels, label_map),
        "non_mainstream": [label for label in products_ordered if "非主流" in status_by_label.get(label, set())],
        "custom": products_ordered,
    }


def _filter_rows_by_product_labels(rows: List[Dict[str, Any]], product_list: List[str]) -> tuple:
    label_map, products_ordered = _integrated_product_labels(rows)
    if not product_list:
        return rows, label_map, products_ordered

    selected_rows = []
    selected_labels = set()
    for row in rows:
        key = (row["source_country"], row["product"], row["category"])
        label = label_map[key]
        if label in product_list or row["product"] in product_list:
            selected_rows.append(row)
            selected_labels.add(label)
    return (
        selected_rows,
        label_map,
        [label for label in products_ordered if label in selected_labels],
    )


def _parse_integrated_excel(file_path):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet_name = '整合明细'
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {'rows': [], 'errors': [{'row': 0, 'message': f'未找到「{sheet_name}」sheet'}], 'summary': {}}

    ws = wb[sheet_name]
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [str(hval).strip() if hval is not None else '' for hval in header_values]

    col_map = {}
    for idx, header in enumerate(headers):
        field = _INTEGRATED_FIELDS.get(header)
        if field:
            col_map[idx + 1] = field

    if not col_map:
        wb.close()
        return {'rows': [], 'errors': [{'row': 0, 'message': '整合明细表头无法识别'}], 'summary': {}}

    rows = []
    errors = []
    seen_keys = set()
    duplicate_key_count = 0
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        value_parse_error = False
        for col_idx, field in col_map.items():
            raw = row_values[col_idx - 1] if col_idx <= len(row_values) else None
            if field == 'business_year':
                try:
                    row_data[field] = int(raw) if raw is not None else None
                except (ValueError, TypeError):
                    row_data[field] = None
            elif field == 'business_week':
                try:
                    row_data[field] = int(raw) if raw is not None else None
                except (ValueError, TypeError):
                    row_data[field] = None
            elif field == 'is_calculable':
                val = str(raw).strip() if raw is not None else ''
                row_data[field] = 1 if val == '是' else 0
            elif field == 'value':
                try:
                    row_data[field] = float(raw) if raw is not None else None
                except (ValueError, TypeError):
                    row_data[field] = None
                    value_parse_error = True
            elif field in ('week_start', 'week_end', 'display_date'):
                row_data[field] = _normalize_date_value(raw)
            elif field == 'metric_type':
                val = str(raw).strip() if raw else ''
                row_data[field] = _METRIC_TYPE_CN.get(val, val)
            else:
                row_data[field] = str(raw).strip() if raw else ''

        missing = [f for f in _REQUIRED_INTEGRATED_FIELDS if not row_data.get(f)]
        if missing:
            errors.append({'row': row_idx, 'message': f'缺少必填字段: {", ".join(missing)}'})
            continue
        if row_data.get('metric_type') not in _VALID_INTEGRATED_METRIC_TYPES:
            errors.append({'row': row_idx, 'message': f'数据类型无效: {row_data.get("metric_type")}'})
            continue
        if value_parse_error or row_data.get('value') is None:
            errors.append({'row': row_idx, 'message': '数值不是数字'})
            continue
        row_data = _normalize_integrated_point(row_data)
        business_key = (
            row_data.get('week_start'),
            row_data.get('metric_type'),
            row_data.get('source_country'),
            row_data.get('product'),
            row_data.get('category'),
        )
        if business_key in seen_keys:
            duplicate_key_count += 1
        else:
            seen_keys.add(business_key)
        rows.append(row_data)

    summary = _summarize_integrated_rows(rows)
    summary['duplicate_key_count'] = duplicate_key_count
    wb.close()
    return {'rows': rows, 'errors': errors, 'summary': summary}


def _summarize_integrated_rows(rows):
    if not rows:
        return {
            'total_points': 0, 'inventory_count': 0, 'shipment_count': 0,
            'arrival_count': 0, 'apparent_demand_count': 0, 'product_count': 0, 'category_count': 0,
            'country_count': 0, 'week_count': 0, 'null_count': 0,
            'date_min': '', 'date_max': '', 'years': [],
        }
    products = set()
    categories = set()
    countries = set()
    weeks = set()
    metric_counts = {'inventory': 0, 'shipment': 0, 'arrival': 0, 'apparent_demand': 0}
    null_count = 0
    dates = []
    for row in rows:
        products.add(row.get('product', ''))
        categories.add(row.get('category', ''))
        countries.add(row.get('source_country', ''))
        weeks.add(row.get('week_start', ''))
        mt = row.get('metric_type', '')
        if mt in metric_counts:
            metric_counts[mt] += 1
        if row.get('value') is None:
            null_count += 1
        d = row.get('display_date', '')
        if d:
            dates.append(d)
    years = sorted(set(w[:4] for w in weeks if w and len(w) >= 4))
    return {
        'total_points': len(rows),
        'inventory_count': metric_counts['inventory'],
        'shipment_count': metric_counts['shipment'],
        'arrival_count': metric_counts['arrival'],
        'apparent_demand_count': metric_counts['apparent_demand'],
        'product_count': len([p for p in products if p]),
        'category_count': len([c for c in categories if c]),
        'country_count': len([c for c in countries if c]),
        'week_count': len([w for w in weeks if w]),
        'null_count': null_count,
        'date_min': min(dates) if dates else '',
        'date_max': max(dates) if dates else '',
        'years': years,
    }


def _integrated_preview_cache_path(preview_id: str) -> Path:
    safe_id = "".join(ch for ch in preview_id if ch.isalnum() or ch in ("-", "_"))
    if not safe_id:
        raise HTTPException(status_code=400, detail="预检缓存编号无效")
    return _INTEGRATED_PREVIEW_CACHE_DIR / f"{safe_id}.json"


def _integrated_preview_file_path(preview_id: str, file_name: str) -> Path:
    safe_id = "".join(ch for ch in preview_id if ch.isalnum() or ch in ("-", "_"))
    suffix = Path(file_name or "").suffix.lower() or ".xlsx"
    if suffix not in {".xlsx", ".xls"}:
        suffix = ".xlsx"
    if not safe_id:
        raise HTTPException(status_code=400, detail="预检缓存编号无效")
    return _INTEGRATED_PREVIEW_FILE_DIR / f"{safe_id}{suffix}"


def _save_integrated_preview_cache(result, file_name, source_path: Optional[str] = None):
    _INTEGRATED_PREVIEW_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    preview_id = uuid.uuid4().hex
    path = _integrated_preview_cache_path(preview_id)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(
            {
                "file_name": file_name,
                "errors": result["errors"],
                "summary": result["summary"],
                "sample_rows": result["rows"][:20],
                "source_path": source_path,
            },
            fh,
            ensure_ascii=False,
        )
    return preview_id


def _load_integrated_preview_cache(preview_id: str):
    path = _integrated_preview_cache_path(preview_id)
    if not path.exists():
        raise HTTPException(status_code=400, detail="预检结果已失效，请重新选择文件预检")
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


async def _save_integrated_preview_upload(request: Request, file_name: str) -> Path:
    _INTEGRATED_PREVIEW_FILE_DIR.mkdir(parents=True, exist_ok=True)
    preview_id = uuid.uuid4().hex
    path = _integrated_preview_file_path(preview_id, file_name)
    with path.open("wb") as fh:
        async for chunk in request.stream():
            if chunk:
                fh.write(chunk)
    if path.stat().st_size == 0:
        path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="请先选择整合 Excel 文件")
    return path


def _run_integrated_preview_job(job_id: str, source_path: str, file_name: str) -> None:
    job = _INTEGRATED_PREVIEW_JOBS[job_id]
    try:
        job.update({"status": "running", "stage": "parsing", "message": "正在解析 Excel"})
        result = _parse_integrated_excel(Path(source_path))
        preview_id = _save_integrated_preview_cache(result, file_name, source_path)
        job.update({
            "status": "succeeded",
            "stage": "done",
            "message": "预检完成",
            "preview_id": preview_id,
            "file_name": file_name,
            "summary": result["summary"],
            "errors": result["errors"],
            "sample_count": min(len(result["rows"]), 20),
            "sample_rows": result["rows"][:20],
            "finished_at": datetime.utcnow().isoformat(),
        })
    except Exception as exc:
        job.update({
            "status": "failed",
            "stage": "error",
            "message": str(exc),
            "finished_at": datetime.utcnow().isoformat(),
        })


def _create_integrated_preview_job(
    background_tasks: BackgroundTasks,
    source_path: str,
    file_name: str,
) -> Dict[str, Any]:
    job_id = uuid.uuid4().hex
    job = {
        "job_id": job_id,
        "status": "queued",
        "stage": "queued",
        "message": "已上传，等待后台预检",
        "file_name": file_name,
        "created_at": datetime.utcnow().isoformat(),
    }
    _INTEGRATED_PREVIEW_JOBS[job_id] = job
    background_tasks.add_task(_run_integrated_preview_job, job_id, source_path, file_name)
    return job



def _import_integrated_points(rows, file_name, user_name):
    """Replace integrated data with the uploaded standard Excel in one batch."""
    rows = [_normalize_integrated_point(row) for row in rows]
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(cur, "DELETE FROM dv_integrated_points")
        db._exec(cur, "DELETE FROM dv_integration_batches")
        batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'committed', ?, ?, ?, ?)""",
            (
                file_name,
                len(rows),
                sum(1 for r in rows if r.get('metric_type') == 'apparent_demand'),
                json.dumps({"source": "integrated_import", "inserted": len(rows)}, ensure_ascii=False),
                user_name,
            ),
        )
        values = [
            (
                batch_id,
                row['week_start'],
                row.get('week_end', ''),
                row.get('business_year'),
                row.get('business_week'),
                row.get('week_label', ''),
                row.get('display_date', ''),
                row['metric_type'],
                row['source_country'],
                row['product'],
                row['category'],
                row.get('mainstream_status', ''),
                row.get('value'),
                row.get('unit', '万吨'),
                row.get('source_file', ''),
                row.get('source_sheet', ''),
                row.get('source_section', ''),
                row.get('is_calculable', 0),
                row.get('validation_status', ''),
                row.get('note', ''),
            )
            for row in rows
        ]
        insert_sql = """INSERT INTO dv_integrated_points
           (batch_id, week_start, week_end, business_year, business_week,
            week_label, display_date, metric_type, source_country, product,
            category, mainstream_status, value, unit, source_file,
            source_sheet, source_section, is_calculable, validation_status, note)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
        if values:
            if db._is_pg():
                from psycopg2.extras import execute_values
                execute_values(
                    cur,
                    insert_sql.replace("?", "%s").replace(
                        "(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        "%s",
                    ),
                    values,
                    page_size=1000,
                )
            else:
                db._executemany(cur, insert_sql, values)
        conn.commit()
    return batch_id


def _run_integrated_import_job(job_id: str, source_path: str, file_name: str, user_name: str) -> None:
    job = _INTEGRATED_IMPORT_JOBS[job_id]
    try:
        job.update({"status": "running", "stage": "parsing", "message": "正在解析 Excel"})
        result = _parse_integrated_excel(Path(source_path))
        if result["errors"]:
            job.update({
                "status": "failed",
                "stage": "validation",
                "message": f"整合 Excel 存在 {len(result['errors'])} 条错误，无法导入",
                "errors": result["errors"][:50],
                "summary": result["summary"],
                "finished_at": datetime.utcnow().isoformat(),
            })
            return

        job.update({
            "stage": "importing",
            "message": f"正在写入 {len(result['rows'])} 条数据",
            "summary": result["summary"],
        })
        batch_id = _import_integrated_points(result["rows"], file_name, user_name)
        job.update({
            "status": "succeeded",
            "stage": "done",
            "message": f"已导入 {len(result['rows'])} 条数据",
            "batch_id": batch_id,
            "summary": result["summary"],
            "finished_at": datetime.utcnow().isoformat(),
        })
    except Exception as exc:
        job.update({
            "status": "failed",
            "stage": "error",
            "message": str(exc),
            "finished_at": datetime.utcnow().isoformat(),
        })


def _create_integrated_import_job(
    background_tasks: BackgroundTasks,
    source_path: str,
    file_name: str,
    user_name: str,
) -> Dict[str, Any]:
    job_id = uuid.uuid4().hex
    job = {
        "job_id": job_id,
        "status": "queued",
        "stage": "queued",
        "message": "已加入后台导入队列",
        "file_name": file_name,
        "created_at": datetime.utcnow().isoformat(),
    }
    _INTEGRATED_IMPORT_JOBS[job_id] = job
    background_tasks.add_task(_run_integrated_import_job, job_id, source_path, file_name, user_name)
    return job


def _recalculate_integrated_apparent_demand(cur, batch_id: int) -> int:
    db._exec(cur, "DELETE FROM dv_integrated_points WHERE metric_type = 'apparent_demand'")
    source_rows = db._exec(
        cur,
        """SELECT *
           FROM dv_integrated_points
           WHERE metric_type IN ('inventory', 'arrival')
             AND is_calculable = 1
           ORDER BY source_country, product, category, business_year, business_week, id""",
    ).fetchall()

    by_series: Dict[tuple, Dict[tuple, Dict[str, Any]]] = {}
    for row in source_rows:
        item = _normalize_integrated_point(_row_to_dict(row))
        series_key = (item["source_country"], item["product"], item["category"], item["mainstream_status"])
        week_key = (item["business_year"], item["business_week"])
        by_series.setdefault(series_key, {})[(week_key, item["metric_type"])] = item

    inserted = 0
    for values in by_series.values():
        previous_inventory = None
        for week_key in sorted({week_key for week_key, _metric_type in values.keys()}):
            inventory = values.get((week_key, "inventory"))
            arrival = values.get((week_key, "arrival"))
            if inventory and arrival and previous_inventory:
                value = float(arrival["value"] or 0) + float(previous_inventory["value"] or 0) - float(inventory["value"] or 0)
                base = arrival
                db._exec(
                    cur,
                    """INSERT INTO dv_integrated_points
                       (batch_id, week_start, week_end, business_year, business_week, week_label,
                        display_date, metric_type, source_country, product,
                        category, mainstream_status, value, unit, source_file, source_sheet,
                        source_section, is_calculable, validation_status, note)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        batch_id,
                        base["week_start"],
                        base["week_end"],
                        base["business_year"],
                        base["business_week"],
                        base["week_label"],
                        base["week_start"],
                        "apparent_demand",
                        base["source_country"],
                        base["product"],
                        base["category"],
                        base["mainstream_status"],
                        value,
                        base.get("unit") or "万吨",
                        "系统计算",
                        "表需",
                        "表需",
                        1,
                        "ok",
                        "表需=到港+上周库存-本周库存",
                    ),
                )
                inserted += 1
            if inventory:
                previous_inventory = inventory
    return inserted


def _save_integrated_points(points: List[Dict[str, Any]], file_names: List[str], user_name: str) -> int:
    points = [_normalize_integrated_point(point) for point in points]
    with db.connect() as conn:
        cur = conn.cursor()
        batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_integration_batches
               (file_names, status, point_count, apparent_demand_count, validation_summary, created_by)
               VALUES (?, 'completed', ?, ?, ?, ?)""",
            (
                ", ".join(file_names),
                0,
                sum(1 for point in points if point["metric_type"] == "apparent_demand"),
                json.dumps({"source": "mysteel", "inserted": 0, "updated": 0, "skipped": 0}, ensure_ascii=False),
                user_name,
            ),
        )
        merge_summary = {
            "source": "mysteel",
            "input_points": len(points),
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "skipped_same_value": 0,
            "skipped_blank_overwrite": 0,
        }
        for point in points:
            _migrate_legacy_integrated_product_key(cur, point)
            _migrate_integrated_mainstream_status(cur, point)
            key_params = (
                point["metric_type"],
                point["source_country"],
                point["product"],
                point["category"],
                point["mainstream_status"],
                point["business_year"],
                point["business_week"],
            )
            existing = db._exec(
                cur,
                """SELECT *
                   FROM dv_integrated_points
                   WHERE metric_type = ?
                     AND source_country = ?
                     AND product = ?
                     AND category = ?
                     AND mainstream_status = ?
                     AND business_year = ?
                     AND business_week = ?
                   ORDER BY id DESC
                   LIMIT 1""",
                key_params,
            ).fetchone()
            new_value = point.get("value")
            if existing:
                old_value = existing["value"]
                if new_value is None and old_value is not None:
                    merge_summary["skipped"] += 1
                    merge_summary["skipped_blank_overwrite"] += 1
                    continue
                if new_value == old_value:
                    merge_summary["skipped"] += 1
                    merge_summary["skipped_same_value"] += 1
                    continue
                db._exec(
                    cur,
                    """UPDATE dv_integrated_points
                       SET batch_id = ?, week_start = ?, week_end = ?, week_label = ?,
                           display_date = ?, value = ?, unit = ?, source_file = ?,
                           source_sheet = ?, source_section = ?, is_calculable = ?,
                           validation_status = ?, note = ?
                       WHERE id = ?""",
                    (
                        batch_id,
                        point["week_start"],
                        point["week_end"],
                        point["week_label"],
                        point["display_date"],
                        new_value,
                        point["unit"],
                        point["source_file"],
                        point["source_sheet"],
                        point["source_section"],
                        point["is_calculable"],
                        point["validation_status"],
                        point["note"],
                        existing["id"],
                    ),
                )
                merge_summary["updated"] += 1
                continue
            db._exec(
                cur,
                """INSERT INTO dv_integrated_points
                   (batch_id, week_start, week_end, business_year, business_week, week_label,
                    display_date, metric_type, source_country, product,
                    category, mainstream_status, value, unit, source_file, source_sheet,
                    source_section, is_calculable, validation_status, note)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    batch_id,
                    point["week_start"],
                    point["week_end"],
                    point["business_year"],
                    point["business_week"],
                    point["week_label"],
                    point["display_date"],
                    point["metric_type"],
                    point["source_country"],
                    point["product"],
                    point["category"],
                    point["mainstream_status"],
                    point["value"],
                    point["unit"],
                    point["source_file"],
                    point["source_sheet"],
                    point["source_section"],
                    point["is_calculable"],
                    point["validation_status"],
                    point["note"],
                ),
            )
            merge_summary["inserted"] += 1
        apparent_demand_count = _recalculate_integrated_apparent_demand(cur, batch_id)
        current_total = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integrated_points").fetchone()["c"]
        db._exec(
            cur,
            """UPDATE dv_integration_batches
               SET point_count = ?, apparent_demand_count = ?, validation_summary = ?
               WHERE id = ?""",
            (current_total, apparent_demand_count, json.dumps(merge_summary, ensure_ascii=False), batch_id),
        )
        conn.commit()
    return batch_id


INTEGRATED_EXPORT_COLUMNS = [
    ("week_start", "统计周一"),
    ("week_end", "统计周日"),
    ("business_year", "业务年份"),
    ("business_week", "业务周次"),
    ("week_label", "周次标签"),
    ("display_date", "展示日期"),
    ("metric_type", "数据类型"),
    ("source_country", "来源/国家"),
    ("product", "品种"),
    ("category", "种类"),
    ("mainstream_status", "主流/非主流"),
    ("value", "数值"),
    ("unit", "单位"),
    ("source_file", "来源文件"),
    ("source_sheet", "来源Sheet"),
    ("source_section", "来源区域"),
    ("is_calculable", "是否参与表需"),
    ("validation_status", "校验状态"),
    ("note", "备注"),
]

METRIC_SHEETS = [
    ("shipment", "发运"),
    ("arrival", "到港"),
    ("inventory", "库存"),
    ("apparent_demand", "表需"),
]


def _metric_label(metric: str) -> str:
    if metric == "inventory":
        return "库存"
    if metric == "shipment":
        return "发运"
    if metric == "arrival":
        return "到港"
    if metric == "apparent_demand":
        return "表需"
    return metric


def _export_cell_value(item: Dict[str, Any], key: str) -> Any:
    if key == "metric_type":
        return _metric_label(item[key])
    if key == "is_calculable":
        return "是" if item[key] else "否"
    return item[key]


def _ensure_week_fields(item: Dict[str, Any]) -> Dict[str, Any]:
    item = _normalize_integrated_point(item)
    if item.get("business_year") and item.get("business_week") and item.get("week_label") and item.get("week_end"):
        return item
    week_start_value = item.get("week_start")
    if week_start_value:
        week_start_date = date.fromisoformat(week_start_value)
        business_week = compute_business_week(week_start_date)
        item["week_end"] = item.get("week_end") or _week_end(week_start_date).isoformat()
        item["business_year"] = item.get("business_year") or business_week["year"]
        item["business_week"] = item.get("business_week") or business_week["week_no"]
        item["week_label"] = item.get("week_label") or f"{business_week['year']} W{business_week['week_no']:02d}"
    return item


def _append_integrated_rows(sheet, rows: List[Any], include_metric_type: bool = True, header_style=None) -> None:
    columns = INTEGRATED_EXPORT_COLUMNS if include_metric_type else [
        (key, label) for key, label in INTEGRATED_EXPORT_COLUMNS if key != "metric_type"
    ]
    header_labels = [label for _key, label in columns]
    if header_style:
        from openpyxl.cell import WriteOnlyCell
        header_font, header_fill = header_style
        header_cells = []
        for label in header_labels:
            cell = WriteOnlyCell(sheet, value=label)
            cell.font = header_font
            cell.fill = header_fill
            header_cells.append(cell)
        sheet.append(header_cells)
    else:
        sheet.append(header_labels)
    for row in rows:
        item = _ensure_week_fields(_row_to_dict(row))
        sheet.append([_export_cell_value(item, key) for key, _label in columns])


def _load_integration_batch_summary(batch_id: int) -> Dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        batch = db._exec(
            cur,
            "SELECT * FROM dv_integration_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        rows = db._exec(
            cur,
            "SELECT * FROM dv_integrated_points ORDER BY week_start, metric_type, source_country, category, product, id",
        ).fetchall()
    summary = _summarize_integrated_rows([_row_to_dict(row) for row in rows])
    merge_summary = {}
    if batch and batch["validation_summary"]:
        try:
            merge_summary = json.loads(batch["validation_summary"])
        except json.JSONDecodeError:
            merge_summary = {}
    return {
        "batch": _row_to_dict(batch),
        "summary": summary,
        "merge_summary": merge_summary,
    }


def build_integrated_workbook_bytes() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    with db.connect() as conn:
        cur = conn.cursor()
        batch = db._exec(
            cur,
            "SELECT * FROM dv_integration_batches ORDER BY created_at DESC, id DESC LIMIT 1",
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="暂无可导出的整合结果")
        _sync_integrated_mainstream_statuses(cur)
        batch = db._exec(
            cur,
            "SELECT * FROM dv_integration_batches ORDER BY created_at DESC, id DESC LIMIT 1",
        ).fetchone()
        rows = db._exec(
            cur,
            """SELECT *
               FROM dv_integrated_points
               ORDER BY week_start, metric_type, source_country, category, product, id""",
        ).fetchall()
        metrics = db._exec(
            cur,
            """SELECT metric_type, COUNT(*) AS c
               FROM dv_integrated_points
               GROUP BY metric_type""",
        ).fetchall()

    wb = openpyxl.Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_style = (header_font, header_fill)

    ws = wb.create_sheet("整合明细")
    _append_integrated_rows(ws, rows, include_metric_type=True, header_style=header_style)

    metric_sheets = []
    for metric_type, sheet_name in METRIC_SHEETS:
        sheet = wb.create_sheet(sheet_name)
        metric_rows = (row for row in rows if row["metric_type"] == metric_type)
        _append_integrated_rows(sheet, metric_rows, include_metric_type=False, header_style=header_style)
        metric_sheets.append(sheet)

    info = wb.create_sheet("批次信息")
    metric_counts = {row["metric_type"]: row["c"] for row in metrics}
    info_rows = [
        ("批次ID", batch["id"]),
        ("来源文件", batch["file_names"]),
        ("状态", batch["status"]),
        ("数据点总数", batch["point_count"]),
        ("库存条数", metric_counts.get("inventory", 0)),
        ("发运条数", metric_counts.get("shipment", 0)),
        ("到港条数", metric_counts.get("arrival", 0)),
        ("表需条数", metric_counts.get("apparent_demand", 0)),
        ("创建人", batch["created_by"]),
        ("创建时间", batch["created_at"]),
    ]
    for item in info_rows:
        info.append(item)

    width_map = {
        "A": 14, "B": 14, "C": 10, "D": 10, "E": 12, "F": 14,
        "G": 12, "H": 16, "I": 18, "J": 12, "K": 14, "L": 12,
        "M": 10, "N": 24, "O": 18, "P": 18, "Q": 14, "R": 12, "S": 24,
    }
    for sheet in [ws, *metric_sheets]:
        sheet.freeze_panes = "A2"
        for col, width in width_map.items():
            sheet.column_dimensions[col].width = width
    info.freeze_panes = "A2"
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 48

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


# ── 业务周计算 ────────────────────────────────────────────────────────
# 规则：周一为起点；1 月 1 日所在周为第 1 周


def compute_business_week(d: date) -> Dict[str, Any]:
    """计算业务周：周一起始，1月1日所在周为 W01。

    修正跨年规则：若该周包含下一年的 1 月 1 日，则归属下一年 W01。
    例：2024-12-30 → 2025 W01；2022-12-26 → 2023 W01。
    """
    weekday = d.weekday()  # 0=Mon … 6=Sun
    week_start = d - timedelta(days=weekday)
    week_end = week_start + timedelta(days=6)

    jan1 = date(d.year, 1, 1)
    jan1_weekday = jan1.weekday()
    jan1_week_start = jan1 - timedelta(days=jan1_weekday)

    if week_start < jan1_week_start:
        prev_jan1 = date(d.year - 1, 1, 1)
        prev_jan1_weekday = prev_jan1.weekday()
        prev_jan1_week_start = prev_jan1 - timedelta(days=prev_jan1_weekday)
        week_no = ((week_start - prev_jan1_week_start).days // 7) + 1
        result = {
            "year": d.year - 1,
            "week_no": week_no,
            "week_start_date": week_start.isoformat(),
            "week_end_date": week_end.isoformat(),
        }
    else:
        week_no = ((week_start - jan1_week_start).days // 7) + 1
        result = {
            "year": d.year,
            "week_no": week_no,
            "week_start_date": week_start.isoformat(),
            "week_end_date": week_end.isoformat(),
        }

    # 跨年修正：若该周包含下一年的 1 月 1 日，则归属下一年 W01
    next_jan1 = date(d.year + 1, 1, 1)
    if week_end >= next_jan1:
        result["year"] = d.year + 1
        result["week_no"] = 1

    return result

# ── Excel 解析 ────────────────────────────────────────────────────────


def parse_excel(file_path: str) -> Dict[str, List[Dict]]:
    """读取发运 sheet 和卡粉库存 sheet，返回 {shipment: [...], inventory: [...]}。"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    result: Dict[str, List[Dict]] = {"shipment": [], "inventory": []}

    if SHIPMENT_SHEET in wb.sheetnames:
        ws = wb[SHIPMENT_SHEET]
        for row_idx in range(2, ws.max_row + 1):
            d = _to_date(ws.cell(row=row_idx, column=SHIPMENT_DATE_COL).value)
            v = _to_float(ws.cell(row=row_idx, column=SHIPMENT_VAL_COL).value)
            if d:
                result["shipment"].append({"date": d.isoformat(), "value": v})

    if INVENTORY_SHEET in wb.sheetnames:
        ws = wb[INVENTORY_SHEET]
        for row_idx in range(2, ws.max_row + 1):
            d = _to_date(ws.cell(row=row_idx, column=INVENTORY_DATE_COL).value)
            v = _to_float(ws.cell(row=row_idx, column=INVENTORY_VAL_COL).value)
            if d:
                result["inventory"].append({"date": d.isoformat(), "value": v})

    wb.close()
    return result


# ── 周匹配合并 ────────────────────────────────────────────────────────
# 同周且日期差 ≤ 4 天 → 共享 week_key


def _upsert_week_key(cur, year: int, week_no: int, week_start_date: str,
                     week_end_date: str, shipment_date: Optional[str],
                     inventory_date: Optional[str], display_date: str) -> int:
    sd = shipment_date or ""
    id_ = inventory_date or ""
    row = db._exec(
        cur,
        """SELECT id FROM dv_week_keys
           WHERE year = ? AND week_no = ? AND shipment_date = ? AND inventory_date = ?""",
        (year, week_no, sd, id_),
    ).fetchone()
    if row:
        return row["id"]
    return db._last_insert_id(
        cur,
        """INSERT INTO dv_week_keys
           (year, week_no, week_start_date, week_end_date, shipment_date, inventory_date, display_date)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (year, week_no, week_start_date, week_end_date, sd, id_, display_date),
    )


def match_and_merge_weeks(parsed: Dict[str, List[Dict]], conn) -> Dict[str, Any]:
    shipments = parsed["shipment"]
    inventories = parsed["inventory"]

    for row in shipments:
        row["_bw"] = compute_business_week(date.fromisoformat(row["date"]))
        row["_date_obj"] = date.fromisoformat(row["date"])
    for row in inventories:
        row["_bw"] = compute_business_week(date.fromisoformat(row["date"]))

    week_key_map: Dict[str, int] = {}
    week_keys: List[Dict] = []
    pairs: List[Dict] = []
    bw_wk_map: Dict[tuple, int] = {}  # (year, week_no) -> week_key_id, 防同 ISO 周重复
    cur = conn.cursor()

    used_shipment: set = set()
    used_inventory: set = set()

    for si, s_row in enumerate(shipments):
        if si in used_shipment:
            continue
        best_match = _find_nearest_inventory(s_row, inventories, used_inventory, max_gap_days=4)
        if best_match is not None:
            ii, i_row = best_match
            used_shipment.add(si)
            used_inventory.add(ii)
            s_bw = s_row["_bw"]
            s_date = s_row["_date_obj"]
            i_date = date.fromisoformat(i_row["date"])
            key_str = f"{s_bw['year']}_{s_bw['week_no']}_{s_row['date']}_{i_row['date']}"
            wk_id = week_key_map.get(key_str)
            if wk_id is None:
                # 合并周使用发运的 ISO 周作为 week_no（发运是财务报告周基准）
                wk_id = _upsert_week_key(
                    cur, s_bw["year"], s_bw["week_no"],
                    s_bw["week_start_date"], s_bw["week_end_date"],
                    s_row["date"], i_row["date"], s_row["date"],
                )
                week_key_map[key_str] = wk_id
                week_keys.append({
                    "id": wk_id, "year": s_bw["year"], "week_no": s_bw["week_no"],
                    "display_date": s_row["date"], "shipment_date": s_row["date"],
                    "inventory_date": i_row["date"],
                })
                bw_wk_map[(s_bw["year"], s_bw["week_no"])] = wk_id
            pairs.append({"shipment_row": s_row, "inventory_row": i_row, "week_key_id": wk_id})

    for si, s_row in enumerate(shipments):
        if si in used_shipment:
            continue
        s_bw = s_row["_bw"]
        key_str = f"{s_bw['year']}_{s_bw['week_no']}_{s_row['date']}_none"
        wk_id = week_key_map.get(key_str)
        if wk_id is None:
            # 已有合并周覆盖了此 ISO 周，复用其 week_key_id 避免重复
            existing_wk = bw_wk_map.get((s_bw["year"], s_bw["week_no"]))
            if existing_wk is not None:
                continue
            wk_id = _upsert_week_key(
                cur, s_bw["year"], s_bw["week_no"],
                s_bw["week_start_date"], s_bw["week_end_date"],
                s_row["date"], None, s_row["date"],
            )
            week_key_map[key_str] = wk_id
            week_keys.append({
                "id": wk_id, "year": s_bw["year"], "week_no": s_bw["week_no"],
                "display_date": s_row["date"], "shipment_date": s_row["date"],
                "inventory_date": None,
            })
        pairs.append({"shipment_row": s_row, "inventory_row": None, "week_key_id": wk_id})

    for ii, i_row in enumerate(inventories):
        if ii in used_inventory:
            continue
        i_bw = i_row["_bw"]
        key_str = f"{i_bw['year']}_{i_bw['week_no']}_none_{i_row['date']}"
        wk_id = week_key_map.get(key_str)
        if wk_id is None:
            # 已有合并周覆盖了此 ISO 周，复用其 week_key_id 避免重复
            existing_wk = bw_wk_map.get((i_bw["year"], i_bw["week_no"]))
            if existing_wk is not None:
                continue
            wk_id = _upsert_week_key(
                cur, i_bw["year"], i_bw["week_no"],
                i_bw["week_start_date"], i_bw["week_end_date"],
                None, i_row["date"], i_row["date"],
            )
            week_key_map[key_str] = wk_id
            week_keys.append({
                "id": wk_id, "year": i_bw["year"], "week_no": i_bw["week_no"],
                "display_date": i_row["date"], "shipment_date": None,
                "inventory_date": i_row["date"],
            })
        pairs.append({"shipment_row": None, "inventory_row": i_row, "week_key_id": wk_id})

    conn.commit()
    return {"week_keys": week_keys, "pairs": pairs}


def _find_nearest_inventory(s_row, inventories, used_inventory, max_gap_days=4):
    """在未匹配库存中找到最接近发运日期的记录，不超过 max_gap_days 天。

    优先匹配发运日期之后 0~max_gap_days 天内的库存（常规 周日→周二=2天）；
    若无，取之前最近的。
    返回 (inventory_index, inventory_row) 或 None。
    """
    s_date = date.fromisoformat(s_row["date"])
    best_after = None   # (idx, row, gap)
    best_before = None  # (idx, row, gap)

    for ii, i_row in enumerate(inventories):
        if ii in used_inventory:
            continue
        i_date = date.fromisoformat(i_row["date"])
        gap = (i_date - s_date).days
        if abs(gap) <= max_gap_days:
            if gap >= 0 and (best_after is None or gap < best_after[2]):
                best_after = (ii, i_row, gap)
            elif gap < 0 and (best_before is None or abs(gap) < best_before[2]):
                best_before = (ii, i_row, abs(gap))

    if best_after:
        return (best_after[0], best_after[1])
    if best_before:
        return (best_before[0], best_before[1])
    return None


# ── 表需重算 ──────────────────────────────────────────────────────────
# 表需(t) = 发运(t-2) + 库存(t-1) - 库存(t)


def recalc_apparent_demand(conn) -> None:
    """重新计算所有表需数据点。

    表需(t) = 发运(t-2) + 库存(t-1) - 库存(t)

    使用 (year, week_no) 二维索引替代位置偏移，避免 weeks 数组中
    合并周/纯发运周/纯库存周交错排列导致的「上一周」指错行问题。
    """
    cur = conn.cursor()
    weeks = db._exec(
        cur,
        """SELECT wk.id, wk.year, wk.week_no, wk.display_date
           FROM dv_week_keys wk ORDER BY wk.display_date""",
    ).fetchall()
    if not weeks:
        return

    all_points = db._exec(
        cur,
        """SELECT id, week_key_id, metric_type, display_value, is_manual_override
           FROM dv_data_points
           WHERE product = ? AND metric_type IN ('shipment','inventory','apparent_demand')""",
        (PRODUCT,),
    ).fetchall()

    # week_key_id -> (year, week_no)
    week_id_info: Dict[int, tuple] = {wk["id"]: (wk["year"], wk["week_no"]) for wk in weeks}

    # 按 (year, week_no) 索引数据点，每种指标独立
    metric_by_week: Dict[str, Dict[tuple, dict]] = {
        "shipment": {}, "inventory": {}, "apparent_demand": {},
    }
    for pt in all_points:
        key = week_id_info.get(pt["week_key_id"])
        if key is not None:
            metric_by_week[pt["metric_type"]][key] = dict(pt)

    sorted_ship_keys: List[tuple] = sorted(metric_by_week["shipment"].keys())
    sorted_inv_keys: List[tuple] = sorted(metric_by_week["inventory"].keys())

    def _nth_prev_in_sorted(cur: tuple, sorted_keys: List[tuple], n: int = 1) -> tuple | None:
        """返回 sorted_keys 中小于 cur 的第 n 个 key（跳过缺口）。"""
        import bisect
        idx = bisect.bisect_left(sorted_keys, cur)
        target = idx - n
        return sorted_keys[target] if target >= 0 else None

    for wk in weeks:
        wid = wk["id"]
        cur_key = (wk["year"], wk["week_no"])
        ad_pt = metric_by_week["apparent_demand"].get(cur_key, {})
        if ad_pt.get("is_manual_override"):
            continue

        # 按排序位置查找前驱数据点（跳过 ISO 周号缺口）
        sp_key = _nth_prev_in_sorted(cur_key, sorted_ship_keys, 2)
        shipment_val = float(metric_by_week["shipment"][sp_key]["display_value"]) if sp_key else 0.0

        ip_key = _nth_prev_in_sorted(cur_key, sorted_inv_keys, 1)
        inv_prev = float(metric_by_week["inventory"][ip_key]["display_value"]) if ip_key else 0.0

        ip_t = metric_by_week["inventory"].get(cur_key)
        inv_t = float(ip_t["display_value"]) if (ip_t and ip_t.get("display_value") is not None) else 0.0

        demand = shipment_val + inv_prev - inv_t

        # 缺少任一输入数据点则标记为缺失
        is_missing = sp_key is None or ip_key is None or ip_t is None

        if not ad_pt:
            db._exec(
                cur,
                """INSERT INTO dv_data_points
                   (week_key_id, product, metric_type, imported_value, calculated_value,
                    manual_value, display_value, is_manual_override, is_missing_filled,
                    source, created_at)
                   VALUES (?, ?, 'apparent_demand', NULL, ?, NULL, ?, 0, ?, '自动计算', CURRENT_TIMESTAMP)""",
                (wid, PRODUCT, demand, demand, 1 if is_missing else 0),
            )
        else:
            db._exec(
                cur,
                """UPDATE dv_data_points
                   SET calculated_value = ?, display_value = ?, is_missing_filled = ?,
                       source = '自动计算', updated_at = CURRENT_TIMESTAMP
                   WHERE id = ?""",
                (demand, demand, 1 if is_missing else 0, ad_pt["id"]),
            )
def _now_expr() -> str:
    """返回 cross-db 的当前时间表达式。"""
    return "CURRENT_TIMESTAMP"



# ── GET /api/data-visualization/years ──────────────────────────────────


@router.get("/data-visualization/integration/local-preview")
async def integration_local_preview(user=Depends(dv_current_user)):
    file_paths = _local_mysteel_files()
    result = integrate_mysteel_files(file_paths)
    return {
        "files": [path.name for path in file_paths],
        "summary": result["summary"],
    }


@router.post("/data-visualization/integration/local-commit")
async def integration_local_commit(user=Depends(dv_current_user)):
    dv_require_edit("data_visualization_integration", user)
    file_paths = _local_mysteel_files()
    result = integrate_mysteel_files(file_paths)
    batch_id = _save_integrated_points(result["points"], [path.name for path in file_paths], user["name"])
    return {
        "ok": True,
        "batch_id": batch_id,
        "files": [path.name for path in file_paths],
        "summary": result["summary"],
    }


@router.post("/data-visualization/integration/preview")
async def integration_upload_preview(payload: IntegrationFilesRequest, user=Depends(dv_current_user)):
    tmp_paths = _write_uploads_to_tmp(payload.files)
    try:
        result = integrate_mysteel_files(tmp_paths)
        source_names = {path.name: item.file_name for path, item in zip(tmp_paths, payload.files)}
        for point in result["summary"]["samples"]:
            point["source_file"] = source_names.get(point["source_file"], point["source_file"])
    finally:
        _cleanup_tmp(tmp_paths)
    return {
        "files": [item.file_name for item in payload.files],
        "summary": result["summary"],
    }


@router.post("/data-visualization/integration/commit")
async def integration_upload_commit(payload: IntegrationFilesRequest, user=Depends(dv_current_user)):
    dv_require_edit("data_visualization_integration", user)
    tmp_paths = _write_uploads_to_tmp(payload.files)
    try:
        result = integrate_mysteel_files(tmp_paths)
        source_names = {path.name: item.file_name for path, item in zip(tmp_paths, payload.files)}
        for point in result["points"]:
            point["source_file"] = source_names.get(point["source_file"], point["source_file"])
        for point in result["summary"]["samples"]:
            point["source_file"] = source_names.get(point["source_file"], point["source_file"])
        batch_id = _save_integrated_points(result["points"], [item.file_name for item in payload.files], user["name"])
    finally:
        _cleanup_tmp(tmp_paths)
    batch_summary = _load_integration_batch_summary(batch_id)
    return {
        "ok": True,
        "batch_id": batch_id,
        "files": [item.file_name for item in payload.files],
        "summary": batch_summary["summary"],
        "merge_summary": batch_summary["merge_summary"],
    }


@router.get("/data-visualization/integration/latest")
async def integration_latest(user=Depends(dv_current_user)):
    with db.connect() as conn:
        cur = conn.cursor()
        batch = db._exec(
            cur,
            "SELECT * FROM dv_integration_batches ORDER BY created_at DESC, id DESC LIMIT 1",
        ).fetchone()
        rows = db._exec(
            cur,
            """SELECT metric_type, COUNT(*) AS c
               FROM dv_integrated_points GROUP BY metric_type ORDER BY metric_type""",
        ).fetchall()
        point_rows = db._exec(
            cur,
            "SELECT * FROM dv_integrated_points ORDER BY week_start, metric_type, source_country, category, product, id",
        ).fetchall()
    merge_summary = {}
    if batch and batch["validation_summary"]:
        try:
            merge_summary = json.loads(batch["validation_summary"])
        except json.JSONDecodeError:
            merge_summary = {}
    return {
        "batch": _row_to_dict(batch),
        "metrics": [_row_to_dict(row) for row in rows],
        "summary": _summarize_integrated_rows([_row_to_dict(row) for row in point_rows]),
        "merge_summary": merge_summary,
    }


@router.get("/data-visualization/integration/export")
async def integration_export(user=Depends(dv_current_user)):
    content = build_integrated_workbook_bytes()
    filename = f"iron_ore_integrated_{date.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/data-visualization/years")
async def get_years(user=Depends(dv_current_user)):
    with db.connect() as conn:
        cur = conn.cursor()
        integrated_rows = db._exec(
            cur,
            "SELECT DISTINCT SUBSTR(week_start, 1, 4) AS year FROM dv_integrated_points ORDER BY year"
        ).fetchall()
        if integrated_rows:
            return {"years": [int(r["year"]) for r in integrated_rows if r["year"]]}
        rows = db._exec(
            cur,
            "SELECT DISTINCT year FROM dv_week_keys ORDER BY year"
        ).fetchall()
    return {"years": [r["year"] for r in rows]}


# ── POST /api/data-visualization/import/preview ───────────────────────

@router.get("/data-visualization/filters")
async def get_filters(user=Depends(dv_current_user)):
    """返回整合结果中可用的筛选选项。"""
    with db.connect() as conn:
        cur = conn.cursor()
        integrated_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integrated_points").fetchone()["c"]
        if integrated_count:
            _sync_integrated_mainstream_statuses(cur)
            filter_rows = [
                _normalize_integrated_point(_row_to_dict(row))
                for row in db._exec(
                    cur,
                    """SELECT DISTINCT source_country, product, category, mainstream_status
                       FROM dv_integrated_points
                       ORDER BY product, category, source_country""",
                ).fetchall()
            ]
            product_pools = _integrated_product_pool_labels(filter_rows)
            products = product_pools["custom"]
            categories = [r["val"] for r in db._exec(cur,
                "SELECT DISTINCT category AS val FROM dv_integrated_points ORDER BY val").fetchall()]
            countries = [r["val"] for r in db._exec(cur,
                "SELECT DISTINCT source_country AS val FROM dv_integrated_points ORDER BY val").fetchall()]
            mainstreams = [r["val"] for r in db._exec(cur,
                "SELECT DISTINCT mainstream_status AS val FROM dv_integrated_points ORDER BY val").fetchall()]
            years = [r["year"] for r in db._exec(cur,
                "SELECT DISTINCT business_year AS year FROM dv_integrated_points ORDER BY year").fetchall() if r["year"]]
            return {
                "products": products,
                "categories": categories,
                "source_countries": countries,
                "mainstream_statuses": mainstreams,
                "years": years,
                "product_pools": {
                    "mainstream": product_pools["mainstream"],
                    "non_mainstream": product_pools["non_mainstream"],
                    "aggregate": ["主流矿合计", "非主流矿合计"],
                    "custom": product_pools["custom"],
                },
            }
        # fallback to old dv_data_points
        products = [r["product"] for r in db._exec(cur,
            "SELECT DISTINCT product FROM dv_data_points ORDER BY product").fetchall()]
        years = [r["year"] for r in db._exec(cur,
            "SELECT DISTINCT year FROM dv_week_keys ORDER BY year").fetchall()]
        return {
            "products": products or list(DV_PRODUCTS),
            "categories": [],
            "source_countries": [],
            "mainstream_statuses": [],
            "years": years,
            "product_pools": {
                "mainstream": products or list(DV_PRODUCTS),
                "non_mainstream": [],
                "aggregate": [],
                "custom": products or list(DV_PRODUCTS),
            },
        }


# ── POST /api/data-visualization/import/integrated/preview ─────────────
@router.post('/data-visualization/import/integrated/preview')

async def import_integrated_preview(
    payload: ImportRequest,
    user=Depends(dv_current_user),
):
    dv_require_edit('data_visualization_data', user)
    if not payload.file_data:
        raise HTTPException(status_code=400, detail="请先选择整合 Excel 文件")

    import base64
    file_bytes = base64.b64decode(payload.file_data)
    _INTEGRATED_PREVIEW_FILE_DIR.mkdir(parents=True, exist_ok=True)
    preview_file_id = uuid.uuid4().hex
    tmp_path = _integrated_preview_file_path(preview_file_id, payload.file_name)
    with tmp_path.open("wb") as tmp:
        tmp.write(file_bytes)

    result = _parse_integrated_excel(tmp_path)
    preview_id = _save_integrated_preview_cache(result, payload.file_name, str(tmp_path))

    return {
        'preview_id': preview_id,
        'file_name': payload.file_name,
        'summary': result['summary'],
        'errors': result['errors'],
        'sample_count': min(len(result['rows']), 20),
        'sample_rows': result['rows'][:20],
    }


@router.post('/data-visualization/import/integrated/preview-file')
async def import_integrated_preview_file(
    request: Request,
    background_tasks: BackgroundTasks,
    file_name: str,
    user=Depends(dv_current_user),
):
    dv_require_edit('data_visualization_data', user)
    path = await _save_integrated_preview_upload(request, file_name)
    return _create_integrated_preview_job(background_tasks, str(path), file_name)


@router.get('/data-visualization/import/integrated/preview-jobs/{job_id}')
async def import_integrated_preview_job_status(
    job_id: str,
    user=Depends(dv_current_user),
):
    dv_require_edit('data_visualization_data', user)
    job = _INTEGRATED_PREVIEW_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="预检任务不存在或已失效")
    return job


# ── POST /api/data-visualization/import/integrated/commit ──────────────

@router.post('/data-visualization/import/integrated/commit')
async def import_integrated_commit(
    payload: ImportRequest,
    background_tasks: BackgroundTasks,
    user=Depends(dv_current_user),
):
    dv_require_edit('data_visualization_data', user)

    if payload.preview_id:
        cached = _load_integrated_preview_cache(payload.preview_id)
        file_name = cached.get('file_name') or payload.file_name
        source_path = cached.get('source_path')
        if source_path:
            path = Path(source_path)
            if not path.exists():
                raise HTTPException(status_code=400, detail="预检文件已失效，请重新选择文件预检")
            return _create_integrated_import_job(
                background_tasks,
                str(path),
                file_name,
                user['name'],
            )
        else:
            result = {
                'rows': cached.get('rows', []),
                'errors': cached.get('errors', []),
                'summary': cached.get('summary', {}),
            }
    else:
        if not payload.file_data:
            raise HTTPException(status_code=400, detail="预检结果已失效，请重新选择文件预检")
        import base64
        file_bytes = base64.b64decode(payload.file_data)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        try:
            result = _parse_integrated_excel(tmp_path)
        finally:
            os.unlink(tmp_path)
        file_name = payload.file_name

    rows = result['rows']
    if result['errors']:
        raise HTTPException(status_code=400, detail=f'整合 Excel 存在 {len(result["errors"])} 条错误，无法导入')

    batch_id = _import_integrated_points(rows, file_name, user['name'])
    return {
        'batch_id': batch_id,
        'summary': result['summary'],
        'message': f'已导入 {len(rows)} 条数据',
    }


@router.get('/data-visualization/import/integrated/jobs/{job_id}')
async def import_integrated_job_status(
    job_id: str,
    user=Depends(dv_current_user),
):
    dv_require_edit('data_visualization_data', user)
    job = _INTEGRATED_IMPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="导入任务不存在或已失效")
    return job


# ── POST /api/data-visualization/import/preview ───────────────────────
@router.post("/data-visualization/import/preview")

async def import_preview(
    payload: ImportRequest,
    user=Depends(dv_current_user),
):
    dv_require_edit("data_visualization_data", user)

    import base64
    file_bytes = base64.b64decode(payload.file_data)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        parsed = parse_excel(tmp_path)
    finally:
        os.unlink(tmp_path)

    shipments = parsed["shipment"]
    inventories = parsed["inventory"]
    all_dates = sorted(set([r["date"] for r in shipments] + [r["date"] for r in inventories]))

    null_count = sum(1 for r in shipments if r["value"] is None) + sum(
        1 for r in inventories if r["value"] is None
    )

    with db.connect() as conn:
        merged = match_and_merge_weeks(parsed, conn)
        cur = conn.cursor()

        manual_protected = []
        history_changes = []
        for pair in merged["pairs"]:
            wk_id = pair["week_key_id"]
            for metric_type, row_key in [("shipment", "shipment_row"), ("inventory", "inventory_row")]:
                row = pair.get(row_key)
                if row is None or row["value"] is None:
                    continue
                existing = db._exec(
                    cur,
                    """SELECT id, display_value, is_manual_override
                       FROM dv_data_points
                       WHERE week_key_id = ? AND product = ? AND metric_type = ?""",
                    (wk_id, PRODUCT, metric_type),
                ).fetchone()
                if existing:
                    if existing["is_manual_override"]:
                        manual_protected.append({
                            "data_point_id": existing["id"],
                            "metric_type": metric_type,
                            "week_key_id": wk_id,
                            "date": row["date"],
                            "current_value": existing["display_value"],
                            "new_value": row["value"],
                        })
                    elif existing["display_value"] != row["value"]:
                        history_changes.append({
                            "metric_type": metric_type,
                            "week_key_id": wk_id,
                            "date": row["date"],
                            "current_value": existing["display_value"],
                            "new_value": row["value"],
                            "is_manual_protected": False,
                        })

    insert_count = sum(
        1 for p in merged["pairs"]
        if (p.get("shipment_row") and p["shipment_row"]["value"] is not None)
        or (p.get("inventory_row") and p["inventory_row"]["value"] is not None)
    )

    return {
        "file_name": payload.file_name,
        "metric_types": "inventory,shipment",
        "date_start": all_dates[0] if all_dates else "",
        "date_end": all_dates[-1] if all_dates else "",
        "total_rows": len(shipments) + len(inventories),
        "insert_count": insert_count,
        "overwrite_count": len(history_changes),
        "null_count": null_count,
        "error_count": 0,
        "manual_protected_count": len(manual_protected),
        "anomalies": [],
        "history_changes": history_changes,
        "manual_protected": manual_protected,
    }


# ── POST /api/data-visualization/import/commit ────────────────────────

@router.post("/data-visualization/import/commit")
async def import_commit(
    payload: ImportRequest,
    user=Depends(dv_current_user),
):
    dv_require_edit("data_visualization_data", user)

    import base64
    file_bytes = base64.b64decode(payload.file_data)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        parsed = parse_excel(tmp_path)
    finally:
        os.unlink(tmp_path)

    overwrite_ids = set(payload.overwrite_manual_ids)
    stats = {"insert_count": 0, "overwrite_count": 0, "error_count": 0, "manual_protected_count": 0}
    user_name = user["name"]

    with db.connect() as conn:
        cur = conn.cursor()
        merged = match_and_merge_weeks(parsed, conn)

        all_dates = sorted(
            set([r["date"] for r in parsed["shipment"]] + [r["date"] for r in parsed["inventory"]])
        )
        batch_id = db._last_insert_id(
            cur,
            """INSERT INTO dv_import_batches
               (file_name, metric_types, date_start, date_end, status, created_by)
               VALUES (?, 'inventory,shipment', ?, ?, 'processing', ?)""",
            (payload.file_name, all_dates[0] if all_dates else "",
             all_dates[-1] if all_dates else "", user_name),
        )

        for pair in merged["pairs"]:
            wk_id = pair["week_key_id"]
            for metric_type, row_key in [("shipment", "shipment_row"), ("inventory", "inventory_row")]:
                row = pair.get(row_key)
                if row is None or row["value"] is None:
                    continue
                value = row["value"]

                existing = db._exec(
                    cur,
                    """SELECT id, display_value, is_manual_override
                       FROM dv_data_points
                       WHERE week_key_id = ? AND product = ? AND metric_type = ?""",
                    (wk_id, PRODUCT, metric_type),
                ).fetchone()

                if existing:
                    dp_id = existing["id"]
                    if existing["is_manual_override"]:
                        if dp_id in overwrite_ids:
                            db._exec(
                                cur,
                                """UPDATE dv_data_points
                                   SET imported_value = ?, display_value = ?, is_manual_override = 0,
                                       manual_value = NULL, source = 'Excel导入覆盖人工修正',
                                       source_batch_id = ?, updated_by = ?
                                   WHERE id = ?""",
                                (value, value, batch_id, user_name, dp_id),
                            )
                            stats["overwrite_count"] += 1
                        else:
                            stats["manual_protected_count"] += 1
                            continue
                    else:
                        old_val = existing["display_value"]
                        db._exec(
                            cur,
                            """UPDATE dv_data_points
                               SET imported_value = ?, display_value = ?, source = '导入',
                                   source_batch_id = ?, updated_by = ?
                               WHERE id = ?""",
                            (value, value, batch_id, user_name, dp_id),
                        )
                        if old_val != value:
                            stats["overwrite_count"] += 1
                else:
                    db._last_insert_id(
                        cur,
                        """INSERT INTO dv_data_points
                           (week_key_id, product, metric_type, imported_value, display_value,
                            is_manual_override, is_missing_filled, source, source_batch_id, created_by)
                           VALUES (?, ?, ?, ?, ?, 0, 0, '导入', ?, ?)""",
                        (wk_id, PRODUCT, metric_type, value, value, batch_id, user_name),
                    )
                    stats["insert_count"] += 1

        db._exec(
            cur,
            """UPDATE dv_import_batches
               SET insert_count = ?, overwrite_count = ?, manual_protected_count = ?,
                   status = 'completed'
               WHERE id = ?""",
            (stats["insert_count"], stats["overwrite_count"], stats["manual_protected_count"], batch_id),
        )

        recalc_apparent_demand(conn)

    return {"ok": True, "batch_id": batch_id, "stats": stats}
# ── GET /api/data-visualization/table ─────────────────────────────────

@router.get("/data-visualization/table")
async def get_table(
    metric: str = Query(..., pattern="^(inventory|shipment|arrival|apparent_demand)$"),
    years: str = "",
    products: str = "",
    categories: str = "",
    source_countries: str = "",
    mainstream_status: str = "",
    product_pool: str = "",
    user=Depends(dv_current_user),
):
    year_list: List[int] = []
    years_empty_requested = False
    if years:
        for part in years.split(","):
            part = part.strip()
            if part:
                if part == "__EMPTY__":
                    years_empty_requested = True
                    continue
                try:
                    year_list.append(int(part))
                except ValueError:
                    pass

    product_list = _split_filter_values(products) if products else []
    category_list = _split_filter_values(categories) if categories else []
    country_list = _split_filter_values(source_countries) if source_countries else []
    mainstream_list = _split_filter_values(mainstream_status) if mainstream_status else []
    empty_filter_requested = years_empty_requested or "__EMPTY__" in product_list or "__EMPTY__" in category_list or "__EMPTY__" in country_list or "__EMPTY__" in mainstream_list
    product_list = [item for item in product_list if item != "__EMPTY__"]
    category_list = [item for item in category_list if item != "__EMPTY__"]
    country_list = [item for item in country_list if item != "__EMPTY__"]
    mainstream_list = [item for item in mainstream_list if item != "__EMPTY__"]

    with db.connect() as conn:
        cur = conn.cursor()
        integrated_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integrated_points").fetchone()["c"]
        if integrated_count:
            _sync_integrated_mainstream_statuses(cur)
            if empty_filter_requested:
                return {"metric": metric, "products": [], "data": []}
            aggregate_labels, aggregate_statuses = _aggregate_labels_and_statuses(product_list, []) if product_pool == "aggregate" else ([], [])
            if product_pool == "aggregate" and not aggregate_labels:
                return {"metric": metric, "products": [], "data": []}
            sql = """SELECT week_start, week_label, business_year, business_week,
                            display_date, source_country, product, category,
                            mainstream_status, value, is_calculable, validation_status, note
                     FROM dv_integrated_points WHERE metric_type = ?"""
            params: List[Any] = [metric]
            if year_list:
                placeholders = ",".join("?" for _ in year_list)
                sql += f" AND business_year IN ({placeholders})"
                params.extend(year_list)
            if category_list:
                placeholders = ",".join("?" for _ in category_list)
                sql += f" AND category IN ({placeholders})"
                params.extend(category_list)
            if country_list:
                placeholders = ",".join("?" for _ in country_list)
                sql += f" AND source_country IN ({placeholders})"
                params.extend(country_list)
            effective_mainstream_list = (
                aggregate_statuses
                if product_pool == "aggregate"
                else _effective_mainstream_statuses(product_pool, mainstream_list)
            )
            sql += " ORDER BY week_start, product, category"
            rows = [
                _normalize_integrated_point(_row_to_dict(row))
                for row in db._exec(cur, sql, tuple(params)).fetchall()
            ]
            if effective_mainstream_list:
                rows = [
                    row for row in rows
                    if (row["mainstream_status"] or "非主流") in effective_mainstream_list
                ]

            if product_pool == "aggregate":
                products_ordered = aggregate_labels
                week_map: Dict[str, Dict] = {}
                for row in rows:
                    ws = _week_map_key(row)
                    if ws not in week_map:
                        week_map[ws] = {
                            "date": row["display_date"] or row["week_start"],
                            "week": row["week_label"] or f"{row['business_year']} W{row['business_week']:02d}",
                        }
                        for p in products_ordered:
                            week_map[ws][p] = {
                                "id": None, "value": None,
                                "is_manual_override": False, "is_missing_filled": False,
                                "source": None, "updated_by": None, "updated_at": None,
                            }
                    status = row["mainstream_status"] or "非主流"
                    label = "主流矿合计" if status == "主流" else "非主流矿合计"
                    if label not in products_ordered:
                        continue
                    current = week_map[ws][label]["value"] or 0
                    week_map[ws][label] = {
                        "id": None,
                        "value": current + float(row["value"] or 0),
                        "is_manual_override": False,
                        "is_missing_filled": False,
                        "source": "整合数据汇总",
                        "updated_by": row["validation_status"],
                        "updated_at": None,
                    }
                return {"metric": metric, "products": products_ordered, "data": list(week_map.values())}

            label_map, products_ordered = _integrated_product_labels(rows)
            if product_list:
                rows, label_map, products_ordered = _filter_rows_by_product_labels(rows, product_list)
            if _uses_mainstream_product_order(product_pool, effective_mainstream_list):
                products_ordered = _ordered_mainstream_labels(products_ordered, label_map)

            week_map: Dict[str, Dict] = {}
            for row in rows:
                ws = _week_map_key(row)
                if ws not in week_map:
                    week_map[ws] = {
                        "date": row["display_date"] or row["week_start"],
                        "week": row["week_label"] or f"{row['business_year']} W{row['business_week']:02d}",
                    }
                    for p in products_ordered:
                        week_map[ws][p] = {
                            "id": None, "value": None,
                            "is_manual_override": False, "is_missing_filled": False,
                            "source": None, "updated_by": None, "updated_at": None,
                        }
                label = label_map[(row["source_country"], row["product"], row["category"])]
                week_map[ws][label] = {
                    "id": None,
                    "value": row["value"],
                    "is_manual_override": False,
                    "is_missing_filled": not bool(row["is_calculable"]) and metric == "apparent_demand",
                    "source": f"{row['source_country']} / {row['category']} / {row['mainstream_status']}",
                    "updated_by": row["validation_status"],
                    "updated_at": row["note"],
                }
            return {"metric": metric, "products": products_ordered, "data": list(week_map.values())}

        # fallback: old dv_data_points path
        base_sql = """SELECT wk.display_date, wk.year, wk.week_no, dp.id, dp.product,
                        dp.display_value, dp.is_manual_override, dp.is_missing_filled,
                        dp.source, dp.updated_by, dp.updated_at
                 FROM dv_data_points dp
                 JOIN dv_week_keys wk ON wk.id = dp.week_key_id
                 WHERE dp.metric_type = ?"""
        params: List[Any] = [metric]
        if year_list:
            placeholders = ",".join("?" for _ in year_list)
            base_sql += f" AND wk.year IN ({placeholders})"
            params.extend(year_list)
        if product_list:
            placeholders = ",".join("?" for _ in product_list)
            base_sql += f" AND dp.product IN ({placeholders})"
            params.extend(product_list)
        base_sql += " ORDER BY wk.display_date, dp.product"
        rows = db._exec(cur, base_sql, tuple(params)).fetchall()

    products = product_list if product_list else list(DV_PRODUCTS)
    week_map: Dict[str, Dict] = {}
    for r in rows:
        key = r["display_date"]
        if key not in week_map:
            week_map[key] = {
                "date": r["display_date"],
                "week": f"{r['year']} W{r['week_no']:02d}",
            }
            for p in products:
                week_map[key][p] = {"id": None, "value": None, "is_manual_override": False, "is_missing_filled": False, "source": None, "updated_by": None, "updated_at": None}
        week_map[key][r["product"]] = {
            "id": r["id"],
            "value": r["display_value"],
            "is_manual_override": bool(r["is_manual_override"]),
            "is_missing_filled": bool(r["is_missing_filled"]),
            "source": r["source"],
            "updated_by": r["updated_by"],
            "updated_at": r["updated_at"],
        }

    return {
        "metric": metric,
        "products": products,
        "data": list(week_map.values()),
    }


@router.put("/data-visualization/value")
async def update_value(
    payload: ManualEditRequest,
    user=Depends(dv_current_user),
):
    dv_require_edit("data_visualization_data", user)

    with db.connect() as conn:
        cur = conn.cursor()
        existing = db._exec(
            cur,
            """SELECT id, display_value, metric_type
               FROM dv_data_points WHERE id = ?""",
            (payload.data_point_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="数据点不存在")

        old_val = existing["display_value"]
        metric_type = existing["metric_type"]
        user_name = user["name"]

        db._exec(
            cur,
            """UPDATE dv_data_points
               SET manual_value = ?, display_value = ?, is_manual_override = 1,
                   source = '手工修改', updated_by = ?
               WHERE id = ?""",
            (payload.new_value, payload.new_value, user_name, payload.data_point_id),
        )

        if metric_type in ("shipment", "inventory"):
            recalc_apparent_demand(conn)

    return {"ok": True, "data_point_id": payload.data_point_id, "new_value": payload.new_value}


# ── GET /api/data-visualization/chart ─────────────────────────────────

@router.get("/data-visualization/chart")
async def get_chart(
    metric: str = Query(..., pattern="^(inventory|shipment|arrival|apparent_demand)$"),
    years: str = "",
    products: str = "",
    categories: str = "",
    source_countries: str = "",
    mainstream_status: str = "",
    product_pool: str = "",
    user=Depends(dv_current_user),
):
    year_list: List[int] = []
    years_empty_requested = False
    if years:
        for part in years.split(","):
            part = part.strip()
            if part:
                if part == "__EMPTY__":
                    years_empty_requested = True
                    continue
                try:
                    year_list.append(int(part))
                except ValueError:
                    pass

    product_list = _split_filter_values(products) if products else []
    category_list = _split_filter_values(categories) if categories else []
    country_list = _split_filter_values(source_countries) if source_countries else []
    mainstream_list = _split_filter_values(mainstream_status) if mainstream_status else []
    empty_filter_requested = years_empty_requested or "__EMPTY__" in product_list or "__EMPTY__" in category_list or "__EMPTY__" in country_list or "__EMPTY__" in mainstream_list
    product_list = [item for item in product_list if item != "__EMPTY__"]
    category_list = [item for item in category_list if item != "__EMPTY__"]
    country_list = [item for item in country_list if item != "__EMPTY__"]
    mainstream_list = [item for item in mainstream_list if item != "__EMPTY__"]

    with db.connect() as conn:
        cur = conn.cursor()
        integrated_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_integrated_points").fetchone()["c"]
        if integrated_count:
            _sync_integrated_mainstream_statuses(cur)
            if empty_filter_requested:
                return {"metric": metric, "series": {}}
            aggregate_labels, aggregate_statuses = _aggregate_labels_and_statuses(product_list, []) if product_pool == "aggregate" else ([], [])
            if product_pool == "aggregate" and not aggregate_labels:
                return {"metric": metric, "series": {}}
            params_i: List[Any] = [metric]
            sql_i = """SELECT week_start, display_date, business_year, business_week,
                              source_country, product, category, mainstream_status, value,
                              is_calculable, validation_status
                       FROM dv_integrated_points WHERE metric_type = ?"""
            if year_list:
                placeholders_y = ",".join("?" for _ in year_list)
                sql_i += f" AND business_year IN ({placeholders_y})"
                params_i.extend(year_list)
            if category_list:
                placeholders_c = ",".join("?" for _ in category_list)
                sql_i += f" AND category IN ({placeholders_c})"
                params_i.extend(category_list)
            if country_list:
                placeholders_n = ",".join("?" for _ in country_list)
                sql_i += f" AND source_country IN ({placeholders_n})"
                params_i.extend(country_list)
            effective_mainstream_list = (
                aggregate_statuses
                if product_pool == "aggregate"
                else _effective_mainstream_statuses(product_pool, mainstream_list)
            )
            sql_i += " ORDER BY product, category, week_start"
            rows_i = [
                _normalize_integrated_point(_row_to_dict(row))
                for row in db._exec(cur, sql_i, tuple(params_i)).fetchall()
            ]
            if effective_mainstream_list:
                rows_i = [
                    row for row in rows_i
                    if (row["mainstream_status"] or "非主流") in effective_mainstream_list
                ]
            if product_pool == "aggregate":
                aggregate: Dict[tuple, Dict[str, Any]] = {}
                for row in rows_i:
                    status = row["mainstream_status"] or "非主流"
                    label = "主流矿合计" if status == "主流" else "非主流矿合计"
                    if label not in aggregate_labels:
                        continue
                    year = str(row["business_year"]) if row["business_year"] else row["week_start"][:4]
                    key = (label, year, row["business_week"])
                    if key not in aggregate:
                        aggregate[key] = {
                            "week_no": row["business_week"] if row["business_week"] else 0,
                            "display_date": row["display_date"] or row["week_start"],
                            "value": 0.0,
                            "is_manual_override": False,
                            "is_missing_filled": False,
                            "_has_value": False,
                        }
                    is_missing = row["value"] is None or (not bool(row["is_calculable"]) and metric == "apparent_demand")
                    if is_missing:
                        aggregate[key]["is_missing_filled"] = True
                    else:
                        aggregate[key]["value"] += float(row["value"])
                        aggregate[key]["_has_value"] = True
                result_agg: Dict[str, Dict[str, List[Dict]]] = {}
                for (label, year, _week_no), item in sorted(aggregate.items()):
                    has_value = item.pop("_has_value", False)
                    if item["is_missing_filled"] or not has_value:
                        item["value"] = None
                    result_agg.setdefault(label, {}).setdefault(year, []).append(item)
                return {"metric": metric, "series": result_agg}

            label_map_i, _products_ordered_i = _integrated_product_labels(rows_i)
            if product_list:
                rows_i, label_map_i, _products_ordered_i = _filter_rows_by_product_labels(rows_i, product_list)
            if _uses_mainstream_product_order(product_pool, effective_mainstream_list):
                _products_ordered_i = _ordered_mainstream_labels(_products_ordered_i, label_map_i)
            result_i: Dict[str, Dict[str, List[Dict]]] = {}
            for row in rows_i:
                label = label_map_i[(row["source_country"], row["product"], row["category"])]
                year = str(row["business_year"]) if row["business_year"] else row["week_start"][:4]
                if label not in result_i:
                    result_i[label] = {}
                if year not in result_i[label]:
                    result_i[label][year] = []
                result_i[label][year].append({
                    "week_no": row["business_week"] if row["business_week"] else 0,
                    "display_date": row["display_date"] or row["week_start"],
                    "value": row["value"],
                    "is_manual_override": False,
                    "is_missing_filled": not bool(row["is_calculable"]) and metric == "apparent_demand",
                })
            ordered_result_i = {
                label: result_i[label]
                for label in _products_ordered_i
                if label in result_i
            }
            return {"metric": metric, "series": ordered_result_i}

        params: List[Any] = [metric]
        sql = """SELECT wk.year, wk.week_no, wk.display_date, dp.product,
                        dp.display_value, dp.is_manual_override, dp.is_missing_filled
                 FROM dv_data_points dp
                 JOIN dv_week_keys wk ON wk.id = dp.week_key_id
                 WHERE dp.metric_type = ?"""

        if product_list:
            placeholders_p = ",".join("?" for _ in product_list)
            sql += f" AND dp.product IN ({placeholders_p})"
            params.extend(product_list)

        if year_list:
            placeholders_y = ",".join("?" for _ in year_list)
            sql += f" AND wk.year IN ({placeholders_y})"
            params.extend(year_list)

        sql += " ORDER BY dp.product, wk.year, wk.week_no"
        rows = db._exec(cur, sql, tuple(params)).fetchall()

    by_product_year: Dict[str, Dict[int, List[Dict]]] = {}
    for r in rows:
        prod = r["product"]
        yr = r["year"]
        if prod not in by_product_year:
            by_product_year[prod] = {}
        if yr not in by_product_year[prod]:
            by_product_year[prod][yr] = []
        by_product_year[prod][yr].append({
            "week_no": r["week_no"],
            "display_date": r["display_date"],
            "value": r["display_value"],
            "is_manual_override": bool(r["is_manual_override"]),
            "is_missing_filled": bool(r["is_missing_filled"]),
        })

    # Sort by year within each product
    result: Dict[str, Dict[str, List[Dict]]] = {}
    for prod, by_year in by_product_year.items():
        result[prod] = {str(k): v for k, v in sorted(by_year.items())}

    return {"metric": metric, "series": result}



# ── GET /api/data-visualization/import-batches ────────────────────────

@router.get("/data-visualization/import-batches")
async def get_import_batches(user=Depends(dv_current_user)):
    with db.connect() as conn:
        cur = conn.cursor()
        rows = db._exec(
            cur,
            """SELECT id, file_name, metric_types, date_start, date_end,
                      insert_count, overwrite_count, error_count, manual_protected_count,
                      status, created_by, created_at
        FROM dv_import_batches ORDER BY created_at DESC LIMIT 50""",
        ).fetchall()
    return {"batches": [_row_to_dict(r) for r in rows]}


# ── Seed test data ─────────────────────────────────────────────────

def seed_dv_data():
    """从真实 Excel 数据播种 DV 表（发运 + 卡粉库存 + 表需）。
    
    数据来源: 建龙/期货组/本地数据库/副本铁矿data base.xlsx
    使用 {parse_excel + match_and_merge_weeks + recalc_apparent_demand} 流程，
    替代旧的随机合成数据。
    """
    from .dv_seed_data import DV_SEED_SHIPMENT, DV_SEED_INVENTORY

    with db.connect() as conn:
        cur = conn.cursor()
        existing_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_week_keys").fetchone()["c"]
        if existing_count > 0:
            old_seed = db._exec(
                cur, "SELECT COUNT(*) AS c FROM dv_data_points WHERE source = '种子数据'"
            ).fetchone()["c"]
            if old_seed > 0:
                print("[seed_dv_data] 检测到旧种子数据，清除后重新生成（真实数据）...")
                db._exec(cur, "DELETE FROM dv_change_log")
                db._exec(cur, "DELETE FROM dv_data_points")
                db._exec(cur, "DELETE FROM dv_week_keys")
                db._exec(cur, "DELETE FROM dv_import_batches")
            else:
                return

        # 构建 parse_excel 兼容的 dict
        parsed = {
            "shipment": [{"date": s[0], "value": s[1]} for s in DV_SEED_SHIPMENT],
            "inventory": [{"date": i[0], "value": i[1]} for i in DV_SEED_INVENTORY],
        }

        result = match_and_merge_weeks(parsed, conn)
        week_keys = result["week_keys"]
        pairs = result["pairs"]

        total_weeks = len(week_keys)
        print(f"[seed_dv_data] 真实数据：{total_weeks} 个业务周")

        # 插入发运和库存数据点
        for pair in pairs:
            wk_id = pair["week_key_id"]
            if pair["shipment_row"]:
                s_val = pair["shipment_row"]["value"]
                db._exec(
                    cur,
                    """INSERT INTO dv_data_points
                       (week_key_id, product, metric_type, imported_value, calculated_value,
                        display_value, source, created_by)
                       VALUES (?, ?, 'shipment', ?, ?, ?, '种子数据', 'system')""",
                    (wk_id, PRODUCT, s_val, s_val, s_val),
                )
            if pair["inventory_row"]:
                i_val = pair["inventory_row"]["value"]
                db._exec(
                    cur,
                    """INSERT INTO dv_data_points
                       (week_key_id, product, metric_type, imported_value, calculated_value,
                        display_value, source, created_by)
                       VALUES (?, ?, 'inventory', ?, ?, ?, '种子数据', 'system')""",
                    (wk_id, PRODUCT, i_val, i_val, i_val),
                )

        # 计算表需
        recalc_apparent_demand(conn)

        point_count = db._exec(cur, "SELECT COUNT(*) AS c FROM dv_data_points WHERE source = '种子数据'").fetchone()["c"]
        print(f"[seed_dv_data] 已完成：{total_weeks} 周，共 {point_count} 条数据点")
