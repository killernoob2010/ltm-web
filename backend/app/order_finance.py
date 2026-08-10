"""订单融资进度监控。

Excel 台账解析、导入、查询和管理端计划字段维护。
"""
from __future__ import annotations

import json
import hashlib
import logging
import re
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import xlrd
from fastapi import APIRouter, Depends, Header, HTTPException, Request
from pydantic import BaseModel

from . import db
from .permissions import require_permission


router = APIRouter()
logger = logging.getLogger(__name__)

SUBSIDIARIES = ["东钢", "北满", "承德", "抚顺", "西林", "阿城"]
LOCAL_DEFAULT_LEDGER_DIR = Path("/Users/wangjingze/建龙/贸易处/订单融资合同汇总")
LOCAL_DEFAULT_LEDGER_WORKBOOK = Path("/Users/wangjingze/建龙/贸易处/YOLANDA和香港建龙出口钢材信用证台账.xlsx")
ORDER_FINANCE_MODULE = "order_finance_progress"
ORDER_FINANCE_CAPITAL_MODULE = "order_finance_capital"
TARGET_XLSX_SHEETS = ("订单", "额度", "预警")
ORDER_VESSEL_SHEET = "26.8.5钢材出口情况表"
ORDER_VESSEL_CURRENT_R1_SHEET = "26.8.10钢材出口情况表"
ORDER_VESSEL_EXPECTED_SHA256 = "53b9a51aa2febe5118980cc32ba50a4821b2e02959394e47a74c17ce70a3e247"

ORDER_VESSEL_HEADERS = {
    "steel_mill": "出口方（钢厂）",
    "export_user": "使用方（终端用户或合同签署方）",
    "cargo": "货物",
    "vessel": "船名（IMO）",
    "quantity_mt": "合同量（吨）",
    "loading_port": "装港港口",
    "loading_port_arrival_date": "船到装港日期",
    "planned_berth_date": "计划靠泊日期",
    "discharge_port": "卸港港口",
    "estimated_discharge_date": "预计到卸港日期",
    "document_status": "单据情况",
    "repayment_due_date": "还款到期日",
    "loan_amount": "借款金额（元）",
    "remark": "备注",
    "business_no": "业务编号",
    "route_distance_nm": "航线距离（海里）",
    "eta_start_date": "卸港ETA起算日",
    "estimated_speed_knots": "估算船速（节）",
    "eta_basis": "ETA计算依据",
    "route_source": "航线来源",
}

ORDER_VESSEL_HEADER_ALIASES = {
    **{field: (label,) for field, label in ORDER_VESSEL_HEADERS.items()},
    "export_user": (
        "使用方（终端用户或合同签署方）",
        "最终业务去向/终端客户",
        "最终业务去向",
        "终端客户",
        "最终贸易合作方",
    ),
    "repayment_due_date": ("还款到期日", "汇报还款到期日"),
}

ORDER_VESSEL_SNAPSHOT_FIELDS = [
    "source_version", "source_date", "source_file_name", "source_sheet_name",
    "source_sha256", "source_row", "business_no", "steel_mill", "export_user",
    "cargo", "vessel", "quantity_mt", "loading_port",
    "loading_port_arrival_date", "planned_berth_date", "discharge_port",
    "estimated_discharge_date", "document_status", "repayment_due_date",
    "loan_amount", "loan_amount_note", "remark", "route_distance_nm",
    "eta_start_date", "estimated_speed_knots", "eta_basis", "route_source",
    "final_destination_status", "final_destination_source",
    "reporting_due_date_source", "email_due_values_json", "email_due_source",
    "email_due_source_date", "preview_status",
]

DEFAULT_BANK_LIMITS = [
    {
        "bank": "中信唐山",
        "limit": 200000000,
        "note": "流贷-限非中信银行融资主体",
        "lc_requirement": "可接受FCR",
        "bill_requirement": "无限制",
        "finance_ratio": "",
        "term": "",
    },
    {
        "bank": "OCBC",
        "limit": 6000 * 7.2 * 10000,
        "note": "订单融资-东钢 / 打包贷款-集团内",
        "lc_requirement": "可接受FCR",
        "bill_requirement": "to order或客户的银行",
        "finance_ratio": "",
        "term": "",
    },
    {
        "bank": "UOB",
        "limit": 2000 * 7.2 * 10000,
        "note": "订单融资-集团内钢厂",
        "lc_requirement": "不能接受FCR",
        "bill_requirement": "to order或客户的银行",
        "finance_ratio": "",
        "term": "",
    },
    {
        "bank": "918ING银行（新加坡）",
        "limit": 1000 * 7.2 * 10000,
        "note": "订单融资-天津、集团内钢厂",
        "lc_requirement": "可接受FCR",
        "bill_requirement": "to order、客户银行、客户均可",
        "finance_ratio": "",
        "term": "",
    },
    {
        "bank": "918ING银行（香港）",
        "limit": 3000 * 7.2 * 10000,
        "note": "订单融资-天津、集团内钢厂",
        "lc_requirement": "可接受FCR",
        "bill_requirement": "to order、客户银行、客户均可",
        "finance_ratio": "",
        "term": "",
    },
]

FACT_FIELDS = [
    "business_key", "subsidiary", "source_file", "source_sheet", "source_row_start",
    "source_row_end", "source_snapshot_date", "product_name", "purchase_contract_no",
    "system_contract_no", "buyer", "seller", "overseas_entity", "terminal_customer",
    "contract_date", "trade_term", "origin_port", "destination_port",
    "contract_quantity_mt", "contract_currency", "contract_amount", "finance_bank",
    "finance_amount_expected", "finance_amount_actual", "repaid_amount",
    "remaining_credit_amount", "finance_drawdown_date", "finance_due_date",
    "finance_days", "finance_status", "latest_shipment_date", "lc_latest_shipment_date",
    "vessel_voyage", "bill_of_lading_date", "bill_of_lading_no",
    "document_submission_date", "collection_date", "actual_shipped_quantity_mt",
    "actual_goods_amount", "tail_amount", "tail_payment_date", "executor",
    "business_status", "risk_level", "remark", "sales_contracts_json",
    "settlement_json", "corrections_json", "import_warnings_json", "source_json",
]

MANAGEMENT_FIELDS = {
    "planned_drawdown_date",
    "planned_finance_amount",
    "amount_adjustment_note",
    "repayment_requirement",
    "repayment_requirement_status",
    "next_action",
    "next_follow_up_date",
    "manager_note",
    "port_confirmed_date",
    "port_confirmed_by",
    "port_confirmed_at",
    "shipment_confirmed_date",
    "shipment_confirmed_by",
    "shipment_confirmed_at",
}


class ImportLocalRequest(BaseModel):
    directory: str = str(LOCAL_DEFAULT_LEDGER_WORKBOOK)


class ManagementUpdateRequest(BaseModel):
    planned_drawdown_date: Optional[str] = None
    planned_finance_amount: Optional[float] = None
    amount_adjustment_note: Optional[str] = None
    repayment_requirement: Optional[str] = None
    repayment_requirement_status: Optional[str] = None
    next_action: Optional[str] = None
    next_follow_up_date: Optional[str] = None
    manager_note: Optional[str] = None


class ShipmentConfirmationRequest(BaseModel):
    confirmed: bool = True
    shipment_confirmed_date: Optional[str] = None


class PortConfirmationRequest(BaseModel):
    confirmed: bool = True
    port_confirmed_date: Optional[str] = None


class ContractReminderRequest(BaseModel):
    manager_note: Optional[str] = None
    next_follow_up_date: Optional[str] = None


class ManualOrderFinanceRequest(BaseModel):
    subsidiary: str
    product_name: Optional[str] = None
    purchase_contract_no: Optional[str] = None
    system_contract_no: Optional[str] = None
    terminal_customer: Optional[str] = None
    contract_quantity_mt: Optional[float] = None
    contract_currency: Optional[str] = "CNY"
    contract_amount: Optional[float] = None
    finance_bank: Optional[str] = None
    finance_amount_expected: Optional[float] = None
    finance_amount_actual: Optional[float] = None
    finance_drawdown_date: Optional[str] = None
    finance_due_date: Optional[str] = None
    latest_shipment_date: Optional[str] = None
    bill_of_lading_date: Optional[str] = None
    collection_date: Optional[str] = None
    executor: Optional[str] = None
    planned_drawdown_date: Optional[str] = None
    planned_finance_amount: Optional[float] = None
    amount_adjustment_note: Optional[str] = None
    repayment_requirement: Optional[str] = None
    repayment_requirement_status: Optional[str] = None
    next_action: Optional[str] = None
    next_follow_up_date: Optional[str] = None
    manager_note: Optional[str] = None


class DuplicateOrderFinanceError(ValueError):
    def __init__(self, existing: Dict[str, Any]):
        super().__init__("已存在相同子公司和合同号的订单融资记录")
        self.existing = existing


async def order_finance_current_user(authorization: Optional[str] = Header(default=None)):
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.removeprefix("Bearer ").strip()
    user = db.get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="请先登录")
    return user


def order_finance_require_edit(user: dict):
    require_permission(user, "order_finance.records", "edit")


def order_finance_require_import(user: dict):
    require_permission(user, "order_finance.records", "import")


def order_finance_require_view(user: dict):
    require_permission(user, "order_finance.records", "view")


def _cell_value(book, cell) -> Any:
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
            return dt.date().isoformat() if dt.time() == datetime.min.time() else dt.isoformat(sep=" ")
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return int(cell.value) if float(cell.value).is_integer() else round(float(cell.value), 6)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    value = str(cell.value).strip()
    return value.replace("\n", " / ") if value else None


def _row_values(book, sheet, row_idx: int) -> List[Any]:
    return [_cell_value(book, sheet.cell(row_idx, col)) for col in range(sheet.ncols)]


def _normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _to_float(value: Any) -> Optional[float]:
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("CNY", "").replace("USD", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value: Any) -> Optional[int]:
    number = _to_float(value)
    return int(number) if number is not None else None


def _normalize_date(value: Any) -> str:
    if value in (None, "", "-"):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for sep in ("/", "."):
        text = text.replace(sep, "-")
    if " " in text and len(text.split()) > 1:
        return text
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError:
        return text


def _normalize_xlsx_date(value: Any) -> str:
    if value in (None, "", "-", 0, "0"):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    return _normalize_date(value)


def _parse_date(value: Any) -> Optional[date]:
    text = _normalize_date(value)
    if not text or len(text) < 10:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _effective_finance_due(new_due: Any, original_due: Any, extension_days: int = 0) -> str:
    normalized_new = _normalize_xlsx_date(new_due)
    if normalized_new:
        return normalized_new
    normalized_original = _normalize_xlsx_date(original_due)
    parsed_original = _parse_date(normalized_original)
    if parsed_original and extension_days > 0:
        return (parsed_original + timedelta(days=extension_days)).isoformat()
    return normalized_original


def _find_col(headers: List[Any], *needles: str) -> Optional[int]:
    for needle in needles:
        for idx, header in enumerate(headers):
            if header and needle in str(header):
                return idx
    return None


def _get(row: List[Any], idx: Optional[int]) -> Any:
    return row[idx] if idx is not None and idx < len(row) else None


def _is_sequence(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return False


def _subsidiary_from_filename(filename: str) -> str:
    for name in SUBSIDIARIES:
        if name in filename:
            return name
    return Path(filename).stem


def _currency_amount(row: List[Any], amount_col: Optional[int]) -> tuple[str, Optional[float]]:
    value = _get(row, amount_col)
    next_value = _get(row, amount_col + 1 if amount_col is not None else None)
    text = _normalize_text(value).upper()
    if text in {"CNY", "USD", "AED", "MYR"}:
        return text, _to_float(next_value)
    return "", _to_float(value)


def _derive_bank(product: Any, finance_bank: Any, remark: Any) -> str:
    explicit = _normalize_text(finance_bank)
    if explicit and len(explicit) <= 20:
        return explicit
    text = f"{_normalize_text(product)} {_normalize_text(remark)}".upper()
    for bank in ["中信", "UOB", "OCBC", "ING", "邮储"]:
        if bank.upper() in text:
            return bank
    return explicit


def _terminal_customer(children: List[Dict[str, Any]], buyer: Any) -> str:
    # The contract chain is ordered from the first resale to the last.  The
    # business destination is therefore the last actual user, not the first
    # buyer appearing after our own contract.
    for child in reversed(children):
        candidate = _normalize_text(child.get("buyer"))
        if candidate and not _is_final_destination_middle_party(candidate):
            return candidate
    main_buyer = _normalize_text(buyer)
    if main_buyer and not _is_final_destination_middle_party(main_buyer):
        return main_buyer
    return ""


def _overseas_entity(row: List[Any], children: List[Dict[str, Any]], buyer_col: Optional[int], seller_col: Optional[int]) -> str:
    candidates = [_get(row, buyer_col), _get(row, seller_col)]
    for child in children:
        candidates.extend([child.get("buyer"), child.get("seller")])
    for candidate in candidates:
        text = _normalize_text(candidate)
        upper = text.upper()
        if "YOLANDA" in upper or "HONG KONG" in upper or "SINGAPORE" in upper or "建龍" in text or "建龙" in text:
            return text
    return ""


def _business_key(subsidiary: str, purchase_contract: Any, system_contract: Any, source_file: str, row_no: int) -> str:
    purchase = _normalize_text(purchase_contract)
    system = _normalize_text(system_contract)
    if purchase or system:
        return "|".join([subsidiary, purchase, system])
    return "|".join([subsidiary, source_file, str(row_no)])


def _manual_business_key(record: Dict[str, Any]) -> str:
    subsidiary = _normalize_text(record.get("subsidiary"))
    purchase = _normalize_text(record.get("purchase_contract_no"))
    system = _normalize_text(record.get("system_contract_no"))
    if purchase or system:
        return _business_key(subsidiary, purchase, system, "手动新增", 0)
    return "|".join([subsidiary, "手动新增", uuid4().hex])


def _build_warnings(record: Dict[str, Any]) -> List[Dict[str, str]]:
    warnings: List[Dict[str, str]] = []
    for field in ("contract_date", "finance_drawdown_date", "finance_due_date", "latest_shipment_date", "bill_of_lading_date"):
        text = _normalize_text(record.get(field))
        parsed = _parse_date(text)
        if text and not parsed:
            warnings.append({"field": field, "level": "高", "message": f"日期无法识别：{text}"})
            continue
        if parsed and (parsed.year < 2024 or parsed.year > 2028):
            warnings.append({"field": field, "level": "高", "message": f"日期年份异常：{text}"})
    drawdown = _parse_date(record.get("finance_drawdown_date"))
    due = _parse_date(record.get("finance_due_date"))
    if drawdown and due and due < drawdown:
        warnings.append({"field": "finance_due_date", "level": "高", "message": "融资到期日早于放款日期"})
    return warnings


def _is_data_quality_warning(warning: Dict[str, Any]) -> bool:
    return _normalize_text(warning.get("field")) != "excel_alert"


def _data_quality_warning_count(records: List[Dict[str, Any]]) -> int:
    return sum(
        1
        for record in records
        for warning in _json_loads(record.get("import_warnings_json"), [])
        if _is_data_quality_warning(warning)
    )


def derive_business_status(record: Dict[str, Any]) -> Dict[str, str]:
    remark = _normalize_text(record.get("remark"))
    drawdown = _normalize_date(record.get("finance_drawdown_date"))
    due = _parse_date(record.get("finance_due_date"))
    bill = _normalize_date(record.get("bill_of_lading_date"))
    collection = _normalize_date(record.get("collection_date"))
    today = date.today()

    if "已结算" in remark:
        return {"business_status": "已结算", "risk_level": "低", "next_action": "无"}
    if not drawdown:
        return {"business_status": "待放款", "risk_level": "中", "next_action": "确认放款计划"}
    if not bill:
        if due and due <= today:
            return {"business_status": "需展期确认", "risk_level": "高", "next_action": "确认是否延期融资"}
        if due and (due - today).days <= 7:
            return {"business_status": "已放款待集港", "risk_level": "高", "next_action": "跟进集港进度"}
        if due and (due - today).days <= 30:
            return {"business_status": "已放款待集港", "risk_level": "中", "next_action": "跟进集港进度"}
        return {"business_status": "已放款待集港", "risk_level": "中", "next_action": "跟进集港进度"}
    if not collection:
        return {"business_status": "已装船待回款", "risk_level": "中", "next_action": "跟进交单和回款"}
    return {"business_status": "已回款待结算", "risk_level": "低", "next_action": "确认结算"}


def _columns(headers: List[Any]) -> Dict[str, Optional[int]]:
    return {
        "product": _find_col(headers, "货物品名", "货物"),
        "purchase_contract": _find_col(headers, "东钢合同号", "北满采/销合同号", "承德采/销合同号", "抚顺采/销合同号", "西林采/销合同号", "阿城采/销合同号"),
        "system_contract": _find_col(headers, "YOLANDA合同号", "Yolanda/Jianlong采/销合同号", "Yolanda采/销合同号", "YOLANDA采/销合同号"),
        "lc_no": _find_col(headers, "LC NO"),
        "buyer": _find_col(headers, "买方"),
        "seller": _find_col(headers, "卖方"),
        "contract_date": _find_col(headers, "合同日期"),
        "trade_term": _find_col(headers, "价格条款"),
        "latest_ship": _find_col(headers, "最迟装船期"),
        "origin": _find_col(headers, "起运港"),
        "destination": _find_col(headers, "目的港", "卸港"),
        "quantity": _find_col(headers, "合同数量"),
        "amount": _find_col(headers, "合同总金额"),
        "loan_expected": _find_col(headers, "应放款金额", "融资金额", "放款金额"),
        "loan_actual": _find_col(headers, "实际放款金额"),
        "repaid": _find_col(headers, "已还款金额"),
        "remaining": _find_col(headers, "剩余额度"),
        "loan_date": _find_col(headers, "放款日期"),
        "loan_due": _find_col(headers, "放款到期日期", "放款到期日", "融资到期日"),
        "finance_days": _find_col(headers, "融资天数"),
        "payment_method": _find_col(headers, "付款方式"),
        "ship_qty": _find_col(headers, "实际出货数量", "装船数量"),
        "actual_amount": _find_col(headers, "实际出货金额", "实际货物金额"),
        "tail_amount": _find_col(headers, "尾款", "采购应退款金额"),
        "tail_date": _find_col(headers, "尾款付款日期"),
        "lc_date": _find_col(headers, "开证日"),
        "lc_bank": _find_col(headers, "开证行"),
        "lc_latest_ship": _find_col(headers, "LC-LSD船期", "LC 最迟装期"),
        "forwarder": _find_col(headers, "货代"),
        "bl_date": _find_col(headers, "提单日期"),
        "vessel": _find_col(headers, "船名航次", "船名", "船期"),
        "bl_no": _find_col(headers, "提单号"),
        "doc_date": _find_col(headers, "交单日期"),
        "collection_date": _find_col(headers, "回款日期"),
        "owner": _find_col(headers, "执行人员", "执行"),
        "status": _find_col(headers, "状态"),
        "remark": _find_col(headers, "备注"),
    }


def parse_order_finance_workbook(path: Path) -> Dict[str, Any]:
    book = xlrd.open_workbook(str(path), formatting_info=True)
    sheet = book.sheet_by_index(0)
    headers = _row_values(book, sheet, 1)
    rows = [_row_values(book, sheet, row_idx) for row_idx in range(sheet.nrows)]
    cols = _columns(headers)
    subsidiary = _subsidiary_from_filename(path.name)
    snapshot_date = date.today().isoformat()

    primary_indices = []
    for row_idx, row in enumerate(rows[2:], start=2):
        first = _get(row, 0)
        product = _normalize_text(_get(row, cols["product"]))
        if _is_sequence(first) and product and product.upper() not in {"TOTAL", "合计"}:
            primary_indices.append(row_idx)
    primary_indices.append(sheet.nrows)

    records = []
    for idx, row_idx in enumerate(primary_indices[:-1]):
        row = rows[row_idx]
        next_row_idx = primary_indices[idx + 1]
        children = []
        for child_idx in range(row_idx + 1, next_row_idx):
            child = rows[child_idx]
            contract = _normalize_text(_get(child, cols["purchase_contract"]))
            system_contract = _normalize_text(_get(child, cols["system_contract"]))
            buyer = _normalize_text(_get(child, cols["buyer"]))
            seller = _normalize_text(_get(child, cols["seller"]))
            lc_no = _normalize_text(_get(child, cols["lc_no"]))
            if not any([contract, system_contract, buyer, seller, lc_no]):
                continue
            child_currency, child_amount = _currency_amount(child, cols["amount"])
            children.append(
                {
                    "source_row": child_idx + 1,
                    "contract": contract,
                    "system_contract": system_contract,
                    "lc_no": lc_no,
                    "buyer": buyer,
                    "seller": seller,
                    "currency": child_currency,
                    "amount": child_amount,
                    "lc_date": _normalize_date(_get(child, cols["lc_date"])),
                    "lc_bank": _normalize_text(_get(child, cols["lc_bank"])),
                    "lc_latest_shipment_date": _normalize_date(_get(child, cols["lc_latest_ship"])),
                    "bill_of_lading_date": _normalize_date(_get(child, cols["bl_date"])),
                    "collection_date": _normalize_date(_get(child, cols["collection_date"])),
                    "remark": _normalize_text(_get(child, cols["remark"])),
                }
            )

        currency, amount = _currency_amount(row, cols["amount"])
        purchase_contract = _normalize_text(_get(row, cols["purchase_contract"]))
        system_contract = _normalize_text(_get(row, cols["system_contract"]))
        remark = _normalize_text(_get(row, cols["remark"])) or _normalize_text(_get(row, cols["status"]))
        record = {
            "business_key": _business_key(subsidiary, purchase_contract, system_contract, path.name, row_idx + 1),
            "subsidiary": subsidiary,
            "source_file": path.name,
            "source_sheet": sheet.name,
            "source_row_start": row_idx + 1,
            "source_row_end": next_row_idx,
            "source_snapshot_date": snapshot_date,
            "product_name": _normalize_text(_get(row, cols["product"])),
            "purchase_contract_no": purchase_contract,
            "system_contract_no": system_contract,
            "buyer": _normalize_text(_get(row, cols["buyer"])),
            "seller": _normalize_text(_get(row, cols["seller"])),
            "contract_date": _normalize_date(_get(row, cols["contract_date"])),
            "trade_term": _normalize_text(_get(row, cols["trade_term"])),
            "origin_port": _normalize_text(_get(row, cols["origin"])),
            "destination_port": _normalize_text(_get(row, cols["destination"])),
            "contract_quantity_mt": _to_float(_get(row, cols["quantity"])),
            "contract_currency": currency,
            "contract_amount": amount,
            "finance_amount_expected": _to_float(_get(row, cols["loan_expected"])),
            "finance_amount_actual": _to_float(_get(row, cols["loan_actual"])),
            "repaid_amount": _to_float(_get(row, cols["repaid"])),
            "remaining_credit_amount": _to_float(_get(row, cols["remaining"])),
            "finance_drawdown_date": _normalize_date(_get(row, cols["loan_date"])),
            "finance_due_date": _normalize_date(_get(row, cols["loan_due"])),
            "finance_days": _to_int(_get(row, cols["finance_days"])),
            "latest_shipment_date": _normalize_date(_get(row, cols["latest_ship"])),
            "lc_latest_shipment_date": _normalize_date(_get(row, cols["lc_latest_ship"])),
            "vessel_voyage": _normalize_text(_get(row, cols["vessel"])),
            "bill_of_lading_date": _normalize_date(_get(row, cols["bl_date"])),
            "bill_of_lading_no": _normalize_text(_get(row, cols["bl_no"])),
            "document_submission_date": _normalize_date(_get(row, cols["doc_date"])),
            "collection_date": _normalize_date(_get(row, cols["collection_date"])),
            "actual_shipped_quantity_mt": _to_float(_get(row, cols["ship_qty"])),
            "actual_goods_amount": _to_float(_get(row, cols["actual_amount"])),
            "tail_amount": _to_float(_get(row, cols["tail_amount"])),
            "tail_payment_date": _normalize_date(_get(row, cols["tail_date"])),
            "executor": _normalize_text(_get(row, cols["owner"])),
            "remark": remark,
            "sales_contracts_json": json.dumps(children, ensure_ascii=False),
            "settlement_json": "{}",
            "corrections_json": "[]",
            "source_json": json.dumps({"headers": headers, "row": row, "children": children}, ensure_ascii=False, default=str),
        }
        record["terminal_customer"] = _terminal_customer(children, record["buyer"])
        record["overseas_entity"] = _overseas_entity(row, children, cols["buyer"], cols["seller"])
        record["finance_bank"] = _derive_bank(record["product_name"], _get(row, cols["lc_bank"]), remark)
        derived = derive_business_status(record)
        record.update(derived)
        record["finance_status"] = record["business_status"]
        warnings = _build_warnings(record)
        record["import_warnings_json"] = json.dumps(warnings, ensure_ascii=False)
        records.append(record)

    return {
        "file": path.name,
        "sheet": sheet.name,
        "records": records,
        "summary": {"record_count": len(records), "warning_count": _data_quality_warning_count(records)},
    }


def _clean_xlsx_value(value: Any) -> Any:
    if value in (None, "", 0, "0"):
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text and text not in {"0", "-", "—"} else None
    return value


def _xlsx_text(value: Any) -> str:
    value = _clean_xlsx_value(value)
    return "" if value is None else str(value).strip()


def _xlsx_float(value: Any, scale: float = 1.0) -> Optional[float]:
    value = _clean_xlsx_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value) / scale, 6)
    text = str(value).replace(",", "").strip()
    try:
        return round(float(text) / scale, 6)
    except ValueError:
        return None


def _xlsx_entity(item_no: str) -> str:
    return "香港建龙" if item_no.startswith("H-") else "YOLANDA"


def _xlsx_business_key_base(record: Dict[str, Any], source_file: str, source_row: int) -> str:
    parts = [
        _normalize_text(record.get("subsidiary")),
        _normalize_text(record.get("purchase_contract_no")),
        _normalize_text(record.get("system_contract_no")),
    ]
    if any(parts[1:]):
        return "|".join(parts)
    return "|".join([parts[0], source_file, str(source_row)])


def _xlsx_business_key_suffix(record: Dict[str, Any], source_row: int) -> str:
    return "|".join([
        _normalize_text(record.get("finance_bank")),
        _normalize_text(record.get("finance_drawdown_date")),
        str(record.get("finance_amount_actual") or ""),
        str(source_row),
    ])


def _xlsx_row_record(path: Path, sheet_name: str, headers: List[str], values: tuple[Any, ...], row_idx: int) -> Optional[Dict[str, Any]]:
    row = dict(zip(headers, values))
    item_no = _xlsx_text(row.get("项次"))
    if not item_no:
        return None
    subsidiary = _xlsx_text(row.get("供应商简称")) or _subsidiary_from_filename(_xlsx_text(row.get("供应商")))
    finance_due = _normalize_xlsx_date(row.get("新到期日")) or _normalize_xlsx_date(row.get("原到期日"))
    repay_date = _normalize_xlsx_date(row.get("还款日"))
    loan_status = _xlsx_text(row.get("贷款状态"))
    lc_contract = _xlsx_text(row.get("双方合同号"))
    source_date = ""
    if "-2025-" in item_no:
        source_date = "2025-01-01"
    elif "-2026-" in item_no:
        source_date = "2026-01-01"
    else:
        source_date = date.today().isoformat()
    record = {
        "subsidiary": subsidiary,
        "source_file": path.name,
        "source_sheet": sheet_name,
        "source_row_start": row_idx,
        "source_row_end": row_idx,
        "source_snapshot_date": source_date,
        "product_name": _xlsx_text(row.get("品名")),
        "purchase_contract_no": _xlsx_text(row.get("合同编号")),
        "system_contract_no": _xlsx_text(row.get("系统合同号")),
        "buyer": _xlsx_text(row.get("合同买方")),
        "seller": _xlsx_text(row.get("供应商")),
        "overseas_entity": _xlsx_entity(item_no),
        "terminal_customer": _xlsx_text(row.get("合同买方")),
        "contract_date": _normalize_xlsx_date(row.get("付款日期")),
        "trade_term": _xlsx_text(row.get("LC类型")),
        "origin_port": _xlsx_text(row.get("起运港")),
        "destination_port": _xlsx_text(row.get("目的港")),
        "contract_quantity_mt": _xlsx_float(row.get("合同数量(吨)")),
        "contract_currency": _xlsx_text(row.get("合同币别")) or "CNY",
        "contract_amount": _xlsx_float(row.get("付款金额")),
        "finance_bank": _xlsx_text(row.get("贷款行")),
        "finance_amount_expected": _xlsx_float(row.get("贷款人民币金额")),
        "finance_amount_actual": _xlsx_float(row.get("贷款人民币金额")),
        "repaid_amount": _xlsx_float(row.get("贷款人民币金额")) if repay_date or loan_status == "已还款" else None,
        "remaining_credit_amount": None,
        "finance_drawdown_date": _normalize_xlsx_date(row.get("借款日期")),
        "finance_due_date": finance_due,
        "finance_days": None,
        "finance_status": loan_status,
        "latest_shipment_date": _normalize_xlsx_date(row.get("最迟装船日")),
        "lc_latest_shipment_date": _normalize_xlsx_date(row.get("LC有效期")),
        "vessel_voyage": _xlsx_text(row.get("船名")),
        "bill_of_lading_date": "",
        "bill_of_lading_no": "",
        "document_submission_date": _normalize_xlsx_date(row.get("交单日期")),
        "collection_date": _normalize_xlsx_date(row.get("收汇日期")),
        "actual_shipped_quantity_mt": None,
        "actual_goods_amount": _xlsx_float(row.get("交单金额")),
        "tail_amount": None,
        "tail_payment_date": "",
        "executor": "",
        "remark": _xlsx_text(row.get("情况说明")) or loan_status,
        "sales_contracts_json": json.dumps([{
            "item_no": item_no,
            "contract": lc_contract,
            "lc_no": _xlsx_text(row.get("信用证编号")),
            "lc_bank": _xlsx_text(row.get("开证银行")),
            "lc_amount": _xlsx_float(row.get("信用证金额")),
            "lc_issue_date": _normalize_xlsx_date(row.get("开证日期")),
            "lc_expiry_date": _normalize_xlsx_date(row.get("LC有效期")),
            "lc_type": _xlsx_text(row.get("LC类型")),
            "transferable": _xlsx_text(row.get("是否可转让")),
            "receiving_bank": _xlsx_text(row.get("收证行")),
            "discount_date": _normalize_xlsx_date(row.get("贴现日期")),
        }], ensure_ascii=False),
        "settlement_json": "{}",
        "corrections_json": "[]",
        "source_json": json.dumps({"headers": headers, "row": list(values), "item_no": item_no}, ensure_ascii=False, default=str),
    }
    record["business_key"] = _xlsx_business_key_base(record, path.name, row_idx)
    derived = derive_business_status(record)
    if loan_status == "已还款" or repay_date:
        derived = {"business_status": "已结算", "risk_level": "低", "next_action": "无"}
    record.update(derived)
    record["finance_status"] = loan_status or record["business_status"]
    warnings = _build_warnings(record)
    record["import_warnings_json"] = json.dumps(warnings, ensure_ascii=False)
    return record


def _compact_header(value: Any) -> str:
    return _xlsx_text(value).replace("\n", "").replace(" ", "")


def _find_xlsx_header_row(sheet, aliases: tuple[str, ...] = ("项次",)) -> tuple[int, List[Any]]:
    normalized_aliases = {_compact_header(alias) for alias in aliases}
    for row_idx, values in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), start=1):
        headers = [_compact_header(value) for value in values]
        if any(alias in headers for alias in normalized_aliases):
            return row_idx, list(values)
    raise ValueError(f"{sheet.title}页签未找到项次表头")


def _row_alias(row: Dict[str, Any], *aliases: str) -> Any:
    compact = {_compact_header(key): value for key, value in row.items()}
    for alias in aliases:
        key = _compact_header(alias)
        if key in compact:
            return compact[key]
    return None


def _normalized_order_status(value: Any) -> str:
    text = _xlsx_text(value)
    if "结案" in text or text in {"已完成", "已结算"}:
        return "结案"
    if "存续" in text or text in {"进行中", "未结案"}:
        return "存续"
    return text


def _alerts_grouped_by_item(sheet) -> Dict[str, List[Dict[str, str]]]:
    alerts: Dict[str, List[Dict[str, str]]] = {}
    rows = [list(values) for values in sheet.iter_rows(values_only=True)]
    for index, values in enumerate(rows):
        compact = [_compact_header(value) for value in values]
        if "项次" not in compact:
            continue
        item_col = compact.index("项次")
        title = ""
        if index > 0:
            title = next((_xlsx_text(value) for value in rows[index - 1] if _xlsx_text(value)), "")
        row_index = index + 1
        while row_index < len(rows):
            next_values = rows[row_index]
            next_compact = [_compact_header(value) for value in next_values]
            if "项次" in next_compact:
                break
            item_no = _xlsx_text(next_values[item_col] if item_col < len(next_values) else None)
            if not item_no:
                break
            if item_no.startswith("#"):
                row_index += 1
                continue
            message_parts = [_xlsx_text(value) for value in next_values if _xlsx_text(value)]
            message = title or "Excel预警"
            if message_parts:
                message = f"{message}：{' / '.join(message_parts[1:])}" if len(message_parts) > 1 else message
            alerts.setdefault(item_no, []).append({
                "field": "excel_alert",
                "level": "高",
                "message": message,
                "source_sheet": sheet.title,
                "source_row": str(row_index + 1),
            })
            row_index += 1
    return alerts


def _quota_label_rows(sheet) -> Dict[str, List[int]]:
    labels: Dict[str, List[int]] = {}
    for row_idx in range(1, sheet.max_row + 1):
        label = _compact_header(sheet.cell(row_idx, 1).value)
        if label:
            labels.setdefault(label, []).append(row_idx)
    return labels


def _quota_numeric_value(sheet, row_indices: List[int], col_idx: int) -> Optional[float]:
    value = None
    for row_idx in row_indices:
        parsed = _xlsx_float(sheet.cell(row_idx, col_idx).value)
        if parsed is not None:
            value = parsed
    return value


def _parse_quota_sheet(book) -> Dict[str, Any]:
    if "额度" not in book.sheetnames:
        return {"banks": [], "total_credit": 0.0, "used_credit": 0.0, "available_credit": 0.0}
    sheet = book["额度"]
    labels = _quota_label_rows(sheet)
    condition_rows = labels.get("限定工厂", [])
    if not condition_rows:
        return {"banks": [], "total_credit": 0.0, "used_credit": 0.0, "available_credit": 0.0}
    bank_header_row = condition_rows[0] - 1
    unit_text = " ".join(
        _xlsx_text(sheet.cell(row, col).value)
        for row in range(1, min(sheet.max_row, 5) + 1)
        for col in range(1, min(sheet.max_column, 12) + 1)
    )
    multiplier = 10000.0 if "万元" in unit_text else 1.0
    banks = []
    for col_idx in range(2, sheet.max_column + 1):
        bank = _xlsx_text(sheet.cell(bank_header_row, col_idx).value)
        if not bank:
            continue
        limit = _quota_numeric_value(sheet, labels.get("授信额度", []), col_idx)
        used = _quota_numeric_value(sheet, labels.get("目前占用额度", []), col_idx)
        available = _quota_numeric_value(sheet, labels.get("目前可用额度", []), col_idx)
        if limit is None and used is None and available is None:
            continue
        banks.append({
            "bank": bank,
            "limit": (limit or 0.0) * multiplier,
            "used": used * multiplier if used is not None else None,
            "available": available * multiplier if available is not None else None,
            "note": _xlsx_text(sheet.cell(condition_rows[0], col_idx).value),
            "lc_requirement": _xlsx_text(sheet.cell(labels.get("信用证要求", [condition_rows[0]])[0], col_idx).value),
            "bill_requirement": _xlsx_text(sheet.cell(labels.get("提单要求", [condition_rows[0]])[0], col_idx).value),
            "finance_ratio": _xlsx_text(sheet.cell(labels.get("订单融资比例", [condition_rows[0]])[0], col_idx).value),
            "term": _xlsx_text(sheet.cell(labels.get("期限", [condition_rows[0]])[0], col_idx).value),
        })
    total_credit = sum(float(bank["limit"] or 0) for bank in banks)
    used_values = [bank["used"] for bank in banks if bank["used"] is not None]
    used_credit = sum(float(value) for value in used_values) if used_values else None
    available_values = [bank["available"] for bank in banks if bank["available"] is not None]
    return {
        "banks": banks,
        "total_credit": total_credit,
        "used_credit": used_credit,
        "available_credit": sum(available_values) if len(available_values) == len(banks) else (total_credit - used_credit if used_credit is not None else None),
        "unit": "元",
    }


def _order_sheet_record(
    path: Path,
    sheet_name: str,
    headers: List[Any],
    values: tuple[Any, ...],
    row_idx: int,
    alerts_by_item: Dict[str, List[Dict[str, str]]],
) -> Optional[Dict[str, Any]]:
    row = dict(zip(headers, values))
    item_no = _xlsx_text(_row_alias(row, "项次", "订单项次"))
    if not item_no or item_no.startswith("#") or item_no in {"合计", "TOTAL"}:
        return None
    supplier_short = _xlsx_text(_row_alias(row, "供应商简称", "钢厂", "发货方", "供应商"))
    supplier_full = _xlsx_text(_row_alias(row, "供应商", "发货方", "钢厂"))
    finance_amount = _xlsx_float(_row_alias(row, "贷款人民币金额", "融资金额", "放款金额"))
    status = _normalized_order_status(_row_alias(row, "状态", "订单状态", "存续/结案", "贷款状态"))
    repay_date = _normalize_xlsx_date(_row_alias(row, "还款日", "还款日期"))
    original_due = _normalize_xlsx_date(_row_alias(row, "原到期日"))
    extension_days = _to_int(_row_alias(row, "展期天数")) or 0
    new_due = _normalize_xlsx_date(_row_alias(row, "新到期日", "融资到期日", "到期日"))
    finance_due = _effective_finance_due(
        new_due,
        original_due,
        extension_days,
    )
    latest_shipment_date = _normalize_xlsx_date(_row_alias(row, "最迟装船日", "最晚装船日"))
    bill_date = _normalize_xlsx_date(_row_alias(row, "提单日", "提单日期"))
    document_date = _normalize_xlsx_date(_row_alias(row, "交单日", "交单日期", "银行交单日"))
    source_date = f"{item_no.split('-')[1]}-01-01" if len(item_no.split("-")) > 2 and item_no.split("-")[1].isdigit() else date.today().isoformat()
    source_meta = {
        "item_no": item_no,
        "headers": [str(header or "") for header in headers],
        "row": list(values),
        "finance_rate": _xlsx_float(_row_alias(row, "利率")),
        "original_due_date": original_due,
        "new_due_date": new_due,
        "extension_days": extension_days,
        "order_status": status,
        "alerts": alerts_by_item.get(item_no, []),
    }
    record = {
        "business_key": f"ITEM|{item_no}|1",
        "subsidiary": supplier_short or supplier_full or "未填供应商",
        "source_file": path.name,
        "source_sheet": sheet_name,
        "source_row_start": row_idx,
        "source_row_end": row_idx,
        "source_snapshot_date": source_date,
        "product_name": _xlsx_text(_row_alias(row, "品种材质", "品名", "品种", "材质")),
        "purchase_contract_no": _xlsx_text(_row_alias(row, "合同编号", "合同号", "合同")),
        "system_contract_no": _xlsx_text(_row_alias(row, "系统合同号")),
        "buyer": _xlsx_text(_row_alias(row, "合同买方", "买方")),
        "seller": supplier_full or supplier_short,
        "overseas_entity": _xlsx_entity(item_no),
        "terminal_customer": _xlsx_text(_row_alias(row, "合同买方", "终端客户", "客户")),
        "contract_date": _normalize_xlsx_date(_row_alias(row, "合同日期")),
        "trade_term": _xlsx_text(_row_alias(row, "贸易条款", "价格条款")),
        "origin_port": _xlsx_text(_row_alias(row, "起运港")),
        "destination_port": _xlsx_text(_row_alias(row, "目的港", "卸港")),
        "contract_quantity_mt": _xlsx_float(_row_alias(row, "合同数量(吨)", "吨数", "合同数量", "数量")),
        "contract_currency": _xlsx_text(_row_alias(row, "合同币别", "币种")) or "CNY",
        "contract_amount": _xlsx_float(_row_alias(row, "合同金额", "付款金额")),
        "finance_bank": _xlsx_text(_row_alias(row, "贷款行", "融资银行")),
        "finance_amount_expected": finance_amount,
        "finance_amount_actual": finance_amount,
        "repaid_amount": finance_amount if repay_date else None,
        "remaining_credit_amount": None,
        "finance_drawdown_date": _normalize_xlsx_date(_row_alias(row, "借款日期", "借款日", "放款日期")),
        "finance_due_date": finance_due,
        "finance_days": _to_int(_row_alias(row, "实际融资周期", "融资天数")),
        "finance_status": status,
        "latest_shipment_date": latest_shipment_date,
        "lc_latest_shipment_date": "",
        "vessel_voyage": "",
        "bill_of_lading_date": bill_date,
        "bill_of_lading_no": _xlsx_text(_row_alias(row, "提单号")),
        "document_submission_date": document_date,
        "collection_date": _normalize_xlsx_date(_row_alias(row, "收汇日期", "收汇日")),
        "actual_shipped_quantity_mt": None,
        "actual_goods_amount": None,
        "tail_amount": None,
        "tail_payment_date": repay_date,
        "executor": _xlsx_text(_row_alias(row, "执行人员", "负责人")),
        "business_status": status,
        "risk_level": "低" if status == "结案" else "中",
        "remark": _xlsx_text(_row_alias(row, "情况说明", "备注")),
        "sales_contracts_json": "[]",
        "settlement_json": "{}",
        "corrections_json": "[]",
        "source_json": json.dumps(source_meta, ensure_ascii=False, default=str),
    }
    warnings = _build_warnings(record) + list(alerts_by_item.get(item_no, []))
    if not status:
        warnings.append({"field": "business_status", "level": "高", "message": "存续/结案状态为空"})
    record["import_warnings_json"] = json.dumps(warnings, ensure_ascii=False)
    return record


def _parse_order_sheet(book, path: Path, alerts_by_item: Dict[str, List[Dict[str, str]]], capital: Dict[str, Any]) -> List[Dict[str, Any]]:
    sheet = book["订单"]
    header_row, headers = _find_xlsx_header_row(sheet)
    records: List[Dict[str, Any]] = []
    by_item: Dict[str, List[Dict[str, Any]]] = {}
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, values_only=True), start=header_row + 1):
        record = _order_sheet_record(path, sheet.title, headers, values, row_idx, alerts_by_item)
        if not record:
            continue
        item_no = _item_no(record)
        siblings = by_item.setdefault(item_no, [])
        record["business_key"] = f"ITEM|{item_no}|{len(siblings) + 1}"
        siblings.append(record)
        records.append(record)
    active_amount = sum(
        float(record.get("finance_amount_actual") or record.get("finance_amount_expected") or 0)
        for record in records
        if record.get("business_status") == "存续"
    )
    quota_used = float(capital.get("used_credit") or 0)
    amount_multiplier = quota_used / active_amount if active_amount and quota_used else 1.0
    if 5000 <= amount_multiplier <= 15000:
        for record in records:
            for field in ("finance_amount_expected", "finance_amount_actual", "repaid_amount"):
                if record.get(field) is not None:
                    record[field] = float(record[field]) * 10000
            source = _json_loads(record.get("source_json"), {})
            source["finance_amount_unit"] = "万元"
            record["source_json"] = json.dumps(source, ensure_ascii=False, default=str)
    if records and capital.get("banks"):
        source = _json_loads(records[0].get("source_json"), {})
        source["workbook_capital"] = capital
        records[0]["source_json"] = json.dumps(source, ensure_ascii=False, default=str)
    return records


def parse_order_finance_xlsx_workbook(path: Path) -> Dict[str, Any]:
    book = load_workbook(path, data_only=True, read_only=False)
    if "订单" not in book.sheetnames:
        raise ValueError("Excel 缺少必需的订单页签")
    sheets = {name: name in book.sheetnames for name in TARGET_XLSX_SHEETS}
    alerts_by_item = _alerts_grouped_by_item(book["预警"]) if sheets["预警"] else {}
    capital = _parse_quota_sheet(book)
    records = _parse_order_sheet(book, path, alerts_by_item, capital)
    return {
        "file": path.name,
        "sheet": "订单",
        "sheets": sheets,
        "capital": capital,
        "records": records,
        "summary": {"record_count": len(records), "warning_count": _data_quality_warning_count(records)},
    }


def parse_order_finance_directory(directory: Path | str) -> Dict[str, Any]:
    base = Path(directory)
    if not base.exists():
        raise ValueError(f"目录不存在：{base}")
    if base.is_file():
        files = [base]
    else:
        files = sorted(
            path for path in base.iterdir()
            if path.suffix.lower() in {".xls", ".xlsx"} and not path.name.startswith("~$")
        )
    records: List[Dict[str, Any]] = []
    file_results = []
    for path in files:
        if path.suffix.lower() == ".xlsx":
            result = parse_order_finance_xlsx_workbook(path)
        else:
            result = parse_order_finance_workbook(path)
        records.extend(result["records"])
        file_results.append({
            "file": result["file"],
            "sheet": result["sheet"],
            "sheets": result.get("sheets"),
            **result["summary"],
        })
    return {
        "records": records,
        "files": file_results,
        "summary": {
            "files_read": len(files),
            "record_count": len(records),
            "warning_count": sum(item["warning_count"] for item in file_results),
        },
    }


def _json_or_empty(value: Any, empty: str = "{}") -> str:
    if value in (None, ""):
        return empty
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def _row_to_dict(row: Any) -> Dict[str, Any]:
    return dict(row) if row else {}


def _normalize_order_vessel_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return " / ".join(
        part.strip()
        for part in re.split(r"[\r\n]+", str(value))
        if part.strip()
    )


def _resolve_order_vessel_sheet(book):
    for sheet_name in (ORDER_VESSEL_CURRENT_R1_SHEET, ORDER_VESSEL_SHEET):
        if sheet_name in book.sheetnames:
            return book[sheet_name]
    candidates = [name for name in book.sheetnames if name.endswith("钢材出口情况表")]
    if len(candidates) == 1:
        return book[candidates[0]]
    raise ValueError("Excel 缺少唯一可识别的钢材出口情况表页签")


def _resolve_order_vessel_headers(sheet) -> Dict[str, int]:
    headers = {
        _normalize_order_vessel_text(sheet.cell(2, column).value): column
        for column in range(1, sheet.max_column + 1)
        if _normalize_order_vessel_text(sheet.cell(2, column).value)
    }
    resolved: Dict[str, int] = {}
    missing: List[str] = []
    for field, aliases in ORDER_VESSEL_HEADER_ALIASES.items():
        column = next((headers[alias] for alias in aliases if alias in headers), None)
        if column is None:
            missing.append(aliases[0])
        else:
            resolved[field] = column
    if missing:
        raise ValueError("Excel 缺少字段：" + "、".join(missing))
    return resolved


def parse_order_vessel_snapshot(path: Path | str) -> Dict[str, Any]:
    workbook_path = Path(path)
    if not workbook_path.exists() or not workbook_path.is_file():
        raise ValueError(f"Excel 文件不存在：{workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("订单与船舶快照仅支持 .xlsx 文件")

    source_sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    book = load_workbook(workbook_path, data_only=True, read_only=False)
    sheet = _resolve_order_vessel_sheet(book)

    title = _normalize_order_vessel_text(sheet.cell(1, 1).value)
    title_date = re.search(r"(\d{4})[./年-](\d{1,2})[./月-](\d{1,2})", title)
    if not title_date:
        raise ValueError("Excel 标题缺少可识别的来源日期")
    source_date = date(*(int(part) for part in title_date.groups())).isoformat()
    source_version = f"{source_date}:{source_sha256[:12]}"

    headers = _resolve_order_vessel_headers(sheet)

    records: List[Dict[str, Any]] = []
    seen_business_numbers: set[str] = set()
    for row_number in range(3, sheet.max_row + 1):
        raw = {
            field: sheet.cell(row_number, column).value
            for field, column in headers.items()
        }
        business_no = _normalize_order_vessel_text(raw["business_no"])
        steel_mill = _normalize_order_vessel_text(raw["steel_mill"])
        if not business_no or steel_mill == "总计":
            continue
        if business_no in seen_business_numbers:
            raise ValueError(f"业务编号重复：{business_no}")
        seen_business_numbers.add(business_no)

        loan_amount = _to_float(raw["loan_amount"])
        loan_amount_note = "" if loan_amount is not None else _normalize_order_vessel_text(raw["loan_amount"])
        destination = _final_business_destination(raw["export_user"])
        records.append({
            "source_version": source_version,
            "source_date": source_date,
            "source_file_name": workbook_path.name,
            "source_sheet_name": sheet.title,
            "source_sha256": source_sha256,
            "source_row": row_number,
            "business_no": business_no,
            "steel_mill": steel_mill,
            "export_user": _normalize_order_vessel_text(raw["export_user"]),
            "cargo": _normalize_order_vessel_text(raw["cargo"]),
            "vessel": _normalize_order_vessel_text(raw["vessel"]),
            "quantity_mt": _to_float(raw["quantity_mt"]),
            "loading_port": _normalize_order_vessel_text(raw["loading_port"]),
            "loading_port_arrival_date": _normalize_date(raw["loading_port_arrival_date"]),
            "planned_berth_date": _normalize_date(raw["planned_berth_date"]),
            "discharge_port": _normalize_order_vessel_text(raw["discharge_port"]),
            "estimated_discharge_date": _normalize_date(raw["estimated_discharge_date"]),
            "document_status": _normalize_order_vessel_text(raw["document_status"]),
            "repayment_due_date": _normalize_date(raw["repayment_due_date"]),
            "loan_amount": loan_amount,
            "loan_amount_note": loan_amount_note,
            "remark": _normalize_order_vessel_text(raw["remark"]),
            "route_distance_nm": _to_float(raw["route_distance_nm"]),
            "eta_start_date": _normalize_date(raw["eta_start_date"]),
            "estimated_speed_knots": _to_float(raw["estimated_speed_knots"]),
            "eta_basis": _normalize_order_vessel_text(raw["eta_basis"]),
            "route_source": _normalize_order_vessel_text(raw["route_source"]),
            "final_destination_status": destination["status"],
            "final_destination_source": destination["source"],
            "reporting_due_date_source": "当前确认R1",
            "email_due_values_json": "[]",
            "email_due_source": "",
            "email_due_source_date": "",
            "preview_status": "shadow",
        })

    if not records:
        raise ValueError("Excel 中没有可导入的业务记录")
    return {
        "source": {
            "date": source_date,
            "version": source_version,
            "file_name": workbook_path.name,
            "sheet_name": sheet.title,
            "sha256": source_sha256,
        },
        "records": records,
        "summary": {
            "record_count": len(records),
            "quantity_mt": sum(float(row.get("quantity_mt") or 0) for row in records),
            "loan_amount": sum(float(row.get("loan_amount") or 0) for row in records),
            "non_financing_count": sum(1 for row in records if row.get("loan_amount_note")),
        },
    }


def _order_vessel_values_equal(existing: Dict[str, Any], incoming: Dict[str, Any]) -> bool:
    numeric_fields = {
        "source_row", "quantity_mt", "loan_amount", "route_distance_nm",
        "estimated_speed_knots",
    }
    for field in ORDER_VESSEL_SNAPSHOT_FIELDS:
        left = existing.get(field)
        right = incoming.get(field)
        if field in numeric_fields:
            left = _to_float(left)
            right = _to_float(right)
        else:
            left = "" if left is None else str(left)
            right = "" if right is None else str(right)
        if left != right:
            return False
    return True


def apply_order_vessel_snapshot(
    records: List[Dict[str, Any]],
    *,
    imported_by: str = "",
) -> Dict[str, int]:
    if not records:
        raise ValueError("订单与船舶快照不能为空")
    source_versions = {_normalize_text(row.get("source_version")) for row in records}
    if len(source_versions) != 1 or "" in source_versions:
        raise ValueError("一次导入只能包含一个有效来源版本")
    if len({_normalize_text(row.get("business_no")) for row in records}) != len(records):
        raise ValueError("订单与船舶快照包含重复业务编号")

    source_version = next(iter(source_versions))
    inserted = updated = unchanged = 0
    with db.connect() as conn:
        cur = conn.cursor()
        for incoming in records:
            existing_row = db._exec(
                cur,
                "SELECT * FROM order_vessel_snapshots WHERE source_version = ? AND business_no = ?",
                (source_version, incoming["business_no"]),
            ).fetchone()
            existing = _row_to_dict(existing_row)
            if existing and _order_vessel_values_equal(existing, incoming) and int(existing.get("is_active") or 0) == 1:
                unchanged += 1
                continue
            values = [incoming.get(field) for field in ORDER_VESSEL_SNAPSHOT_FIELDS]
            if existing:
                assignments = ", ".join(f"{field} = ?" for field in ORDER_VESSEL_SNAPSHOT_FIELDS)
                db._exec(
                    cur,
                    f"UPDATE order_vessel_snapshots SET {assignments}, is_active = 1, imported_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    tuple(values + [imported_by, existing["id"]]),
                )
                updated += 1
            else:
                insert_fields = ORDER_VESSEL_SNAPSHOT_FIELDS + ["is_active", "imported_by"]
                placeholders = ", ".join("?" for _ in insert_fields)
                db._exec(
                    cur,
                    f"INSERT INTO order_vessel_snapshots ({', '.join(insert_fields)}) VALUES ({placeholders})",
                    tuple(values + [1, imported_by]),
                )
                inserted += 1
        deactivated = db._exec(
            cur,
            "UPDATE order_vessel_snapshots SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE is_active = 1 AND source_version <> ?",
            (source_version,),
        ).rowcount
    return {
        "inserted": inserted,
        "updated": updated,
        "unchanged": unchanged,
        "deactivated": max(int(deactivated or 0), 0),
    }


def apply_order_vessel_email_due_checks(
    checks: List[Dict[str, Any]],
) -> Dict[str, int]:
    business_numbers = [_normalize_text(check.get("business_no")) for check in checks]
    if not checks or any(not value for value in business_numbers):
        raise ValueError("邮件核对结果必须包含业务编号")
    if len(set(business_numbers)) != len(business_numbers):
        raise ValueError("邮件核对结果包含重复业务编号")

    updated = unchanged = 0
    with db.connect() as conn:
        cur = conn.cursor()
        for check, business_no in zip(checks, business_numbers):
            row = db._exec(
                cur,
                "SELECT * FROM order_vessel_snapshots WHERE is_active = 1 AND business_no = ?",
                (business_no,),
            ).fetchone()
            existing = _row_to_dict(row)
            if not existing:
                raise ValueError(f"邮件核对业务编号未命中当前R1：{business_no}")
            email_values = []
            for value in (check.get("email_due_dates") or []):
                normalized = _normalize_date(value)
                if normalized and not _parse_date(normalized):
                    raise ValueError(f"邮件核对日期格式不正确：{business_no}")
                if normalized and normalized not in email_values:
                    email_values.append(normalized)
            email_values.sort()
            values_json = json.dumps(email_values, ensure_ascii=False, separators=(",", ":"))
            email_source = _normalize_text(check.get("source"))
            email_source_date = _normalize_date(check.get("source_date"))
            if not email_source or not email_source_date:
                raise ValueError(f"邮件核对缺少来源或日期：{business_no}")
            current_values = _email_due_values(existing)
            if (
                current_values == email_values
                and _normalize_text(existing.get("email_due_source")) == email_source
                and _normalize_date(existing.get("email_due_source_date")) == email_source_date
            ):
                unchanged += 1
                continue
            db._exec(
                cur,
                """
                UPDATE order_vessel_snapshots
                SET email_due_values_json = ?, email_due_source = ?, email_due_source_date = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (values_json, email_source, email_source_date, existing["id"]),
            )
            updated += 1
    return {"updated": updated, "unchanged": unchanged}


def import_order_vessel_snapshot(
    path: Path | str,
    *,
    apply: bool = False,
    imported_by: str = "",
    expected_sha256: Optional[str] = None,
) -> Dict[str, Any]:
    parsed = parse_order_vessel_snapshot(path)
    actual_sha256 = parsed["source"]["sha256"]
    if expected_sha256 and actual_sha256.lower() != expected_sha256.lower():
        raise ValueError("Excel SHA-256 与锁定的定稿版本不一致")
    result = {**parsed["source"], **parsed["summary"], "applied": apply}
    if apply:
        result["changes"] = apply_order_vessel_snapshot(parsed["records"], imported_by=imported_by)
    return result


def list_order_vessel_snapshots() -> List[Dict[str, Any]]:
    with db.connect() as conn:
        cur = conn.cursor()
        rows = db._exec(
            cur,
            "SELECT * FROM order_vessel_snapshots WHERE is_active = 1 ORDER BY source_date DESC, source_row, id",
        ).fetchall()
    return [_row_to_dict(row) for row in rows]


def _serialize_record(record: Dict[str, Any]) -> Dict[str, Any]:
    item = dict(record)
    for field, empty in (
        ("sales_contracts_json", "[]"),
        ("settlement_json", "{}"),
        ("management_plan_json", "{}"),
        ("manual_change_log_json", "[]"),
        ("corrections_json", "[]"),
        ("import_warnings_json", "[]"),
        ("source_json", "{}"),
    ):
        item[field] = _json_or_empty(item.get(field), empty)
    return item


def upsert_order_finance_records(records: List[Dict[str, Any]], imported_by: str = "") -> Dict[str, int]:
    inserted = 0
    updated = 0
    with db.connect() as conn:
        cur = conn.cursor()
        for raw in records:
            record = _serialize_record(raw)
            existing = db._exec(
                cur,
                "SELECT id FROM order_finance_progress WHERE business_key = ?",
                (record["business_key"],),
            ).fetchone()
            if existing:
                assignments = ", ".join(f"{field} = ?" for field in FACT_FIELDS)
                params = [record.get(field) for field in FACT_FIELDS]
                params.append(existing["id"])
                db._exec(
                    cur,
                    f"UPDATE order_finance_progress SET {assignments}, is_archived = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    tuple(params),
                )
                updated += 1
            else:
                insert_fields = FACT_FIELDS + [
                    "planned_drawdown_date", "planned_finance_amount", "amount_adjustment_note",
                    "repayment_requirement", "repayment_requirement_status", "next_action",
                    "next_follow_up_date", "manager_note", "manual_override_fields",
                    "shipment_confirmed_date", "shipment_confirmed_by", "shipment_confirmed_at",
                    "management_plan_json", "manual_change_log_json",
                ]
                values = [record.get(field) for field in FACT_FIELDS]
                values.extend([
                    None, None, "", "", "", record.get("next_action", ""), "", "",
                    "[]", None, None, None, "{}", "[]",
                ])
                placeholders = ", ".join("?" for _ in insert_fields)
                db._exec(
                    cur,
                    f"INSERT INTO order_finance_progress ({', '.join(insert_fields)}) VALUES ({placeholders})",
                    tuple(values),
                )
                inserted += 1
    return {"inserted": inserted, "updated": updated}


def _fact_values_equal(existing: Dict[str, Any], incoming: Dict[str, Any]) -> bool:
    return all(existing.get(field) == incoming.get(field) for field in FACT_FIELDS)


def snapshot_business_keys_hash(records: List[Dict[str, Any]]) -> str:
    keys = sorted({_normalize_text(row.get("business_key")) for row in records})
    return hashlib.sha256("\n".join(keys).encode("utf-8")).hexdigest()


def list_order_finance_fact_snapshot_records() -> List[Dict[str, Any]]:
    field_sql = ", ".join(FACT_FIELDS)
    with db.connect() as conn:
        cur = conn.cursor()
        rows = db._exec(
            cur,
            f"""SELECT {field_sql}
                FROM order_finance_progress
                WHERE is_archived = 0 AND source_file != '手动新增'
                ORDER BY business_key""",
        ).fetchall()
    return [_row_to_dict(row) for row in rows]


def order_finance_facts_hash(records: List[Dict[str, Any]]) -> str:
    normalized = [
        {field: row.get(field) for field in FACT_FIELDS}
        for row in sorted(
            records,
            key=lambda item: str(item.get("business_key") or ""),
        )
    ]
    payload = json.dumps(
        normalized,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def get_active_synced_business_keys() -> set[str]:
    with db.connect() as conn:
        cur = conn.cursor()
        rows = db._exec(
            cur,
            """SELECT business_key FROM order_finance_progress
               WHERE is_archived = 0 AND source_file != '手动新增'""",
        ).fetchall()
    return {_normalize_text(dict(row).get("business_key")) for row in rows}


def record_pending_order_finance_shrink(
    source_version: str,
    business_keys_hash: str,
    record_count: int,
    attempt_slot: Optional[str],
) -> None:
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(
            cur,
            """UPDATE order_finance_sync_status
               SET pending_source_version = ?, pending_business_keys_hash = ?,
                   pending_record_count = ?,
                   last_attempt_slot = COALESCE(?, last_attempt_slot),
                   updated_at = CURRENT_TIMESTAMP WHERE id = 1""",
            (source_version, business_keys_hash, record_count, attempt_slot),
        )


def apply_order_finance_snapshot(
    records: List[Dict[str, Any]],
    imported_by: str = "",
    sync_success_at: Optional[str] = None,
    source_version: Optional[str] = None,
    attempt_slot: Optional[str] = None,
) -> Dict[str, int]:
    del imported_by
    inserted = 0
    updated = 0
    archived = 0
    serialized = [_serialize_record(record) for record in records]
    incoming_keys = {record["business_key"] for record in serialized}

    with db.connect() as conn:
        cur = conn.cursor()
        existing_rows = db._exec(cur, "SELECT * FROM order_finance_progress").fetchall()
        existing_by_key = {
            row["business_key"]: _row_to_dict(row)
            for row in existing_rows
        }

        for record in serialized:
            existing = existing_by_key.get(record["business_key"])
            if existing:
                if not _fact_values_equal(existing, record) or existing.get("is_archived"):
                    for key_date_field in ("document_submission_date", "tail_payment_date"):
                        if existing.get(key_date_field) and not record.get(key_date_field):
                            logger.warning(
                                "order_finance_key_date_cleared",
                                extra={
                                    "business_key": record["business_key"],
                                    "key_date_field": key_date_field,
                                },
                            )
                    assignments = ", ".join(f"{field} = ?" for field in FACT_FIELDS)
                    params = [record.get(field) for field in FACT_FIELDS]
                    params.append(existing["id"])
                    db._exec(
                        cur,
                        f"UPDATE order_finance_progress SET {assignments}, "
                        "is_archived = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                        tuple(params),
                    )
                    updated += 1
                continue

            insert_fields = FACT_FIELDS + [
                "planned_drawdown_date", "planned_finance_amount", "amount_adjustment_note",
                "repayment_requirement", "repayment_requirement_status", "next_action",
                "next_follow_up_date", "manager_note", "manual_override_fields",
                "shipment_confirmed_date", "shipment_confirmed_by", "shipment_confirmed_at",
                "management_plan_json", "manual_change_log_json",
            ]
            values = [record.get(field) for field in FACT_FIELDS]
            values.extend([
                None, None, "", "", "", record.get("next_action", ""), "", "",
                "[]", None, None, None, "{}", "[]",
            ])
            placeholders = ", ".join("?" for _ in insert_fields)
            db._exec(
                cur,
                f"INSERT INTO order_finance_progress ({', '.join(insert_fields)}) "
                f"VALUES ({placeholders})",
                tuple(values),
            )
            inserted += 1

        for existing in existing_by_key.values():
            if (
                not existing.get("is_archived")
                and existing.get("source_file") != "手动新增"
                and existing["business_key"] not in incoming_keys
            ):
                db._exec(
                    cur,
                    """UPDATE order_finance_progress
                       SET is_archived = 1, updated_at = CURRENT_TIMESTAMP
                       WHERE id = ?""",
                    (existing["id"],),
                )
                archived += 1

        changed_count = inserted + updated + archived
        if sync_success_at is not None:
            db._exec(
                cur,
                """UPDATE order_finance_sync_status
                   SET last_success_at = ?, changed_count = ?, source_version = ?,
                       last_attempt_slot = COALESCE(?, last_attempt_slot),
                       pending_source_version = NULL,
                       pending_business_keys_hash = NULL,
                       pending_record_count = 0,
                       updated_at = CURRENT_TIMESTAMP
                   WHERE id = 1""",
                (sync_success_at, changed_count, source_version, attempt_slot),
            )
        else:
            db._exec(
                cur,
                """UPDATE order_finance_sync_status
                   SET source_version = NULL, pending_source_version = NULL,
                       pending_business_keys_hash = NULL, pending_record_count = 0,
                       updated_at = CURRENT_TIMESTAMP
                   WHERE id = 1""",
            )

    return {
        "inserted": inserted,
        "updated": updated,
        "archived": archived,
        "changed_count": changed_count,
    }


def get_order_finance_sync_status() -> Dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(
            cur,
            """SELECT last_success_at, changed_count, source_version, last_attempt_slot,
                      pending_source_version, pending_business_keys_hash, pending_record_count
               FROM order_finance_sync_status WHERE id = 1""",
        ).fetchone()
    status = _row_to_dict(row)
    return {
        "last_success_at": status.get("last_success_at"),
        "changed_count": int(status.get("changed_count") or 0),
        "source_version": status.get("source_version"),
        "last_attempt_slot": status.get("last_attempt_slot"),
        "pending_source_version": status.get("pending_source_version"),
        "pending_business_keys_hash": status.get("pending_business_keys_hash"),
        "pending_record_count": int(status.get("pending_record_count") or 0),
    }


def claim_order_finance_sync_slot(slot_key: str) -> bool:
    with db.connect() as conn:
        cur = conn.cursor()
        result = db._exec(
            cur,
            """UPDATE order_finance_sync_status
               SET last_attempt_slot = ?, updated_at = CURRENT_TIMESTAMP
               WHERE id = 1 AND COALESCE(last_attempt_slot, '') != ?""",
            (slot_key, slot_key),
        )
        return int(getattr(result, "rowcount", 0) or 0) == 1


def record_unchanged_order_finance_sync(
    sync_success_at: str,
    source_version: str,
    attempt_slot: str,
) -> None:
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(
            cur,
            """UPDATE order_finance_sync_status
               SET last_success_at = ?, changed_count = 0, source_version = ?,
                   last_attempt_slot = ?, pending_source_version = NULL,
                   pending_business_keys_hash = NULL, pending_record_count = 0,
                   updated_at = CURRENT_TIMESTAMP
               WHERE id = 1""",
            (sync_success_at, source_version, attempt_slot),
        )


def archive_existing_excel_order_finance_records() -> int:
    with db.connect() as conn:
        cur = conn.cursor()
        result = db._exec(
            cur,
            """
            UPDATE order_finance_progress
            SET is_archived = 1, updated_at = CURRENT_TIMESTAMP
            WHERE is_archived = 0 AND source_file != '手动新增'
            """,
        )
    return int(getattr(result, "rowcount", 0) or 0)


def import_order_finance_directory(directory: Path | str, imported_by: str = "") -> Dict[str, Any]:
    parsed = parse_order_finance_directory(directory)
    changes = apply_order_finance_snapshot(parsed["records"], imported_by=imported_by)
    parsed["summary"].update(changes)
    return parsed


async def import_order_finance_upload(request: Request, file_name: str, imported_by: str = "") -> Dict[str, Any]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise ValueError("请选择 .xlsx 或 .xls 格式的订单融资台账")
    file_bytes = await request.body()
    if not file_bytes:
        raise ValueError("上传文件为空")
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)
        parsed = parse_order_finance_directory(tmp_path)
        for record in parsed["records"]:
            record["source_file"] = file_name
            source = _json_loads(record.get("source_json"), {})
            if isinstance(source, dict):
                source["uploaded_file_name"] = file_name
                record["source_json"] = json.dumps(source, ensure_ascii=False, default=str)
        for item in parsed.get("files", []):
            item["file"] = file_name
        changes = apply_order_finance_snapshot(parsed["records"], imported_by=imported_by)
        parsed["summary"].update(changes)
        return parsed
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink()


ORDER_FINANCE_LIST_FIELDS = [
    "id", "business_key", "subsidiary", "source_file", "source_sheet", "source_row_start",
    "source_snapshot_date", "product_name", "purchase_contract_no", "system_contract_no",
    "overseas_entity", "terminal_customer", "contract_quantity_mt", "contract_currency", "contract_amount",
    "finance_bank", "finance_amount_expected", "finance_amount_actual", "finance_drawdown_date",
    "finance_due_date", "latest_shipment_date", "port_confirmed_date", "port_confirmed_by",
    "port_confirmed_at", "shipment_confirmed_date", "shipment_confirmed_by",
    "shipment_confirmed_at", "vessel_voyage", "bill_of_lading_date",
    "document_submission_date", "collection_date", "actual_shipped_quantity_mt", "executor", "business_status",
    "risk_level", "planned_drawdown_date", "planned_finance_amount", "amount_adjustment_note",
    "repayment_requirement", "repayment_requirement_status", "next_action",
    "next_follow_up_date", "manager_note", "tail_payment_date", "sales_contracts_json",
    "import_warnings_json", "source_json", "created_at", "updated_at",
]


def _clamp_limit(limit: int) -> int:
    return max(1, min(limit or 5000, 5000))


def list_order_finance_records_page(limit: int = 5000, offset: int = 0) -> Dict[str, Any]:
    limit = _clamp_limit(limit)
    offset = max(0, offset or 0)
    field_sql = ", ".join(ORDER_FINANCE_LIST_FIELDS)
    with db.connect() as conn:
        cur = conn.cursor()
        total_row = db._exec(
            cur,
            "SELECT COUNT(*) AS c FROM order_finance_progress WHERE is_archived = 0",
        ).fetchone()
        rows = db._exec(
            cur,
            f"""
            SELECT {field_sql}
            FROM order_finance_progress
            WHERE is_archived = 0
            ORDER BY
                CASE risk_level WHEN '高' THEN 1 WHEN '中' THEN 2 ELSE 3 END,
                COALESCE(finance_due_date, '9999-12-31'),
                id
            LIMIT ? OFFSET ?
            """,
            (limit, offset),
        ).fetchall()
    records = [_row_to_dict(row) for row in rows]
    total = int(total_row["c"] or 0)
    return {
        "records": records,
        "pagination": {
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_more": offset + len(records) < total,
        },
    }


def list_order_finance_records() -> List[Dict[str, Any]]:
    return list_order_finance_records_page(limit=5000, offset=0)["records"]


def get_order_finance_record(record_id: int) -> Dict[str, Any]:
    with db.connect() as conn:
        cur = conn.cursor()
        row = db._exec(cur, "SELECT * FROM order_finance_progress WHERE id = ?", (record_id,)).fetchone()
    if not row:
        raise KeyError(record_id)
    return _row_to_dict(row)


def find_order_finance_duplicates(payload: Dict[str, Any], exclude_id: Optional[int] = None) -> Dict[str, Any]:
    candidate = dict(payload)
    candidate["subsidiary"] = _normalize_text(candidate.get("subsidiary"))
    candidate["purchase_contract_no"] = _normalize_text(candidate.get("purchase_contract_no"))
    candidate["system_contract_no"] = _normalize_text(candidate.get("system_contract_no"))
    exact = None
    if candidate["purchase_contract_no"] or candidate["system_contract_no"]:
        business_key = _manual_business_key(candidate)
        with db.connect() as conn:
            cur = conn.cursor()
            row = db._exec(
                cur,
                "SELECT * FROM order_finance_progress WHERE business_key = ? AND is_archived = 0",
                (business_key,),
            ).fetchone()
        exact = _row_to_dict(row)
        if exact and exclude_id and exact.get("id") == exclude_id:
            exact = None

    target_amount = (
        _to_float(candidate.get("planned_finance_amount"))
        or _to_float(candidate.get("finance_amount_actual"))
        or _to_float(candidate.get("finance_amount_expected"))
        or _to_float(candidate.get("contract_amount"))
    )
    target_customer = _normalize_text(candidate.get("terminal_customer")).upper()
    target_due = _normalize_date(candidate.get("finance_due_date"))
    similar = []
    if candidate["subsidiary"] and target_customer and target_due and target_amount is not None:
        for row in list_order_finance_records():
            if exclude_id and row.get("id") == exclude_id:
                continue
            if row.get("subsidiary") != candidate["subsidiary"]:
                continue
            row_customer = _normalize_text(row.get("terminal_customer")).upper()
            row_due = _normalize_date(row.get("finance_due_date"))
            row_amount = (
                _to_float(row.get("planned_finance_amount"))
                or _to_float(row.get("finance_amount_actual"))
                or _to_float(row.get("finance_amount_expected"))
                or _to_float(row.get("contract_amount"))
            )
            if row_customer == target_customer and row_due == target_due and row_amount == target_amount:
                similar.append(row)
    return {"exact": exact, "similar": similar[:5]}


def create_manual_order_finance_record(payload: Dict[str, Any], created_by: str = "") -> Dict[str, Any]:
    record = dict(payload)
    record["subsidiary"] = _normalize_text(record.get("subsidiary"))
    if not record["subsidiary"]:
        raise ValueError("子公司不能为空")
    duplicates = find_order_finance_duplicates(record)
    if duplicates["exact"]:
        raise DuplicateOrderFinanceError(duplicates["exact"])

    management_values = {field: record.get(field) for field in MANAGEMENT_FIELDS if record.get(field) not in (None, "")}
    manual_next_action = _normalize_text(record.get("next_action"))
    record.update(
        {
            "business_key": _manual_business_key(record),
            "source_file": "手动新增",
            "source_sheet": "",
            "source_row_start": None,
            "source_row_end": None,
            "source_snapshot_date": date.today().isoformat(),
            "purchase_contract_no": _normalize_text(record.get("purchase_contract_no")),
            "system_contract_no": _normalize_text(record.get("system_contract_no")),
            "terminal_customer": _normalize_text(record.get("terminal_customer")),
            "product_name": _normalize_text(record.get("product_name")),
            "contract_currency": _normalize_text(record.get("contract_currency")) or "CNY",
            "contract_quantity_mt": _to_float(record.get("contract_quantity_mt")),
            "contract_amount": _to_float(record.get("contract_amount")),
            "finance_bank": _normalize_text(record.get("finance_bank")),
            "finance_amount_expected": _to_float(record.get("finance_amount_expected")),
            "finance_amount_actual": _to_float(record.get("finance_amount_actual")),
            "finance_drawdown_date": _normalize_date(record.get("finance_drawdown_date")),
            "finance_due_date": _normalize_date(record.get("finance_due_date")),
            "latest_shipment_date": _normalize_date(record.get("latest_shipment_date")),
            "bill_of_lading_date": _normalize_date(record.get("bill_of_lading_date")),
            "collection_date": _normalize_date(record.get("collection_date")),
            "executor": _normalize_text(record.get("executor")),
            "remark": "手动新增",
            "sales_contracts_json": "[]",
            "settlement_json": "{}",
            "corrections_json": "[]",
            "source_json": json.dumps({"created_by": created_by, "source": "manual"}, ensure_ascii=False),
        }
    )
    derived = derive_business_status(record)
    record["business_status"] = derived["business_status"]
    record["risk_level"] = derived["risk_level"]
    record["finance_status"] = derived["business_status"]
    warnings = _build_warnings(record)
    record["import_warnings_json"] = json.dumps(warnings, ensure_ascii=False)

    insert_fields = FACT_FIELDS + [
        "planned_drawdown_date", "planned_finance_amount", "amount_adjustment_note",
        "repayment_requirement", "repayment_requirement_status", "next_action",
        "next_follow_up_date", "manager_note", "manual_override_fields",
        "shipment_confirmed_date", "shipment_confirmed_by", "shipment_confirmed_at",
        "management_plan_json", "manual_change_log_json",
    ]
    values = [record.get(field) for field in FACT_FIELDS]
    values.extend([
        _normalize_date(record.get("planned_drawdown_date")),
        _to_float(record.get("planned_finance_amount")),
        _normalize_text(record.get("amount_adjustment_note")),
        _normalize_text(record.get("repayment_requirement")),
        _normalize_text(record.get("repayment_requirement_status")),
        manual_next_action or derived["next_action"],
        _normalize_date(record.get("next_follow_up_date")),
        _normalize_text(record.get("manager_note")),
        json.dumps(sorted(management_values), ensure_ascii=False),
        _normalize_date(record.get("shipment_confirmed_date")),
        _normalize_text(record.get("shipment_confirmed_by")),
        _normalize_text(record.get("shipment_confirmed_at")),
        "{}",
        "[]",
    ])
    placeholders = ", ".join("?" for _ in insert_fields)
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(
            cur,
            f"INSERT INTO order_finance_progress ({', '.join(insert_fields)}) VALUES ({placeholders})",
            tuple(values),
        )
        record_id = db.last_insert_id(conn)
    return get_order_finance_record(record_id)


def update_management_fields(record_id: int, changes: Dict[str, Any], updated_by: str = "") -> Dict[str, Any]:
    allowed = {key: value for key, value in changes.items() if key in MANAGEMENT_FIELDS}
    if not allowed:
        return get_order_finance_record(record_id)
    before = get_order_finance_record(record_id)
    existing_log = json.loads(before.get("manual_change_log_json") or "[]")
    log_items = []
    for key, value in allowed.items():
        if before.get(key) != value:
            log_items.append({
                "field": key,
                "before": before.get(key),
                "after": value,
                "updated_by": updated_by,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            })
    override_fields = sorted(set(json.loads(before.get("manual_override_fields") or "[]")) | set(allowed))
    existing_log.extend(log_items)
    allowed["manual_override_fields"] = json.dumps(override_fields, ensure_ascii=False)
    allowed["manual_change_log_json"] = json.dumps(existing_log, ensure_ascii=False)

    assignments = ", ".join(f"{field} = ?" for field in allowed)
    params = list(allowed.values()) + [record_id]
    with db.connect() as conn:
        cur = conn.cursor()
        db._exec(
            cur,
            f"UPDATE order_finance_progress SET {assignments}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            tuple(params),
        )
    return get_order_finance_record(record_id)


def set_shipment_confirmation(
    item_no: str,
    confirmed: bool,
    shipment_confirmed_date: Optional[str] = None,
    updated_by: str = "",
) -> Dict[str, Any]:
    normalized_item = _normalize_text(item_no)
    matching = [row for row in list_order_finance_records() if _item_no(row) == normalized_item]
    if not matching:
        raise KeyError(normalized_item)
    if confirmed:
        normalized_date = _normalize_date(shipment_confirmed_date or date.today().isoformat())
        if not _parse_date(normalized_date):
            raise ValueError("实际装船日格式不正确")
        changes = {
            "shipment_confirmed_date": normalized_date,
            "shipment_confirmed_by": updated_by,
            "shipment_confirmed_at": datetime.now().isoformat(timespec="seconds"),
        }
    else:
        changes = {
            "shipment_confirmed_date": None,
            "shipment_confirmed_by": None,
            "shipment_confirmed_at": None,
        }
    for row in matching:
        update_management_fields(row["id"], changes, updated_by=updated_by)
    return {"item_no": normalized_item, "confirmed": confirmed, "updated": len(matching)}


def set_port_confirmation(
    item_no: str,
    confirmed: bool,
    port_confirmed_date: Optional[str] = None,
    updated_by: str = "",
) -> Dict[str, Any]:
    normalized_item = _normalize_text(item_no)
    matching = [row for row in list_order_finance_records() if _item_no(row) == normalized_item]
    if not matching:
        raise KeyError(normalized_item)
    if confirmed:
        normalized_date = _normalize_date(port_confirmed_date or date.today().isoformat())
        if not _parse_date(normalized_date):
            raise ValueError("实际集港日格式不正确")
        changes = {
            "port_confirmed_date": normalized_date,
            "port_confirmed_by": updated_by,
            "port_confirmed_at": datetime.now().isoformat(timespec="seconds"),
        }
    else:
        changes = {
            "port_confirmed_date": None,
            "port_confirmed_by": None,
            "port_confirmed_at": None,
        }
    for row in matching:
        update_management_fields(row["id"], changes, updated_by=updated_by)
    return {"item_no": normalized_item, "confirmed": confirmed, "updated": len(matching)}


def set_contract_reminder(
    item_no: str,
    manager_note: Optional[str] = None,
    next_follow_up_date: Optional[str] = None,
    updated_by: str = "",
) -> Dict[str, Any]:
    normalized_item = _normalize_text(item_no)
    matching = [row for row in list_order_finance_records() if _item_no(row) == normalized_item]
    if not matching:
        raise KeyError(normalized_item)
    normalized_note = _normalize_text(manager_note)
    normalized_date = _normalize_date(next_follow_up_date)
    if normalized_date and not _parse_date(normalized_date):
        raise ValueError("跟进日期格式不正确")
    stored_date = normalized_date or None
    changes = {"manager_note": normalized_note, "next_follow_up_date": stored_date}
    for row in matching:
        update_management_fields(row["id"], changes, updated_by=updated_by)
    return {
        "item_no": normalized_item,
        "manager_note": normalized_note,
        "next_follow_up_date": normalized_date,
        "updated": len(matching),
    }


def summarize_order_finance(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    active = [row for row in records if _normalize_text(row.get("business_status")) not in {"结案", "已完成", "已结算"}]
    due_soon = 0
    today = date.today()
    for row in active:
        due = _parse_date(row.get("finance_due_date"))
        if due and 0 <= (due - today).days <= 30:
            due_soon += 1
    finance_balance = sum(
        _to_float(row.get("planned_finance_amount")) or _to_float(row.get("finance_amount_actual")) or _to_float(row.get("finance_amount_expected")) or 0
        for row in active
    )
    return {
        "total_count": len(records),
        "active_count": len(active),
        "finance_balance": finance_balance,
        "due_30d_count": due_soon,
        "high_risk_count": len([row for row in records if row.get("risk_level") == "高"]),
    }


def _money_value(*values: Any) -> float:
    for value in values:
        parsed = _to_float(value)
        if parsed is not None:
            return parsed
    return 0.0


def _json_loads(value: Any, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return fallback


def _group_key(row: Dict[str, Any]) -> str:
    item_no = _item_no(row)
    if item_no and row.get("source_file") != "手动新增":
        return f"ITEM|{item_no}"
    parts = str(row.get("business_key") or "").split("|")
    if len(parts) >= 3:
        return "|".join(parts[:3])
    return "|".join([
        _normalize_text(row.get("subsidiary")),
        _normalize_text(row.get("purchase_contract_no")),
        _normalize_text(row.get("system_contract_no")),
    ])


def _item_no(row: Dict[str, Any]) -> str:
    source = _json_loads(row.get("source_json"), {})
    return _normalize_text(source.get("item_no")) or _normalize_text(row.get("purchase_contract_no")) or str(row.get("id"))


def _business_timestamp(value: Any) -> str:
    text = _normalize_text(value)
    return re.sub(r"\.\d+(?=Z|[+-]\d{2}:?\d{2}|$)", "", text) if text else ""


def _current_document_date(rows: List[Dict[str, Any]]) -> str:
    ignored = {"", "无", "0", "-", "未交单", "无需交单"}
    dates = sorted(
        value
        for row in rows
        if (value := _normalize_text(row.get("document_submission_date"))) not in ignored
    )
    return dates[-1] if dates else ""


ORDER_VESSEL_PROCESS_STATUS_TEXT = {
    "complete": "完成",
    "current": "当前",
    "pending": "待处理",
    "na": "不涉及融资",
}


def _order_vessel_exporter(business_no: str, rows: List[Dict[str, Any]]) -> str:
    exporters = []
    for row in rows:
        value = _normalize_text(row.get("overseas_entity"))
        if value and value not in exporters:
            exporters.append(value)
    if exporters:
        return " / ".join(exporters)
    return _xlsx_entity(business_no) if business_no else ""


FINAL_DESTINATION_ALIASES = {
    "JIANLONGUAE": "JIANLONG MIDDLE EAST STEEL TRADING-L.L.C",
    "JIANLONGMIDDLEEAST": "JIANLONG MIDDLE EAST STEEL TRADING-L.L.C",
    "JIANLONGMIDDLEEASTSTEELTRADINGLLC": "JIANLONG MIDDLE EAST STEEL TRADING-L.L.C",
    "MOLYCOPSG": "MOLYCOP SINGAPORE TRADING PTE LTD",
    "MOLYCOPSINGAPORE": "MOLYCOP SINGAPORE TRADING PTE LTD",
    "MOLYCOPSINGAPORETRADINGPTELTD": "MOLYCOP SINGAPORE TRADING PTE LTD",
    "SAMSUNGCTCORPORATION": "SAMSUNG C AND T CORPORATION",
    "SAMSUNGCANDTCORPORATION": "SAMSUNG C AND T CORPORATION",
    "COLAKOGLU": "COLAKOGLU METALURJI A.S.",
    "COLAKOGLUMETALURJIAS": "COLAKOGLU METALURJI A.S.",
}
FINAL_DESTINATION_MIDDLE_PARTIES = {
    "YOLANDA",
    "SINGAPOREYOLANDAPTELTD",
    "建龙国贸",
    "天津建龙",
    "香港建龙",
    "HONGKONGJIANLONG",
}
FINAL_DESTINATION_CONFIRMED = set(FINAL_DESTINATION_ALIASES.values())


def _company_alias_key(value: Any) -> str:
    text = _normalize_text(value).upper().replace("&", "AND")
    return re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", text)


def _is_final_destination_middle_party(value: Any) -> bool:
    key = _company_alias_key(value)
    return bool(key) and ("YOLANDA" in key or key in FINAL_DESTINATION_MIDDLE_PARTIES)


def _final_business_destination(value: Any) -> Dict[str, str]:
    raw = _normalize_text(value)
    key = _company_alias_key(raw)
    canonical = FINAL_DESTINATION_ALIASES.get(key)
    if canonical:
        return {"value": canonical, "status": "confirmed", "source": "当前确认R1"}
    if not raw or _is_final_destination_middle_party(raw):
        return {"value": "待确认", "status": "pending", "source": "待业务确认"}
    if raw in FINAL_DESTINATION_CONFIRMED:
        return {"value": raw, "status": "confirmed", "source": "当前确认R1"}
    return {"value": "待确认", "status": "pending", "source": "待业务确认"}


def _email_due_values(record: Dict[str, Any]) -> List[str]:
    raw_values = _json_loads(record.get("email_due_values_json"), [])
    if not isinstance(raw_values, list):
        return []
    values = []
    for value in raw_values:
        normalized = _normalize_date(value)
        if normalized and normalized not in values:
            values.append(normalized)
    return sorted(values)


def _classify_due_date_comparison(r1_value: Any, email_values: List[Any]) -> str:
    reporting_due = _normalize_date(r1_value)
    normalized_email = sorted({
        normalized
        for value in email_values
        if (normalized := _normalize_date(value))
    })
    if len(normalized_email) > 1:
        return "multiple_email_dates"
    if not reporting_due and not normalized_email:
        return "missing_both"
    if not reporting_due:
        return "missing_r1"
    if not normalized_email:
        return "missing_email"
    return "consistent" if reporting_due == normalized_email[0] else "conflict"


REPAYMENT_RISK_LABELS = {
    "repaid": "已回款",
    "actual_overdue": "实际逾期",
    "shipping_schedule_conflict": "预计船期冲突",
    "date_missing": "日期缺失",
    "source_conflict": "来源冲突",
    "no_schedule_comparison": "暂无船期比较",
    "normal": "正常",
    "not_applicable": "不涉及融资",
}


def _order_vessel_repayment_risks(
    record: Dict[str, Any],
    matched_rows: List[Dict[str, Any]],
    current_date: date,
) -> List[str]:
    if "不涉及融资" in _normalize_text(record.get("loan_amount_note")):
        return ["not_applicable"]
    paid_count = sum(1 for row in matched_rows if _normalize_text(row.get("tail_payment_date")))
    if matched_rows and paid_count == len(matched_rows):
        return ["repaid"]

    due_text = _normalize_date(record.get("reporting_repayment_due_date") or record.get("repayment_due_date"))
    due = _parse_date(due_text)
    states: List[str] = []
    if not due:
        states.append("date_missing")

    comparison_status = _normalize_text(record.get("due_date_comparison_status"))
    if comparison_status not in {"", "consistent"} and comparison_status != "missing_both":
        states.append("source_conflict")

    schedule_dates = [
        parsed
        for value in (
            record.get("loading_port_arrival_date"),
            record.get("planned_berth_date"),
        )
        if (parsed := _parse_date(value))
    ]
    if due:
        if current_date > due:
            states.append("actual_overdue")
        if schedule_dates:
            if any(due < schedule_date for schedule_date in schedule_dates):
                states.append("shipping_schedule_conflict")
        else:
            states.append("no_schedule_comparison")
        if "actual_overdue" not in states and "shipping_schedule_conflict" not in states:
            states.append("normal")
    return states


def _order_vessel_process_node(
    key: str,
    label: str,
    base_status: str,
    value: str,
    *,
    status_text: str = "",
    alerts: Optional[List[Dict[str, str]]] = None,
) -> Dict[str, Any]:
    return {
        "key": key,
        "label": label,
        "base_status": base_status,
        "value": value,
        "status_text": status_text or ORDER_VESSEL_PROCESS_STATUS_TEXT[base_status],
        "alerts": alerts or [],
    }


def _build_order_vessel_process(
    record: Dict[str, Any],
    matched_rows: List[Dict[str, Any]],
    current_date: date,
) -> Dict[str, Any]:
    document_status = _normalize_text(record.get("document_status"))
    document_date = _normalize_text(record.get("document_date"))
    due_date_text = _normalize_text(record.get("repayment_due_date"))
    due_date = _parse_date(due_date_text)
    arrival_date_text = _normalize_text(record.get("loading_port_arrival_date"))
    arrival_date = _parse_date(arrival_date_text)
    berth_date_text = _normalize_text(record.get("planned_berth_date"))
    berth_date = _parse_date(berth_date_text)
    repayment_dates = sorted(
        value
        for row in matched_rows
        if (value := _normalize_text(row.get("tail_payment_date")))
    )
    paid_count = sum(1 for row in matched_rows if _normalize_text(row.get("tail_payment_date")))
    all_paid = bool(matched_rows) and paid_count == len(matched_rows)
    payment_started = paid_count > 0
    non_financing = "不涉及融资" in _normalize_text(record.get("loan_amount_note"))
    document_complete = bool(document_date) or document_status in {"已交单", "无需交单"}
    prior_flow_complete = document_complete or payment_started

    arrival_alerts: List[Dict[str, str]] = []
    if not arrival_date_text:
        arrival_alerts.append({"kind": "missing", "text": "日期缺失"})
    if prior_flow_complete or (arrival_date and arrival_date < current_date):
        arrival_base = "complete"
        arrival_status = "完成"
    else:
        arrival_base = "current"
        arrival_status = "当前 · 待确认" if arrival_date_text else "当前 · 待补资料"

    berth_alerts: List[Dict[str, str]] = []
    if not berth_date_text:
        berth_alerts.append({"kind": "missing", "text": "日期缺失"})
    if prior_flow_complete:
        berth_base = "complete"
        berth_status = "完成 · 流程已推进"
    elif arrival_base == "complete":
        berth_base = "current"
        berth_status = "当前 · 待确认"
        if berth_date and berth_date < current_date:
            berth_alerts.append({"kind": "abnormal", "text": "计划日已过，待确认"})
    else:
        berth_base = "pending"
        berth_status = "待处理"

    document_alerts: List[Dict[str, str]] = []
    if document_status == "无需交单":
        document_base = "complete"
        document_value = "无需交单"
        document_status_text = "完成 · 无需交单"
    elif document_complete or payment_started:
        document_base = "complete"
        document_value = document_date or "已交单"
        document_status_text = "完成"
        if not document_date:
            document_alerts.append({"kind": "missing", "text": "交单日期缺失"})
    else:
        document_base = "pending"
        document_value = "尚未交单" if document_status else "状态未提供"
        document_status_text = "待处理"
        document_deadline = due_date - timedelta(days=15) if due_date else None
        if document_deadline and document_deadline <= current_date and not non_financing:
            document_alerts.append({
                "kind": "abnormal",
                "text": f"交单节点已到（{document_deadline.isoformat()}）",
            })

    repayment_alerts: List[Dict[str, str]] = []
    if non_financing:
        repayment_base = "na"
        repayment_value = "不适用"
        repayment_status = "不涉及融资"
    elif all_paid:
        repayment_base = "complete"
        repayment_value = repayment_dates[-1] if repayment_dates else "已回款"
        repayment_status = "完成"
    elif document_complete or payment_started:
        repayment_base = "current"
        repayment_value = f"到期 {due_date_text}" if due_date_text else "到期日未提供"
        repayment_status = "当前 · 待回款" if not paid_count else f"当前 · 已回款 {paid_count}/{len(matched_rows)} 笔"
    else:
        repayment_base = "pending"
        repayment_value = f"到期 {due_date_text}" if due_date_text else "到期日未提供"
        repayment_status = "待处理"
    repayment_risk_states = list(record.get("repayment_risk_states") or [])
    schedule_values = [
        value
        for value in (arrival_date_text, berth_date_text)
        if value
    ]
    if "date_missing" in repayment_risk_states:
        repayment_alerts.append({"kind": "missing", "text": "汇报还款到期日缺失"})
    if "source_conflict" in repayment_risk_states:
        repayment_alerts.append({"kind": "abnormal", "text": "来源冲突：R1与邮件待确认"})
    if "actual_overdue" in repayment_risk_states:
        repayment_alerts.append({
            "kind": "abnormal",
            "text": f"实际逾期：今天晚于汇报还款日 {due_date_text}",
        })
    if "shipping_schedule_conflict" in repayment_risk_states:
        repayment_alerts.append({
            "kind": "abnormal",
            "text": f"预计船期冲突：汇报还款日 {due_date_text} 早于 {' / '.join(schedule_values)}",
        })
    if "no_schedule_comparison" in repayment_risk_states:
        repayment_alerts.append({"kind": "missing", "text": "暂无船期比较，不能判定船期安全"})

    nodes = [
        _order_vessel_process_node(
            "arrival", "船到装港", arrival_base,
            arrival_date_text or "日期未提供",
            status_text=arrival_status,
            alerts=arrival_alerts,
        ),
        _order_vessel_process_node(
            "berth", "计划靠泊", berth_base,
            berth_date_text or "日期未提供",
            status_text=berth_status,
            alerts=berth_alerts,
        ),
        _order_vessel_process_node(
            "document", "交单", document_base, document_value,
            status_text=document_status_text,
            alerts=document_alerts,
        ),
        _order_vessel_process_node(
            "repayment", "还款", repayment_base, repayment_value,
            status_text=repayment_status,
            alerts=repayment_alerts,
        ),
    ]
    nodes[-1]["risk_states"] = repayment_risk_states
    nodes[-1]["risk_labels"] = [
        REPAYMENT_RISK_LABELS[state]
        for state in repayment_risk_states
    ]

    missing_fields = []
    for field, label in (
        ("exporter", "出口方"),
        ("steel_mill", "钢厂"),
        ("loading_port", "装港"),
        ("discharge_port", "卸港"),
        ("export_user", "最终业务去向/终端客户"),
    ):
        if not _normalize_text(record.get(field)):
            missing_fields.append(label)
    if _normalize_text(record.get("final_destination_status")) == "pending":
        missing_fields.append("最终业务去向/终端客户")
    for node, label in zip(nodes, ("船到装港日期", "计划靠泊日期", "交单日期", "汇报还款到期日")):
        if node["key"] == "repayment":
            is_missing = "date_missing" in repayment_risk_states
        else:
            is_missing = any(alert["kind"] == "missing" for alert in node["alerts"])
        if is_missing:
            missing_fields.append(label)
    missing_fields = list(dict.fromkeys(missing_fields))
    abnormal_count = sum(
        1 for node in nodes for alert in node["alerts"] if alert["kind"] == "abnormal"
    )
    status_set = {node["base_status"] for node in nodes}
    status_set.update(alert["kind"] for node in nodes for alert in node["alerts"])
    if missing_fields:
        status_set.add("missing")
    status_order = ("complete", "current", "pending", "abnormal", "missing", "na")
    status_values = [value for value in status_order if value in status_set]
    current_node = next((node for node in nodes if node["base_status"] == "current"), None)

    if non_financing:
        overall_tone = "na"
        overall_label = "融资：不涉及"
        if abnormal_count:
            overall_label += f" · 异常 {abnormal_count}"
        elif missing_fields:
            overall_label += f" · 缺失 {len(missing_fields)}"
    elif abnormal_count:
        overall_tone = "abnormal"
        prefix = f"当前：{current_node['label']}" if current_node else "流程"
        overall_label = f"{prefix} · 异常 {abnormal_count}"
    elif missing_fields:
        overall_tone = "missing"
        prefix = f"当前：{current_node['label']}" if current_node else "流程"
        overall_label = f"{prefix} · 缺失 {len(missing_fields)}"
    elif current_node:
        overall_tone = "current"
        overall_label = f"当前：{current_node['label']}"
    else:
        overall_tone = "complete"
        overall_label = "流程完成"

    return {
        "nodes": nodes,
        "status_values": status_values,
        "missing_fields": missing_fields,
        "abnormal_count": abnormal_count,
        "overall": {
            "tone": overall_tone,
            "label": overall_label,
            "current_key": current_node["key"] if current_node else "",
        },
    }


def build_order_vessel_overview(
    snapshots: Optional[List[Dict[str, Any]]] = None,
    finance_records: Optional[List[Dict[str, Any]]] = None,
    today: Optional[date] = None,
) -> Dict[str, Any]:
    snapshots = snapshots if snapshots is not None else list_order_vessel_snapshots()
    finance_records = finance_records if finance_records is not None else list_order_finance_records()
    current_date = today or date.today()
    finance_by_business_no: Dict[str, List[Dict[str, Any]]] = {}
    for row in finance_records:
        business_no = _item_no(row)
        if business_no:
            finance_by_business_no.setdefault(business_no, []).append(row)

    records = []
    for snapshot in snapshots:
        business_no = _normalize_text(snapshot.get("business_no"))
        matched_rows = finance_by_business_no.get(business_no, [])
        current_amount = sum(
            _money_value(
                row.get("finance_amount_actual"),
                row.get("finance_amount_expected"),
                row.get("planned_finance_amount"),
            )
            for row in matched_rows
        )
        funding_due_dates = sorted(
            value
            for row in matched_rows
            if (value := _normalize_text(row.get("finance_due_date"))) not in {"", "无", "0", "-"}
        )
        document_date = _current_document_date(matched_rows)
        if document_date:
            document_status = "已交单"
        elif matched_rows:
            document_status = _normalize_text(snapshot.get("document_status")) or "未交单"
        else:
            document_status = _normalize_text(snapshot.get("document_status"))

        source_amount = _to_float(snapshot.get("loan_amount"))
        loan_amount = current_amount if matched_rows else source_amount
        finance_updated_at = max(
            (_business_timestamp(row.get("updated_at")) for row in matched_rows),
            default="",
        )
        exporter = _order_vessel_exporter(business_no, matched_rows)
        destination = _final_business_destination(snapshot.get("export_user"))
        reporting_due_date = _normalize_date(snapshot.get("repayment_due_date"))
        email_due_values = _email_due_values(snapshot)
        due_date_comparison_status = _classify_due_date_comparison(
            reporting_due_date,
            email_due_values,
        )
        record = {
            "business_no": business_no,
            "exporter": exporter,
            "steel_mill": _normalize_text(snapshot.get("steel_mill")),
            "export_user": destination["value"],
            "final_business_destination": destination["value"],
            "final_destination_status": destination["status"],
            "final_destination_source": destination["source"],
            "cargo": _normalize_text(snapshot.get("cargo")),
            "vessel": _normalize_text(snapshot.get("vessel")),
            "quantity_mt": _to_float(snapshot.get("quantity_mt")),
            "loading_port": _normalize_text(snapshot.get("loading_port")),
            "loading_port_arrival_date": _normalize_text(snapshot.get("loading_port_arrival_date")),
            "planned_berth_date": _normalize_text(snapshot.get("planned_berth_date")),
            "discharge_port": _normalize_text(snapshot.get("discharge_port")),
            "estimated_discharge_date": _normalize_text(snapshot.get("estimated_discharge_date")),
            "document_status": document_status,
            "document_date": document_date,
            "repayment_due_date": reporting_due_date,
            "reporting_repayment_due_date": reporting_due_date,
            "reporting_due_date_source": _normalize_text(snapshot.get("reporting_due_date_source")) or "当前确认R1",
            "email_reporting_due_dates": email_due_values,
            "email_due_date_source": _normalize_text(snapshot.get("email_due_source")),
            "email_due_date_source_date": _normalize_date(snapshot.get("email_due_source_date")),
            "due_date_comparison_status": due_date_comparison_status,
            "funding_execution_due_date": funding_due_dates[0] if funding_due_dates else "",
            "funding_execution_due_dates": funding_due_dates,
            "preview_status": _normalize_text(snapshot.get("preview_status")) or "shadow",
            "loan_amount": loan_amount,
            "loan_amount_note": "" if loan_amount is not None else _normalize_text(snapshot.get("loan_amount_note")),
            "remark": _normalize_text(snapshot.get("remark")),
            "route_distance_nm": _to_float(snapshot.get("route_distance_nm")),
            "eta_start_date": _normalize_text(snapshot.get("eta_start_date")),
            "estimated_speed_knots": _to_float(snapshot.get("estimated_speed_knots")),
            "eta_basis": _normalize_text(snapshot.get("eta_basis")),
            "route_source": _normalize_text(snapshot.get("route_source")),
            "finance_match": bool(matched_rows),
            "finance_record_count": len(matched_rows),
            "finance_updated_at": finance_updated_at,
            "business_follow_source": "当前确认R1 + 订单融资资金事实" if matched_rows else "当前确认R1影子快照",
        }
        record["repayment_risk_states"] = _order_vessel_repayment_risks(record, matched_rows, current_date)
        record["repayment_risk_labels"] = [
            REPAYMENT_RISK_LABELS[state]
            for state in record["repayment_risk_states"]
        ]
        record["process"] = _build_order_vessel_process(record, matched_rows, current_date)
        records.append(record)

    source_row = snapshots[0] if snapshots else {}
    sync_status = get_order_finance_sync_status()
    return {
        "source": {
            "date": _normalize_text(source_row.get("source_date")),
            "version": _normalize_text(source_row.get("source_version")),
            "file_name": _normalize_text(source_row.get("source_file_name")),
            "sheet_name": _normalize_text(source_row.get("source_sheet_name")),
            "imported_at": _business_timestamp(source_row.get("updated_at") or source_row.get("created_at")),
        },
        "finance_sync": {
            "last_success_at": _business_timestamp(sync_status.get("last_success_at")),
            "matched_count": sum(1 for row in records if row["finance_match"]),
            "unmatched_count": sum(1 for row in records if not row["finance_match"]),
        },
        "summary": {
            "total_orders": len(records),
            "total_quantity_mt": sum(float(row.get("quantity_mt") or 0) for row in records),
            "financed_orders": sum(1 for row in records if _to_float(row.get("loan_amount")) is not None),
            "total_loan_amount": sum(float(row.get("loan_amount") or 0) for row in records),
            "pending_document_count": sum(1 for row in records if row.get("document_status") == "未交单"),
            "no_document_required_count": sum(1 for row in records if row.get("document_status") == "无需交单"),
        },
        "records": records,
    }


def _lc_info(row: Dict[str, Any]) -> Dict[str, Any]:
    items = _json_loads(row.get("sales_contracts_json"), [])
    if isinstance(items, list) and items:
        return items[0] or {}
    return {}


def _is_completed_group(rows: List[Dict[str, Any]]) -> bool:
    business_statuses = [_normalize_text(row.get("business_status")) for row in rows]
    if any(status == "存续" for status in business_statuses):
        return False
    explicit = [status for status in business_statuses if status]
    return bool(explicit) and all(status in {"结案", "已完成", "已结算"} for status in explicit)


def _group_has_value(rows: List[Dict[str, Any]], field: str) -> bool:
    return any(bool(_normalize_text(row.get(field))) for row in rows)


def _group_shipment_completed(rows: List[Dict[str, Any]]) -> bool:
    return _group_has_value(rows, "shipment_confirmed_date") or _group_has_value(rows, "document_submission_date")


def _row_is_paid(row: Dict[str, Any]) -> bool:
    return bool(_normalize_text(row.get("tail_payment_date")))


def _group_document_date(rows: List[Dict[str, Any]]) -> str:
    dates = sorted(
        row["document_submission_date"]
        for row in rows
        if row.get("document_submission_date")
    )
    return dates[-1] if dates else ""


def _row_document_deadline(row: Dict[str, Any]) -> str:
    due = _parse_date(row.get("finance_due_date"))
    return (due - timedelta(days=15)).isoformat() if due else ""


def _group_stage(rows: List[Dict[str, Any]]) -> str:
    if _is_completed_group(rows):
        return "已完成"
    has_loan = any(row.get("finance_drawdown_date") or _money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected")) for row in rows)
    has_document = bool(_group_document_date(rows))
    if rows and all(_row_is_paid(row) for row in rows):
        return "已回款待结案"
    if has_document:
        return "已交单待回款"
    if _group_has_value(rows, "shipment_confirmed_date"):
        return "已装船待交单"
    if _group_has_value(rows, "port_confirmed_date"):
        return "已集港待装船"
    return "已放款待集港" if has_loan else "待放款"


def _days_to(value: Any) -> Optional[int]:
    parsed = _parse_date(value)
    if not parsed:
        return None
    return (parsed - date.today()).days


def _warning_indicator(warning: Dict[str, Any]) -> str:
    field = _normalize_text(warning.get("field"))
    message = _normalize_text(warning.get("message"))
    if field == "latest_shipment_date" or "最迟装船" in message:
        return "shipment"
    if field == "document_submission_date" or "交单" in message:
        return "document"
    if field == "finance_due_date" or any(text in message for text in ("融资到期", "贷款到期", "还款到期", "回款")):
        return "payment"
    return "reminder"


def _group_indicator_risks(rows: List[Dict[str, Any]], stage: str) -> Dict[str, str]:
    risks = {"shipment": "低", "document": "低", "payment": "低", "reminder": "低"}
    if stage == "已完成":
        return risks
    if stage == "已回款待结案":
        return risks
    if stage == "已交单待回款":
        unpaid_rows = [row for row in rows if not _row_is_paid(row)]
        due_days = [
            _days_to(row.get("finance_due_date"))
            for row in unpaid_rows
            if row.get("finance_due_date")
        ]
        risks["payment"] = (
            "高" if any(days is not None and days <= 0 for days in due_days) else "中"
        )
        return risks
    if stage == "已集港待装船":
        shipment_days = [
            _days_to(row.get("latest_shipment_date"))
            for row in rows
            if row.get("latest_shipment_date")
        ]
        valid_days = [days for days in shipment_days if days is not None]
        risks["shipment"] = "高" if not valid_days or min(valid_days) <= 2 else "中"
        return risks
    shipment_completed = _group_shipment_completed(rows)
    warnings = [
        warning
        for row in rows
        for warning in _json_loads(row.get("import_warnings_json"), [])
    ]
    for warning in warnings:
        indicator = _warning_indicator(warning)
        if indicator == "shipment" and shipment_completed:
            continue
        risks[indicator] = "高"

    if not shipment_completed:
        shipment_days = [_days_to(row.get("latest_shipment_date")) for row in rows if row.get("latest_shipment_date")]
        min_shipment = min([item for item in shipment_days if item is not None], default=None)
        if min_shipment is None or min_shipment < 0:
            risks["shipment"] = "高"
        elif min_shipment <= 10 and risks["shipment"] != "高":
            risks["shipment"] = "中"

    follow_up_days = [_days_to(row.get("next_follow_up_date")) for row in rows if row.get("next_follow_up_date")]
    if any(item is not None and item <= 10 for item in follow_up_days):
        risks["reminder"] = "中"

    unpaid_rows = [row for row in rows if not _row_is_paid(row)]
    if not _group_document_date(rows):
        deadline_days = [
            _days_to(deadline)
            for row in unpaid_rows
            if (deadline := _row_document_deadline(row))
        ]
        if any(item is not None and item <= 0 for item in deadline_days):
            risks["document"] = "中"

    due_days = [_days_to(row.get("finance_due_date")) for row in unpaid_rows if row.get("finance_due_date")]
    if any(item is not None and item <= 7 for item in due_days):
        risks["payment"] = "高"
    elif any(item is not None and item <= 30 for item in due_days):
        risks["payment"] = "中"
    return risks


def _group_risk(indicator_risks: Dict[str, str], stage: str) -> str:
    if stage == "已完成":
        return "已完成"
    if "高" in indicator_risks.values():
        return "高"
    if "中" in indicator_risks.values():
        return "中"
    return "低"


def _group_weekly_focus_reasons(rows: List[Dict[str, Any]], stage: str, risk: str) -> List[str]:
    if stage == "已完成":
        return []
    if stage == "已回款待结案":
        return []
    if stage == "已交单待回款":
        return ["high_risk"] if risk == "高" else []
    if stage == "已集港待装船":
        return ["high_risk"] if risk == "高" else []
    reasons = []
    if risk == "高":
        reasons.append("high_risk")
    if not _group_shipment_completed(rows):
        shipment_days = [_days_to(row.get("latest_shipment_date")) for row in rows if row.get("latest_shipment_date")]
        if any(item is not None and 0 <= item <= 10 for item in shipment_days):
            reasons.append("shipment_follow_up")
    if not _group_document_date(rows):
        unpaid_rows = [row for row in rows if not _row_is_paid(row)]
        deadline_days = [
            _days_to(deadline)
            for row in unpaid_rows
            if (deadline := _row_document_deadline(row))
        ]
        if any(item is not None and item <= 0 for item in deadline_days):
            reasons.append("document_follow_up")
    follow_up_days = [_days_to(row.get("next_follow_up_date")) for row in rows if row.get("next_follow_up_date")]
    if any(item is not None and item <= 10 for item in follow_up_days):
        reasons.append("manual_follow_up")
    return reasons


def _group_repayment_timing(rows: List[Dict[str, Any]]) -> str:
    deltas = []
    for row in rows:
        due = _parse_date(row.get("finance_due_date"))
        repaid = _parse_date(row.get("tail_payment_date"))
        if due and repaid:
            deltas.append((repaid - due).days)
    if not deltas:
        return ""
    latest_delta = max(deltas)
    if latest_delta > 0:
        return f"逾期 {latest_delta} 天回款"
    if latest_delta < 0:
        return f"提前 {abs(latest_delta)} 天回款"
    return "按期回款"


def _group_next_action(rows: List[Dict[str, Any]], stage: str, risk: str) -> str:
    if stage == "已集港待装船":
        return "确认装船状态或交单状态" if risk == "高" else "跟进装船并及时确认状态"
    manual = next((_normalize_text(row.get("next_action")) for row in rows if _normalize_text(row.get("next_action"))), "")
    if manual and manual != "无":
        return manual
    if stage == "已完成":
        return "已闭环，保留历史查询"
    if stage == "待放款":
        return "确认贷款行、金额和借款日期"
    if stage == "已放款待集港":
        return "优先联系工厂确认集港进度" if risk == "高" else "跟进集港进度"
    if stage == "已装船待交单":
        return "跟进交单并确认交单日"
    if stage == "已交单待回款":
        return "跟进回款并确认回款日"
    if stage == "已回款待结案":
        return "确认订单结案状态"
    return "确认当前订单状态"


def _build_progress_group(group_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    rows = sorted(group_rows, key=lambda row: (row.get("finance_due_date") or "9999-12-31", row.get("id") or 0))
    first = rows[0]
    lc = _lc_info(first)
    stage = _group_stage(rows)
    indicator_risks = _group_indicator_risks(rows, stage)
    risk = _group_risk(indicator_risks, stage)
    finance_total = sum(_money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount")) for row in rows)
    due_dates = sorted([row.get("finance_due_date") for row in rows if row.get("finance_due_date")])
    unpaid_rows = [row for row in rows if not _row_is_paid(row)]
    missing_due_count = (
        sum(1 for row in unpaid_rows if not _normalize_text(row.get("finance_due_date")))
        if stage == "已交单待回款"
        else 0
    )
    unpaid_due_dates = sorted([row.get("finance_due_date") for row in unpaid_rows if row.get("finance_due_date")])
    document_deadlines = sorted([
        deadline for row in unpaid_rows if (deadline := _row_document_deadline(row))
    ])
    latest_shipment_dates = sorted([row.get("latest_shipment_date") for row in rows if row.get("latest_shipment_date")])
    document_dates = sorted([row.get("document_submission_date") for row in rows if row.get("document_submission_date")])
    bill_dates = sorted([row.get("bill_of_lading_date") for row in rows if row.get("bill_of_lading_date")])
    repay_dates = sorted([row.get("tail_payment_date") for row in rows if row.get("tail_payment_date")])
    shipment_confirmed_dates = sorted([row.get("shipment_confirmed_date") for row in rows if row.get("shipment_confirmed_date")])
    shipment_confirmed_at = sorted([row.get("shipment_confirmed_at") for row in rows if row.get("shipment_confirmed_at")])
    port_confirmed_dates = sorted([row.get("port_confirmed_date") for row in rows if row.get("port_confirmed_date")])
    port_confirmed_at = sorted([row.get("port_confirmed_at") for row in rows if row.get("port_confirmed_at")])
    follow_up_dates = sorted([row.get("next_follow_up_date") for row in rows if row.get("next_follow_up_date")])
    manager_note = next((_normalize_text(row.get("manager_note")) for row in rows if _normalize_text(row.get("manager_note"))), "")
    weekly_focus_reasons = _group_weekly_focus_reasons(rows, stage, risk)
    paid_count = len(rows) - len(unpaid_rows)
    if paid_count == len(rows):
        payment_progress = f"已回款 {paid_count}/{len(rows)}笔"
    elif paid_count:
        payment_progress = f"部分回款 {paid_count}/{len(rows)}笔"
    else:
        payment_progress = f"待回款 0/{len(rows)}笔"
    warnings = []
    for row in rows:
        warnings.extend(_json_loads(row.get("import_warnings_json"), []))
    return {
        "id": _group_key(first),
        "item_no": _item_no(first),
        "entity": first.get("overseas_entity") or "",
        "subsidiary": first.get("subsidiary") or "",
        "contract_no": first.get("purchase_contract_no") or "",
        "system_contract_no": first.get("system_contract_no") or "",
        "product": first.get("product_name") or "",
        "quantity": first.get("contract_quantity_mt"),
        "terminal_customer": first.get("terminal_customer") or "",
        "issuing_bank": lc.get("lc_bank") or "",
        "lc_no": lc.get("lc_no") or "",
        "lc_amount": lc.get("lc_amount"),
        "lc_expiry_date": lc.get("lc_expiry_date") or first.get("lc_latest_shipment_date") or "",
        "lc_type": lc.get("lc_type") or "",
        "transferable": lc.get("transferable") or "",
        "receiving_bank": lc.get("receiving_bank") or "",
        "latest_shipment_date": latest_shipment_dates[0] if latest_shipment_dates else "",
        "port_confirmed_date": port_confirmed_dates[-1] if port_confirmed_dates else "",
        "port_confirmed_by": next((row.get("port_confirmed_by") for row in rows if row.get("port_confirmed_by")), ""),
        "port_confirmed_at": port_confirmed_at[-1] if port_confirmed_at else "",
        "shipment_completed": _group_shipment_completed(rows),
        "shipment_basis": "document" if document_dates else "manual" if shipment_confirmed_dates else "",
        "shipment_confirmed_date": shipment_confirmed_dates[-1] if shipment_confirmed_dates else "",
        "shipment_confirmed_by": next((row.get("shipment_confirmed_by") for row in rows if row.get("shipment_confirmed_by")), ""),
        "shipment_confirmed_at": shipment_confirmed_at[-1] if shipment_confirmed_at else "",
        "vessel": next((row.get("vessel_voyage") for row in rows if row.get("vessel_voyage")), ""),
        "latest_due_date": due_dates[0] if due_dates else "",
        "payment_due_date": unpaid_due_dates[0] if unpaid_due_dates else (due_dates[0] if due_dates else ""),
        "document_deadline": document_deadlines[0] if document_deadlines else "",
        "bill_date": bill_dates[-1] if bill_dates else "",
        "document_date": document_dates[-1] if document_dates else "",
        "repay_date": repay_dates[-1] if repay_dates else "",
        "payment_progress": payment_progress,
        "repayment_timing": _group_repayment_timing(rows),
        "stage": stage,
        "risk": risk,
        "indicator_risks": indicator_risks,
        "manager_note": manager_note,
        "next_follow_up_date": follow_up_dates[0] if follow_up_dates else "",
        "is_weekly_focus": bool(weekly_focus_reasons),
        "weekly_focus_reasons": weekly_focus_reasons,
        "next_action": _group_next_action(rows, stage, risk),
        "total_finance": finance_total,
        "financing_count": len(rows),
        "data_issue_count": (
            len([warning for warning in warnings if _is_data_quality_warning(warning)])
            + missing_due_count
        ),
        "source_file": first.get("source_file") or "",
        "source_sheet": first.get("source_sheet") or "",
        "source_row_start": min((row.get("source_row_start") or 0 for row in rows), default=0),
        "source_row_end": max((row.get("source_row_end") or row.get("source_row_start") or 0 for row in rows), default=0),
        "financings": [
            {
                "id": row.get("id"),
                "bank": row.get("finance_bank") or "",
                "amount": _money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount")),
                "borrow_date": row.get("finance_drawdown_date") or "",
                "original_due_date": _json_loads(row.get("source_json"), {}).get("original_due_date") or "",
                "new_due_date": _json_loads(row.get("source_json"), {}).get("new_due_date") or "",
                "extension_days": _json_loads(row.get("source_json"), {}).get("extension_days") or 0,
                "due_date": row.get("finance_due_date") or "",
                "rate": _json_loads(row.get("source_json"), {}).get("finance_rate"),
                "bill_date": row.get("bill_of_lading_date") or "",
                "document_date": row.get("document_submission_date") or "",
                "repay_date": row.get("tail_payment_date") or "",
                "payment_state": "已回款" if _row_is_paid(row) else "待回款",
                "status": row.get("finance_status") or row.get("business_status") or "",
                "next_action": row.get("next_action") or "",
                "source_file": row.get("source_file") or "",
                "source_sheet": row.get("source_sheet") or "",
                "source_row_start": row.get("source_row_start") or 0,
            }
            for row in rows
        ],
    }


def build_order_finance_progress_view(records: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    use_persisted_records = records is None
    records = records if records is not None else list_order_finance_records()
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for row in records:
        groups.setdefault(_group_key(row), []).append(row)
    contracts = [_build_progress_group(rows) for rows in groups.values()]
    contracts.sort(key=lambda item: (
        0 if item["risk"] == "高" else 1 if item["risk"] == "中" else 2 if item["risk"] == "低" else 3,
        item.get("latest_due_date") or "9999-12-31",
        item.get("item_no") or "",
    ))
    open_contracts = [item for item in contracts if item["stage"] != "已完成"]
    completed_contracts = sorted(
        (item for item in contracts if item["stage"] == "已完成"),
        key=lambda item: (
            bool(item.get("latest_due_date")),
            item.get("latest_due_date") or "",
            item.get("item_no") or "",
        ),
        reverse=True,
    )
    contracts = open_contracts + completed_contracts
    summary = {
        "open_contracts": len(open_contracts),
        "active_finance": sum(item["total_finance"] for item in open_contracts),
        "due_7d": len([item for item in open_contracts if (days := _days_to(item.get("payment_due_date"))) is not None and 0 <= days <= 7]),
        "due_30d": len([item for item in open_contracts if (days := _days_to(item.get("payment_due_date"))) is not None and 0 <= days <= 30]),
        "focus_risk": len([item for item in open_contracts if item["is_weekly_focus"]]),
        "pending_drawdown": len([item for item in open_contracts if item["stage"] == "待放款"]),
        "financed_uncollected": len([item for item in open_contracts if item["stage"] == "已放款待集港"]),
        "collected_unshipped": len([item for item in open_contracts if item["stage"] == "已集港待装船"]),
        "shipped_undocumented": len([item for item in open_contracts if item["stage"] == "已装船待交单"]),
        "documented_unpaid": len([item for item in open_contracts if item["stage"] == "已交单待回款"]),
        "paid_unclosed": len([item for item in open_contracts if item["stage"] == "已回款待结案"]),
        "completed": len([item for item in contracts if item["stage"] == "已完成"]),
        "data_issues": sum(item["data_issue_count"] for item in open_contracts),
        "total_contracts": len(contracts),
    }
    sync_status = get_order_finance_sync_status() if use_persisted_records else {}
    return {
        "summary": summary,
        "contracts": contracts,
        "sync_status": {
            "last_success_at": sync_status.get("last_success_at"),
            "changed_count": int(sync_status.get("changed_count") or 0),
        },
    }


def _sum_by(records: List[Dict[str, Any]], field: str) -> List[Dict[str, Any]]:
    totals: Dict[str, float] = {}
    for row in records:
        key = _normalize_text(row.get(field)) or "未填"
        totals[key] = totals.get(key, 0.0) + _money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount"))
    return [{"name": key, "amount": value} for key, value in sorted(totals.items(), key=lambda item: item[1], reverse=True)]


def _workbook_capital_metadata(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    for row in records:
        source = _json_loads(row.get("source_json"), {})
        capital = source.get("workbook_capital") if isinstance(source, dict) else None
        if isinstance(capital, dict) and capital.get("banks"):
            return capital
    return {}


def build_order_finance_capital_view(records: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    records = records if records is not None else list_order_finance_records()
    open_rows = [row for row in records if _group_stage([row]) != "已完成"]
    bank_used: Dict[str, float] = {}
    for row in open_rows:
        bank = _normalize_text(row.get("finance_bank")) or "未填贷款行"
        bank_used[bank] = bank_used.get(bank, 0.0) + _money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount"))
    capital_metadata = _workbook_capital_metadata(records)
    quota_banks = capital_metadata.get("banks") or DEFAULT_BANK_LIMITS
    bank_limit_map = {row["bank"]: row for row in quota_banks}
    all_banks = sorted(set(bank_limit_map) | set(bank_used))
    bank_usage = []
    for bank in all_banks:
        limit_row = bank_limit_map.get(bank, {"bank": bank, "limit": 0, "note": "", "lc_requirement": "", "bill_requirement": "", "finance_ratio": "", "term": ""})
        order_used = bank_used.get(bank, 0.0)
        limit = float(limit_row.get("limit") or 0)
        used = float(limit_row.get("used") if limit_row.get("used") is not None else order_used)
        available = limit_row.get("available")
        if available is None and limit:
            available = limit - used
        bank_usage.append({
            **limit_row,
            "used": used,
            "available": available,
            "usage_rate": used / limit if limit else None,
            "order_used": order_used,
            "difference": used - order_used,
        })
    bank_usage.sort(key=lambda item: item["used"], reverse=True)
    total_credit = float(capital_metadata.get("total_credit") or sum(float(row.get("limit") or 0) for row in quota_banks))
    order_used_credit = sum(bank_used.values())
    used_credit = float(capital_metadata.get("used_credit") if capital_metadata.get("used_credit") is not None else order_used_credit)
    available_credit = float(capital_metadata.get("available_credit") if capital_metadata.get("available_credit") is not None else total_credit - used_credit)
    buckets = [
        {"label": "7天内", "min": 0, "max": 7},
        {"label": "8-30天", "min": 8, "max": 30},
        {"label": "31-60天", "min": 31, "max": 60},
        {"label": "60天以上", "min": 61, "max": 99999},
        {"label": "已逾期", "min": -99999, "max": -1},
    ]
    due_buckets = []
    for bucket in buckets:
        rows = [row for row in open_rows if (days := _days_to(row.get("finance_due_date"))) is not None and bucket["min"] <= days <= bucket["max"]]
        due_buckets.append({
            "label": bucket["label"],
            "count": len(rows),
            "amount": sum(_money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount")) for row in rows),
        })
    supplier_usage = _sum_by(open_rows, "subsidiary")
    entity_usage = _sum_by(open_rows, "overseas_entity")
    due_30_amount = sum(bucket["amount"] for bucket in due_buckets if bucket["label"] in {"7天内", "8-30天"})
    largest_bank = max((row["used"] for row in bank_usage), default=0)
    largest_supplier = max((row["amount"] for row in supplier_usage), default=0)
    return {
        "summary": {
            "total_credit": total_credit,
            "used_credit": used_credit,
            "available_credit": available_credit,
            "order_used_credit": order_used_credit,
            "usage_difference": used_credit - order_used_credit,
            "utilization_rate": used_credit / total_credit if total_credit else 0,
            "near_limit_banks": len([row for row in bank_usage if row.get("usage_rate") is not None and row["usage_rate"] >= 0.9]),
            "due_30_amount": due_30_amount,
            "largest_bank_share": largest_bank / used_credit if used_credit else 0,
            "largest_supplier_share": largest_supplier / used_credit if used_credit else 0,
        },
        "bank_usage": bank_usage,
        "entity_usage": entity_usage,
        "supplier_usage": supplier_usage,
        "due_buckets": due_buckets,
        "bank_details": [
            {
                "bank": row.get("finance_bank") or "未填贷款行",
                "item_no": _item_no(row),
                "contract_no": row.get("purchase_contract_no") or "",
                "subsidiary": row.get("subsidiary") or "",
                "amount": _money_value(row.get("finance_amount_actual"), row.get("finance_amount_expected"), row.get("planned_finance_amount")),
                "due_date": row.get("finance_due_date") or "",
                "status": row.get("finance_status") or row.get("business_status") or "",
            }
            for row in open_rows
        ],
    }


@router.post("/order-finance/import-local")
def import_order_finance_local(request: ImportLocalRequest, user: dict = Depends(order_finance_current_user)):
    order_finance_require_import(user)
    try:
        result = import_order_finance_directory(Path(request.directory), imported_by=user["name"])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return result


@router.post("/order-finance/import-file")
async def import_order_finance_file(
    request: Request,
    file_name: str,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_import(user)
    try:
        return await import_order_finance_upload(request, file_name=file_name, imported_by=user["name"])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/order-finance/records")
def order_finance_records(
    limit: int = 5000,
    offset: int = 0,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_view(user)
    result = list_order_finance_records_page(limit=limit, offset=offset)
    return {"summary": summarize_order_finance(result["records"]), **result}


@router.get("/order-finance/progress")
def order_finance_progress(user: dict = Depends(order_finance_current_user)):
    order_finance_require_view(user)
    return build_order_finance_progress_view()


@router.get("/order-finance/vessel-overview")
def order_finance_vessel_overview(user: dict = Depends(order_finance_current_user)):
    order_finance_require_view(user)
    return build_order_vessel_overview()


@router.get("/order-finance/capital")
def order_finance_capital(user: dict = Depends(order_finance_current_user)):
    require_permission(user, "order_finance.capital", "view")
    return build_order_finance_capital_view()


@router.get("/order-finance/records/{record_id}")
def order_finance_record(record_id: int, user: dict = Depends(order_finance_current_user)):
    order_finance_require_view(user)
    try:
        return get_order_finance_record(record_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="记录不存在") from exc


@router.post("/order-finance/records/manual")
def order_finance_create_manual(request: ManualOrderFinanceRequest, user: dict = Depends(order_finance_current_user)):
    order_finance_require_edit(user)
    changes = request.model_dump(exclude_unset=True) if hasattr(request, "model_dump") else request.dict(exclude_unset=True)
    try:
        duplicates = find_order_finance_duplicates(changes)
        record = create_manual_order_finance_record(changes, created_by=user["name"])
    except DuplicateOrderFinanceError as exc:
        raise HTTPException(status_code=409, detail={"message": str(exc), "existing": exc.existing}) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"record": record, "duplicate_candidates": duplicates["similar"]}


@router.patch("/order-finance/records/{record_id}/management")
def order_finance_update_management(
    record_id: int,
    request: ManagementUpdateRequest,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_edit(user)
    try:
        changes = request.model_dump(exclude_unset=True) if hasattr(request, "model_dump") else request.dict(exclude_unset=True)
        return update_management_fields(record_id, changes, updated_by=user["name"])
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="记录不存在") from exc


@router.patch("/order-finance/contracts/{item_no}/shipment-confirmation")
def order_finance_shipment_confirmation(
    item_no: str,
    request: ShipmentConfirmationRequest,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_edit(user)
    try:
        return set_shipment_confirmation(
            item_no,
            confirmed=request.confirmed,
            shipment_confirmed_date=request.shipment_confirmed_date,
            updated_by=user["name"],
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="项次不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.patch("/order-finance/contracts/{item_no}/port-confirmation")
def order_finance_port_confirmation(
    item_no: str,
    request: PortConfirmationRequest,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_edit(user)
    try:
        return set_port_confirmation(
            item_no,
            confirmed=request.confirmed,
            port_confirmed_date=request.port_confirmed_date,
            updated_by=user["name"],
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="项次不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.patch("/order-finance/contracts/{item_no}/reminder")
def order_finance_contract_reminder(
    item_no: str,
    request: ContractReminderRequest,
    user: dict = Depends(order_finance_current_user),
):
    order_finance_require_edit(user)
    try:
        return set_contract_reminder(
            item_no,
            manager_note=request.manager_note,
            next_follow_up_date=request.next_follow_up_date,
            updated_by=user["name"],
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="项次不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
